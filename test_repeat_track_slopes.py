import csv
import math
from pathlib import Path
from collections import defaultdict

import numpy as np
from openpyxl import load_workbook


# ================================================================
# CONFIGURATION
# ================================================================

CSV_FILE = Path("gnssir_tide_arc_analysis.csv")
TIDE_FILE = Path("marconi_tides_sherwood.xlsx")

REFERENCE_HEIGHT = 18.665

MIN_OBSERVATIONS = 3

AZ_SECTORS = {
    "ALL": (0, 360),
    "100-150": (100, 150),
    "100-130": (100, 130),
    "110-130": (110, 130),
    "110-120": (110, 120),
}


# ================================================================
# UTILITIES
# ================================================================

def finite(x):

    try:
        x = float(x)

        if math.isfinite(x):
            return x

    except (TypeError, ValueError):
        pass

    return None


def correlation(x, y):

    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 2:
        return float("nan")

    if np.std(x) == 0 or np.std(y) == 0:
        return float("nan")

    return float(
        np.corrcoef(x, y)[0, 1]
    )


def linear_fit(x, y):

    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 2:
        return float("nan"), float("nan"), float("nan")

    slope, intercept = np.polyfit(
        x,
        y,
        1,
    )

    predicted = (
        intercept
        + slope * x
    )

    residuals = y - predicted

    rms = np.sqrt(
        np.mean(
            residuals ** 2
        )
    )

    return (
        float(slope),
        float(intercept),
        float(rms),
    )


# ================================================================
# LOAD TIDE MODELS
# ================================================================

def load_tide_models():

    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb.active

    header = [
        cell.value
        for cell in ws[1]
    ]

    models = [
        "EOT20_heightm",
        "GOT5.5_heightm",
        "GOT5.6_heightm",
        "FES2022_heightm",
    ]

    if "time" not in header:
        raise RuntimeError(
            "Tide workbook has no time column."
        )

    time_col = header.index("time")

    model_cols = {}

    for model in models:

        if model not in header:
            raise RuntimeError(
                f"Missing tide model: {model}"
            )

        model_cols[model] = header.index(model)

    times = []

    values = {
        model: []
        for model in models
    }

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):

        t = row[time_col]

        if t is None:
            continue

        valid = True

        row_values = {}

        for model in models:

            value = finite(
                row[model_cols[model]]
            )

            if value is None:

                valid = False
                break

            row_values[model] = value

        if not valid:
            continue

        times.append(t)

        for model in models:

            values[model].append(
                row_values[model]
            )

    if len(times) < 2:

        raise RuntimeError(
            "Not enough tide data."
        )

    epoch = np.array(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    interpolators = {}

    for model in models:

        arr = np.asarray(
            values[model],
            dtype=float,
        )

        def make_interp(arr):

            def interp(query_time):

                q = (
                    query_time
                    - times[0]
                ).total_seconds()

                if (
                    q < epoch[0]
                    or q > epoch[-1]
                ):

                    return float("nan")

                return float(
                    np.interp(
                        q,
                        epoch,
                        arr,
                    )
                )

            return interp

        interpolators[model] = make_interp(
            arr
        )

    print()
    print("Tide models:")

    for model in models:
        print(f"  {model}")

    print(
        f"Tide points: {len(times)}"
    )

    print(
        f"Tide coverage: "
        f"{times[0]} through {times[-1]}"
    )

    return interpolators


# ================================================================
# LOAD GNSS-IR CSV
# ================================================================

def load_records():

    records = []

    with open(
        CSV_FILE,
        newline="",
    ) as f:

        reader = csv.DictReader(f)

        for row in reader:

            try:

                dt = (
                    row[
                        "solution_time_utc"
                    ]
                )

                from datetime import datetime

                dt = datetime.fromisoformat(
                    dt
                )

                sat = int(
                    float(
                        row["sat"]
                    )
                )

                freq = int(
                    float(
                        row["freq"]
                    )
                )

                rh = float(
                    row["RH_m"]
                )

                az = float(
                    row["Azim"]
                )

                if not all(
                    math.isfinite(x)
                    for x in (
                        rh,
                        az,
                    )
                ):
                    continue

                records.append(
                    {
                        "dt": dt,
                        "sat": sat,
                        "freq": freq,
                        "rh": rh,
                        "az": az,
                    }
                )

            except (
                KeyError,
                ValueError,
                TypeError,
            ):

                continue

    return records


# ================================================================
# MAIN
# ================================================================

print()
print("=" * 80)
print(
    "GNSS-IR REPEAT TRACK / TIDE SLOPE TEST"
)
print("=" * 80)

print(
    f"CSV: {CSV_FILE}"
)

print(
    f"Tide file: {TIDE_FILE}"
)

print(
    f"Reference height: "
    f"{REFERENCE_HEIGHT:.3f} m"
)

print(
    f"Minimum observations per track: "
    f"{MIN_OBSERVATIONS}"
)

tide_models = load_tide_models()

records = load_records()

print()
print(
    f"GNSS-IR records loaded: "
    f"{len(records)}"
)


# ================================================================
# ADD TIDE VALUES
# ================================================================

for r in records:

    for model, interp in tide_models.items():

        r[model] = interp(
            r["dt"]
        )


# ================================================================
# GROUP BY SATELLITE + FREQUENCY
# ================================================================

groups = defaultdict(list)

for r in records:

    groups[
        (
            r["sat"],
            r["freq"],
        )
    ].append(r)


print()
print(
    f"Unique satellite/frequency "
    f"tracks: {len(groups)}"
)


# ================================================================
# ANALYZE TRACKS
# ================================================================

all_results = []

for sector_name, (
    az_lo,
    az_hi,
) in AZ_SECTORS.items():

    print()
    print("=" * 80)
    print(
        f"SECTOR: {sector_name}"
    )
    print("=" * 80)

    sector_results = []

    for (
        sat,
        freq,
    ), group in sorted(
        groups.items()
    ):

        group = [

            r

            for r in group

            if (
                az_lo
                <= r["az"]
                <= az_hi
            )

        ]

        if len(group) < MIN_OBSERVATIONS:
            continue

        group.sort(
            key=lambda r: r["dt"]
        )

        az_mean = np.mean(
            [
                r["az"]
                for r in group
            ]
        )

        for model in tide_models:

            x = np.array(
                [
                    r[model]
                    for r in group
                ],
                dtype=float,
            )

            y = np.array(
                [
                    r["rh"]
                    for r in group
                ],
                dtype=float,
            )

            valid = (
                np.isfinite(x)
                & np.isfinite(y)
            )

            x = x[valid]
            y = y[valid]

            if len(x) < MIN_OBSERVATIONS:
                continue

            r_value = correlation(
                x,
                y,
            )

            slope, intercept, rms = (
                linear_fit(
                    x,
                    y,
                )
            )

            result = {
                "sector": sector_name,
                "sat": sat,
                "freq": freq,
                "n": len(x),
                "az_mean": az_mean,
                "model": model,
                "r": r_value,
                "slope": slope,
                "intercept": intercept,
                "rms": rms,
            }

            sector_results.append(
                result
            )

            all_results.append(
                result
            )

    if not sector_results:

        print(
            "No repeat tracks with "
            "enough observations."
        )

        continue

    print()
    print(
        "SAT FREQ  N   AZmean   MODEL             "
        "r       SLOPE      RMS"
    )

    print("-" * 80)

    for r in sector_results:

        print(
            f"{r['sat']:3d} "
            f"{r['freq']:4d} "
            f"{r['n']:2d} "
            f"{r['az_mean']:7.2f} "
            f"{r['model']:17s} "
            f"{r['r']:+7.3f} "
            f"{r['slope']:+9.3f} "
            f"{r['rms']*100:7.2f} cm"
        )


# ================================================================
# FOCUS ON PHYSICALLY INTERESTING TRACKS
# ================================================================

print()
print("=" * 80)
print(
    "PHYSICALLY INTERESTING TRACKS"
)
print("=" * 80)

interesting = [

    r

    for r in all_results

    if (
        r["model"]
        == "EOT20_heightm"
        and r["slope"] < 0
    )

]

interesting.sort(
    key=lambda r:
    abs(r["slope"] + 1.0)
)


if interesting:

    print()
    print(
        "Negative slopes, closest to "
        "the expected -1 m/m response:"
    )

    print()
    print(
        "SECTOR      SAT FREQ N   AZ     "
        "r       SLOPE     RMS"
    )

    print("-" * 80)

    for r in interesting:

        print(
            f"{r['sector']:10s} "
            f"{r['sat']:3d} "
            f"{r['freq']:4d} "
            f"{r['n']:2d} "
            f"{r['az_mean']:6.1f} "
            f"{r['r']:+7.3f} "
            f"{r['slope']:+8.3f} "
            f"{r['rms']*100:7.2f} cm"
        )

else:

    print(
        "No negative-slope repeat tracks "
        "with sufficient observations."
    )


# ================================================================
# ALL TRACKS WITH >= 3 OBSERVATIONS
# ================================================================

print()
print("=" * 80)
print(
    "ALL REPEAT TRACKS — EOT20"
)
print("=" * 80)

eot = [

    r

    for r in all_results

    if r["model"]
    == "EOT20_heightm"

]

eot.sort(
    key=lambda r:
    r["slope"]
)

print()
print(
    "SECTOR      SAT FREQ N   AZ      r       "
    "SLOPE      RMS"
)

print("-" * 80)

for r in eot:

    print(
        f"{r['sector']:10s} "
        f"{r['sat']:3d} "
        f"{r['freq']:4d} "
        f"{r['n']:2d} "
        f"{r['az_mean']:7.2f} "
        f"{r['r']:+7.3f} "
        f"{r['slope']:+9.3f} "
        f"{r['rms']*100:7.2f} cm"
    )


# ================================================================
# WRITE CSV
# ================================================================

output = Path(
    "repeat_track_tide_slopes.csv"
)

fieldnames = [
    "sector",
    "sat",
    "freq",
    "n",
    "az_mean",
    "model",
    "r",
    "slope_m_per_m",
    "intercept_m",
    "rms_m",
]

with open(
    output,
    "w",
    newline="",
) as f:

    writer = csv.DictWriter(
        f,
        fieldnames=fieldnames,
    )

    writer.writeheader()

    for r in all_results:

        writer.writerow(
            {
                "sector": r["sector"],
                "sat": r["sat"],
                "freq": r["freq"],
                "n": r["n"],
                "az_mean": r["az_mean"],
                "model": r["model"],
                "r": r["r"],
                "slope_m_per_m": r[
                    "slope"
                ],
                "intercept_m": r[
                    "intercept"
                ],
                "rms_m": r["rms"],
            }
        )


# ================================================================
# DONE
# ================================================================

print()
print("=" * 80)
print("DONE")
print("=" * 80)

print()
print(
    f"Results written to: {output}"
)

print()
print(
    "Expected physical response:"
)

print(
    "  tide rises -> RH decreases"
)

print(
    "  expected RH/tide slope < 0"
)

print(
    "A slope near -1 would be especially "
    "interesting, but should not be expected "
    "a priori."
)
