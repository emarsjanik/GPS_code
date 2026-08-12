import csv
import math
import re
from pathlib import Path
from datetime import datetime, timedelta
from collections import defaultdict

import numpy as np
from openpyxl import load_workbook


# ================================================================
# CONFIGURATION
# ================================================================

DOY1 = 204
DOY2 = 207
YEAR = 2026

SUCCESS_DIR = Path(
    "products/refl_code/logs/usgs/ocean90_150/2026"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

CSV_FILE = Path(
    "gnssir_tide_arc_analysis.csv"
)

MATCH_LIMIT_SEC = 90.0

REFERENCE_HEIGHT = 18.665

AZ_SECTORS = {
    "ALL": (100.0, 150.0),
    "100-130": (100.0, 130.0),
    "110-120": (110.0, 120.0),
}


# ================================================================
# UTILITIES
# ================================================================

def finite(value):

    try:
        x = float(value)

        if math.isfinite(x):
            return x

    except (TypeError, ValueError):

        pass

    return None


def parse_datetime(value):

    if value is None:
        return None

    value = str(value).strip()

    if not value:
        return None

    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def pearson(x, y):

    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    valid = (
        np.isfinite(x)
        & np.isfinite(y)
    )

    x = x[valid]
    y = y[valid]

    if len(x) < 2:
        return float("nan")

    if np.std(x) == 0:
        return float("nan")

    if np.std(y) == 0:
        return float("nan")

    return float(
        np.corrcoef(x, y)[0, 1]
    )


# ================================================================
# TIDE MODEL
# ================================================================

def load_tide_model():

    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb.active

    header = [
        cell.value
        for cell in ws[1]
    ]

    required = [
        "time",
        "EOT20_heightm",
        "GOT5.5_heightm",
        "GOT5.6_heightm",
        "FES2022_heightm",
    ]

    columns = {}

    for name in required:

        if name not in header:

            raise RuntimeError(
                f"Missing tide-model column: {name}"
            )

        columns[name] = header.index(name)

    times = []
    model_values = defaultdict(list)

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):

        t = row[columns["time"]]

        if not isinstance(t, datetime):
            continue

        values = []

        valid = True

        for name in required[1:]:

            value = finite(
                row[columns[name]]
            )

            if value is None:

                valid = False
                break

            values.append(value)

        if not valid:
            continue

        times.append(t)

        for name, value in zip(
            required[1:],
            values,
        ):

            model_values[name].append(
                value
            )

    if len(times) < 2:

        raise RuntimeError(
            "Insufficient tide-model data."
        )

    matrix = np.array(
        [
            model_values[name]
            for name in required[1:]
        ],
        dtype=float,
    )

    ensemble = np.mean(
        matrix,
        axis=0,
    )

    epoch = np.array(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    def tide_at(query_time):

        q = (
            query_time - times[0]
        ).total_seconds()

        if q < epoch[0]:
            return float("nan")

        if q > epoch[-1]:
            return float("nan")

        return float(
            np.interp(
                q,
                epoch,
                ensemble,
            )
        )

    print()
    print("Tide model columns:")

    for name in required[1:]:
        print(f"  {name}")

    print(
        f"Tide points loaded: {len(times)}"
    )

    print(
        f"Tide coverage: "
        f"{times[0]} through {times[-1]}"
    )

    return tide_at


# ================================================================
# LOAD GNSS-IR ARC CSV
# ================================================================

def load_arc_csv():

    records = []

    with open(
        CSV_FILE,
        newline="",
    ) as f:

        reader = csv.DictReader(f)

        for row in reader:

            dt = parse_datetime(
                row.get(
                    "solution_time_utc"
                )
            )

            if dt is None:
                continue

            sat = finite(
                row.get("sat")
            )

            freq = finite(
                row.get("freq")
            )

            rh = finite(
                row.get("RH_m")
            )

            start = parse_datetime(
                row.get(
                    "time_start_utc"
                )
            )

            end = parse_datetime(
                row.get(
                    "time_end_utc"
                )
            )

            if (
                sat is None
                or freq is None
                or rh is None
                or start is None
                or end is None
            ):
                continue

            records.append(
                {
                    "dt": dt,
                    "sat": int(sat),
                    "freq": int(freq),
                    "rh": rh,
                    "start": start,
                    "end": end,
                    "az": finite(
                        row.get("Azim")
                    ),
                    "amp": finite(
                        row.get("Amp")
                    ),
                    "pkn": finite(
                        row.get("PkNoise")
                    ),
                }
            )

    print(
        f"CSV arc records loaded: "
        f"{len(records)}"
    )

    return records


# ================================================================
# LOAD SUCCESS LOGS
# ================================================================

SUCCESS_PATTERN = re.compile(
    r"SUCCESS\s+Azimuth\s+([0-9.]+)\s+"
    r"(?:Sat|Satellite)\s+(\d+)\s+"
    r"RH\s+([-+0-9.]+)\s+m\s+"
    r"PkNoise\s+([-+0-9.]+)\s+"
    r"Amp\s+([-+0-9.]+)\s+"
    r"Fr\s*(\d+)\s+"
    r"UTC\s+([0-9:]+)\s+"
    r"DT\s+([0-9.]+)"
)


def load_success_records():

    records = []

    for doy in range(
        DOY1,
        DOY2 + 1,
    ):

        path = (
            SUCCESS_DIR
            / f"{doy}_gnssir.txt"
        )

        if not path.exists():

            print(
                f"WARNING: missing {path}"
            )

            continue

        day = (
            datetime(YEAR, 1, 1)
            + timedelta(days=doy - 1)
        ).date()

        with open(path) as f:

            for line in f:

                m = SUCCESS_PATTERN.search(
                    line
                )

                if not m:
                    continue

                az = float(m.group(1))
                sat = int(m.group(2))
                rh = float(m.group(3))
                pkn = float(m.group(4))
                amp = float(m.group(5))
                freq = int(m.group(6))
                utc = m.group(7)
                dt_seconds = float(
                    m.group(8)
                )

                parts = utc.split(":")

                if len(parts) == 2:

                    hh = int(parts[0])
                    mm = int(parts[1])
                    ss = 0

                elif len(parts) == 3:

                    hh = int(parts[0])
                    mm = int(parts[1])
                    ss = int(float(parts[2]))

                else:

                    continue

                dt = datetime.combine(
                    day,
                    datetime.min.time(),
                ) + timedelta(
                    hours=hh,
                    minutes=mm,
                    seconds=ss,
                )

                records.append(
                    {
                        "dt": dt,
                        "doy": doy,
                        "utc": utc,
                        "az": az,
                        "sat": sat,
                        "freq": freq,
                        "rh": rh,
                        "pkn": pkn,
                        "amp": amp,
                        "dt_arc_min": dt_seconds,
                    }
                )

    return records


# ================================================================
# MATCH SUCCESS TO ARC CSV
# ================================================================

def match_records(
    success,
    arc_csv,
):

    matched = []
    unmatched = []

    for s in success:

        candidates = [

            r

            for r in arc_csv

            if (
                r["sat"] == s["sat"]
                and r["freq"] == s["freq"]
                and r["dt"].date()
                == s["dt"].date()
            )

        ]

        if not candidates:

            unmatched.append(s)
            continue

        best = min(
            candidates,
            key=lambda r:
            abs(
                (
                    r["dt"]
                    - s["dt"]
                ).total_seconds()
            ),
        )

        delta = abs(
            (
                best["dt"]
                - s["dt"]
            ).total_seconds()
        )

        if delta > MATCH_LIMIT_SEC:

            unmatched.append(s)
            continue

        r = dict(s)

        r["csv_dt"] = best["dt"]
        r["arc_start"] = best["start"]
        r["arc_end"] = best["end"]

        r["csv_az"] = best["az"]
        r["csv_amp"] = best["amp"]
        r["csv_pkn"] = best["pkn"]

        r["match_delta_sec"] = delta

        matched.append(r)

    return matched, unmatched


# ================================================================
# CALCULATE TIDE DURING EACH ARC
# ================================================================

def calculate_arc_values(
    records,
    tide_at,
):

    output = []

    for r in records:

        tide_start = tide_at(
            r["arc_start"]
        )

        tide_solution = tide_at(
            r["dt"]
        )

        tide_end = tide_at(
            r["arc_end"]
        )

        if not all(
            math.isfinite(x)
            for x in (
                tide_start,
                tide_solution,
                tide_end,
            )
        ):

            continue

        tide_change = (
            tide_end
            - tide_start
        )

        tide_start_to_solution = (
            tide_solution
            - tide_start
        )

        tide_solution_to_end = (
            tide_end
            - tide_solution
        )

        gnss_water = (
            REFERENCE_HEIGHT
            - r["rh"]
        )

        r = dict(r)

        r["tide_start"] = tide_start
        r["tide_solution"] = tide_solution
        r["tide_end"] = tide_end

        r["tide_change"] = tide_change
        r["tide_start_to_solution"] = (
            tide_start_to_solution
        )
        r["tide_solution_to_end"] = (
            tide_solution_to_end
        )

        r["gnss_water"] = gnss_water

        r["arc_duration_min"] = (
            (
                r["arc_end"]
                - r["arc_start"]
            ).total_seconds()
            / 60.0
        )

        output.append(r)

    return output


# ================================================================
# SAME-SATELLITE / SAME-FREQUENCY DIFFERENCE TEST
# ================================================================

def repeat_test(
    records,
    sector_name,
    az_lo,
    az_hi,
):

    subset = [

        r

        for r in records

        if (
            az_lo
            <= r["az"]
            <= az_hi
        )

    ]

    groups = defaultdict(list)

    for r in subset:

        groups[
            (
                r["sat"],
                r["freq"],
            )
        ].append(r)

    pairs = []

    for key, group in groups.items():

        group.sort(
            key=lambda x: x["dt"]
        )

        if len(group) < 2:
            continue

        for a, b in zip(
            group[:-1],
            group[1:],
        ):

            delta_rh = (
                b["rh"]
                - a["rh"]
            )

            delta_tide = (
                b["tide_solution"]
                - a["tide_solution"]
            )

            delta_gnss = (
                b["gnss_water"]
                - a["gnss_water"]
            )

            pairs.append(
                {
                    "sat": key[0],
                    "freq": key[1],
                    "t1": a["dt"],
                    "t2": b["dt"],
                    "az1": a["az"],
                    "az2": b["az"],
                    "rh1": a["rh"],
                    "rh2": b["rh"],
                    "tide1": a[
                        "tide_solution"
                    ],
                    "tide2": b[
                        "tide_solution"
                    ],
                    "delta_rh": delta_rh,
                    "delta_tide": delta_tide,
                    "delta_gnss": delta_gnss,
                    "hours": (
                        b["dt"]
                        - a["dt"]
                    ).total_seconds()
                    / 3600.0,
                }
            )

    print()
    print("=" * 80)
    print(
        f"REPEAT-SATELLITE TEST: "
        f"{sector_name}"
    )
    print("=" * 80)

    print(
        f"Records in sector: "
        f"{len(subset)}"
    )

    print(
        f"Repeated satellite/frequency pairs: "
        f"{len(pairs)}"
    )

    if not pairs:

        print(
            "No repeated satellite/frequency "
            "observations."
        )

        return pairs

    delta_rh = [
        p["delta_rh"]
        for p in pairs
    ]

    delta_tide = [
        p["delta_tide"]
        for p in pairs
    ]

    delta_gnss = [
        p["delta_gnss"]
        for p in pairs
    ]

    print()
    print(
        f"Delta RH vs delta tide: "
        f"{pearson(delta_rh, delta_tide):+.4f}"
    )

    print(
        f"Delta GNSS water vs delta tide: "
        f"{pearson(delta_gnss, delta_tide):+.4f}"
    )

    print()
    print(
        "SAT FREQ  TIME1       TIME2       "
        "dRH       dTIDE      hours"
    )

    print("-" * 80)

    for p in pairs:

        print(
            f"{p['sat']:3d} "
            f"{p['freq']:4d} "
            f"{p['t1'].strftime('%m-%d %H:%M')} "
            f"{p['t2'].strftime('%m-%d %H:%M')} "
            f"{p['delta_rh']:+8.3f} "
            f"{p['delta_tide']:+9.3f} "
            f"{p['hours']:6.1f}"
        )

    return pairs


# ================================================================
# MAIN
# ================================================================

print()
print("=" * 80)
print(
    "GNSS-IR INDIVIDUAL-ARC / REPEAT-SATELLITE "
    "TIDE RESPONSE TEST"
)
print("=" * 80)

print(
    f"DOY range       : {DOY1}-{DOY2}"
)

print(
    f"SUCCESS logs    : {SUCCESS_DIR}"
)

print(
    f"Tide workbook   : {TIDE_FILE}"
)

print(
    f"Arc CSV         : {CSV_FILE}"
)

print(
    f"Reference H     : {REFERENCE_HEIGHT:.3f} m"
)

print(
    f"Match tolerance : "
    f"{MATCH_LIMIT_SEC:.0f} sec"
)


# ------------------------------------------------
# Load data
# ------------------------------------------------

tide_at = load_tide_model()

arc_csv = load_arc_csv()

success = load_success_records()

print()
print(
    f"SUCCESS records : {len(success)}"
)

matched, unmatched = match_records(
    success,
    arc_csv,
)

print(
    f"Matched         : {len(matched)}"
)

print(
    f"Unmatched       : {len(unmatched)}"
)

records = calculate_arc_values(
    matched,
    tide_at,
)

print(
    f"Arc tide records: {len(records)}"
)


# ------------------------------------------------
# Individual arc information
# ------------------------------------------------

print()
print("=" * 80)
print("INDIVIDUAL ARC TIDE CONDITIONS")
print("=" * 80)

for r in sorted(
    records,
    key=lambda x: x["dt"],
):

    print(
        f"{r['dt'].strftime('%m-%d %H:%M:%S')} "
        f"Az={r['az']:6.1f} "
        f"sat={r['sat']:3d} "
        f"freq={r['freq']:3d} "
        f"RH={r['rh']:7.3f} "
        f"Tstart={r['tide_start']:+7.3f} "
        f"Tsol={r['tide_solution']:+7.3f} "
        f"Tend={r['tide_end']:+7.3f} "
        f"dT={r['tide_change']:+7.3f} "
        f"arc={r['arc_duration_min']:5.1f}m"
    )


# ------------------------------------------------
# Basic epoch correlations
# ------------------------------------------------

print()
print("=" * 80)
print("EPOCH CORRELATIONS")
print("=" * 80)

rh = np.array(
    [
        r["rh"]
        for r in records
    ]
)

tide_start = np.array(
    [
        r["tide_start"]
        for r in records
    ]
)

tide_solution = np.array(
    [
        r["tide_solution"]
        for r in records
    ]
)

tide_end = np.array(
    [
        r["tide_end"]
        for r in records
    ]
)

tide_change = np.array(
    [
        r["tide_change"]
        for r in records
    ]
)

gnss_water = np.array(
    [
        r["gnss_water"]
        for r in records
    ]
)

print(
    f"RH vs tide START     : "
    f"{pearson(rh, tide_start):+.4f}"
)

print(
    f"RH vs tide SOLUTION  : "
    f"{pearson(rh, tide_solution):+.4f}"
)

print(
    f"RH vs tide END       : "
    f"{pearson(rh, tide_end):+.4f}"
)

print(
    f"RH vs within-arc dT  : "
    f"{pearson(rh, tide_change):+.4f}"
)

print(
    f"GNSS water vs tide   : "
    f"{pearson(gnss_water, tide_solution):+.4f}"
)


# ------------------------------------------------
# Repeat satellite tests
# ------------------------------------------------

all_pairs = []

for name, (lo, hi) in AZ_SECTORS.items():

    pairs = repeat_test(
        records,
        name,
        lo,
        hi,
    )

    for p in pairs:

        p2 = dict(p)
        p2["sector"] = name
        all_pairs.append(p2)


# ------------------------------------------------
# Write detailed CSV
# ------------------------------------------------

out_csv = Path(
    "individual_arc_tide_response_results.csv"
)

fieldnames = [
    "doy",
    "dt",
    "sat",
    "freq",
    "az",
    "rh",
    "gnss_water",
    "arc_start",
    "arc_end",
    "arc_duration_min",
    "tide_start",
    "tide_solution",
    "tide_end",
    "tide_change",
    "tide_start_to_solution",
    "tide_solution_to_end",
    "match_delta_sec",
]

with open(
    out_csv,
    "w",
    newline="",
) as f:

    writer = csv.DictWriter(
        f,
        fieldnames=fieldnames,
    )

    writer.writeheader()

    for r in records:

        writer.writerow(
            {
                "doy": r["doy"],
                "dt": r["dt"].isoformat(),
                "sat": r["sat"],
                "freq": r["freq"],
                "az": r["az"],
                "rh": r["rh"],
                "gnss_water": r[
                    "gnss_water"
                ],
                "arc_start": r[
                    "arc_start"
                ].isoformat(),
                "arc_end": r[
                    "arc_end"
                ].isoformat(),
                "arc_duration_min": r[
                    "arc_duration_min"
                ],
                "tide_start": r[
                    "tide_start"
                ],
                "tide_solution": r[
                    "tide_solution"
                ],
                "tide_end": r[
                    "tide_end"
                ],
                "tide_change": r[
                    "tide_change"
                ],
                "tide_start_to_solution": r[
                    "tide_start_to_solution"
                ],
                "tide_solution_to_end": r[
                    "tide_solution_to_end"
                ],
                "match_delta_sec": r[
                    "match_delta_sec"
                ],
            }
        )


# ------------------------------------------------
# Summary
# ------------------------------------------------

print()
print("=" * 80)
print("INTERPRETATION")
print("=" * 80)

print(
    """
The individual-arc tide values above show how much
the modeled water level changed while each GNSS-IR
arc was being observed.

The repeat-satellite test is more important than simply
searching for the highest correlation among arbitrary
azimuth bins.

For a fixed satellite/frequency observed repeatedly,
an increasing tide should generally correspond to a
decreasing RH because:

    GNSS water level = H - RH

Therefore:

    dRH / dtide < 0

would be the expected physical sign.

The repeat test is still exploratory because the number
of repeated satellite/frequency observations is small.
"""
)

print()
print(
    f"Detailed results written to: {out_csv}"
)

print()
print("=" * 80)
print("DONE")
print("=" * 80)
