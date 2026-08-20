#!/usr/bin/env python3
"""
Marconi selected-track TRUE TOPOBATHY / FRESNEL validation.

This is the physical-validation stage for the experimental multi-track
GNSS-R product.

It uses:
  * the current GPS L1 GNSS-R result files
  * the current +0.242 m datum hypothesis
  * the published EOT20 model in marconi_tides_sherwood.xlsx
  * the USGS 1-m Marconi topobathy raster:
        2021022FA_Marconi_topobathy_1m.tif

Selected tracks are NOT hard-coded. They are reselected using the current
experimental-product criteria:

  N >= 10
  |corrected median GNSS-R - EOT20| <= 0.05 m
  corrected RMS <= 0.20 m
  tide correlation r >= 0.93

Fresnel mathematics
-------------------
This follows the current clean-start repository's geometry implementation,
which explicitly reproduces the gnssrefl Fresnel-zone equations and calls
get_wavelength(freq, sat). This avoids the GLONASS-frequency API issue that
was encountered earlier in the project.

For the current GPS L1 tracks:
  wavelength is obtained from gnssrefl.

For each selected observation and each representative elevation:
  5, 9, 13 degrees

we calculate two footprints:

  OBSERVED:
      RH_observed = GNSS-IR reflector height

  TIDE-PREDICTED:
      MSL_NAVD88 = -0.242 m
      water_surface_NAVD88 = EOT20 + MSL_NAVD88
      RH_predicted = H_ortho - water_surface_NAVD88
                   = H_ortho - EOT20 + 0.242 m

The topobathy raster is EPSG:6348 (NAD83(2011) / UTM zone 19N).

Wet classification:
  raster elevation <= water surface elevation

For each Fresnel ellipse we compute:
  total footprint area
  finite DEM-covered footprint area
  wet fraction
  dry fraction
  mean DEM elevation
  minimum/maximum DEM elevation
  center coordinate

A footprint is NOT declared "confirmed ocean" from one number alone.
The script reports:
  STRONG_WATER_FOOTPRINT
  WATER_DOMINANT
  MIXED_SHORELINE
  LAND_DOMINANT
  OUTSIDE_RASTER

using the 9-degree predicted footprint plus robustness across 5/9/13 degrees.

It produces:
  marconi_selected_topobathy_validation.csv
  marconi_selected_topobathy_track_summary.csv
  marconi_selected_topobathy_summary.txt
  marconi_selected_topobathy.kml
  marconi_selected_topobathy_plots/
      01_selected_tracks_topobathy_map.png
      02_predicted_wet_fraction_by_track.png
      03_observed_vs_predicted_wet_fraction.png
      04_selected_track_residuals_and_geometry.png

This is a geometry validation product only. It does not modify the
gnssrefl production JSON or the GNSS-IR processing.
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from xml.sax.saxutils import escape

import matplotlib.pyplot as plt
import numpy as np
import rasterio
from pyproj import Transformer
from openpyxl import load_workbook
from gnssrefl.gnss_frequencies import get_wavelength


# =====================================================================
# CONFIGURATION
# =====================================================================

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

TOPOBATHY_FILE = Path(
    "2021022FA_Marconi_topobathy_1m.tif"
)

H_ORTHO_M = 18.665

# Independent datum hypothesis already tested:
# MSL approximately -0.242 m NAVD88 at the site/region.
MSL_NAVD88_M = -0.242

PRIMARY_TIDE = "EOT20_heightm"

STATION_LAT = 41.8928243333
STATION_LON = -69.9633227139

AZ_TOL_DEG = 3.0
MIN_TRACK_N = 10

MAX_ABS_CORRECTED_MEDIAN_M = 0.05
MAX_CORRECTED_RMS_M = 0.20
MIN_TIDE_R = 0.93

# Representative elevation grid.  9 degrees is the center of the
# 5-13 degree production window; 5 and 13 test footprint robustness.
TEST_ELEVATIONS_DEG = [5.0, 9.0, 13.0]

# Wet-footprint classification thresholds.
WATER_THRESHOLD = 0.80
LAND_THRESHOLD = 0.20

# DEM values can contain NaN even though GeoTIFF nodata metadata is None.
FINITE_ONLY = True

OUT_CSV = Path(
    "marconi_selected_topobathy_validation.csv"
)

OUT_TRACKS = Path(
    "marconi_selected_topobathy_track_summary.csv"
)

OUT_SUMMARY = Path(
    "marconi_selected_topobathy_summary.txt"
)

OUT_KML = Path(
    "marconi_selected_topobathy.kml"
)

PLOT_DIR = Path(
    "marconi_selected_topobathy_plots"
)


# =====================================================================
# HELPERS
# =====================================================================

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def utc_datetime(year, doy, utc_hours):
    return (
        datetime(year, 1, 1)
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


def circular_az_diff(a, b):
    d = abs(a - b) % 360.0
    return min(d, 360.0 - d)


def correlation(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if len(x) < 3:
        return math.nan
    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan
    return float(np.corrcoef(x, y)[0, 1])


def rms(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(np.sqrt(np.mean(x ** 2)))


# =====================================================================
# TIDE
# =====================================================================

def load_tide():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]

    time_col = header.index("time")
    tide_col = header.index(PRIMARY_TIDE)

    times = []
    values = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_col]
        v = finite(row[tide_col])

        if not isinstance(t, datetime):
            continue
        if v is None:
            continue

        times.append(t)
        values.append(v)

    wb.close()

    return (
        np.asarray(
            times,
            dtype="datetime64[ms]",
        ),
        np.asarray(
            values,
            dtype=float,
        ),
    )


def tide_at(
    tide_times,
    tide_values,
    dt,
):
    x = tide_times.astype("int64")
    q = np.datetime64(
        dt,
        "ms",
    ).astype("int64")

    if q < x[0] or q > x[-1]:
        return math.nan

    return float(
        np.interp(
            q,
            x,
            tide_values,
        )
    )


# =====================================================================
# GNSS-R RESULTS
# =====================================================================

def load_results():
    rows = []

    for path in sorted(
        RESULT_DIR.glob("*.txt")
    ):
        try:
            int(path.stem)
        except Exception:
            continue

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if not line or line.startswith("%"):
                continue

            c = line.split()
            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc_hours = float(c[4])
                az = float(c[5])
                amp = float(c[6])
                emin = float(c[7])
                emax = float(c[8])
                nobs = int(float(c[9]))
                freq = int(float(c[10]))
                rise = int(float(c[11]))
                pkn = float(c[13])
                delt = float(c[14])
            except Exception:
                continue

            if freq != 1:
                continue

            dt = utc_datetime(
                year,
                doy,
                utc_hours,
            )

            raw_wl = H_ORTHO_M - rh
            corrected_wl = raw_wl - MSL_NAVD88_M

            rows.append({
                "datetime": dt,
                "year": year,
                "doy": doy,
                "sat": sat,
                "freq": freq,
                "rise": rise,
                "az": az,
                "RH_m": rh,
                "raw_wl_m": raw_wl,
                "corrected_wl_m": corrected_wl,
                "Amp": amp,
                "PkNoise": pkn,
                "NumbOf": nobs,
                "DelT_min": delt,
                "emin_deg": emin,
                "emax_deg": emax,
            })

    rows.sort(
        key=lambda r:
            r["datetime"]
    )

    return rows


def attach_tide(
    rows,
    tide_times,
    tide_values,
):
    for row in rows:
        row["EOT20_m"] = tide_at(
            tide_times,
            tide_values,
            row["datetime"],
        )

        row["raw_residual_m"] = (
            row["raw_wl_m"]
            - (
                row["EOT20_m"]
                + MSL_NAVD88_M
            )
        )

        row["corrected_residual_m"] = (
            row["corrected_wl_m"]
            - row["EOT20_m"]
        )

    return [
        row
        for row in rows
        if math.isfinite(row["EOT20_m"])
    ]


# =====================================================================
# TRACK SELECTION
# =====================================================================

def build_tracks(rows):
    base = defaultdict(list)

    for row in rows:
        base[
            (
                row["sat"],
                row["freq"],
                row["rise"],
            )
        ].append(row)

    tracks = []

    for _, group in base.items():
        group = sorted(
            group,
            key=lambda r:
                r["az"],
        )

        current = []
        previous_az = None

        for row in group:

            if (
                previous_az is None
                or circular_az_diff(
                    row["az"],
                    previous_az,
                ) <= AZ_TOL_DEG
            ):
                current.append(row)
            else:
                if len(current) >= MIN_TRACK_N:
                    tracks.append(current)
                current = [row]

            previous_az = row["az"]

        if len(current) >= MIN_TRACK_N:
            tracks.append(current)

    return tracks


def summarize_track(group):
    wl = np.asarray(
        [
            r["corrected_wl_m"]
            for r in group
        ],
        dtype=float,
    )

    tide = np.asarray(
        [
            r["EOT20_m"]
            for r in group
        ],
        dtype=float,
    )

    residual = wl - tide

    r = correlation(
        wl,
        tide,
    )

    slope = float(
        np.polyfit(
            tide,
            wl,
            1,
        )[0]
    )

    return {
        "sat": group[0]["sat"],
        "freq": group[0]["freq"],
        "rise": group[0]["rise"],
        "n": len(group),
        "n_days": len(
            {
                r0["doy"]
                for r0 in group
            }
        ),
        "az_mean_deg": float(
            np.mean(
                [
                    r0["az"]
                    for r0 in group
                ]
            )
        ),
        "az_sd_deg": float(
            np.std(
                [
                    r0["az"]
                    for r0 in group
                ]
            )
        ),
        "corrected_median_m": float(
            np.median(residual)
        ),
        "corrected_rms_m": rms(residual),
        "tide_r": r,
        "tide_slope": slope,
    }


def is_selected(summary):
    return (
        summary["n"] >= MIN_TRACK_N
        and abs(
            summary["corrected_median_m"]
        ) <= MAX_ABS_CORRECTED_MEDIAN_M
        and summary["corrected_rms_m"]
        <= MAX_CORRECTED_RMS_M
        and summary["tide_r"]
        >= MIN_TIDE_R
    )


# =====================================================================
# FRESNEL GEOMETRY
# =====================================================================

def wavelength_for_arc(freq, sat):
    return float(
        get_wavelength(
            int(freq),
            sat=int(sat),
        )
    )


def fresnel_zone_local(
    freq,
    sat,
    elevation_deg,
    reflector_height_m,
    azimuth_deg,
):
    """
    Current clean-start repository geometry implementation.

    Returns a local EN polygon centered at the receiver:
      x = East
      y = North
    """
    if reflector_height_m <= 0:
        raise ValueError(
            f"RH must be >0; got {reflector_height_m}"
        )

    wavelength = wavelength_for_arc(
        freq,
        sat,
    )

    delta = wavelength / 2.0
    e_rad = math.radians(
        elevation_deg
    )
    sin_e = math.sin(e_rad)

    if sin_e <= 0:
        raise ValueError(
            "Elevation must be positive."
        )

    B = math.sqrt(
        (
            2.0
            * delta
            * reflector_height_m
            / sin_e
        )
        + (
            delta
            / sin_e
        ) ** 2
    )

    A = B / sin_e

    center = (
        reflector_height_m
        + delta / sin_e
    ) / math.tan(e_rad)

    # Same orientation convention as gnssrefl.makeFresnelEllipse.
    angle = (
        360.0
        - float(azimuth_deg)
        + 90.0
    )

    rtheta = math.radians(angle)

    theta = np.deg2rad(
        np.arange(
            0.0,
            375.0,
            5.0,
        )
    )

    x0 = A * np.cos(theta)
    y0 = B * np.sin(theta)

    cos_r = math.cos(rtheta)
    sin_r = math.sin(rtheta)

    east = (
        cos_r * x0
        - sin_r * y0
    )

    north = (
        sin_r * x0
        + cos_r * y0
    )

    east += center * cos_r
    north += center * sin_r

    return {
        "wavelength_m": wavelength,
        "A_m": float(A),
        "B_m": float(B),
        "center_m": float(center),
        "east_m": east,
        "north_m": north,
    }


# =====================================================================
# TOPOBATHY SAMPLING
# =====================================================================

def raster_indices_from_xy(
    raster,
    eastings,
    northings,
):
    rows, cols = raster.index(
        eastings,
        northings,
    )
    return (
        np.asarray(rows),
        np.asarray(cols),
    )


def sample_dem_polygon(
    raster,
    polygon_east,
    polygon_north,
    water_surface_navd88_m,
):
    """
    Sample the 1-m DEM cells whose centers fall inside a Fresnel polygon.
    Returns wet/dry statistics relative to the supplied NAVD88 water surface.
    """
    from rasterio.windows import Window
    from matplotlib.path import Path as MplPath

    min_e = float(np.min(polygon_east))
    max_e = float(np.max(polygon_east))
    min_n = float(np.min(polygon_north))
    max_n = float(np.max(polygon_north))

    row_a, col_a = raster.index(min_e, max_n)
    row_b, col_b = raster.index(max_e, min_n)

    rmin = max(0, min(row_a, row_b) - 2)
    rmax = min(raster.height - 1, max(row_a, row_b) + 2)
    cmin = max(0, min(col_a, col_b) - 2)
    cmax = min(raster.width - 1, max(col_a, col_b) + 2)

    if rmin > rmax or cmin > cmax:
        return {
            "status": "OUTSIDE_RASTER",
            "coverage_fraction": 0.0,
            "wet_fraction": math.nan,
            "dry_fraction": math.nan,
            "mean_dem_m": math.nan,
            "min_dem_m": math.nan,
            "max_dem_m": math.nan,
            "n_cells": 0,
        }

    width = cmax - cmin + 1
    height = rmax - rmin + 1

    data = np.asarray(
        raster.read(
            1,
            window=Window(
                cmin,
                rmin,
                width,
                height,
            ),
        ),
        dtype=float,
    )

    if data.shape != (height, width):
        raise RuntimeError(
            f"Unexpected DEM window shape: got {data.shape}, "
            f"expected {(height, width)}"
        )

    rows_idx = np.arange(rmin, rmax + 1)
    cols_idx = np.arange(cmin, cmax + 1)
    cc, rr = np.meshgrid(cols_idx, rows_idx)

    xs, ys = raster.xy(
        rr,
        cc,
        offset="center",
    )

    xs = np.asarray(xs, dtype=float).reshape(data.shape)
    ys = np.asarray(ys, dtype=float).reshape(data.shape)

    polygon = np.column_stack(
        [
            np.asarray(polygon_east, dtype=float),
            np.asarray(polygon_north, dtype=float),
        ]
    )

    inside_flat = MplPath(polygon).contains_points(
        np.column_stack(
            [
                xs.ravel(),
                ys.ravel(),
            ]
        )
    )

    expected = data.size

    if inside_flat.size != expected:
        raise RuntimeError(
            "Point-in-polygon result size does not match DEM window: "
            f"points={inside_flat.size}, DEM={data.shape}"
        )

    inside = np.asarray(
        inside_flat,
        dtype=bool,
    ).reshape(data.shape)

    footprint_mask = inside & np.isfinite(data)
    n_cells = int(np.sum(footprint_mask))

    if n_cells == 0:
        return {
            "status": "OUTSIDE_RASTER",
            "coverage_fraction": 0.0,
            "wet_fraction": math.nan,
            "dry_fraction": math.nan,
            "mean_dem_m": math.nan,
            "min_dem_m": math.nan,
            "max_dem_m": math.nan,
            "n_cells": 0,
        }

    dem = data[footprint_mask]
    wet = dem <= water_surface_navd88_m
    wet_fraction = float(np.mean(wet))

    footprint_area = polygon_area(
        np.asarray(polygon_east, dtype=float),
        np.asarray(polygon_north, dtype=float),
    )
    cell_area = abs(
        float(raster.res[0])
        * float(raster.res[1])
    )
    covered_area = n_cells * cell_area

    coverage_fraction = float(
        min(
            1.0,
            max(
                0.0,
                covered_area / footprint_area
                if footprint_area > 0
                else 0.0,
            ),
        )
    )

    return {
        "status": "VALID",
        "coverage_fraction": coverage_fraction,
        "wet_fraction": wet_fraction,
        "dry_fraction": 1.0 - wet_fraction,
        "mean_dem_m": float(np.mean(dem)),
        "min_dem_m": float(np.min(dem)),
        "max_dem_m": float(np.max(dem)),
        "n_cells": n_cells,
    }



# =====================================================================
# CLASSIFICATION
# =====================================================================

def classify_robust(
    fractions,
):
    """
    Use the 5/9/13 degree predicted wet fractions.

    STRONG_WATER:
      9° >=80% and all 3 elevations >=70%

    WATER_DOMINANT:
      9° >=80%

    LAND_DOMINANT:
      9° <=20%

    MIXED_SHORELINE:
      otherwise
    """
    f5 = fractions.get(5.0, math.nan)
    f9 = fractions.get(9.0, math.nan)
    f13 = fractions.get(13.0, math.nan)

    if not math.isfinite(f9):
        return "OUTSIDE_RASTER"

    if (
        f9 >= WATER_THRESHOLD
        and all(
            math.isfinite(v)
            and v >= 0.70
            for v in [f5, f9, f13]
        )
    ):
        return "STRONG_WATER_FOOTPRINT"

    if f9 >= WATER_THRESHOLD:
        return "WATER_DOMINANT"

    if f9 <= LAND_THRESHOLD:
        return "LAND_DOMINANT"

    return "MIXED_SHORELINE"


# =====================================================================
# PROCESS
# =====================================================================

def process_observation(
    row,
    raster,
    station_e,
    station_n,
):
    tide = row["EOT20_m"]

    # Physical water surface in DEM/NAVD88:
    # EOT20 + MSL_NAVD88.
    predicted_water_surface = (
        tide
        + MSL_NAVD88_M
    )

    # Equivalent reflector height expected at the GNSS-R antenna.
    predicted_rh = (
        H_ORTHO_M
        - predicted_water_surface
    )

    result = {
        "datetime": row["datetime"].isoformat(),
        "sat": row["sat"],
        "freq": row["freq"],
        "rise": row["rise"],
        "az_deg": row["az"],
        "RH_observed_m": row["RH_m"],
        "EOT20_m": tide,
        "water_surface_NAVD88_m":
            predicted_water_surface,
        "RH_predicted_m":
            predicted_rh,
    }

    observed_fractions = {}
    predicted_fractions = {}

    for elevation in TEST_ELEVATIONS_DEG:

        obs = fresnel_zone_local(
            row["freq"],
            row["sat"],
            elevation,
            row["RH_m"],
            row["az"],
        )

        pred = fresnel_zone_local(
            row["freq"],
            row["sat"],
            elevation,
            predicted_rh,
            row["az"],
        )

        # Fresnel east/north coordinates are LOCAL OFFSETS from the
        # GNSS-R station. Convert them to absolute raster CRS coordinates
        # before sampling the DEM.
        obs_abs_east = station_e + obs["east_m"]
        obs_abs_north = station_n + obs["north_m"]

        pred_abs_east = station_e + pred["east_m"]
        pred_abs_north = station_n + pred["north_m"]

        obs_metrics = sample_dem_polygon(
            raster,
            obs_abs_east,
            obs_abs_north,
            predicted_water_surface,
        )

        pred_metrics = sample_dem_polygon(
            raster,
            pred_abs_east,
            pred_abs_north,
            predicted_water_surface,
        )

        tag = str(int(elevation))

        result[
            f"obs_{tag}_wet_fraction"
        ] = obs_metrics["wet_fraction"]

        result[
            f"obs_{tag}_coverage_fraction"
        ] = obs_metrics["coverage_fraction"]

        result[
            f"pred_{tag}_wet_fraction"
        ] = pred_metrics["wet_fraction"]

        result[
            f"pred_{tag}_coverage_fraction"
        ] = pred_metrics["coverage_fraction"]

        result[
            f"pred_{tag}_A_m"
        ] = pred["A_m"]

        result[
            f"pred_{tag}_B_m"
        ] = pred["B_m"]

        result[
            f"pred_{tag}_center_m"
        ] = pred["center_m"]

        result[
            f"pred_{tag}_mean_dem_m"
        ] = pred_metrics["mean_dem_m"]

        result[
            f"pred_{tag}_min_dem_m"
        ] = pred_metrics["min_dem_m"]

        result[
            f"pred_{tag}_max_dem_m"
        ] = pred_metrics["max_dem_m"]

        observed_fractions[
            elevation
        ] = obs_metrics[
            "wet_fraction"
        ]

        predicted_fractions[
            elevation
        ] = pred_metrics[
            "wet_fraction"
        ]

    result["obs_9_class"] = classify_robust(
        observed_fractions
    )

    result["pred_9_class"] = classify_robust(
        predicted_fractions
    )

    return result


# =====================================================================
# KML
# =====================================================================

def local_to_lonlat(
    transformer,
    east,
    north,
):
    lon, lat = transformer.transform(
        east,
        north,
    )
    return (
        float(lon),
        float(lat),
    )


def ellipse_kml_coords(
    station_e,
    station_n,
    east,
    north,
    to_wgs84,
):
    coords = []

    for e, n in zip(
        east,
        north,
    ):
        lon, lat = local_to_lonlat(
            to_wgs84,
            station_e + e,
            station_n + n,
        )

        coords.append(
            f"{lon:.7f},{lat:.7f},0"
        )

    return " ".join(coords)


# =====================================================================
# PLOTS
# =====================================================================

def plot_map(
    raster,
    station_e,
    station_n,
    selected_rows,
    geometry_results,
    to_wgs84,
):
    # Show DEM in the native map projection.  Since it is a 1-m raster,
    # use a modest downsampling for plotting.
    data = raster.read(1)

    finite_data = data[
        np.isfinite(data)
    ]

    vmin = (
        np.percentile(
            finite_data,
            2,
        )
        if len(finite_data)
        else -2
    )

    vmax = (
        np.percentile(
            finite_data,
            98,
        )
        if len(finite_data)
        else 2
    )

    fig, ax = plt.subplots(
        figsize=(12, 10)
    )

    extent = [
        raster.bounds.left,
        raster.bounds.right,
        raster.bounds.bottom,
        raster.bounds.top,
    ]

    ax.imshow(
        data,
        extent=extent,
        origin="upper",
        vmin=vmin,
        vmax=vmax,
        aspect="equal",
    )

    groups = {}

    # Only plot one representative footprint per selected track:
    # the first available observation at 9° predicted geometry.
    for row in selected_rows:
        key = (
            row["sat"],
            row["freq"],
            row["rise"],
        )

        if key in groups:
            continue

        geom = geometry_results[
            row["datetime"].isoformat()
        ]

        pred = geom["pred_9_geometry"]

        groups[key] = (
            row,
            pred,
        )

    for key, (row, pred) in groups.items():

        # Convert local EN to raster/map coordinates.
        east = (
            station_e
            + pred["east_m"]
        )

        north = (
            station_n
            + pred["north_m"]
        )

        ax.plot(
            east,
            north,
            linewidth=2,
            label=(
                f"PRN {row['sat']} "
                f"{'R' if row['rise'] == 1 else 'S'} "
                f"Az {row['az']:.1f}°"
            ),
        )

    ax.scatter(
        station_e,
        station_n,
        s=60,
        marker="x",
        label="GNSS-R station",
    )

    ax.set_title(
        "Marconi Selected GNSS-R 9° Predicted Fresnel Footprints"
    )
    ax.set_xlabel(
        "Easting (m)"
    )
    ax.set_ylabel(
        "Northing (m)"
    )
    ax.grid(
        True,
        alpha=0.2,
    )
    ax.legend(
        fontsize=8,
    )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "01_selected_tracks_topobathy_map.png",
        dpi=200,
    )

    plt.close(fig)


def plot_wet_fraction(
    track_summaries,
):
    labels = []
    values = []

    for item in track_summaries:

        labels.append(
            f"PRN {item['sat']} "
            f"{'R' if item['rise']==1 else 'S'} "
            f"{item['az_mean_deg']:.1f}°"
        )

        values.append(
            item["pred_9_wet_fraction"]
        )

    y = np.arange(
        len(labels)
    )

    fig, ax = plt.subplots(
        figsize=(10, 5)
    )

    ax.barh(
        y,
        values,
    )

    ax.axvline(
        WATER_THRESHOLD,
        linestyle="--",
        linewidth=1.5,
        label="80% water threshold",
    )

    ax.axvline(
        LAND_THRESHOLD,
        linestyle="--",
        linewidth=1.5,
        label="20% water threshold",
    )

    ax.set_yticks(y)
    ax.set_yticklabels(labels)

    ax.set_xlim(
        0,
        1,
    )

    ax.set_xlabel(
        "9° predicted Fresnel wet fraction"
    )

    ax.set_title(
        "Selected Track Predicted Water Footprint"
    )

    ax.grid(
        True,
        axis="x",
        alpha=0.2,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "02_predicted_wet_fraction_by_track.png",
        dpi=200,
    )

    plt.close(fig)


def plot_observed_predicted(
    track_summaries,
):
    labels = []
    observed = []
    predicted = []

    for item in track_summaries:

        labels.append(
            f"PRN {item['sat']} "
            f"{'R' if item['rise']==1 else 'S'} "
            f"{item['az_mean_deg']:.1f}°"
        )

        observed.append(
            item["obs_9_wet_fraction"]
        )

        predicted.append(
            item["pred_9_wet_fraction"]
        )

    y = np.arange(
        len(labels)
    )

    fig, ax = plt.subplots(
        figsize=(10, 6)
    )

    ax.plot(
        observed,
        y,
        "o-",
        label="Observed-RH footprint",
    )

    ax.plot(
        predicted,
        y,
        "o-",
        label="Tide-predicted-RH footprint",
    )

    ax.set_yticks(y)
    ax.set_yticklabels(labels)

    ax.set_xlim(
        0,
        1,
    )

    ax.set_xlabel(
        "Wet fraction"
    )

    ax.set_title(
        "Observed vs Tide-Predicted Fresnel Wet Fraction"
    )

    ax.grid(
        True,
        axis="x",
        alpha=0.2,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "03_observed_vs_predicted_wet_fraction.png",
        dpi=200,
    )

    plt.close(fig)


def plot_residual_geometry(
    selected_rows,
    geometry_results,
):
    residual = []
    wet = []
    labels = []

    for row in selected_rows:

        result = geometry_results[
            row["datetime"].isoformat()
        ]

        if not math.isfinite(
            row["corrected_residual_m"]
        ):
            continue

        if not math.isfinite(
            result["pred_9_wet_fraction"]
        ):
            continue

        residual.append(
            100
            * row["corrected_residual_m"]
        )

        wet.append(
            result["pred_9_wet_fraction"]
        )

        labels.append(
            row["sat"]
        )

    fig, ax = plt.subplots(
        figsize=(9, 7)
    )

    ax.scatter(
        wet,
        residual,
        s=20,
        alpha=0.65,
    )

    ax.axhline(
        0,
        linewidth=1.3,
    )

    ax.axvline(
        WATER_THRESHOLD,
        linestyle="--",
        linewidth=1.3,
    )

    ax.set_xlabel(
        "Predicted 9° Fresnel wet fraction"
    )

    ax.set_ylabel(
        "Corrected GNSS-R − EOT20 (cm)"
    )

    ax.set_title(
        "GNSS-R Residual vs Physical Water-Footprint Fraction"
    )

    ax.grid(
        True,
        alpha=0.2,
    )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "04_selected_track_residuals_and_geometry.png",
        dpi=200,
    )

    plt.close(fig)


# =====================================================================
# MAIN
# =====================================================================

def main():
    print()
    print("=" * 96)
    print(
        "MARCONI TRUE TOPOBATHY / FRESNEL VALIDATION OF SELECTED TRACKS"
    )
    print("=" * 96)

    print(
        "Topobathy:",
        TOPOBATHY_FILE,
    )

    print(
        "Datum / MSL NAVD88 hypothesis:",
        f"{MSL_NAVD88_M:+.3f} m",
    )

    if not TOPOBATHY_FILE.exists():
        raise SystemExit(
            f"Missing {TOPOBATHY_FILE}"
        )

    PLOT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    tide_times, tide_values = load_tide()
    rows = load_results()

    rows = attach_tide(
        rows,
        tide_times,
        tide_values,
    )

    track_groups = build_tracks(
        rows
    )

    summaries = [
        summarize_track(group)
        for group in track_groups
    ]

    summaries = [
        summary
        for summary in summaries
        if is_selected(summary)
    ]

    print(
        "Selected statistical tracks:",
        len(summaries),
    )

    for summary in summaries:
        print(
            f"  PRN {summary['sat']:2d} "
            f"{'RISING' if summary['rise']==1 else 'SETTING':7s} "
            f"Az={summary['az_mean_deg']:7.2f} "
            f"N={summary['n']:2d} "
            f"r={summary['tide_r']:+.4f} "
            f"RMS={summary['corrected_rms_m']*100:.1f} cm"
        )

    # Map station into raster CRS.
    with rasterio.open(
        TOPOBATHY_FILE
    ) as raster:

        print()
        print(
            "Topobathy CRS:",
            raster.crs,
        )
        print(
            "Resolution:",
            raster.res,
        )
        print(
            "Bounds:",
            raster.bounds,
        )

        to_raster = Transformer.from_crs(
            "EPSG:4326",
            raster.crs,
            always_xy=True,
        )

        station_e, station_n = to_raster.transform(
            STATION_LON,
            STATION_LAT,
        )

        print(
            "Station in raster CRS:",
            f"E={station_e:.3f}",
            f"N={station_n:.3f}",
        )
        print(
            "IMPORTANT: Fresnel polygons will be sampled after converting",
            "local EN offsets to absolute raster CRS coordinates.",
        )

        # Verify station is covered.
        if not (
            raster.bounds.left
            <= station_e
            <= raster.bounds.right
            and
            raster.bounds.bottom
            <= station_n
            <= raster.bounds.top
        ):
            raise RuntimeError(
                "Station is outside topobathy raster bounds."
            )

        # WGS84 export transformer. The raster CRS is NAD83(2011),
        # but for this short local KML map, the standard geodetic transform
        # is sufficient.
        to_wgs84 = Transformer.from_crs(
            raster.crs,
            "EPSG:4326",
            always_xy=True,
        )

        geometry_rows = []
        geometry_lookup = {}

        # Need the actual group membership for selected tracks.
        selected_groups = []

        for group in track_groups:
            summary = summarize_track(group)

            if is_selected(summary):
                selected_groups.append(group)

        selected_rows = [
            row
            for group in selected_groups
            for row in group
        ]

        # Process each selected observation.
        for row in selected_rows:

            result = process_observation(
                row,
                raster,
                station_e,
                station_n,
            )

            result[
                "corrected_residual_m"
            ] = row[
                "corrected_residual_m"
            ]

            result[
                "track_key"
            ] = (
                row["sat"],
                row["freq"],
                row["rise"],
            )

            # Save geometry arrays internally for representative map plots.
            pred9 = fresnel_zone_local(
                row["freq"],
                row["sat"],
                9.0,
                (
                    H_ORTHO_M
                    - (
                        row["EOT20_m"]
                        + MSL_NAVD88_M
                    )
                ),
                row["az"],
            )

            result[
                "pred_9_geometry"
            ] = pred9

            geometry_lookup[
                row["datetime"].isoformat()
            ] = result

            geometry_rows.append(
                result
            )

        # Representative per-track summaries.
        track_summary_records = []

        for group in selected_groups:

            summary = summarize_track(
                group
            )

            # Find all geometry rows for this track.
            rows_for_track = [
                r
                for r in geometry_rows
                if (
                    r["sat"]
                    == summary["sat"]
                    and
                    r["freq"]
                    == summary["freq"]
                    and
                    r["rise"]
                    == summary["rise"]
                    and
                    circular_az_diff(
                        r["az_deg"],
                        summary["az_mean_deg"],
                    ) <= AZ_TOL_DEG
                )
            ]

            obs9 = np.asarray(
                [
                    r["obs_9_wet_fraction"]
                    for r in rows_for_track
                    if math.isfinite(
                        r["obs_9_wet_fraction"]
                    )
                ],
                dtype=float,
            )

            pred5 = np.asarray(
                [
                    r["pred_5_wet_fraction"]
                    for r in rows_for_track
                    if math.isfinite(
                        r["pred_5_wet_fraction"]
                    )
                ],
                dtype=float,
            )

            pred9 = np.asarray(
                [
                    r["pred_9_wet_fraction"]
                    for r in rows_for_track
                    if math.isfinite(
                        r["pred_9_wet_fraction"]
                    )
                ],
                dtype=float,
            )

            pred13 = np.asarray(
                [
                    r["pred_13_wet_fraction"]
                    for r in rows_for_track
                    if math.isfinite(
                        r["pred_13_wet_fraction"]
                    )
                ],
                dtype=float,
            )

            summary = dict(summary)

            summary[
                "obs_9_wet_fraction"
            ] = (
                float(np.mean(obs9))
                if len(obs9)
                else math.nan
            )

            summary[
                "pred_5_wet_fraction"
            ] = (
                float(np.mean(pred5))
                if len(pred5)
                else math.nan
            )

            summary[
                "pred_9_wet_fraction"
            ] = (
                float(np.mean(pred9))
                if len(pred9)
                else math.nan
            )

            summary[
                "pred_13_wet_fraction"
            ] = (
                float(np.mean(pred13))
                if len(pred13)
                else math.nan
            )

            # Track-level physical class.
            f9 = summary[
                "pred_9_wet_fraction"
            ]

            if not math.isfinite(f9):
                phys_class = (
                    "OUTSIDE_RASTER"
                )
            elif f9 >= WATER_THRESHOLD:
                phys_class = (
                    "WATER_DOMINANT"
                )
            elif f9 <= LAND_THRESHOLD:
                phys_class = (
                    "LAND_DOMINANT"
                )
            else:
                phys_class = (
                    "MIXED_SHORELINE"
                )

            summary[
                "geometry_class"
            ] = phys_class

            track_summary_records.append(
                summary
            )

        # Write geometry CSV.
        if geometry_rows:
            fields = list(
                geometry_rows[0].keys()
            )

            # Geometry arrays cannot be serialized.
            fields = [
                f
                for f in fields
                if f
                not in {
                    "pred_9_geometry",
                    "track_key",
                }
            ]

            with open(
                OUT_CSV,
                "w",
                newline="",
            ) as f:

                writer = csv.DictWriter(
                    f,
                    fieldnames=fields,
                )

                writer.writeheader()

                for row in geometry_rows:
                    output = {
                        f: row.get(
                            f,
                            "",
                        )
                        for f in fields
                    }
                    writer.writerow(output)

        # Track summary CSV.
        if track_summary_records:

            fields = list(
                track_summary_records[0].keys()
            )

            with open(
                OUT_TRACKS,
                "w",
                newline="",
            ) as f:

                writer = csv.DictWriter(
                    f,
                    fieldnames=fields,
                )

                writer.writeheader()
                writer.writerows(
                    track_summary_records
                )

        # KML.
        with open(
            OUT_KML,
            "w",
        ) as f:

            f.write(
                '<?xml version="1.0" encoding="UTF-8"?>\n'
            )
            f.write(
                '<kml xmlns="http://www.opengis.net/kml/2.2">\n'
            )
            f.write("<Document>\n")
            f.write(
                "<name>Marconi selected GNSS-R topobathy validation</name>\n"
            )

            for group in selected_groups:

                row = group[0]

                pred9 = fresnel_zone_local(
                    row["freq"],
                    row["sat"],
                    9.0,
                    (
                        H_ORTHO_M
                        - (
                            row["EOT20_m"]
                            + MSL_NAVD88_M
                        )
                    ),
                    row["az"],
                )

                coords = (
                    ellipse_kml_coords(
                        station_e,
                        station_n,
                        pred9["east_m"],
                        pred9["north_m"],
                        to_wgs84,
                    )
                )

                label = (
                    f"PRN {row['sat']} "
                    f"{'Rising' if row['rise']==1 else 'Setting'} "
                    f"Az {row['az']:.1f}°"
                )

                f.write(
                    "<Placemark>\n"
                )
                f.write(
                    f"<name>{escape(label)}</name>\n"
                )
                f.write(
                    "<Style><LineStyle>"
                    "<width>3</width>"
                    "</LineStyle></Style>\n"
                )
                f.write(
                    "<Polygon><outerBoundaryIs>"
                    "<LinearRing><coordinates>"
                    f"{coords}"
                    "</coordinates></LinearRing>"
                    "</outerBoundaryIs></Polygon>\n"
                )
                f.write(
                    "</Placemark>\n"
                )

            lon, lat = to_wgs84.transform(
                station_e,
                station_n,
            )

            f.write(
                "<Placemark>\n"
            )
            f.write(
                "<name>GNSS-R station</name>\n"
            )
            f.write(
                "<Point><coordinates>"
                f"{lon:.7f},{lat:.7f},0"
                "</coordinates></Point>\n"
            )
            f.write(
                "</Placemark>\n"
            )

            f.write(
                "</Document></kml>\n"
            )

        # Summary text.
        lines = [
            "MARCONI TRUE TOPOBATHY / FRESNEL VALIDATION",
            "=" * 96,
            "",
            f"Topobathy: {TOPOBATHY_FILE}",
            f"CRS: {raster.crs}",
            f"Station latitude: {STATION_LAT}",
            f"Station longitude: {STATION_LON}",
            f"H_ortho: {H_ORTHO_M:.3f} m",
            f"MSL NAVD88 hypothesis: {MSL_NAVD88_M:+.3f} m",
            f"Primary tide model: {PRIMARY_TIDE}",
            "",
            "The tide-predicted water surface is:",
            "    water_surface_NAVD88 = EOT20 + MSL_NAVD88",
            "",
            "The equivalent tide-predicted reflector height is:",
            "    RH_predicted = H_ortho - water_surface_NAVD88",
            "",
            "SELECTED TRACK GEOMETRY",
            "-" * 96,
        ]

        for item in track_summary_records:
            lines.append(
                f"PRN {item['sat']:2d} "
                f"{'RISING' if item['rise']==1 else 'SETTING':7s} "
                f"Az={item['az_mean_deg']:7.2f} "
                f"N={item['n']:2d} "
                f"r={item['tide_r']:+.4f} "
                f"RMS={item['corrected_rms_m']*100:.1f} cm "
                f"bias={item['corrected_median_m']*100:+.1f} cm "
                f"obs9wet={item['obs_9_wet_fraction']:.3f} "
                f"pred5wet={item['pred_5_wet_fraction']:.3f} "
                f"pred9wet={item['pred_9_wet_fraction']:.3f} "
                f"pred13wet={item['pred_13_wet_fraction']:.3f} "
                f"class={item['geometry_class']}"
            )

        lines += [
            "",
            "INTERPRETATION",
            "-" * 96,
            "This is a physical footprint validation, not a statistical",
            "tide-validation score.",
            "",
            "A track is not called confirmed ocean solely because the",
            "GNSS-R time series follows EOT20. The independent topobathy",
            "footprint must also be water-dominated.",
            "",
            "The current +0.242 m datum hypothesis is used only to derive",
            "the tide-predicted NAVD88 water surface for geometry.",
            "",
            "The raster is EPSG:6348 and is sampled at its 1-m cell centers.",
        ]

        OUT_SUMMARY.write_text(
            "\n".join(lines)
            + "\n"
        )

        # Plots.
        plot_map(
            raster,
            station_e,
            station_n,
            selected_rows,
            geometry_lookup,
            to_wgs84,
        )

        plot_wet_fraction(
            track_summary_records
        )

        plot_observed_predicted(
            track_summary_records
        )

        plot_residual_geometry(
            selected_rows,
            geometry_lookup,
        )

    print()
    print("=" * 96)
    print("OUTPUTS")
    print("=" * 96)
    print(
        "Observation geometry CSV:",
        OUT_CSV.resolve(),
    )
    print(
        "Track summary CSV:",
        OUT_TRACKS.resolve(),
    )
    print(
        "Summary:",
        OUT_SUMMARY.resolve(),
    )
    print(
        "KML:",
        OUT_KML.resolve(),
    )
    print(
        "Plots:",
        PLOT_DIR.resolve(),
    )

    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
