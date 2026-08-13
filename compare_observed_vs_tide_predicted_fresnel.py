#!/usr/bin/env python3
"""
compare_observed_vs_tide_predicted_fresnel.py

Geometry diagnostic for Marconi GNSS-IR.

For each arc:
  OBSERVED geometry:
      reflector height = retrieved RH_m

  TIDE-PREDICTED geometry:
      reflector height = H_reference - tide_model

The key fix in this version is that it DOES NOT call
gnssrefl.refl_zones.FresnelZone() because that function calls
get_wavelength(f) without the satellite number and therefore fails
for GLONASS frequency codes 101+ in this installed gnssrefl version.

Instead, we reproduce the FresnelZone equations directly and call:
    get_wavelength(freq, sat)
which is the working API in this installation.

Outputs:
  observed_vs_tide_predicted_geometry.csv
  observed_vs_tide_predicted_geometry_summary.txt
"""

from __future__ import annotations

import csv
import math
from datetime import datetime
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from gnssrefl.gnss_frequencies import get_wavelength


CSV_FILE = Path("gnssir_tide_arc_analysis.csv")
TIDE_FILE = Path("marconi_tides_sherwood.xlsx")

REFERENCE_H_M = 18.665

WATER_BEARING_DEG = 83.06
WATER_DISTANCE_M = 71.78

ELEVATIONS = [5.0, 10.0, 15.0]

OPEN_THRESHOLD = 0.80
LAND_THRESHOLD = 0.20

OUT_CSV = Path("observed_vs_tide_predicted_geometry.csv")
OUT_SUMMARY = Path(
    "observed_vs_tide_predicted_geometry_summary.txt"
)

EARTH_M = 6371000.0
LAT = 41.8928243333
LON = -69.9633227139


def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def get_wavelength_for_arc(freq, sat):
    """
    Use the exact working gnssrefl API in this installation.

    GPS and fixed-frequency systems ignore sat internally.
    GLONASS codes require sat.
    """
    return float(
        get_wavelength(
            int(freq),
            sat=int(sat),
        )
    )


def fresnel_zone_local(freq, sat, elevation_deg, h_m, azimuth_deg):
    """
    Reproduce the FresnelZone + makeFresnelEllipse mathematics used
    by gnssrefl, but use get_wavelength(freq, sat) explicitly.
    """

    if h_m <= 0:
        raise ValueError(
            f"Reflector height must be > 0 for this geometry test; "
            f"got {h_m}"
        )

    wavelength = get_wavelength_for_arc(
        freq,
        sat,
    )

    delta = wavelength / 2.0
    e_rad = math.radians(float(elevation_deg))
    sin_e = math.sin(e_rad)

    if sin_e <= 0:
        raise ValueError("Invalid elevation angle.")

    B = math.sqrt(
        (2.0 * delta * h_m / sin_e)
        + (delta / sin_e) ** 2
    )

    A = B / sin_e

    center = (
        h_m + delta / sin_e
    ) / math.tan(e_rad)

    # Exact orientation convention from gnssrefl.makeFresnelEllipse().
    angle = 360.0 - float(azimuth_deg) + 90.0
    rtheta = math.radians(angle)

    theta = np.deg2rad(
        np.arange(0.0, 375.0, 15.0)
    )

    x = A * np.cos(theta)
    y = B * np.sin(theta)

    R = np.array(
        [
            [math.cos(rtheta), -math.sin(rtheta)],
            [math.sin(rtheta), math.cos(rtheta)],
        ]
    )

    x, y = np.dot(
        R,
        np.array([x, y]),
    )

    xcenter = center * math.cos(rtheta)
    ycenter = center * math.sin(rtheta)

    x += xcenter
    y += ycenter

    poly = [
        (float(xx), float(yy))
        for xx, yy in zip(x, y)
    ]

    return {
        "wavelength_m": wavelength,
        "A_m": float(A),
        "B_m": float(B),
        "center_m": float(center),
        "poly": poly,
    }


def shoreline_signed_coordinate(east, north):
    """
    Positive means on the water side of the reference shoreline plane.

    The plane is perpendicular to the 83.06-degree station-to-water
    bearing and anchored at 71.78 m from the station.
    """
    b = math.radians(WATER_BEARING_DEG)

    along = (
        east * math.sin(b)
        + north * math.cos(b)
    )

    return along - WATER_DISTANCE_M


def polygon_area(poly):
    if len(poly) < 3:
        return 0.0

    area = 0.0

    for i in range(len(poly)):
        x1, y1 = poly[i]
        x2, y2 = poly[(i + 1) % len(poly)]
        area += x1 * y2 - x2 * y1

    return abs(area) * 0.5


def clip_to_water(poly):
    """
    Sutherland-Hodgman clipping against the water-side half plane:
        shoreline_signed_coordinate >= 0
    """
    if not poly:
        return []

    out = []

    def signed(p):
        return shoreline_signed_coordinate(
            p[0], p[1]
        )

    for i, current in enumerate(poly):
        previous = poly[i - 1]

        sc = signed(current)
        sp = signed(previous)

        current_inside = sc >= 0.0
        previous_inside = sp >= 0.0

        if current_inside and previous_inside:
            out.append(current)

        elif previous_inside and not current_inside:
            denom = sp - sc
            if abs(denom) > 1e-12:
                t = sp / denom
                out.append(
                    (
                        previous[0]
                        + t * (current[0] - previous[0]),
                        previous[1]
                        + t * (current[1] - previous[1]),
                    )
                )

        elif not previous_inside and current_inside:
            denom = sp - sc
            if abs(denom) > 1e-12:
                t = sp / denom
                out.append(
                    (
                        previous[0]
                        + t * (current[0] - previous[0]),
                        previous[1]
                        + t * (current[1] - previous[1]),
                    )
                )
            out.append(current)

    return out


def footprint_metrics(poly):
    area = polygon_area(poly)

    if area <= 0:
        return {
            "water_fraction": math.nan,
            "center_signed_m": math.nan,
            "min_signed_m": math.nan,
            "max_signed_m": math.nan,
            "area_m2": math.nan,
        }

    water_area = polygon_area(
        clip_to_water(poly)
    )

    center_e = float(
        np.mean([p[0] for p in poly])
    )
    center_n = float(
        np.mean([p[1] for p in poly])
    )

    signed = np.array(
        [
            shoreline_signed_coordinate(
                x, y
            )
            for x, y in poly
        ],
        dtype=float,
    )

    return {
        "water_fraction": float(
            water_area / area
        ),
        "center_signed_m": float(
            shoreline_signed_coordinate(
                center_e,
                center_n,
            )
        ),
        "min_signed_m": float(
            np.min(signed)
        ),
        "max_signed_m": float(
            np.max(signed)
        ),
        "area_m2": float(area),
    }


def load_tide_interpolators():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [
        c.value for c in ws[1]
    ]

    models = [
        "EOT20_heightm",
        "GOT5.5_heightm",
        "GOT5.6_heightm",
        "FES2022_heightm",
    ]

    time_idx = header.index("time")
    col = {m: header.index(m) for m in models}

    times = []
    values = {m: [] for m in models}

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_idx]
        if not isinstance(t, datetime):
            continue

        ok = True
        temp = {}

        for m in models:
            v = finite(row[col[m]])
            if v is None:
                ok = False
                break
            temp[m] = v

        if not ok:
            continue

        times.append(t)
        for m in models:
            values[m].append(temp[m])

    wb.close()

    if len(times) < 2:
        raise RuntimeError(
            "Insufficient tide-model points."
        )

    epoch = np.array(
        [
            (t - times[0]).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    interpolators = {}

    for m in models:
        arr = np.asarray(values[m], dtype=float)

        def make_interp(arr):
            def interp(dt):
                x = (
                    dt - times[0]
                ).total_seconds()

                if x < epoch[0] or x > epoch[-1]:
                    return math.nan

                return float(
                    np.interp(
                        x,
                        epoch,
                        arr,
                    )
                )
            return interp

        interpolators[m] = make_interp(arr)

    return times, interpolators


def load_records(interpolators):
    records = []

    with open(
        CSV_FILE,
        newline="",
    ) as f:
        reader = csv.DictReader(f)

        required = [
            "solution_time_utc",
            "sat",
            "freq",
            "RH_m",
            "Azim",
        ]

        missing = [
            c for c in required
            if c not in reader.fieldnames
        ]

        if missing:
            raise RuntimeError(
                f"Missing required CSV columns: {missing}"
            )

        for row_num, r in enumerate(
            reader,
            start=2,
        ):
            try:
                dt = datetime.fromisoformat(
                    r["solution_time_utc"]
                )

                sat = int(float(r["sat"]))
                freq = int(float(r["freq"]))
                rh = finite(r["RH_m"])
                az = finite(r["Azim"])

                if rh is None or az is None:
                    continue

                tide = {
                    model: interpolators[model](dt)
                    for model in interpolators
                }

                if not all(
                    math.isfinite(v)
                    for v in tide.values()
                ):
                    continue

                records.append(
                    {
                        "row_num": row_num,
                        "dt": dt,
                        "sat": sat,
                        "freq": freq,
                        "rh": rh,
                        "az": az,
                        "tide": tide,
                    }
                )

            except Exception:
                continue

    return records


def main():
    print()
    print("=" * 90)
    print(
        "OBSERVED-RH VS TIDE-PREDICTED-OCEAN FRESNEL GEOMETRY"
    )
    print("=" * 90)

    print(
        f"Station: {LAT if 'LAT' in globals() else 41.8928243333}, "
        f"{LON if 'LON' in globals() else -69.9633227139}"
    )
    print(
        f"Reference H: {REFERENCE_H_M:.3f} m"
    )
    print(
        f"Reference water bearing: {WATER_BEARING_DEG:.2f}°"
    )
    print(
        f"Reference shoreline distance: {WATER_DISTANCE_M:.2f} m"
    )

    if not CSV_FILE.exists():
        raise SystemExit(
            f"ERROR: {CSV_FILE} not found."
        )

    if not TIDE_FILE.exists():
        raise SystemExit(
            f"ERROR: {TIDE_FILE} not found."
        )

    tide_times, interpolators = (
        load_tide_interpolators()
    )

    records = load_records(
        interpolators
    )

    print(
        f"Usable records: {len(records)}"
    )

    models = list(
        interpolators.keys()
    )

    results = []

    for r in records:
        for model in models:
            tide = r["tide"][model]

            predicted_ocean_h = (
                REFERENCE_H_M - tide
            )

            diff = (
                r["rh"]
                - predicted_ocean_h
            )

            row = {
                "row_num": r["row_num"],
                "dt": r["dt"].isoformat(),
                "sat": r["sat"],
                "freq": r["freq"],
                "rh_m": r["rh"],
                "az_deg": r["az"],
                "tide_model": model,
                "tide_m": tide,
                "predicted_ocean_reflector_h_m": predicted_ocean_h,
                "RH_minus_predicted_ocean_H_m": diff,
            }

            for elev in ELEVATIONS:
                obs = fresnel_zone_local(
                    r["freq"],
                    r["sat"],
                    elev,
                    r["rh"],
                    r["az"],
                )

                pred = fresnel_zone_local(
                    r["freq"],
                    r["sat"],
                    elev,
                    predicted_ocean_h,
                    r["az"],
                )

                obs_m = footprint_metrics(
                    obs["poly"]
                )

                pred_m = footprint_metrics(
                    pred["poly"]
                )

                tag = str(int(elev))

                row[
                    f"obs_{tag}_water_fraction"
                ] = obs_m["water_fraction"]

                row[
                    f"obs_{tag}_center_signed_m"
                ] = obs_m["center_signed_m"]

                row[
                    f"pred_{tag}_water_fraction"
                ] = pred_m["water_fraction"]

                row[
                    f"pred_{tag}_center_signed_m"
                ] = pred_m["center_signed_m"]

                row[
                    f"pred_{tag}_A_m"
                ] = pred["A_m"]

                row[
                    f"pred_{tag}_B_m"
                ] = pred["B_m"]

                row[
                    f"pred_{tag}_center_m"
                ] = pred["center_m"]

            results.append(row)

    fields = list(results[0].keys())

    with open(
        OUT_CSV,
        "w",
        newline="",
    ) as f:
        w = csv.DictWriter(
            f,
            fieldnames=fields,
        )
        w.writeheader()
        w.writerows(results)

    print()
    print("=" * 90)
    print("KEY RESULT")
    print("=" * 90)

    with open(
        OUT_SUMMARY,
        "w",
    ) as f:
        f.write(
            "OBSERVED-RH VS TIDE-PREDICTED-OCEAN GEOMETRY\n"
        )
        f.write("=" * 90 + "\n")

        for model in models:
            subset = [
                x for x in results
                if x["tide_model"] == model
            ]

            diffs = np.array(
                [
                    x[
                        "RH_minus_predicted_ocean_H_m"
                    ]
                    for x in subset
                ],
                dtype=float,
            )

            pred10 = np.array(
                [
                    x[
                        "pred_10_water_fraction"
                    ]
                    for x in subset
                ],
                dtype=float,
            )

            mean_diff = float(
                np.mean(diffs)
            )
            median_diff = float(
                np.median(diffs)
            )
            rms = float(
                np.sqrt(
                    np.mean(diffs ** 2)
                )
            )

            n_water = int(
                np.sum(
                    pred10 >= OPEN_THRESHOLD
                )
            )

            n_land = int(
                np.sum(
                    pred10 <= LAND_THRESHOLD
                )
            )

            print()
            print(model)
            print(
                "  mean discrepancy:",
                f"{mean_diff:+.3f} m"
            )
            print(
                "  median discrepancy:",
                f"{median_diff:+.3f} m"
            )
            print(
                "  RMS discrepancy:",
                f"{rms:.3f} m"
            )
            print(
                "  predicted 10° zone >=80% water:",
                f"{n_water}/{len(pred10)}"
            )
            print(
                "  predicted 10° zone <=20% water:",
                f"{n_land}/{len(pred10)}"
            )

            f.write(
                f"\n{model}\n"
            )
            f.write(
                f"mean discrepancy: {mean_diff:+.6f} m\n"
            )
            f.write(
                f"median discrepancy: {median_diff:+.6f} m\n"
            )
            f.write(
                f"RMS discrepancy: {rms:.6f} m\n"
            )
            f.write(
                f"predicted 10 deg zone >=80% water: "
                f"{n_water}/{len(pred10)}\n"
            )
            f.write(
                f"predicted 10 deg zone <=20% water: "
                f"{n_land}/{len(pred10)}\n"
            )

    print()
    print(
        "Outputs:"
    )
    print(
        f"  {OUT_CSV}"
    )
    print(
        f"  {OUT_SUMMARY}"
    )
    print()
    print("DONE")


if __name__ == "__main__":
    main()
