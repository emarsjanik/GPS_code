#!/usr/bin/env python3
"""
merge_tide_files.py

Joins the two tide model files into one continuous series, so
comparisons can span the whole GNSS record rather than stopping
where one file ends.

THE TIMESTAMP OFFSET

The two files contain the SAME EOT20 model output, offset by exactly
one 10-minute row. Confirmed directly: shifting one by -10 minutes
makes the maximum difference over 144 overlapping points exactly
0.00000 m.

They cannot both be right. A timeshift scan of the older file
against this station's own GNSS data found best agreement at about
+0.1 hours rather than 0, which points to the older file being the
misaligned one -- the newer file's timestamps are what the GNSS data
prefers.

That inference is not airtight: the scan cannot cleanly separate
+0.1 from +0.2 hours (RMS 0.0864 vs 0.0866 m, within noise). So the
shift is applied to the older file by default but exposed as
--shift-old-minutes, and can be set to 0 to merge them as-is if
Chris confirms the opposite.

Getting this wrong matters. At peak flow the water here moves about
0.6 m/hour, so 10 minutes is roughly 10 cm -- comparable to the
entire measurement error being studied. Merging without correcting
would embed that into every subsequent comparison.

WHICH VALUES WIN IN THE OVERLAP

The newer file's, since its alignment is the one the GNSS data
supports. The overlap is a single day, so this affects very little
either way.

Usage:
    python3 merge_tide_files.py \\
        --old marconi_tides_sherwood_20260715-20260831.xlsx \\
        --new marconi_tides_sherwood.xlsx \\
        --output marconi_tides_merged.xlsx
"""

from __future__ import annotations

import argparse
from datetime import datetime, timedelta
from pathlib import Path


def load(path: Path):
    """Returns {timestamp: {column: value}} plus the column order."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    if "time" not in header:
        raise SystemExit(f"{path}: no 'time' column (found {header})")
    ti = header.index("time")

    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        t = r[ti]
        if not isinstance(t, datetime):
            continue
        rows[t] = {header[i]: r[i] for i in range(len(header)) if i != ti}
    wb.close()
    return rows, [h for h in header if h != "time"]


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--old", required=True, help="the earlier file")
    p.add_argument("--new", required=True, help="the later file")
    p.add_argument("--output", required=True)
    p.add_argument("--shift-old-minutes", type=float, default=10.0,
                   help="minutes to add to the older file's timestamps so they "
                        "align with the newer file (default 10; use 0 to merge "
                        "without correction)")
    args = p.parse_args()

    old_rows, old_cols = load(Path(args.old))
    new_rows, new_cols = load(Path(args.new))

    print(f"old: {len(old_rows)} rows, columns {old_cols}")
    print(f"new: {len(new_rows)} rows, columns {new_cols}")

    shift = timedelta(minutes=args.shift_old_minutes)
    if args.shift_old_minutes:
        print(f"\nShifting the older file's timestamps by "
              f"{args.shift_old_minutes:+.0f} minutes to match the newer file.")
        old_rows = {t + shift: v for t, v in old_rows.items()}

    # Verify the correction actually worked, using whichever value
    # column both files share. A merge that silently leaves a 10 cm
    # step at the join is worse than no merge.
    shared_cols = [c for c in old_cols if c in new_cols]
    overlap = sorted(set(old_rows) & set(new_rows))
    if overlap and shared_cols:
        col = shared_cols[0]
        diffs = []
        for t in overlap:
            a, b = old_rows[t].get(col), new_rows[t].get(col)
            if isinstance(a, (int, float)) and isinstance(b, (int, float)):
                diffs.append(abs(a - b))
        if diffs:
            worst = max(diffs)
            print(f"\nOverlap check on '{col}': {len(diffs)} shared timestamps, "
                  f"largest difference {worst:.5f} m")
            if worst > 0.005:
                print("  WARNING: the two files still disagree after shifting.")
                print("  Check --shift-old-minutes before trusting this merge.")
            else:
                print("  They agree; the shift is correct.")
    elif not overlap:
        print("\nNo overlapping timestamps -- the files abut rather than overlap.")
        print("  The shift cannot be verified from the data itself.")

    # Newer values win where both exist.
    merged = dict(old_rows)
    merged.update(new_rows)

    columns = list(dict.fromkeys(old_cols + new_cols))

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "merged"
    ws.append(["time"] + columns)
    for t in sorted(merged):
        row = merged[t]
        ws.append([t] + [row.get(c) for c in columns])
    wb.save(args.output)

    times = sorted(merged)
    print(f"\nWrote {args.output}")
    print(f"  {len(merged)} rows, {times[0]} to {times[-1]}")
    print(f"  columns: {columns}")
    missing = [c for c in columns if c not in old_cols or c not in new_cols]
    if missing:
        print(f"  note: {missing} exist in only one file, so are blank "
              f"outside that file's range")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
