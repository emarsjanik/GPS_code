#!/usr/bin/env python3
"""
timeshift_scan.py

Tests whether a timezone/time-offset misalignment explains poor
agreement between the GNSS-IR spline output (gnssrefl's own header
confirms this is UTC) and the tide model file (timezone not
independently confirmed).

Scans a range of candidate time shifts, recomputes the de-meaned RMS
and correlation at each, and reports which shift (if any) produces a
clear, sharp improvement -- a sharp peak at a specific, physically
sensible offset (e.g. exactly 4 or 5 hours, matching EDT/EST vs UTC)
is strong, direct evidence of a real timezone mismatch, as opposed
to gradual, shift-insensitive noise.

Usage:
    python3 timeshift_scan.py \\
        --spline-file products/refl_code/Files/usgs/usgs_spline_out.txt \\
        --tide-file marconi_tides_sherwood.xlsx \\
        --tide-value-col EOT20_heightm
"""

from __future__ import annotations

import argparse
import math
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np


def load_spline_output(path: Path):
    times, values = [], []
    with open(path, errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            cols = line.split()
            if len(cols) < 9:
                continue
            try:
                year = int(float(cols[2]))
                month = int(float(cols[3]))
                day = int(float(cols[4]))
                hour = int(float(cols[5]))
                minute = int(float(cols[6]))
                second = int(float(cols[7]))
                water_level = float(cols[8])
            except (ValueError, IndexError):
                continue
            try:
                dt = datetime(year, month, day, hour, minute, second)
            except ValueError:
                continue
            times.append(dt)
            values.append(water_level)
    return times, np.asarray(values, float)


def load_tide_reference(path: Path, time_col: str, value_col: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    time_idx = header.index(time_col)
    value_idx = header.index(value_col)

    times, values = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[time_idx]
        if not isinstance(t, datetime):
            continue
        try:
            v = float(row[value_idx])
        except (TypeError, ValueError):
            continue
        if not math.isfinite(v):
            continue
        times.append(t)
        values.append(v)
    wb.close()
    return times, np.asarray(values, float)


def interpolate_reference(ref_times, ref_values, query_times):
    ref_epoch = np.array([(t - ref_times[0]).total_seconds() for t in ref_times])
    query_epoch = np.array([(t - ref_times[0]).total_seconds() for t in query_times])
    order = np.argsort(ref_epoch)
    return np.interp(
        query_epoch, ref_epoch[order], ref_values[order],
        left=np.nan, right=np.nan,
    )


def evaluate_shift(spline_times, spline_values, tide_times, tide_values, shift_hours):
    shifted_query_times = [t - timedelta(hours=shift_hours) for t in spline_times]
    tide_at_query = interpolate_reference(tide_times, tide_values, shifted_query_times)

    valid = np.isfinite(tide_at_query) & np.isfinite(spline_values)
    n_valid = int(np.sum(valid))
    if n_valid < 20:
        return n_valid, float("nan"), float("nan")

    gnss = spline_values[valid]
    tide = tide_at_query[valid]

    diff = gnss - tide
    demeaned_rms = float(np.std(diff))
    corr = float(np.corrcoef(gnss, tide)[0, 1]) if np.std(gnss) > 0 and np.std(tide) > 0 else float("nan")
    return n_valid, demeaned_rms, corr


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--spline-file", required=True)
    p.add_argument("--tide-file", required=True)
    p.add_argument("--tide-time-col", default="time")
    p.add_argument("--tide-value-col", required=True)
    p.add_argument("--max-shift-hours", type=float, default=14.0)
    p.add_argument("--step-hours", type=float, default=0.5)
    args = p.parse_args()

    spline_times, spline_values = load_spline_output(Path(args.spline_file))
    tide_times, tide_values = load_tide_reference(
        Path(args.tide_file), args.tide_time_col, args.tide_value_col
    )
    print(f"Loaded {len(spline_times)} GNSS-IR spline points, {len(tide_times)} tide model points")
    print()
    print(f"{'shift (h)':>10} {'n':>6} {'de-meaned RMS (m)':>20} {'correlation':>14}")
    print("-" * 55)

    results = []
    shift = -args.max_shift_hours
    while shift <= args.max_shift_hours + 1e-9:
        n, rms, corr = evaluate_shift(spline_times, spline_values, tide_times, tide_values, shift)
        results.append((shift, n, rms, corr))
        marker = ""
        # Flag physically meaningful US timezone offsets for easy scanning
        if abs(abs(shift) - 4.0) < 0.01 or abs(abs(shift) - 5.0) < 0.01:
            marker = "  <-- EDT/EST vs UTC"
        print(f"{shift:>10.1f} {n:>6d} {rms:>20.4f} {corr:>14.4f}{marker}")
        shift += args.step_hours

    valid_results = [r for r in results if not math.isnan(r[3])]
    if valid_results:
        best = max(valid_results, key=lambda r: abs(r[3]))
        print()
        print("=" * 55)
        print(f"Best-agreement shift: {best[0]:+.1f} hours")
        print(f"  n={best[1]}, de-meaned RMS={best[2]:.4f} m, correlation={best[3]:+.4f}")
        print("=" * 55)
        print()
        print("If this best shift is close to a round number like -4, -5,")
        print("+4, or +5 hours, and correlation there is dramatically")
        print("stronger (larger magnitude, more negative or more positive)")
        print("than at shift=0, this is strong, direct evidence of a real")
        print("timezone misalignment between the two files -- not proof of")
        print("anything else being wrong with the underlying data.")


if __name__ == "__main__":
    main()
