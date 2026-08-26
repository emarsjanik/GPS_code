#!/usr/bin/env python3
"""
plot_gnssir_vs_tide.py

Generates a line graph directly overlaying the GNSS-IR water level
(gnssrefl's evenly-sampled spline output) against the tide model, on
a shared time and value axis -- so the real agreement in shape, and
the real, still-unresolved constant offset between them, are both
visible directly, honestly, without artificially aligning the two
curves to hide the open question.

Confirmed, real file format used here (from direct inspection
earlier in this project): gnssrefl's own evenly-sampled spline
output file has the date in columns 3-8 (YYYY MM DD HH MM SS,
1-indexed) and column 9 is the already-computed quasi-sea-level
value (Hortho - RH), in meters.

Usage:
    python3 plot_gnssir_vs_tide.py \\
        --spline-file products/refl_code/Files/usgs/usgs_spline_out.txt \\
        --tide-file marconi_tides_sherwood.xlsx \\
        --tide-value-col EOT20_heightm \\
        --output comparison_plot.png
"""

from __future__ import annotations

import argparse
import math
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")  # write directly to a file, no display needed
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np


def load_spline_output(path: Path):
    times, values = [], []
    with open(path, errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            cols = line.split()
            if len(cols) < 9:
                continue
            try:
                year = int(float(cols[2]))
                month = int(float(cols[3]))
                day = int(float(cols[4]))
                hour = int(float(cols[5]))
                minute = int(float(cols[6]))
                second = int(float(cols[7]))
                water_level = float(cols[8])
            except (ValueError, IndexError):
                continue
            try:
                dt = datetime(year, month, day, hour, minute, second)
            except ValueError:
                continue
            times.append(dt)
            values.append(water_level)
    return times, np.asarray(values, float)


def load_tide_reference(path: Path, time_col: str, value_col: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    time_idx = header.index(time_col)
    value_idx = header.index(value_col)

    times, values = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[time_idx]
        if not isinstance(t, datetime):
            continue
        try:
            v = float(row[value_idx])
        except (TypeError, ValueError):
            continue
        if not math.isfinite(v):
            continue
        times.append(t)
        values.append(v)
    wb.close()
    return times, np.asarray(values, float)


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--spline-file", required=True)
    p.add_argument("--tide-file", required=True)
    p.add_argument("--tide-time-col", default="time")
    p.add_argument("--tide-value-col", required=True)
    p.add_argument("--output", default="gnssir_vs_tide.png")
    p.add_argument("--start-date", default=None, help="YYYY-MM-DD, optional, restricts the plotted window")
    p.add_argument("--end-date", default=None, help="YYYY-MM-DD, optional, restricts the plotted window")
    args = p.parse_args()

    spline_times, spline_values = load_spline_output(Path(args.spline_file))
    print(f"Loaded {len(spline_times)} GNSS-IR spline points")
    if not spline_times:
        raise SystemExit("No spline points loaded -- check --spline-file path/format.")

    tide_times, tide_values = load_tide_reference(
        Path(args.tide_file), args.tide_time_col, args.tide_value_col
    )
    print(f"Loaded {len(tide_times)} tide model points")

    if args.start_date:
        start = datetime.strptime(args.start_date, "%Y-%m-%d")
        mask = [t >= start for t in spline_times]
        spline_times = [t for t, m in zip(spline_times, mask) if m]
        spline_values = spline_values[mask]
        tide_mask = [t >= start for t in tide_times]
        tide_times = [t for t, m in zip(tide_times, tide_mask) if m]
        tide_values = tide_values[tide_mask]

    if args.end_date:
        end = datetime.strptime(args.end_date, "%Y-%m-%d")
        mask = [t <= end for t in spline_times]
        spline_times = [t for t, m in zip(spline_times, mask) if m]
        spline_values = spline_values[mask]
        tide_mask = [t <= end for t in tide_times]
        tide_times = [t for t, m in zip(tide_times, tide_mask) if m]
        tide_values = tide_values[tide_mask]

    fig, ax = plt.subplots(figsize=(16, 6))

    ax.plot(spline_times, spline_values, color="tab:blue", linewidth=1.2,
            label="GNSS-IR water level (this station)")
    ax.plot(tide_times, tide_values, color="tab:orange", linewidth=1.0,
            alpha=0.85, label=f"Tide model ({args.tide_value_col})")

    ax.set_xlabel("Date")
    ax.set_ylabel("Water level (m)")
    ax.set_title("GNSS-IR Water Level vs. Tide Model\n"
                  "(shown on a shared, unadjusted axis -- any constant vertical "
                  "offset between the two references is intentionally left visible, not corrected)")
    ax.legend(loc="upper right")
    ax.grid(True, alpha=0.3)

    ax.xaxis.set_major_locator(mdates.AutoDateLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    fig.autofmt_xdate()

    fig.tight_layout()
    fig.savefig(args.output, dpi=150)
    print(f"Plot saved to: {args.output}")


if __name__ == "__main__":
    main()

