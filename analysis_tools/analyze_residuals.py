#!/usr/bin/env python3
"""
analyze_residuals.py

The headline agreement figures (correlation, mean absolute
deviation) say how well the GNSS-IR water level matches the tide
model on average. They say nothing about whether the disagreement is
random scatter or has structure -- and the difference matters:

  - Random scatter is measurement noise. It averages down, and the
    only way to reduce it is more retrievals.

  - Structure means something systematic is unmodelled. A residual
    that grows with tidal range, or with tidal rate of change, or
    that varies by time of day, points at a specific physical or
    processing cause that could be corrected rather than merely
    averaged over.

This checks for the structures worth ruling out at a coastal
GNSS-IR site:

  1. Against tide height    -- a residual that grows toward high or
                               low water suggests the reflecting
                               surface moves outside the assumed
                               geometry at range extremes.

  2. Against tidal rate     -- gnssrefl's RHdot correction handles
                               a moving surface during an arc. A
                               residual correlated with rate of
                               change suggests it is
                               under-correcting.

  3. Against time of day    -- would point at a diurnal effect:
                               thermal, atmospheric, or a satellite
                               geometry that repeats daily.

  4. Over the record        -- a trend suggests drift: antenna
                               settling, a changing datum, or
                               seasonal water density.

  5. Distribution shape     -- heavy tails mean a few bad retrievals
                               dominate; a normal spread means the
                               noise is well behaved.

Usage:
    python3 analyze_residuals.py \\
        --spline-file products/refl_code/Files/usgs/usgs_spline_out.txt \\
        --tide-file marconi_tides_sherwood.xlsx \\
        --tide-value-col EOT20_heightm
"""

from __future__ import annotations

import argparse
import math
from datetime import datetime
from pathlib import Path

import numpy as np


def load_spline(path: Path):
    times, values = [], []
    with open(path, errors="replace") as f:
        for line in f:
            if line.startswith("%") or not line.strip():
                continue
            c = line.split()
            if len(c) < 9:
                continue
            try:
                dt = datetime(int(float(c[2])), int(float(c[3])), int(float(c[4])),
                              int(float(c[5])), int(float(c[6])), int(float(c[7])))
                v = float(c[8])
            except (ValueError, IndexError):
                continue
            times.append(dt)
            values.append(v)
    return times, np.asarray(values, float)


def load_tide(path: Path, time_col: str, value_col: str):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    ti, vi = header.index(time_col), header.index(value_col)
    times, values = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[ti]
        if not isinstance(t, datetime):
            continue
        try:
            v = float(row[vi])
        except (TypeError, ValueError):
            continue
        if math.isfinite(v):
            times.append(t)
            values.append(v)
    wb.close()
    return times, np.asarray(values, float)


def interp(ref_t, ref_v, q_t):
    base = ref_t[0]
    rx = np.array([(t - base).total_seconds() for t in ref_t])
    qx = np.array([(t - base).total_seconds() for t in q_t])
    o = np.argsort(rx)
    return np.interp(qx, rx[o], ref_v[o], left=np.nan, right=np.nan)


def describe_corr(r: float) -> str:
    a = abs(r)
    if a < 0.1:
        return "negligible -- no meaningful structure"
    if a < 0.25:
        return "weak"
    if a < 0.5:
        return "moderate -- worth investigating"
    return "strong -- something systematic is unmodelled"


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--spline-file", required=True)
    p.add_argument("--tide-file", required=True)
    p.add_argument("--tide-time-col", default="time")
    p.add_argument("--tide-value-col", required=True)
    args = p.parse_args()

    st, sv = load_spline(Path(args.spline_file))
    tt, tv = load_tide(Path(args.tide_file), args.tide_time_col, args.tide_value_col)

    tide_at = interp(tt, tv, st)
    ok = np.isfinite(tide_at) & np.isfinite(sv)

    times = [t for t, m in zip(st, ok) if m]
    gnss = sv[ok]
    tide = tide_at[ok]

    if len(gnss) < 50:
        print("Too few overlapping points for a meaningful analysis.")
        return 1

    resid = gnss - tide
    offset = resid.mean()
    r = resid - offset          # de-meaned: the datum offset is not the subject here

    print("=" * 70)
    print("  RESIDUAL STRUCTURE: GNSS-IR minus tide model")
    print("=" * 70)
    print(f"  {len(r)} paired points, {times[0].date()} to {times[-1].date()}")
    print(f"  Mean offset removed: {offset:+.3f} m")
    print(f"  Residual spread: {r.std():.3f} m std, "
          f"{np.abs(r).mean():.3f} m mean absolute")
    print()

    # --- 1. Against tide height ---
    c = np.corrcoef(tide, np.abs(r))[0, 1]
    print(f"  1. |residual| vs tide height:      r = {c:+.3f}  ({describe_corr(c)})")

    # Compare the extremes against the middle explicitly: a
    # correlation can hide a U-shape, where error grows at BOTH ends.
    lo, hi = np.percentile(tide, [25, 75])
    mid_mask = (tide > lo) & (tide < hi)
    ext_mask = ~mid_mask
    print(f"       mid-range tides:  {np.abs(r[mid_mask]).mean():.3f} m mean abs error")
    print(f"       high/low extremes: {np.abs(r[ext_mask]).mean():.3f} m mean abs error")
    print()

    # --- 2. Against tidal rate of change ---
    secs = np.array([(t - times[0]).total_seconds() for t in times])
    rate = np.gradient(tide, secs) * 3600.0     # m/hour
    c = np.corrcoef(np.abs(rate), np.abs(r))[0, 1]
    print(f"  2. |residual| vs tidal rate:       r = {c:+.3f}  ({describe_corr(c)})")
    slack = np.abs(rate) < np.percentile(np.abs(rate), 25)
    fast = np.abs(rate) > np.percentile(np.abs(rate), 75)
    print(f"       near slack water: {np.abs(r[slack]).mean():.3f} m mean abs error")
    print(f"       fastest flow:     {np.abs(r[fast]).mean():.3f} m mean abs error")
    print()

    # --- 3. Time of day ---
    hours = np.array([t.hour + t.minute / 60 for t in times])
    print("  3. Residual by time of day (UTC):")
    worst_h, worst_v = None, 0.0
    for h0 in range(0, 24, 4):
        m = (hours >= h0) & (hours < h0 + 4)
        if m.sum() < 10:
            continue
        v = np.abs(r[m]).mean()
        bar = "#" * int(v * 200)
        print(f"       {h0:02d}-{h0+4:02d}h  {v:.3f} m  {bar}")
        if v > worst_v:
            worst_h, worst_v = h0, v
    spread = max(np.abs(r[(hours >= h) & (hours < h + 4)]).mean()
                 for h in range(0, 24, 4)
                 if ((hours >= h) & (hours < h + 4)).sum() >= 10)
    best = min(np.abs(r[(hours >= h) & (hours < h + 4)]).mean()
               for h in range(0, 24, 4)
               if ((hours >= h) & (hours < h + 4)).sum() >= 10)
    ratio = spread / best if best > 0 else float("nan")
    if ratio > 1.5:
        print(f"       -> {ratio:.1f}x variation across the day: a diurnal effect")
    else:
        print(f"       -> {ratio:.1f}x variation: no meaningful diurnal pattern")
    print()

    # --- 4. Drift over the record ---
    days = secs / 86400.0
    slope, intercept = np.polyfit(days, r, 1)
    total = slope * (days[-1] - days[0])
    print(f"  4. Drift over the record:          {slope*1000:+.2f} mm/day "
          f"({total*100:+.1f} cm over {days[-1]-days[0]:.0f} days)")
    if abs(total) > 0.02:
        print("       -> a real trend: antenna settling, datum change, or seasonal")
    else:
        print("       -> negligible; the offset is stable")
    print()

    # --- 5. Distribution ---
    within1 = 100 * np.mean(np.abs(r) < r.std())
    within2 = 100 * np.mean(np.abs(r) < 2 * r.std())
    print("  5. Distribution shape:")
    print(f"       within 1 sd: {within1:.0f}%  (normal would be 68%)")
    print(f"       within 2 sd: {within2:.0f}%  (normal would be 95%)")
    worst = np.sort(np.abs(r))[-int(len(r) * 0.01):]
    print(f"       worst 1% of points average {worst.mean():.3f} m off")
    if within2 < 93:
        print("       -> heavy tails: a minority of bad retrievals dominates the error")
    else:
        print("       -> well-behaved; no small group of outliers dominating")

    print()
    print("=" * 70)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
