#!/usr/bin/env python3
"""
export_timeseries.py

Exports the actual timeseries data behind the GNSS-IR vs. tide model
comparison plot as a single, clean CSV -- one row per GNSS-IR spline
point, with the tide model's value at that same moment interpolated
alongside it. Uses the same loading and interpolation logic already
verified in plot_gnssir_vs_tide.py and compare_to_tide_deviation.py,
not new, separately-untested parsing.

Usage:
    python3 export_timeseries.py \\
        --spline-file products/refl_code/Files/usgs/usgs_spline_out.txt \\
        --tide-file marconi_tides_sherwood.xlsx \\
        --tide-value-col EOT20_heightm \\
        --output timeseries_export.csv
"""

from __future__ import annotations

import argparse
import csv
import math
from datetime import datetime
from pathlib import Path

import numpy as np


def load_spline_output(path: Path):
    times, values = [], []
    with open(path, errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            cols = line.split()
            if len(cols) < 9:
                continue
            try:
                year = int(float(cols[2]))
                month = int(float(cols[3]))
                day = int(float(cols[4]))
                hour = int(float(cols[5]))
                minute = int(float(cols[6]))
                second = int(float(cols[7]))
                water_level = float(cols[8])
            except (ValueError, IndexError):
                continue
            try:
                dt = datetime(year, month, day, hour, minute, second)
            except ValueError:
                continue
            times.append(dt)
            values.append(water_level)
    return times, np.asarray(values, float)


def load_tide_reference(path: Path, time_col: str, value_col: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    time_idx = header.index(time_col)
    value_idx = header.index(value_col)

    times, values = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[time_idx]
        if not isinstance(t, datetime):
            continue
        try:
            v = float(row[value_idx])
        except (TypeError, ValueError):
            continue
        if not math.isfinite(v):
            continue
        times.append(t)
        values.append(v)
    wb.close()
    return times, np.asarray(values, float)


def interpolate_reference(ref_times, ref_values, query_times):
    ref_epoch = np.array([(t - ref_times[0]).total_seconds() for t in ref_times])
    query_epoch = np.array([(t - ref_times[0]).total_seconds() for t in query_times])
    order = np.argsort(ref_epoch)
    return np.interp(
        query_epoch, ref_epoch[order], ref_values[order],
        left=np.nan, right=np.nan,
    )


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--spline-file", required=True)
    p.add_argument("--tide-file", required=True)
    p.add_argument("--tide-time-col", default="time")
    p.add_argument("--tide-value-col", required=True)
    p.add_argument("--output", default="timeseries_export.csv")
    args = p.parse_args()

    spline_times, spline_values = load_spline_output(Path(args.spline_file))
    print(f"Loaded {len(spline_times)} GNSS-IR spline points")
    if not spline_times:
        raise SystemExit("No spline points loaded -- check --spline-file path/format.")

    tide_times, tide_values = load_tide_reference(
        Path(args.tide_file), args.tide_time_col, args.tide_value_col
    )
    print(f"Loaded {len(tide_times)} tide model points")

    tide_at_spline_times = interpolate_reference(tide_times, tide_values, spline_times)

    with open(args.output, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["timestamp_utc", "gnss_ir_water_level_m", f"tide_model_{args.tide_value_col}"])
        for t, gnss_v, tide_v in zip(spline_times, spline_values, tide_at_spline_times):
            tide_str = f"{tide_v:.4f}" if np.isfinite(tide_v) else ""
            writer.writerow([t.strftime("%Y-%m-%d %H:%M:%S"), f"{gnss_v:.4f}", tide_str])

    n_with_tide = int(np.sum(np.isfinite(tide_at_spline_times)))
    print(f"Wrote {len(spline_times)} rows to {args.output}")
    print(f"  ({n_with_tide} rows have a matching tide model value; the rest are outside the tide file's time range)")


if __name__ == "__main__":
    main()
