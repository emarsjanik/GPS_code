#!/usr/bin/env python3
"""
Marconi datum-correction validation -- clean version.

This version is intentionally written from scratch to avoid the previous
function/variable name collision.

It compares:
    RAW GNSS-R = 18.665 - RH
against
    EOT20 at the exact GNSS-R observation time

and:
    CORRECTED GNSS-R = RAW GNSS-R + 0.242 m

The +0.242 m value is an external hypothesis test, not fitted here.

Outputs:
  marconi_datum_test_clean.csv
  marconi_datum_test_clean_summary.txt
  marconi_datum_test_clean_plots/
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

H_ORTHO_M = 18.665
DATUM_SHIFT_M = 0.242
AZ_TOL_DEG = 3.0
MIN_TRACK_N = 4

OUT_CSV = Path(
    "marconi_datum_test_clean.csv"
)

OUT_SUMMARY = Path(
    "marconi_datum_test_clean_summary.txt"
)

OUT_PLOTS = Path(
    "marconi_datum_test_clean_plots"
)

TIDE_COLUMN = "EOT20_heightm"


# ---------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------

def finite(value):
    try:
        x = float(value)
        return x if math.isfinite(x) else None
    except Exception:
        return None


def to_datetime(year, doy, utc_hours):
    return (
        datetime(year, 1, 1)
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


def circular_difference(a, b):
    d = abs(a - b) % 360.0
    return min(d, 360.0 - d)


def correlation(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 3:
        return math.nan

    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(np.corrcoef(x, y)[0, 1])


def rms(values):
    values = np.asarray(values, dtype=float)
    if len(values) == 0:
        return math.nan
    return float(np.sqrt(np.mean(values**2)))


def mae(values):
    values = np.asarray(values, dtype=float)
    if len(values) == 0:
        return math.nan
    return float(np.mean(np.abs(values)))


def mad(values):
    values = np.asarray(values, dtype=float)
    if len(values) == 0:
        return math.nan
    med = np.median(values)
    return float(
        1.4826 * np.median(np.abs(values - med))
    )


# ---------------------------------------------------------------------
# LOAD EOT20
# ---------------------------------------------------------------------

def load_eot20():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]
    header = [cell.value for cell in ws[1]]

    time_index = header.index("time")
    tide_index = header.index(TIDE_COLUMN)

    times = []
    values = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_index]
        v = finite(row[tide_index])

        if not isinstance(t, datetime):
            continue

        if v is None:
            continue

        times.append(t)
        values.append(v)

    wb.close()

    if len(times) < 2:
        raise RuntimeError(
            "Not enough EOT20 points."
        )

    return (
        np.asarray(times, dtype="datetime64[ms]"),
        np.asarray(values, dtype=float),
    )


def interpolate_tide(
    tide_times,
    tide_values,
    dt,
):
    x = tide_times.astype("int64")

    query = np.datetime64(
        dt,
        "ms",
    ).astype("int64")

    if query < x[0] or query > x[-1]:
        return math.nan

    return float(
        np.interp(
            query,
            x,
            tide_values,
        )
    )


# ---------------------------------------------------------------------
# LOAD GNSS-R
# ---------------------------------------------------------------------

def load_gnssr():
    rows = []

    for path in sorted(
        RESULT_DIR.glob("*.txt")
    ):
        try:
            int(path.stem)
        except Exception:
            continue

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if not line or line.startswith("%"):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc_hours = float(c[4])
                az = float(c[5])
                freq = int(float(c[10]))
                rise = int(float(c[11]))
                amp = float(c[6])
                pkn = float(c[13])
                nobs = int(float(c[9]))
            except Exception:
                continue

            if freq != 1:
                continue

            dt = to_datetime(
                year,
                doy,
                utc_hours,
            )

            raw_wl = H_ORTHO_M - rh
            corrected_wl = raw_wl + DATUM_SHIFT_M

            rows.append(
                {
                    "datetime": dt,
                    "doy": doy,
                    "sat": sat,
                    "freq": freq,
                    "rise": rise,
                    "az": az,
                    "RH_m": rh,
                    "raw_wl_m": raw_wl,
                    "corrected_wl_m": corrected_wl,
                    "Amp": amp,
                    "PkNoise": pkn,
                    "NumbOf": nobs,
                }
            )

    rows.sort(
        key=lambda row: row["datetime"]
    )

    return rows


# ---------------------------------------------------------------------
# ADD EXACT-TIME TIDE
# ---------------------------------------------------------------------

def attach_tide(rows, tide_times, tide_values):
    valid_rows = []

    for row in rows:
        tide = interpolate_tide(
            tide_times,
            tide_values,
            row["datetime"],
        )

        if not math.isfinite(tide):
            continue

        row["eot20_m"] = tide
        row["raw_residual_m"] = (
            row["raw_wl_m"] - tide
        )
        row["corrected_residual_m"] = (
            row["corrected_wl_m"] - tide
        )

        valid_rows.append(row)

    return valid_rows


# ---------------------------------------------------------------------
# POPULATION STATISTICS
# ---------------------------------------------------------------------

def summarize_residuals(
    residuals,
    water_levels,
    tide_levels,
):
    residuals = np.asarray(
        residuals,
        dtype=float,
    )

    water_levels = np.asarray(
        water_levels,
        dtype=float,
    )

    tide_levels = np.asarray(
        tide_levels,
        dtype=float,
    )

    slope = math.nan
    r = math.nan

    if len(residuals) >= 3:
        r = correlation(
            water_levels,
            tide_levels,
        )
        slope = float(
            np.polyfit(
                tide_levels,
                water_levels,
                1,
            )[0]
        )

    return {
        "n": len(residuals),
        "mean_m": float(np.mean(residuals)),
        "median_m": float(np.median(residuals)),
        "sd_m": float(np.std(residuals)),
        "rms_m": rms(residuals),
        "mae_m": mae(residuals),
        "mad_m": mad(residuals),
        "r": r,
        "slope": slope,
    }


# ---------------------------------------------------------------------
# TRACK CLUSTERING
# ---------------------------------------------------------------------

def build_tracks(rows):
    buckets = defaultdict(list)

    for row in rows:
        buckets[
            (
                row["sat"],
                row["freq"],
                row["rise"],
            )
        ].append(row)

    tracks = []

    for key, group in buckets.items():

        group = sorted(
            group,
            key=lambda row: row["az"],
        )

        current = []
        previous_az = None

        for row in group:

            if (
                previous_az is None
                or circular_difference(
                    row["az"],
                    previous_az,
                ) <= AZ_TOL_DEG
            ):
                current.append(row)
            else:
                if len(current) >= MIN_TRACK_N:
                    tracks.append(current)
                current = [row]

            previous_az = row["az"]

        if len(current) >= MIN_TRACK_N:
            tracks.append(current)

    return tracks


def summarize_track(track):
    raw = np.asarray(
        [
            row["raw_residual_m"]
            for row in track
        ],
        dtype=float,
    )

    corrected = np.asarray(
        [
            row["corrected_residual_m"]
            for row in track
        ],
        dtype=float,
    )

    tide = np.asarray(
        [
            row["eot20_m"]
            for row in track
        ],
        dtype=float,
    )

    raw_wl = np.asarray(
        [
            row["raw_wl_m"]
            for row in track
        ],
        dtype=float,
    )

    corrected_wl = np.asarray(
        [
            row["corrected_wl_m"]
            for row in track
        ],
        dtype=float,
    )

    az = np.asarray(
        [
            row["az"]
            for row in track
        ],
        dtype=float,
    )

    days = sorted(
        {
            row["doy"]
            for row in track
        }
    )

    return {
        "sat": track[0]["sat"],
        "freq": track[0]["freq"],
        "rise": track[0]["rise"],
        "n": len(track),
        "n_days": len(days),
        "az_mean_deg": float(np.mean(az)),
        "az_sd_deg": float(np.std(az)),
        "raw_median_m": float(np.median(raw)),
        "corrected_median_m": float(np.median(corrected)),
        "raw_mean_m": float(np.mean(raw)),
        "corrected_mean_m": float(np.mean(corrected)),
        "raw_sd_m": float(np.std(raw)),
        "corrected_sd_m": float(np.std(corrected)),
        "raw_rms_m": rms(raw),
        "corrected_rms_m": rms(corrected),
        "raw_mae_m": mae(raw),
        "corrected_mae_m": mae(corrected),
        "raw_r": correlation(raw_wl, tide),
        "corrected_r": correlation(corrected_wl, tide),
        "raw_slope": float(
            np.polyfit(
                tide,
                raw_wl,
                1,
            )[0]
        ),
        "corrected_slope": float(
            np.polyfit(
                tide,
                corrected_wl,
                1,
            )[0]
        ),
    }


# ---------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------

def write_csv(rows):
    fields = [
        "datetime",
        "doy",
        "sat",
        "freq",
        "rise",
        "az",
        "RH_m",
        "raw_wl_m",
        "corrected_wl_m",
        "eot20_m",
        "raw_residual_m",
        "corrected_residual_m",
        "Amp",
        "PkNoise",
        "NumbOf",
    ]

    with open(
        OUT_CSV,
        "w",
        newline="",
    ) as handle:

        writer = csv.DictWriter(
            handle,
            fieldnames=fields,
        )

        writer.writeheader()

        for row in rows:

            output = {}

            for field in fields:
                value = row[field]

                if isinstance(value, datetime):
                    value = value.isoformat()

                output[field] = value

            writer.writerow(output)


# ---------------------------------------------------------------------
# PLOTS
# ---------------------------------------------------------------------

def make_overlay(
    tide_times,
    tide_values,
    rows,
):
    OUT_PLOTS.mkdir(
        parents=True,
        exist_ok=True,
    )

    start = min(
        row["datetime"]
        for row in rows
    )

    end = max(
        row["datetime"]
        for row in rows
    )

    tide_mask = (
        (tide_times >= np.datetime64(start))
        & (tide_times <= np.datetime64(end))
    )

    fig, ax = plt.subplots(
        figsize=(15, 7)
    )

    ax.plot(
        tide_times[tide_mask],
        tide_values[tide_mask],
        linewidth=2.5,
        label="EOT20 tide model",
    )

    ax.plot(
        [row["datetime"] for row in rows],
        [row["raw_wl_m"] for row in rows],
        linewidth=1.1,
        marker="o",
        markersize=2.5,
        label="GNSS-R raw",
    )

    ax.plot(
        [row["datetime"] for row in rows],
        [row["corrected_wl_m"] for row in rows],
        linewidth=1.1,
        marker="o",
        markersize=2.5,
        label=f"GNSS-R +{DATUM_SHIFT_M:.3f} m",
    )

    ax.set_title(
        "Marconi GNSS-R vs EOT20: Raw and Datum-Tested"
    )
    ax.set_xlabel("UTC")
    ax.set_ylabel("Water-level elevation / anomaly (m)")
    ax.grid(True, alpha=0.25)
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_PLOTS / "01_raw_vs_corrected_overlay.png",
        dpi=200,
    )

    plt.close(fig)


def make_histogram(rows):
    raw = np.asarray(
        [
            row["raw_residual_m"] * 100.0
            for row in rows
        ]
    )

    corrected = raw + DATUM_SHIFT_M * 100.0

    fig, ax = plt.subplots(
        figsize=(10, 6)
    )

    ax.hist(
        raw,
        bins=30,
        alpha=0.55,
        label="Raw",
    )

    ax.hist(
        corrected,
        bins=30,
        alpha=0.55,
        label=f"+{DATUM_SHIFT_M:.3f} m test",
    )

    ax.axvline(
        0,
        linewidth=1.5,
    )

    ax.set_title(
        "GNSS-R − EOT20 Residual Distribution"
    )
    ax.set_xlabel("Residual (cm)")
    ax.set_ylabel("Count")
    ax.grid(True, axis="y", alpha=0.25)
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_PLOTS / "02_residual_histogram.png",
        dpi=200,
    )

    plt.close(fig)


def make_residual_vs_tide(rows):
    x = np.asarray(
        [
            row["eot20_m"]
            for row in rows
        ]
    )

    raw = np.asarray(
        [
            row["raw_residual_m"] * 100.0
            for row in rows
        ]
    )

    corrected = raw + DATUM_SHIFT_M * 100.0

    coeff = np.polyfit(
        x,
        corrected,
        1,
    )

    xx = np.linspace(
        x.min(),
        x.max(),
        200,
    )

    fig, ax = plt.subplots(
        figsize=(9, 7)
    )

    ax.scatter(
        x,
        corrected,
        s=18,
        alpha=0.5,
        label="Corrected residual",
    )

    ax.plot(
        xx,
        coeff[0] * xx + coeff[1],
        linewidth=2,
        label=f"trend slope={coeff[0]:+.3f}",
    )

    ax.axhline(
        np.median(corrected),
        linestyle="--",
        linewidth=1.5,
        label=(
            f"median={np.median(corrected):+.1f} cm"
        ),
    )

    ax.axhline(
        0,
        linewidth=1.2,
    )

    ax.set_title(
        "Datum-Tested GNSS-R Residual vs EOT20"
    )
    ax.set_xlabel("EOT20 at GNSS-R observation (m)")
    ax.set_ylabel("Corrected GNSS-R − EOT20 (cm)")
    ax.grid(True, alpha=0.25)
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_PLOTS / "03_corrected_residual_vs_tide.png",
        dpi=200,
    )

    plt.close(fig)


def make_track_offsets(track_summaries):
    track_summaries = sorted(
        track_summaries,
        key=lambda item:
            item["raw_median_m"],
    )

    labels = [
        (
            f"PRN {item['sat']} "
            f"{'R' if item['rise'] == 1 else 'S'} "
            f"{item['az_mean_deg']:.1f}°"
        )
        for item in track_summaries
    ]

    raw = np.asarray(
        [
            100.0 * item["raw_median_m"]
            for item in track_summaries
        ]
    )

    corrected = np.asarray(
        [
            100.0 * item["corrected_median_m"]
            for item in track_summaries
        ]
    )

    y = np.arange(
        len(labels)
    )

    fig, ax = plt.subplots(
        figsize=(12, 9)
    )

    ax.plot(
        raw,
        y,
        "o-",
        label="Raw median offset",
    )

    ax.plot(
        corrected,
        y,
        "o-",
        label=f"+{DATUM_SHIFT_M:.3f} m test",
    )

    ax.axvline(
        0,
        linewidth=1.5,
    )

    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.set_xlabel(
        "GNSS-R − EOT20 median offset (cm)"
    )
    ax.set_title(
        "Persistent Track Datum Test"
    )
    ax.grid(
        True,
        axis="x",
        alpha=0.25,
    )
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_PLOTS / "04_track_offsets.png",
        dpi=200,
    )

    plt.close(fig)


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():
    print()
    print("=" * 92)
    print(
        "MARCONI DATUM TEST -- CLEAN VERSION"
    )
    print("=" * 92)
    print(
        f"EOT20 model: {TIDE_COLUMN}"
    )
    print(
        f"H_ortho: {H_ORTHO_M:.3f} m"
    )
    print(
        f"Independent test shift: +{DATUM_SHIFT_M:.3f} m"
    )

    tide_times, tide_values = load_eot20()
    gnssr_rows = load_gnssr()

    rows = attach_tide(
        gnssr_rows,
        tide_times,
        tide_values,
    )

    if not rows:
        raise SystemExit(
            "No overlapping GNSS-R/EOT20 observations."
        )

    raw_residuals = np.asarray(
        [
            row["raw_residual_m"]
            for row in rows
        ]
    )

    corrected_residuals = np.asarray(
        [
            row["corrected_residual_m"]
            for row in rows
        ]
    )

    raw_wl = np.asarray(
        [
            row["raw_wl_m"]
            for row in rows
        ]
    )

    corrected_wl = np.asarray(
        [
            row["corrected_wl_m"]
            for row in rows
        ]
    )

    tide_at_obs = np.asarray(
        [
            row["eot20_m"]
            for row in rows
        ]
    )

    raw_summary = summarize_residuals(
        raw_residuals,
        raw_wl,
        tide_at_obs,
    )

    corrected_summary = summarize_residuals(
        corrected_residuals,
        corrected_wl,
        tide_at_obs,
    )

    print()
    print("=" * 92)
    print("RAW VS DATUM-CORRECTED")
    print("=" * 92)

    for label, summary in [
        ("RAW", raw_summary),
        ("CORRECTED", corrected_summary),
    ]:

        print()
        print(label)
        print(
            f"  mean   = {summary['mean_m']:+.4f} m"
            f" ({summary['mean_m']*100:+.2f} cm)"
        )
        print(
            f"  median = {summary['median_m']:+.4f} m"
            f" ({summary['median_m']*100:+.2f} cm)"
        )
        print(
            f"  RMS    = {summary['rms_m']:.4f} m"
            f" ({summary['rms_m']*100:.2f} cm)"
        )
        print(
            f"  MAE    = {summary['mae_m']:.4f} m"
        )
        print(
            f"  MAD    = {summary['mad_m']:.4f} m"
        )
        print(
            f"  r      = {summary['r']:+.4f}"
        )
        print(
            f"  slope  = {summary['slope']:+.4f}"
        )

    track_groups = build_tracks(
        rows
    )

    track_summaries = [
        summarize_track(group)
        for group in track_groups
    ]

    track_summaries.sort(
        key=lambda item:
            item["raw_median_m"],
    )

    print()
    print("=" * 92)
    print("PERSISTENT TRACK DATUM TEST")
    print("=" * 92)

    for item in track_summaries:
        print(
            f"PRN {item['sat']:2d} "
            f"{'RISING' if item['rise'] == 1 else 'SETTING':7s} "
            f"Az={item['az_mean_deg']:7.2f} "
            f"N={item['n']:2d} "
            f"raw_med={item['raw_median_m']:+.3f} m "
            f"corr_med={item['corrected_median_m']:+.3f} m "
            f"raw_RMS={item['raw_rms_m']*100:.1f} cm "
            f"corr_RMS={item['corrected_rms_m']*100:.1f} cm "
            f"raw_r={item['raw_r']:+.4f} "
            f"corr_r={item['corrected_r']:+.4f}"
        )

    write_csv(rows)

    summary_lines = [
        "MARCONI DATUM TEST -- CLEAN VERSION",
        "=" * 92,
        f"EOT20 model: {TIDE_COLUMN}",
        f"H_ortho: {H_ORTHO_M:.3f} m",
        f"Independent shift: +{DATUM_SHIFT_M:.3f} m",
        "",
        "POPULATION",
        "-" * 92,
        f"Raw mean: {raw_summary['mean_m']:+.4f} m",
        f"Raw median: {raw_summary['median_m']:+.4f} m",
        f"Raw RMS: {raw_summary['rms_m']:.4f} m",
        f"Raw MAE: {raw_summary['mae_m']:.4f} m",
        f"Corrected mean: {corrected_summary['mean_m']:+.4f} m",
        f"Corrected median: {corrected_summary['median_m']:+.4f} m",
        f"Corrected RMS: {corrected_summary['rms_m']:.4f} m",
        f"Corrected MAE: {corrected_summary['mae_m']:.4f} m",
        f"Correlation: {corrected_summary['r']:+.4f}",
        f"Slope: {corrected_summary['slope']:+.4f}",
        "",
        "TRACKS",
        "-" * 92,
    ]

    for item in track_summaries:
        summary_lines.append(
            f"PRN {item['sat']:2d} "
            f"{'RISING' if item['rise'] == 1 else 'SETTING':7s} "
            f"Az={item['az_mean_deg']:.2f} "
            f"N={item['n']} "
            f"raw_med={item['raw_median_m']:+.4f} "
            f"corr_med={item['corrected_median_m']:+.4f} "
            f"raw_RMS={item['raw_rms_m']:.4f} "
            f"corr_RMS={item['corrected_rms_m']:.4f}"
        )

    summary_lines += [
        "",
        "INTERPRETATION",
        "-" * 92,
        "The +0.242 m value is an external hypothesis test.",
        "It is not fitted to these GNSS-R observations.",
        "A constant vertical shift changes the offset metrics but not",
        "correlation or regression slope.",
    ]

    OUT_SUMMARY.write_text(
        "\n".join(summary_lines)
        + "\n"
    )

    print()
    print("Generating plots...")
    make_overlay(
        tide_times,
        tide_values,
        rows,
    )
    make_histogram(rows)
    make_residual_vs_tide(rows)
    make_track_offsets(track_summaries)

    print()
    print("=" * 92)
    print("OUTPUTS")
    print("=" * 92)
    print("CSV:", OUT_CSV.resolve())
    print("Summary:", OUT_SUMMARY.resolve())
    print("Plots:", OUT_PLOTS.resolve())
    print("DONE")


if __name__ == "__main__":
    main()
