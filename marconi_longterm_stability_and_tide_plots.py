#!/usr/bin/env python3
"""
Marconi long-term GNSS-R stability + tide plotting.

This script analyzes the extended 17-23 m GPS L1 experiment over every
successful daily result file currently present in:

  products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13/

It is designed to answer:

1. Which repeated GPS L1 tracks remain stable over the longer interval?
2. Which tracks follow the confirmed tide model most closely?
3. Does the absolute GNSS-R water level
       WL_GNSSR = 18.665 - RH
   remain approximately tied to the tide with one constant offset?
4. Which tracks deserve the expensive Fresnel/topobathy validation?
5. What does the complete GNSS-R population look like when plotted
   against the tide over the entire available observation period?

DEFAULT PRIMARY TIDE MODEL
--------------------------
EOT20_heightm

The script also loads GOT5.5, GOT5.6, and FES2022 for comparison and
creates an ensemble mean, but EOT20 is the primary plotting/calibration
reference. Change PRIMARY_TIDE_MODEL below if the confirmed model is
different.

IMPORTANT
---------
This is a diagnostic long-term analysis. It does not modify the GNSS-IR
configuration or any production JSON.

Outputs
-------
marconi_longterm_track_stability.csv
marconi_longterm_track_stability_summary.txt

plots/
  marconi_all_gnssr_vs_tide_timeseries.png
  marconi_all_gnssr_vs_tide_scatter.png
  marconi_top_tracks_vs_tide.png
  marconi_top_track_residuals.png
  marconi_track_calibration_stability.png
  marconi_daily_population_vs_tide.png

The script uses matplotlib only and does not require seaborn.
"""

from __future__ import annotations

import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

OUT_CSV = Path(
    "marconi_longterm_track_stability.csv"
)

OUT_SUMMARY = Path(
    "marconi_longterm_track_stability_summary.txt"
)

PLOT_DIR = Path(
    "marconi_longterm_plots"
)

H_ORTHO_M = 18.665

PRIMARY_TIDE_MODEL = "EOT20_heightm"

MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]

# Minimum data required for a repeated track to be ranked.
MIN_TRACK_N = 5

# "Top" tracks for detailed plots.
TOP_TRACKS_FOR_PLOTS = 8


# ---------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 3:
        return math.nan

    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(
        np.corrcoef(x, y)[0, 1]
    )


def rms(x):
    x = np.asarray(x, dtype=float)

    if len(x) == 0:
        return math.nan

    return float(
        np.sqrt(
            np.mean(x ** 2)
        )
    )


def mae(x):
    x = np.asarray(x, dtype=float)

    if len(x) == 0:
        return math.nan

    return float(
        np.mean(
            np.abs(x)
        )
    )


def utc_hours_to_datetime(
    year,
    doy,
    utc_hours,
):
    return (
        datetime(
            year,
            1,
            1,
        )
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


# ---------------------------------------------------------------------
# TIDE DATA
# ---------------------------------------------------------------------

def load_tide_data():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [
        c.value
        for c in ws[1]
    ]

    time_col = header.index(
        "time"
    )

    model_cols = {
        model:
            header.index(model)
        for model in MODELS
    }

    times = []
    values = {
        model: []
        for model in MODELS
    }

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):

        t = row[time_col]

        if not isinstance(
            t,
            datetime,
        ):
            continue

        tmp = {}
        good = True

        for model in MODELS:

            v = finite(
                row[
                    model_cols[model]
                ]
            )

            if v is None:
                good = False
                break

            tmp[model] = v

        if not good:
            continue

        times.append(t)

        for model in MODELS:
            values[model].append(
                tmp[model]
            )

    wb.close()

    epoch = np.asarray(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    arrays = {
        model:
            np.asarray(
                values[model],
                dtype=float,
            )
        for model in MODELS
    }

    def tide_at(
        dt,
        model,
    ):
        x = (
            dt - times[0]
        ).total_seconds()

        if (
            x < epoch[0]
            or x > epoch[-1]
        ):
            return math.nan

        return float(
            np.interp(
                x,
                epoch,
                arrays[model],
            )
        )

    return times, tide_at


# ---------------------------------------------------------------------
# GNSS-IR DATA
# ---------------------------------------------------------------------

def load_gnss_results():
    rows = []

    result_files = sorted(
        RESULT_DIR.glob("*.txt")
    )

    if not result_files:
        raise SystemExit(
            f"No result files found in:\n{RESULT_DIR}"
        )

    for path in result_files:

        try:
            doy_from_name = int(
                path.stem
            )
        except Exception:
            continue

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(
                    float(c[0])
                )

                doy = int(
                    float(c[1])
                )

                rh = float(
                    c[2]
                )

                sat = int(
                    float(c[3])
                )

                utc_hours = float(
                    c[4]
                )

                az = float(
                    c[5]
                )

                amp = float(
                    c[6]
                )

                emin = float(
                    c[7]
                )

                emax = float(
                    c[8]
                )

                nobs = int(
                    float(c[9])
                )

                freq = int(
                    float(c[10])
                )

                rise = int(
                    float(c[11])
                )

                pkn = float(
                    c[13]
                )

                delt = float(
                    c[14]
                )

            except Exception:
                continue

            # This experiment is GPS L1 only.
            if freq != 1:
                continue

            dt = utc_hours_to_datetime(
                year,
                doy,
                utc_hours,
            )

            wl = (
                H_ORTHO_M
                - rh
            )

            rows.append(
                {
                    "year": year,
                    "doy": doy,
                    "datetime_utc": dt,
                    "sat": sat,
                    "freq": freq,
                    "RH_m": rh,
                    "GNSS_WL_m": wl,
                    "az_deg": az,
                    "Amp": amp,
                    "PkNoise": pkn,
                    "eminO_deg": emin,
                    "emaxO_deg": emax,
                    "NumbOf": nobs,
                    "rise": rise,
                    "DelT_min": delt,
                    "source_file":
                        str(path),
                }
            )

    rows.sort(
        key=lambda r:
            r["datetime_utc"]
    )

    return rows


# ---------------------------------------------------------------------
# ADD TIDES
# ---------------------------------------------------------------------

def add_tides(
    rows,
    tide_at,
):
    for row in rows:

        for model in MODELS:
            row[model] = tide_at(
                row["datetime_utc"],
                model,
            )

        vals = [
            row[model]
            for model in MODELS
            if math.isfinite(
                row[model]
            )
        ]

        row[
            "TIDE_ENSEMBLE_m"
        ] = (
            float(
                np.mean(vals)
            )
            if vals
            else math.nan
        )

        row[
            "PRIMARY_TIDE_m"
        ] = row[
            PRIMARY_TIDE_MODEL
        ]

        row[
            "raw_GNSSR_minus_tide_m"
        ] = (
            row["GNSS_WL_m"]
            - row["PRIMARY_TIDE_m"]
            if math.isfinite(
                row["PRIMARY_TIDE_m"]
            )
            else math.nan
        )


# ---------------------------------------------------------------------
# TRACK ANALYSIS
# ---------------------------------------------------------------------

def group_tracks(rows):
    groups = defaultdict(list)

    for row in rows:
        groups[
            (
                row["sat"],
                row["freq"],
            )
        ].append(row)

    for key in groups:
        groups[key].sort(
            key=lambda r:
                r["datetime_utc"]
        )

    return groups


def analyze_track(
    key,
    group,
):
    wl = np.asarray(
        [
            r["GNSS_WL_m"]
            for r in group
        ],
        dtype=float,
    )

    tide = np.asarray(
        [
            r["PRIMARY_TIDE_m"]
            for r in group
        ],
        dtype=float,
    )

    az = np.asarray(
        [
            r["az_deg"]
            for r in group
        ],
        dtype=float,
    )

    pkn = np.asarray(
        [
            r["PkNoise"]
            for r in group
        ],
        dtype=float,
    )

    amp = np.asarray(
        [
            r["Amp"]
            for r in group
        ],
        dtype=float,
    )

    valid = (
        np.isfinite(wl)
        & np.isfinite(tide)
    )

    if np.sum(valid) >= 3:

        w = wl[valid]
        t = tide[valid]

        r = pearson(w, t)

        coeff = np.polyfit(
            t,
            w,
            1,
        )

        slope = float(
            coeff[0]
        )

        intercept = float(
            coeff[1]
        )

        C = float(
            np.mean(
                w - t
            )
        )

        residual = (
            w
            - C
            - t
        )

        unit_rms_cm = (
            rms(residual)
            * 100
        )

        free_rms_cm = (
            rms(
                w
                - (
                    slope * t
                    + intercept
                )
            )
            * 100
        )

        unit_mae_cm = (
            mae(residual)
            * 100
        )

    else:
        r = math.nan
        slope = math.nan
        intercept = math.nan
        C = math.nan
        unit_rms_cm = math.nan
        free_rms_cm = math.nan
        unit_mae_cm = math.nan

    az_std = float(
        np.std(az)
    )

    pkn_mean = float(
        np.mean(pkn)
    )

    amp_mean = float(
        np.mean(amp)
    )

    unique_days = sorted(
        {
            r0["doy"]
            for r0 in group
        }
    )

    duration_days = (
        (
            group[-1]["datetime_utc"]
            - group[0]["datetime_utc"]
        ).total_seconds()
        / 86400.0
    )

    # A transparent long-term diagnostic score.
    corr_score = (
        abs(r)
        if math.isfinite(r)
        else 0.0
    )

    slope_score = (
        max(
            0.0,
            1.0
            - min(
                1.0,
                abs(
                    slope - 1.0
                )
                if math.isfinite(slope)
                else 1.0,
            ),
        )
    )

    rms_score = (
        max(
            0.0,
            1.0
            - min(
                1.0,
                unit_rms_cm / 20.0,
            ),
        )
        if math.isfinite(
            unit_rms_cm
        )
        else 0.0
    )

    pkn_score = max(
        0.0,
        min(
            1.0,
            (pkn_mean - 2.8)
            / 1.2,
        ),
    )

    amp_score = max(
        0.0,
        min(
            1.0,
            amp_mean / 50.0,
        ),
    )

    az_score = max(
        0.0,
        min(
            1.0,
            1.0
            - az_std / 2.0,
        ),
    )

    persistence_score = max(
        0.0,
        min(
            1.0,
            duration_days / 20.0,
        ),
    )

    score = (
        0.25 * corr_score
        + 0.20 * slope_score
        + 0.15 * rms_score
        + 0.10 * pkn_score
        + 0.10 * amp_score
        + 0.10 * az_score
        + 0.10 * persistence_score
    )

    return {
        "sat": key[0],
        "freq": key[1],
        "track":
            f"SAT{key[0]}_FREQ{key[1]}",
        "n": len(group),
        "n_days":
            len(unique_days),
        "doy_first":
            min(unique_days),
        "doy_last":
            max(unique_days),
        "duration_days":
            duration_days,
        "az_mean_deg":
            float(np.mean(az)),
        "az_std_deg":
            az_std,
        "RH_mean_m":
            float(np.mean(wl)),
        "RH_sd_m":
            float(np.std(
                [
                    r0["RH_m"]
                    for r0 in group
                ]
            )),
        "PkNoise_mean":
            pkn_mean,
        "Amp_mean":
            amp_mean,
        "tide_r":
            r,
        "tide_slope":
            slope,
        "tide_intercept_m":
            intercept,
        "C_unit_slope_m":
            C,
        "unit_slope_RMS_cm":
            unit_rms_cm,
        "unit_slope_MAE_cm":
            unit_mae_cm,
        "free_fit_RMS_cm":
            free_rms_cm,
        "longterm_score":
            score,
    }


# ---------------------------------------------------------------------
# DAILY POPULATION
# ---------------------------------------------------------------------

def daily_population(rows):
    by_day = defaultdict(list)

    for row in rows:
        by_day[
            row["datetime_utc"].date()
        ].append(row)

    out = []

    for day, group in sorted(
        by_day.items()
    ):

        wl = np.asarray(
            [
                r["GNSS_WL_m"]
                for r in group
            ],
            dtype=float,
        )

        tide = np.asarray(
            [
                r["PRIMARY_TIDE_m"]
                for r in group
            ],
            dtype=float,
        )

        valid = (
            np.isfinite(wl)
            & np.isfinite(tide)
        )

        if np.sum(valid) >= 3:

            daily_corr = pearson(
                wl[valid],
                tide[valid],
            )

            daily_bias = float(
                np.median(
                    wl[valid]
                    - tide[valid]
                )
            )

        else:
            daily_corr = math.nan
            daily_bias = math.nan

        out.append(
            {
                "date": day,
                "n": len(group),
                "gnss_median":
                    float(
                        np.median(wl)
                    ),
                "gnss_mean":
                    float(
                        np.mean(wl)
                    ),
                "tide_at_obs_median":
                    float(
                        np.median(
                            tide
                        )
                    ),
                "daily_bias_median":
                    daily_bias,
                "daily_r":
                    daily_corr,
            }
        )

    return out


# ---------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------

def write_track_csv(
    records,
):
    fields = [
        "sat",
        "freq",
        "track",
        "n",
        "n_days",
        "doy_first",
        "doy_last",
        "duration_days",
        "az_mean_deg",
        "az_std_deg",
        "RH_mean_m",
        "RH_sd_m",
        "PkNoise_mean",
        "Amp_mean",
        "tide_r",
        "tide_slope",
        "tide_intercept_m",
        "C_unit_slope_m",
        "unit_slope_RMS_cm",
        "unit_slope_MAE_cm",
        "free_fit_RMS_cm",
        "longterm_score",
    ]

    with open(
        OUT_CSV,
        "w",
        newline="",
    ) as f:

        import csv

        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )

        writer.writeheader()

        for row in records:
            writer.writerow({
                field:
                    row.get(
                        field,
                        math.nan,
                    )
                for field in fields
            })


# ---------------------------------------------------------------------
# PLOTS
# ---------------------------------------------------------------------

def plot_all_gnssr_vs_tide(
    rows,
):
    PLOT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    rows_sorted = sorted(
        rows,
        key=lambda r:
            r["datetime_utc"],
    )

    times = [
        r["datetime_utc"]
        for r in rows_sorted
    ]

    gnss = np.asarray(
        [
            r["GNSS_WL_m"]
            for r in rows_sorted
        ],
        dtype=float,
    )

    tide = np.asarray(
        [
            r["PRIMARY_TIDE_m"]
            for r in rows_sorted
        ],
        dtype=float,
    )

    fig, ax = plt.subplots(
        figsize=(14, 7)
    )

    ax.plot(
        times,
        tide,
        linewidth=2,
        label=PRIMARY_TIDE_MODEL,
    )

    ax.scatter(
        times,
        gnss,
        s=14,
        alpha=0.55,
        label="All GPS L1 GNSS-R arcs",
    )

    ax.set_title(
        "Marconi GNSS-R Water Level vs Confirmed Tide Model"
    )

    ax.set_ylabel(
        "Water-level anomaly (m)"
    )

    ax.set_xlabel(
        "UTC"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "marconi_all_gnssr_vs_tide_timeseries.png",
        dpi=180,
    )

    plt.close(fig)


def plot_all_scatter(
    rows,
):
    x = np.asarray(
        [
            r["PRIMARY_TIDE_m"]
            for r in rows
        ],
        dtype=float,
    )

    y = np.asarray(
        [
            r["GNSS_WL_m"]
            for r in rows
        ],
        dtype=float,
    )

    valid = (
        np.isfinite(x)
        & np.isfinite(y)
    )

    x = x[valid]
    y = y[valid]

    if len(x) < 3:
        return

    coeff = np.polyfit(
        x,
        y,
        1,
    )

    xx = np.linspace(
        np.min(x),
        np.max(x),
        200,
    )

    yy = (
        coeff[0] * xx
        + coeff[1]
    )

    r = pearson(
        x,
        y,
    )

    fig, ax = plt.subplots(
        figsize=(8, 8)
    )

    ax.scatter(
        x,
        y,
        s=18,
        alpha=0.6,
    )

    ax.plot(
        xx,
        yy,
        linewidth=2,
        label=(
            f"fit slope={coeff[0]:.3f}, "
            f"r={r:.3f}"
        ),
    )

    ax.plot(
        xx,
        xx,
        linestyle="--",
        linewidth=1.5,
        label="1:1",
    )

    ax.set_title(
        "All GPS L1 GNSS-R Water Level vs Tide"
    )

    ax.set_xlabel(
        PRIMARY_TIDE_MODEL
        + " (m)"
    )

    ax.set_ylabel(
        "GNSS-R water level (m)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "marconi_all_gnssr_vs_tide_scatter.png",
        dpi=180,
    )

    plt.close(fig)


def plot_top_tracks(
    rows,
    track_records,
):
    ranked = [
        r
        for r in track_records
        if r["n"] >= MIN_TRACK_N
    ]

    ranked.sort(
        key=lambda r:
            r["longterm_score"],
        reverse=True,
    )

    selected = ranked[
        :TOP_TRACKS_FOR_PLOTS
    ]

    colors = plt.cm.tab10(
        np.linspace(
            0,
            1,
            len(selected),
        )
    )

    fig, ax = plt.subplots(
        figsize=(15, 8)
    )

    # Tide line on same figure.
    times = [
        r["datetime_utc"]
        for r in rows
    ]

    tide = np.asarray(
        [
            r["PRIMARY_TIDE_m"]
            for r in rows
        ],
        dtype=float,
    )

    valid_tide = np.isfinite(
        tide
    )

    if np.any(valid_tide):

        tide_times = [
            times[i]
            for i in np.where(
                valid_tide
            )[0]
        ]

        tide_values = tide[
            valid_tide
        ]

        ax.plot(
            tide_times,
            tide_values,
            color="black",
            linewidth=2.5,
            label=PRIMARY_TIDE_MODEL,
            zorder=1,
        )

    for color, tr in zip(
        colors,
        selected,
    ):

        subset = [
            row
            for row in rows
            if (
                row["sat"]
                == tr["sat"]
                and row["freq"]
                == tr["freq"]
            )
        ]

        subset.sort(
            key=lambda r:
                r["datetime_utc"]
        )

        t = [
            r["datetime_utc"]
            for r in subset
        ]

        y = np.asarray(
            [
                r["GNSS_WL_m"]
                for r in subset
            ],
            dtype=float,
        )

        ax.plot(
            t,
            y,
            marker="o",
            linewidth=1.4,
            markersize=4,
            color=color,
            label=(
                f"PRN {tr['sat']} "
                f"(N={tr['n']}, "
                f"r={tr['tide_r']:.3f})"
            ),
            zorder=2,
        )

    ax.set_title(
        "Top Repeated GNSS-R Tracks vs Confirmed Tide"
    )

    ax.set_ylabel(
        "Water-level anomaly (m)"
    )

    ax.set_xlabel(
        "UTC"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend(
        ncol=2,
        fontsize=8,
    )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "marconi_top_tracks_vs_tide.png",
        dpi=180,
    )

    plt.close(fig)


def plot_residuals(
    rows,
    track_records,
):
    ranked = sorted(
        [
            r
            for r in track_records
            if r["n"] >= MIN_TRACK_N
        ],
        key=lambda r:
            r["longterm_score"],
        reverse=True,
    )[:TOP_TRACKS_FOR_PLOTS]

    fig, ax = plt.subplots(
        figsize=(15, 8)
    )

    colors = plt.cm.tab10(
        np.linspace(
            0,
            1,
            len(ranked),
        )
    )

    for color, tr in zip(
        colors,
        ranked,
    ):

        C = tr[
            "C_unit_slope_m"
        ]

        subset = [
            r
            for r in rows
            if (
                r["sat"]
                == tr["sat"]
                and r["freq"]
                == tr["freq"]
            )
        ]

        subset.sort(
            key=lambda r:
                r["datetime_utc"]
        )

        times = [
            r["datetime_utc"]
            for r in subset
        ]

        residual = np.asarray(
            [
                r["GNSS_WL_m"]
                - C
                - r["PRIMARY_TIDE_m"]
                for r in subset
            ],
            dtype=float,
        )

        ax.plot(
            times,
            residual * 100,
            marker="o",
            markersize=3,
            linewidth=1,
            color=color,
            label=(
                f"PRN {tr['sat']} "
                f"RMS={tr['unit_slope_RMS_cm']:.1f} cm"
            ),
        )

    ax.axhline(
        0,
        linewidth=1.5,
    )

    ax.set_title(
        "Top GNSS-R Tracks: Unit-Slope Residuals"
    )

    ax.set_ylabel(
        "GNSS-R - tide residual (cm)"
    )

    ax.set_xlabel(
        "UTC"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend(
        ncol=2,
        fontsize=8,
    )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "marconi_top_track_residuals.png",
        dpi=180,
    )

    plt.close(fig)


def plot_calibration_stability(
    rows,
    track_records,
):
    ranked = sorted(
        [
            r
            for r in track_records
            if r["n"] >= MIN_TRACK_N
        ],
        key=lambda r:
            r["longterm_score"],
        reverse=True,
    )[:TOP_TRACKS_FOR_PLOTS]

    fig, ax = plt.subplots(
        figsize=(15, 8)
    )

    colors = plt.cm.tab10(
        np.linspace(
            0,
            1,
            len(ranked),
        )
    )

    for color, tr in zip(
        colors,
        ranked,
    ):

        subset = [
            r
            for r in rows
            if (
                r["sat"]
                == tr["sat"]
                and r["freq"]
                == tr["freq"]
            )
        ]

        subset.sort(
            key=lambda r:
                r["datetime_utc"]
        )

        times = [
            r["datetime_utc"]
            for r in subset
        ]

        c = np.asarray(
            [
                r["GNSS_WL_m"]
                - r["PRIMARY_TIDE_m"]
                for r in subset
            ],
            dtype=float,
        )

        ax.plot(
            times,
            c * 100,
            marker="o",
            markersize=3,
            linewidth=1,
            color=color,
            label=(
                f"PRN {tr['sat']}"
            ),
        )

    ax.axhline(
        0,
        linewidth=1.5,
    )

    ax.set_title(
        "GNSS-R Minus Tide: Track Calibration Offset"
    )

    ax.set_ylabel(
        "GNSS-R water level - tide (cm)"
    )

    ax.set_xlabel(
        "UTC"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend(
        ncol=2,
        fontsize=8,
    )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "marconi_track_calibration_stability.png",
        dpi=180,
    )

    plt.close(fig)


def plot_daily_population(
    daily,
):
    if not daily:
        return

    times = [
        r["date"]
        for r in daily
    ]

    tide = np.asarray(
        [
            r["tide_at_obs_median"]
            for r in daily
        ],
        dtype=float,
    )

    gnss = np.asarray(
        [
            r["gnss_median"]
            for r in daily
        ],
        dtype=float,
    )

    fig, ax = plt.subplots(
        figsize=(14, 7)
    )

    ax.plot(
        times,
        tide,
        linewidth=2,
        label=PRIMARY_TIDE_MODEL,
    )

    ax.plot(
        times,
        gnss,
        marker="o",
        linewidth=1.2,
        label="Daily GNSS-R median",
    )

    ax.set_title(
        "Daily GNSS-R Population Median vs Tide"
    )

    ax.set_ylabel(
        "Water-level anomaly (m)"
    )

    ax.set_xlabel(
        "Date"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "marconi_daily_population_vs_tide.png",
        dpi=180,
    )

    plt.close(fig)


# ---------------------------------------------------------------------
# SUMMARY
# ---------------------------------------------------------------------

def write_summary(
    track_records,
    rows,
    daily,
):
    ranked = sorted(
        track_records,
        key=lambda r:
            r["longterm_score"],
        reverse=True,
    )

    lines = []

    lines.append(
        "MARCONI LONG-TERM GNSS-R STABILITY / TIDE ANALYSIS"
    )
    lines.append(
        "=" * 100
    )

    lines.append(
        f"Primary tide model: {PRIMARY_TIDE_MODEL}"
    )

    lines.append(
        f"Station H_ortho: {H_ORTHO_M:.3f} m"
    )

    lines.append(
        f"GNSS-R result directory: {RESULT_DIR}"
    )

    if rows:
        lines.append(
            f"GNSS-R observations: {len(rows)}"
        )

        lines.append(
            f"Observation interval: "
            f"{min(r['datetime_utc'] for r in rows)} "
            f"through "
            f"{max(r['datetime_utc'] for r in rows)}"
        )

    lines.append(
        ""
    )

    lines.append(
        "TOP LONG-TERM TRACKS"
    )
    lines.append(
        "-" * 100
    )

    for i, r in enumerate(
        ranked[:20],
        start=1,
    ):

        lines.append(
            f"{i:2d}. "
            f"SAT={r['sat']:3d} "
            f"FREQ={r['freq']:3d} "
            f"N={r['n']:3d} "
            f"days={r['n_days']:2d} "
            f"Az={r['az_mean_deg']:6.2f}±{r['az_std_deg']:.2f} "
            f"r={r['tide_r']:+.4f} "
            f"slope={r['tide_slope']:+.4f} "
            f"C={r['C_unit_slope_m']:+.3f} "
            f"unitRMS={r['unit_slope_RMS_cm']:.2f}cm "
            f"score={r['longterm_score']:.3f}"
        )

    lines.append(
        ""
    )

    lines.append(
        "INTERPRETATION"
    )
    lines.append(
        "The primary diagnostic water-level estimate is H_ortho - RH."
    )
    lines.append(
        "The unit-slope calibration constant C is computed separately"
    )
    lines.append(
        "for each track as mean(GNSS_WL - primary tide)."
    )
    lines.append(
        "This is a track-diagnostic statistic, not a production datum"
    )
    lines.append(
        "calibration. Track rankings should be confirmed against physical"
    )
    lines.append(
        "Fresnel/topobathy geometry before being called ocean tracks."
    )

    lines.append(
        ""
    )
    lines.append(
        "PLOTS"
    )
    lines.append(
        "  marconi_all_gnssr_vs_tide_timeseries.png"
    )
    lines.append(
        "  marconi_all_gnssr_vs_tide_scatter.png"
    )
    lines.append(
        "  marconi_top_tracks_vs_tide.png"
    )
    lines.append(
        "  marconi_top_track_residuals.png"
    )
    lines.append(
        "  marconi_track_calibration_stability.png"
    )
    lines.append(
        "  marconi_daily_population_vs_tide.png"
    )

    OUT_SUMMARY.write_text(
        "\n".join(lines)
        + "\n"
    )


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():
    print()
    print("=" * 100)
    print(
        "MARCONI LONG-TERM GNSS-R STABILITY + TIDE ANALYSIS"
    )
    print("=" * 100)

    print(
        "Primary tide model:",
        PRIMARY_TIDE_MODEL,
    )

    print(
        "Result directory:",
        RESULT_DIR,
    )

    tide_times, tide_at = (
        load_tide_data()
    )

    rows = load_gnss_results()

    add_tides(
        rows,
        tide_at,
    )

    print(
        "GNSS-R observations:",
        len(rows),
    )

    if rows:
        print(
            "Observation interval:",
            min(
                r["datetime_utc"]
                for r in rows
            ),
            "through",
            max(
                r["datetime_utc"]
                for r in rows
            ),
        )

    groups = group_tracks(
        rows
    )

    repeated = {
        key: group
        for key, group in groups.items()
        if len(group) >= MIN_TRACK_N
    }

    print(
        "Unique satellite/frequency tracks:",
        len(groups),
    )

    print(
        f"Repeated tracks N >= {MIN_TRACK_N}:",
        len(repeated),
    )

    track_records = [
        analyze_track(
            key,
            group,
        )
        for key, group in repeated.items()
    ]

    track_records.sort(
        key=lambda r:
            r["longterm_score"],
        reverse=True,
    )

    print()
    print(
        "=" * 100
    )
    print(
        "TOP LONG-TERM TRACKS"
    )
    print(
        "=" * 100
    )

    for i, r in enumerate(
        track_records[:20],
        start=1,
    ):
        print(
            f"{i:2d} "
            f"SAT={r['sat']:3d} "
            f"N={r['n']:3d} "
            f"days={r['n_days']:2d} "
            f"Az={r['az_mean_deg']:6.2f} "
            f"r={r['tide_r']:+.4f} "
            f"slope={r['tide_slope']:+.4f} "
            f"RMS={r['unit_slope_RMS_cm']:.2f}cm "
            f"C={r['C_unit_slope_m']:+.3f} "
            f"score={r['longterm_score']:.3f}"
        )

    write_track_csv(
        track_records
    )

    daily = daily_population(
        rows
    )

    write_summary(
        track_records,
        rows,
        daily,
    )

    print()
    print(
        "Generating plots..."
    )

    plot_all_gnssr_vs_tide_timeseries(
        rows
    )

    plot_all_scatter(
        rows
    )

    plot_top_tracks(
        rows,
        track_records,
    )

    plot_residuals(
        rows,
        track_records,
    )

    plot_calibration_stability(
        rows,
        track_records,
    )

    plot_daily_population(
        daily
    )

    print()
    print(
        "=" * 100
    )
    print(
        "OUTPUTS"
    )
    print(
        "=" * 100
    )

    print(
        "CSV:",
        OUT_CSV.resolve()
    )

    print(
        "Summary:",
        OUT_SUMMARY.resolve()
    )

    print(
        "Plots:",
        PLOT_DIR.resolve()
    )

    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
