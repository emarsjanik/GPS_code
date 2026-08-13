#!/usr/bin/env python3
"""
Marconi GNSS-R leave-one-track-out cross-validation + all-track ranking.

Purpose
-------
This is the next validation stage after the successful 17-23 m GPS L1
experiment.

It performs TWO analyses:

A) LEAVE-ONE-TRACK-OUT CROSS-VALIDATION
---------------------------------------
Tracks:
  PRN 26 / L1 / ~91°
  PRN 21 / L1 / ~93°
  PRN 16 / L1 / ~96°

For each held-out track:
  * estimate one calibration constant C from the OTHER TWO TRACKS
  * apply C without refitting to the held-out track
  * report correlation, free slope, bias, RMS, MAE
  * compare each of the four tide models and the ensemble

This tests whether a common vertical calibration transfers between
independent satellite geometries.

B) ALL-TRACK CANDIDATE RANKING
------------------------------
Scans all GPS L1 successful arcs in the current:
    ocean17_23_l1_e5_13
result files for DOY 204-207.

Ranks repeated satellite/frequency tracks using:
  * number of observations
  * azimuth stability
  * RH stability
  * PkNoise
  * amplitude
  * RH-vs-tide correlation
  * free slope closeness to +1 for GNSS water level
  * fixed-calibration RMS
  * topobathy 13-degree wet fraction, if
    marconi_topobathy_fresnel_validation.csv exists

The ranking is diagnostic. It does NOT alter production settings.

Inputs
------
GNSS-IR:
  products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13/{204..207}.txt

Tide workbook:
  marconi_tides_sherwood.xlsx

Optional geometry validation:
  marconi_topobathy_fresnel_validation.csv

Station orthometric height:
  18.665 m

Outputs
-------
marconi_leave_one_track_out.csv
marconi_leave_one_track_out_summary.txt
marconi_all_track_ranking.csv
marconi_all_track_ranking_summary.txt
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

TOPO_CSV = Path(
    "marconi_topobathy_fresnel_validation.csv"
)

DOYS = [204, 205, 206, 207]

H_ORTHO_M = 18.665

MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]

# Initial high-confidence validation tracks.
CORE_TRACKS = {
    (26, 1): "PRN26_AZ91",
    (21, 1): "PRN21_AZ93",
    (16, 1): "PRN16_AZ96",
}

# Minimum observations for a repeated-track candidate.
MIN_TRACK_N = 3

# For ranking, we use the most localized Fresnel validation available.
GEOMETRY_ELEVATION = 13.0


# ---------------------------------------------------------------------
# MATH HELPERS
# ---------------------------------------------------------------------

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 3:
        return math.nan

    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(
        np.corrcoef(x, y)[0, 1]
    )


def rms(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(
        np.sqrt(
            np.mean(x ** 2)
        )
    )


def mae(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(
        np.mean(
            np.abs(x)
        )
    )


def utc_hours_to_datetime(year, doy, utc_hours):
    return (
        datetime(
            year,
            1,
            1,
        )
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


# ---------------------------------------------------------------------
# TIDES
# ---------------------------------------------------------------------

def load_tides():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [
        cell.value
        for cell in ws[1]
    ]

    time_col = header.index(
        "time"
    )

    model_cols = {
        model:
        header.index(model)
        for model in MODELS
    }

    times = []
    values = {
        model: []
        for model in MODELS
    }

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_col]

        if not isinstance(
            t,
            datetime,
        ):
            continue

        row_values = {}
        good = True

        for model in MODELS:

            v = finite(
                row[
                    model_cols[model]
                ]
            )

            if v is None:
                good = False
                break

            row_values[
                model
            ] = v

        if not good:
            continue

        times.append(t)

        for model in MODELS:
            values[
                model
            ].append(
                row_values[
                    model
                ]
            )

    wb.close()

    if len(times) < 2:
        raise RuntimeError(
            "Insufficient tide-model points."
        )

    epoch = np.asarray(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    arrays = {
        model:
        np.asarray(
            values[model],
            dtype=float,
        )
        for model in MODELS
    }

    def tide_at(
        dt,
        model,
    ):
        x = (
            dt - times[0]
        ).total_seconds()

        if (
            x < epoch[0]
            or x > epoch[-1]
        ):
            return math.nan

        return float(
            np.interp(
                x,
                epoch,
                arrays[model],
            )
        )

    return tide_at


# ---------------------------------------------------------------------
# GNSS-IR RESULT PARSING
# ---------------------------------------------------------------------

def load_gnss_rows():
    rows = []

    for doy in DOYS:

        path = (
            RESULT_DIR
            / f"{doy}.txt"
        )

        if not path.exists():
            raise SystemExit(
                f"Missing result file: {path}"
            )

        for line in path.read_text(
            errors="replace",
        ).splitlines():

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(
                    float(c[0])
                )
                doy2 = int(
                    float(c[1])
                )
                rh = float(
                    c[2]
                )
                sat = int(
                    float(c[3])
                )
                utc_hours = float(
                    c[4]
                )
                az = float(
                    c[5]
                )
                amp = float(
                    c[6]
                )
                emin = float(
                    c[7]
                )
                emax = float(
                    c[8]
                )
                nobs = int(
                    float(c[9])
                )
                freq = int(
                    float(c[10])
                )
                rise = int(
                    float(c[11])
                )
                edot = float(
                    c[12]
                )
                pkn = float(
                    c[13]
                )
                delt = float(
                    c[14]
                )
                mjd = float(
                    c[15]
                )
                refr = float(
                    c[16]
                )
            except Exception:
                continue

            dt = utc_hours_to_datetime(
                year,
                doy2,
                utc_hours,
            )

            wl = (
                H_ORTHO_M
                - rh
            )

            rows.append(
                {
                    "year": year,
                    "doy": doy2,
                    "datetime_utc": dt,
                    "sat": sat,
                    "freq": freq,
                    "RH_m": rh,
                    "GNSS_water_level_m": wl,
                    "az_deg": az,
                    "Amp": amp,
                    "PkNoise": pkn,
                    "eminO_deg": emin,
                    "emaxO_deg": emax,
                    "NumbOf": nobs,
                    "rise": rise,
                    "EdotF": edot,
                    "DelT_min": delt,
                    "MJD": mjd,
                    "refr": refr,
                }
            )

    rows.sort(
        key=lambda r: (
            r["sat"],
            r["freq"],
            r["datetime_utc"],
        )
    )

    return rows


def add_tide_values(rows, tide_at):
    for row in rows:
        for model in MODELS:
            row[model] = tide_at(
                row["datetime_utc"],
                model,
            )

        vals = [
            row[m]
            for m in MODELS
            if math.isfinite(
                row[m]
            )
        ]

        row[
            "TIDE_ENSEMBLE_m"
        ] = (
            float(
                np.mean(vals)
            )
            if vals
            else math.nan
        )


# ---------------------------------------------------------------------
# OPTIONAL TOPOBATHY GEOMETRY
# ---------------------------------------------------------------------

def load_geometry_scores():
    """
    Return mean 13-degree wet fraction by satellite/frequency.

    The topobathy CSV contains rows for each tide model and the ensemble.
    We use only:
        tide_model = ENSEMBLE_MEAN
        test_elevation_deg = 13
    """

    result = {}

    if not TOPO_CSV.exists():
        return result

    with open(
        TOPO_CSV,
        newline="",
    ) as f:
        reader = csv.DictReader(f)

        groups = defaultdict(list)

        for row in reader:

            try:
                elev = float(
                    row[
                        "test_elevation_deg"
                    ]
                )

                sat = int(
                    row["sat"]
                )

                freq = int(
                    row["freq"]
                )

                model = row[
                    "tide_model"
                ]

                wet = float(
                    row["wet_fraction"]
                )

            except Exception:
                continue

            if (
                abs(
                    elev
                    - GEOMETRY_ELEVATION
                )
                > 1e-9
            ):
                continue

            if model != "ENSEMBLE_MEAN":
                continue

            if not math.isfinite(wet):
                continue

            groups[
                (sat, freq)
            ].append(
                wet
            )

    for key, values in groups.items():
        if values:
            result[key] = {
                "geometry_wet_mean":
                    float(
                        np.mean(
                            values
                        )
                    ),
                "geometry_wet_min":
                    float(
                        np.min(
                            values
                        )
                    ),
                "geometry_wet_max":
                    float(
                        np.max(
                            values
                        )
                    ),
                "geometry_n":
                    len(values),
            }

    return result


# ---------------------------------------------------------------------
# TRACK METRICS
# ---------------------------------------------------------------------

def track_groups(rows):
    groups = defaultdict(list)

    for row in rows:
        groups[
            (
                row["sat"],
                row["freq"],
            )
        ].append(row)

    for key in groups:
        groups[key].sort(
            key=lambda r:
                r["datetime_utc"]
        )

    return groups


def fixed_calibration_metrics(
    calibration_rows,
    test_rows,
    model,
):
    """
    Fit C from calibration rows only:
        GNSS_WL - tide = C

    Then apply fixed C to test rows.
    """
    diffs = np.array(
        [
            r["GNSS_water_level_m"]
            - r[model]
            for r in calibration_rows
            if math.isfinite(
                r[model]
            )
        ],
        dtype=float,
    )

    if len(diffs) == 0:
        return {
            "C_m": math.nan,
            "r": math.nan,
            "slope": math.nan,
            "bias_m": math.nan,
            "rms_cm": math.nan,
            "mae_cm": math.nan,
            "n": len(test_rows),
        }

    C = float(
        np.mean(diffs)
    )

    gnss = np.array(
        [
            r["GNSS_water_level_m"]
            for r in test_rows
        ],
        dtype=float,
    )

    tide = np.array(
        [
            r[model]
            for r in test_rows
        ],
        dtype=float,
    )

    valid = (
        np.isfinite(gnss)
        & np.isfinite(tide)
    )

    gnss = gnss[valid]
    tide = tide[valid]

    if len(gnss) < 3:
        return {
            "C_m": C,
            "r": math.nan,
            "slope": math.nan,
            "bias_m": math.nan,
            "rms_cm": math.nan,
            "mae_cm": math.nan,
            "n": len(gnss),
        }

    residual = (
        gnss
        - C
        - tide
    )

    return {
        "C_m": C,
        "r": pearson(
            gnss,
            tide,
        ),
        "slope": float(
            np.polyfit(
                tide,
                gnss,
                1,
            )[0]
        ),
        "bias_m": float(
            np.mean(
                residual
            )
        ),
        "rms_cm": rms(
            residual
        ) * 100.0,
        "mae_cm": mae(
            residual
        ) * 100.0,
        "n": len(gnss),
    }


def all_track_metrics(
    rows,
    geometry_scores,
):
    groups = track_groups(rows)

    out = []

    for key, subset in groups.items():

        if len(subset) < MIN_TRACK_N:
            continue

        sat, freq = key

        RH = np.array(
            [
                r["RH_m"]
                for r in subset
            ],
            dtype=float,
        )

        WL = np.array(
            [
                r["GNSS_water_level_m"]
                for r in subset
            ],
            dtype=float,
        )

        az = np.array(
            [
                r["az_deg"]
                for r in subset
            ],
            dtype=float,
        )

        pkn = np.array(
            [
                r["PkNoise"]
                for r in subset
            ],
            dtype=float,
        )

        amp = np.array(
            [
                r["Amp"]
                for r in subset
            ],
            dtype=float,
        )

        tide = np.array(
            [
                r["TIDE_ENSEMBLE_m"]
                for r in subset
            ],
            dtype=float,
        )

        valid = (
            np.isfinite(WL)
            & np.isfinite(tide)
        )

        if np.sum(valid) >= 3:

            corr = pearson(
                WL[valid],
                tide[valid],
            )

            slope = float(
                np.polyfit(
                    tide[valid],
                    WL[valid],
                    1,
                )[0]
            )

            free_rms = rms(
                (
                    WL[valid]
                    - (
                        np.polyval(
                            np.polyfit(
                                tide[valid],
                                WL[valid],
                                1,
                            ),
                            tide[valid],
                        )
                    )
                )
            ) * 100.0

        else:
            corr = math.nan
            slope = math.nan
            free_rms = math.nan

        # Repeatability: same satellite/frequency over days.
        az_std = float(
            np.std(az)
        )

        rh_std = float(
            np.std(RH)
        )

        # High-quality signal statistics.
        mean_pkn = float(
            np.mean(pkn)
        )

        mean_amp = float(
            np.mean(amp)
        )

        # Common absolute calibration diagnostic:
        # C from THIS track only, not used for the all-track rank.
        C = float(
            np.mean(
                WL[valid]
                - tide[valid]
            )
        )

        residual = (
            WL[valid]
            - C
            - tide[valid]
        )

        fixed_rms = (
            rms(residual)
            * 100.0
        )

        g = geometry_scores.get(
            key,
            {}
        )

        wet = g.get(
            "geometry_wet_mean",
            math.nan,
        )

        # Ranking score:
        #
        # 30% tide correlation
        # 20% closeness of slope to +1
        # 15% geometry water fraction
        # 15% PkNoise
        # 10% azimuth repeatability
        # 10% RH/tide fixed-slope residual
        #
        # This is deliberately transparent and diagnostic.
        corr_score = (
            max(
                0.0,
                min(
                    1.0,
                    abs(corr)
                    if math.isfinite(corr)
                    else 0.0,
                ),
            )
        )

        slope_score = (
            max(
                0.0,
                1.0
                - min(
                    1.0,
                    abs(
                        slope
                        - 1.0
                    )
                    if math.isfinite(
                        slope
                    )
                    else 1.0,
                ),
            )
        )

        if math.isfinite(wet):
            geometry_score = max(
                0.0,
                min(
                    1.0,
                    wet,
                ),
            )
        else:
            geometry_score = 0.5

        pkn_score = max(
            0.0,
            min(
                1.0,
                (
                    mean_pkn
                    - 2.8
                )
                / 2.0,
            ),
        )

        az_score = max(
            0.0,
            min(
                1.0,
                1.0
                - (
                    az_std
                    / 2.0
                ),
            ),
        )

        fixed_score = max(
            0.0,
            min(
                1.0,
                1.0
                - (
                    fixed_rms
                    / 20.0
                ),
            ),
        )

        score = (
            0.30 * corr_score
            + 0.20 * slope_score
            + 0.15 * geometry_score
            + 0.15 * pkn_score
            + 0.10 * az_score
            + 0.10 * fixed_score
        )

        out.append(
            {
                "sat": sat,
                "freq": freq,
                "track": (
                    f"SAT{sat}_FREQ{freq}"
                ),
                "n": len(subset),
                "doy_first": subset[0]["doy"],
                "doy_last": subset[-1]["doy"],
                "az_mean_deg":
                    float(np.mean(az)),
                "az_std_deg":
                    az_std,
                "RH_mean_m":
                    float(np.mean(RH)),
                "RH_std_m":
                    rh_std,
                "PkNoise_mean":
                    mean_pkn,
                "Amp_mean":
                    mean_amp,
                "tide_corr_r":
                    corr,
                "free_slope":
                    slope,
                "free_fit_RMS_cm":
                    free_rms,
                "track_C_m":
                    C,
                "fixed_unit_slope_RMS_cm":
                    fixed_rms,
                "geometry13_wet_mean":
                    wet,
                "geometry13_wet_min":
                    g.get(
                        "geometry_wet_min",
                        math.nan,
                    ),
                "geometry13_wet_max":
                    g.get(
                        "geometry_wet_max",
                        math.nan,
                    ),
                "diagnostic_score":
                    score,
            }
        )

    out.sort(
        key=lambda r:
            r["diagnostic_score"],
        reverse=True,
    )

    return out


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():

    print()
    print("=" * 100)
    print(
        "MARCONI LEAVE-ONE-TRACK-OUT + ALL-TRACK GNSS-R VALIDATION"
    )
    print("=" * 100)

    tide_at = load_tides()

    rows = load_gnss_rows()

    if not rows:
        raise SystemExit(
            "No GNSS-R result rows found."
        )

    add_tide_values(
        rows,
        tide_at,
    )

    print(
        f"GNSS-R rows loaded: {len(rows)}"
    )

    geometry_scores = (
        load_geometry_scores()
    )

    if geometry_scores:
        print(
            f"Geometry scores loaded for "
            f"{len(geometry_scores)} satellite/frequency tracks."
        )
    else:
        print(
            "No topobathy geometry CSV found; "
            "ranking will use neutral geometry score."
        )

    groups = track_groups(
        rows
    )

    print(
        f"Unique satellite/frequency tracks: {len(groups)}"
    )

    # -----------------------------------------------------------------
    # LEAVE-ONE-TRACK-OUT
    # -----------------------------------------------------------------

    print()
    print("=" * 100)
    print(
        "LEAVE-ONE-TRACK-OUT CROSS-VALIDATION"
    )
    print("=" * 100)

    loo_rows = []

    for held_key, held_label in CORE_TRACKS.items():

        held = groups.get(
            held_key,
            [],
        )

        calibration = []

        for key in CORE_TRACKS:

            if key == held_key:
                continue

            calibration.extend(
                groups.get(
                    key,
                    [],
                )
            )

        print()
        print(
            f"HOLD OUT: {held_label}"
        )

        print(
            "Calibration tracks:"
        )

        for key in CORE_TRACKS:

            if key == held_key:
                continue

            print(
                f"  {CORE_TRACKS[key]} "
                f"N={len(groups.get(key, []))}"
            )

        print(
            f"Validation N={len(held)}"
        )

        for model in MODELS + [
            "TIDE_ENSEMBLE_m"
        ]:

            metrics = fixed_calibration_metrics(
                calibration,
                held,
                model,
            )

            print(
                f"  {model:18s}"
                f" C={metrics['C_m']:+.4f} m"
                f" r={metrics['r']:+.4f}"
                f" slope={metrics['slope']:+.4f}"
                f" RMS={metrics['rms_cm']:.2f} cm"
                f" MAE={metrics['mae_cm']:.2f} cm"
            )

            loo_rows.append(
                {
                    "held_out_track":
                        held_label,
                    "held_sat":
                        held_key[0],
                    "held_freq":
                        held_key[1],
                    "calibration_tracks":
                        ";".join(
                            CORE_TRACKS[k]
                            for k in CORE_TRACKS
                            if k != held_key
                        ),
                    "model":
                        model,
                    "n_calibration":
                        len(calibration),
                    "n_validation":
                        metrics["n"],
                    "calibration_C_m":
                        metrics["C_m"],
                    "validation_r":
                        metrics["r"],
                    "validation_slope":
                        metrics["slope"],
                    "validation_bias_m":
                        metrics["bias_m"],
                    "validation_RMS_cm":
                        metrics["rms_cm"],
                    "validation_MAE_cm":
                        metrics["mae_cm"],
                }
            )

    with open(
        "marconi_leave_one_track_out.csv",
        "w",
        newline="",
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=list(
                loo_rows[0].keys()
            ),
        )

        writer.writeheader()
        writer.writerows(
            loo_rows
        )

    # Summary.
    loo_summary = []

    loo_summary.append(
        "MARCONI LEAVE-ONE-TRACK-OUT CROSS-VALIDATION"
    )
    loo_summary.append(
        "=" * 100
    )
    loo_summary.append(
        "Calibration is estimated from two tracks and applied unchanged"
    )
    loo_summary.append(
        "to the held-out third track."
    )
    loo_summary.append(
        ""
    )

    for row in loo_rows:
        loo_summary.append(
            f"HOLD {row['held_out_track']:30s}"
            f" {row['model']:18s}"
            f" C={row['calibration_C_m']:+.4f} m"
            f" r={row['validation_r']:+.4f}"
            f" slope={row['validation_slope']:+.4f}"
            f" RMS={row['validation_RMS_cm']:.2f} cm"
            f" MAE={row['validation_MAE_cm']:.2f} cm"
        )

    Path(
        "marconi_leave_one_track_out_summary.txt"
    ).write_text(
        "\n".join(loo_summary)
        + "\n"
    )

    # -----------------------------------------------------------------
    # ALL TRACK RANKING
    # -----------------------------------------------------------------

    ranking = all_track_metrics(
        rows,
        geometry_scores,
    )

    with open(
        "marconi_all_track_ranking.csv",
        "w",
        newline="",
    ) as f:

        fields = list(
            ranking[0].keys()
        )

        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )

        writer.writeheader()
        writer.writerows(
            ranking
        )

    rank_summary = []

    rank_summary.append(
        "MARCONI ALL-TRACK GNSS-R CANDIDATE RANKING"
    )
    rank_summary.append(
        "=" * 100
    )
    rank_summary.append(
        "Only repeated satellite/frequency tracks with N >= "
        f"{MIN_TRACK_N} are ranked."
    )
    rank_summary.append(
        "Ranking uses tide response, slope closeness, quality, azimuth"
    )
    rank_summary.append(
        "repeatability, fixed-slope residual, and optional 13-degree"
    )
    rank_summary.append(
        "topobathy wet fraction."
    )
    rank_summary.append(
        ""
    )

    rank_summary.append(
        "TOP 25 TRACKS"
    )
    rank_summary.append(
        "-" * 100
    )

    for i, r in enumerate(
        ranking[:25],
        start=1,
    ):

        rank_summary.append(
            f"{i:2d}. "
            f"SAT={r['sat']:3d} "
            f"FREQ={r['freq']:3d} "
            f"N={r['n']:2d} "
            f"Az={r['az_mean_deg']:6.2f}±{r['az_std_deg']:.2f} "
            f"RH={r['RH_mean_m']:.3f}±{r['RH_std_m']:.3f} "
            f"r={r['tide_corr_r']:+.4f} "
            f"slope={r['free_slope']:+.4f} "
            f"PkN={r['PkNoise_mean']:.2f} "
            f"Amp={r['Amp_mean']:.1f} "
            f"wet13={r['geometry13_wet_mean']:.3f} "
            f"RMS={r['fixed_unit_slope_RMS_cm']:.2f}cm "
            f"score={r['diagnostic_score']:.3f}"
        )

    Path(
        "marconi_all_track_ranking_summary.txt"
    ).write_text(
        "\n".join(rank_summary)
        + "\n"
    )

    # -----------------------------------------------------------------
    # Console final
    # -----------------------------------------------------------------

    print()
    print("=" * 100)
    print(
        "TOP ALL-TRACK CANDIDATES"
    )
    print("=" * 100)

    for i, r in enumerate(
        ranking[:20],
        start=1,
    ):

        print(
            f"{i:2d} "
            f"SAT={r['sat']:3d} "
            f"FREQ={r['freq']:3d} "
            f"N={r['n']:2d} "
            f"Az={r['az_mean_deg']:6.2f} "
            f"r={r['tide_corr_r']:+.4f} "
            f"slope={r['free_slope']:+.4f} "
            f"wet13={r['geometry13_wet_mean']:.3f} "
            f"RMS={r['fixed_unit_slope_RMS_cm']:.2f}cm "
            f"score={r['diagnostic_score']:.3f}"
        )

    print()
    print(
        "Outputs:"
    )
    print(
        "  marconi_leave_one_track_out.csv"
    )
    print(
        "  marconi_leave_one_track_out_summary.txt"
    )
    print(
        "  marconi_all_track_ranking.csv"
    )
    print(
        "  marconi_all_track_ranking_summary.txt"
    )

    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
