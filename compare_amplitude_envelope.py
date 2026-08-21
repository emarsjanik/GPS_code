#!/usr/bin/env python3
"""
compare_amplitude_envelope.py

Computes a rolling amplitude envelope (a proxy for the spring-neap
tidal cycle) for both the GNSS-IR spline output and the tide model,
and plots them together -- to check directly whether the two
series' spring-neap cycles are actually synchronized in timing and
magnitude, rather than trying to eyeball this from a raw overlay
plot where fast tidal oscillation makes slow amplitude modulation
hard to see directly.

The envelope is computed as a rolling standard deviation over a
window sized to span roughly one full semidiurnal tidal cycle
(~12.4 hours) but short compared to the ~14-day spring-neap cycle,
so it tracks the spring-neap envelope without being swamped by the
individual tidal oscillation itself.

Usage:
    python3 compare_amplitude_envelope.py \\
        --spline-file products/refl_code/Files/usgs/usgs_spline_out.txt \\
        --tide-file marconi_tides_sherwood.xlsx \\
        --tide-value-col EOT20_heightm \\
        --output amplitude_envelope.png
"""

from __future__ import annotations

import argparse
import math
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np


def load_spline_output(path: Path):
    times, values = [], []
    with open(path, errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            cols = line.split()
            if len(cols) < 9:
                continue
            try:
                year = int(float(cols[2]))
                month = int(float(cols[3]))
                day = int(float(cols[4]))
                hour = int(float(cols[5]))
                minute = int(float(cols[6]))
                second = int(float(cols[7]))
                water_level = float(cols[8])
            except (ValueError, IndexError):
                continue
            try:
                dt = datetime(year, month, day, hour, minute, second)
            except ValueError:
                continue
            times.append(dt)
            values.append(water_level)
    return times, np.asarray(values, float)


def load_tide_reference(path: Path, time_col: str, value_col: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    time_idx = header.index(time_col)
    value_idx = header.index(value_col)

    times, values = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[time_idx]
        if not isinstance(t, datetime):
            continue
        try:
            v = float(row[value_idx])
        except (TypeError, ValueError):
            continue
        if not math.isfinite(v):
            continue
        times.append(t)
        values.append(v)
    wb.close()
    return times, np.asarray(values, float)


def rolling_amplitude_envelope(times, values, window_hours=24.0, step_hours=3.0):
    """
    Computes a rolling amplitude proxy: for each step, the (max-min)
    of all points within +/- window_hours/2 of that step's center
    time. A window of ~24h comfortably spans 1-2 full semidiurnal
    cycles, smoothing out individual-cycle noise while still
    resolving the much slower ~14-day spring-neap modulation.
    """
    if not times:
        return [], []

    t0 = times[0]
    epoch = np.array([(t - t0).total_seconds() / 3600.0 for t in times])  # hours

    start = epoch.min()
    end = epoch.max()

    step_centers = np.arange(start, end, step_hours)
    envelope_times, envelope_values = [], []

    half_window = window_hours / 2.0

    for center in step_centers:
        mask = (epoch >= center - half_window) & (epoch <= center + half_window)
        if np.sum(mask) < 4:
            continue
        window_vals = values[mask]
        envelope_values.append(float(window_vals.max() - window_vals.min()))
        envelope_times.append(t0 + timedelta(hours=float(center)))

    return envelope_times, np.asarray(envelope_values, float)


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--spline-file", required=True)
    p.add_argument("--tide-file", required=True)
    p.add_argument("--tide-time-col", default="time")
    p.add_argument("--tide-value-col", required=True)
    p.add_argument("--output", default="amplitude_envelope.png")
    args = p.parse_args()

    spline_times, spline_values = load_spline_output(Path(args.spline_file))
    print(f"Loaded {len(spline_times)} GNSS-IR spline points")

    tide_times, tide_values = load_tide_reference(
        Path(args.tide_file), args.tide_time_col, args.tide_value_col
    )
    print(f"Loaded {len(tide_times)} tide model points")

    gnss_env_times, gnss_env = rolling_amplitude_envelope(spline_times, spline_values)
    tide_env_times, tide_env = rolling_amplitude_envelope(tide_times, tide_values)

    print(f"GNSS-IR envelope:  {len(gnss_env)} points, range {gnss_env.min():.3f}-{gnss_env.max():.3f} m" if len(gnss_env) else "GNSS-IR envelope: empty")
    print(f"Tide model envelope: {len(tide_env)} points, range {tide_env.min():.3f}-{tide_env.max():.3f} m" if len(tide_env) else "Tide model envelope: empty")

    fig, ax = plt.subplots(figsize=(16, 6))

    ax.plot(gnss_env_times, gnss_env, color="tab:blue", linewidth=1.5,
            label="GNSS-IR amplitude envelope (24h window)")
    ax.plot(tide_env_times, tide_env, color="tab:orange", linewidth=1.5,
            label="Tide model amplitude envelope (24h window)")

    ax.set_xlabel("Date")
    ax.set_ylabel("Amplitude (peak-to-peak, m, 24h rolling window)")
    ax.set_title("Spring-Neap Amplitude Envelope: GNSS-IR vs. Tide Model\n"
                  "(if these two curves rise and fall together, the spring-neap cycles are "
                  "synchronized; if they diverge, that's a real, site-specific finding)")
    ax.legend(loc="upper right")
    ax.grid(True, alpha=0.3)

    ax.xaxis.set_major_locator(mdates.AutoDateLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y-%m-%d"))
    fig.autofmt_xdate()

    fig.tight_layout()
    fig.savefig(args.output, dpi=150)
    print(f"Plot saved to: {args.output}")


if __name__ == "__main__":
    main()
