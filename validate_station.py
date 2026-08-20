#!/usr/bin/env python3
"""
validate_station.py

Generalized GNSS-IR site validation tool. Bundles four diagnostic
checks developed and proven during an extended real-world
investigation at one site, generalized here so any GNSS-IR
researcher can run the same methodology at their own station:

  1. reflection-zone : builds a Fresnel-footprint KML for the
                        station's real RH range and azimuth mask, to
                        visually/quantitatively check the footprint
                        actually falls where you think it does
                        (open water, bare soil, etc.) using real
                        satellite imagery -- not an assumption.

  2. azimuth          : tests whether a track's correlation with an
                         external reference signal (e.g. a tide
                         model) depends on azimuth. A real,
                         geometry-specific signal (e.g. water
                         reflection) should show strong correlation
                         concentrated near the true bearing to the
                         target and weak/absent correlation
                         elsewhere. A shared, non-physical drift
                         (atmospheric, instrumental, systematic
                         processing artifact) instead shows uniform
                         correlation across azimuth regardless of
                         geometry -- this check distinguishes the two
                         using the data's own internal structure,
                         not external assumption.

  3. elevation        : tests whether that same relationship depends
                         on elevation angle instead, by re-running
                         gnssir across several narrower elevation
                         sub-windows and comparing. Distinguishes a
                         fixed structural reflector (azimuth-specific,
                         elevation-independent) from an atmospheric/
                         multipath-geometry effect (elevation-
                         dependent, since atmospheric path length and
                         multipath delay both scale with elevation).

  4. refraction        : tests whether gnssrefl's tropospheric
                         refraction correction is responsible for a
                         found pattern, by re-running gnssir with
                         refraction correction disabled and comparing
                         the azimuth-independence result directly.

IMPORTANT -- what this tool does NOT do
-----------------------------------------
This tool tests METHODOLOGY, not conclusions. A track showing strong,
azimuth-independent correlation with your reference signal is
evidence of a shared, non-geometry-specific effect -- it does not by
itself tell you what that effect IS (atmospheric delay, an
instrumental artifact, a processing bug, or something else). Checks
2-4 are designed to rule candidate explanations in or out one at a
time; ruling out refraction (check 4) does not mean the remaining
pattern has no cause, only that this specific, tested cause is not
it.

Station identity, coordinates, and RH/elevation configuration are
read automatically from your existing station.json wherever
possible, so this tool does not duplicate that configuration --
only the reference-signal file (e.g. a tide model) and which checks
to run need to be specified explicitly.

USAGE
-----
Run everything:
    python3 validate_station.py --checks all \\
        --doy1 196 --doy2 221 \\
        --reference-file tides.xlsx --reference-time-col time \\
        --reference-value-col EOT20_heightm

Run just the azimuth-independence check:
    python3 validate_station.py --checks azimuth \\
        --doy1 196 --doy2 221 \\
        --reference-file tides.xlsx --reference-time-col time \\
        --reference-value-col EOT20_heightm

Run just the reflection-zone KML (no reference file needed):
    python3 validate_station.py --checks reflection-zone

Requires the gnssrefl virtual environment active (REFL_CODE, EXE,
ORBITS set) for checks that invoke gnssir directly (elevation,
refraction), and simplekml installed for the reflection-zone check.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import subprocess
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np


# ---------------------------------------------------------------------
# SHARED CONFIGURATION LOADING
# ---------------------------------------------------------------------

def load_station_config(station_json_path):
    """
    Loads station.json and returns a dict of the fields this tool
    needs, with clear errors for anything genuinely required but
    missing. Mirrors the field names used throughout the station/
    pipeline (gnssrefl_processor.py), so this tool reads the exact
    same configuration the real pipeline uses -- no separate,
    parallel config to keep in sync.
    """
    if not station_json_path.exists():
        print(f"ERROR: station.json not found at {station_json_path}")
        sys.exit(1)

    with open(station_json_path) as f:
        raw = json.load(f)

    station_code = raw.get("gnssrefl_station_code") or raw.get("station_id", "")[:4].lower()

    return {
        "station_code": station_code,
        "lat": raw.get("latitude"),
        "lon": raw.get("longitude"),
        "h_ortho": raw.get("gnssrefl_orthometric_height"),
        "rh_min": raw.get("gnssrefl_reflector_height_min"),
        "rh_max": raw.get("gnssrefl_reflector_height_max"),
        "e_min": raw.get("gnssrefl_elevation_min", 5.0),
        "e_max": raw.get("gnssrefl_elevation_max", 15.0),
        "raw": raw,
    }


def find_refl_code():
    refl_code = os.environ.get("REFL_CODE")
    if not refl_code:
        print("ERROR: REFL_CODE environment variable is not set.")
        print("Activate the gnssrefl venv and export REFL_CODE/EXE/ORBITS first.")
        sys.exit(1)
    return Path(refl_code)


# ---------------------------------------------------------------------
# SHARED: GNSS-IR RESULT FILE PARSING
# ---------------------------------------------------------------------

def parse_gnssir_results(result_dir, doys, freq_filter=1):
    """
    Parses standard gnssrefl result files (year doy RH sat UTCtime
    Azim Amp eminO emaxO NumbOf freq rise EdotF PkNoise DelT MJD
    refr), the same real, confirmed column layout used throughout
    this project. Returns a flat list of observation dicts.
    """
    rows = []
    for doy in doys:
        path = result_dir / f"{doy}.txt"
        if not path.exists():
            continue
        with open(path, errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("%"):
                    continue
                cols = line.split()
                if len(cols) < 17:
                    continue
                try:
                    year = int(float(cols[0]))
                    doy2 = int(float(cols[1]))
                    rh = float(cols[2])
                    sat = int(float(cols[3]))
                    utc_hours = float(cols[4])
                    az = float(cols[5])
                    emin = float(cols[7])
                    emax = float(cols[8])
                    freq = int(float(cols[10]))
                except (ValueError, IndexError):
                    continue
                if freq_filter is not None and freq != freq_filter:
                    continue
                day = datetime(year, 1, 1) + timedelta(days=doy2 - 1)
                dt = day + timedelta(hours=utc_hours)
                rows.append({
                    "doy": doy2, "sat": sat, "az": az, "rh": rh,
                    "elev_min": emin, "elev_max": emax, "dt": dt,
                })
    return rows


# ---------------------------------------------------------------------
# SHARED: REFERENCE SIGNAL (E.G. TIDE MODEL) LOADING
# ---------------------------------------------------------------------

def load_reference_signal(path, time_col, value_col):
    """
    Loads an external reference time series (e.g. a tide model) from
    an .xlsx file for correlation testing. Column names are
    explicitly configurable, not assumed, since this tool is meant
    to work with any researcher's own reference data format.
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]

    if time_col not in header:
        print(f"ERROR: time column {time_col!r} not found. Available columns: {header}")
        sys.exit(1)
    if value_col not in header:
        print(f"ERROR: value column {value_col!r} not found. Available columns: {header}")
        sys.exit(1)

    time_idx = header.index(time_col)
    value_idx = header.index(value_col)

    times, values = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[time_idx]
        if not isinstance(t, datetime):
            continue
        try:
            v = float(row[value_idx])
        except (TypeError, ValueError):
            continue
        if not math.isfinite(v):
            continue
        times.append(t)
        values.append(v)
    wb.close()

    if len(times) < 2:
        print("ERROR: insufficient reference signal data after loading.")
        sys.exit(1)

    epoch = np.array([(t - times[0]).total_seconds() for t in times])
    values = np.asarray(values, float)

    def value_at(dt):
        x = (dt - times[0]).total_seconds()
        if x < epoch[0] or x > epoch[-1]:
            return float("nan")
        return float(np.interp(x, epoch, values))

    return times, values, value_at


def pearson(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if len(x) < 5:
        return float("nan")
    if np.std(x) == 0 or np.std(y) == 0:
        return float("nan")
    return float(np.corrcoef(x, y)[0, 1])


# ---------------------------------------------------------------------
# CHECK 1: REFLECTION-ZONE / FRESNEL FOOTPRINT KML
# ---------------------------------------------------------------------

def check_reflection_zone(args, cfg):
    import simplekml
    from gnssrefl.refl_zones import makeEllipse_latlon

    if cfg["lat"] is None or cfg["lon"] is None:
        print("ERROR: latitude/longitude not found in station.json.")
        sys.exit(1)

    rh_min = args.rh_min or cfg["rh_min"]
    rh_max = args.rh_max or cfg["rh_max"]
    if rh_min is None or rh_max is None:
        print("ERROR: RH range not available (set --rh-min/--rh-max or "
              "gnssrefl_reflector_height_min/max in station.json).")
        sys.exit(1)
    rh_typical = (rh_min + rh_max) / 2.0

    center_az = args.center_azimuth
    if center_az is None:
        print("ERROR: --center-azimuth is required for the reflection-zone check "
              "(the true, measured bearing toward your target reflector).")
        sys.exit(1)

    half_width = args.half_width
    az_step = args.az_step
    elevation = args.elevation
    freq = args.freq

    n_steps = int(round((2 * half_width) / az_step))
    azimuths = [(center_az + off) % 360 for off in np.linspace(-half_width, half_width, n_steps + 1)]

    rh_values = {"low_water_max_reach": rh_max, "typical": rh_typical, "high_water_min_reach": rh_min}
    rh_colors = {"low_water_max_reach": simplekml.Color.red, "typical": simplekml.Color.yellow,
                 "high_water_min_reach": simplekml.Color.blue}

    kml = simplekml.Kml()
    shared_styles = {}
    for label, color in rh_colors.items():
        style = simplekml.Style()
        style.linestyle.color = color
        style.linestyle.width = 2
        style.polystyle.color = simplekml.Color.changealphaint(35, color)
        shared_styles[label] = style

    for label, rh_value in rh_values.items():
        folder = kml.newfolder(name=f"RH = {rh_value:.2f}m ({label})")
        for az in azimuths:
            lng, lat = makeEllipse_latlon(freq, elevation, rh_value, az, cfg["lat"], cfg["lon"])
            coords = [(float(x), float(y)) for x, y in zip(lng, lat)]
            p = folder.newpolygon(name=f"AZ {az:.1f} EL {elevation}")
            p.outerboundaryis = coords
            p.style = shared_styles[label]

    station_pt = kml.newpoint(name="GNSS station")
    station_pt.coords = [(cfg["lon"], cfg["lat"])]
    station_pt.style.iconstyle.icon.href = "http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png"

    out_path = Path(args.output_kml)
    kml.save(str(out_path))

    print()
    print("=" * 78)
    print("CHECK 1: REFLECTION-ZONE / FRESNEL FOOTPRINT")
    print("=" * 78)
    print(f"RH range tested : {rh_min:.2f} to {rh_max:.2f} m")
    print(f"Center azimuth  : {center_az:.1f} deg  (+/- {half_width:.0f} deg)")
    print(f"Elevation       : {elevation} deg")
    print(f"KML written to  : {out_path.resolve()}")
    print()
    print("NEXT STEP: open this KML in Google Earth over real satellite imagery.")
    print("Check directly whether the red (lowest-water, farthest-reaching)")
    print("footprints actually reach your intended target. If they fall short,")
    print("no azimuth adjustment will fix it -- physical antenna reach is the")
    print("limiting factor.")


# ---------------------------------------------------------------------
# CHECK 2: AZIMUTH-INDEPENDENCE
# ---------------------------------------------------------------------

def check_azimuth(args, cfg, refl_code):
    if not args.reference_file:
        print("ERROR: --reference-file is required for the azimuth check.")
        sys.exit(1)

    if args.extension:
        result_dir = refl_code / str(args.year) / "results" / cfg["station_code"] / args.extension
    else:
        result_dir = refl_code / str(args.year) / "results" / cfg["station_code"]
    doys = list(range(args.doy1, args.doy2 + 1))

    print()
    print("=" * 78)
    print("CHECK 2: AZIMUTH-INDEPENDENCE TEST")
    print("=" * 78)
    print(f"Reading results from: {result_dir}")
    if not args.extension:
        print("(Using PRODUCTION results -- pass --extension to test an "
              "isolated experimental config instead)")

    rows = parse_gnssir_results(result_dir, doys)
    print(f"Loaded {len(rows)} raw observations")

    _, _, value_at = load_reference_signal(
        Path(args.reference_file), args.reference_time_col, args.reference_value_col
    )

    by_sat = defaultdict(list)
    for r in rows:
        by_sat[r["sat"]].append(r)

    results = []
    for sat, observations in sorted(by_sat.items()):
        if len(observations) < args.min_track_obs:
            continue
        rh = np.array([o["rh"] for o in observations])
        az = np.array([o["az"] for o in observations])
        ref = np.array([value_at(o["dt"]) for o in observations])
        valid = np.isfinite(ref)
        if np.sum(valid) < args.min_track_obs:
            continue
        r_corr = pearson(rh[valid], ref[valid])
        if not math.isfinite(r_corr):
            continue
        results.append({"sat": sat, "n": int(np.sum(valid)), "az_mean": float(np.mean(az)), "r": r_corr})

    results.sort(key=lambda r: r["az_mean"])

    print()
    print(f"{'sat':>4} {'az':>7} {'n':>4} {'r':>8}")
    for r in results:
        print(f"{r['sat']:>4} {r['az_mean']:7.2f} {r['n']:>4} {r['r']:8.4f}")

    if len(results) < 6:
        print()
        print(f"Only {len(results)} tracks survived (need >= 6 for a meaningful "
              f"azimuth-correlation estimate). Not reporting a summary statistic.")
        return

    az = np.array([r["az_mean"] for r in results])
    rvals = np.array([abs(r["r"]) for r in results])
    summary_corr = float(np.corrcoef(az, rvals)[0, 1])

    print()
    print(f"corr(azimuth, |r|) = {summary_corr:+.4f}")
    print()
    print("INTERPRETATION")
    print("A value near zero, with strong correlation appearing at azimuths")
    print("far from your true target bearing (and/or weak correlation near it),")
    print("indicates a shared, non-geometry-specific effect -- NOT genuine")
    print("target-specific sensing. A strong, clearly negative relationship")
    print("(correlation concentrated near the true bearing, weak elsewhere) is")
    print("consistent with genuine, geometry-specific sensing.")


# ---------------------------------------------------------------------
# CHECK 3: ELEVATION-WINDOW SENSITIVITY
# ---------------------------------------------------------------------

def run_gnssir_extension(station_code, year, doys, extension, freq=1):
    for doy in doys:
        subprocess.run(
            ["gnssir", station_code, str(year), str(doy),
             "-extension", extension, "-fr", str(freq)],
            capture_output=True, text=True, check=False,
        )


def make_extension_config(base_json_path, extension_name, refl_code, station_code, overrides):
    with open(base_json_path) as f:
        base_config = json.load(f)
    config = dict(base_config)
    config.update(overrides)
    dst_dir = refl_code / "input" / station_code / extension_name
    dst_dir.mkdir(parents=True, exist_ok=True)
    with open(dst_dir / f"{station_code}.json", "w") as f:
        json.dump(config, f, indent=4)


def check_elevation(args, cfg, refl_code):
    if not args.reference_file:
        print("ERROR: --reference-file is required for the elevation check.")
        sys.exit(1)

    e_min = args.e_min or cfg["e_min"]
    e_max = args.e_max or cfg["e_max"]
    doys = list(range(args.doy1, args.doy2 + 1))
    station_code = cfg["station_code"]

    base_json = refl_code / "input" / station_code / f"{station_code}.json"
    if not base_json.exists():
        print(f"ERROR: base gnssir config not found at {base_json}. Run the "
              f"normal pipeline at least once first.")
        sys.exit(1)

    midpoint = (e_min + e_max) / 2.0
    windows = [("lower_half", e_min, midpoint), ("upper_half", midpoint, e_max)]

    print()
    print("=" * 78)
    print("CHECK 3: ELEVATION-WINDOW SENSITIVITY TEST")
    print("=" * 78)
    print(f"Full elevation range: {e_min}-{e_max} deg, split at {midpoint:.1f} deg")
    print()
    print("NOTE: windows narrower than roughly 4 deg often produce zero")
    print("successful retrievals -- a genuine LSP-resolution limit, not a bug.")
    print("If a window below shows 'insufficient data', this is the likely cause.")

    for label, lo, hi in windows:
        extension = f"validate_elev_{label}"
        print()
        print(f"--- {label}: {lo:.1f}-{hi:.1f} deg ---")
        make_extension_config(base_json, extension, refl_code, station_code,
                               {"e1": float(lo), "e2": float(hi)})
        run_gnssir_extension(station_code, args.year, doys, extension)

        result_dir = refl_code / str(args.year) / "results" / station_code / extension
        rows = parse_gnssir_results(result_dir, doys)

        if not rows:
            print("  No successful retrievals in this window.")
            continue

        _, _, value_at = load_reference_signal(
            Path(args.reference_file), args.reference_time_col, args.reference_value_col
        )

        by_sat = defaultdict(list)
        for r in rows:
            by_sat[r["sat"]].append(r)

        results = []
        for sat, observations in sorted(by_sat.items()):
            if len(observations) < args.min_track_obs:
                continue
            rh = np.array([o["rh"] for o in observations])
            ref = np.array([value_at(o["dt"]) for o in observations])
            valid = np.isfinite(ref)
            if np.sum(valid) < args.min_track_obs:
                continue
            r_corr = pearson(rh[valid], ref[valid])
            if math.isfinite(r_corr):
                results.append(r_corr)

        print(f"  {len(rows)} observations, {len(results)} usable tracks")
        if len(results) >= 3:
            print(f"  mean |r| across tracks: {np.mean(np.abs(results)):.4f}")
        else:
            print("  Too few tracks in this window for a meaningful summary.")

    print()
    print("INTERPRETATION")
    print("Compare the two windows' results directly. A meaningful difference")
    print("between them suggests an elevation-dependent effect. Similar results")
    print("in both windows argue against elevation-dependence as the driver.")


# ---------------------------------------------------------------------
# CHECK 4: REFRACTION ON/OFF COMPARISON
# ---------------------------------------------------------------------

def check_refraction(args, cfg, refl_code):
    if not args.reference_file:
        print("ERROR: --reference-file is required for the refraction check.")
        sys.exit(1)

    doys = list(range(args.doy1, args.doy2 + 1))
    station_code = cfg["station_code"]

    base_json = refl_code / "input" / station_code / f"{station_code}.json"
    if not base_json.exists():
        print(f"ERROR: base gnssir config not found at {base_json}. Run the "
              f"normal pipeline at least once first.")
        sys.exit(1)

    extension = "validate_norefraction"

    print()
    print("=" * 78)
    print("CHECK 4: REFRACTION CORRECTION ON/OFF COMPARISON")
    print("=" * 78)

    make_extension_config(base_json, extension, refl_code, station_code, {"refraction": False})
    run_gnssir_extension(station_code, args.year, doys, extension)

    result_dir = refl_code / str(args.year) / "results" / station_code / extension
    rows = parse_gnssir_results(result_dir, doys)

    if not rows:
        print("No successful retrievals with refraction correction disabled.")
        return

    _, _, value_at = load_reference_signal(
        Path(args.reference_file), args.reference_time_col, args.reference_value_col
    )

    by_sat = defaultdict(list)
    for r in rows:
        by_sat[r["sat"]].append(r)

    results = []
    for sat, observations in sorted(by_sat.items()):
        if len(observations) < args.min_track_obs:
            continue
        rh = np.array([o["rh"] for o in observations])
        az = np.array([o["az"] for o in observations])
        ref = np.array([value_at(o["dt"]) for o in observations])
        valid = np.isfinite(ref)
        if np.sum(valid) < args.min_track_obs:
            continue
        r_corr = pearson(rh[valid], ref[valid])
        if math.isfinite(r_corr):
            results.append({"sat": sat, "az_mean": float(np.mean(az)), "r": r_corr})

    print(f"{len(rows)} observations, {len(results)} usable tracks (refraction OFF)")

    if len(results) >= 6:
        az = np.array([r["az_mean"] for r in results])
        rvals = np.array([abs(r["r"]) for r in results])
        summary_corr = float(np.corrcoef(az, rvals)[0, 1])
        print(f"corr(azimuth, |r|) with refraction OFF: {summary_corr:+.4f}")
        print()
        print("Compare this directly against the same statistic from the")
        print("'azimuth' check (refraction correction ON, gnssrefl's normal")
        print("behavior). A similar value in both cases means refraction")
        print("correction is not the driver of whatever pattern was found.")
        print("A meaningfully different value implicates refraction correction.")
    else:
        print("Too few tracks to compute a meaningful summary statistic.")


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--station-json", default="station/resources/station.json")
    p.add_argument("--checks", default="all",
                   help="comma-separated: reflection-zone,azimuth,elevation,refraction,all")
    p.add_argument("--year", type=int, default=datetime.now().year)
    p.add_argument("--doy1", type=int, help="start day of year (required for azimuth/elevation/refraction)")
    p.add_argument("--doy2", type=int, help="end day of year (required for azimuth/elevation/refraction)")

    p.add_argument("--reference-file", default=None, help="external reference signal .xlsx (e.g. tide model)")
    p.add_argument("--reference-time-col", default="time")
    p.add_argument("--reference-value-col", default=None)
    p.add_argument("--extension", default=None,
                   help="gnssir -extension subdirectory to read results from for the "
                        "azimuth check (default: production results, NOT any "
                        "experimental/widened-range config). Use this to point at an "
                        "existing, already-processed experimental result set instead "
                        "of production.")

    p.add_argument("--min-track-obs", type=int, default=5)

    # reflection-zone specific
    p.add_argument("--center-azimuth", type=float, default=None)
    p.add_argument("--half-width", type=float, default=90.0)
    p.add_argument("--az-step", type=float, default=5.0)
    p.add_argument("--elevation", type=float, default=5.0)
    p.add_argument("--freq", type=int, default=1)
    p.add_argument("--rh-min", type=float, default=None)
    p.add_argument("--rh-max", type=float, default=None)
    p.add_argument("--e-min", type=float, default=None)
    p.add_argument("--e-max", type=float, default=None)
    p.add_argument("--output-kml", default="reflection_zone.kml")

    return p.parse_args()


def main():
    args = parse_args()

    checks = set(c.strip() for c in args.checks.split(","))
    if "all" in checks:
        checks = {"reflection-zone", "azimuth", "elevation", "refraction"}

    cfg = load_station_config(Path(args.station_json))

    needs_gnssir = checks & {"elevation", "refraction"}
    needs_results = checks & {"azimuth", "elevation", "refraction"}

    refl_code = None
    if needs_results:
        refl_code = find_refl_code()

    if needs_results and (args.doy1 is None or args.doy2 is None):
        print("ERROR: --doy1 and --doy2 are required for azimuth/elevation/refraction checks.")
        sys.exit(1)

    if needs_results and not args.reference_value_col:
        print("ERROR: --reference-value-col is required for azimuth/elevation/refraction checks.")
        sys.exit(1)

    print()
    print("=" * 78)
    print("GNSS-IR STATION VALIDATION")
    print("=" * 78)
    print(f"Station code : {cfg['station_code']}")
    print(f"Checks       : {sorted(checks)}")

    if "reflection-zone" in checks:
        check_reflection_zone(args, cfg)

    if "azimuth" in checks:
        check_azimuth(args, cfg, refl_code)

    if "elevation" in checks:
        check_elevation(args, cfg, refl_code)

    if "refraction" in checks:
        check_refraction(args, cfg, refl_code)

    print()
    print("=" * 78)
    print("DONE")
    print("=" * 78)


if __name__ == "__main__":
    main()
