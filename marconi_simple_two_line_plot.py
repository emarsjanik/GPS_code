#!/usr/bin/env python3
"""
Marconi simple two-line GNSS-R validation plot.

Exactly two plotted lines:

1. EOT20 tide model:
   continuous line using all tide-model timestamps.

2. GNSS-R estimated water level:
   one line connecting all GPS L1 GNSS-R observations in chronological
   order.

GNSS-R water level:
    WL = H_ortho - RH
    H_ortho = 18.665 m

IMPORTANT:
The GNSS-R line is a visualization of the chronological observations.
Because it combines different satellites/reflection tracks, it should
not be interpreted as a physically continuous single-track water-level
time series. It is included here because the requested visualization
is a direct two-line overlay.
"""

from pathlib import Path
from datetime import datetime, timedelta
import math

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

OUT_FILE = Path(
    "marconi_simple_gnssr_vs_eot20_two_lines.png"
)

H_ORTHO = 18.665


def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def utc_datetime(year, doy, utc_hours):
    return (
        datetime(year, 1, 1)
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


def load_eot20():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [c.value for c in ws[1]]

    time_col = header.index("time")
    eot20_col = header.index("EOT20_heightm")

    times = []
    values = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_col]
        v = finite(row[eot20_col])

        if not isinstance(t, datetime):
            continue

        if v is None:
            continue

        times.append(t)
        values.append(v)

    wb.close()

    return times, np.asarray(values, dtype=float)


def load_gnssr():
    rows = []

    for path in sorted(
        RESULT_DIR.glob("*.txt")
    ):
        try:
            int(path.stem)
        except Exception:
            continue

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy = int(float(c[1]))
                rh = float(c[2])
                utc = float(c[4])
                freq = int(float(c[10]))
            except Exception:
                continue

            if freq != 1:
                continue

            dt = utc_datetime(
                year,
                doy,
                utc,
            )

            rows.append(
                (
                    dt,
                    H_ORTHO - rh,
                )
            )

    rows.sort(key=lambda x: x[0])

    return rows


def main():
    tide_times, tide_values = load_eot20()
    gnssr = load_gnssr()

    if not tide_times:
        raise SystemExit(
            "No EOT20 tide data found."
        )

    if not gnssr:
        raise SystemExit(
            "No GPS L1 GNSS-R data found."
        )

    fig, ax = plt.subplots(
        figsize=(15, 7)
    )

    # LINE 1: continuous tide model
    ax.plot(
        tide_times,
        tide_values,
        linewidth=2.5,
        color="red",
        label="EOT20 tide model",
    )

    # LINE 2: chronological GNSS-R estimates
    ax.plot(
        [x[0] for x in gnssr],
        [x[1] for x in gnssr],
        linewidth=1.8,
        marker="o",
        markersize=2.8,
        color="blue",
        label="GNSS-R estimated water level",
    )

    ax.set_title(
        "Marconi GNSS-R Water Level vs EOT20"
    )

    ax.set_xlabel(
        "UTC"
    )

    ax.set_ylabel(
        "Water-level elevation / anomaly (m)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend(
        loc="upper right"
    )

    fig.tight_layout()

    fig.savefig(
        OUT_FILE,
        dpi=200,
    )

    plt.close(fig)

    print()
    print("=" * 80)
    print(
        "MARCONI TWO-LINE GNSS-R / EOT20 VALIDATION PLOT"
    )
    print("=" * 80)
    print(
        "EOT20 tide points :",
        len(tide_times),
    )
    print(
        "GNSS-R points     :",
        len(gnssr),
    )
    print()
    print(
        "GNSS-R is plotted as one chronological line."
    )
    print(
        "This is a visualization only; it combines different reflection tracks."
    )
    print()
    print(
        "Saved:",
        OUT_FILE.resolve(),
    )
    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
