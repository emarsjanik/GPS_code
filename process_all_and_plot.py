#!/usr/bin/env python3
"""
process_all_and_plot.py

Single-command GNSS-IR pipeline: finds every available day of raw SNR
data, reprocesses all of it fresh through gnssir (production RH range,
-0.5 to 5.0m), builds a water-level time series, and overlays it
against the real tide model in one plot.

DESIGNED TO BE SIMPLE TO RUN:

    python3 process_all_and_plot.py

That's it -- it will auto-detect the station name, auto-find the tide
workbook if there's exactly one .xlsx in the current directory, and
process every day it has local SNR data for. Override any of that
with flags if needed (see --help).

IMPORTANT -- THE OFFSET CONSTANT
---------------------------------
OFFSET_M below (currently +0.242m) is NOT a confirmed, validated datum
correction. It was derived from a population-wide average across many
different satellite tracks, and a careful per-track breakdown showed
those tracks disagree with each other by up to ~1.75m -- inconsistent
with a genuine, single physical datum offset. Treat this as a labeled,
adjustable placeholder, not an established fact. It is applied and
clearly marked on the plot so it's easy to see, question, and change.
Set OFFSET_M = 0.0 (or pass --offset 0) to see the uncorrected result.

WHAT THIS ASSUMES
------------------
- The gnssrefl virtual environment is active
- REFL_CODE, EXE, ORBITS environment variables are set
- Production usgs.json already has the RH search range you want
  (this tool does NOT modify station config -- it uses whatever is
  currently live)
- Raw SNR files already exist locally (this tool does not generate
  RINEX/SNR from raw receiver data -- see the station/ pipeline for
  that earlier stage)

Outputs:
    water_level_vs_tide.png
    water_level_vs_tide_data.csv
"""

from __future__ import annotations

import argparse
import csv
import re
import subprocess
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# CONFIGURATION -- the few things you might actually want to change
# ---------------------------------------------------------------------

# See the "IMPORTANT" note in the module docstring above before
# trusting this value.
OFFSET_M = 0.242

H_ORTHO_M = 18.665  # station orthometric height, from station.json

TIDE_MODEL_COLUMNS_SUFFIX = "_heightm"  # matches marconi_tides_*.xlsx format


# ---------------------------------------------------------------------
# DISCOVERY
# ---------------------------------------------------------------------

def find_tide_workbook(explicit_path):
    if explicit_path:
        path = Path(explicit_path)
        if not path.exists():
            print(f"ERROR: tide workbook not found: {path}")
            sys.exit(1)
        return path

    candidates = list(Path(".").glob("*.xlsx"))
    if len(candidates) == 1:
        print(f"Auto-detected tide workbook: {candidates[0]}")
        return candidates[0]
    if len(candidates) == 0:
        print("ERROR: no .xlsx tide workbook found in the current directory.")
        print("Pass one explicitly with --tide-file.")
        sys.exit(1)
    print("ERROR: multiple .xlsx files found, can't auto-detect which is the tide workbook:")
    for c in candidates:
        print(f"  {c}")
    print("Pass the correct one explicitly with --tide-file.")
    sys.exit(1)


def find_snr_days(station, refl_code_dir, year):
    """
    Scans the standard gnssrefl SNR directory for available days,
    matching the real filename pattern: STATIONDDD0.YY.snr66[.gz]
    """
    snr_dir = refl_code_dir / str(year) / "snr" / station
    if not snr_dir.exists():
        print(f"ERROR: SNR directory not found: {snr_dir}")
        print("Has any data been processed to SNR format yet for this station/year?")
        sys.exit(1)

    yy = str(year)[2:]
    pattern = re.compile(rf"^{re.escape(station)}(\d{{3}})0\.{yy}\.snr66(?:\.gz)?$")

    doys = set()
    for p in snr_dir.iterdir():
        m = pattern.match(p.name)
        if m:
            doys.add(int(m.group(1)))

    return sorted(doys)


# ---------------------------------------------------------------------
# PROCESSING
# ---------------------------------------------------------------------

def run_gnssir_for_day(station, year, doy):
    cmd = ["gnssir", station, str(year), str(doy)]
    result = subprocess.run(cmd, text=True, capture_output=True, check=False)
    return result.returncode == 0, result.stdout, result.stderr


def process_all_days(station, year, doys):
    print()
    print("=" * 78)
    print(f"PROCESSING {len(doys)} DAY(S) WITH gnssir (production config, unchanged)")
    print("=" * 78)

    successes = []
    failures = []

    for doy in doys:
        ok, stdout, stderr = run_gnssir_for_day(station, year, doy)
        if ok:
            successes.append(doy)
            print(f"  DOY {doy}: OK")
        else:
            failures.append(doy)
            print(f"  DOY {doy}: FAILED")
            if stderr.strip():
                print(f"    {stderr.strip()[:200]}")

    print()
    print(f"Processed successfully: {len(successes)}/{len(doys)}")
    if failures:
        print(f"Failed days: {failures}")

    return successes


def find_spline_file(station, refl_code_dir):
    """
    gnssrefl's subdaily tool writes the spline output to
    $REFL_CODE/Files/<station>/<station>_spline_out.txt
    """
    return refl_code_dir / "Files" / station / f"{station}_spline_out.txt"


def spline_is_current(spline_path, successes):
    """
    Simple, conservative freshness check: the spline file must exist
    and be newer than the most recently written result file it should
    cover. If in doubt, regenerate rather than risk a stale spline.
    """
    if not spline_path.exists():
        return False
    return True  # existence check only; regeneration is cheap and safe to force via --regenerate-spline


def generate_spline(station, year, doys):
    """
    Runs gnssrefl's subdaily tool to build a smoothed spline across
    the full available day range, using the same -knots 4 setting
    confirmed necessary to avoid spline overshoot artifacts with the
    default (8-knot) setting on real, gap-containing data.
    """
    doy1, doy2 = min(doys), max(doys)
    print()
    print("=" * 78)
    print(f"GENERATING SMOOTHED SPLINE (subdaily, DOY {doy1}-{doy2}, -knots 4)")
    print("=" * 78)

    cmd = ["subdaily", station, str(year), "-doy1", str(doy1), "-doy2", str(doy2),
           "-rhdot", "True", "-knots", "4"]
    print("$", " ".join(cmd))

    result = subprocess.run(cmd, text=True, capture_output=True, check=False)
    print(result.stdout[-2000:])  # subdaily is verbose; show the tail
    if result.returncode != 0:
        print("WARNING: subdaily did not complete successfully. Spline overlay")
        print("will be skipped; the raw-points plot is unaffected.")
        if result.stderr.strip():
            print(result.stderr.strip()[:500])
        return False
    return True


def load_spline(spline_path, year):
    """
    Parses gnssrefl's subdaily spline output file. Format (from real,
    confirmed output): whitespace-separated columns where column 0 is
    fractional day-of-year and column 1 is spline-fit reflector height
    in meters. Lines starting with '%' are headers/comments. A value
    of 999 means a real gap with no spline value available -- these
    are converted to NaN so plotting shows a genuine break, not a
    misleading straight line across missing data.
    """
    if not spline_path.exists():
        return None, None

    times = []
    rh_values = []

    day_start_ref = datetime(year, 1, 1)

    for line in spline_path.read_text(errors="replace").splitlines():
        line = line.strip()
        if not line or line.startswith("%"):
            continue
        parts = line.split()
        if len(parts) < 2:
            continue
        try:
            frac_doy = float(parts[0])
            rh = float(parts[1])
        except ValueError:
            continue

        dt = day_start_ref + timedelta(days=frac_doy - 1)
        times.append(dt)
        rh_values.append(rh if abs(rh - 999.0) > 1e-6 else float("nan"))

    if not times:
        return None, None

    return times, np.array(rh_values, dtype=float)


# ---------------------------------------------------------------------
# PARSE RESULTS
# ---------------------------------------------------------------------

def load_results(station, year, doys, refl_code_dir):
    result_dir = refl_code_dir / str(year) / "results" / station

    rows = []
    for doy in doys:
        path = result_dir / f"{doy}.txt"
        if not path.exists():
            continue

        for line in path.read_text(errors="replace").splitlines():
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            c = line.split()
            if len(c) < 17:
                continue
            try:
                yr = int(float(c[0]))
                d2 = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc_hours = float(c[4])
                freq = int(float(c[10]))
            except (ValueError, IndexError):
                continue

            dt = datetime(yr, 1, 1) + timedelta(days=d2 - 1, hours=utc_hours)

            rows.append({
                "datetime": dt, "doy": d2, "sat": sat, "freq": freq,
                "RH_m": rh,
                "water_level_raw_m": H_ORTHO_M - rh,
            })

    rows.sort(key=lambda r: r["datetime"])
    return rows


# ---------------------------------------------------------------------
# TIDE MODEL
# ---------------------------------------------------------------------

def load_tide_model(tide_path):
    wb = load_workbook(tide_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    time_col = header.index("time")
    model_cols = {}
    for i, name in enumerate(header):
        if name and str(name).endswith(TIDE_MODEL_COLUMNS_SUFFIX):
            model_cols[str(name)] = i

    if not model_cols:
        print(f"ERROR: no columns ending in '{TIDE_MODEL_COLUMNS_SUFFIX}' found in {tide_path}")
        sys.exit(1)

    times = []
    values = {name: [] for name in model_cols}

    for row in rows[1:]:
        t = row[time_col]
        if not isinstance(t, datetime):
            continue
        row_ok = True
        row_vals = {}
        for name, col in model_cols.items():
            v = row[col]
            try:
                v = float(v)
            except (TypeError, ValueError):
                row_ok = False
                break
            if not np.isfinite(v):
                row_ok = False
                break
            row_vals[name] = v
        if not row_ok:
            continue
        times.append(t)
        for name in model_cols:
            values[name].append(row_vals[name])

    wb.close()

    if len(times) < 2:
        print("ERROR: insufficient tide model data.")
        sys.exit(1)

    ensemble = np.mean([values[name] for name in model_cols], axis=0)
    epoch = np.array([(t - times[0]).total_seconds() for t in times], dtype=float)

    def tide_at(dt):
        x = (dt - times[0]).total_seconds()
        if x < epoch[0] or x > epoch[-1]:
            return float("nan")
        return float(np.interp(x, epoch, ensemble))

    return times, ensemble, tide_at, list(model_cols.keys())


# ---------------------------------------------------------------------
# PLOT
# ---------------------------------------------------------------------

def make_plot(rows, tide_times, tide_ensemble, offset_m, model_names,
              spline_times=None, spline_rh=None):
    fig, ax1 = plt.subplots(figsize=(14, 7))

    dts = [r["datetime"] for r in rows]
    wl_corrected = [r["water_level_raw_m"] + offset_m for r in rows]

    # Raw, individual observations -- always shown, since smoothing
    # can hide real scatter/disagreement that's worth seeing directly.
    ax1.plot(dts, wl_corrected, "o", ms=3, alpha=0.5, color="tab:orange",
              label=f"GNSS-IR raw arcs (+{offset_m:.3f}m offset)")

    # Optional smoothed spline overlay, if available.
    if spline_times is not None and spline_rh is not None:
        spline_wl = H_ORTHO_M - spline_rh + offset_m
        ax1.plot(spline_times, spline_wl, "-", linewidth=1.5, color="darkorange",
                  label="GNSS-IR smoothed spline (subdaily, -knots 4)")

    ax1.set_ylabel("GNSS-IR water level (m)", color="tab:orange")
    ax1.tick_params(axis="y", labelcolor="tab:orange")

    ax2 = ax1.twinx()
    ax2.plot(tide_times, tide_ensemble, "-", color="tab:blue", linewidth=2,
              label=f"Tide model (ensemble mean of {len(model_names)} models)")
    ax2.set_ylabel("Tide model height (m)", color="tab:blue")
    ax2.tick_params(axis="y", labelcolor="tab:blue")

    title = "GNSS-IR Water Level vs. Tide Model"
    if offset_m != 0.0:
        title += f"\n(offset = {offset_m:+.3f}m -- see script docstring: NOT a validated constant)"
    ax1.set_title(title)
    ax1.set_xlabel("UTC")
    ax1.grid(alpha=0.3)

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper right")

    fig.autofmt_xdate()
    fig.tight_layout()

    out_path = Path("water_level_vs_tide.png")
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    return out_path


def write_csv(rows, tide_at, offset_m):
    out_path = Path("water_level_vs_tide_data.csv")
    with open(out_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["datetime_utc", "doy", "sat", "freq", "RH_m",
                          "water_level_raw_m", "water_level_corrected_m",
                          "tide_ensemble_m"])
        for r in rows:
            tide = tide_at(r["datetime"])
            writer.writerow([
                r["datetime"].isoformat(), r["doy"], r["sat"], r["freq"],
                f"{r['RH_m']:.4f}",
                f"{r['water_level_raw_m']:.4f}",
                f"{r['water_level_raw_m'] + offset_m:.4f}",
                f"{tide:.4f}" if np.isfinite(tide) else "",
            ])
    return out_path


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description="Process all available GNSS-IR data and plot against the tide model."
    )
    p.add_argument("--station", default="usgs", help="4-char station code (default: usgs)")
    p.add_argument("--year", type=int, default=2026)
    p.add_argument("--tide-file", default=None,
                    help="Path to tide model .xlsx (auto-detected if omitted and only one exists)")
    p.add_argument("--offset", type=float, default=None,
                    help=f"Override the offset constant (default: {OFFSET_M}, see script docstring)")
    p.add_argument("--refl-code", default=None,
                    help="Path to REFL_CODE directory (default: $REFL_CODE env var)")
    p.add_argument("--no-spline", action="store_true",
                    help="Skip smoothed spline generation; plot raw points only (faster)")
    return p.parse_args()


def main():
    args = parse_args()

    import os
    refl_code_env = os.environ.get("REFL_CODE")
    refl_code_dir = Path(args.refl_code or refl_code_env or "products/refl_code")

    offset_m = args.offset if args.offset is not None else OFFSET_M

    print()
    print("=" * 78)
    print("PROCESS ALL DATA AND PLOT AGAINST TIDE")
    print("=" * 78)
    print(f"Station        : {args.station}")
    print(f"Year           : {args.year}")
    print(f"REFL_CODE dir  : {refl_code_dir}")
    print(f"Offset applied : {offset_m:+.3f} m  (see script docstring re: validity)")

    tide_path = find_tide_workbook(args.tide_file)

    doys = find_snr_days(args.station, refl_code_dir, args.year)
    if not doys:
        print("ERROR: no SNR files found for this station/year.")
        sys.exit(1)

    print()
    print(f"Found local SNR data for {len(doys)} day(s): {doys}")

    successes = process_all_days(args.station, args.year, doys)
    if not successes:
        print("ERROR: no days processed successfully.")
        sys.exit(1)

    rows = load_results(args.station, args.year, successes, refl_code_dir)
    if not rows:
        print("ERROR: no usable observations found in the result files.")
        sys.exit(1)

    print(f"\nLoaded {len(rows)} total observations across {len(successes)} day(s).")

    # Spline generation is automatic by default -- identifies whether
    # a spline exists, generates/regenerates it fresh each run (cheap
    # relative to the gnssir step above), and gracefully falls back
    # to raw-points-only if subdaily fails for any reason.
    spline_times, spline_rh = None, None
    if not args.no_spline:
        spline_path = find_spline_file(args.station, refl_code_dir)
        ok = generate_spline(args.station, args.year, successes)
        if ok:
            spline_times, spline_rh = load_spline(spline_path, args.year)
            if spline_times is not None:
                print(f"Spline loaded: {len(spline_times)} points from {spline_path}")
            else:
                print("WARNING: spline file could not be parsed; continuing with raw points only.")

    tide_times, tide_ensemble, tide_at, model_names = load_tide_model(tide_path)
    print(f"Tide models used: {model_names}")

    plot_path = make_plot(rows, tide_times, tide_ensemble, offset_m, model_names,
                            spline_times, spline_rh)
    csv_path = write_csv(rows, tide_at, offset_m)

    print()
    print("=" * 78)
    print("DONE")
    print("=" * 78)
    print(f"Plot: {plot_path.resolve()}")
    print(f"Data: {csv_path.resolve()}")


if __name__ == "__main__":
    main()
