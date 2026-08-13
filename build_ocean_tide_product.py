#!/usr/bin/env python3
"""
build_ocean_tide_product.py

Build a scientifically controlled Marconi GNSS-IR ocean-water-level
product without using gnssrefl's subdaily spline as the primary
water-level estimate.

DESIGN PRINCIPLES
-----------------
1. Start from individual, QC-passing GNSS-IR arc solutions.
2. Do NOT combine all azimuths, satellites, and frequencies into one
   spline.
3. Default to GPS L1 (freq=1) only for the primary product.
4. Default to the corrected Marconi ocean-facing sector 35-135 deg.
   This is deliberately conservative and is based on the repository's
   reflection-zone investigation; change it only after checking the
   Fresnel-zone KML.
5. Convert reflector height to water-level convention as:
       water_level = H_ortho - RH
   but emphasize anomaly comparisons because absolute vertical datum
   compatibility with the tide model must be independently established.
6. Compare GNSS-IR water-level anomaly with each tide model separately
   and with their ensemble mean.
7. Do NOT spline-fit the primary product.
8. Produce an auditable CSV containing every included/excluded arc and
   the reason for exclusion.
9. Also produce a track-normalized anomaly product for diagnostic use.
   A track is satellite+frequency; it is only normalized if it has at
   least MIN_TRACK_OBS observations.

INPUT
-----
Default:
    gnssir_tide_arc_analysis.csv
    marconi_tides_sherwood.xlsx

The CSV is expected to contain:
    sat, freq, solution_time_utc, RH_m, Azim, Amp, PkNoise,
    delT_min, eminO, emaxO

OUTPUT
------
ocean_gnssir_arc_product.csv
ocean_gnssir_excluded_arcs.csv
ocean_gnssir_30min_binned.csv
ocean_gnssir_summary.txt
ocean_gnssir_anomaly_vs_tide.png
ocean_gnssir_absolute_vs_tide.png

This script is intentionally a product-builder, not a replacement
for gnssrefl itself.
"""

from __future__ import annotations

import argparse
import csv
import math
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# DEFAULT SCIENTIFIC SETTINGS
# ---------------------------------------------------------------------

DEFAULT_CSV = Path("gnssir_tide_arc_analysis.csv")
DEFAULT_TIDE = Path("marconi_tides_sherwood.xlsx")

REFERENCE_HEIGHT_M = 18.665

# Conservative ocean-facing sector from the repository's corrected
# Marconi reflection-zone investigation.
DEFAULT_AZ1 = 35.0
DEFAULT_AZ2 = 135.0

# Primary product: GPS L1 only.
DEFAULT_FREQ = 1

# These reproduce the established production QC gates.
DEFAULT_MIN_PKNOISE = 2.8
DEFAULT_MIN_AMP = 5.0
DEFAULT_MAX_DELT_MIN = 40.0

# Original GNSS-IR elevation window. These are arc solution limits,
# not a guarantee that the Fresnel zone is water.
DEFAULT_EMIN = 5.0
DEFAULT_EMAX = 15.0

# Track normalization is deliberately conservative.
MIN_TRACK_OBS = 3

# Output aggregation only; this is NOT a spline.
BIN_MINUTES = 30


@dataclass
class Record:
    source_row: int
    doy: int
    dt: datetime
    sat: int
    freq: int
    rh_m: float
    az: float
    amp: float
    pkn: float
    delt_min: float
    emin: float
    emax: float

    tide_eot20: float = math.nan
    tide_got55: float = math.nan
    tide_got56: float = math.nan
    tide_fes2022: float = math.nan

    water_level_m: float = math.nan
    water_anomaly_m: float = math.nan
    track_anomaly_m: float = math.nan


def finite(value):
    try:
        x = float(value)
        return x if math.isfinite(x) else None
    except (TypeError, ValueError):
        return None


def pearson(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    m = np.isfinite(x) & np.isfinite(y)
    x = x[m]
    y = y[m]

    if len(x) < 3:
        return math.nan
    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(np.corrcoef(x, y)[0, 1])


def linear_fit(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    m = np.isfinite(x) & np.isfinite(y)
    x = x[m]
    y = y[m]

    if len(x) < 3:
        return math.nan, math.nan, math.nan

    slope, intercept = np.polyfit(x, y, 1)
    resid = y - (slope * x + intercept)
    rms = float(np.sqrt(np.mean(resid ** 2)))

    return float(slope), float(intercept), rms


def parse_args():
    p = argparse.ArgumentParser(
        description="Build controlled ocean-only GNSS-IR water-level products."
    )

    p.add_argument("--csv", type=Path, default=DEFAULT_CSV)
    p.add_argument("--tide", type=Path, default=DEFAULT_TIDE)

    p.add_argument("--az1", type=float, default=DEFAULT_AZ1)
    p.add_argument("--az2", type=float, default=DEFAULT_AZ2)

    p.add_argument(
        "--freq",
        type=int,
        default=DEFAULT_FREQ,
        help="Primary frequency code; default 1 = GPS L1.",
    )

    p.add_argument(
        "--all-frequencies",
        action="store_true",
        help="Diagnostic mode: retain all frequencies instead of freq=1.",
    )

    p.add_argument("--min-pknoise", type=float, default=DEFAULT_MIN_PKNOISE)
    p.add_argument("--min-amp", type=float, default=DEFAULT_MIN_AMP)
    p.add_argument("--max-delt-min", type=float, default=DEFAULT_MAX_DELT_MIN)

    p.add_argument("--emin", type=float, default=DEFAULT_EMIN)
    p.add_argument("--emax", type=float, default=DEFAULT_EMAX)

    p.add_argument("--bin-minutes", type=int, default=BIN_MINUTES)

    p.add_argument(
        "--no-track-normalization",
        action="store_true",
        help="Skip satellite+frequency track normalization.",
    )

    return p.parse_args()


def load_tide_models(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])

    required = [
        "time",
        "EOT20_heightm",
        "GOT5.5_heightm",
        "GOT5.6_heightm",
        "FES2022_heightm",
    ]

    for name in required:
        if name not in header:
            raise RuntimeError(
                f"Tide workbook is missing required column: {name}"
            )

    cols = {name: header.index(name) for name in required}

    times = []
    values = {name: [] for name in required[1:]}

    for row in rows[1:]:
        t = row[cols["time"]]

        if not isinstance(t, datetime):
            continue

        vals = {}
        valid = True

        for name in required[1:]:
            v = finite(row[cols[name]])
            if v is None:
                valid = False
                break
            vals[name] = v

        if not valid:
            continue

        times.append(t)
        for name in required[1:]:
            values[name].append(vals[name])

    wb.close()

    if len(times) < 2:
        raise RuntimeError("Insufficient tide-model points.")

    epoch = np.array(
        [(t - times[0]).total_seconds() for t in times],
        dtype=float,
    )

    interpolators = {}

    for name in values:
        arr = np.asarray(values[name], dtype=float)

        def make_interp(arr):
            def interp(query_dt):
                q = (query_dt - times[0]).total_seconds()
                if q < epoch[0] or q > epoch[-1]:
                    return math.nan
                return float(np.interp(q, epoch, arr))
            return interp

        interpolators[name] = make_interp(arr)

    return times, interpolators


def load_records(path):
    records = []

    with open(path, newline="") as f:
        reader = csv.DictReader(f)

        required = [
            "sat",
            "freq",
            "solution_time_utc",
            "RH_m",
            "Azim",
            "Amp",
            "PkNoise",
            "delT_min",
            "eminO",
            "emaxO",
        ]

        missing = [x for x in required if x not in reader.fieldnames]
        if missing:
            raise RuntimeError(
                f"CSV missing required columns: {missing}"
            )

        for source_row, row in enumerate(reader, start=2):
            try:
                dt = datetime.fromisoformat(
                    row["solution_time_utc"]
                )

                sat = int(float(row["sat"]))
                freq = int(float(row["freq"]))

                rh = finite(row["RH_m"])
                az = finite(row["Azim"])
                amp = finite(row["Amp"])
                pkn = finite(row["PkNoise"])
                delt = finite(row["delT_min"])
                emin = finite(row["eminO"])
                emax = finite(row["emaxO"])

                if any(
                    x is None
                    for x in [rh, az, amp, pkn, delt, emin, emax]
                ):
                    continue

                doy = int(
                    (dt.date() - datetime(dt.year, 1, 1).date()).days + 1
                )

                records.append(
                    Record(
                        source_row=source_row,
                        doy=doy,
                        dt=dt,
                        sat=sat,
                        freq=freq,
                        rh_m=rh,
                        az=az,
                        amp=amp,
                        pkn=pkn,
                        delt_min=delt,
                        emin=emin,
                        emax=emax,
                    )
                )

            except (ValueError, TypeError, KeyError):
                continue

    return records


def azimuth_in_sector(az, az1, az2):
    az = az % 360.0
    az1 = az1 % 360.0
    az2 = az2 % 360.0

    if az1 <= az2:
        return az1 <= az <= az2

    return az >= az1 or az <= az2


def assign_tides(records, interpolators):
    for r in records:
        r.tide_eot20 = interpolators["EOT20_heightm"](r.dt)
        r.tide_got55 = interpolators["GOT5.5_heightm"](r.dt)
        r.tide_got56 = interpolators["GOT5.6_heightm"](r.dt)
        r.tide_fes2022 = interpolators["FES2022_heightm"](r.dt)

        r.water_level_m = REFERENCE_HEIGHT_M - r.rh_m


def filter_records(records, args):
    included = []
    excluded = []

    for r in records:
        reason = None

        if not args.all_frequencies and r.freq != args.freq:
            reason = "not_primary_frequency"

        elif not azimuth_in_sector(r.az, args.az1, args.az2):
            reason = "outside_ocean_azimuth"

        elif r.pkn < args.min_pknoise:
            reason = "pknoise_below_threshold"

        elif r.amp < args.min_amp:
            reason = "amplitude_below_threshold"

        elif r.delt_min > args.max_delt_min:
            reason = "arc_too_long"

        elif r.emin < args.emin or r.emax > args.emax:
            reason = "elevation_window_mismatch"

        elif not all(
            math.isfinite(x)
            for x in [
                r.tide_eot20,
                r.tide_got55,
                r.tide_got56,
                r.tide_fes2022,
            ]
        ):
            reason = "outside_tide_coverage"

        if reason is None:
            included.append(r)
        else:
            excluded.append((r, reason))

    return included, excluded


def apply_anomalies(records):
    if not records:
        return

    median_water = float(
        np.median([r.water_level_m for r in records])
    )

    for r in records:
        r.water_anomaly_m = (
            r.water_level_m
            - median_water
        )

    groups = defaultdict(list)

    for r in records:
        groups[(r.sat, r.freq)].append(r)

    track_medians = {
        key: float(
            np.median(
                [r.water_level_m for r in group]
            )
        )
        for key, group in groups.items()
        if len(group) >= MIN_TRACK_OBS
    }

    for r in records:
        key = (r.sat, r.freq)

        if key in track_medians:
            r.track_anomaly_m = (
                r.water_level_m
                - track_medians[key]
            )
        else:
            r.track_anomaly_m = math.nan


def bin_records(records, minutes):
    if not records:
        return []

    grouped = defaultdict(list)

    for r in records:
        total_minutes = (
            r.dt.hour * 60
            + r.dt.minute
        )

        bin_min = (
            total_minutes // minutes
        ) * minutes

        dt_bin = r.dt.replace(
            hour=bin_min // 60,
            minute=bin_min % 60,
            second=0,
            microsecond=0,
        )

        grouped[dt_bin].append(r)

    output = []

    for dt_bin, group in sorted(grouped.items()):
        water = np.array(
            [r.water_level_m for r in group],
            dtype=float,
        )

        anomaly = np.array(
            [r.water_anomaly_m for r in group],
            dtype=float,
        )

        track = np.array(
            [r.track_anomaly_m for r in group],
            dtype=float,
        )

        tides = {
            "EOT20": np.mean(
                [r.tide_eot20 for r in group]
            ),
            "GOT5.5": np.mean(
                [r.tide_got55 for r in group]
            ),
            "GOT5.6": np.mean(
                [r.tide_got56 for r in group]
            ),
            "FES2022": np.mean(
                [r.tide_fes2022 for r in group]
            ),
        }

        output.append(
            {
                "datetime": dt_bin,
                "n": len(group),
                "water_level_median": float(np.median(water)),
                "water_anomaly_median": float(np.median(anomaly)),
                "track_anomaly_median": (
                    float(np.nanmedian(track))
                    if np.any(np.isfinite(track))
                    else math.nan
                ),
                **{
                    f"tide_{k}": float(v)
                    for k, v in tides.items()
                },
                "satellites": ",".join(
                    str(x)
                    for x in sorted(
                        set(r.sat for r in group)
                    )
                ),
                "frequencies": ",".join(
                    str(x)
                    for x in sorted(
                        set(r.freq for r in group)
                    )
                ),
            }
        )

    return output


def write_arc_csv(path, records):
    fields = [
        "source_row",
        "datetime_utc",
        "doy",
        "sat",
        "freq",
        "azimuth_deg",
        "RH_m",
        "water_level_m",
        "water_anomaly_m",
        "track_anomaly_m",
        "Amp",
        "PkNoise",
        "delT_min",
        "eminO",
        "emaxO",
        "tide_EOT20_m",
        "tide_GOT5.5_m",
        "tide_GOT5.6_m",
        "tide_FES2022_m",
        "tide_ensemble_m",
    ]

    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        for r in sorted(records, key=lambda x: x.dt):
            tide_ensemble = float(
                np.mean(
                    [
                        r.tide_eot20,
                        r.tide_got55,
                        r.tide_got56,
                        r.tide_fes2022,
                    ]
                )
            )

            w.writerow(
                {
                    "source_row": r.source_row,
                    "datetime_utc": r.dt.isoformat(),
                    "doy": r.doy,
                    "sat": r.sat,
                    "freq": r.freq,
                    "azimuth_deg": f"{r.az:.4f}",
                    "RH_m": f"{r.rh_m:.6f}",
                    "water_level_m": f"{r.water_level_m:.6f}",
                    "water_anomaly_m": f"{r.water_anomaly_m:.6f}",
                    "track_anomaly_m": (
                        f"{r.track_anomaly_m:.6f}"
                        if math.isfinite(r.track_anomaly_m)
                        else ""
                    ),
                    "Amp": f"{r.amp:.4f}",
                    "PkNoise": f"{r.pkn:.4f}",
                    "delT_min": f"{r.delt_min:.4f}",
                    "eminO": f"{r.emin:.4f}",
                    "emaxO": f"{r.emax:.4f}",
                    "tide_EOT20_m": f"{r.tide_eot20:.6f}",
                    "tide_GOT5.5_m": f"{r.tide_got55:.6f}",
                    "tide_GOT5.6_m": f"{r.tide_got56:.6f}",
                    "tide_FES2022_m": f"{r.tide_fes2022:.6f}",
                    "tide_ensemble_m": f"{tide_ensemble:.6f}",
                }
            )


def write_excluded_csv(path, excluded):
    fields = [
        "source_row",
        "datetime_utc",
        "doy",
        "sat",
        "freq",
        "azimuth_deg",
        "RH_m",
        "Amp",
        "PkNoise",
        "delT_min",
        "eminO",
        "emaxO",
        "reason",
    ]

    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        for r, reason in sorted(
            excluded,
            key=lambda x: x[0].dt,
        ):
            w.writerow(
                {
                    "source_row": r.source_row,
                    "datetime_utc": r.dt.isoformat(),
                    "doy": r.doy,
                    "sat": r.sat,
                    "freq": r.freq,
                    "azimuth_deg": f"{r.az:.4f}",
                    "RH_m": f"{r.rh_m:.6f}",
                    "Amp": f"{r.amp:.4f}",
                    "PkNoise": f"{r.pkn:.4f}",
                    "delT_min": f"{r.delt_min:.4f}",
                    "eminO": f"{r.emin:.4f}",
                    "emaxO": f"{r.emax:.4f}",
                    "reason": reason,
                }
            )


def write_binned_csv(path, binned):
    if not binned:
        Path(path).write_text("")
        return

    fields = list(binned[0].keys())

    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(
            [
                {
                    k: (
                        v.isoformat()
                        if isinstance(v, datetime)
                        else v
                    )
                    for k, v in row.items()
                }
                for row in binned
            ]
        )


def model_metrics(records, value_key):
    gnss = np.array(
        [getattr(r, value_key) for r in records],
        dtype=float,
    )

    models = {
        "EOT20": np.array(
            [r.tide_eot20 for r in records], dtype=float
        ),
        "GOT5.5": np.array(
            [r.tide_got55 for r in records], dtype=float
        ),
        "GOT5.6": np.array(
            [r.tide_got56 for r in records], dtype=float
        ),
        "FES2022": np.array(
            [r.tide_fes2022 for r in records], dtype=float
        ),
    }

    result = {}

    for name, tide in models.items():
        m = np.isfinite(gnss) & np.isfinite(tide)

        if m.sum() < 3:
            result[name] = None
            continue

        g = gnss[m]
        t = tide[m]

        offset = float(np.mean(g - t))
        residual = g - (t + offset)

        slope, intercept, rms_fit = linear_fit(t, g)

        result[name] = {
            "n": int(m.sum()),
            "r": pearson(g, t),
            "bias_offset_m": offset,
            "bias_removed_rms_cm": float(
                np.sqrt(np.mean(residual ** 2)) * 100.0
            ),
            "slope": slope,
            "fit_intercept": intercept,
            "fit_rms_cm": rms_fit * 100.0,
        }

    return result


def write_summary(path, args, raw_count, included, excluded, metrics):
    with open(path, "w") as f:
        f.write("=" * 80 + "\n")
        f.write("CONTROLLED MARCONI OCEAN GNSS-IR PRODUCT\n")
        f.write("=" * 80 + "\n\n")

        f.write("INPUTS\n")
        f.write(f"CSV: {args.csv}\n")
        f.write(f"Tide workbook: {args.tide}\n\n")

        f.write("PRIMARY FILTERS\n")
        f.write(f"Frequency: {'ALL' if args.all_frequencies else args.freq}\n")
        f.write(
            f"Ocean azimuth: {args.az1:.1f} to {args.az2:.1f} deg\n"
        )
        f.write(
            f"PkNoise >= {args.min_pknoise:.2f}\n"
        )
        f.write(
            f"Amp >= {args.min_amp:.2f}\n"
        )
        f.write(
            f"Arc duration <= {args.max_delt_min:.1f} min\n"
        )
        f.write(
            f"Elevation window: {args.emin:.1f}-{args.emax:.1f} deg\n"
        )
        f.write(
            f"Orthometric/reference height: "
            f"{REFERENCE_HEIGHT_M:.3f} m\n\n"
        )

        f.write("COUNTS\n")
        f.write(f"Input usable records: {raw_count}\n")
        f.write(f"Included records: {len(included)}\n")
        f.write(f"Excluded records: {len(excluded)}\n\n")

        f.write("INCLUSION BY SAT/FREQ\n")
        counts = Counter(
            (r.sat, r.freq)
            for r in included
        )
        for (sat, freq), n in sorted(counts.items()):
            f.write(
                f"  sat={sat:3d} freq={freq:3d}: {n}\n"
            )

        f.write("\nEXCLUSION REASONS\n")
        reasons = Counter(
            reason
            for _, reason in excluded
        )
        for reason, n in reasons.most_common():
            f.write(
                f"  {reason}: {n}\n"
            )

        f.write("\nMODEL COMPARISON: RAW WATER-LEVEL ANOMALY\n")
        f.write(
            "(absolute datum is not assumed; constant offset is removed)\n\n"
        )

        for model, values in metrics.items():
            if values is None:
                f.write(
                    f"{model}: insufficient data\n"
                )
                continue

            f.write(
                f"{model}: "
                f"n={values['n']} "
                f"r={values['r']:+.4f} "
                f"offset={values['bias_offset_m']:+.4f} m "
                f"bias_removed_RMS={values['bias_removed_rms_cm']:.2f} cm "
                f"slope={values['slope']:+.4f} "
                f"fit_RMS={values['fit_rms_cm']:.2f} cm\n"
            )

        f.write("\nIMPORTANT\n")
        f.write(
            "This product does NOT use gnssrefl subdaily as its primary "
            "water-level estimator. Each individual GNSS-IR arc remains "
            "auditable. No spline is used in the primary product.\n"
        )
        f.write(
            "Water-level convention is H_ortho - RH. Absolute comparison "
            "to the tide datum requires independent vertical-datum validation. "
            "The anomaly comparison is the preferred first diagnostic.\n"
        )


def make_plots(records, binned, args):
    # --------------------------------------------------------------
    # Plot 1: individual anomaly observations vs tide anomaly
    # --------------------------------------------------------------

    if not records:
        return

    gnss = np.array(
        [r.water_anomaly_m for r in records]
    )

    tide = np.array(
        [
            np.mean(
                [
                    r.tide_eot20,
                    r.tide_got55,
                    r.tide_got56,
                    r.tide_fes2022,
                ]
            )
            for r in records
        ]
    )

    tide_anom = tide - np.median(tide)

    fig, ax = plt.subplots(
        figsize=(14, 7)
    )

    ax.axhline(
        0.0,
        linewidth=1.0,
    )

    ax.plot(
        [r.dt for r in records],
        gnss,
        marker="o",
        linestyle="None",
        markersize=5,
        label="GNSS-IR ocean arc water anomaly",
    )

    ax.plot(
        [r.dt for r in records],
        tide_anom,
        linewidth=2.0,
        label="Tide-model ensemble anomaly",
    )

    ax.set_xlabel("UTC")
    ax.set_ylabel("Water-level anomaly (m)")
    ax.set_title(
        "Marconi — Controlled Ocean GNSS-IR vs Tide Anomaly\n"
        f"GPS L1, azimuth {args.az1:g}–{args.az2:g}°, "
        f"no spline"
    )

    ax.grid(True, alpha=0.3)
    ax.legend()

    fig.autofmt_xdate()
    fig.tight_layout()

    fig.savefig(
        "ocean_gnssir_anomaly_vs_tide.png",
        dpi=200,
    )

    plt.close(fig)

    # --------------------------------------------------------------
    # Plot 2: absolute values, clearly labeled as datum-dependent
    # --------------------------------------------------------------

    fig, ax1 = plt.subplots(
        figsize=(14, 7)
    )

    ax1.plot(
        [r.dt for r in records],
        [r.water_level_m for r in records],
        marker="o",
        linestyle="-",
        linewidth=1.0,
        label="GNSS-IR H_ortho - RH",
    )

    ax1.set_ylabel(
        "GNSS-IR water-level convention (m)"
    )

    tide_mean = [
        np.mean(
            [
                r.tide_eot20,
                r.tide_got55,
                r.tide_got56,
                r.tide_fes2022,
            ]
        )
        for r in records
    ]

    ax2 = ax1.twinx()

    ax2.plot(
        [r.dt for r in records],
        tide_mean,
        linewidth=2.0,
        label="Tide-model ensemble mean",
    )

    ax2.set_ylabel(
        "Tide-model height (m)"
    )

    ax1.set_xlabel("UTC")

    ax1.set_title(
        "Marconi — Absolute GNSS-IR Water-Level Convention vs Tide Model\n"
        "Separate axes: absolute vertical datums are NOT assumed equivalent"
    )

    ax1.grid(True, alpha=0.3)

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()

    ax1.legend(
        lines1 + lines2,
        labels1 + labels2,
        loc="best",
    )

    fig.autofmt_xdate()
    fig.tight_layout()

    fig.savefig(
        "ocean_gnssir_absolute_vs_tide.png",
        dpi=200,
    )

    plt.close(fig)


def main():
    args = parse_args()

    print()
    print("=" * 80)
    print("CONTROLLED MARCONI OCEAN GNSS-IR PRODUCT")
    print("=" * 80)

    print(f"Input CSV       : {args.csv}")
    print(f"Tide workbook   : {args.tide}")
    print(
        f"Ocean azimuth   : {args.az1:.1f}–{args.az2:.1f}°"
    )
    print(
        f"Frequency       : "
        f"{'ALL' if args.all_frequencies else args.freq}"
    )
    print(
        f"PkNoise minimum : {args.min_pknoise:.2f}"
    )
    print(
        f"Amplitude min   : {args.min_amp:.2f}"
    )
    print(
        f"Max arc length  : {args.max_delt_min:.1f} min"
    )
    print(
        f"Elevation       : {args.emin:.1f}–{args.emax:.1f}°"
    )
    print(
        f"Reference H     : {REFERENCE_HEIGHT_M:.3f} m"
    )

    if not args.csv.exists():
        raise SystemExit(
            f"ERROR: input CSV not found: {args.csv}"
        )

    if not args.tide.exists():
        raise SystemExit(
            f"ERROR: tide workbook not found: {args.tide}"
        )

    tide_times, interpolators = load_tide_models(
        args.tide
    )

    print()
    print(
        f"Tide coverage: {tide_times[0]} through {tide_times[-1]}"
    )
    print(
        f"Tide points: {len(tide_times)}"
    )

    raw = load_records(args.csv)

    print(
        f"Input usable CSV records: {len(raw)}"
    )

    assign_tides(
        raw,
        interpolators
    )

    included, excluded = filter_records(
        raw,
        args,
    )

    apply_anomalies(
        included
    )

    print()
    print("=" * 80)
    print("FILTER RESULT")
    print("=" * 80)

    print(
        f"Included: {len(included)}"
    )
    print(
        f"Excluded: {len(excluded)}"
    )

    print()
    print("Included SAT/FREQ:")
    for key, n in sorted(
        Counter(
            (r.sat, r.freq)
            for r in included
        ).items()
    ):
        print(
            f"  sat={key[0]:3d} "
            f"freq={key[1]:3d} "
            f"n={n}"
        )

    print()
    print("Exclusion reasons:")
    for reason, n in Counter(
        x[1] for x in excluded
    ).most_common():
        print(
            f"  {reason}: {n}"
        )

    if len(included) < 3:
        raise SystemExit(
            "ERROR: fewer than 3 records survived. "
            "Do not interpret the product."
        )

    # --------------------------------------------------------------
    # Metrics: raw water anomaly
    # --------------------------------------------------------------

    metrics = model_metrics(
        included,
        "water_anomaly_m",
    )

    print()
    print("=" * 80)
    print("TIDE COMPARISON — WATER-LEVEL ANOMALY")
    print("=" * 80)

    for model, values in metrics.items():
        if values is None:
            continue

        print(
            f"{model:8s} "
            f"n={values['n']:3d} "
            f"r={values['r']:+.4f} "
            f"bias-removed RMS="
            f"{values['bias_removed_rms_cm']:7.2f} cm "
            f"slope={values['slope']:+.4f}"
        )

    # --------------------------------------------------------------
    # Bin observations for a diagnostic 30-minute product.
    # --------------------------------------------------------------

    binned = bin_records(
        included,
        args.bin_minutes,
    )

    # Track-normalized diagnostic product if possible.
    if not args.no_track_normalization:
        valid_track = [
            r
            for r in included
            if math.isfinite(
                r.track_anomaly_m
            )
        ]

        print()
        print(
            f"Track-normalized records: "
            f"{len(valid_track)}"
        )

    # --------------------------------------------------------------
    # Write files.
    # --------------------------------------------------------------

    write_arc_csv(
        "ocean_gnssir_arc_product.csv",
        included,
    )

    write_excluded_csv(
        "ocean_gnssir_excluded_arcs.csv",
        excluded,
    )

    write_binned_csv(
        "ocean_gnssir_30min_binned.csv",
        binned,
    )

    write_summary(
        "ocean_gnssir_summary.txt",
        args,
        len(raw),
        included,
        excluded,
        metrics,
    )

    make_plots(
        included,
        binned,
        args,
    )

    print()
    print("=" * 80)
    print("OUTPUTS")
    print("=" * 80)

    for name in [
        "ocean_gnssir_arc_product.csv",
        "ocean_gnssir_excluded_arcs.csv",
        "ocean_gnssir_30min_binned.csv",
        "ocean_gnssir_summary.txt",
        "ocean_gnssir_anomaly_vs_tide.png",
        "ocean_gnssir_absolute_vs_tide.png",
    ]:
        print(f"  {name}")

    print()
    print("=" * 80)
    print("IMPORTANT")
    print("=" * 80)
    print(
        "The primary product is individual ocean GNSS-IR arc solutions."
    )
    print(
        "No subdaily spline is used."
    )
    print(
        "GPS L1 is kept separate from other frequencies by default."
    )
    print(
        "Absolute GNSS-IR water level uses H_ortho - RH, but the first"
        " scientific comparison is made on anomalies after removing a"
        " constant median, because vertical-datum equivalence has not"
        " yet been independently demonstrated."
    )
    print(
        "The 35-135 degree mask is a conservative starting point based"
        " on the corrected Marconi reflection-zone investigation; it is"
        " NOT a substitute for a true Fresnel-footprint/coastline test."
    )
    print()
    print("DONE")


if __name__ == "__main__":
    main()
