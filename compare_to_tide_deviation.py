#!/usr/bin/env python3
"""
compare_to_tide_deviation.py

Computes actual deviation statistics between the GNSS-IR-derived
water level (gnssrefl's own evenly-sampled spline output) and an
external tide reference -- not just correlation (whether the shapes
track together), but the real magnitude of disagreement.

Reports two numbers, deliberately kept separate:

    RAW RMS difference
        The literal difference between the two series. Includes
        whatever unresolved vertical-datum offset exists between
        gnssrefl_orthometric_height (not independently confirmed to
        be NAVD88-referenced) and the tide model's own reference --
        so this number conflates "does the tidal SIGNAL agree" with
        "do the two vertical datums happen to agree", which have not
        been established to be the same thing.

    DE-MEANED RMS difference (mean offset removed first)
        Isolates whether the SHAPE and AMPLITUDE of the two signals
        agree, independent of any constant vertical offset between
        them. This is the more interpretable number until the datum
        question is separately resolved.

Confirmed, real file format used here (from direct inspection
earlier in this project): gnssrefl's own evenly-sampled spline
output file has the date in columns 3-8 (YYYY MM DD HH MM SS,
1-indexed / columns 2-7 0-indexed) and column 9 (0-indexed column 8)
is the already-computed quasi-sea-level value (Hortho - RH), in
meters -- NOT a fractional day-of-year, which an earlier version of
a different tool in this project incorrectly assumed.

Usage:
    python3 compare_to_tide_deviation.py \\
        --spline-file products/refl_code/Files/usgs/usgs_spline_out.txt \\
        --tide-file marconi_tides_sherwood.xlsx \\
        --tide-time-col time \\
        --tide-value-col EOT20_heightm
"""

from __future__ import annotations

import argparse
import math
from datetime import datetime
from pathlib import Path

import numpy as np


def load_spline_output(path: Path):
    """
    Parses gnssrefl's evenly-sampled spline output file using the
    confirmed real column layout: columns 3-8 (1-indexed) are
    YYYY MM DD HH MM SS, column 9 is the quasi-sea-level value
    (Hortho - RH), in meters.
    """
    times, values = [], []

    with open(path, errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            cols = line.split()
            if len(cols) < 9:
                continue
            try:
                year = int(float(cols[2]))
                month = int(float(cols[3]))
                day = int(float(cols[4]))
                hour = int(float(cols[5]))
                minute = int(float(cols[6]))
                second = int(float(cols[7]))
                water_level = float(cols[8])
            except (ValueError, IndexError):
                continue

            try:
                dt = datetime(year, month, day, hour, minute, second)
            except ValueError:
                continue

            times.append(dt)
            values.append(water_level)

    return times, np.asarray(values, float)


def load_tide_reference(path: Path, time_col: str, value_col: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]

    if time_col not in header:
        raise SystemExit(f"Time column {time_col!r} not found. Available: {header}")
    if value_col not in header:
        raise SystemExit(f"Value column {value_col!r} not found. Available: {header}")

    time_idx = header.index(time_col)
    value_idx = header.index(value_col)

    times, values = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[time_idx]
        if not isinstance(t, datetime):
            continue
        try:
            v = float(row[value_idx])
        except (TypeError, ValueError):
            continue
        if not math.isfinite(v):
            continue
        times.append(t)
        values.append(v)
    wb.close()

    return times, np.asarray(values, float)


def interpolate_reference(ref_times, ref_values, query_times):
    """Linear interpolation of the reference signal to arbitrary query times."""
    ref_epoch = np.array([(t - ref_times[0]).total_seconds() for t in ref_times])
    query_epoch = np.array([(t - ref_times[0]).total_seconds() for t in query_times])

    order = np.argsort(ref_epoch)
    ref_epoch_sorted = ref_epoch[order]
    ref_values_sorted = ref_values[order]

    interpolated = np.interp(
        query_epoch, ref_epoch_sorted, ref_values_sorted,
        left=np.nan, right=np.nan,
    )
    return interpolated


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--spline-file", required=True)
    p.add_argument("--tide-file", required=True)
    p.add_argument("--tide-time-col", default="time")
    p.add_argument("--tide-value-col", required=True)
    args = p.parse_args()

    spline_times, spline_values = load_spline_output(Path(args.spline_file))
    print(f"Loaded {len(spline_times)} GNSS-IR spline points")
    if not spline_times:
        raise SystemExit("No spline points loaded -- check --spline-file path/format.")

    tide_times, tide_values = load_tide_reference(
        Path(args.tide_file), args.tide_time_col, args.tide_value_col
    )
    print(f"Loaded {len(tide_times)} tide model points")

    tide_at_spline_times = interpolate_reference(tide_times, tide_values, spline_times)

    valid = np.isfinite(tide_at_spline_times) & np.isfinite(spline_values)
    n_valid = int(np.sum(valid))
    print(f"{n_valid} points have both a GNSS-IR value and a tide-model value to compare")

    if n_valid < 5:
        raise SystemExit("Too few overlapping points for a meaningful comparison.")

    gnss = spline_values[valid]
    tide = tide_at_spline_times[valid]

    diff = gnss - tide
    raw_rms = float(np.sqrt(np.mean(diff**2)))
    mean_offset = float(np.mean(diff))
    demeaned_diff = diff - mean_offset
    demeaned_rms = float(np.sqrt(np.mean(demeaned_diff ** 2)))
    demeaned_mad = float(np.mean(np.abs(demeaned_diff)))
    demeaned_median_ad = float(np.median(np.abs(demeaned_diff)))
    corr = float(np.corrcoef(gnss, tide)[0, 1]) if np.std(gnss) > 0 and np.std(tide) > 0 else float("nan")

    print()
    print("=" * 64)
    print("DEVIATION: GNSS-IR water level vs. tide model")
    print("=" * 64)
    print(f"Points compared         : {n_valid}")
    print(f"Mean offset (GNSS - tide): {mean_offset:+.3f} m")
    print(f"RAW RMS difference       : {raw_rms:.3f} m")
    print(f"  (includes any unresolved vertical-datum offset -- see script docstring)")
    print(f"DE-MEANED RMS difference : {demeaned_rms:.3f} m")
    print(f"  (offset removed; squares each error before averaging, so this weights")
    print(f"   occasional larger deviations more heavily than a plain average would)")
    print(f"DE-MEANED MEAN ABSOLUTE DEVIATION : {demeaned_mad:.3f} m  ({demeaned_mad*100:.1f} cm)")
    print(f"  (offset removed; this is the literal, defensible number for a claim like")
    print(f'   "on average, the GPS data tracks the tide model within {demeaned_mad*100:.1f} cm")')
    print(f"DE-MEANED MEDIAN ABSOLUTE DEVIATION: {demeaned_median_ad:.3f} m  ({demeaned_median_ad*100:.1f} cm)")
    print(f"  (less sensitive to any remaining outliers than the mean above)")
    print(f"Correlation              : {corr:+.4f}")
    print("=" * 64)


if __name__ == "__main__":
    main()
