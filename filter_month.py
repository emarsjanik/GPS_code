#!/usr/bin/env python3
"""
filter_month.py

Extracts a single month's rows from this project's whole-record
timeseries files, so each month's folder on the archive server
contains genuinely that month's data rather than a copy of the
entire record.

Handles two real gnssrefl output formats, both confirmed against
actual files from this station:

  usgs_spline_out.txt
      Evenly-sampled spline, 30-minute steps.
      Columns (0-indexed): 2=year, 3=month, 4=day, 8=water level.

  usgs_<year>_subdaily_edit.txt
      Individual satellite-arc retrievals that passed quality
      control. Columns (0-indexed): 0=year, 17=MM, 18=DD.

Comment lines (starting with %) are preserved verbatim at the top of
each output file -- they carry the column descriptions and the
gnssrefl version, which a future reader of an archived file needs.

Optionally also writes a merged CSV pairing the GNSS-IR water level
with an interpolated tide-model value, using the same logic as
analysis_tools/export_timeseries.py.

Usage:
    python3 filter_month.py --year 2026 --month 8 \\
        --products-dir products/refl_code/Files/usgs \\
        --output-dir /tmp/august2026

    # with a tide comparison CSV as well
    python3 filter_month.py --year 2026 --month 8 \\
        --products-dir products/refl_code/Files/usgs \\
        --output-dir /tmp/august2026 \\
        --tide-file marconi_tides_sherwood.xlsx \\
        --tide-value-col EOT20_heightm
"""

from __future__ import annotations

import argparse
import csv
import math
from datetime import datetime
from pathlib import Path


MONTH_NAMES = {
    1: "january", 2: "february", 3: "march", 4: "april",
    5: "may", 6: "june", 7: "july", 8: "august",
    9: "september", 10: "october", 11: "november", 12: "december",
}


def month_folder(year: int, month: int) -> str:
    return f"{MONTH_NAMES[month]}{year}"


def filter_by_columns(src: Path, dest: Path, year: int, month: int,
                      year_col: int, month_col: int) -> int:
    """
    Copies src to dest, keeping comment lines and only those data
    rows whose year/month columns match. Returns the number of data
    rows kept.
    """
    kept = 0
    with open(src, errors="replace") as fin, open(dest, "w") as fout:
        for line in fin:
            if line.startswith("%"):
                fout.write(line)
                continue
            cols = line.split()
            if len(cols) <= max(year_col, month_col):
                continue
            try:
                row_year = int(float(cols[year_col]))
                row_month = int(float(cols[month_col]))
            except ValueError:
                continue
            if row_year == year and row_month == month:
                fout.write(line)
                kept += 1
    return kept


def load_spline(path: Path):
    times, values = [], []
    with open(path, errors="replace") as f:
        for line in f:
            if line.startswith("%") or not line.strip():
                continue
            cols = line.split()
            if len(cols) < 9:
                continue
            try:
                dt = datetime(
                    int(float(cols[2])), int(float(cols[3])), int(float(cols[4])),
                    int(float(cols[5])), int(float(cols[6])), int(float(cols[7])),
                )
                values.append(float(cols[8]))
                times.append(dt)
            except (ValueError, IndexError):
                continue
    return times, values


def load_tide(path: Path, time_col: str, value_col: str):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    ti, vi = header.index(time_col), header.index(value_col)

    times, values = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[ti]
        if not isinstance(t, datetime):
            continue
        try:
            v = float(row[vi])
        except (TypeError, ValueError):
            continue
        if math.isfinite(v):
            times.append(t)
            values.append(v)
    wb.close()
    return times, values


def interp_tide(tide_times, tide_values, query_times):
    """Linear interpolation without requiring numpy."""
    if not tide_times:
        return [None] * len(query_times)

    pairs = sorted(zip(tide_times, tide_values))
    ts = [p[0] for p in pairs]
    vs = [p[1] for p in pairs]

    out = []
    for q in query_times:
        if q < ts[0] or q > ts[-1]:
            out.append(None)
            continue
        lo, hi = 0, len(ts) - 1
        while hi - lo > 1:
            mid = (lo + hi) // 2
            if ts[mid] <= q:
                lo = mid
            else:
                hi = mid
        span = (ts[hi] - ts[lo]).total_seconds()
        if span == 0:
            out.append(vs[lo])
        else:
            frac = (q - ts[lo]).total_seconds() / span
            out.append(vs[lo] + frac * (vs[hi] - vs[lo]))
    return out


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--year", type=int, required=True)
    p.add_argument("--month", type=int, required=True)
    p.add_argument("--products-dir", required=True)
    p.add_argument("--output-dir", required=True)
    p.add_argument("--station-code", default="usgs")
    p.add_argument("--tide-file")
    p.add_argument("--tide-time-col", default="time")
    p.add_argument("--tide-value-col")
    args = p.parse_args()

    if args.month not in MONTH_NAMES:
        raise SystemExit(f"Invalid month: {args.month}")

    products = Path(args.products_dir)
    outdir = Path(args.output_dir)
    outdir.mkdir(parents=True, exist_ok=True)

    label = month_folder(args.year, args.month)
    print(f"Filtering to {label}")
    print()

    wrote_any = False

    # --- spline output ---
    spline_src = products / f"{args.station_code}_spline_out.txt"
    if spline_src.exists():
        dest = outdir / f"{args.station_code}_spline_out_{label}.txt"
        n = filter_by_columns(spline_src, dest, args.year, args.month,
                              year_col=2, month_col=3)
        if n > 0:
            print(f"  {dest.name}: {n} rows")
            wrote_any = True
        else:
            dest.unlink(missing_ok=True)
            print(f"  {spline_src.name}: no rows for this month -- skipped")
    else:
        print(f"  {spline_src.name}: not found -- skipped")

    # --- subdaily retrievals ---
    subdaily_src = products / f"{args.station_code}_{args.year}_subdaily_edit.txt"
    if subdaily_src.exists():
        dest = outdir / f"{args.station_code}_subdaily_{label}.txt"
        n = filter_by_columns(subdaily_src, dest, args.year, args.month,
                              year_col=0, month_col=17)
        if n > 0:
            print(f"  {dest.name}: {n} rows")
            wrote_any = True
        else:
            dest.unlink(missing_ok=True)
            print(f"  {subdaily_src.name}: no rows for this month -- skipped")
    else:
        print(f"  {subdaily_src.name}: not found -- skipped")

    # --- merged CSV with tide comparison ---
    if args.tide_file and args.tide_value_col and spline_src.exists():
        times, values = load_spline(spline_src)
        sel = [(t, v) for t, v in zip(times, values)
               if t.year == args.year and t.month == args.month]

        if sel:
            tide_times, tide_values = load_tide(
                Path(args.tide_file), args.tide_time_col, args.tide_value_col)
            tide_at = interp_tide(tide_times, tide_values, [t for t, _ in sel])

            dest = outdir / f"{args.station_code}_timeseries_{label}.csv"
            with open(dest, "w", newline="") as f:
                w = csv.writer(f)
                w.writerow(["timestamp_utc", "gnss_ir_water_level_m",
                            f"tide_model_{args.tide_value_col}"])
                for (t, v), tv in zip(sel, tide_at):
                    w.writerow([t.strftime("%Y-%m-%d %H:%M:%S"),
                                f"{v:.4f}",
                                "" if tv is None else f"{tv:.4f}"])
            matched = sum(1 for tv in tide_at if tv is not None)
            print(f"  {dest.name}: {len(sel)} rows ({matched} with a tide value)")
            wrote_any = True

    print()
    if wrote_any:
        print(f"Wrote to: {outdir}")
    else:
        print("No data found for this month -- nothing written.")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
