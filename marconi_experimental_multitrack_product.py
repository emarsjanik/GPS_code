#!/usr/bin/env python3
"""
Marconi experimental multi-track GNSS-R water-level product.

Purpose
-------
Build an experimental, datum-corrected Marconi GNSS-R water-level product
from the strongest persistent tracks identified in the current analysis.

Current independent datum test:
    +0.242 m

Raw GNSS-R water level:
    18.665 - RH

Corrected GNSS-R water level:
    18.665 - RH + 0.242

PRIMARY TRACK SELECTION
-----------------------
A track is retained when all of the following are true:

    observations N >= 10
    |corrected median GNSS-R - EOT20| <= 0.05 m
    corrected RMS <= 0.20 m
    tide correlation r >= 0.93

These thresholds are deliberately conservative and are based on the
current Marconi analysis, not on a fitted production calibration.

Based on the current results supplied in the conversation, the expected
primary tracks are approximately:

    PRN 29 rising  ~113 deg
    PRN 9  rising  ~124 deg
    PRN 4  rising   ~96 deg

The script DOES NOT hard-code those tracks. It evaluates the current
result files and selects tracks automatically using the criteria above.

Products
--------
1. Individual corrected observations:
       marconi_experimental_multitrack_observations.csv

2. Selected track summary:
       marconi_experimental_multitrack_track_summary.csv

3. Daily combined product:
       marconi_experimental_multitrack_daily.csv

4. Summary:
       marconi_experimental_multitrack_summary.txt

5. Plots:
       marconi_experimental_multitrack_plots/
         01_corrected_tracks_vs_eot20.png
         02_daily_multitrack_vs_eot20.png
         03_track_residuals_vs_eot20.png
         04_selected_track_offsets.png

Important
---------
This is an EXPERIMENTAL product. It does not claim that the selected
tracks have independently confirmed ocean Fresnel geometry. Geometry
validation remains an explicit scientific filter to apply next.

No GNSS-IR processing parameters or production JSON files are modified.
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


# =====================================================================
# CONFIGURATION
# =====================================================================

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

H_ORTHO_M = 18.665
DATUM_SHIFT_M = 0.242

PRIMARY_TIDE = "EOT20_heightm"

AZ_TOL_DEG = 3.0

# Experimental product selection criteria.
MIN_TRACK_N = 10
MAX_ABS_CORRECTED_MEDIAN_M = 0.05
MAX_CORRECTED_RMS_M = 0.20
MIN_TIDE_R = 0.93

OUT_OBS = Path(
    "marconi_experimental_multitrack_observations.csv"
)

OUT_TRACKS = Path(
    "marconi_experimental_multitrack_track_summary.csv"
)

OUT_DAILY = Path(
    "marconi_experimental_multitrack_daily.csv"
)

OUT_SUMMARY = Path(
    "marconi_experimental_multitrack_summary.txt"
)

PLOT_DIR = Path(
    "marconi_experimental_multitrack_plots"
)


# =====================================================================
# HELPERS
# =====================================================================

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def utc_datetime(year, doy, utc_hours):
    return (
        datetime(year, 1, 1)
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


def circular_az_diff(a, b):
    d = abs(a - b) % 360.0
    return min(d, 360.0 - d)


def correlation(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 3:
        return math.nan

    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(
        np.corrcoef(x, y)[0, 1]
    )


def rms(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(
        np.sqrt(np.mean(x ** 2))
    )


def mae(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(
        np.mean(np.abs(x))
    )


# =====================================================================
# TIDE
# =====================================================================

def load_tide():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]

    time_col = header.index("time")
    tide_col = header.index(PRIMARY_TIDE)

    times = []
    values = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_col]
        v = finite(row[tide_col])

        if not isinstance(t, datetime):
            continue

        if v is None:
            continue

        times.append(t)
        values.append(v)

    wb.close()

    return (
        np.asarray(
            times,
            dtype="datetime64[ms]",
        ),
        np.asarray(
            values,
            dtype=float,
        ),
    )


def tide_at(
    tide_times,
    tide_values,
    dt,
):
    x = tide_times.astype("int64")
    q = np.datetime64(
        dt,
        "ms",
    ).astype("int64")

    if q < x[0] or q > x[-1]:
        return math.nan

    return float(
        np.interp(
            q,
            x,
            tide_values,
        )
    )


# =====================================================================
# GNSS-R RESULTS
# =====================================================================

def load_results():
    rows = []

    for path in sorted(
        RESULT_DIR.glob("*.txt")
    ):
        try:
            int(path.stem)
        except Exception:
            continue

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if not line or line.startswith("%"):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc_hours = float(c[4])
                az = float(c[5])
                amp = float(c[6])
                freq = int(float(c[10]))
                rise = int(float(c[11]))
                pkn = float(c[13])
                nobs = int(float(c[9]))
                delt = float(c[14])
            except Exception:
                continue

            if freq != 1:
                continue

            dt = utc_datetime(
                year,
                doy,
                utc_hours,
            )

            raw_wl = H_ORTHO_M - rh
            corrected_wl = (
                raw_wl + DATUM_SHIFT_M
            )

            rows.append({
                "datetime": dt,
                "year": year,
                "doy": doy,
                "sat": sat,
                "freq": freq,
                "rise": rise,
                "az": az,
                "RH_m": rh,
                "raw_wl_m": raw_wl,
                "corrected_wl_m": corrected_wl,
                "Amp": amp,
                "PkNoise": pkn,
                "NumbOf": nobs,
                "DelT_min": delt,
            })

    rows.sort(
        key=lambda r:
            r["datetime"]
    )

    return rows


def add_tide_values(
    rows,
    tide_times,
    tide_values,
):
    for row in rows:

        row["EOT20_m"] = tide_at(
            tide_times,
            tide_values,
            row["datetime"],
        )

        row["raw_residual_m"] = (
            row["raw_wl_m"]
            - row["EOT20_m"]
        )

        row["corrected_residual_m"] = (
            row["corrected_wl_m"]
            - row["EOT20_m"]
        )


# =====================================================================
# TRACK CLUSTERING
# =====================================================================

def build_track_groups(rows):
    base = defaultdict(list)

    for row in rows:
        base[
            (
                row["sat"],
                row["freq"],
                row["rise"],
            )
        ].append(row)

    tracks = []

    for key, group in base.items():

        group = sorted(
            group,
            key=lambda r:
                r["az"],
        )

        current = []
        prev_az = None

        for row in group:

            if (
                prev_az is None
                or circular_az_diff(
                    row["az"],
                    prev_az,
                ) <= AZ_TOL_DEG
            ):
                current.append(row)

            else:

                if len(current) >= MIN_TRACK_N:
                    tracks.append(current)

                current = [row]

            prev_az = row["az"]

        if len(current) >= MIN_TRACK_N:
            tracks.append(current)

    return tracks


# =====================================================================
# TRACK STATISTICS / SELECTION
# =====================================================================

def summarize_track(group):
    wl = np.asarray(
        [
            r["corrected_wl_m"]
            for r in group
        ],
        dtype=float,
    )

    tide = np.asarray(
        [
            r["EOT20_m"]
            for r in group
        ],
        dtype=float,
    )

    residual = (
        wl - tide
    )

    az = np.asarray(
        [
            r["az"]
            for r in group
        ],
        dtype=float,
    )

    valid = (
        np.isfinite(wl)
        & np.isfinite(tide)
        & np.isfinite(residual)
    )

    wl = wl[valid]
    tide = tide[valid]
    residual = residual[valid]

    if len(wl) >= 3:

        r = correlation(
            wl,
            tide,
        )

        slope = float(
            np.polyfit(
                tide,
                wl,
                1,
            )[0]
        )

    else:

        r = math.nan
        slope = math.nan

    return {
        "sat": group[0]["sat"],
        "freq": group[0]["freq"],
        "rise": group[0]["rise"],
        "n": len(wl),
        "n_days": len(
            {
                r0["doy"]
                for r0 in group
            }
        ),
        "az_mean_deg": float(
            np.mean(az)
        ),
        "az_sd_deg": float(
            np.std(az)
        ),
        "corrected_median_m": float(
            np.median(residual)
        ),
        "corrected_mean_m": float(
            np.mean(residual)
        ),
        "corrected_rms_m": rms(
            residual
        ),
        "corrected_mae_m": mae(
            residual
        ),
        "tide_r": r,
        "tide_slope": slope,
        "mean_amp": float(
            np.mean(
                [
                    r0["Amp"]
                    for r0 in group
                ]
            )
        ),
        "mean_pkn": float(
            np.mean(
                [
                    r0["PkNoise"]
                    for r0 in group
                ]
            )
        ),
    }


def is_selected(summary):
    return (
        summary["n"]
        >= MIN_TRACK_N
        and abs(
            summary[
                "corrected_median_m"
            ]
        )
        <= MAX_ABS_CORRECTED_MEDIAN_M
        and summary[
            "corrected_rms_m"
        ]
        <= MAX_CORRECTED_RMS_M
        and summary[
            "tide_r"
        ] >= MIN_TIDE_R
    )


# =====================================================================
# DAILY COMBINATION
# =====================================================================

def make_daily_product(
    selected_rows,
):
    by_day = defaultdict(list)

    for row in selected_rows:
        by_day[
            row["datetime"].date()
        ].append(row)

    daily = []

    for day, group in sorted(
        by_day.items()
    ):

        wl = np.asarray(
            [
                r["corrected_wl_m"]
                for r in group
            ],
            dtype=float,
        )

        tide = np.asarray(
            [
                r["EOT20_m"]
                for r in group
            ],
            dtype=float,
        )

        residual = (
            wl - tide
        )

        daily.append({
            "date": day,
            "n_observations": len(group),
            "n_tracks": len(
                {
                    (
                        r["sat"],
                        r["freq"],
                        r["rise"],
                    )
                    for r in group
                }
            ),
            "GNSSR_median_m": float(
                np.median(wl)
            ),
            "GNSSR_mean_m": float(
                np.mean(wl)
            ),
            "GNSSR_sd_m": float(
                np.std(wl)
            ),
            "EOT20_median_at_obs_m": float(
                np.median(tide)
            ),
            "median_residual_m": float(
                np.median(residual)
            ),
            "residual_mad_m": float(
                1.4826
                * np.median(
                    np.abs(
                        residual
                        - np.median(residual)
                    )
                )
            ),
        })

    return daily


# =====================================================================
# CSV OUTPUTS
# =====================================================================

def write_observation_csv(
    rows,
):
    fields = [
        "datetime",
        "year",
        "doy",
        "sat",
        "freq",
        "rise",
        "az",
        "RH_m",
        "raw_wl_m",
        "corrected_wl_m",
        "EOT20_m",
        "raw_residual_m",
        "corrected_residual_m",
        "Amp",
        "PkNoise",
        "NumbOf",
        "DelT_min",
        "selected_track",
    ]

    with open(
        OUT_OBS,
        "w",
        newline="",
    ) as handle:

        writer = csv.DictWriter(
            handle,
            fieldnames=fields,
        )

        writer.writeheader()

        for row in rows:

            out = {
                field:
                    row.get(field, "")
                for field in fields
            }

            if isinstance(
                out["datetime"],
                datetime,
            ):
                out["datetime"] = (
                    out["datetime"].isoformat()
                )

            writer.writerow(out)


def write_track_csv(
    summaries,
):
    if not summaries:
        return

    fields = list(
        summaries[0].keys()
    )

    with open(
        OUT_TRACKS,
        "w",
        newline="",
    ) as handle:

        writer = csv.DictWriter(
            handle,
            fieldnames=fields,
        )

        writer.writeheader()
        writer.writerows(
            summaries
        )


def write_daily_csv(
    daily,
):
    if not daily:
        return

    fields = list(
        daily[0].keys()
    )

    with open(
        OUT_DAILY,
        "w",
        newline="",
    ) as handle:

        writer = csv.DictWriter(
            handle,
            fieldnames=fields,
        )

        writer.writeheader()

        for row in daily:

            out = dict(row)

            if isinstance(
                out["date"],
                datetime,
            ):
                out["date"] = (
                    out["date"].date().isoformat()
                )
            else:
                out["date"] = str(
                    out["date"]
                )

            writer.writerow(out)


# =====================================================================
# PLOTS
# =====================================================================

def plot_corrected_tracks(
    tide_times,
    tide_values,
    selected_rows,
):
    fig, ax = plt.subplots(
        figsize=(16, 8)
    )

    if len(tide_times):

        start = min(
            r["datetime"]
            for r in selected_rows
        )

        end = max(
            r["datetime"]
            for r in selected_rows
        )

        mask = (
            (tide_times >= np.datetime64(start))
            & (tide_times <= np.datetime64(end))
        )

        ax.plot(
            tide_times[mask],
            tide_values[mask],
            linewidth=2.5,
            label="EOT20 tide model",
        )

    groups = defaultdict(list)

    for row in selected_rows:
        key = (
            row["sat"],
            row["freq"],
            row["rise"],
        )
        groups[key].append(row)

    for key, group in sorted(
        groups.items()
    ):

        group.sort(
            key=lambda r:
                r["datetime"]
        )

        az_mean = float(
            np.mean(
                [
                    r["az"]
                    for r in group
                ]
            )
        )

        label = (
            f"PRN {key[0]} "
            f"{'R' if key[2] == 1 else 'S'} "
            f"{az_mean:.1f}°"
        )

        ax.plot(
            [
                r["datetime"]
                for r in group
            ],
            [
                r["corrected_wl_m"]
                for r in group
            ],
            marker="o",
            markersize=3,
            linewidth=1.2,
            label=label,
        )

    ax.set_title(
        "Marconi Experimental GNSS-R Water Level vs EOT20"
    )

    ax.set_xlabel("UTC")

    ax.set_ylabel(
        "Corrected GNSS-R water level / EOT20 (m)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend(
        ncol=2,
        fontsize=8,
    )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "01_corrected_tracks_vs_eot20.png",
        dpi=200,
    )

    plt.close(fig)


def plot_daily(
    tide_times,
    tide_values,
    daily,
):
    dates = np.asarray(
        [
            np.datetime64(
                row["date"],
                "D",
            )
            for row in daily
        ]
    )

    gnssr = np.asarray(
        [
            row["GNSSR_median_m"]
            for row in daily
        ],
        dtype=float,
    )

    tide = np.asarray(
        [
            row["EOT20_median_at_obs_m"]
            for row in daily
        ],
        dtype=float,
    )

    fig, ax = plt.subplots(
        figsize=(15, 7)
    )

    ax.plot(
        dates,
        gnssr,
        marker="o",
        linewidth=1.8,
        label="Daily multi-track GNSS-R",
    )

    ax.plot(
        dates,
        tide,
        marker="o",
        linewidth=1.8,
        label="EOT20 at GNSS-R observation times",
    )

    ax.set_title(
        "Marconi Experimental Daily Multi-Track GNSS-R vs EOT20"
    )

    ax.set_xlabel("Date")

    ax.set_ylabel(
        "Water-level elevation / anomaly (m)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "02_daily_multitrack_vs_eot20.png",
        dpi=200,
    )

    plt.close(fig)


def plot_track_residuals(
    selected_rows,
):
    fig, ax = plt.subplots(
        figsize=(15, 7)
    )

    groups = defaultdict(list)

    for row in selected_rows:
        key = (
            row["sat"],
            row["rise"],
        )
        groups[key].append(row)

    for key, group in sorted(
        groups.items()
    ):

        group.sort(
            key=lambda r:
                r["datetime"]
        )

        ax.plot(
            [
                r["datetime"]
                for r in group
            ],
            [
                100
                * r[
                    "corrected_residual_m"
                ]
                for r in group
            ],
            marker="o",
            markersize=3,
            linewidth=1.2,
            label=(
                f"PRN {key[0]} "
                f"{'R' if key[1] == 1 else 'S'}"
            ),
        )

    ax.axhline(
        0,
        linewidth=1.5,
    )

    ax.set_title(
        "Selected Track Residuals After +0.242 m Datum Test"
    )

    ax.set_xlabel("UTC")

    ax.set_ylabel(
        "Corrected GNSS-R − EOT20 (cm)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend(
        ncol=2,
    )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "03_track_residuals_vs_eot20.png",
        dpi=200,
    )

    plt.close(fig)


def plot_offsets(
    selected_summaries,
):
    labels = []
    vals = []

    for item in selected_summaries:
        labels.append(
            f"PRN {item['sat']} "
            f"{'R' if item['rise'] == 1 else 'S'} "
            f"{item['az_mean_deg']:.1f}°"
        )
        vals.append(
            item[
                "corrected_median_m"
            ]
            * 100
        )

    y = np.arange(
        len(labels)
    )

    fig, ax = plt.subplots(
        figsize=(10, 5)
    )

    ax.barh(
        y,
        vals,
    )

    ax.axvline(
        0,
        linewidth=1.5,
    )

    ax.set_yticks(y)
    ax.set_yticklabels(labels)

    ax.set_xlabel(
        "Corrected median GNSS-R − EOT20 (cm)"
    )

    ax.set_title(
        "Selected Multi-Track Residual Bias"
    )

    ax.grid(
        True,
        axis="x",
        alpha=0.25,
    )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR
        / "04_selected_track_offsets.png",
        dpi=200,
    )

    plt.close(fig)


# =====================================================================
# SUMMARY
# =====================================================================

def write_summary(
    all_summaries,
    selected_summaries,
    daily,
    selected_rows,
):
    lines = [
        "MARCONI EXPERIMENTAL MULTI-TRACK GNSS-R WATER-LEVEL PRODUCT",
        "=" * 96,
        "",
        "GNSS-R configuration:",
        "  GPS L1 only",
        "  RH 17-23 m",
        "  elevation 5-13 deg",
        "",
        f"H_ortho = {H_ORTHO_M:.3f} m",
        f"datum test = +{DATUM_SHIFT_M:.3f} m",
        "",
        "TRACK SELECTION CRITERIA:",
        f"  N >= {MIN_TRACK_N}",
        f"  |corrected median offset| <= {MAX_ABS_CORRECTED_MEDIAN_M:.3f} m",
        f"  corrected RMS <= {MAX_CORRECTED_RMS_M:.3f} m",
        f"  tide r >= {MIN_TIDE_R}",
        "",
        f"Total repeated tracks considered: {len(all_summaries)}",
        f"Selected tracks: {len(selected_summaries)}",
        f"Selected observations: {len(selected_rows)}",
        f"Daily product days: {len(daily)}",
        "",
        "SELECTED TRACKS",
        "-" * 96,
    ]

    for item in selected_summaries:

        lines.append(
            f"PRN {item['sat']:2d} "
            f"{'RISING' if item['rise']==1 else 'SETTING':7s} "
            f"Az={item['az_mean_deg']:7.2f} "
            f"N={item['n']:2d} "
            f"days={item['n_days']:2d} "
            f"r={item['tide_r']:+.4f} "
            f"slope={item['tide_slope']:+.4f} "
            f"median={item['corrected_median_m']*100:+.2f} cm "
            f"RMS={item['corrected_rms_m']*100:.2f} cm"
        )

    lines += [
        "",
        "OTHER REPEATED TRACKS",
        "-" * 96,
    ]

    rejected = [
        item
        for item in all_summaries
        if not any(
            (
                item["sat"] == selected["sat"]
                and item["freq"] == selected["freq"]
                and item["rise"] == selected["rise"]
                and abs(
                    item["az_mean_deg"]
                    - selected["az_mean_deg"]
                ) <= AZ_TOL_DEG
            )
            for selected in selected_summaries
        )
    ]

    for item in rejected:
        lines.append(
            f"PRN {item['sat']:2d} "
            f"{'RISING' if item['rise']==1 else 'SETTING':7s} "
            f"Az={item['az_mean_deg']:7.2f} "
            f"N={item['n']:2d} "
            f"r={item['tide_r']:+.4f} "
            f"corr_median={item['corrected_median_m']*100:+.1f} cm "
            f"corr_RMS={item['corrected_rms_m']*100:.1f} cm"
        )

    lines += [
        "",
        "SCIENTIFIC STATUS",
        "-" * 96,
        "This is an experimental multi-track GNSS-R product.",
        "The +0.242 m vertical transformation is retained as an external",
        "datum hypothesis test and is not fitted to this dataset.",
        "The selected tracks have strong tidal response and low residual",
        "bias/RMS under the current thresholds.",
        "Independent Fresnel/topobathy water-footprint validation is still",
        "required before declaring the selected tracks confirmed open-water",
        "reflectors or treating this as a final operational product.",
    ]

    OUT_SUMMARY.write_text(
        "\n".join(lines)
        + "\n"
    )


# =====================================================================
# MAIN
# =====================================================================

def main():
    print()
    print("=" * 96)
    print(
        "MARCONI EXPERIMENTAL MULTI-TRACK GNSS-R WATER-LEVEL PRODUCT"
    )
    print("=" * 96)

    print(
        f"Datum test: +{DATUM_SHIFT_M:.3f} m"
    )

    tide_times, tide_values = load_tide()
    rows = load_results()

    add_tide_values(
        rows,
        tide_times,
        tide_values,
    )

    valid_rows = [
        row
        for row in rows
        if math.isfinite(
            row["EOT20_m"]
        )
    ]

    print(
        "Total GPS L1 GNSS-R observations:",
        len(valid_rows),
    )

    track_groups = build_track_groups(
        valid_rows
    )

    all_summaries = [
        summarize_track(group)
        for group in track_groups
    ]

    all_summaries.sort(
        key=lambda item:
            (
                -item["tide_r"],
                item["corrected_rms_m"],
            )
    )

    selected_summaries = [
        item
        for item in all_summaries
        if is_selected(item)
    ]

    print(
        "Repeated tracks:",
        len(all_summaries),
    )

    print(
        "Selected tracks:",
        len(selected_summaries),
    )

    print()
    print("=" * 96)
    print(
        "SELECTED EXPERIMENTAL TRACKS"
    )
    print("=" * 96)

    for item in selected_summaries:

        print(
            f"PRN {item['sat']:2d} "
            f"{'RISING' if item['rise']==1 else 'SETTING':7s} "
            f"Az={item['az_mean_deg']:7.2f} "
            f"N={item['n']:2d} "
            f"days={item['n_days']:2d} "
            f"r={item['tide_r']:+.4f} "
            f"slope={item['tide_slope']:+.4f} "
            f"bias={item['corrected_median_m']*100:+.1f} cm "
            f"RMS={item['corrected_rms_m']*100:.1f} cm"
        )

    selected_keys = {
        (
            item["sat"],
            item["freq"],
            item["rise"],
            round(
                item["az_mean_deg"],
                1,
            ),
        )
        for item in selected_summaries
    }

    selected_rows = []

    for row in valid_rows:

        for item in selected_summaries:

            if (
                row["sat"]
                == item["sat"]
                and row["freq"]
                == item["freq"]
                and row["rise"]
                == item["rise"]
                and circular_az_diff(
                    row["az"],
                    item["az_mean_deg"],
                ) <= AZ_TOL_DEG
            ):
                row["selected_track"] = True
                selected_rows.append(
                    row
                )
                break

    for row in valid_rows:
        if "selected_track" not in row:
            row["selected_track"] = False

    daily = make_daily_product(
        selected_rows
    )

    print()
    print(
        "Selected observations:",
        len(selected_rows),
    )

    print(
        "Daily product days:",
        len(daily),
    )

    write_observation_csv(
        valid_rows
    )

    write_track_csv(
        all_summaries
    )

    write_daily_csv(
        daily
    )

    write_summary(
        all_summaries,
        selected_summaries,
        daily,
        selected_rows,
    )

    PLOT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    if selected_rows:

        plot_corrected_tracks(
            tide_times,
            tide_values,
            selected_rows,
        )

        plot_daily(
            tide_times,
            tide_values,
            daily,
        )

        plot_track_residuals(
            selected_rows
        )

        plot_offsets(
            selected_summaries
        )

    print()
    print("=" * 96)
    print("OUTPUTS")
    print("=" * 96)
    print(
        "Observations:",
        OUT_OBS.resolve(),
    )
    print(
        "Track summary:",
        OUT_TRACKS.resolve(),
    )
    print(
        "Daily product:",
        OUT_DAILY.resolve(),
    )
    print(
        "Summary:",
        OUT_SUMMARY.resolve(),
    )
    print(
        "Plots:",
        PLOT_DIR.resolve(),
    )

    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
