#!/usr/bin/env python3

import re
import math
from pathlib import Path
from datetime import datetime, timedelta

import openpyxl
import numpy as np


STATION = "usgs"
YEAR = 2026
DOY1 = 204
DOY2 = 207

REFERENCE_HEIGHT_M = 18.665

TIDE_FILE = Path("marconi_tides_sherwood.xlsx")

LOG_DIR = Path(
    "products/refl_code/logs/usgs/2026"
)


SUCCESS_PATTERN = re.compile(
    r"SUCCESS\s+"
    r"Azimuth\s+([0-9.]+)\s+"
    r"Sat\s+(\d+)\s+"
    r"RH\s+([-+0-9.]+)\s+m\s+"
    r"PkNoise\s+([-+0-9.]+)\s+"
    r"Amp\s+([-+0-9.]+)\s+"
    r"Fr\s+(\d+)\s+"
    r"UTC\s+([0-9:.]+)\s+"
    r"DT\s+([0-9.]+)"
)


def read_tide_models(path):

    print(f"Reading tide model file: {path}")

    wb = openpyxl.load_workbook(
        path,
        read_only=True,
        data_only=True
    )

    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)

    header = next(rows)

    columns = {}

    for i, name in enumerate(header):

        if name and name.endswith("_heightm"):

            model = name[:-len("_heightm")]

            columns[model] = i

    times = []
    values = {
        name: []
        for name in columns
    }

    for row in rows:

        if row[0] is None:
            continue

        times.append(row[0])

        for name, index in columns.items():

            values[name].append(
                float(row[index])
            )

    wb.close()

    print(
        f"Loaded {len(times)} tide points."
    )

    print(
        "Models:",
        list(values.keys())
    )

    return times, values


def make_interpolators(times, values):

    epoch = np.array(
        [
            (t - times[0]).total_seconds()
            for t in times
        ],
        dtype=float
    )

    interpolators = {}

    for name, series in values.items():

        y = np.array(
            series,
            dtype=float
        )

        def interp(
            query,
            epoch=epoch,
            y=y
        ):

            seconds = (
                query - times[0]
            ).total_seconds()

            return float(
                np.interp(
                    seconds,
                    epoch,
                    y
                )
            )

        interpolators[name] = interp

    return interpolators


def parse_utc_string(utc):

    """
    Accept all UTC formats appearing in the GNSS-IR logs:

        HH:MM
        HH:MM:SS
        HH:MM:SS.sss

    Returns:
        hours, minutes, seconds
    """

    parts = utc.split(":")

    if len(parts) == 2:

        hours = int(parts[0])
        minutes = int(parts[1])
        seconds = 0.0

    elif len(parts) == 3:

        hours = int(parts[0])
        minutes = int(parts[1])
        seconds = float(parts[2])

    else:

        raise ValueError(
            f"Unrecognized UTC time: {utc}"
        )

    return hours, minutes, seconds


def utc_datetime(year, doy, utc):

    start = (
        datetime(year, 1, 1)
        + timedelta(days=doy - 1)
    )

    hours, minutes, seconds = (
        parse_utc_string(utc)
    )

    return (
        start
        + timedelta(
            hours=hours,
            minutes=minutes,
            seconds=seconds
        )
    )


def parse_success_records():

    records = []

    for doy in range(DOY1, DOY2 + 1):

        path = (
            LOG_DIR /
            f"{doy}_gnssir.txt"
        )

        print()
        print(
            f"Reading {path}"
        )

        count = 0

        with open(path) as f:

            for line in f:

                match = SUCCESS_PATTERN.search(
                    line
                )

                if not match:
                    continue

                (
                    az,
                    sat,
                    rh,
                    pkn,
                    amp,
                    freq,
                    utc,
                    dt
                ) = match.groups()

                records.append(
                    {
                        "doy": doy,
                        "azimuth": float(az),
                        "sat": int(sat),
                        "RH": float(rh),
                        "PkNoise": float(pkn),
                        "Amplitude": float(amp),
                        "freq": int(freq),
                        "UTC": utc,
                        "delT": float(dt),
                    }
                )

                count += 1

        print(
            f"Production SUCCESS: {count}"
        )

    return records


def main():

    print()
    print("=" * 80)
    print("PRODUCTION GNSS-IR SUCCESS VS TIDE MODEL")
    print("=" * 80)

    print(
        f"Station             : {STATION}"
    )

    print(
        f"DOY                 : "
        f"{DOY1}-{DOY2}"
    )

    print(
        f"Reference height    : "
        f"{REFERENCE_HEIGHT_M:.3f} m"
    )

    print(
        "Time basis          : UTC"
    )

    print()

    times, tide_values = (
        read_tide_models(
            TIDE_FILE
        )
    )

    interpolators = (
        make_interpolators(
            times,
            tide_values
        )
    )

    records = (
        parse_success_records()
    )

    print()
    print("=" * 80)
    print("TOTAL PRODUCTION SUCCESS")
    print("=" * 80)

    print(
        f"SUCCESS records: "
        f"{len(records)}"
    )

    if not records:

        print(
            "ERROR: no SUCCESS records found."
        )

        return

    for record in records:

        dt = utc_datetime(
            YEAR,
            record["doy"],
            record["UTC"]
        )

        record["datetime"] = dt

        record[
            "GNSS_water_level"
        ] = (
            REFERENCE_HEIGHT_M
            - record["RH"]
        )

        for name, interp in (
            interpolators.items()
        ):

            record[
                f"tide_{name}"
            ] = interp(dt)

    model_names = list(
        interpolators.keys()
    )

    print()
    print("=" * 80)
    print("MODEL COMPARISON")
    print("=" * 80)

    gnss = np.array(
        [
            r["GNSS_water_level"]
            for r in records
        ]
    )

    for model in model_names:

        tide = np.array(
            [
                r[f"tide_{model}"]
                for r in records
            ]
        )

        residual = (
            gnss - tide
        )

        correlation = (
            np.corrcoef(
                gnss,
                tide
            )[0, 1]
        )

        rms = math.sqrt(
            np.mean(
                residual ** 2
            )
        )

        bias = np.mean(
            residual
        )

        centered = (
            residual - bias
        )

        centered_rms = math.sqrt(
            np.mean(
                centered ** 2
            )
        )

        print()
        print(
            f"{model}"
        )

        print(
            f"  Correlation      : "
            f"{correlation:+.4f}"
        )

        print(
            f"  Bias             : "
            f"{bias:+.4f} m"
        )

        print(
            f"  RMS              : "
            f"{rms:.4f} m"
        )

        print(
            f"  Bias-removed RMS : "
            f"{centered_rms:.4f} m"
        )

    print()
    print("=" * 80)
    print("FIRST 20 PRODUCTION SUCCESS SOLUTIONS")
    print("=" * 80)

    for r in records[:20]:

        tide_values_here = [
            r[f"tide_{m}"]
            for m in model_names
        ]

        tide_mean = np.mean(
            tide_values_here
        )

        residual = (
            r["GNSS_water_level"]
            - tide_mean
        )

        print(
            f"{r['datetime']} "
            f"sat={r['sat']:3d} "
            f"freq={r['freq']:3d} "
            f"RH={r['RH']:7.3f} "
            f"GNSS={r['GNSS_water_level']:8.3f} "
            f"TIDE={tide_mean:8.3f} "
            f"RES={residual*100:+8.2f} cm"
        )

    print()
    print("=" * 80)
    print("DONE")
    print("=" * 80)


if __name__ == "__main__":
    main()
