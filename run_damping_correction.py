#!/usr/bin/env python3
"""
run_damping_correction.py

Runs the damping-aware correction pipeline (damping_correction.py)
against real arc data for a given station/day range, with a real
cross-validation step against gnssrefl's own already-computed EdotF
values before trusting anything further.

Usage:
    python3 run_damping_correction.py <station> <year> <doy1> [doy2] <tide_models_xlsx>

Example:
    python3 run_damping_correction.py usgs 2026 204 207 marconi_tides_sherwood.xlsx

Requires the gnssrefl virtual environment active and the usual
REFL_CODE/EXE/ORBITS environment variables set.
"""

from __future__ import annotations

import sys
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl

import damping_correction as dc

MJD_EPOCH = datetime(1858, 11, 17)

# Our real, full frequency list (matches station.json's own
# gnssrefl_all_frequencies=True config) -- each arc's own real
# wavelength is now taken directly from gnssrefl's own already-
# computed cf field (2*cf), not a hardcoded per-frequency table, so
# this list only needs to match which frequencies to request from
# extract_arcs, not their physical properties.
REAL_FREQUENCIES = [1, 5, 20, 101, 102, 201, 205, 206, 207, 208, 301, 302, 305, 306, 307, 308]


def read_tide_models(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    model_columns = {}
    for i, name in enumerate(header):
        if name and name.endswith("_heightm"):
            model_columns[name[: -len("_heightm")]] = i
    times = []
    model_values = {name: [] for name in model_columns}
    for row in rows[1:]:
        times.append(row[0])
        for model_name, col_index in model_columns.items():
            model_values[model_name].append(row[col_index])
    return times, model_values


def build_ensemble_interpolator(times, model_values):
    ensemble = np.mean(np.array([model_values[name] for name in model_values]), axis=0)
    epoch_seconds = np.array([(t - times[0]).total_seconds() for t in times], dtype=float)

    def interpolate(query_dt: datetime) -> float:
        query_seconds = (query_dt - times[0]).total_seconds()
        return float(np.interp(query_seconds, epoch_seconds, ensemble))

    return interpolate


def main():
    if len(sys.argv) not in (5, 6):
        print(f"Usage: {sys.argv[0]} <station> <year> <doy1> [doy2] <tide_models_xlsx>")
        sys.exit(1)

    station = sys.argv[1]
    year = int(sys.argv[2])
    doy1 = int(sys.argv[3])

    if len(sys.argv) == 6:
        doy2 = int(sys.argv[4])
        xlsx_path = Path(sys.argv[5])
    else:
        doy2 = doy1
        xlsx_path = Path(sys.argv[4])

    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} does not exist.")
        sys.exit(1)

    try:
        from gnssrefl.extract_arcs import extract_arcs_from_station
    except ImportError:
        print("ERROR: could not import gnssrefl.extract_arcs -- make sure the")
        print("gnssrefl virtual environment is activated.")
        sys.exit(1)

    # --- Gather real arcs across every real GNSS frequency ---
    day_start_ref = datetime(year, 1, 1)
    arc_times_hours = []
    arc_rh = []
    arc_elevation = []
    arc_dsnr = []
    arc_edot = []
    arc_wavelength_m = []

    # Cross-validation accumulators
    edotf_mine = []
    edotf_gnssrefl_meta = []
    edotf_gnssrefl_result = []

    for doy in range(doy1, doy2 + 1):
        print(f"Extracting all-frequency arcs for {station} {year} doy {doy}...")
        arcs = extract_arcs_from_station(
            station, year, doy,
            freq=REAL_FREQUENCIES,
            e1=5.0, e2=15.0,
            azlist=[100.0, 130.0, 150.0, 215.0],
            polyV=2, pele=[5, 30],
            attach_results=True, buffer_hours=2,
        )
        # Confirmed necessary against real data: "gnssir_processing_results
        # is not None" only confirms SOME result got attached, not that
        # it's a genuine, QC-passing match -- real data showed 65-72% of
        # "matched" arcs actually had a NaN RH value inside an otherwise
        # non-None results dict. Requiring a real, finite RH is the
        # correct, stricter filter. Confirmed separately (see chat):
        # these are genuinely the same arcs our real pipeline's own
        # peak2noise QC already rejects -- not a bug, not overly strict.
        matched = [
            (m, d) for m, d in arcs
            if m["gnssir_processing_results"] is not None
            and np.isfinite(m["gnssir_processing_results"]["RH"])
        ]
        print(f"  {len(arcs)} arcs extracted, {len(matched)} passed QC")

        day_start = day_start_ref + timedelta(days=doy - 1)

        for meta, data in matched:
            result = meta["gnssir_processing_results"]

            # cf is confirmed (directly, against a known GPS L1 case)
            # to equal wavelength/2 for every frequency, including
            # correctly handling GLONASS's per-satellite FDMA
            # wavelength variation automatically, since gnssrefl
            # already accounts for it internally.
            wavelength_m = 2 * meta["cf"]

            # Cross-validate: my classical_edotf() vs gnssrefl's own
            # two independent reportings of essentially the same
            # quantity (meta's own edot_factor, and the result's
            # EdotF from the actual gnssir run).
            my_edotf = dc.classical_edotf(data["ele"], data["edot"])
            edotf_mine.append(my_edotf)
            edotf_gnssrefl_meta.append(meta["edot_factor"])
            edotf_gnssrefl_result.append(result["EdotF"])

            # Real, absolute time for this arc: MJD -> hours since the
            # first day in our range, for use in the window solver.
            arc_mjd = result["MJD"]
            arc_dt = MJD_EPOCH + timedelta(days=arc_mjd)
            hours_since_start = (arc_dt - day_start_ref).total_seconds() / 3600

            arc_times_hours.append(hours_since_start)
            arc_rh.append(result["RH"])
            arc_elevation.append(data["ele"])
            arc_dsnr.append(data["snr"])
            arc_edot.append(data["edot"])
            arc_wavelength_m.append(wavelength_m)

    if len(arc_times_hours) < 10:
        print(f"\nOnly {len(arc_times_hours)} matched arcs found -- too few to proceed.")
        sys.exit(1)

    arc_times_hours = np.array(arc_times_hours)
    arc_rh = np.array(arc_rh)
    arc_wavelength_m = np.array(arc_wavelength_m)

    print(f"\n{len(arc_times_hours)} total arcs gathered across days {doy1}-{doy2} "
          f"(all frequencies combined).")

    # --- Real cross-validation against gnssrefl's own EdotF values ---
    edotf_mine = np.array(edotf_mine)
    edotf_gnssrefl_meta = np.array(edotf_gnssrefl_meta)
    edotf_gnssrefl_result = np.array(edotf_gnssrefl_result)

    diff_vs_meta = np.abs(edotf_mine - edotf_gnssrefl_meta)
    diff_vs_result = np.abs(edotf_mine - edotf_gnssrefl_result)

    print("\n--- Cross-validation: my classical_edotf() vs gnssrefl's own values ---")
    print(f"  vs meta['edot_factor']:          mean diff = {diff_vs_meta.mean():.6f}, "
          f"max diff = {diff_vs_meta.max():.6f}")
    print(f"  vs gnssir_processing_results['EdotF']: mean diff = {diff_vs_result.mean():.6f}, "
          f"max diff = {diff_vs_result.max():.6f}")
    print("  (Small values here confirm our own edot-based calculation matches")
    print("  gnssrefl's own real, already-validated computation -- if these are")
    print("  large, something about our unit/definition assumptions is wrong")
    print("  and the damping-aware results below should not yet be trusted.)")

    # --- Run the full damping-aware pipeline ---
    window_length_hours = 1.0
    window_shift_hours = 1 / 6
    print("\nRunning the damping-aware correction pipeline...")
    results = dc.run_damping_aware_correction(
        arc_times_hours, arc_rh, arc_elevation, arc_dsnr, arc_edot, arc_wavelength_m,
        window_length_hours=window_length_hours, window_shift_hours=window_shift_hours,
    )

    if not results:
        print("No windows could be solved (need at least 2 arcs per window).")
        sys.exit(1)

    print(f"\n{len(results)} windows solved.")
    lambdas = np.array([r["damping_lambda"] for r in results])
    print(f"Fitted damping lambda across windows: mean={lambdas.mean():.4f}, "
          f"min={lambdas.min():.4f}, max={lambdas.max():.4f}")

    # --- Compare against the real tide models ---
    print(f"\nReading tide models from {xlsx_path}...")
    tide_times, tide_values = read_tide_models(xlsx_path)
    tide_at = build_ensemble_interpolator(tide_times, tide_values)

    window_dts = [day_start_ref + timedelta(hours=r["window_center_hours"]) for r in results]
    tide_at_windows = np.array([tide_at(dt) for dt in window_dts])
    h_w_values = np.array([r["h_w"] for r in results])

    # Fit scale+offset to compare shapes on a common footing (same
    # approach used throughout tonight's other comparison scripts).
    A = np.vstack([tide_at_windows, np.ones_like(tide_at_windows)]).T
    a, b = np.linalg.lstsq(A, h_w_values, rcond=None)[0]
    calibrated_tide = a * tide_at_windows + b
    rms_vs_tide = float(np.sqrt(np.mean((h_w_values - calibrated_tide) ** 2)))
    print(f"\nRMS of damping-aware corrected h_w vs. calibrated tide model shape: "
          f"{rms_vs_tide:.4f} m")

    # --- Plot ---
    # Confirmed necessary, same real bug as compare_to_tide_models.py's
    # earlier fix: most candidate windows don't have the minimum 2 arcs
    # needed to solve (with e.g. only 38 total arcs spread across 4
    # days), leaving real gaps between clusters of solved windows.
    # Without this, matplotlib connects across those gaps with a
    # straight line for BOTH curves (since they share the same x-axis
    # window centers) -- producing a misleading sawtooth appearance
    # rather than genuine blank gaps.
    window_hours_arr = np.array([r["window_center_hours"] for r in results])
    gap_threshold_hours = window_shift_hours * 3  # generous margin over the expected ~10min spacing

    h_w_plot = h_w_values.astype(float).copy()
    tide_plot = tide_at_windows.astype(float).copy()
    window_dts_plot = list(window_dts)

    insert_at = []
    for i in range(1, len(window_hours_arr)):
        if window_hours_arr[i] - window_hours_arr[i - 1] > gap_threshold_hours:
            insert_at.append(i)

    if insert_at:
        print(f"\nFound {len(insert_at)} real gap(s) between solved windows -- "
              "inserting breaks so the plot doesn't draw a misleading "
              "straight line across missing data.")
        for offset, idx in enumerate(insert_at):
            insert_pos = idx + offset
            mid_dt = window_dts_plot[insert_pos - 1] + (window_dts_plot[insert_pos] - window_dts_plot[insert_pos - 1]) / 2
            window_dts_plot.insert(insert_pos, mid_dt)
            h_w_plot = np.insert(h_w_plot, insert_pos, np.nan)
            tide_plot = np.insert(tide_plot, insert_pos, np.nan)

    fig, ax1 = plt.subplots(figsize=(12, 5))
    ax1.plot(window_dts_plot, h_w_plot, "o-", color="tab:blue", ms=4, label="Damping-aware corrected (h_w)")
    ax1.set_ylabel("Reflector height (m)", color="tab:blue")
    ax1.invert_yaxis()
    ax1.tick_params(axis="y", labelcolor="tab:blue")

    ax2 = ax1.twinx()
    ax2.plot(window_dts_plot, tide_plot, "--", color="tab:red", alpha=0.7, label="Tide model (ensemble mean)")
    ax2.set_ylabel("Tide model predicted height (m)", color="tab:red")
    ax2.tick_params(axis="y", labelcolor="tab:red")

    fig.autofmt_xdate()
    ax1.set_title(f"{station} {year} doy {doy1}-{doy2}: damping-aware corrected water level "
                  f"vs. real tide models\n(RMS = {rms_vs_tide:.4f} m)")
    fig.tight_layout()
    output_path = Path("damping_correction_result.png")
    fig.savefig(output_path, dpi=150)
    print(f"\nSaved comparison plot to: {output_path}")


if __name__ == "__main__":
    main()
