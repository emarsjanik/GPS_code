#!/usr/bin/env python3
"""
Marconi long-term GNSS-R track analysis v2.

Fixes the first long-term script and, importantly, separates true
reflection tracks instead of pooling every observation of a satellite
and frequency together.

Track identity is now:
    satellite + frequency + rise/setting + azimuth cluster

Azimuth clustering is done with a configurable tolerance (default 3 deg).
This prevents a satellite that appears repeatedly in different reflection
directions from being incorrectly treated as one long-term physical track.

Primary tide model:
    EOT20_heightm

Water level:
    WL_GNSSR = H_ortho - RH

Outputs:
    marconi_longterm_track_stability_v2.csv
    marconi_longterm_track_stability_v2_summary.txt
    marconi_longterm_plots_v2/
      all_gnssr_vs_tide_timeseries.png
      all_gnssr_vs_tide_scatter.png
      top_tracks_vs_tide.png
      top_track_residuals.png
      track_calibration_stability.png
      daily_population_vs_tide.png
      track_count_by_day.png
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)
TIDE_FILE = Path("marconi_tides_sherwood.xlsx")

OUT_CSV = Path("marconi_longterm_track_stability_v2.csv")
OUT_SUMMARY = Path("marconi_longterm_track_stability_v2_summary.txt")
PLOT_DIR = Path("marconi_longterm_plots_v2")

H_ORTHO_M = 18.665
PRIMARY_TIDE_MODEL = "EOT20_heightm"
MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]

MIN_TRACK_N = 4
AZ_CLUSTER_TOL_DEG = 3.0
TOP_TRACKS_FOR_PLOTS = 10


def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if len(x) < 3 or np.std(x) == 0 or np.std(y) == 0:
        return math.nan
    return float(np.corrcoef(x, y)[0, 1])


def rms(x):
    x = np.asarray(x, float)
    return float(np.sqrt(np.mean(x ** 2))) if len(x) else math.nan


def mae(x):
    x = np.asarray(x, float)
    return float(np.mean(np.abs(x))) if len(x) else math.nan


def utc_hours_to_datetime(year, doy, utc_hours):
    return (
        datetime(year, 1, 1)
        + timedelta(days=doy - 1, hours=float(utc_hours))
    )


def load_tide_data():
    wb = load_workbook(TIDE_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    time_col = header.index("time")
    model_cols = {m: header.index(m) for m in MODELS}

    times = []
    values = {m: [] for m in MODELS}
    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[time_col]
        if not isinstance(t, datetime):
            continue
        tmp = {}
        good = True
        for m in MODELS:
            v = finite(row[model_cols[m]])
            if v is None:
                good = False
                break
            tmp[m] = v
        if not good:
            continue
        times.append(t)
        for m in MODELS:
            values[m].append(tmp[m])
    wb.close()

    epoch = np.asarray(
        [(t - times[0]).total_seconds() for t in times],
        float,
    )
    arrays = {m: np.asarray(values[m], float) for m in MODELS}

    def tide_at(dt, model):
        x = (dt - times[0]).total_seconds()
        if x < epoch[0] or x > epoch[-1]:
            return math.nan
        return float(np.interp(x, epoch, arrays[model]))

    return times, tide_at


def load_results():
    rows = []
    files = sorted(RESULT_DIR.glob("*.txt"))
    if not files:
        raise SystemExit(f"No result files found in {RESULT_DIR}")

    for path in files:
        try:
            int(path.stem)
        except Exception:
            continue

        for line in path.read_text(errors="replace").splitlines():
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            c = line.split()
            if len(c) < 17:
                continue
            try:
                year = int(float(c[0]))
                doy = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc_hours = float(c[4])
                az = float(c[5])
                amp = float(c[6])
                emin = float(c[7])
                emax = float(c[8])
                nobs = int(float(c[9]))
                freq = int(float(c[10]))
                rise = int(float(c[11]))
                pkn = float(c[13])
                delt = float(c[14])
            except Exception:
                continue

            if freq != 1:
                continue

            dt = utc_hours_to_datetime(year, doy, utc_hours)
            rows.append({
                "year": year,
                "doy": doy,
                "datetime_utc": dt,
                "sat": sat,
                "freq": freq,
                "RH_m": rh,
                "GNSS_WL_m": H_ORTHO_M - rh,
                "az_deg": az,
                "Amp": amp,
                "PkNoise": pkn,
                "eminO_deg": emin,
                "emaxO_deg": emax,
                "NumbOf": nobs,
                "rise": rise,
                "DelT_min": delt,
                "source_file": str(path),
            })

    rows.sort(key=lambda r: r["datetime_utc"])
    return rows


def add_tides(rows, tide_at):
    for row in rows:
        for m in MODELS:
            row[m] = tide_at(row["datetime_utc"], m)
        vals = [row[m] for m in MODELS if math.isfinite(row[m])]
        row["TIDE_ENSEMBLE_m"] = float(np.mean(vals)) if vals else math.nan
        row["PRIMARY_TIDE_m"] = row[PRIMARY_TIDE_MODEL]


def circular_az_diff(a, b):
    d = abs(a - b) % 360.0
    return min(d, 360.0 - d)


def cluster_track_rows(rows):
    """
    Cluster within sat/freq/rise by azimuth. Clusters are formed from
    sorted azimuths with gaps <= AZ_CLUSTER_TOL_DEG.
    """
    base = defaultdict(list)
    for r in rows:
        base[(r["sat"], r["freq"], r["rise"])].append(r)

    tracks = []
    for key, group in base.items():
        group = sorted(group, key=lambda r: r["az_deg"])
        current = []
        prev_az = None

        for r in group:
            if prev_az is None or circular_az_diff(r["az_deg"], prev_az) <= AZ_CLUSTER_TOL_DEG:
                current.append(r)
            else:
                tracks.append(current)
                current = [r]
            prev_az = r["az_deg"]

        if current:
            tracks.append(current)

    tracks = [g for g in tracks if len(g) >= MIN_TRACK_N]
    for i, g in enumerate(tracks, 1):
        pass
    return tracks


def analyze_track(group, track_id):
    wl = np.asarray([r["GNSS_WL_m"] for r in group], float)
    tide = np.asarray([r["PRIMARY_TIDE_m"] for r in group], float)
    az = np.asarray([r["az_deg"] for r in group], float)
    pkn = np.asarray([r["PkNoise"] for r in group], float)
    amp = np.asarray([r["Amp"] for r in group], float)

    valid = np.isfinite(wl) & np.isfinite(tide)
    w = wl[valid]
    t = tide[valid]

    if len(w) >= 3:
        r = pearson(w, t)
        coeff = np.polyfit(t, w, 1)
        slope = float(coeff[0])
        intercept = float(coeff[1])
        C = float(np.mean(w - t))
        resid = w - C - t
        unit_rms_cm = rms(resid) * 100
        free_rms_cm = rms(w - (slope * t + intercept)) * 100
        unit_mae_cm = mae(resid) * 100
    else:
        r = slope = intercept = C = math.nan
        unit_rms_cm = free_rms_cm = unit_mae_cm = math.nan

    days = sorted({r0["doy"] for r0 in group})
    duration = (
        group[-1]["datetime_utc"] - group[0]["datetime_utc"]
    ).total_seconds() / 86400.0

    # Duration is supporting evidence, not the dominant score.
    corr_score = abs(r) if math.isfinite(r) else 0.0
    slope_score = (
        max(0.0, 1.0 - min(1.0, abs(slope - 1.0)))
        if math.isfinite(slope) else 0.0
    )
    rms_score = (
        max(0.0, 1.0 - min(1.0, unit_rms_cm / 25.0))
        if math.isfinite(unit_rms_cm) else 0.0
    )
    pkn_score = max(0.0, min(1.0, (float(np.mean(pkn)) - 2.8) / 1.2))
    amp_score = max(0.0, min(1.0, float(np.mean(amp)) / 50.0))
    az_std = float(np.std(az))
    az_score = max(0.0, min(1.0, 1.0 - az_std / 2.0))
    persistence = max(0.0, min(1.0, len(days) / 20.0))

    score = (
        0.30 * corr_score
        + 0.20 * slope_score
        + 0.20 * rms_score
        + 0.10 * pkn_score
        + 0.08 * amp_score
        + 0.07 * az_score
        + 0.05 * persistence
    )

    return {
        "track_id": track_id,
        "sat": group[0]["sat"],
        "freq": group[0]["freq"],
        "rise": group[0]["rise"],
        "n": len(group),
        "n_days": len(days),
        "doy_first": min(days),
        "doy_last": max(days),
        "duration_days": duration,
        "az_mean_deg": float(np.mean(az)),
        "az_std_deg": az_std,
        "RH_mean_m": float(np.mean([r0["RH_m"] for r0 in group])),
        "RH_sd_m": float(np.std([r0["RH_m"] for r0 in group])),
        "PkNoise_mean": float(np.mean(pkn)),
        "Amp_mean": float(np.mean(amp)),
        "tide_r": r,
        "tide_slope": slope,
        "tide_intercept_m": intercept,
        "C_unit_slope_m": C,
        "unit_slope_RMS_cm": unit_rms_cm,
        "unit_slope_MAE_cm": unit_mae_cm,
        "free_fit_RMS_cm": free_rms_cm,
        "longterm_score": score,
    }


def write_track_csv(records):
    fields = list(records[0].keys())
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(records)


def get_track_group(rows, rec):
    return [
        r for r in rows
        if (
            r["sat"] == rec["sat"]
            and r["freq"] == rec["freq"]
            and r["rise"] == rec["rise"]
            and circular_az_diff(
                r["az_deg"],
                rec["az_mean_deg"],
            ) <= AZ_CLUSTER_TOL_DEG
        )
    ]


def plot_all_gnssr_vs_tide_timeseries(rows):
    PLOT_DIR.mkdir(parents=True, exist_ok=True)

    rows = sorted(rows, key=lambda r: r["datetime_utc"])
    times = [r["datetime_utc"] for r in rows]
    gnss = np.asarray([r["GNSS_WL_m"] for r in rows], float)
    tide = np.asarray([r["PRIMARY_TIDE_m"] for r in rows], float)

    fig, ax = plt.subplots(figsize=(15, 7))
    ax.plot(times, tide, linewidth=2, label=PRIMARY_TIDE_MODEL)
    ax.scatter(times, gnss, s=11, alpha=0.40, label="All GPS L1 GNSS-R arcs")
    ax.set_title("Marconi: All GPS L1 GNSS-R Water Level vs EOT20")
    ax.set_xlabel("UTC")
    ax.set_ylabel("Water-level anomaly (m)")
    ax.grid(True, alpha=0.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "all_gnssr_vs_tide_timeseries.png", dpi=180)
    plt.close(fig)


def plot_all_gnssr_vs_tide_scatter(rows):
    x = np.asarray([r["PRIMARY_TIDE_m"] for r in rows], float)
    y = np.asarray([r["GNSS_WL_m"] for r in rows], float)
    valid = np.isfinite(x) & np.isfinite(y)
    x, y = x[valid], y[valid]
    if len(x) < 3:
        return

    coeff = np.polyfit(x, y, 1)
    r = pearson(x, y)
    xx = np.linspace(x.min(), x.max(), 200)
    yy = coeff[0] * xx + coeff[1]

    fig, ax = plt.subplots(figsize=(8, 8))
    ax.scatter(x, y, s=15, alpha=0.45)
    ax.plot(xx, yy, linewidth=2, label=f"fit slope={coeff[0]:.3f}, r={r:.3f}")
    ax.plot(xx, xx, linestyle="--", linewidth=1.5, label="1:1")
    ax.set_title("All GPS L1 GNSS-R Water Level vs EOT20")
    ax.set_xlabel("EOT20 tide (m)")
    ax.set_ylabel("GNSS-R water level (m)")
    ax.grid(True, alpha=0.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "all_gnssr_vs_tide_scatter.png", dpi=180)
    plt.close(fig)


def plot_top_tracks_vs_tide(rows, records):
    PLOT_DIR.mkdir(parents=True, exist_ok=True)

    selected = sorted(
        records,
        key=lambda r: r["longterm_score"],
        reverse=True,
    )[:TOP_TRACKS_FOR_PLOTS]

    fig, ax = plt.subplots(figsize=(15, 8))

    all_rows = sorted(rows, key=lambda r: r["datetime_utc"])
    tide_times = [r["datetime_utc"] for r in all_rows]
    tide = np.asarray([r["PRIMARY_TIDE_m"] for r in all_rows], float)
    valid = np.isfinite(tide)
    ax.plot(
        [tide_times[i] for i in np.where(valid)[0]],
        tide[valid],
        linewidth=2.5,
        label=PRIMARY_TIDE_MODEL,
        zorder=1,
    )

    colors = plt.cm.tab10(np.linspace(0, 1, len(selected)))

    for color, rec in zip(colors, selected):
        group = sorted(
            get_track_group(rows, rec),
            key=lambda r: r["datetime_utc"],
        )
        ax.plot(
            [r["datetime_utc"] for r in group],
            [r["GNSS_WL_m"] for r in group],
            marker="o",
            markersize=3.5,
            linewidth=1.2,
            color=color,
            label=(
                f"PRN {rec['sat']} "
                f"Az {rec['az_mean_deg']:.1f} "
                f"N={rec['n']} "
                f"r={rec['tide_r']:.3f}"
            ),
        )

    ax.set_title("Top Long-Term GNSS-R Tracks vs EOT20 Tide")
    ax.set_xlabel("UTC")
    ax.set_ylabel("Water-level anomaly (m)")
    ax.grid(True, alpha=0.25)
    ax.legend(ncol=2, fontsize=8)
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "top_tracks_vs_tide.png", dpi=180)
    plt.close(fig)


def plot_top_track_residuals(rows, records):
    selected = sorted(
        records,
        key=lambda r: r["longterm_score"],
        reverse=True,
    )[:TOP_TRACKS_FOR_PLOTS]

    fig, ax = plt.subplots(figsize=(15, 8))
    colors = plt.cm.tab10(np.linspace(0, 1, len(selected)))

    for color, rec in zip(colors, selected):
        group = sorted(
            get_track_group(rows, rec),
            key=lambda r: r["datetime_utc"],
        )
        c = rec["C_unit_slope_m"]
        residual_cm = np.asarray(
            [
                (r["GNSS_WL_m"] - c - r["PRIMARY_TIDE_m"]) * 100.0
                for r in group
            ],
            float,
        )
        ax.plot(
            [r["datetime_utc"] for r in group],
            residual_cm,
            marker="o",
            markersize=3,
            linewidth=1,
            color=color,
            label=f"PRN {rec['sat']} Az {rec['az_mean_deg']:.1f}",
        )

    ax.axhline(0, linewidth=1.5)
    ax.set_title("Top Long-Term Track Unit-Slope Residuals")
    ax.set_xlabel("UTC")
    ax.set_ylabel("GNSS-R - EOT20 after track C (cm)")
    ax.grid(True, alpha=0.25)
    ax.legend(ncol=2, fontsize=8)
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "top_track_residuals.png", dpi=180)
    plt.close(fig)


def plot_calibration_stability(rows, records):
    selected = sorted(
        records,
        key=lambda r: r["longterm_score"],
        reverse=True,
    )[:TOP_TRACKS_FOR_PLOTS]

    fig, ax = plt.subplots(figsize=(15, 8))
    colors = plt.cm.tab10(np.linspace(0, 1, len(selected)))

    for color, rec in zip(colors, selected):
        group = sorted(
            get_track_group(rows, rec),
            key=lambda r: r["datetime_utc"],
        )
        offsets_cm = [
            (r["GNSS_WL_m"] - r["PRIMARY_TIDE_m"]) * 100.0
            for r in group
        ]
        ax.plot(
            [r["datetime_utc"] for r in group],
            offsets_cm,
            marker="o",
            markersize=3,
            linewidth=1,
            color=color,
            label=f"PRN {rec['sat']} Az {rec['az_mean_deg']:.1f}",
        )

    ax.axhline(0, linewidth=1.5)
    ax.set_title("GNSS-R Minus EOT20: Calibration Offset Stability")
    ax.set_xlabel("UTC")
    ax.set_ylabel("GNSS-R water level - EOT20 (cm)")
    ax.grid(True, alpha=0.25)
    ax.legend(ncol=2, fontsize=8)
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "track_calibration_stability.png", dpi=180)
    plt.close(fig)


def plot_daily_population(rows):
    by_day = defaultdict(list)
    for r in rows:
        by_day[r["datetime_utc"].date()].append(r)

    days = sorted(by_day)
    gnss_med = []
    tide_med = []
    n = []

    for day in days:
        g = by_day[day]
        gnss_med.append(float(np.median([r["GNSS_WL_m"] for r in g])))
        tide_med.append(float(np.median([r["PRIMARY_TIDE_m"] for r in g])))
        n.append(len(g))

    fig, ax = plt.subplots(figsize=(15, 7))
    ax.plot(days, tide_med, linewidth=2, label=PRIMARY_TIDE_MODEL)
    ax.plot(days, gnss_med, marker="o", linewidth=1.2, label="Daily GNSS-R median")
    ax.set_title("Daily GNSS-R Population Median vs EOT20")
    ax.set_xlabel("Date")
    ax.set_ylabel("Water-level anomaly (m)")
    ax.grid(True, alpha=0.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "daily_population_vs_tide.png", dpi=180)
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(15, 4))
    ax.bar(days, n)
    ax.set_title("GNSS-R Successful Arcs per Day")
    ax.set_xlabel("Date")
    ax.set_ylabel("Number of L1 solutions")
    ax.grid(True, axis="y", alpha=0.25)
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "track_count_by_day.png", dpi=180)
    plt.close(fig)


def write_summary(records, rows):
    lines = [
        "MARCONI LONG-TERM GNSS-R STABILITY V2",
        "=" * 100,
        f"Primary tide model: {PRIMARY_TIDE_MODEL}",
        f"Station H_ortho: {H_ORTHO_M:.3f} m",
        f"Azimuth cluster tolerance: {AZ_CLUSTER_TOL_DEG:.1f} deg",
        f"Minimum track observations: {MIN_TRACK_N}",
        f"Total GPS L1 arcs: {len(rows)}",
        "",
        "TRACK RANKING",
        "-" * 100,
    ]

    for i, r in enumerate(records[:30], 1):
        lines.append(
            f"{i:2d}. SAT={r['sat']:3d} FREQ={r['freq']:1d} "
            f"rise={r['rise']:2d} N={r['n']:3d} days={r['n_days']:2d} "
            f"Az={r['az_mean_deg']:6.2f}±{r['az_std_deg']:.2f} "
            f"r={r['tide_r']:+.4f} slope={r['tide_slope']:+.4f} "
            f"C={r['C_unit_slope_m']:+.3f} "
            f"RMS={r['unit_slope_RMS_cm']:.2f}cm "
            f"score={r['longterm_score']:.3f}"
        )

    lines += [
        "",
        "IMPORTANT INTERPRETATION",
        "The previous analysis pooled all observations from a satellite/frequency.",
        "This version separates tracks by rise/setting and azimuth cluster.",
        "This is essential because one satellite can produce multiple physical",
        "reflection geometries during the long observation period.",
        "",
        "The first long-term result showing PRN 26 dropping from ~4 cm RMS over",
        "the four-day test to ~26 cm RMS over the full interval is therefore NOT",
        "automatically evidence that PRN 26 stopped working. The pooled track may",
        "contain multiple geometries or reflection modes. V2 tests that explicitly.",
    ]

    OUT_SUMMARY.write_text(
        "\n".join(lines) + "\n"
    )


def main():
    print()
    print("=" * 100)
    print("MARCONI LONG-TERM GNSS-R STABILITY V2")
    print("=" * 100)
    print("Primary tide:", PRIMARY_TIDE_MODEL)
    print("Azimuth cluster tolerance:", AZ_CLUSTER_TOL_DEG, "deg")

    tide_times, tide_at = load_tide_data()
    rows = load_results()
    add_tides(rows, tide_at)

    print("GNSS-R observations:", len(rows))
    print(
        "Observation interval:",
        min(r["datetime_utc"] for r in rows),
        "through",
        max(r["datetime_utc"] for r in rows),
    )

    clustered = cluster_track_rows(rows)

    records = []
    for i, group in enumerate(
        clustered,
        start=1,
    ):
        records.append(
            analyze_track(
                group,
                i,
            )
        )

    records.sort(
        key=lambda r:
            r["longterm_score"],
        reverse=True,
    )

    print()
    print("=" * 100)
    print("TOP CLUSTERED LONG-TERM TRACKS")
    print("=" * 100)

    for i, r in enumerate(records[:25], 1):
        print(
            f"{i:2d} SAT={r['sat']:3d} "
            f"rise={r['rise']:2d} "
            f"N={r['n']:3d} "
            f"days={r['n_days']:2d} "
            f"Az={r['az_mean_deg']:6.2f}±{r['az_std_deg']:.2f} "
            f"r={r['tide_r']:+.4f} "
            f"slope={r['tide_slope']:+.4f} "
            f"RMS={r['unit_slope_RMS_cm']:.2f}cm "
            f"C={r['C_unit_slope_m']:+.3f} "
            f"score={r['longterm_score']:.3f}"
        )

    write_track_csv(records)
    write_summary(records, rows)

    PLOT_DIR.mkdir(parents=True, exist_ok=True)

    print()
    print("Generating plots...")
    plot_all_gnssr_vs_tide_timeseries(rows)
    plot_all_gnssr_vs_tide_scatter(rows)
    plot_top_tracks_vs_tide(rows, records)
    plot_top_track_residuals(rows, records)
    plot_calibration_stability(rows, records)
    plot_daily_population(rows)

    print()
    print("=" * 100)
    print("OUTPUTS")
    print("=" * 100)
    print("CSV:", OUT_CSV.resolve())
    print("Summary:", OUT_SUMMARY.resolve())
    print("Plots:", PLOT_DIR.resolve())
    print("DONE")


if __name__ == "__main__":
    main()
