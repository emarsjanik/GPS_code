#!/usr/bin/env python3
from __future__ import annotations
import csv, math, os, re, sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
try:
    from openpyxl import load_workbook
except ImportError:
    print('ERROR: openpyxl is required.'); sys.exit(1)

BASE_DIR=Path.home()/'GNSS'/'v4.1'
REFL_CODE=Path(os.environ.get('REFL_CODE',str(BASE_DIR/'products'/'refl_code')))
YEAR=2026; STATION='usgs'; CONFIG='ocean17_23_l1_e5_13'
TIDE_FILE=BASE_DIR/'marconi_tides_sherwood.xlsx'
RESULT_DIR=REFL_CODE/str(YEAR)/'results'/STATION/CONFIG
DATUM_OFFSET_M=0.242; PRIMARY_TIDE_MODEL='EOT20_heightm'
MIN_OBS=14; MIN_DAYS=14; MIN_TIDE_R=.90; SLOPE_MIN=.85; SLOPE_MAX=1.15; MAX_UNIT_RMS_M=.30; MAX_AZ_SD_DEG=1.0
NEAR_R_MARGIN=.025; NEAR_SLOPE_MARGIN=.05; NEAR_RMS_MARGIN_M=.05; NEAR_AZ_MARGIN_DEG=.50; NEAR_COUNT_MARGIN=2
OUT_DIR=BASE_DIR/'marconi_longterm_v7'; PLOT_DIR=OUT_DIR/'track_plots'; RESID_DIR=OUT_DIR/'residual_plots'
OUT_CSV=OUT_DIR/'marconi_longterm_track_diagnostics_v7.csv'; OUT_POP_CSV=OUT_DIR/'marconi_longterm_population_statistics_v7.csv'; OUT_SUMMARY=OUT_DIR/'marconi_longterm_summary_v7.txt'

def banner(s): print('\n'+'='*100+'\n'+s+'\n'+'='*100)
def fnum(v):
    try:
        x=float(v); return x if math.isfinite(x) else math.nan
    except: return math.nan
def normalize_dt(v):
    if isinstance(v,datetime): return v.replace(tzinfo=None)
    if hasattr(v,'to_pydatetime'): return v.to_pydatetime().replace(tzinfo=None)
    if isinstance(v,str):
        for fmt in ('%Y-%m-%d %H:%M:%S','%Y-%m-%d %H:%M','%Y/%m/%d %H:%M:%S','%Y/%m/%d %H:%M','%m/%d/%Y %H:%M:%S','%m/%d/%Y %H:%M'):
            try:return datetime.strptime(v.strip(),fmt)
            except:pass
        try:return datetime.fromisoformat(v.strip().replace('Z',''))
        except:return None
    return None
def azdiff(a,b): return abs((a-b+180)%360-180)
def circular_mean(v):
    a=np.asarray(v,float); a=a[np.isfinite(a)]
    if not len(a): return math.nan
    r=np.deg2rad(a); return float(np.rad2deg(np.arctan2(np.mean(np.sin(r)),np.mean(np.cos(r))))%360)
def az_sd(v):
    a=np.asarray(v,float); a=a[np.isfinite(a)]
    if len(a)<2:return 0.0
    m=circular_mean(a); return float(np.std([azdiff(x,m) for x in a],ddof=1))
def regression(x,y):
    x=np.asarray(x,float); y=np.asarray(y,float); ok=np.isfinite(x)&np.isfinite(y); x=x[ok]; y=y[ok]
    if len(x)<2 or np.std(x)==0:return math.nan,math.nan,math.nan
    s,i=np.polyfit(x,y,1); return float(s),float(i),float(np.corrcoef(x,y)[0,1])
def stats(v):
    v=np.asarray(v,float); v=v[np.isfinite(v)]
    if not len(v): return dict(n=0,mean_bias=math.nan,mean_abs_deviation=math.nan,median_abs_deviation=math.nan,rms=math.nan)
    return dict(n=len(v),mean_bias=float(np.mean(v)),mean_abs_deviation=float(np.mean(np.abs(v))),median_abs_deviation=float(np.median(np.abs(v))),rms=float(np.sqrt(np.mean(v*v))))

def parse_result(p):
    rows=[]
    with open(p,'r',errors='replace') as f:
        for line in f:
            z=line.strip()
            if not z or z.startswith('%') or z.startswith('#'):continue
            a=z.split()
            if len(a)<17:continue
            try:y=int(a[0]); doy=int(a[1]); rh=float(a[2]); sat=int(a[3]); utc=float(a[4]); az=float(a[5]); freq=int(a[10]); rise=int(a[11])
            except:continue
            if y!=YEAR or not all(math.isfinite(x) for x in (rh,utc,az)):continue
            dt=datetime(y,1,1)+timedelta(days=doy-1,hours=utc)
            rows.append(dict(datetime_utc=dt,doy=doy,sat=sat,rise=rise,azimuth=az,GNSS_WL_m=rh,freq=freq,source=p))
    return rows

def load_results():
    banner('LOADING ESTABLISHED GNSS-IR RESULT FILES')
    files=sorted(p for p in RESULT_DIR.glob('*.txt') if re.fullmatch(r'\d{3}\.txt',p.name)) if RESULT_DIR.exists() else []
    print('Selected configuration:',CONFIG); print('Selected directory:',RESULT_DIR); print('Daily result files selected:',len(files))
    rows=[]
    for p in files:
        r=[x for x in parse_result(p) if x['freq']==1]
        if r: print(f'{p.stem:>4} : {len(r):5d} observations'); rows.extend(r)
    u={}
    for r in rows:u[(r['datetime_utc'],r['sat'],r['rise'],round(r['azimuth'],5),round(r['GNSS_WL_m'],6))]=r
    rows=sorted(u.values(),key=lambda x:x['datetime_utc']); print('Unique GPS L1 GNSS-R observations:',len(rows)); return rows

def find_col(headers,cands):
    low={str(h).strip().lower():i for i,h in enumerate(headers) if h is not None}
    for c in cands:
        if c.lower() in low:return low[c.lower()]
    for i,h in enumerate(headers):
        if any(c.lower() in str(h).strip().lower() for c in cands):return i
    return None

def load_tide():
    banner('LOADING MARCONI / SHERWOOD TIDE DATA')
    wb=load_workbook(TIDE_FILE,data_only=True,read_only=True); ws=wb['in'] if 'in' in wb.sheetnames else wb[wb.sheetnames[0]]; rows=list(ws.iter_rows(values_only=True))
    hi=0; headers=[str(x).strip() if x is not None else '' for x in rows[0]]
    for i,row in enumerate(rows[:20]):
        h=' '.join(str(x).lower() for x in row if x is not None)
        if 'eot20_heightm' in h or 'timestamp' in h or 'datetime' in h:hi=i;headers=[str(x).strip() if x is not None else '' for x in row];break
    di=find_col(headers,['datetime','date_time','timestamp','time','date']); ti=find_col(headers,[PRIMARY_TIDE_MODEL,'EOT20','heightm','tide','height'])
    di=0 if di is None else di; ti=1 if ti is None and len(headers)>1 else (0 if ti is None else ti)
    tide=[]
    for row in rows[hi+1:]:
        if len(row)<=max(di,ti):continue
        dt=normalize_dt(row[di]); val=fnum(row[ti])
        if dt is not None and math.isfinite(val):tide.append((dt,val))
    tide.sort(); print('Tide sheet:',ws.title); print('Tide records:',len(tide)); print('Tide interval:',tide[0][0],'through',tide[-1][0]); print('Tide column:',headers[ti]); return tide

def interp(tide,dt):
    if dt<tide[0][0] or dt>tide[-1][0]:return math.nan
    lo,hi=0,len(tide)-1
    while lo<=hi:
        m=(lo+hi)//2
        if tide[m][0]<dt:lo=m+1
        elif tide[m][0]>dt:hi=m-1
        else:return tide[m][1]
    i=max(1,lo); t0,y0=tide[i-1]; t1,y1=tide[i]; sec=(t1-t0).total_seconds()
    return math.nan if sec<=0 else y0+(y1-y0)*(dt-t0).total_seconds()/sec

def match(rows,tide):
    out=[]
    for r in rows:
        t=interp(tide,r['datetime_utc'])
        if math.isfinite(t):
            q=dict(r);q['tide_m']=t;q['raw_residual_m']=q['GNSS_WL_m']-t;q['plus_residual_m']=q['GNSS_WL_m']+DATUM_OFFSET_M-t;out.append(q)
    print('GNSS-R observations matched to tide:',len(out));
    if out:print('Observation interval:',min(x['datetime_utc'] for x in out),'through',max(x['datetime_utc'] for x in out))
    return out

def cluster(rows):
    groups=defaultdict(list)
    for r in rows:groups[(r['sat'],r['rise'])].append(r)
    tracks=[]
    for key,vals in sorted(groups.items()):
        cs=[]
        for r in sorted(vals,key=lambda x:x['datetime_utc']):
            hit=False
            for c in cs:
                if azdiff(r['azimuth'],circular_mean([x['azimuth'] for x in c]))<=3.0:c.append(r);hit=True;break
            if not hit:cs.append([r])
        tracks.extend(cs)
    return tracks

def evaluate(track,tid):
    tide=np.array([r['tide_m'] for r in track]);gnss=np.array([r['GNSS_WL_m'] for r in track]);slope,intercept,r=regression(tide,gnss);res=gnss-tide;rms=float(np.sqrt(np.mean(res[np.isfinite(res)]**2)));days=len(set(x['datetime_utc'].date() for x in track)); azsd=az_sd([x['azimuth'] for x in track]);fail=[]
    if len(track)<MIN_OBS:fail.append('N')
    if days<MIN_DAYS:fail.append('DAYS')
    if not math.isfinite(r) or r<MIN_TIDE_R:fail.append('R')
    if not math.isfinite(slope) or not(SLOPE_MIN<=slope<=SLOPE_MAX):fail.append('SLOPE')
    if not math.isfinite(rms) or rms>MAX_UNIT_RMS_M:fail.append('RMS')
    if azsd>MAX_AZ_SD_DEG:fail.append('AZ')
    near=[]
    for f in fail:
        if f=='N' and len(track)>=MIN_OBS-NEAR_COUNT_MARGIN:near.append(f)
        elif f=='DAYS' and days>=MIN_DAYS-NEAR_COUNT_MARGIN:near.append(f)
        elif f=='R' and math.isfinite(r) and r>=MIN_TIDE_R-NEAR_R_MARGIN:near.append(f)
        elif f=='SLOPE' and math.isfinite(slope) and (SLOPE_MIN-NEAR_SLOPE_MARGIN<=slope<SLOPE_MIN or SLOPE_MAX<slope<=SLOPE_MAX+NEAR_SLOPE_MARGIN):near.append(f)
        elif f=='RMS' and rms<=MAX_UNIT_RMS_M+NEAR_RMS_MARGIN_M:near.append(f)
        elif f=='AZ' and azsd<=MAX_AZ_SD_DEG+NEAR_AZ_MARGIN_DEG:near.append(f)
    cat='GOOD' if not fail else ('NEAR-GOOD' if len(fail)==1 and len(near)==1 else 'BAD')
    return dict(track_id=tid,sat=track[0]['sat'],rise=track[0]['rise'],n=len(track),days=days,azmean=circular_mean([x['azimuth'] for x in track]),azsd=azsd,r=r,slope=slope,intercept=intercept,rms_m=rms,fail=','.join(fail) or 'PASS',near_reason=','.join(near),category=cat,track=track,raw_stats=stats(res),plus_stats=stats(res+DATUM_OFFSET_M))

def population(good):
    raw=[r['raw_residual_m'] for d in good for r in d['track']];plus=[r['plus_residual_m'] for d in good for r in d['track']];return stats(raw),stats(plus)

def plot_track(d):
    tr=sorted(d['track'],key=lambda r:r['datetime_utc']);t=[r['datetime_utc'] for r in tr]; tide=np.array([r['tide_m'] for r in tr]);g=np.array([r['GNSS_WL_m'] for r in tr]);p=g+DATUM_OFFSET_M; safe=f"track_{d['track_id']:02d}_sat{d['sat']}_r{d['rise']:+d}_{d['category'].lower().replace('-','_')}"
    fig,ax=plt.subplots(figsize=(12,6));ax.plot(t,tide,'o-',label='Tide model');ax.plot(t,g,'o-',label='GNSS-R raw');ax.plot(t,p,'o-',label='GNSS-R +0.242 m');ax.set_title(f"Track {d['track_id']:02d} | SAT {d['sat']} rise {d['rise']:+d} | {d['category']}\nN={d['n']} days={d['days']} r={d['r']:.4f} slope={d['slope']:.4f} RMS={d['rms_m']*100:.2f} cm");ax.set_ylabel('Height (m)');ax.set_xlabel('UTC');ax.grid(True,alpha=.25);ax.legend();fig.autofmt_xdate();fig.tight_layout();fig.savefig(PLOT_DIR/(safe+'.png'),dpi=160);plt.close(fig)
    fig,ax=plt.subplots(figsize=(12,5));ax.axhline(0,linewidth=1);ax.plot(t,(g-tide)*100,'o-',label='Raw GNSS-R − tide');ax.plot(t,(p-tide)*100,'o-',label='GNSS-R +0.242 m − tide');ax.set_title(f'Residuals | Track {d["track_id"]:02d} | SAT {d["sat"]} rise {d["rise"]:+d}');ax.set_ylabel('Residual (cm)');ax.set_xlabel('UTC');ax.grid(True,alpha=.25);ax.legend();fig.autofmt_xdate();fig.tight_layout();fig.savefig(RESID_DIR/(safe+'_residuals.png'),dpi=160);plt.close(fig)

def write(diags,raw,plus):
    OUT_DIR.mkdir(parents=True,exist_ok=True);fields=['track_id','category','sat','rise','n','days','azmean','azsd','r','slope','intercept','rms_m','fail','near_reason','raw_mean_bias_m','raw_mean_abs_deviation_m','raw_median_abs_deviation_m','raw_rms_m','plus_mean_bias_m','plus_mean_abs_deviation_m','plus_median_abs_deviation_m','plus_rms_m']
    with open(OUT_CSV,'w',newline='') as f:
        w=csv.DictWriter(f,fieldnames=fields);w.writeheader()
        for d in diags:
            a,b=d['raw_stats'],d['plus_stats'];w.writerow({'track_id':d['track_id'],'category':d['category'],'sat':d['sat'],'rise':d['rise'],'n':d['n'],'days':d['days'],'azmean':d['azmean'],'azsd':d['azsd'],'r':d['r'],'slope':d['slope'],'intercept':d['intercept'],'rms_m':d['rms_m'],'fail':d['fail'],'near_reason':d['near_reason'],'raw_mean_bias_m':a['mean_bias'],'raw_mean_abs_deviation_m':a['mean_abs_deviation'],'raw_median_abs_deviation_m':a['median_abs_deviation'],'raw_rms_m':a['rms'],'plus_mean_bias_m':b['mean_bias'],'plus_mean_abs_deviation_m':b['mean_abs_deviation'],'plus_median_abs_deviation_m':b['median_abs_deviation'],'plus_rms_m':b['rms']})
    with open(OUT_POP_CSV,'w',newline='') as f:
        w=csv.writer(f);w.writerow(['population','n_observations','mean_bias_m','mean_abs_deviation_m','median_abs_deviation_m','rms_m']);w.writerow(['GOOD raw',raw['n'],raw['mean_bias'],raw['mean_abs_deviation'],raw['median_abs_deviation'],raw['rms']]);w.writerow(['GOOD +0.242m',plus['n'],plus['mean_bias'],plus['mean_abs_deviation'],plus['median_abs_deviation'],plus['rms']])
    with open(OUT_SUMMARY,'w') as f:
        f.write('MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V7\n'+'='*80+'\n\n');f.write(f'Result directory: {RESULT_DIR}\nTide file: {TIDE_FILE}\nPrimary tide model: {PRIMARY_TIDE_MODEL}\nDiagnostic offset: +{DATUM_OFFSET_M:.3f} m\n\n');f.write(f'GOOD criteria: N>={MIN_OBS}, days>={MIN_DAYS}, r>={MIN_TIDE_R}, slope={SLOPE_MIN}-{SLOPE_MAX}, RMS<={MAX_UNIT_RMS_M*100:.1f} cm, az SD<={MAX_AZ_SD_DEG:.1f} deg\n\n');cnt=defaultdict(int)
        for d in diags:cnt[d['category']]+=1
        f.write(f"GOOD: {cnt['GOOD']}\nNEAR-GOOD: {cnt['NEAR-GOOD']}\nBAD: {cnt['BAD']}\n\n")
        for name,s in [('RAW',raw),('+0.242 m',plus)]:f.write(f"{name}: N={s['n']}, mean bias={s['mean_bias']*100:.3f} cm, mean abs deviation={s['mean_abs_deviation']*100:.3f} cm, median abs deviation={s['median_abs_deviation']*100:.3f} cm, RMS={s['rms']*100:.3f} cm\n")
        f.write('\nTRACKS\n');
        for d in diags:f.write(f"Track {d['track_id']:02d} SAT={d['sat']} rise={d['rise']:+d} {d['category']} N={d['n']} days={d['days']} Az={d['azmean']:.2f}±{d['azsd']:.2f} r={d['r']:.4f} slope={d['slope']:.4f} RMS={d['rms_m']*100:.2f} cm fail={d['fail']} near={d['near_reason']}\n")
    print('CSV:',OUT_CSV);print('Population CSV:',OUT_POP_CSV);print('Summary:',OUT_SUMMARY)

def main():
    OUT_DIR.mkdir(parents=True,exist_ok=True);PLOT_DIR.mkdir(parents=True,exist_ok=True);RESID_DIR.mkdir(parents=True,exist_ok=True)
    banner('MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V7');print('Base directory:',BASE_DIR);print('REFL_CODE:',REFL_CODE);print('Result directory:',RESULT_DIR);print('GNSS reference height: 18.665 m\nDiagnostic offset: +0.242 m');print('GOOD thresholds unchanged from V5. +0.242 m is diagnostic only.')
    rows=load_results()
    if not rows:return 2
    tide=load_tide();matched=match(rows,tide)
    if not matched:return 3
    banner('TRACK-BY-TRACK GOOD / NEAR-GOOD / BAD DIAGNOSTICS');ds=[evaluate(t,i+1) for i,t in enumerate(cluster(matched))];order={'GOOD':0,'NEAR-GOOD':1,'BAD':2};ds.sort(key=lambda d:(order[d['category']],-(d['r'] if math.isfinite(d['r']) else -999)))
    for i,d in enumerate(ds,1):d['track_id']=i
    print('ID SAT R N DAYS AZmean AZsd r slope RMScm CATEGORY FAIL');print('-'*100)
    for d in ds:print(f"{d['track_id']:2d} {d['sat']:3d} {d['rise']:2d} {d['n']:2d} {d['days']:4d} {d['azmean']:7.2f} {d['azsd']:5.2f} {d['r']:7.4f} {d['slope']:7.4f} {d['rms_m']*100:7.2f} {d['category']:9s} {d['fail']}")
    good=[d for d in ds if d['category']=='GOOD'];near=[d for d in ds if d['category']=='NEAR-GOOD'];bad=[d for d in ds if d['category']=='BAD'];raw,plus=population(good)
    banner('GOOD-TRACK POPULATION STATISTICS');print('GOOD tracks:',len(good));print('GOOD observations:',raw['n']);
    for label,s in [('RAW',raw),('GNSS-R +0.242 m',plus)]:print(f"\n{label}\n  mean bias                 = {s['mean_bias']*100:.3f} cm\n  mean absolute deviation  = {s['mean_abs_deviation']*100:.3f} cm\n  median absolute deviation = {s['median_abs_deviation']*100:.3f} cm\n  RMS                       = {s['rms']*100:.3f} cm")
    banner('GENERATING TRACK PLOTS');
    for d in ds:plot_track(d)
    write(ds,raw,plus);banner('DONE');print('GOOD:',len(good),' NEAR-GOOD:',len(near),' BAD:',len(bad));return 0
if __name__=='__main__':raise SystemExit(main())
