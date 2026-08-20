#!/usr/bin/env python3
"""
Marconi top-10 repeated-track Fresnel + tide + topobathy validator.

Purpose
-------
Automatically:
  1. Load all GPS L1 GNSS-IR result rows from the controlled
     17-23 m / 5-13 deg experiment.
  2. Rank repeated satellite tracks by tide response, amplitude
     response, repeatability and GNSS-IR quality.
  3. Select the top N repeated tracks (default 10).
  4. Build first-Fresnel ellipses at 5, 9 and 13 degrees using each
     arc's actual recovered RH and azimuth.
  5. Sample the USGS 1-m Marconi topobathy surface beneath each ellipse.
  6. Compare the footprint against the modeled tide surface using the
     same datum-sensitive methodology already established.
  7. Produce a unified track table with:
        - tide correlation
        - free slope
        - fixed-slope RMS
        - PkNoise
        - amplitude
        - azimuth stability
        - RH stability
        - 5/9/13 deg topobathy wet fraction
        - integrated diagnostic class
  8. Produce a KML containing all selected track footprints.

IMPORTANT
---------
This is a DISCOVERY / VALIDATION product, not a production algorithm.

Track selection uses the GNSS-IR observations first. Geometry is then
used as an independent validation layer.

A track is NOT called "confirmed ocean" merely because it has a high
tide correlation. The integrated class requires both strong temporal
behavior and physical footprint evidence.

The tide-model / USGS topobathy datum relationship is still treated as
a diagnostic sensitivity issue. DATUM_SHIFT_M defaults to 0.0 m and
should not be tuned to improve the classification.
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import rasterio
from rasterio.mask import mask
from pyproj import Transformer
from shapely.geometry import Polygon, mapping
from shapely.ops import transform as shapely_transform
from openpyxl import load_workbook
import simplekml

from gnssrefl.refl_zones import makeEllipse_latlon


# ---------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TOPO_FILE = Path(
    "2021022FA_Marconi_topobathy_1m.tif"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

OUT_RANK = Path(
    "marconi_top10_integrated_track_ranking.csv"
)

OUT_SUMMARY = Path(
    "marconi_top10_integrated_track_ranking_summary.txt"
)

OUT_KML = Path(
    "marconi_top10_integrated_fresnel.kml"
)

LAT = 41.8928243333
LON = -69.9633227139
H_ORTHO_M = 18.665

DOYS = [204, 205, 206, 207]

MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]

# Use ensemble mean tide for the geometry classification.
# The previous analysis showed the four models are very similar.
GEOM_TIDE_MODEL = "ENSEMBLE_MEAN"

ELEVATIONS = [5.0, 9.0, 13.0]

# Diagnostic datum shift. DO NOT tune this from the results.
DATUM_SHIFT_M = 0.0

NEAR_WATER_MARGIN_M = 0.15

TOP_N = 10

MIN_TRACK_N = 3

# Ranking weights before geometry:
# 30% tide correlation
# 20% slope closeness to +1
# 15% fixed-slope residual
# 15% PkNoise
# 10% amplitude
# 10% azimuth repeatability
#
# Geometry is deliberately kept separate until after the statistical
# ranking so that unknown geometry does not silently penalize tracks.
WEIGHT_CORR = 0.30
WEIGHT_SLOPE = 0.20
WEIGHT_RMS = 0.15
WEIGHT_PKN = 0.15
WEIGHT_AMP = 0.10
WEIGHT_AZ = 0.10


# ---------------------------------------------------------------------
# MATH HELPERS
# ---------------------------------------------------------------------

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 3:
        return math.nan

    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(np.corrcoef(x, y)[0, 1])


def rms(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(np.sqrt(np.mean(x ** 2)))


def mae(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(np.mean(np.abs(x)))


def utc_hours_to_datetime(year, doy, utc_hours):
    return (
        datetime(year, 1, 1)
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


# ---------------------------------------------------------------------
# TIDE DATA
# ---------------------------------------------------------------------

def load_tides():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [c.value for c in ws[1]]

    time_col = header.index("time")

    model_cols = {
        model: header.index(model)
        for model in MODELS
    }

    times = []
    values = {model: [] for model in MODELS}

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_col]

        if not isinstance(t, datetime):
            continue

        tmp = {}
        good = True

        for model in MODELS:
            v = finite(
                row[model_cols[model]]
            )
            if v is None:
                good = False
                break
            tmp[model] = v

        if not good:
            continue

        times.append(t)

        for model in MODELS:
            values[model].append(
                tmp[model]
            )

    wb.close()

    epoch = np.asarray(
        [
            (t - times[0]).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    arrays = {
        model: np.asarray(values[model], dtype=float)
        for model in MODELS
    }

    def tide_at(dt, model):
        x = (
            dt - times[0]
        ).total_seconds()

        if x < epoch[0] or x > epoch[-1]:
            return math.nan

        return float(
            np.interp(
                x,
                epoch,
                arrays[model],
            )
        )

    return tide_at


# ---------------------------------------------------------------------
# GNSS-IR RESULT DATA
# ---------------------------------------------------------------------

def load_gnss_rows(tide_at):
    rows = []

    for doy in DOYS:

        path = RESULT_DIR / f"{doy}.txt"

        if not path.exists():
            raise SystemExit(
                f"Missing GNSS-IR result file: {path}"
            )

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if not line or line.startswith("%"):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy2 = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc_hours = float(c[4])
                az = float(c[5])
                amp = float(c[6])
                emin = float(c[7])
                emax = float(c[8])
                nobs = int(float(c[9]))
                freq = int(float(c[10]))
                rise = int(float(c[11]))
                edot = float(c[12])
                pkn = float(c[13])
                delt = float(c[14])
                mjd = float(c[15])
                refr = float(c[16])
            except Exception:
                continue

            # This integrated experiment is specifically GPS L1.
            if freq != 1:
                continue

            dt = utc_hours_to_datetime(
                year,
                doy2,
                utc_hours,
            )

            tide_vals = {
                model: tide_at(
                    dt,
                    model,
                )
                for model in MODELS
            }

            finite_tides = [
                v for v in tide_vals.values()
                if math.isfinite(v)
            ]

            tide_ensemble = (
                float(np.mean(finite_tides))
                if finite_tides
                else math.nan
            )

            wl = H_ORTHO_M - rh

            rows.append(
                {
                    "year": year,
                    "doy": doy2,
                    "datetime_utc": dt,
                    "sat": sat,
                    "freq": freq,
                    "RH_m": rh,
                    "GNSS_WL_m": wl,
                    "az_deg": az,
                    "Amp": amp,
                    "PkNoise": pkn,
                    "emin": emin,
                    "emax": emax,
                    "NumbOf": nobs,
                    "rise": rise,
                    "EdotF": edot,
                    "DelT_min": delt,
                    "MJD": mjd,
                    "refr": refr,
                    **tide_vals,
                    "TIDE_ENSEMBLE_m": tide_ensemble,
                }
            )

    rows.sort(
        key=lambda r: (
            r["sat"],
            r["datetime_utc"],
        )
    )

    return rows


# ---------------------------------------------------------------------
# TRACK STATISTICS
# ---------------------------------------------------------------------

def group_tracks(rows):
    groups = defaultdict(list)

    for row in rows:
        groups[
            (row["sat"], row["freq"])
        ].append(row)

    for key in groups:
        groups[key].sort(
            key=lambda r: r["datetime_utc"]
        )

    return groups


def normalize_score(value, low, high):
    if not math.isfinite(value):
        return 0.0

    if high == low:
        return 0.0

    return max(
        0.0,
        min(
            1.0,
            (value - low) / (high - low),
        ),
    )


def compute_track_score(group):
    wl = np.asarray(
        [r["GNSS_WL_m"] for r in group],
        dtype=float,
    )

    tide = np.asarray(
        [r["TIDE_ENSEMBLE_m"] for r in group],
        dtype=float,
    )

    az = np.asarray(
        [r["az_deg"] for r in group],
        dtype=float,
    )

    pkn = np.asarray(
        [r["PkNoise"] for r in group],
        dtype=float,
    )

    amp = np.asarray(
        [r["Amp"] for r in group],
        dtype=float,
    )

    valid = (
        np.isfinite(wl)
        & np.isfinite(tide)
    )

    if np.sum(valid) >= 3:
        r = pearson(
            wl[valid],
            tide[valid],
        )

        slope = float(
            np.polyfit(
                tide[valid],
                wl[valid],
                1,
            )[0]
        )

        C = float(
            np.mean(
                wl[valid]
                - tide[valid]
            )
        )

        residual = (
            wl[valid]
            - C
            - tide[valid]
        )

        unit_rms_cm = rms(
            residual
        ) * 100.0
    else:
        r = math.nan
        slope = math.nan
        C = math.nan
        unit_rms_cm = math.nan

    # Direct, transparent scores.
    corr_score = (
        abs(r)
        if math.isfinite(r)
        else 0.0
    )

    slope_score = (
        max(
            0.0,
            1.0
            - min(
                1.0,
                abs(
                    slope - 1.0
                )
                if math.isfinite(slope)
                else 1.0,
            ),
        )
    )

    rms_score = (
        max(
            0.0,
            1.0
            - min(
                1.0,
                unit_rms_cm / 20.0,
            ),
        )
        if math.isfinite(unit_rms_cm)
        else 0.0
    )

    # PkNoise >4 gets the maximum score.
    pkn_score = max(
        0.0,
        min(
            1.0,
            (float(np.mean(pkn)) - 2.8) / 1.2,
        ),
    )

    # Amplitude >=50 gets the maximum score.
    amp_score = max(
        0.0,
        min(
            1.0,
            float(np.mean(amp)) / 50.0,
        ),
    )

    # Stable azimuth is a desirable repeat-track property.
    az_std = float(
        np.std(az)
    )

    az_score = max(
        0.0,
        min(
            1.0,
            1.0 - az_std / 2.0,
        ),
    )

    score = (
        WEIGHT_CORR * corr_score
        + WEIGHT_SLOPE * slope_score
        + WEIGHT_RMS * rms_score
        + WEIGHT_PKN * pkn_score
        + WEIGHT_AMP * amp_score
        + WEIGHT_AZ * az_score
    )

    return {
        "n": len(group),
        "r": r,
        "slope": slope,
        "C_m": C,
        "unit_rms_cm": unit_rms_cm,
        "az_mean_deg": float(np.mean(az)),
        "az_std_deg": az_std,
        "RH_mean_m": float(np.mean([
            r0["RH_m"] for r0 in group
        ])),
        "RH_std_m": float(np.std([
            r0["RH_m"] for r0 in group
        ])),
        "PkNoise_mean": float(np.mean(pkn)),
        "Amp_mean": float(np.mean(amp)),
        "score_pre_geometry": score,
    }


# ---------------------------------------------------------------------
# FRESNEL / TOPOBATHY
# ---------------------------------------------------------------------

def build_fresnel_polygon(row, elevation):
    lon, lat = makeEllipse_latlon(
        row["freq"],
        elevation,
        row["RH_m"],
        row["az_deg"],
        LAT,
        LON,
    )

    return Polygon(
        list(
            zip(
                map(float, lon),
                map(float, lat),
            )
        )
    )


def raster_values(src, polygon):
    to_raster = Transformer.from_crs(
        "EPSG:4326",
        src.crs,
        always_xy=True,
    )

    poly_raster = shapely_transform(
        lambda x, y, z=None:
            to_raster.transform(x, y),
        polygon,
    )

    try:
        arr, _ = mask(
            src,
            [mapping(poly_raster)],
            crop=True,
            filled=False,
        )
    except ValueError:
        return np.array([], dtype=float)

    return np.asarray(
        arr[0].compressed(),
        dtype=float,
    )


def track_geometry_metrics(
    group,
    src,
):
    result = {}

    for elev in ELEVATIONS:

        wet_fracs = []
        near_fracs = []
        dry_fracs = []
        cell_counts = []

        for row in group:

            poly = build_fresnel_polygon(
                row,
                elev,
            )

            z = raster_values(
                src,
                poly,
            )

            if len(z) == 0:
                continue

            tide = row[
                "TIDE_ENSEMBLE_m"
            ]

            if not math.isfinite(tide):
                continue

            water_level = (
                tide
                + DATUM_SHIFT_M
            )

            wet = float(
                np.mean(
                    z <= water_level
                )
            )

            near = float(
                np.mean(
                    (
                        z > water_level
                    )
                    & (
                        z <=
                        water_level
                        + NEAR_WATER_MARGIN_M
                    )
                )
            )

            dry = float(
                np.mean(
                    z >
                    water_level
                    + NEAR_WATER_MARGIN_M
                )
            )

            wet_fracs.append(wet)
            near_fracs.append(near)
            dry_fracs.append(dry)
            cell_counts.append(len(z))

        if wet_fracs:
            result[
                f"wet_{int(elev)}_mean"
            ] = float(np.mean(wet_fracs))

            result[
                f"wet_{int(elev)}_min"
            ] = float(np.min(wet_fracs))

            result[
                f"wet_{int(elev)}_max"
            ] = float(np.max(wet_fracs))

            result[
                f"near_{int(elev)}_mean"
            ] = float(np.mean(near_fracs))

            result[
                f"dry_{int(elev)}_mean"
            ] = float(np.mean(dry_fracs))

            result[
                f"geometry_n_{int(elev)}"
            ] = len(wet_fracs)

        else:
            result[
                f"wet_{int(elev)}_mean"
            ] = math.nan

            result[
                f"wet_{int(elev)}_min"
            ] = math.nan

            result[
                f"wet_{int(elev)}_max"
            ] = math.nan

            result[
                f"near_{int(elev)}_mean"
            ] = math.nan

            result[
                f"dry_{int(elev)}_mean"
            ] = math.nan

            result[
                f"geometry_n_{int(elev)}"
            ] = 0

    return result


def integrated_class(metrics):
    wet13 = metrics.get(
        "wet_13_mean",
        math.nan,
    )

    r = metrics.get(
        "r",
        math.nan,
    )

    slope = metrics.get(
        "slope",
        math.nan,
    )

    rms_cm = metrics.get(
        "unit_rms_cm",
        math.nan,
    )

    n = metrics.get(
        "n",
        0,
    )

    if n >= 4 and (
        math.isfinite(r)
        and abs(r) >= 0.99
        and math.isfinite(slope)
        and 0.85 <= slope <= 1.15
        and math.isfinite(rms_cm)
        and rms_cm <= 10.0
        and math.isfinite(wet13)
        and wet13 >= 0.70
    ):
        return "A_STRONG_OCEAN_CANDIDATE"

    if n >= 3 and (
        math.isfinite(r)
        and abs(r) >= 0.98
        and math.isfinite(slope)
        and 0.75 <= slope <= 1.30
        and math.isfinite(rms_cm)
        and rms_cm <= 15.0
    ):
        if math.isfinite(wet13) and wet13 >= 0.50:
            return "B_PROBABLE_OCEAN_OR_MIXED_SHORELINE"
        return "C_TIDE_RESPONSIVE_GEOMETRY_UNCONFIRMED"

    if n >= 3 and (
        math.isfinite(r)
        and abs(r) >= 0.90
    ):
        return "C_TIDE_RESPONSIVE_GEOMETRY_UNCONFIRMED"

    return "D_WEAK_OR_NON_TIDAL"


# ---------------------------------------------------------------------
# KML
# ---------------------------------------------------------------------

def kml_color_for_class(classification):
    if classification == "A_STRONG_OCEAN_CANDIDATE":
        return simplekml.Color.green

    if classification == "B_PROBABLE_OCEAN_OR_MIXED_SHORELINE":
        return simplekml.Color.yellow

    if classification == "C_TIDE_RESPONSIVE_GEOMETRY_UNCONFIRMED":
        return simplekml.Color.orange

    return simplekml.Color.red


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():

    print()
    print("=" * 100)
    print(
        "MARCONI TOP-10 INTEGRATED GNSS-R / FRESNEL / TOPOBATHY TEST"
    )
    print("=" * 100)

    if not TOPO_FILE.exists():
        raise SystemExit(
            f"Missing topobathy file: {TOPO_FILE}"
        )

    tide_at = load_tides()

    rows = load_gnss_rows(
        tide_at
    )

    groups = group_tracks(
        rows
    )

    repeated = {
        key: group
        for key, group in groups.items()
        if len(group) >= MIN_TRACK_N
    }

    print(
        f"GNSS-R rows: {len(rows)}"
    )

    print(
        f"Repeated tracks (N >= {MIN_TRACK_N}): "
        f"{len(repeated)}"
    )

    # -------------------------------------------------------------
    # Statistical first-stage ranking.
    # -------------------------------------------------------------

    first_stage = []

    for key, group in repeated.items():

        stats = compute_track_score(
            group
        )

        first_stage.append(
            {
                "sat": key[0],
                "freq": key[1],
                "track": (
                    f"SAT{key[0]}_FREQ{key[1]}"
                ),
                **stats,
            }
        )

    first_stage.sort(
        key=lambda r:
            r["score_pre_geometry"],
        reverse=True,
    )

    selected = first_stage[
        :TOP_N
    ]

    print()
    print(
        "TOP 10 BEFORE GEOMETRY"
    )
    print(
        "-" * 100
    )

    for i, r in enumerate(
        selected,
        start=1,
    ):
        print(
            f"{i:2d} "
            f"SAT={r['sat']:3d} "
            f"N={r['n']:2d} "
            f"Az={r['az_mean_deg']:6.2f} "
            f"r={r['r']:+.4f} "
            f"slope={r['slope']:+.4f} "
            f"RMS={r['unit_rms_cm']:.2f}cm "
            f"PkN={r['PkNoise_mean']:.2f} "
            f"Amp={r['Amp_mean']:.1f}"
        )

    # -------------------------------------------------------------
    # Geometry for selected tracks.
    # -------------------------------------------------------------

    print()
    print(
        "RUNNING TRUE TOPOBATHY GEOMETRY"
    )

    kml = simplekml.Kml()

    station = kml.newpoint(
        name="Marconi GNSS station"
    )
    station.coords = [
        (LON, LAT)
    ]

    final_rows = []

    with rasterio.open(
        TOPO_FILE
    ) as src:

        for rank, selected_row in enumerate(
            selected,
            start=1,
        ):

            key = (
                selected_row["sat"],
                selected_row["freq"],
            )

            group = repeated[key]

            geom = track_geometry_metrics(
                group,
                src,
            )

            combined = {
                **selected_row,
                **geom,
            }

            classification = integrated_class(
                combined
            )

            combined[
                "integrated_class"
            ] = classification

            # Second-stage integrated score:
            # geometry is now included, with 13-deg wet fraction
            # as the primary physical discriminator.
            wet13 = combined.get(
                "wet_13_mean",
                math.nan,
            )

            if math.isfinite(wet13):
                geom_score = max(
                    0.0,
                    min(
                        1.0,
                        wet13,
                    ),
                )
            else:
                geom_score = 0.0

            combined[
                "integrated_score"
            ] = (
                0.75
                * selected_row[
                    "score_pre_geometry"
                ]
                + 0.25
                * geom_score
            )

            final_rows.append(
                combined
            )

            print()
            print(
                f"{rank}. SAT={key[0]} "
                f"N={selected_row['n']} "
                f"class={classification}"
            )

            for elev in ELEVATIONS:
                print(
                    f"   EL={int(elev):2d}° "
                    f"wet="
                    f"{combined.get(f'wet_{int(elev)}_mean', math.nan):.3f} "
                    f"near="
                    f"{combined.get(f'near_{int(elev)}_mean', math.nan):.3f} "
                    f"dry="
                    f"{combined.get(f'dry_{int(elev)}_mean', math.nan):.3f}"
                )

            # KML folder
            folder = kml.newfolder(
                name=(
                    f"{rank:02d} "
                    f"SAT {key[0]} FREQ {key[1]} "
                    f"{classification}"
                )
            )

            for row in group:

                for elev in ELEVATIONS:

                    poly = build_fresnel_polygon(
                        row,
                        elev,
                    )

                    p = folder.newpolygon(
                        name=(
                            f"{row['datetime_utc'].strftime('%m-%d %H:%M')} "
                            f"EL={int(elev)} "
                            f"RH={row['RH_m']:.3f}"
                        )
                    )

                    p.outerboundaryis = [
                        (
                            float(x),
                            float(y),
                        )
                        for x, y in poly.exterior.coords
                    ]

                    style = simplekml.Style()

                    color = (
                        kml_color_for_class(
                            classification
                        )
                    )

                    style.polystyle.color = (
                        simplekml.Color.changealphaint(
                            80,
                            color,
                        )
                    )

                    style.linestyle.color = color
                    style.linestyle.width = 2

                    p.style = style

    # Sort final ranking by integrated score.
    final_rows.sort(
        key=lambda r:
            r["integrated_score"],
        reverse=True,
    )

    # -------------------------------------------------------------
    # CSV
    # -------------------------------------------------------------

    fields = [
        "sat",
        "freq",
        "track",
        "n",
        "az_mean_deg",
        "az_std_deg",
        "RH_mean_m",
        "RH_std_m",
        "PkNoise_mean",
        "Amp_mean",
        "r",
        "slope",
        "C_m",
        "unit_rms_cm",
        "score_pre_geometry",
        "wet_5_mean",
        "wet_9_mean",
        "wet_13_mean",
        "wet_13_min",
        "wet_13_max",
        "near_13_mean",
        "dry_13_mean",
        "geometry_n_13",
        "integrated_score",
        "integrated_class",
    ]

    with open(
        OUT_RANK,
        "w",
        newline="",
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )

        writer.writeheader()

        for row in final_rows:
            writer.writerow({
                field:
                    row.get(
                        field,
                        math.nan,
                    )
                for field in fields
            })

    # -------------------------------------------------------------
    # Summary
    # -------------------------------------------------------------

    summary = []

    summary.append(
        "MARCONI TOP-10 INTEGRATED GNSS-R / FRESNEL / TOPOBATHY TEST"
    )
    summary.append(
        "=" * 100
    )
    summary.append(
        f"Top N: {TOP_N}"
    )
    summary.append(
        f"Minimum repeated-track N: {MIN_TRACK_N}"
    )
    summary.append(
        f"Geometry elevations: {ELEVATIONS}"
    )
    summary.append(
        f"DATUM_SHIFT_M: {DATUM_SHIFT_M:+.2f} m"
    )
    summary.append(
        ""
    )

    summary.append(
        "CLASSIFICATION"
    )
    summary.append(
        "A = strong ocean candidate: N>=4, |r|>=0.99,"
    )
    summary.append(
        "    slope 0.85-1.15, unit RMS<=10 cm, wet13>=70%"
    )
    summary.append(
        "B = probable ocean / mixed shoreline"
    )
    summary.append(
        "C = tide-responsive but geometry not confirmed"
    )
    summary.append(
        "D = weak or non-tidal"
    )
    summary.append(
        ""
    )

    summary.append(
        "FINAL RANKING"
    )
    summary.append(
        "-" * 100
    )

    for i, row in enumerate(
        final_rows,
        start=1,
    ):

        summary.append(
            f"{i:2d}. "
            f"SAT={row['sat']:3d} "
            f"N={row['n']:2d} "
            f"Az={row['az_mean_deg']:.2f} "
            f"r={row['r']:+.4f} "
            f"slope={row['slope']:+.4f} "
            f"RMS={row['unit_rms_cm']:.2f}cm "
            f"wet13={row['wet_13_mean']:.3f} "
            f"class={row['integrated_class']} "
            f"score={row['integrated_score']:.3f}"
        )

    OUT_SUMMARY.write_text(
        "\n".join(summary)
        + "\n"
    )

    kml.save(
        str(OUT_KML)
    )

    # -------------------------------------------------------------
    # Console
    # -------------------------------------------------------------

    print()
    print("=" * 100)
    print(
        "FINAL INTEGRATED TOP-10"
    )
    print("=" * 100)

    for i, row in enumerate(
        final_rows,
        start=1,
    ):

        print(
            f"{i:2d} "
            f"SAT={row['sat']:3d} "
            f"N={row['n']:2d} "
            f"Az={row['az_mean_deg']:6.2f} "
            f"r={row['r']:+.4f} "
            f"slope={row['slope']:+.4f} "
            f"RMS={row['unit_rms_cm']:.2f}cm "
            f"wet13={row['wet_13_mean']:.3f} "
            f"class={row['integrated_class']} "
            f"score={row['integrated_score']:.3f}"
        )

    print()
    print(
        "Outputs:"
    )
    print(
        f"  {OUT_RANK}"
    )
    print(
        f"  {OUT_SUMMARY}"
    )
    print(
        f"  {OUT_KML}"
    )

    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
