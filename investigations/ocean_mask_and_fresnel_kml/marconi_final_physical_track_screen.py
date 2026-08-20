#!/usr/bin/env python3
"""
MARCONI FINAL PHYSICAL TRACK SCREEN / EXPERIMENTAL WATER-LEVEL PRODUCT

Purpose
-------
Use the current statistical track selection plus the 1-m Marconi topobathy
to create a conservative physical-screening product.

This version fixes two important interpretation issues:

1. Raster coverage is treated explicitly.
   A missing part of a Fresnel footprint is NEVER treated as land or water.

2. Geometry classes are conservative:
      A_ROBUST_WATER_WITH_SHORELINE
      B_WATER_DOMINANT_LOW_ANGLE
      C_INSUFFICIENT_RASTER_COVERAGE
      D_LAND_DOMINANT
      E_MIXED_SHORELINE

The selected tracks are first screened statistically using the same current
experimental criteria:

    N >= 10
    |corrected median bias| <= 5 cm
    corrected RMS <= 20 cm
    tide r >= 0.93

GNSS-R corrected water level:
    WL = H_ortho - RH + 0.242 m

Tide-predicted water surface:
    water_surface_NAVD88 = EOT20 - 0.242 m

The +0.242 m transformation remains an independent hypothesis test. It is
not optimized in this script.

PHYSICAL SCREEN
---------------
For each selected observation, 5°, 9°, and 13° predicted Fresnel ellipses
are sampled against the 1-m DEM.

For each elevation we calculate:

    coverage_fraction
    wet_fraction
    dry_fraction
    DEM statistics

A footprint is "adequately covered" only when >=90% of its polygon area
is represented by finite raster cells.

Track-level geometry classification uses the median wet/coverage fractions
over all observations for that track.

IMPORTANT
---------
This is not a claim that the raster vertical datum is independently proven.
It is a physical-footprint screen using the current -0.242 m MSL/NAVD88
hypothesis.

For a track with inadequate raster coverage, the script reports the required
minimum bounding box of the observed 5/9/13-degree predicted footprints so
the next external raster/topobathy acquisition can cover it.

Outputs
-------
    marconi_final_physical_track_screen.csv
    marconi_final_physical_track_summary.csv
    marconi_final_physical_track_summary.txt
    marconi_final_physical_track.kml
    marconi_final_physical_track_plots/
        01_physical_track_screen.png
        02_selected_tracks_vs_eot20.png
        03_geometry_coverage.png
        04_wet_fraction_by_elevation.png
        05_required_raster_extension.png

The final experimental water-level product contains only tracks classified
A or B. Tracks classified C/D/E remain in the diagnostics but are NOT
included in the final combined product.
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from xml.sax.saxutils import escape

import matplotlib.pyplot as plt
import numpy as np
import rasterio
from matplotlib.path import Path as MplPath
from openpyxl import load_workbook
from pyproj import Transformer
from rasterio.windows import Window
from gnssrefl.gnss_frequencies import get_wavelength


# =====================================================================
# CONFIGURATION
# =====================================================================

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

TOPOBATHY_FILE = Path(
    "2021022FA_Marconi_topobathy_1m.tif"
)

H_ORTHO_M = 18.665
DATUM_SHIFT_M = 0.242
MSL_NAVD88_M = -DATUM_SHIFT_M

PRIMARY_TIDE = "EOT20_heightm"

STATION_LAT = 41.8928243333
STATION_LON = -69.9633227139

AZ_TOL_DEG = 3.0
MIN_TRACK_N = 10

MAX_ABS_CORRECTED_MEDIAN_M = 0.05
MAX_CORRECTED_RMS_M = 0.20
MIN_TIDE_R = 0.93

TEST_ELEVATIONS_DEG = [5.0, 9.0, 13.0]

MIN_RASTER_COVERAGE = 0.90
WATER_DOMINANT = 0.80
LAND_DOMINANT = 0.20
ROBUST_13_WATER = 0.50

OUT_OBS = Path(
    "marconi_final_physical_track_screen.csv"
)

OUT_TRACKS = Path(
    "marconi_final_physical_track_summary.csv"
)

OUT_SUMMARY = Path(
    "marconi_final_physical_track_summary.txt"
)

OUT_KML = Path(
    "marconi_final_physical_track.kml"
)

OUT_DAILY = Path(
    "marconi_final_physical_daily_product.csv"
)

PLOT_DIR = Path(
    "marconi_final_physical_track_plots"
)


# =====================================================================
# BASIC HELPERS
# =====================================================================

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def utc_datetime(year, doy, utc_hours):
    return (
        datetime(year, 1, 1)
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


def circular_az_diff(a, b):
    d = abs(a - b) % 360.0
    return min(d, 360.0 - d)


def correlation(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if len(x) < 3 or np.std(x) == 0 or np.std(y) == 0:
        return math.nan
    return float(np.corrcoef(x, y)[0, 1])


def rms(x):
    x = np.asarray(x, dtype=float)
    return float(np.sqrt(np.mean(x ** 2))) if len(x) else math.nan


def polygon_area(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    if len(x) < 3:
        return 0.0
    return float(
        abs(
            0.5
            * np.sum(
                x * np.roll(y, -1)
                - y * np.roll(x, -1)
            )
        )
    )


# =====================================================================
# TIDE
# =====================================================================

def load_tide():
    wb = load_workbook(TIDE_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]

    time_col = header.index("time")
    tide_col = header.index(PRIMARY_TIDE)

    times = []
    values = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[time_col]
        v = finite(row[tide_col])

        if isinstance(t, datetime) and v is not None:
            times.append(t)
            values.append(v)

    wb.close()

    return (
        np.asarray(times, dtype="datetime64[ms]"),
        np.asarray(values, dtype=float),
    )


def tide_at(tide_times, tide_values, dt):
    x = tide_times.astype("int64")
    q = np.datetime64(dt, "ms").astype("int64")

    if q < x[0] or q > x[-1]:
        return math.nan

    return float(np.interp(q, x, tide_values))


# =====================================================================
# GNSS-R
# =====================================================================

def load_gnssr():
    rows = []

    for path in sorted(RESULT_DIR.glob("*.txt")):
        try:
            int(path.stem)
        except Exception:
            continue

        for line in path.read_text(errors="replace").splitlines():
            line = line.strip()

            if not line or line.startswith("%"):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc = float(c[4])
                az = float(c[5])
                amp = float(c[6])
                emin = float(c[7])
                emax = float(c[8])
                nobs = int(float(c[9]))
                freq = int(float(c[10]))
                rise = int(float(c[11]))
                pkn = float(c[13])
                delt = float(c[14])
            except Exception:
                continue

            if freq != 1:
                continue

            dt = utc_datetime(year, doy, utc)

            raw_wl = H_ORTHO_M - rh
            corrected_wl = raw_wl + DATUM_SHIFT_M

            rows.append({
                "datetime": dt,
                "year": year,
                "doy": doy,
                "sat": sat,
                "freq": freq,
                "rise": rise,
                "az": az,
                "RH_m": rh,
                "raw_wl_m": raw_wl,
                "corrected_wl_m": corrected_wl,
                "Amp": amp,
                "PkNoise": pkn,
                "NumbOf": nobs,
                "DelT_min": delt,
                "emin_deg": emin,
                "emax_deg": emax,
            })

    rows.sort(key=lambda r: r["datetime"])
    return rows


def attach_tide(rows, tide_times, tide_values):
    out = []

    for row in rows:
        tide = tide_at(
            tide_times,
            tide_values,
            row["datetime"],
        )

        if not math.isfinite(tide):
            continue

        row["EOT20_m"] = tide
        row["corrected_residual_m"] = (
            row["corrected_wl_m"] - tide
        )

        out.append(row)

    return out


# =====================================================================
# TRACKS
# =====================================================================

def build_tracks(rows):
    buckets = defaultdict(list)

    for row in rows:
        buckets[
            (row["sat"], row["freq"], row["rise"])
        ].append(row)

    tracks = []

    for _, group in buckets.items():
        group = sorted(group, key=lambda r: r["az"])

        current = []
        previous = None

        for row in group:
            if (
                previous is None
                or circular_az_diff(
                    row["az"],
                    previous,
                ) <= AZ_TOL_DEG
            ):
                current.append(row)
            else:
                if len(current) >= MIN_TRACK_N:
                    tracks.append(current)
                current = [row]

            previous = row["az"]

        if len(current) >= MIN_TRACK_N:
            tracks.append(current)

    return tracks


def statistical_summary(group):
    wl = np.asarray(
        [r["corrected_wl_m"] for r in group],
        dtype=float,
    )
    tide = np.asarray(
        [r["EOT20_m"] for r in group],
        dtype=float,
    )
    resid = wl - tide

    return {
        "sat": group[0]["sat"],
        "freq": group[0]["freq"],
        "rise": group[0]["rise"],
        "n": len(group),
        "n_days": len({r["doy"] for r in group}),
        "az_mean_deg": float(np.mean([r["az"] for r in group])),
        "az_sd_deg": float(np.std([r["az"] for r in group])),
        "tide_r": correlation(wl, tide),
        "tide_slope": float(np.polyfit(tide, wl, 1)[0]),
        "corrected_median_m": float(np.median(resid)),
        "corrected_rms_m": rms(resid),
    }


def statistical_select(summary):
    return (
        summary["n"] >= MIN_TRACK_N
        and abs(summary["corrected_median_m"])
        <= MAX_ABS_CORRECTED_MEDIAN_M
        and summary["corrected_rms_m"]
        <= MAX_CORRECTED_RMS_M
        and summary["tide_r"] >= MIN_TIDE_R
    )


# =====================================================================
# FRESNEL GEOMETRY
# =====================================================================

def local_fresnel(
    freq,
    sat,
    elevation_deg,
    rh_m,
    az_deg,
):
    if rh_m <= 0:
        raise ValueError(
            f"Invalid reflector height {rh_m}"
        )

    wavelength = float(
        get_wavelength(
            int(freq),
            sat=int(sat),
        )
    )

    delta = wavelength / 2.0
    e = math.radians(elevation_deg)
    sin_e = math.sin(e)

    if sin_e <= 0:
        raise ValueError("Elevation must be positive.")

    b = math.sqrt(
        2.0 * delta * rh_m / sin_e
        + (delta / sin_e) ** 2
    )

    a = b / sin_e

    center = (
        rh_m + delta / sin_e
    ) / math.tan(e)

    # gnssrefl orientation convention
    angle = math.radians(
        360.0 - float(az_deg) + 90.0
    )

    t = np.deg2rad(
        np.arange(
            0.0,
            365.0,
            5.0,
        )
    )

    x0 = a * np.cos(t)
    y0 = b * np.sin(t)

    ce = math.cos(angle)
    se = math.sin(angle)

    east = (
        ce * x0
        - se * y0
        + center * ce
    )

    north = (
        se * x0
        + ce * y0
        + center * se
    )

    return {
        "wavelength_m": wavelength,
        "A_m": float(a),
        "B_m": float(b),
        "center_m": float(center),
        "east_m": east,
        "north_m": north,
    }


# =====================================================================
# DEM SAMPLING
# =====================================================================

def sample_polygon(
    raster,
    polygon_east,
    polygon_north,
    water_surface_m,
):
    min_e = float(np.min(polygon_east))
    max_e = float(np.max(polygon_east))
    min_n = float(np.min(polygon_north))
    max_n = float(np.max(polygon_north))

    row_a, col_a = raster.index(min_e, max_n)
    row_b, col_b = raster.index(max_e, min_n)

    rmin = min(row_a, row_b)
    rmax = max(row_a, row_b)
    cmin = min(col_a, col_b)
    cmax = max(col_a, col_b)

    # If bbox is completely outside raster, report zero coverage.
    if (
        rmax < 0
        or rmin >= raster.height
        or cmax < 0
        or cmin >= raster.width
    ):
        return {
            "coverage_fraction": 0.0,
            "wet_fraction": math.nan,
            "dry_fraction": math.nan,
            "mean_dem_m": math.nan,
            "min_dem_m": math.nan,
            "max_dem_m": math.nan,
            "n_cells": 0,
            "status": "OUTSIDE_RASTER",
        }

    rmin = max(0, rmin)
    rmax = min(raster.height - 1, rmax)
    cmin = max(0, cmin)
    cmax = min(raster.width - 1, cmax)

    width = cmax - cmin + 1
    height = rmax - rmin + 1

    data = np.asarray(
        raster.read(
            1,
            window=Window(
                cmin,
                rmin,
                width,
                height,
            ),
        ),
        dtype=float,
    )

    rows_idx = np.arange(rmin, rmax + 1)
    cols_idx = np.arange(cmin, cmax + 1)
    cc, rr = np.meshgrid(cols_idx, rows_idx)

    xs, ys = raster.xy(
        rr,
        cc,
        offset="center",
    )

    xs = np.asarray(xs, dtype=float).reshape(data.shape)
    ys = np.asarray(ys, dtype=float).reshape(data.shape)

    poly = np.column_stack(
        [
            np.asarray(polygon_east, dtype=float),
            np.asarray(polygon_north, dtype=float),
        ]
    )

    inside = MplPath(poly).contains_points(
        np.column_stack(
            [
                xs.ravel(),
                ys.ravel(),
            ]
        )
    ).reshape(data.shape)

    valid = inside & np.isfinite(data)
    n_cells = int(np.sum(valid))

    total_polygon_area = polygon_area(
        polygon_east,
        polygon_north,
    )

    cell_area = abs(
        float(raster.res[0])
        * float(raster.res[1])
    )

    coverage_fraction = (
        min(
            1.0,
            max(
                0.0,
                (n_cells * cell_area)
                / total_polygon_area
                if total_polygon_area > 0
                else 0.0,
            ),
        )
    )

    if n_cells == 0:
        return {
            "coverage_fraction": coverage_fraction,
            "wet_fraction": math.nan,
            "dry_fraction": math.nan,
            "mean_dem_m": math.nan,
            "min_dem_m": math.nan,
            "max_dem_m": math.nan,
            "n_cells": 0,
            "status": "OUTSIDE_RASTER",
        }

    dem = data[valid]
    wet = dem <= water_surface_m

    return {
        "coverage_fraction": coverage_fraction,
        "wet_fraction": float(np.mean(wet)),
        "dry_fraction": float(1.0 - np.mean(wet)),
        "mean_dem_m": float(np.mean(dem)),
        "min_dem_m": float(np.min(dem)),
        "max_dem_m": float(np.max(dem)),
        "n_cells": n_cells,
        "status": (
            "ADEQUATE"
            if coverage_fraction >= MIN_RASTER_COVERAGE
            else "PARTIAL_RASTER"
        ),
    }


# =====================================================================
# GEOMETRY CLASS
# =====================================================================

def classify_track(track_geometry):
    """
    Conservative classification.

    A:
      adequate coverage at 5, 9, 13
      5 and 9 water-dominant
      13 remains >=50% water

    B:
      adequate 9-degree coverage
      9-degree water-dominant
      13-degree may transition strongly toward land
      (coastal shoreline interaction)

    C:
      9-degree coverage insufficient

    D:
      adequately covered 9-degree footprint is land-dominant

    E:
      adequately covered 9-degree footprint is mixed.
    """
    c5 = track_geometry["coverage_5"]
    c9 = track_geometry["coverage_9"]
    c13 = track_geometry["coverage_13"]

    w5 = track_geometry["wet_5"]
    w9 = track_geometry["wet_9"]
    w13 = track_geometry["wet_13"]

    if not math.isfinite(c9) or c9 < MIN_RASTER_COVERAGE:
        return "C_INSUFFICIENT_RASTER_COVERAGE"

    if not math.isfinite(w9):
        return "C_INSUFFICIENT_RASTER_COVERAGE"

    if w9 <= LAND_DOMINANT:
        return "D_LAND_DOMINANT"

    if (
        w9 >= WATER_DOMINANT
        and math.isfinite(c5)
        and c5 >= MIN_RASTER_COVERAGE
        and math.isfinite(w5)
        and w5 >= WATER_DOMINANT
        and math.isfinite(c13)
        and c13 >= MIN_RASTER_COVERAGE
        and math.isfinite(w13)
        and w13 >= ROBUST_13_WATER
    ):
        return "A_ROBUST_WATER_WITH_SHORELINE"

    if (
        w9 >= WATER_DOMINANT
    ):
        return "B_WATER_DOMINANT_LOW_ANGLE"

    return "E_MIXED_SHORELINE"


# =====================================================================
# PER-OBSERVATION GEOMETRY
# =====================================================================

def process_observation(
    row,
    raster,
    station_e,
    station_n,
):
    tide = row["EOT20_m"]

    water_surface_m = (
        tide
        + MSL_NAVD88_M
    )

    predicted_rh = (
        H_ORTHO_M
        - water_surface_m
    )

    result = {
        "datetime": row["datetime"].isoformat(),
        "sat": row["sat"],
        "freq": row["freq"],
        "rise": row["rise"],
        "az_deg": row["az"],
        "RH_observed_m": row["RH_m"],
        "EOT20_m": tide,
        "water_surface_NAVD88_m": water_surface_m,
        "RH_predicted_m": predicted_rh,
        "corrected_residual_m":
            row["corrected_residual_m"],
    }

    for elev in TEST_ELEVATIONS_DEG:

        tag = str(int(elev))

        geometry = local_fresnel(
            row["freq"],
            row["sat"],
            elev,
            predicted_rh,
            row["az"],
        )

        abs_e = (
            station_e
            + geometry["east_m"]
        )
        abs_n = (
            station_n
            + geometry["north_m"]
        )

        metrics = sample_polygon(
            raster,
            abs_e,
            abs_n,
            water_surface_m,
        )

        result[
            f"coverage_{tag}"
        ] = metrics["coverage_fraction"]

        result[
            f"wet_{tag}"
        ] = metrics["wet_fraction"]

        result[
            f"dry_{tag}"
        ] = metrics["dry_fraction"]

        result[
            f"mean_dem_{tag}"
        ] = metrics["mean_dem_m"]

        result[
            f"min_dem_{tag}"
        ] = metrics["min_dem_m"]

        result[
            f"max_dem_{tag}"
        ] = metrics["max_dem_m"]

        result[
            f"A_{tag}_m"
        ] = geometry["A_m"]

        result[
            f"B_{tag}_m"
        ] = geometry["B_m"]

        result[
            f"center_{tag}_m"
        ] = geometry["center_m"]

        # Save bbox for the 9° footprint for raster-extension mapping.
        if elev == 9.0:
            result["pred9_min_e"] = float(np.min(abs_e))
            result["pred9_max_e"] = float(np.max(abs_e))
            result["pred9_min_n"] = float(np.min(abs_n))
            result["pred9_max_n"] = float(np.max(abs_n))

    return result


# =====================================================================
# TRACK GEOMETRY SUMMARY
# =====================================================================

def median_finite(values):
    values = [
        float(v)
        for v in values
        if v is not None
        and math.isfinite(float(v))
    ]
    if not values:
        return math.nan
    return float(np.median(values))


def summarize_geometry(
    stat_summary,
    geometry_rows,
):
    values = {}

    for elev in [5, 9, 13]:
        values[f"coverage_{elev}"] = median_finite(
            [
                r[f"coverage_{elev}"]
                for r in geometry_rows
            ]
        )
        values[f"wet_{elev}"] = median_finite(
            [
                r[f"wet_{elev}"]
                for r in geometry_rows
            ]
        )

    geom_class = classify_track(values)

    bbox = {}
    for key in [
        "pred9_min_e",
        "pred9_max_e",
        "pred9_min_n",
        "pred9_max_n",
    ]:
        nums = [
            r[key]
            for r in geometry_rows
            if math.isfinite(
                float(r[key])
            )
        ]
        bbox[key] = (
            float(np.min(nums))
            if "min" in key
            else float(np.max(nums))
            if nums
            else math.nan
        )

    return {
        **stat_summary,
        **values,
        **bbox,
        "geometry_class": geom_class,
    }


# =====================================================================
# DAILY FINAL PRODUCT
# =====================================================================

def build_daily_product(
    selected_rows,
):
    groups = defaultdict(list)

    for row in selected_rows:
        groups[
            row["datetime"].date()
        ].append(row)

    daily = []

    for day, group in sorted(
        groups.items()
    ):
        wl = np.asarray(
            [
                r["corrected_wl_m"]
                for r in group
            ],
            dtype=float,
        )
        tide = np.asarray(
            [
                r["EOT20_m"]
                for r in group
            ],
            dtype=float,
        )

        resid = wl - tide

        daily.append({
            "date": str(day),
            "n_observations": len(group),
            "n_tracks": len(
                {
                    (
                        r["sat"],
                        r["freq"],
                        r["rise"],
                    )
                    for r in group
                }
            ),
            "GNSSR_median_m": float(
                np.median(wl)
            ),
            "GNSSR_mean_m": float(
                np.mean(wl)
            ),
            "GNSSR_sd_m": float(
                np.std(wl)
            ),
            "EOT20_median_m": float(
                np.median(tide)
            ),
            "median_residual_m": float(
                np.median(resid)
            ),
            "rms_residual_m": rms(resid),
        })

    return daily


# =====================================================================
# PLOTS
# =====================================================================

def plot_screen(track_summaries):
    labels = []
    vals = []
    classes = []

    for item in track_summaries:
        labels.append(
            f"PRN {item['sat']} "
            f"{'R' if item['rise']==1 else 'S'} "
            f"{item['az_mean_deg']:.1f}°"
        )
        vals.append(
            item["wet_9"]
            if math.isfinite(item["wet_9"])
            else 0.0
        )
        classes.append(
            item["geometry_class"]
        )

    y = np.arange(len(labels))

    fig, ax = plt.subplots(
        figsize=(11, 7)
    )

    ax.barh(
        y,
        vals,
    )

    ax.axvline(
        WATER_DOMINANT,
        linestyle="--",
        linewidth=1.5,
    )

    ax.axvline(
        LAND_DOMINANT,
        linestyle="--",
        linewidth=1.5,
    )

    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.set_xlim(0, 1)

    ax.set_xlabel(
        "Median 9° predicted wet fraction"
    )

    ax.set_title(
        "Marconi Physical Track Screen"
    )

    ax.grid(
        True,
        axis="x",
        alpha=0.25,
    )

    for yy, txt in zip(y, classes):
        ax.text(
            1.01,
            yy,
            txt,
            va="center",
            fontsize=8,
        )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR / "01_physical_track_screen.png",
        dpi=200,
    )

    plt.close(fig)


def plot_final_tracks(
    tide_times,
    tide_values,
    selected_rows,
):
    fig, ax = plt.subplots(
        figsize=(16, 8)
    )

    if selected_rows:

        start = min(
            r["datetime"]
            for r in selected_rows
        )
        end = max(
            r["datetime"]
            for r in selected_rows
        )

        mask = (
            (tide_times >= np.datetime64(start))
            & (tide_times <= np.datetime64(end))
        )

        ax.plot(
            tide_times[mask],
            tide_values[mask],
            linewidth=2.5,
            label="EOT20 tide model",
        )

    by_track = defaultdict(list)

    for row in selected_rows:
        by_track[
            (
                row["sat"],
                row["rise"],
            )
        ].append(row)

    for key, group in sorted(
        by_track.items()
    ):
        group.sort(
            key=lambda r:
                r["datetime"]
        )

        ax.plot(
            [
                r["datetime"]
                for r in group
            ],
            [
                r["corrected_wl_m"]
                for r in group
            ],
            marker="o",
            markersize=3,
            linewidth=1.2,
            label=(
                f"PRN {key[0]} "
                f"{'R' if key[1] == 1 else 'S'}"
            ),
        )

    ax.set_title(
        "Marconi Physically Screened GNSS-R Tracks vs EOT20"
    )
    ax.set_xlabel("UTC")
    ax.set_ylabel(
        "Water-level elevation / anomaly (m)"
    )
    ax.grid(
        True,
        alpha=0.25,
    )
    ax.legend(
        ncol=2,
        fontsize=8,
    )

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR / "02_selected_tracks_vs_eot20.png",
        dpi=200,
    )

    plt.close(fig)


def plot_coverage(track_summaries):
    labels = []
    c5 = []
    c9 = []
    c13 = []

    for item in track_summaries:
        labels.append(
            f"PRN {item['sat']} "
            f"{'R' if item['rise']==1 else 'S'} "
            f"{item['az_mean_deg']:.1f}°"
        )

        c5.append(
            item["coverage_5"]
            if math.isfinite(item["coverage_5"])
            else 0
        )
        c9.append(
            item["coverage_9"]
            if math.isfinite(item["coverage_9"])
            else 0
        )
        c13.append(
            item["coverage_13"]
            if math.isfinite(item["coverage_13"])
            else 0
        )

    y = np.arange(len(labels))

    fig, ax = plt.subplots(
        figsize=(11, 7)
    )

    ax.plot(
        c5,
        y,
        "o-",
        label="5° coverage",
    )
    ax.plot(
        c9,
        y,
        "o-",
        label="9° coverage",
    )
    ax.plot(
        c13,
        y,
        "o-",
        label="13° coverage",
    )

    ax.axvline(
        MIN_RASTER_COVERAGE,
        linestyle="--",
        linewidth=1.5,
    )

    ax.set_yticks(y)
    ax.set_yticklabels(labels)

    ax.set_xlim(0, 1.05)

    ax.set_xlabel(
        "Fraction of Fresnel footprint covered by finite DEM"
    )

    ax.set_title(
        "Marconi Fresnel Raster Coverage"
    )

    ax.grid(
        True,
        axis="x",
        alpha=0.25,
    )
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR / "03_geometry_coverage.png",
        dpi=200,
    )

    plt.close(fig)


def plot_wet_by_elevation(track_summaries):
    labels = []
    w5 = []
    w9 = []
    w13 = []

    for item in track_summaries:
        labels.append(
            f"PRN {item['sat']} "
            f"{'R' if item['rise']==1 else 'S'} "
            f"{item['az_mean_deg']:.1f}°"
        )

        w5.append(
            item["wet_5"]
            if math.isfinite(item["wet_5"])
            else np.nan
        )
        w9.append(
            item["wet_9"]
            if math.isfinite(item["wet_9"])
            else np.nan
        )
        w13.append(
            item["wet_13"]
            if math.isfinite(item["wet_13"])
            else np.nan
        )

    x = np.arange(len(labels))
    width = 0.24

    fig, ax = plt.subplots(
        figsize=(12, 7)
    )

    ax.bar(
        x - width,
        w5,
        width,
        label="5°",
    )
    ax.bar(
        x,
        w9,
        width,
        label="9°",
    )
    ax.bar(
        x + width,
        w13,
        width,
        label="13°",
    )

    ax.axhline(
        WATER_DOMINANT,
        linestyle="--",
        linewidth=1.5,
    )

    ax.set_xticks(x)
    ax.set_xticklabels(
        labels,
        rotation=30,
        ha="right",
    )

    ax.set_ylim(0, 1)

    ax.set_ylabel(
        "Median wet fraction"
    )

    ax.set_title(
        "Marconi Predicted Fresnel Wet Fraction vs Elevation"
    )

    ax.grid(
        True,
        axis="y",
        alpha=0.25,
    )
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR / "04_wet_fraction_by_elevation.png",
        dpi=200,
    )

    plt.close(fig)


def plot_required_extension(
    raster,
    track_summaries,
):
    labels = []
    west_missing = []
    east_missing = []
    south_missing = []
    north_missing = []

    for item in track_summaries:

        if item["geometry_class"] != "C_INSUFFICIENT_RASTER_COVERAGE":
            continue

        labels.append(
            f"PRN {item['sat']} "
            f"{'R' if item['rise']==1 else 'S'} "
            f"{item['az_mean_deg']:.1f}°"
        )

        west_missing.append(
            max(
                0.0,
                raster.bounds.left
                - item["pred9_min_e"],
            )
        )

        east_missing.append(
            max(
                0.0,
                item["pred9_max_e"]
                - raster.bounds.right,
            )
        )

        south_missing.append(
            max(
                0.0,
                raster.bounds.bottom
                - item["pred9_min_n"],
            )
        )

        north_missing.append(
            max(
                0.0,
                item["pred9_max_n"]
                - raster.bounds.top,
            )
        )

    if not labels:
        return

    x = np.arange(len(labels))
    width = 0.2

    fig, ax = plt.subplots(
        figsize=(11, 7)
    )

    ax.bar(
        x - 1.5 * width,
        west_missing,
        width,
        label="West",
    )
    ax.bar(
        x - 0.5 * width,
        east_missing,
        width,
        label="East",
    )
    ax.bar(
        x + 0.5 * width,
        south_missing,
        width,
        label="South",
    )
    ax.bar(
        x + 1.5 * width,
        north_missing,
        width,
        label="North",
    )

    ax.set_xticks(x)
    ax.set_xticklabels(labels)

    ax.set_ylabel(
        "Additional raster extent required (m)"
    )

    ax.set_title(
        "Required Topobathy Extension for Inadequately Covered Tracks"
    )

    ax.grid(
        True,
        axis="y",
        alpha=0.25,
    )
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        PLOT_DIR / "05_required_raster_extension.png",
        dpi=200,
    )

    plt.close(fig)


# =====================================================================
# KML
# =====================================================================

def write_kml(
    raster,
    station_e,
    station_n,
    selected_track_groups,
):
    to_wgs84 = Transformer.from_crs(
        raster.crs,
        "EPSG:4326",
        always_xy=True,
    )

    with open(
        OUT_KML,
        "w",
    ) as f:
        f.write(
            '<?xml version="1.0" encoding="UTF-8"?>\n'
        )
        f.write(
            '<kml xmlns="http://www.opengis.net/kml/2.2">\n'
        )
        f.write("<Document>\n")
        f.write(
            "<name>Marconi physically screened GNSS-R tracks</name>\n"
        )

        for group in selected_track_groups:
            row = group[0]

            predicted_rh = (
                H_ORTHO_M
                - (
                    row["EOT20_m"]
                    + MSL_NAVD88_M
                )
            )

            geom = local_fresnel(
                row["freq"],
                row["sat"],
                9.0,
                predicted_rh,
                row["az"],
            )

            east = station_e + geom["east_m"]
            north = station_n + geom["north_m"]

            coords = []

            for e, n in zip(east, north):
                lon, lat = to_wgs84.transform(
                    e,
                    n,
                )
                coords.append(
                    f"{lon:.7f},{lat:.7f},0"
                )

            label = (
                f"PRN {row['sat']} "
                f"{'Rising' if row['rise']==1 else 'Setting'} "
                f"Az {row['az']:.1f}°"
            )

            f.write("<Placemark>\n")
            f.write(
                f"<name>{escape(label)}</name>\n"
            )
            f.write(
                "<Style><LineStyle>"
                "<width>3</width>"
                "</LineStyle></Style>\n"
            )
            f.write(
                "<Polygon><outerBoundaryIs>"
                "<LinearRing><coordinates>"
                + " ".join(coords)
                + "</coordinates></LinearRing>"
                "</outerBoundaryIs></Polygon>\n"
            )
            f.write("</Placemark>\n")

        lon, lat = to_wgs84.transform(
            station_e,
            station_n,
        )

        f.write("<Placemark>\n")
        f.write("<name>GNSS-R station</name>\n")
        f.write(
            "<Point><coordinates>"
            f"{lon:.7f},{lat:.7f},0"
            "</coordinates></Point>\n"
        )
        f.write("</Placemark>\n")

        f.write("</Document>\n")
        f.write("</kml>\n")


# =====================================================================
# CSV / SUMMARY
# =====================================================================

def write_observation_csv(rows):
    if not rows:
        return

    fields = list(rows[0].keys())

    with open(
        OUT_OBS,
        "w",
        newline="",
    ) as f:
        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )
        writer.writeheader()
        writer.writerows(rows)


def write_track_csv(rows):
    if not rows:
        return

    fields = list(rows[0].keys())

    with open(
        OUT_TRACKS,
        "w",
        newline="",
    ) as f:
        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )
        writer.writeheader()
        writer.writerows(rows)


def write_daily_csv(rows):
    if not rows:
        return

    fields = list(rows[0].keys())

    with open(
        OUT_DAILY,
        "w",
        newline="",
    ) as f:
        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )
        writer.writeheader()
        writer.writerows(rows)


def write_summary(
    track_summaries,
    final_summaries,
    final_rows,
    daily,
):
    lines = [
        "MARCONI FINAL PHYSICAL GNSS-R TRACK SCREEN",
        "=" * 96,
        "",
        f"Topobathy: {TOPOBATHY_FILE}",
        f"Primary tide: {PRIMARY_TIDE}",
        f"H_ortho: {H_ORTHO_M:.3f} m",
        f"Datum transformation: +{DATUM_SHIFT_M:.3f} m",
        "",
        "STATISTICAL SCREEN:",
        f"  N >= {MIN_TRACK_N}",
        f"  |median bias| <= {MAX_ABS_CORRECTED_MEDIAN_M:.3f} m",
        f"  RMS <= {MAX_CORRECTED_RMS_M:.3f} m",
        f"  r >= {MIN_TIDE_R}",
        "",
        "PHYSICAL SCREEN:",
        f"  adequate raster coverage >= {MIN_RASTER_COVERAGE:.2f}",
        f"  9° water-dominant >= {WATER_DOMINANT:.2f}",
        f"  A-class 13° water >= {ROBUST_13_WATER:.2f}",
        "",
        f"Statistically selected tracks: {len(track_summaries)}",
        f"Physically accepted A/B tracks: {len(final_summaries)}",
        f"Final observations: {len(final_rows)}",
        f"Daily product days: {len(daily)}",
        "",
        "TRACK CLASSIFICATION",
        "-" * 96,
    ]

    for item in track_summaries:
        lines.append(
            f"PRN {item['sat']:2d} "
            f"{'RISING' if item['rise']==1 else 'SETTING':7s} "
            f"Az={item['az_mean_deg']:7.2f} "
            f"N={item['n']:2d} "
            f"r={item['tide_r']:+.4f} "
            f"RMS={item['corrected_rms_m']*100:.1f} cm "
            f"bias={item['corrected_median_m']*100:+.1f} cm "
            f"C5={item['coverage_5']:.3f} "
            f"C9={item['coverage_9']:.3f} "
            f"C13={item['coverage_13']:.3f} "
            f"W5={item['wet_5']:.3f} "
            f"W9={item['wet_9']:.3f} "
            f"W13={item['wet_13']:.3f} "
            f"{item['geometry_class']}"
        )

    lines += [
        "",
        "FINAL WATER-LEVEL PRODUCT STATUS",
        "-" * 96,
        "Only A/B tracks are included in the final combined product.",
        "C tracks require a larger topobathy raster.",
        "D/E tracks are not physically accepted under the current rules.",
        "",
        "FINAL TRACKS:",
    ]

    for item in final_summaries:
        lines.append(
            f"  PRN {item['sat']} "
            f"{'R' if item['rise']==1 else 'S'} "
            f"Az={item['az_mean_deg']:.1f}° "
            f"class={item['geometry_class']}"
        )

    OUT_SUMMARY.write_text(
        "\n".join(lines)
        + "\n"
    )


# =====================================================================
# MAIN
# =====================================================================

def main():
    print()
    print("=" * 96)
    print(
        "MARCONI FINAL PHYSICAL TRACK SCREEN / EXPERIMENTAL WATER-LEVEL PRODUCT"
    )
    print("=" * 96)
    print(
        f"Datum test: +{DATUM_SHIFT_M:.3f} m"
    )

    if not TOPOBATHY_FILE.exists():
        raise SystemExit(
            f"Missing topobathy file: {TOPOBATHY_FILE}"
        )

    PLOT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    tide_times, tide_values = load_tide()

    rows = attach_tide(
        load_gnssr(),
        tide_times,
        tide_values,
    )

    tracks = build_tracks(rows)

    statistical_groups = []
    statistical_summaries = []

    for group in tracks:
        summary = statistical_summary(group)
        if statistical_select(summary):
            statistical_groups.append(group)
            statistical_summaries.append(summary)

    print(
        "Statistically selected tracks:",
        len(statistical_groups),
    )

    with rasterio.open(
        TOPOBATHY_FILE
    ) as raster:

        to_raster = Transformer.from_crs(
            "EPSG:4326",
            raster.crs,
            always_xy=True,
        )

        station_e, station_n = to_raster.transform(
            STATION_LON,
            STATION_LAT,
        )

        print(
            "Raster CRS:",
            raster.crs,
        )
        print(
            "Station E/N:",
            f"{station_e:.3f}",
            f"{station_n:.3f}",
        )

        geometry_rows = []
        geometry_by_track = defaultdict(list)

        for group in statistical_groups:

            for row in group:

                result = process_observation(
                    row,
                    raster,
                    station_e,
                    station_n,
                )

                result["track_key"] = (
                    row["sat"],
                    row["freq"],
                    row["rise"],
                )

                geometry_rows.append(
                    result
                )

                geometry_by_track[
                    result["track_key"]
                ].append(
                    result
                )

        final_track_summaries = []

        for summary in statistical_summaries:

            key = (
                summary["sat"],
                summary["freq"],
                summary["rise"],
            )

            geom_rows = geometry_by_track[key]

            geometry_summary = summarize_geometry(
                summary,
                geom_rows,
            )

            final_track_summaries.append(
                geometry_summary
            )

        # Final accepted tracks = A or B.
        accepted_tracks = [
            item
            for item in final_track_summaries
            if item["geometry_class"].startswith(
                ("A_", "B_")
            )
        ]

        accepted_keys = {
            (
                item["sat"],
                item["freq"],
                item["rise"],
            )
            for item in accepted_tracks
        }

        final_rows = [
            row
            for row in rows
            if (
                row["sat"],
                row["freq"],
                row["rise"],
            )
            in accepted_keys
        ]

        daily = build_daily_product(
            final_rows
        )

        # Console report.
        print()
        print("=" * 96)
        print(
            "FINAL PHYSICAL TRACK CLASSIFICATION"
        )
        print("=" * 96)

        for item in final_track_summaries:
            print(
                f"PRN {item['sat']:2d} "
                f"{'RISING' if item['rise']==1 else 'SETTING':7s} "
                f"Az={item['az_mean_deg']:7.2f} "
                f"N={item['n']:2d} "
                f"r={item['tide_r']:+.4f} "
                f"RMS={item['corrected_rms_m']*100:.1f} cm "
                f"W5={item['wet_5']:.3f} "
                f"W9={item['wet_9']:.3f} "
                f"W13={item['wet_13']:.3f} "
                f"C9={item['coverage_9']:.3f} "
                f"class={item['geometry_class']}"
            )

        print()
        print(
            "FINAL ACCEPTED TRACKS:",
            len(accepted_tracks),
        )
        print(
            "FINAL OBSERVATIONS:",
            len(final_rows),
        )
        print(
            "FINAL DAILY DAYS:",
            len(daily),
        )

        # CSVs.
        write_observation_csv(
            geometry_rows
        )

        write_track_csv(
            final_track_summaries
        )

        write_daily_csv(
            daily
        )

        write_summary(
            final_track_summaries,
            accepted_tracks,
            final_rows,
            daily,
        )

        # KML only for accepted tracks.
        accepted_groups = [
            group
            for group in statistical_groups
            if (
                group[0]["sat"],
                group[0]["freq"],
                group[0]["rise"],
            ) in accepted_keys
        ]

        write_kml(
            raster,
            station_e,
            station_n,
            accepted_groups,
        )

        # Plots.
        plot_screen(
            final_track_summaries
        )

        plot_final_tracks(
            tide_times,
            tide_values,
            final_rows,
        )

        plot_coverage(
            final_track_summaries
        )

        plot_wet_by_elevation(
            final_track_summaries
        )

        plot_required_extension(
            raster,
            final_track_summaries
        )

    print()
    print("=" * 96)
    print("OUTPUTS")
    print("=" * 96)
    print(
        "All geometry observations:",
        OUT_OBS.resolve(),
    )
    print(
        "Track summary:",
        OUT_TRACKS.resolve(),
    )
    print(
        "Final daily product:",
        OUT_DAILY.resolve(),
    )
    print(
        "Summary:",
        OUT_SUMMARY.resolve(),
    )
    print(
        "KML:",
        OUT_KML.resolve(),
    )
    print(
        "Plots:",
        PLOT_DIR.resolve(),
    )
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
