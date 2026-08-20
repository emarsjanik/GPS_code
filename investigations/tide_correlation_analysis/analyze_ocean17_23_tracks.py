import csv
import math
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timedelta

import numpy as np
from openpyxl import load_workbook

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

DOYS = [204, 205, 206, 207]

MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]


def finite(x):
    try:
        v = float(x)
        return v if math.isfinite(v) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)

    if len(x) < 3:
        return float("nan")

    if np.std(x) == 0 or np.std(y) == 0:
        return float("nan")

    return float(np.corrcoef(x, y)[0, 1])


# ------------------------------------------------------------
# LOAD TIDE MODELS
# ------------------------------------------------------------

wb = load_workbook(
    TIDE_FILE,
    data_only=True
)

ws = wb[wb.sheetnames[0]]

header = [
    cell.value
    for cell in ws[1]
]

time_col = header.index("time")

model_cols = {
    model: header.index(model)
    for model in MODELS
}

tide_times = []
tide_values = {
    model: []
    for model in MODELS
}

for row in ws.iter_rows(
    min_row=2,
    values_only=True
):

    t = row[time_col]

    if not isinstance(t, datetime):
        continue

    vals = {}

    good = True

    for model in MODELS:

        value = finite(
            row[model_cols[model]]
        )

        if value is None:
            good = False
            break

        vals[model] = value

    if not good:
        continue

    tide_times.append(t)

    for model in MODELS:
        tide_values[model].append(
            vals[model]
        )

wb.close()

epoch = np.array([
    (
        t - tide_times[0]
    ).total_seconds()
    for t in tide_times
])


def tide_at(dt, model):

    x = (
        dt - tide_times[0]
    ).total_seconds()

    if (
        x < epoch[0]
        or x > epoch[-1]
    ):
        return float("nan")

    return float(
        np.interp(
            x,
            epoch,
            np.asarray(
                tide_values[model],
                float
            )
        )
    )


# ------------------------------------------------------------
# LOAD GNSS-IR RESULT FILES
# ------------------------------------------------------------

tracks = defaultdict(list)

for doy in DOYS:

    path = (
        RESULT_DIR
        / f"{doy}.txt"
    )

    with open(
        path,
        errors="replace"
    ) as f:

        for line in f:

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            cols = line.split()

            if len(cols) < 17:
                continue

            try:

                year = int(
                    float(cols[0])
                )

                doy2 = int(
                    float(cols[1])
                )

                rh = float(
                    cols[2]
                )

                sat = int(
                    float(cols[3])
                )

                utc_hours = float(
                    cols[4]
                )

                az = float(
                    cols[5]
                )

                amp = float(
                    cols[6]
                )

                emin = float(
                    cols[7]
                )

                emax = float(
                    cols[8]
                )

                freq = int(
                    float(cols[10])
                )

                rise = int(
                    float(cols[11])
                )

                pkn = float(
                    cols[13]
                )

                delt = float(
                    cols[14]
                )

            except Exception:
                continue

            day = (
                datetime(
                    year,
                    1,
                    1
                )
                + timedelta(
                    days=doy2 - 1
                )
            )

            dt = (
                day
                + timedelta(
                    hours=utc_hours
                )
            )

            # This experiment is GPS L1, but verify it.
            if freq != 1:
                continue

            tracks[
                (sat, freq)
            ].append({
                "doy": doy2,
                "dt": dt,
                "rh": rh,
                "az": az,
                "amp": amp,
                "pkn": pkn,
                "emin": emin,
                "emax": emax,
                "rise": rise,
                "delt": delt,
            })


# ------------------------------------------------------------
# TRACK ANALYSIS
# ------------------------------------------------------------

print()
print("=" * 100)
print(
    "MARCONI 17–23 m GPS L1 REPEATED-TRACK TIDE TEST"
)
print("=" * 100)

print(
    f"Tracks found: {len(tracks)}"
)

print()

results = []

for (
    sat,
    freq
), observations in sorted(
    tracks.items()
):

    if len(observations) < 3:
        continue

    observations.sort(
        key=lambda r: r["dt"]
    )

    rh = np.array([
        r["rh"]
        for r in observations
    ])

    az = np.array([
        r["az"]
        for r in observations
    ])

    print()
    print("-" * 100)

    print(
        f"SAT={sat:3d} FREQ={freq:3d}"
        f"  N={len(observations)}"
        f"  Az={np.mean(az):.2f}"
    )

    for r in observations:

        print(
            f"  DOY={r['doy']}"
            f"  {r['dt'].strftime('%Y-%m-%d %H:%M')}"
            f"  RH={r['rh']:.3f}"
            f"  Az={r['az']:.2f}"
            f"  PkN={r['pkn']:.2f}"
            f"  Amp={r['amp']:.1f}"
        )

    for model in MODELS:

        tide = np.array([
            tide_at(
                r["dt"],
                model
            )
            for r in observations
        ])

        valid = np.isfinite(tide)

        if np.sum(valid) < 3:
            continue

        rhv = rh[valid]
        tv = tide[valid]

        r_corr = pearson(
            rhv,
            tv
        )

        slope = float(
            np.polyfit(
                tv,
                rhv,
                1
            )[0]
        )

        # Physically expected:
        #
        # RH + tide = constant
        #
        offset = float(
            np.mean(
                rhv + tv
            )
        )

        residual = (
            rhv
            + tv
            - offset
        )

        rms_cm = (
            float(
                np.sqrt(
                    np.mean(
                        residual ** 2
                    )
                )
            )
            * 100.0
        )

        results.append({
            "sat": sat,
            "freq": freq,
            "n": len(rhv),
            "az_mean": float(
                np.mean(az)
            ),
            "model": model,
            "r": r_corr,
            "slope": slope,
            "offset": offset,
            "rms_cm": rms_cm,
        })

        print(
            f"  {model:18s}"
            f" r={r_corr:+.4f}"
            f" slope={slope:+.4f}"
            f" offset={offset:+.3f} m"
            f" unit-slope RMS={rms_cm:.2f} cm"
        )


# ------------------------------------------------------------
# RANK TRACKS
# ------------------------------------------------------------

print()
print("=" * 100)
print(
    "MOST PROMISING REPEATED TRACKS"
)
print("=" * 100)

# Use EOT20 for ranking; other models will be nearly identical
# for this short interval.
ranked = [
    r
    for r in results
    if r["model"] == "EOT20_heightm"
    and math.isfinite(r["r"])
]

ranked.sort(
    key=lambda r: (
        -abs(r["r"]),
        r["rms_cm"]
    )
)

for r in ranked[:20]:

    print(
        f"SAT={r['sat']:3d}"
        f"  N={r['n']}"
        f"  Az={r['az_mean']:6.2f}"
        f"  r={r['r']:+.4f}"
        f"  slope={r['slope']:+.4f}"
        f"  offset={r['offset']:+.3f}"
        f"  RMS={r['rms_cm']:.2f} cm"
    )


# ------------------------------------------------------------
# CSV
# ------------------------------------------------------------

out = Path(
    "ocean17_23_repeated_track_tide_results.csv"
)

with open(
    out,
    "w",
    newline=""
) as f:

    writer = csv.DictWriter(
        f,
        fieldnames=[
            "sat",
            "freq",
            "n",
            "az_mean",
            "model",
            "r",
            "slope",
            "offset",
            "rms_cm",
        ]
    )

    writer.writeheader()

    writer.writerows(
        results
    )

print()
print(
    f"Results written to: {out}"
)

print()
print(
    "DONE"
)
