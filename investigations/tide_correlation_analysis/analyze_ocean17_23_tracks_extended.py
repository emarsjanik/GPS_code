#!/usr/bin/env python3
"""
analyze_ocean17_23_tracks_extended.py

Extended-duration version of analyze_ocean17_23_tracks.py, using the
full available 22-day dataset (DOY 196-221, ~July 15 - Aug 9, roughly
1.5 spring-neap cycles) instead of just DOY 204-207.

This is the decisive test for the "genuine water tracking vs shared
artifact" question raised in prior analysis: over just 4 days the real
tide is roughly monotonic, so ANY slowly-drifting quantity (real or
artifact) can show a spuriously strong correlation against it. Over 22
days spanning multiple tidal cycles, the tide genuinely oscillates --
a shared, non-physical drift should NOT track an oscillating signal
this well, while genuine water reflection should.

Also directly tests whether correlation strength depends on azimuth
(water-specific) or is uniform across all satellite directions (shared
artifact) -- the same check that falsified the earlier 4-day result.

Usage:
    python3 analyze_ocean17_23_tracks_extended.py
"""

import csv
import math
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timedelta

import numpy as np
from openpyxl import load_workbook

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

# Every day we confirmed has a real result file for this experimental
# config (196-198, 201, 203-221 -- 202 legitimately had no good
# retrievals, matching earlier observed behavior for that day).
DOYS = [196, 197, 198, 201, 203, 204, 205, 206, 207, 208, 209, 210,
        211, 212, 213, 214, 215, 216, 217, 218, 219, 220, 221]

MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]


def finite(x):
    try:
        v = float(x)
        return v if math.isfinite(v) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if len(x) < 3:
        return float("nan")
    if np.std(x) == 0 or np.std(y) == 0:
        return float("nan")
    return float(np.corrcoef(x, y)[0, 1])


# ------------------------------------------------------------
# LOAD TIDE MODELS
# ------------------------------------------------------------

wb = load_workbook(TIDE_FILE, data_only=True)
ws = wb[wb.sheetnames[0]]
header = [cell.value for cell in ws[1]]
time_col = header.index("time")
model_cols = {model: header.index(model) for model in MODELS}

tide_times = []
tide_values = {model: [] for model in MODELS}

for row in ws.iter_rows(min_row=2, values_only=True):
    t = row[time_col]
    if not isinstance(t, datetime):
        continue
    vals = {}
    good = True
    for model in MODELS:
        value = finite(row[model_cols[model]])
        if value is None:
            good = False
            break
        vals[model] = value
    if not good:
        continue
    tide_times.append(t)
    for model in MODELS:
        tide_values[model].append(vals[model])

wb.close()

epoch = np.array([(t - tide_times[0]).total_seconds() for t in tide_times])

print(f"Tide coverage: {tide_times[0]} through {tide_times[-1]}")
print(f"Tide points: {len(tide_times)}")

# Confirm the tide is genuinely non-monotonic over this longer window
# (a real sanity check, not assumed).
ensemble_tide = np.mean([tide_values[m] for m in MODELS], axis=0)
diffs = np.diff(ensemble_tide)
sign_changes = np.sum(np.diff(np.sign(diffs)) != 0)
print(f"Tide sign changes over full period (direction reversals): {sign_changes}")
print("(A monotonic signal would show 0; multiple tidal cycles should show many)")


def tide_at(dt, model):
    x = (dt - tide_times[0]).total_seconds()
    if x < epoch[0] or x > epoch[-1]:
        return float("nan")
    return float(np.interp(x, epoch, np.asarray(tide_values[model], float)))


# ------------------------------------------------------------
# LOAD GNSS-IR RESULT FILES
# ------------------------------------------------------------

tracks = defaultdict(list)
missing_days = []

for doy in DOYS:
    path = RESULT_DIR / f"{doy}.txt"
    if not path.exists():
        missing_days.append(doy)
        continue

    with open(path, errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            cols = line.split()
            if len(cols) < 17:
                continue
            try:
                year = int(float(cols[0]))
                doy2 = int(float(cols[1]))
                rh = float(cols[2])
                sat = int(float(cols[3]))
                utc_hours = float(cols[4])
                az = float(cols[5])
                amp = float(cols[6])
                emin = float(cols[7])
                emax = float(cols[8])
                freq = int(float(cols[10]))
                rise = int(float(cols[11]))
                pkn = float(cols[13])
                delt = float(cols[14])
            except Exception:
                continue

            day = datetime(year, 1, 1) + timedelta(days=doy2 - 1)
            dt = day + timedelta(hours=utc_hours)

            if freq != 1:
                continue

            tracks[(sat, freq)].append({
                "doy": doy2, "dt": dt, "rh": rh, "az": az,
                "amp": amp, "pkn": pkn, "emin": emin, "emax": emax,
                "rise": rise, "delt": delt,
            })

if missing_days:
    print(f"Days with no result file (skipped): {missing_days}")

# ------------------------------------------------------------
# TRACK ANALYSIS
# ------------------------------------------------------------

print()
print("=" * 100)
print("MARCONI 17-23 m GPS L1 EXTENDED (22-DAY) REPEATED-TRACK TIDE TEST")
print("=" * 100)
print(f"Tracks found: {len(tracks)}")
print()

results = []

for (sat, freq), observations in sorted(tracks.items()):
    if len(observations) < 5:  # require more points given the longer baseline
        continue

    observations.sort(key=lambda r: r["dt"])

    rh = np.array([r["rh"] for r in observations])
    az = np.array([r["az"] for r in observations])

    model = "EOT20_heightm"
    tide = np.array([tide_at(r["dt"], model) for r in observations])
    valid = np.isfinite(tide)

    if np.sum(valid) < 5:
        continue

    rhv = rh[valid]
    tv = tide[valid]

    r_corr = pearson(rhv, tv)
    slope = float(np.polyfit(tv, rhv, 1)[0])

    offset = float(np.mean(rhv + tv))
    residual = rhv + tv - offset
    rms_cm = float(np.sqrt(np.mean(residual ** 2))) * 100.0

    results.append({
        "sat": sat, "freq": freq, "n": len(rhv),
        "az_mean": float(np.mean(az)), "az_std": float(np.std(az)),
        "r": r_corr, "slope": slope, "offset": offset, "rms_cm": rms_cm,
    })

# ------------------------------------------------------------
# RANK AND CHECK AZIMUTH DEPENDENCE
# ------------------------------------------------------------

ranked = [r for r in results if math.isfinite(r["r"])]
ranked.sort(key=lambda r: (-abs(r["r"]), r["rms_cm"]))

print(f"{'sat':>4} {'N':>3} {'az_mean':>8} {'az_std':>7} {'r':>8} {'slope':>8} {'offset':>8} {'RMS_cm':>8}")
for r in ranked:
    print(f"{r['sat']:>4} {r['n']:>3} {r['az_mean']:8.2f} {r['az_std']:7.2f} "
          f"{r['r']:8.4f} {r['slope']:8.4f} {r['offset']:8.3f} {r['rms_cm']:8.2f}")

if len(ranked) >= 4:
    azs = np.array([r["az_mean"] for r in ranked])
    corrs = np.array([abs(r["r"]) for r in ranked])
    az_corr_relationship = float(np.corrcoef(azs, corrs)[0, 1])

    print()
    print("=" * 100)
    print("KEY DIAGNOSTIC: DOES CORRELATION STRENGTH DEPEND ON AZIMUTH?")
    print("=" * 100)
    print(f"corr(azimuth, |r|) = {az_corr_relationship:+.4f}")
    print("(Near zero = uniform pattern across all directions = shared artifact)")
    print("(Strong relationship, peaking near true water bearing ~83 deg = genuine water)")

# ------------------------------------------------------------
# CSV
# ------------------------------------------------------------

out = Path("ocean17_23_extended_repeated_track_tide_results.csv")
with open(out, "w", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=list(results[0].keys()) if results else [])
    if results:
        writer.writeheader()
        writer.writerows(results)

print()
print(f"Results written to: {out}")
print()
print("DONE")
