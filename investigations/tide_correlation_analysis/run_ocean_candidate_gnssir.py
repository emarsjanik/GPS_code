#!/usr/bin/env python3
"""
run_ocean_candidate_gnssir.py

Controlled, NON-PRODUCTION GNSS-IR experiment for Marconi.

It creates an isolated gnssrefl analysis strategy from the CURRENT
$REFL_CODE/input/usgs/usgs.json, changing only:
    reflector-height search: 17-23 m
    elevation:               5-13 deg
    azimuth:                 0-360 deg
    frequency:               GPS L1 only (via gnssir -fr 1)

Then it runs gnssir for DOY 204-207 and parses the resulting per-arc
results files into an auditable CSV.

This deliberately does NOT alter the production JSON.

Outputs:
    products/refl_code/input/usgs/ocean17_23_l1_e5_13/usgs.json
    products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13/*.txt
    ocean17_23_candidate_arcs.csv
    ocean17_23_candidate_tracks.csv
    ocean17_23_candidate_summary.txt

The current repository documents the gnssrefl results columns as:
year, doy, RH, sat, UTCtime, Azim, Amp, eminO, emaxO, NumbOf,
freq, rise, EdotF, PkNoise, DelT, MJD, refr.
"""

from __future__ import annotations

import csv
import json
import math
import os
import shutil
import subprocess
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


YEAR = 2026
DOYS = [204, 205, 206, 207]

STATION = "usgs"

EXTENSION = "ocean17_23_l1_e5_13"

H1 = 17.0
H2 = 23.0
E1 = 5.0
E2 = 13.0

# Candidate discovery: examine the whole azimuth circle.
AZ_LIST = [0.0, 360.0]

# Production-like QC gates used by the current Marconi strategy.
MIN_PKNOISE = 2.8
MIN_AMP = 5.0
MAX_ARC_MIN = 40.0

REFL_CODE = Path(
    os.environ.get(
        "REFL_CODE",
        str(Path.home() / "GNSS/v4.1/products/refl_code"),
    )
)

INPUT_DIR = (
    REFL_CODE
    / "input"
    / STATION
)

SOURCE_JSON = INPUT_DIR / f"{STATION}.json"
TARGET_DIR = INPUT_DIR / EXTENSION
TARGET_JSON = TARGET_DIR / f"{STATION}.json"

RESULT_DIR = (
    REFL_CODE
    / str(YEAR)
    / "results"
    / STATION
    / EXTENSION
)

OUT_ARCS = Path(
    "ocean17_23_candidate_arcs.csv"
)
OUT_TRACKS = Path(
    "ocean17_23_candidate_tracks.csv"
)
OUT_SUMMARY = Path(
    "ocean17_23_candidate_summary.txt"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

TIDE_MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]


def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 3:
        return math.nan

    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(np.corrcoef(x, y)[0, 1])


def ensure_environment():
    os.environ["REFL_CODE"] = str(REFL_CODE)

    # Keep existing EXE / ORBITS if already defined; otherwise gnssrefl
    # will use its usual fallback behavior.
    print(f"REFL_CODE = {os.environ['REFL_CODE']}")


def modify_json_value(data, key, value):
    """
    Modify a top-level JSON key if present.

    Returns True if modified.
    """
    if key in data:
        data[key] = value
        return True
    return False


def build_strategy():
    if not SOURCE_JSON.exists():
        raise SystemExit(
            f"ERROR: current strategy not found:\n  {SOURCE_JSON}"
        )

    with open(SOURCE_JSON) as f:
        data = json.load(f)

    original = json.loads(
        json.dumps(data)
    )

    changed = []

    # These are the current gnssrefl JSON strategy names produced by
    # make_gnssir_input in the Marconi processor.
    for key, value in [
        ("e1", E1),
        ("e2", E2),
        ("h1", H1),
        ("h2", H2),
        ("azlist2", AZ_LIST),
        ("peak2noise", MIN_PKNOISE),
        ("ampl", MIN_AMP),
        ("delTmax", MAX_ARC_MIN),
    ]:
        if modify_json_value(data, key, value):
            changed.append(
                f"{key}={value}"
            )

    # Some installations/configs may use allfreq as a JSON strategy key.
    # We keep the file's existing frequency structure but explicitly call
    # gnssir with -fr 1 below, which is the authoritative override for this
    # experiment.
    if "allfreq" in data:
        data["allfreq"] = False
        changed.append("allfreq=False")

    TARGET_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    with open(TARGET_JSON, "w") as f:
        json.dump(
            data,
            f,
            indent=2,
        )
        f.write("\n")

    print()
    print("=" * 88)
    print("ISOLATED GNSS-IR STRATEGY")
    print("=" * 88)
    print("Source:")
    print(f"  {SOURCE_JSON}")
    print("Target:")
    print(f"  {TARGET_JSON}")
    print()
    print("Changes:")
    for c in changed:
        print(f"  {c}")

    # Verify the key fields actually exist.
    missing = [
        k for k in [
            "e1", "e2", "h1", "h2"
        ]
        if k not in data
    ]

    if missing:
        print()
        print(
            "WARNING: source JSON did not contain expected direct keys:"
        )
        print(" ", missing)
        print(
            "gnssir may still run, but the requested RH/elevation"
            " strategy may not have been applied."
        )

    # Show a compact before/after diff for the science-critical fields.
    print()
    print("Science-critical JSON values:")
    for key in [
        "e1",
        "e2",
        "h1",
        "h2",
        "azlist2",
        "peak2noise",
        "ampl",
        "delTmax",
        "allfreq",
    ]:
        print(
            f"  {key:12s}: "
            f"{original.get(key)!r} -> {data.get(key)!r}"
        )


def run_gnssir():
    RESULT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    for doy in DOYS:
        cmd = [
            "gnssir",
            STATION,
            str(YEAR),
            str(doy),
            "-extension",
            EXTENSION,
            "-fr",
            "1",
            "-nooverwrite",
            "False",
        ]

        print()
        print("=" * 88)
        print(f"RUNNING DOY {doy}")
        print("=" * 88)
        print("$", " ".join(cmd))

        result = subprocess.run(
            cmd,
            text=True,
            capture_output=True,
            check=False,
        )

        print(result.stdout)

        if result.stderr:
            print(result.stderr)

        if result.returncode != 0:
            raise SystemExit(
                f"gnssir failed for DOY {doy} "
                f"with return code {result.returncode}"
            )

        path = RESULT_DIR / f"{doy}.txt"

        if not path.exists():
            raise SystemExit(
                f"gnssir reported success but result file is missing:\n{path}"
            )

        print(
            f"Verified result file: {path}"
        )


def utc_hours_to_datetime(year, doy, utc_hours):
    day = (
        datetime(year, 1, 1)
        + timedelta(days=doy - 1)
    )

    total_seconds = (
        float(utc_hours) * 3600.0
    )

    return day + timedelta(
        seconds=total_seconds
    )


def load_tide_interpolators():
    if not TIDE_FILE.exists():
        raise SystemExit(
            f"ERROR: tide file not found: {TIDE_FILE}"
        )

    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [
        c.value
        for c in ws[1]
    ]

    time_idx = header.index(
        "time"
    )

    model_idx = {
        model: header.index(model)
        for model in TIDE_MODELS
    }

    times = []
    values = {
        model: []
        for model in TIDE_MODELS
    }

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_idx]

        if not isinstance(
            t,
            datetime,
        ):
            continue

        row_values = {}

        valid = True

        for model in TIDE_MODELS:
            value = finite(
                row[model_idx[model]]
            )
            if value is None:
                valid = False
                break
            row_values[model] = value

        if not valid:
            continue

        times.append(t)

        for model in TIDE_MODELS:
            values[model].append(
                row_values[model]
            )

    wb.close()

    epoch = np.array(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    functions = {}

    for model in TIDE_MODELS:
        arr = np.asarray(
            values[model],
            dtype=float,
        )

        def make_interp(arr):
            def interp(dt):
                x = (
                    dt - times[0]
                ).total_seconds()

                if (
                    x < epoch[0]
                    or x > epoch[-1]
                ):
                    return math.nan

                return float(
                    np.interp(
                        x,
                        epoch,
                        arr,
                    )
                )

            return interp

        functions[model] = make_interp(
            arr
        )

    return functions


def parse_result_file(
    path,
    tide_functions,
):
    """
    Parse the documented gnssrefl 4.1.5 result format.

    Columns:
      0 year
      1 doy
      2 RH
      3 sat
      4 UTCtime (hours)
      5 Azim
      6 Amp
      7 eminO
      8 emaxO
      9 NumbOf
     10 freq
     11 rise
     12 EdotF
     13 PkNoise
     14 DelT
     15 MJD
     16 refr
    """

    records = []

    lines = path.read_text(
        errors="replace"
    ).splitlines()

    for line in lines:

        if (
            not line.strip()
            or line.lstrip().startswith("%")
        ):
            continue

        cols = line.split()

        if len(cols) < 17:
            continue

        try:
            year = int(
                float(cols[0])
            )

            doy = int(
                float(cols[1])
            )

            rh = float(
                cols[2]
            )

            sat = int(
                float(cols[3])
            )

            utc_hours = float(
                cols[4]
            )

            az = float(
                cols[5]
            )

            amp = float(
                cols[6]
            )

            emin = float(
                cols[7]
            )

            emax = float(
                cols[8]
            )

            nobs = int(
                float(cols[9])
            )

            freq = int(
                float(cols[10])
            )

            rise = int(
                float(cols[11])
            )

            edot = float(
                cols[12]
            )

            pkn = float(
                cols[13]
            )

            delt = float(
                cols[14]
            )

            mjd = float(
                cols[15]
            )

            refr = float(
                cols[16]
            )

        except (
            ValueError,
            TypeError,
        ):
            continue

        dt = utc_hours_to_datetime(
            year,
            doy,
            utc_hours,
        )

        tide = {
            model: tide_functions[model](dt)
            for model in TIDE_MODELS
        }

        if not all(
            math.isfinite(v)
            for v in tide.values()
        ):
            tide = {
                model: math.nan
                for model in TIDE_MODELS
            }

        # Candidate QC here is deliberately recorded rather than
        # silently changing gnssrefl's own saved result population.
        candidate = (
            H1 <= rh <= H2
            and pkn >= MIN_PKNOISE
            and amp >= MIN_AMP
            and delt <= MAX_ARC_MIN
        )

        records.append(
            {
                "doy": doy,
                "year": year,
                "datetime_utc": dt.isoformat(),
                "sat": sat,
                "freq": freq,
                "RH_m": rh,
                "Azim_deg": az,
                "Amp": amp,
                "PkNoise": pkn,
                "eminO_deg": emin,
                "emaxO_deg": emax,
                "NumbOf": nobs,
                "rise": rise,
                "EdotF": edot,
                "DelT_min": delt,
                "MJD": mjd,
                "refr": refr,
                "candidate": candidate,
                **{
                    f"tide_{model}":
                    tide[model]
                    for model in TIDE_MODELS
                },
            }
        )

    return records


def write_arcs(records):
    if not records:
        raise SystemExit(
            "No parseable GNSS-IR result rows were found."
        )

    fields = list(
        records[0].keys()
    )

    with open(
        OUT_ARCS,
        "w",
        newline="",
    ) as f:
        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )

        writer.writeheader()
        writer.writerows(
            records
        )


def build_track_summary(records):
    candidates = [
        r for r in records
        if r["candidate"]
    ]

    grouped = defaultdict(
        list
    )

    for r in candidates:
        grouped[
            (r["sat"], r["freq"])
        ].append(r)

    rows = []

    for (sat, freq), group in grouped.items():

        rh = np.array(
            [r["RH_m"] for r in group],
            dtype=float,
        )

        az = np.array(
            [r["Azim_deg"] for r in group],
            dtype=float,
        )

        pkn = np.array(
            [r["PkNoise"] for r in group],
            dtype=float,
        )

        amp = np.array(
            [r["Amp"] for r in group],
            dtype=float,
        )

        row = {
            "sat": sat,
            "freq": freq,
            "n": len(group),
            "doys": ",".join(
                str(r["doy"])
                for r in sorted(
                    group,
                    key=lambda x: x["datetime_utc"],
                )
            ),
            "az_mean_deg": float(
                np.mean(az)
            ),
            "az_std_deg": float(
                np.std(az)
            ),
            "RH_mean_m": float(
                np.mean(rh)
            ),
            "RH_std_m": float(
                np.std(rh)
            ),
            "RH_min_m": float(
                np.min(rh)
            ),
            "RH_max_m": float(
                np.max(rh)
            ),
            "PkNoise_mean": float(
                np.mean(pkn)
            ),
            "Amp_mean": float(
                np.mean(amp)
            ),
        }

        for model in TIDE_MODELS:
            tide = np.array(
                [
                    r[f"tide_{model}"]
                    for r in group
                ],
                dtype=float,
            )

            if len(group) >= 3:
                r_corr = pearson(
                    rh,
                    tide,
                )

                if (
                    np.std(tide) > 0
                    and np.std(rh) > 0
                ):
                    slope = float(
                        np.polyfit(
                            tide,
                            rh,
                            1,
                        )[0]
                    )
                else:
                    slope = math.nan

                # Physically constrained slope = -1.
                unit_offset = float(
                    np.mean(
                        rh + tide
                    )
                )

                unit_resid = (
                    rh
                    + tide
                    - unit_offset
                )

                unit_rms_cm = (
                    float(
                        np.sqrt(
                            np.mean(
                                unit_resid ** 2
                            )
                        )
                    )
                    * 100.0
                )

            else:
                r_corr = math.nan
                slope = math.nan
                unit_offset = math.nan
                unit_rms_cm = math.nan

            row[
                f"{model}_r_RH_tide"
            ] = r_corr

            row[
                f"{model}_free_slope"
            ] = slope

            row[
                f"{model}_unit_slope_offset_m"
            ] = unit_offset

            row[
                f"{model}_unit_slope_RMS_cm"
            ] = unit_rms_cm

        rows.append(row)

    rows.sort(
        key=lambda r: (
            -r["n"],
            -r["PkNoise_mean"],
            r["RH_std_m"],
        )
    )

    return rows


def write_track_summary(rows):
    if not rows:
        Path(OUT_TRACKS).write_text("")
        return

    fields = list(
        rows[0].keys()
    )

    with open(
        OUT_TRACKS,
        "w",
        newline="",
    ) as f:
        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )
        writer.writeheader()
        writer.writerows(rows)


def write_summary(
    all_records,
    candidate_records,
    tracks,
):
    with open(
        OUT_SUMMARY,
        "w",
    ) as f:

        f.write(
            "=" * 88
            + "\n"
        )
        f.write(
            "MARCONI 17-23 M GNSS-IR CANDIDATE RECONNAISSANCE\n"
        )
        f.write(
            "=" * 88
            + "\n\n"
        )

        f.write(
            "Strategy:\n"
        )
        f.write(
            f"  RH = {H1:.1f}–{H2:.1f} m\n"
        )
        f.write(
            f"  elevation = {E1:.1f}–{E2:.1f} deg\n"
        )
        f.write(
            "  GPS L1 only\n"
        )
        f.write(
            "  azimuth discovery = 0–360 deg\n"
        )
        f.write(
            f"  PkNoise >= {MIN_PKNOISE}\n"
        )
        f.write(
            f"  Amp >= {MIN_AMP}\n"
        )
        f.write(
            f"  DelT <= {MAX_ARC_MIN} min\n\n"
        )

        f.write(
            f"All parsed result rows: {len(all_records)}\n"
        )
        f.write(
            f"Candidate rows: {len(candidate_records)}\n"
        )

        track_counts = defaultdict(int)

        for r in candidate_records:
            track_counts[
                (r["sat"], r["freq"])
            ] += 1

        f.write(
            f"Repeated candidate tracks (n>=2): "
            f"{sum(n >= 2 for n in track_counts.values())}\n\n"
        )

        f.write(
            "TOP REPEATED TRACKS\n"
        )

        for row in tracks[:20]:
            if row["n"] < 2:
                continue

            f.write(
                f"\nSAT {row['sat']:3d} "
                f"FREQ {row['freq']:3d} "
                f"n={row['n']} "
                f"Az={row['az_mean_deg']:.2f} "
                f"RH={row['RH_mean_m']:.3f} "
                f"RHstd={row['RH_std_m']:.3f} "
                f"PkN={row['PkNoise_mean']:.2f} "
                f"Amp={row['Amp_mean']:.1f}\n"
            )

            for model in TIDE_MODELS:
                f.write(
                    f"  {model}: "
                    f"r={row[f'{model}_r_RH_tide']:+.4f} "
                    f"slope={row[f'{model}_free_slope']:+.4f} "
                    f"unitRMS={row[f'{model}_unit_slope_RMS_cm']:.2f} cm\n"
                )


def main():
    ensure_environment()

    print()
    print("=" * 88)
    print(
        "MARCONI 17-23 M GNSS-IR CANDIDATE EXTRACTION"
    )
    print("=" * 88)

    print(
        f"DOYs: {DOYS}"
    )
    print(
        f"REFL_CODE: {REFL_CODE}"
    )
    print(
        f"Current JSON: {SOURCE_JSON}"
    )
    print(
        f"Target extension: {EXTENSION}"
    )

    build_strategy()

    run_gnssir()

    tide_functions = load_tide_interpolators()

    all_records = []

    for doy in DOYS:
        path = (
            RESULT_DIR
            / f"{doy}.txt"
        )

        records = parse_result_file(
            path,
            tide_functions,
        )

        print()
        print(
            f"DOY {doy}: "
            f"{len(records)} parsed result rows"
        )

        candidates = [
            r for r in records
            if r["candidate"]
        ]

        print(
            f"DOY {doy}: "
            f"{len(candidates)} candidates "
            f"within {H1}-{H2} m and QC"
        )

        all_records.extend(
            records
        )

    candidate_records = [
        r for r in all_records
        if r["candidate"]
    ]

    write_arcs(
        candidate_records
    )

    tracks = build_track_summary(
        candidate_records
    )

    write_track_summary(
        tracks
    )

    write_summary(
        all_records,
        candidate_records,
        tracks,
    )

    print()
    print("=" * 88)
    print("TOP REPEATED CANDIDATE TRACKS")
    print("=" * 88)

    shown = 0

    for row in tracks:

        if row["n"] < 2:
            continue

        print(
            f"SAT={row['sat']:3d} "
            f"FREQ={row['freq']:3d} "
            f"N={row['n']:2d} "
            f"Az={row['az_mean_deg']:6.2f} "
            f"RH={row['RH_mean_m']:7.3f} "
            f"RHstd={row['RH_std_m']:6.3f} "
            f"PkN={row['PkNoise_mean']:5.2f} "
            f"Amp={row['Amp_mean']:6.1f}"
        )

        shown += 1

        if shown >= 20:
            break

    print()
    print(
        "Outputs:"
    )
    print(
        f"  {OUT_ARCS}"
    )
    print(
        f"  {OUT_TRACKS}"
    )
    print(
        f"  {OUT_SUMMARY}"
    )

    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
