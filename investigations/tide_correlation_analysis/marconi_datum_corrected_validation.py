#!/usr/bin/env python3
"""
Marconi datum-corrected GNSS-R validation.

Purpose
-------
Compare the uncorrected and hypothesis-corrected GNSS-R water levels
against the EOT20 tide model.

Uncorrected:
    WL_raw = H_ortho - RH

Hypothesis-corrected:
    WL_corr = H_ortho - RH + DATUM_SHIFT_M

The +0.242 m shift is an INDEPENDENT TEST VALUE / HYPOTHESIS.
It is NOT fitted from the GNSS-R observations in this script.

This script reports how much the correction changes:
  * population mean/median offset
  * MAE / RMS / robust MAD
  * correlation
  * free regression slope
  * persistent-track offsets
  * persistent-track RMS

It also makes direct overlay plots:

  01_raw_vs_corrected_overlay.png
  02_raw_and_corrected_vs_eot20.png
  03_residual_histogram.png
  04_corrected_residual_vs_tide.png
  05_track_offset_comparison.png

The main scientific question is:
    Does an independently justified datum shift remove the dominant
    ~20-30 cm vertical separation without changing the tidal timing?

It does NOT change the GNSS-IR processing, RH values, or production JSON.
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

OUT_CSV = Path(
    "marconi_datum_corrected_validation.csv"
)

OUT_SUMMARY = Path(
    "marconi_datum_corrected_validation_summary.txt"
)

OUT_DIR = Path(
    "marconi_datum_corrected_plots"
)

H_ORTHO_M = 18.665

PRIMARY_MODEL = "EOT20_heightm"

# ---------------------------------------------------------------------
# INDEPENDENT TEST VALUE
# ---------------------------------------------------------------------
#
# This is NOT fitted to GNSS-R.
#
# It is the regional datum hypothesis discussed previously:
# local mean-water reference approximately -0.242 m NAVD88.
#
# We apply the inverse +0.242 m to put the GNSS-R NAVD88 elevation
# approximately into that mean-water reference frame.
#
# Treat this as a hypothesis test, NOT as a final datum definition.
#
DATUM_SHIFT_M = +0.242


AZ_TOL_DEG = 3.0
MIN_TRACK_N = 4


# ---------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def utc_datetime(year, doy, utc_hours):
    return (
        datetime(
            year,
            1,
            1,
        )
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


def circular_az_diff(a, b):
    d = abs(a - b) % 360.0
    return min(d, 360.0 - d)


def pearson(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 3:
        return math.nan

    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(
        np.corrcoef(x, y)[0, 1]
    )


def rms(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(
        np.sqrt(
            np.mean(
                x ** 2
            )
        )
    )


def mae(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(
        np.mean(
            np.abs(x)
        )
    )


def robust_mad(x):
    x = np.asarray(x, dtype=float)

    if len(x) == 0:
        return math.nan

    med = np.median(x)

    return float(
        1.4826
        * np.median(
            np.abs(
                x - med
            )
        )
    )


# ---------------------------------------------------------------------
# TIDE
# ---------------------------------------------------------------------

def load_eot20():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [
        c.value
        for c in ws[1]
    ]

    time_col = header.index("time")
    model_col = header.index(PRIMARY_MODEL)

    times = []
    values = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_col]
        v = finite(
            row[model_col]
        )

        if not isinstance(t, datetime):
            continue

        if v is None:
            continue

        times.append(t)
        values.append(v)

    wb.close()

    return (
        np.asarray(
            times,
            dtype="datetime64[ms]",
        ),
        np.asarray(
            values,
            dtype=float,
        ),
    )


def interpolate_tide(
    tide_times,
    tide_values,
    dt,
):
    x = (
        tide_times
        .astype("int64")
    )

    q = (
        np.datetime64(
            dt,
            "ms",
        )
        .astype("int64")
    )

    if (
        q < x[0]
        or q > x[-1]
    ):
        return math.nan

    return float(
        np.interp(
            q,
            x,
            tide_values,
        )
    )


# ---------------------------------------------------------------------
# GNSS-R RESULTS
# ---------------------------------------------------------------------

def load_gnssr():
    rows = []

    for path in sorted(
        RESULT_DIR.glob("*.txt")
    ):
        try:
            int(path.stem)
        except Exception:
            continue

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc = float(c[4])
                az = float(c[5])
                freq = int(float(c[10]))
                rise = int(float(c[11]))
                amp = float(c[6])
                pkn = float(c[13])
                nobs = int(float(c[9]))
            except Exception:
                continue

            if freq != 1:
                continue

            dt = utc_datetime(
                year,
                doy,
                utc,
            )

            wl_raw = (
                H_ORTHO_M
                - rh
            )

            rows.append(
                {
                    "datetime": dt,
                    "doy": doy,
                    "sat": sat,
                    "freq": freq,
                    "rise": rise,
                    "az": az,
                    "RH_m": rh,
                    "WL_raw_m": wl_raw,
                    "WL_corrected_m":
                        wl_raw
                        + DATUM_SHIFT_M,
                    "Amp": amp,
                    "PkNoise": pkn,
                    "NumbOf": nobs,
                }
            )

    rows.sort(
        key=lambda r:
            r["datetime"]
    )

    return rows


def add_tide(rows, tide_times, tide_values):
    for row in rows:
        row[
            "EOT20_m"
        ] = interpolate_tide(
            tide_times,
            tide_values,
            row["datetime"],
        )

        row[
            "raw_residual_m"
        ] = (
            row["WL_raw_m"]
            - row["EOT20_m"]
        )

        row[
            "corrected_residual_m"
        ] = (
            row["WL_corrected_m"]
            - row["EOT20_m"]
        )


# ---------------------------------------------------------------------
# TRACKS
# ---------------------------------------------------------------------

def cluster_tracks(rows):
    base = defaultdict(list)

    for row in rows:
        base[
            (
                row["sat"],
                row["freq"],
                row["rise"],
            )
        ].append(row)

    tracks = []

    for key, group in base.items():

        group = sorted(
            group,
            key=lambda r:
                r["az"],
        )

        current = []
        prev = None

        for row in group:

            if (
                prev is None
                or circular_az_diff(
                    row["az"],
                    prev,
                ) <= AZ_TOL_DEG
            ):
                current.append(row)
            else:
                if len(current) >= MIN_TRACK_N:
                    tracks.append(current)
                current = [row]

            prev = row["az"]

        if len(current) >= MIN_TRACK_N:
            tracks.append(current)

    return tracks


def track_stats(group):
    raw = np.asarray(
        [
            r["raw_residual_m"]
            for r in group
            if math.isfinite(
                r["raw_residual_m"]
            )
        ],
        dtype=float,
    )

    corrected = (
        raw
        + DATUM_SHIFT_M
    )

    tide = np.asarray(
        [
            r["EOT20_m"]
            for r in group
            if math.isfinite(
                r["EOT20_m"]
            )
        ],
        dtype=float,
    )

    wl_raw = np.asarray(
        [
            r["WL_raw_m"]
            for r in group
            if math.isfinite(
                r["EOT20_m"]
            )
        ],
        dtype=float,
    )

    wl_corrected = (
        wl_raw
        + DATUM_SHIFT_M
    )

    if len(raw) >= 3:

        raw_r = pearson(
            wl_raw,
            tide,
        )

        corr_r = pearson(
            wl_corrected,
            tide,
        )

        raw_slope = float(
            np.polyfit(
                tide,
                wl_raw,
                1,
            )[0]
        )

        corr_slope = float(
            np.polyfit(
                tide,
                wl_corrected,
                1,
            )[0]
        )

    else:

        raw_r = math.nan
        corr_r = math.nan
        raw_slope = math.nan
        corr_slope = math.nan

    days = sorted(
        {
            r["doy"]
            for r in group
        }
    )

    azs = np.asarray(
        [
            r["az"]
            for r in group
        ],
        dtype=float,
    )

    return {
        "sat":
            group[0]["sat"],
        "freq":
            group[0]["freq"],
        "rise":
            group[0]["rise"],
        "n":
            len(group),
        "days":
            len(days),
        "az_mean_deg":
            float(
                np.mean(azs)
            ),
        "az_sd_deg":
            float(
                np.std(azs)
            ),
        "raw_median_offset_m":
            float(
                np.median(raw)
            ),
        "corrected_median_offset_m":
            float(
                np.median(
                    corrected
                )
            ),
        "raw_mean_offset_m":
            float(
                np.mean(raw)
            ),
        "corrected_mean_offset_m":
            float(
                np.mean(corrected)
            ),
        "raw_sd_m":
            float(
                np.std(raw)
            ),
        "corrected_sd_m":
            float(
                np.std(corrected)
            ),
        "raw_RMS_m":
            rms(raw),
        "corrected_RMS_m":
            rms(corrected),
        "raw_MAE_m":
            mae(raw),
        "corrected_MAE_m":
            mae(corrected),
        "raw_r":
            raw_r,
        "corrected_r":
            corr_r,
        "raw_slope":
            raw_slope,
        "corrected_slope":
            corr_slope,
    }


# ---------------------------------------------------------------------
# PLOTS
# ---------------------------------------------------------------------

def plot_overlay(
    tide_times,
    tide_values,
    rows,
):
    valid = [
        r
        for r in rows
        if math.isfinite(
            r["EOT20_m"]
        )
    ]

    start = min(
        r["datetime"]
        for r in valid
    )

    end = max(
        r["datetime"]
        for r in valid
    )

    mask = (
        (tide_times >= np.datetime64(start))
        & (tide_times <= np.datetime64(end))
    )

    fig, ax = plt.subplots(
        figsize=(15, 7)
    )

    ax.plot(
        tide_times[mask],
        tide_values[mask],
        linewidth=2.5,
        label="EOT20 tide model",
    )

    ax.plot(
        [
            r["datetime"]
            for r in valid
        ],
        [
            r["WL_raw_m"]
            for r in valid
        ],
        linewidth=1.2,
        marker="o",
        markersize=2.5,
        label="GNSS-R raw (NAVD88)",
    )

    ax.plot(
        [
            r["datetime"]
            for r in valid
        ],
        [
            r["WL_corrected_m"]
            for r in valid
        ],
        linewidth=1.2,
        marker="o",
        markersize=2.5,
        label=(
            f"GNSS-R +{DATUM_SHIFT_M:.3f} m test"
        ),
    )

    ax.set_title(
        "Marconi GNSS-R vs EOT20 — Raw and +0.242 m Datum Test"
    )

    ax.set_xlabel(
        "UTC"
    )

    ax.set_ylabel(
        "Water-level elevation / anomaly (m)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "01_raw_vs_corrected_overlay.png",
        dpi=200,
    )

    plt.close(fig)


def plot_residual_histogram(rows):
    raw = np.asarray(
        [
            100
            * r["raw_residual_m"]
            for r in rows
            if math.isfinite(
                r["raw_residual_m"]
            )
        ]
    )

    corr = raw + (
        DATUM_SHIFT_M * 100
    )

    fig, ax = plt.subplots(
        figsize=(10, 6)
    )

    ax.hist(
        raw,
        bins=30,
        alpha=0.55,
        label="Raw GNSS-R − EOT20",
    )

    ax.hist(
        corr,
        bins=30,
        alpha=0.55,
        label=(
            f"Corrected (+{DATUM_SHIFT_M:.3f} m)"
        ),
    )

    ax.axvline(
        0,
        linewidth=1.5,
    )

    ax.set_title(
        "GNSS-R − EOT20 Residual Distribution"
    )

    ax.set_xlabel(
        "Residual (cm)"
    )

    ax.set_ylabel(
        "Count"
    )

    ax.grid(
        True,
        axis="y",
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "03_residual_histogram.png",
        dpi=200,
    )

    plt.close(fig)


def plot_corrected_residual_vs_tide(rows):
    x = np.asarray(
        [
            r["EOT20_m"]
            for r in rows
            if math.isfinite(
                r["EOT20_m"]
            )
        ],
        dtype=float,
    )

    raw_y = np.asarray(
        [
            r["raw_residual_m"] * 100
            for r in rows
            if math.isfinite(
                r["EOT20_m"]
            )
        ],
        dtype=float,
    )

    corr_y = (
        raw_y
        + DATUM_SHIFT_M * 100
    )

    raw_fit = np.polyfit(
        x,
        raw_y,
        1,
    )

    corr_fit = np.polyfit(
        x,
        corr_y,
        1,
    )

    xx = np.linspace(
        x.min(),
        x.max(),
        200,
    )

    fig, ax = plt.subplots(
        figsize=(9, 7)
    )

    ax.scatter(
        x,
        corr_y,
        s=18,
        alpha=0.5,
        label="Corrected residual",
    )

    ax.plot(
        xx,
        corr_fit[0] * xx + corr_fit[1],
        linewidth=2,
        label=(
            f"corrected trend slope="
            f"{corr_fit[0]:+.3f}"
        ),
    )

    ax.axhline(
        np.median(corr_y),
        linestyle="--",
        linewidth=1.5,
        label=(
            "corrected median="
            f"{np.median(corr_y):+.1f} cm"
        ),
    )

    ax.axhline(
        0,
        linewidth=1.2,
    )

    ax.set_title(
        "Datum-Corrected GNSS-R Residual vs EOT20"
    )

    ax.set_xlabel(
        "EOT20 tide at observation (m)"
    )

    ax.set_ylabel(
        "Corrected GNSS-R − EOT20 (cm)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "04_corrected_residual_vs_tide.png",
        dpi=200,
    )

    plt.close(fig)


def plot_track_comparison(
    stats,
):
    stats = sorted(
        stats,
        key=lambda x:
            x["raw_median_offset_m"],
    )

    labels = [
        (
            f"PRN {s['sat']} "
            f"{'R' if s['rise'] == 1 else 'S'} "
            f"{s['az_mean_deg']:.1f}°"
        )
        for s in stats
    ]

    raw = [
        100
        * s[
            "raw_median_offset_m"
        ]
        for s in stats
    ]

    corrected = [
        100
        * s[
            "corrected_median_offset_m"
        ]
        for s in stats
    ]

    y = np.arange(
        len(labels)
    )

    fig, ax = plt.subplots(
        figsize=(12, 9)
    )

    ax.plot(
        raw,
        y,
        "o-",
        label="Raw",
    )

    ax.plot(
        corrected,
        y,
        "o-",
        label=(
            f"+{DATUM_SHIFT_M:.3f} m test"
        ),
    )

    ax.axvline(
        0,
        linewidth=1.5,
    )

    ax.set_yticks(y)
    ax.set_yticklabels(labels)

    ax.set_xlabel(
        "Median GNSS-R − EOT20 offset (cm)"
    )

    ax.set_title(
        "Persistent Track Offsets: Before and After Datum Test"
    )

    ax.grid(
        True,
        axis="x",
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "05_track_offset_comparison.png",
        dpi=200,
    )

    plt.close(fig)


# ---------------------------------------------------------------------
# OUTPUTS
# ---------------------------------------------------------------------

def write_csv(
    rows,
):
    fields = [
        "datetime",
        "sat",
        "freq",
        "rise",
        "az",
        "RH_m",
        "WL_raw_m",
        "WL_corrected_m",
        "EOT20_m",
        "raw_residual_m",
        "corrected_residual_m",
        "Amp",
        "PkNoise",
        "NumbOf",
    ]

    with open(
        OUT_CSV,
        "w",
        newline="",
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )

        writer.writeheader()

        for row in rows:

            out = {}

            for field in fields:

                value = row.get(
                    field,
                    "",
                )

                if isinstance(
                    value,
                    datetime,
                ):
                    value = value.isoformat()

                out[field] = value

            writer.writerow(out)


def write_summary(
    raw_stats,
    corrected_stats,
    track_stats,
):
    lines = []

    lines.append(
        "MARCONI DATUM-CORRECTED GNSS-R VALIDATION"
    )

    lines.append(
        "=" * 92
    )

    lines.append(
        f"Primary tide model: {PRIMARY_MODEL}"
    )

    lines.append(
        f"H_ortho: {H_ORTHO_M:.3f} m"
    )

    lines.append(
        f"Independent test shift: +{DATUM_SHIFT_M:.3f} m"
    )

    lines.append(
        ""
    )

    lines.append(
        "POPULATION COMPARISON"
    )

    lines.append(
        "-" * 92
    )

    for name, stats in [
        (
            "RAW",
            raw_stats,
        ),
        (
            "DATUM-CORRECTED",
            corrected_stats,
        ),
    ]:

        lines.append(
            name
        )

        lines.append(
            f"  N       = {stats['n']}"
        )

        lines.append(
            f"  mean    = {stats['mean_m']:+.4f} m"
            f" ({stats['mean_m']*100:+.2f} cm)"
        )

        lines.append(
            f"  median  = {stats['median_m']:+.4f} m"
            f" ({stats['median_m']*100:+.2f} cm)"
        )

        lines.append(
            f"  SD      = {stats['std_m']:.4f} m"
        )

        lines.append(
            f"  RMS     = {stats['rms_m']:.4f} m"
            f" ({stats['rms_m']*100:.2f} cm)"
        )

        lines.append(
            f"  MAE     = {stats['mae_m']:.4f} m"
        )

        lines.append(
            f"  MAD     = {stats['MAD_m']:.4f} m"
        )

        lines.append(
            f"  r       = {stats['r']:+.4f}"
        )

        lines.append(
            f"  slope   = {stats['slope']:+.4f}"
        )

        lines.append(
            ""
        )

    lines.append(
        "TRACK COMPARISON"
    )

    lines.append(
        "-" * 92
    )

    for s in track_stats:

        lines.append(
            f"PRN {s['sat']:2d} "
            f"{'RISING' if s['rise']==1 else 'SETTING':7s} "
            f"Az={s['az_mean_deg']:7.2f} "
            f"N={s['n']:2d} "
            f"raw_med={s['raw_median_offset_m']:+.3f} m "
            f"corr_med={s['corrected_median_offset_m']:+.3f} m "
            f"raw_RMS={s['raw_RMS_m']*100:.1f} cm "
            f"corr_RMS={s['corrected_RMS_m']*100:.1f} cm "
            f"raw_r={s['raw_r']:+.4f} "
            f"corr_r={s['corrected_r']:+.4f} "
            f"raw_slope={s['raw_slope']:+.4f} "
            f"corr_slope={s['corrected_slope']:+.4f}"
        )

    lines.append(
        ""
    )

    lines.append(
        "INTERPRETATION"
    )

    lines.append(
        "-" * 92
    )

    lines.append(
        "The +0.242 m value is treated only as an independent hypothesis"
    )

    lines.append(
        "test. It is not fitted to the GNSS-R observations here."
    )

    lines.append(
        "Because a constant vertical shift does not change correlation"
    )

    lines.append(
        "or regression slope, any improvement should appear primarily"
    )

    lines.append(
        "as a reduction in mean/median offset, MAE, and RMS."
    )

    lines.append(
        "If a large residual remains after the shift, that residual"
    )

    lines.append(
        "represents track-dependent reflection bias, measurement noise,"
    )

    lines.append(
        "model error, or an incorrect datum hypothesis."
    )

    OUT_SUMMARY.write_text(
        "\n".join(lines)
        + "\n"
    )


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():
    print()
    print("=" * 92)
    print(
        "MARCONI DATUM-CORRECTED GNSS-R VALIDATION"
    )
    print("=" * 92)

    print(
        f"Primary tide model: {PRIMARY_MODEL}"
    )

    print(
        f"Independent test shift: +{DATUM_SHIFT_M:.3f} m"
    )

    OUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    tide_times, tide_values = load_eot20()
    rows = load_gnssr()

    add_tide(
        rows,
        tide_times,
        tide_values,
    )

    valid = [
        r
        for r in rows
        if math.isfinite(
            r["EOT20_m"]
        )
    ]

    if not valid:
        raise SystemExit(
            "No exact-time comparisons available."
        )

    raw_resid = np.asarray(
        [
            r["raw_residual_m"]
            for r in valid
        ],
        dtype=float,
    )

    corr_resid = (
        raw_resid
        + DATUM_SHIFT_M
    )

    tide = np.asarray(
        [
            r["EOT20_m"]
            for r in valid
        ],
        dtype=float,
    )

    raw_wl = np.asarray(
        [
            r["WL_raw_m"]
            for r in valid
        ],
        dtype=float,
    )

    corr_wl = (
        raw_wl
        + DATUM_SHIFT_M
    )

    raw_stats = {
        "n":
            len(raw_resid),
        "mean_m":
            float(np.mean(raw_resid)),
        "median_m":
            float(np.median(raw_resid)),
        "std_m":
            float(np.std(raw_resid)),
        "rms_m":
            rms(raw_resid),
        "mae_m":
            mae(raw_resid),
        "MAD_m":
            robust_mad(raw_resid),
        "r":
            pearson(
                raw_wl,
                tide,
            ),
        "slope":
            float(
                np.polyfit(
                    tide,
                    raw_wl,
                    1,
                )[0]
            ),
    }

    corrected_stats = {
        "n":
            len(corr_resid),
        "mean_m":
            float(np.mean(corr_resid)),
        "median_m":
            float(np.median(corr_resid)),
        "std_m":
            float(np.std(corr_resid)),
        "rms_m":
            rms(corr_resid),
        "mae_m":
            mae(corr_resid),
        "MAD_m":
            robust_mad(corr_resid),
        "r":
            pearson(
                corr_wl,
                tide,
            ),
        "slope":
            float(
                np.polyfit(
                    tide,
                    corr_wl,
                    1,
                )[0]
            ),
    }

    print()
    print(
        "=" * 92
    )
    print(
        "RAW vs DATUM-CORRECTED"
    )
    print(
        "=" * 92
    )

    for label, s in [
        (
            "RAW",
            raw_stats,
        ),
        (
            "CORRECTED",
            corrected_stats,
        ),
    ]:

        print()
        print(
            label
        )

        print(
            f"  mean   = {s['mean_m']:+.4f} m"
            f" ({s['mean_m']*100:+.2f} cm)"
        )

        print(
            f"  median = {s['median_m']:+.4f} m"
            f" ({s['median_m']*100:+.2f} cm)"
        )

        print(
            f"  RMS    = {s['rms_m']:.4f} m"
            f" ({s['rms_m']*100:.2f} cm)"
        )

        print(
            f"  MAE    = {s['mae_m']:.4f} m"
        )

        print(
            f"  MAD    = {s['MAD_m']:.4f} m"
        )

        print(
            f"  r      = {s['r']:+.4f}"
        )

        print(
            f"  slope  = {s['slope']:+.4f}"
        )

    tracks = cluster_tracks(
        valid
    )

    track_stats = [
        track_stats(
            group
        )
        for group in tracks
    ]

    print()
    print(
        "=" * 92
    )
    print(
        "PERSISTENT TRACK DATUM TEST"
    )
    print(
        "=" * 92
    )

    for s in track_stats:

        print(
            f"PRN {s['sat']:2d} "
            f"{'RISING' if s['rise']==1 else 'SETTING':7s} "
            f"Az={s['az_mean_deg']:7.2f} "
            f"N={s['n']:2d} "
            f"raw_med={s['raw_median_offset_m']:+.3f} "
            f"corr_med={s['corrected_median_offset_m']:+.3f} "
            f"raw_RMS={s['raw_RMS_m']*100:.1f}cm "
            f"corr_RMS={s['corrected_RMS_m']*100:.1f}cm"
        )

    write_csv(valid)
    write_summary(
        raw_stats,
        corrected_stats,
        track_stats,
    )

    print()
    print(
        "Generating plots..."
    )

    plot_overlay(
        tide_times,
        tide_values,
        valid,
    )

    plot_residual_histogram(
        valid
    )

    plot_corrected_residual_vs_tide(
        valid
    )

    if track_stats:
        plot_track_comparison(
            track_stats
        )

    print()
    print(
        "=" * 92
    )
    print(
        "OUTPUTS"
    )
    print(
        "=" * 92
    )

    print(
        "CSV:",
        OUT_CSV.resolve()
    )

    print(
        "Summary:",
        OUT_SUMMARY.resolve()
    )

    print(
        "Plots:",
        OUT_DIR.resolve()
    )

    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
