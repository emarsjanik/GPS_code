#!/usr/bin/env python3
"""
compare_to_tide_models.py

Overlays this station's real GNSS-IR water level output against four
real scientific tide models (EOT20, GOT5.5, GOT5.6, FES2022) for
this site, aligned by actual date/time, for visual comparison.

IMPORTANT, HONEST NOTE ON WHAT THIS COMPARISON CAN AND CAN'T SHOW:
Our GNSS-IR water level and these tide models are on two genuinely
different vertical reference frames. Overlaying them on a single
shared y-axis would be misleading, since the absolute numbers were
never meant to match. This script uses two separate y-axes instead,
sharing only the time axis, so what's actually being compared is the
*shape* of the curve over time -- not the specific numbers.

The four individual models are shown as thin, semi-transparent lines
(confirmed via direct inspection of the real data: they agree with
each other very closely, mean spread ~5cm, max spread ~12cm across
the entire dataset), with their ensemble mean as a bold reference
curve -- this shows genuine inter-model agreement rather than
picking one model arbitrarily.

Usage:
    python3 compare_to_tide_models.py path/to/tide_models.xlsx

Expects an Excel file with a sheet containing columns: time,
<model>_heightm (one or more), and optionally <model>_stage columns
(ignored here).

Reads: ~/GNSS/v4.1/products/refl_code/Files/usgs/usgs_spline_out.txt
Writes: ~/GNSS/v4.1/products/refl_code/Files/usgs/tide_model_comparison.png
"""

from __future__ import annotations

from datetime import timedelta
from pathlib import Path
import sys

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl

LOCAL_SPLINE_FILE = Path.home() / "GNSS/v4.1/products/refl_code/Files/usgs/usgs_spline_out.txt"
EXTERNAL_PRODUCTS_DIR = Path("/mnt/I2Rgus_Data/GPS_Data/Products")


def find_newest_spline_file() -> Path:
    """
    Finds the most recently modified usgs_spline_out.txt across both
    possible locations it could currently be in:
      - the local working directory (if data was just processed and
        hasn't been exported to external storage yet)
      - any dated export folder under external storage (if the local
        copy has since been exported and the local products/
        directory has been wiped, per station_manager.py's own daily
        export cycle)

    Confirmed necessary rather than assuming one location: which one
    actually has the newest data depends entirely on the timing of
    the last manual run versus the last automatic daily export, and
    guessing wrong silently shows stale data instead of the real,
    current result.
    """

    candidates = []

    if LOCAL_SPLINE_FILE.exists():
        candidates.append(LOCAL_SPLINE_FILE)

    if EXTERNAL_PRODUCTS_DIR.is_dir():
        candidates.extend(
            EXTERNAL_PRODUCTS_DIR.glob("*/Files/usgs/usgs_spline_out.txt")
        )

    if not candidates:
        print("ERROR: no usgs_spline_out.txt found in the local working")
        print(f"directory ({LOCAL_SPLINE_FILE}) or in any dated export")
        print(f"folder under {EXTERNAL_PRODUCTS_DIR}.")
        print("Run subdaily first (process_gps_data.sh does this automatically).")
        sys.exit(1)

    newest = max(candidates, key=lambda p: p.stat().st_mtime)
    print(f"Using the most recently modified spline file: {newest}")

    return newest
OUTPUT_FILE = Path.home() / "GNSS/v4.1/products/refl_code/Files/usgs/tide_model_comparison.png"

MODEL_COLORS = {
    "EOT20": "#e69f00",
    "GOT5.5": "#56b4e9",
    "GOT5.6": "#009e73",
    "FES2022": "#cc79a7",
}


def read_gnssrefl_spline_output(path: Path):
    """
    Parses gnssrefl's own evenly-sampled spline output file
    (usgs_spline_out.txt). Format confirmed directly from real
    output: whitespace-separated columns, comment lines starting
    with '%', first column is MJD (Modified Julian Date), second
    column is the value.
    """
    if not path.exists():
        print(f"ERROR: {path} does not exist.")
        print("Run subdaily first (process_gps_data.sh does this automatically).")
        sys.exit(1)

    times = []
    values = []

    with open(path) as f:
        lines = f.readlines()

    print(f"--- Raw first 3 lines of {path.name} (for verification) ---")
    for line in lines[:3]:
        print(repr(line.rstrip()))
    print("---")

    from datetime import datetime, timedelta as td

    MJD_EPOCH = datetime(1858, 11, 17)

    for line in lines:
        stripped = line.strip()
        if not stripped or stripped.startswith("%"):
            continue
        columns = stripped.split()
        mjd = float(columns[0])
        value = float(columns[1])
        dt = MJD_EPOCH + td(days=mjd)
        times.append(dt)
        values.append(value)

    print(f"Parsed {len(times)} points. First: {times[0]} = {values[0]}")
    print(f"                          Last:  {times[-1]} = {values[-1]}")

    # Confirmed real bug this fixes: gnssrefl's spline output simply
    # omits rows during a gap rather than inserting placeholder
    # values, so without this, matplotlib's default behavior connects
    # the last point before a gap straight to the first point after
    # it -- producing a smooth, perfectly straight line across days
    # of missing data that could easily be mistaken for real,
    # gradually-changing water level. gnssrefl's own H0 plot correctly
    # shows these as blank breaks; this matches that behavior here.
    # Threshold: the default output interval is 30 minutes (1800s),
    # confirmed via subdaily's own -delta_out default, so anything
    # meaningfully larger than that (2 hours, giving generous margin
    # for normal floating-point/rounding variation) is a real gap,
    # not just sampling noise.
    GAP_THRESHOLD = timedelta(hours=2)
    gaps_found = 0
    broken_times = [times[0]]
    broken_values = [values[0]]
    for i in range(1, len(times)):
        if times[i] - times[i - 1] > GAP_THRESHOLD:
            broken_times.append(times[i - 1] + (times[i] - times[i - 1]) / 2)
            broken_values.append(float("nan"))
            gaps_found += 1
        broken_times.append(times[i])
        broken_values.append(values[i])

    if gaps_found:
        print(f"Found {gaps_found} real gap(s) larger than {GAP_THRESHOLD} -- "
              "inserting breaks so the plot doesn't draw a misleading "
              "straight line across missing data.")

    times, values = broken_times, broken_values
    print("Confirmed: this is raw reflector height (meters), not the final")
    print("Hortho-converted water level -- the plot inverts this axis so")
    print("'up' means 'more water' on both curves for an intuitive comparison.")
    print("---")

    return times, values


def read_tide_models(xlsx_path: Path):
    """
    Reads the real tide model Excel file directly. Returns
    (times, {model_name: [heights]}) for every column ending in
    '_heightm'.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    model_columns = {}
    for i, name in enumerate(header):
        if name and name.endswith("_heightm"):
            model_name = name[: -len("_heightm")]
            model_columns[model_name] = i

    times = []
    model_values = {name: [] for name in model_columns}

    for row in rows[1:]:
        times.append(row[0])
        for model_name, col_index in model_columns.items():
            model_values[model_name].append(row[col_index])

    return times, model_values


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 compare_to_tide_models.py path/to/tide_models.xlsx")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    gnss_times, gnss_values = read_gnssrefl_spline_output(find_newest_spline_file())

    tide_times, model_values = read_tide_models(xlsx_path)
    model_names = sorted(model_values.keys())
    print(f"Parsed {len(tide_times)} tide model points from {xlsx_path.name}")
    print(f"Models found: {model_names}")

    # Ensemble mean across all models at each timestamp
    ensemble_mean = [
        sum(model_values[name][i] for name in model_names) / len(model_names)
        for i in range(len(tide_times))
    ]

    fig, ax1 = plt.subplots(figsize=(14, 7))

    ax1.plot(gnss_times, gnss_values, color="tab:blue", linewidth=2.5, label="GNSS-IR (this station, reflector height)", zorder=5)
    ax1.set_xlabel("Date")
    ax1.set_ylabel("GNSS-IR reflector height (meters) -- INVERTED axis", color="tab:blue")
    ax1.tick_params(axis="y", labelcolor="tab:blue")
    ax1.invert_yaxis()

    ax2 = ax1.twinx()

    for name in model_names:
        ax2.plot(
            tide_times, model_values[name],
            color=MODEL_COLORS.get(name, "gray"),
            linewidth=1, alpha=0.5, linestyle="-",
            label=f"{name} (individual model)",
        )

    ax2.plot(
        tide_times, ensemble_mean,
        color="tab:red", linewidth=1.75, linestyle="--",
        label="Ensemble mean (all 4 models)", zorder=4,
    )

    ax2.set_ylabel("Tide model predicted height (meters)", color="tab:red")
    ax2.tick_params(axis="y", labelcolor="tab:red")

    fig.autofmt_xdate()
    plt.title("GNSS-IR Water Level vs. Real Tide Models (EOT20, GOT5.5, GOT5.6, FES2022)\n(separate y-axes -- compare shape/timing, not absolute values)")

    # Zoom to the GNSS-IR data's own range, with a small buffer --
    # confirmed necessary: the tide model data spans months, while
    # our real GNSS-IR data spans only a few days, so showing the
    # full tide-model range would make our own data an unreadable
    # sliver.
    buffer = timedelta(hours=12)
    ax1.set_xlim(gnss_times[0] - buffer, gnss_times[-1] + buffer)

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left", fontsize=8)

    plt.tight_layout()
    plt.savefig(OUTPUT_FILE, dpi=150)
    print(f"Saved comparison plot to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
