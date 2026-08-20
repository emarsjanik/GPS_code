#!/usr/bin/env python3
"""
Marconi GNSS-R absolute water-level calibration test.

Goal
----
Convert validated GNSS-IR reflector heights to an absolute-style water-level
proxy using the station orthometric height:

    WL_GNSSR = H_ortho - RH

Then compare against each tide-model signal.

Calibration strategy
--------------------
1. PRN 26 / GPS L1 is the primary calibration track because the previous
   independent analysis showed:
       r ~= -0.997
       slope ~= -0.988
       unit-slope RMS ~= 3.2 cm

2. Estimate ONE constant offset C from PRN 26:

       WL_GNSSR = TIDE + C

   equivalently:

       H_ortho - RH - TIDE = C

3. Apply that SAME C to PRN 21 and PRN 16 without refitting.

4. Report:
       raw GNSS-minus-tide bias
       calibrated bias
       calibrated RMS
       calibrated MAE
       correlation
       free slope
       unit-slope slope diagnostics

5. Also calculate the ensemble of the four tide models.

This is NOT a production calibration. It is a controlled scientific
validation. In particular, it does not tune C separately for each track.

Inputs
------
GNSS-IR:
  products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13/{204..207}.txt

Tide:
  marconi_tides_sherwood.xlsx

H_ortho:
  18.665 m

Outputs
-------
marconi_absolute_wl_calibration.csv
marconi_absolute_wl_calibration_summary.txt
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

OUT_CSV = Path(
    "marconi_absolute_wl_calibration.csv"
)

OUT_SUMMARY = Path(
    "marconi_absolute_wl_calibration_summary.txt"
)

H_ORTHO_M = 18.665

MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]

TARGETS = {
    (26, 1): "PRN26_AZ91_PRIMARY",
    (21, 1): "PRN21_AZ93_VALIDATION",
    (16, 1): "PRN16_AZ96_VALIDATION",
}

DOYS = [204, 205, 206, 207]


# ---------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 3:
        return math.nan

    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(
        np.corrcoef(x, y)[0, 1]
    )


def rms(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(
        np.sqrt(
            np.mean(x ** 2)
        )
    )


def mae(x):
    x = np.asarray(x, dtype=float)
    if len(x) == 0:
        return math.nan
    return float(
        np.mean(
            np.abs(x)
        )
    )


def utc_hours_to_datetime(
    year,
    doy,
    utc_hours,
):
    day = (
        datetime(
            year,
            1,
            1,
        )
        + timedelta(
            days=doy - 1,
        )
    )

    return day + timedelta(
        hours=float(utc_hours)
    )


# ---------------------------------------------------------------------
# TIDE MODEL
# ---------------------------------------------------------------------

def load_tides():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [
        c.value for c in ws[1]
    ]

    time_col = header.index(
        "time"
    )

    model_cols = {
        model:
        header.index(model)
        for model in MODELS
    }

    times = []
    values = {
        model: []
        for model in MODELS
    }

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_col]

        if not isinstance(
            t,
            datetime,
        ):
            continue

        row_values = {}
        good = True

        for model in MODELS:

            value = finite(
                row[
                    model_cols[model]
                ]
            )

            if value is None:
                good = False
                break

            row_values[
                model
            ] = value

        if not good:
            continue

        times.append(t)

        for model in MODELS:
            values[
                model
            ].append(
                row_values[
                    model
                ]
            )

    wb.close()

    epoch = np.array(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    def tide_at(
        dt,
        model,
    ):
        x = (
            dt - times[0]
        ).total_seconds()

        if (
            x < epoch[0]
            or x > epoch[-1]
        ):
            return math.nan

        return float(
            np.interp(
                x,
                epoch,
                np.asarray(
                    values[model],
                    dtype=float,
                ),
            )
        )

    return (
        times,
        tide_at,
        values,
    )


# ---------------------------------------------------------------------
# GNSS-IR
# ---------------------------------------------------------------------

def load_gnss_rows():
    rows = []

    for doy in DOYS:

        path = (
            RESULT_DIR
            / f"{doy}.txt"
        )

        if not path.exists():
            raise SystemExit(
                f"Missing result file: {path}"
            )

        for line in path.read_text(
            errors="replace",
        ).splitlines():

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(
                    float(c[0])
                )

                doy2 = int(
                    float(c[1])
                )

                rh = float(
                    c[2]
                )

                sat = int(
                    float(c[3])
                )

                utc_hours = float(
                    c[4]
                )

                az = float(
                    c[5]
                )

                amp = float(
                    c[6]
                )

                emin = float(
                    c[7]
                )

                emax = float(
                    c[8]
                )

                freq = int(
                    float(c[10])
                )

                rise = int(
                    float(c[11])
                )

                pkn = float(
                    c[13]
                )

                delt = float(
                    c[14]
                )

                refr = float(
                    c[16]
                )

            except Exception:
                continue

            if (
                sat,
                freq,
            ) not in TARGETS:
                continue

            dt = utc_hours_to_datetime(
                year,
                doy2,
                utc_hours,
            )

            wl_gnssr = (
                H_ORTHO_M
                - rh
            )

            rows.append(
                {
                    "year": year,
                    "doy": doy2,
                    "datetime_utc": dt,
                    "sat": sat,
                    "freq": freq,
                    "track": TARGETS[
                        (sat, freq)
                    ],
                    "RH_m": rh,
                    "H_ortho_m":
                        H_ORTHO_M,
                    "GNSS_water_level_m":
                        wl_gnssr,
                    "az_deg": az,
                    "Amp": amp,
                    "PkNoise": pkn,
                    "eminO_deg": emin,
                    "emaxO_deg": emax,
                    "rise": rise,
                    "DelT_min": delt,
                    "refr_model": refr,
                }
            )

    rows.sort(
        key=lambda r: (
            r["sat"],
            r["datetime_utc"],
        )
    )

    return rows


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():

    print()
    print("=" * 96)
    print(
        "MARCONI ABSOLUTE GNSS-R WATER-LEVEL CALIBRATION"
    )
    print("=" * 96)

    print(
        f"Station H_ortho = {H_ORTHO_M:.3f} m"
    )

    rows = load_gnss_rows()

    if not rows:
        raise SystemExit(
            "No target GNSS-IR observations found."
        )

    tide_times, tide_at, _ = load_tides()

    print(
        f"GNSS-R observations: {len(rows)}"
    )

    print(
        f"Tide coverage: "
        f"{tide_times[0]} through "
        f"{tide_times[-1]}"
    )

    # -----------------------------------------------------------------
    # Add tide values to every observation.
    # -----------------------------------------------------------------

    for row in rows:

        for model in MODELS:
            row[
                model
            ] = tide_at(
                row["datetime_utc"],
                model,
            )

    # Ensemble mean tide.
    for row in rows:
        vals = [
            row[model]
            for model in MODELS
            if math.isfinite(
                row[model]
            )
        ]

        row["TIDE_ENSEMBLE_m"] = (
            float(np.mean(vals))
            if vals
            else math.nan
        )

    # -----------------------------------------------------------------
    # PRIMARY CALIBRATION: PRN 26
    #
    # C = mean(GNSS_water - tide)
    #
    # This is fit ONCE per tide model using PRN 26 only.
    # -----------------------------------------------------------------

    primary_rows = [
        r
        for r in rows
        if (
            r["sat"] == 26
            and r["freq"] == 1
        )
    ]

    if len(primary_rows) < 3:
        raise SystemExit(
            "PRN 26 has fewer than 3 observations."
        )

    calibration = {}

    for model in MODELS + [
        "TIDE_ENSEMBLE_m"
    ]:

        diffs = np.array(
            [
                r["GNSS_water_level_m"]
                - r[model]
                for r in primary_rows
                if math.isfinite(
                    r[model]
                )
            ],
            dtype=float,
        )

        if len(diffs) == 0:
            calibration[model] = math.nan
        else:
            calibration[model] = float(
                np.mean(diffs)
            )

    print()
    print(
        "=" * 96
    )
    print(
        "PRIMARY CALIBRATION: PRN 26"
    )
    print(
        "=" * 96
    )

    for model in calibration:
        print(
            f"{model:18s} "
            f"C = {calibration[model]:+.4f} m"
        )

    # -----------------------------------------------------------------
    # Apply the same C to every track.
    # -----------------------------------------------------------------

    for row in rows:

        row["raw_bias_ensemble_m"] = (
            row["GNSS_water_level_m"]
            - row["TIDE_ENSEMBLE_m"]
        )

        row[
            "calibrated_GNSS_ensemble_m"
        ] = (
            row["GNSS_water_level_m"]
            - calibration[
                "TIDE_ENSEMBLE_m"
            ]
        )

        row[
            "calibrated_residual_ensemble_m"
        ] = (
            row[
                "calibrated_GNSS_ensemble_m"
            ]
            - row["TIDE_ENSEMBLE_m"]
        )

        for model in MODELS:

            row[
                f"raw_bias_{model}_m"
            ] = (
                row[
                    "GNSS_water_level_m"
                ]
                - row[model]
            )

            row[
                f"calibrated_residual_{model}_m"
            ] = (
                row[
                    "GNSS_water_level_m"
                ]
                - calibration[model]
                - row[model]
            )

    # -----------------------------------------------------------------
    # Track diagnostics.
    # -----------------------------------------------------------------

    print()
    print(
        "=" * 96
    )
    print(
        "TRACK DIAGNOSTICS — SAME CALIBRATION APPLIED"
    )
    print(
        "=" * 96
    )

    track_stats = []

    for (
        sat,
        freq,
    ), label in TARGETS.items():

        subset = [
            r
            for r in rows
            if (
                r["sat"] == sat
                and r["freq"] == freq
            )
        ]

        print()
        print(
            label
        )

        for r in subset:
            print(
                f"  "
                f"{r['datetime_utc'].strftime('%m-%d %H:%M')} "
                f"RH={r['RH_m']:.3f} "
                f"GNSS_WL={r['GNSS_water_level_m']:.3f} "
                f"TIDE={r['TIDE_ENSEMBLE_m']:+.3f} "
                f"raw_bias={r['raw_bias_ensemble_m']:+.3f} "
                f"cal_resid={r['calibrated_residual_ensemble_m']:+.3f}"
            )

        gnss = np.array(
            [
                r[
                    "GNSS_water_level_m"
                ]
                for r in subset
            ],
            dtype=float,
        )

        tide = np.array(
            [
                r[
                    "TIDE_ENSEMBLE_m"
                ]
                for r in subset
            ],
            dtype=float,
        )

        resid = (
            gnss
            - calibration[
                "TIDE_ENSEMBLE_m"
            ]
            - tide
        )

        if len(subset) >= 3:

            slope = float(
                np.polyfit(
                    tide,
                    gnss,
                    1,
                )[0]
            )

            corr = pearson(
                gnss,
                tide,
            )

        else:

            slope = math.nan
            corr = math.nan

        track_stats.append(
            {
                "sat": sat,
                "freq": freq,
                "track": label,
                "n": len(subset),
                "corr": corr,
                "slope": slope,
                "raw_bias_mean_m":
                    float(
                        np.mean(
                            gnss - tide
                        )
                    ),
                "calibrated_bias_mean_m":
                    float(
                        np.mean(
                            resid
                        )
                    ),
                "calibrated_rms_cm":
                    rms(resid) * 100.0,
                "calibrated_mae_cm":
                    mae(resid) * 100.0,
            }
        )

        print()
        print(
            f"  r(GNSS water, tide) = "
            f"{corr:+.4f}"
        )

        print(
            f"  free slope           = "
            f"{slope:+.4f}"
        )

        print(
            f"  raw bias mean        = "
            f"{np.mean(gnss - tide):+.4f} m"
        )

        print(
            f"  calibrated bias      = "
            f"{np.mean(resid):+.4f} m"
        )

        print(
            f"  calibrated RMS       = "
            f"{rms(resid)*100:.2f} cm"
        )

        print(
            f"  calibrated MAE       = "
            f"{mae(resid)*100:.2f} cm"
        )

    # -----------------------------------------------------------------
    # Detailed per-model fixed-calibration diagnostics.
    # -----------------------------------------------------------------

    print()
    print(
        "=" * 96
    )
    print(
        "PER-MODEL FIXED-CALIBRATION RESULTS"
    )
    print(
        "=" * 96
    )

    model_stats = []

    for (
        sat,
        freq,
    ), label in TARGETS.items():

        subset = [
            r
            for r in rows
            if (
                r["sat"] == sat
                and r["freq"] == freq
            )
        ]

        for model in MODELS:

            gnss = np.array(
                [
                    r["GNSS_water_level_m"]
                    for r in subset
                ],
                dtype=float,
            )

            tide = np.array(
                [
                    r[model]
                    for r in subset
                ],
                dtype=float,
            )

            resid = (
                gnss
                - calibration[model]
                - tide
            )

            slope = float(
                np.polyfit(
                    tide,
                    gnss,
                    1,
                )[0]
            )

            corr = pearson(
                gnss,
                tide,
            )

            model_stats.append(
                {
                    "track":
                        label,
                    "sat":
                        sat,
                    "freq":
                        freq,
                    "model":
                        model,
                    "n":
                        len(subset),
                    "r":
                        corr,
                    "free_slope":
                        slope,
                    "calibration_C_m":
                        calibration[model],
                    "calibrated_rms_cm":
                        rms(resid) * 100.0,
                    "calibrated_mae_cm":
                        mae(resid) * 100.0,
                }
            )

            print(
                f"{label:26s} "
                f"{model:18s} "
                f"r={corr:+.4f} "
                f"slope={slope:+.4f} "
                f"C={calibration[model]:+.4f} "
                f"RMS={rms(resid)*100:.2f} cm"
            )

    # -----------------------------------------------------------------
    # CSV
    # -----------------------------------------------------------------

    fields = [
        "datetime_utc",
        "doy",
        "sat",
        "freq",
        "track",
        "RH_m",
        "H_ortho_m",
        "GNSS_water_level_m",
        "TIDE_ENSEMBLE_m",
        "raw_bias_ensemble_m",
        "calibrated_GNSS_ensemble_m",
        "calibrated_residual_ensemble_m",
        "az_deg",
        "Amp",
        "PkNoise",
        "eminO_deg",
        "emaxO_deg",
        "rise",
        "DelT_min",
        "refr_model",
    ]

    for model in MODELS:
        fields.extend(
            [
                model,
                f"raw_bias_{model}_m",
                f"calibrated_residual_{model}_m",
            ]
        )

    with open(
        OUT_CSV,
        "w",
        newline="",
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )

        writer.writeheader()

        for row in rows:

            output_row = {}

            for field in fields:
                value = row.get(
                    field,
                    math.nan,
                )

                if isinstance(
                    value,
                    datetime,
                ):
                    value = value.isoformat()

                output_row[field] = value

            writer.writerow(
                output_row
            )

    # -----------------------------------------------------------------
    # SUMMARY
    # -----------------------------------------------------------------

    summary = []

    summary.append(
        "MARCONI ABSOLUTE GNSS-R WATER-LEVEL CALIBRATION"
    )
    summary.append(
        "=" * 96
    )
    summary.append(
        f"H_ortho = {H_ORTHO_M:.3f} m"
    )
    summary.append(
        "Primary calibration track: PRN 26 / GPS L1"
    )
    summary.append(
        "Calibration form: GNSS_WL = TIDE + C"
    )
    summary.append(
        ""
    )

    summary.append(
        "PRIMARY CALIBRATION CONSTANTS"
    )
    summary.append(
        "-" * 96
    )

    for model in calibration:
        summary.append(
            f"{model:18s} "
            f"C = {calibration[model]:+.6f} m"
        )

    summary.append(
        ""
    )

    summary.append(
        "TRACK RESULTS — ENSEMBLE TIDE"
    )
    summary.append(
        "-" * 96
    )

    for s in track_stats:
        summary.append(
            f"{s['track']:30s} "
            f"N={s['n']:2d} "
            f"r={s['corr']:+.5f} "
            f"slope={s['slope']:+.5f} "
            f"raw_bias={s['raw_bias_mean_m']:+.4f} m "
            f"cal_RMS={s['calibrated_rms_cm']:.2f} cm "
            f"cal_MAE={s['calibrated_mae_cm']:.2f} cm"
        )

    summary.append(
        ""
    )

    summary.append(
        "PER-MODEL FIXED-CALIBRATION RESULTS"
    )
    summary.append(
        "-" * 96
    )

    for s in model_stats:
        summary.append(
            f"{s['track']:30s} "
            f"{s['model']:18s} "
            f"r={s['r']:+.5f} "
            f"slope={s['free_slope']:+.5f} "
            f"C={s['calibration_C_m']:+.4f} m "
            f"RMS={s['calibrated_rms_cm']:.2f} cm "
            f"MAE={s['calibrated_mae_cm']:.2f} cm"
        )

    summary.append(
        ""
    )

    summary.append(
        "INTERPRETATION"
    )
    summary.append(
        "The calibration constant is estimated ONLY from PRN 26."
    )
    summary.append(
        "The same constant is then applied unchanged to PRN 21 and PRN 16."
    )
    summary.append(
        "This prevents the calibration from being tuned separately to"
    )
    summary.append(
        "each track."
    )
    summary.append(
        ""
    )
    summary.append(
        "The strongest result would be a near-zero calibrated bias and"
    )
    summary.append(
        "small calibrated RMS on PRN 21 and PRN 16 without refitting."
    )

    OUT_SUMMARY.write_text(
        "\n".join(summary)
        + "\n"
    )

    # -----------------------------------------------------------------
    # FINAL CONSOLE SUMMARY
    # -----------------------------------------------------------------

    print()
    print(
        "=" * 96
    )
    print(
        "FINAL ENSEMBLE RESULT"
    )
    print(
        "=" * 96
    )

    for s in track_stats:
        print(
            f"{s['track']:30s} "
            f"r={s['corr']:+.4f} "
            f"slope={s['slope']:+.4f} "
            f"cal_RMS={s['calibrated_rms_cm']:.2f} cm "
            f"cal_MAE={s['calibrated_mae_cm']:.2f} cm"
        )

    print()
    print(
        "Outputs:"
    )
    print(
        f"  {OUT_CSV}"
    )
    print(
        f"  {OUT_SUMMARY}"
    )

    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
