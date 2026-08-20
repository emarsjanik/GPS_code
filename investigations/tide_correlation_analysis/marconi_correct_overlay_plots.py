#!/usr/bin/env python3
"""
Marconi CORRECT overlay plotting.

Primary figure:
  Continuous EOT20 tide model in one color
  GNSS-R estimated water level (H_ortho - RH) in another color

IMPORTANT:
The tide model is plotted from EVERY native workbook time point in the
requested time range. It is NOT sampled only at GNSS-R observation times.

GNSS-R observations are plotted as POINTS only. They are never connected
across different satellite tracks.

Outputs:
  marconi_overlay_plots/
    01_all_gnssr_vs_eot20_raw.png
    02_all_gnssr_vs_eot20_bias_aligned.png
    03_top_tracks_vs_eot20.png
    04_prn29_vs_eot20.png
    05_prn9_vs_eot20.png
    06_prn25_vs_eot20.png

Also writes:
  marconi_overlay_plot_summary.txt

The raw plot is the primary scientific comparison.
The bias-aligned plot subtracts ONE constant median offset from all GNSS-R
points for visual comparison only; it does not redefine the vertical datum.
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

OUT_DIR = Path(
    "marconi_overlay_plots"
)

SUMMARY = Path(
    "marconi_overlay_plot_summary.txt"
)

H_ORTHO_M = 18.665

PRIMARY_MODEL = "EOT20_heightm"

# Full available overlap with the GNSS-R result set.
# The script will automatically clip this to actual available data.
START_DATE = datetime(2026, 7, 15, 0, 0, 0)
END_DATE = datetime(2026, 8, 10, 0, 0, 0)

# The long-term tracks that have emerged as the most interesting.
# Rising/setting is included so separate geometries are not mixed.
TRACKS = [
    (29, 1, 1, 113.01, "PRN 29 rising ~113°"),
    (9, 1, 1, 124.46, "PRN 9 rising ~124°"),
    (9, 1, -1, 35.02, "PRN 9 setting ~35°"),
    (25, 1, -1, 47.61, "PRN 25 setting ~48°"),
    (21, 1, -1, 92.84, "PRN 21 setting ~93°"),
    (26, 1, -1, 90.71, "PRN 26 setting ~91°"),
]

AZ_TOL = 3.0


# ---------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def circular_az_diff(a, b):
    d = abs(a - b) % 360.0
    return min(d, 360.0 - d)


def utc_hours_to_datetime(year, doy, utc_hours):
    return (
        datetime(year, 1, 1)
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


# ---------------------------------------------------------------------
# TIDE WORKBOOK
# ---------------------------------------------------------------------

def load_tide_series():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [c.value for c in ws[1]]

    time_col = header.index("time")
    model_col = header.index(PRIMARY_MODEL)

    times = []
    values = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[time_col]
        v = finite(row[model_col])

        if not isinstance(t, datetime):
            continue

        if v is None:
            continue

        if (
            t < START_DATE
            or t > END_DATE
        ):
            continue

        times.append(t)
        values.append(v)

    wb.close()

    return (
        np.asarray(times, dtype="datetime64[ms]"),
        np.asarray(values, dtype=float),
    )


# ---------------------------------------------------------------------
# GNSS-R RESULTS
# ---------------------------------------------------------------------

def load_gnssr():
    rows = []

    for path in sorted(
        RESULT_DIR.glob("*.txt")
    ):

        try:
            int(path.stem)
        except Exception:
            continue

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc_hours = float(c[4])
                az = float(c[5])
                amp = float(c[6])
                emin = float(c[7])
                emax = float(c[8])
                freq = int(float(c[10]))
                rise = int(float(c[11]))
                pkn = float(c[13])
            except Exception:
                continue

            if freq != 1:
                continue

            dt = utc_hours_to_datetime(
                year,
                doy,
                utc_hours,
            )

            if (
                dt < START_DATE
                or dt > END_DATE
            ):
                continue

            rows.append({
                "datetime": dt,
                "sat": sat,
                "freq": freq,
                "rise": rise,
                "az": az,
                "RH": rh,
                "WL": H_ORTHO_M - rh,
                "Amp": amp,
                "PkNoise": pkn,
                "emin": emin,
                "emax": emax,
            })

    rows.sort(
        key=lambda r:
            r["datetime"]
    )

    return rows


def select_track(
    rows,
    sat,
    freq,
    rise,
    az_center,
):
    return [
        r for r in rows
        if (
            r["sat"] == sat
            and r["freq"] == freq
            and r["rise"] == rise
            and circular_az_diff(
                r["az"],
                az_center,
            ) <= AZ_TOL
        )
    ]


# ---------------------------------------------------------------------
# INTERPOLATE TIDE TO GNSS-R TIMES
# ---------------------------------------------------------------------

def interpolate_tide(
    tide_times,
    tide_values,
    dt,
):
    x = (
        tide_times.astype("int64")
    )

    q = np.datetime64(
        dt,
        "ms",
    ).astype("int64")

    if q < x[0] or q > x[-1]:
        return math.nan

    return float(
        np.interp(
            q,
            x,
            tide_values,
        )
    )


# ---------------------------------------------------------------------
# PLOTS
# ---------------------------------------------------------------------

def make_all_raw_plot(
    tide_times,
    tide_values,
    gnssr,
):
    fig, ax = plt.subplots(
        figsize=(16, 8),
    )

    ax.plot(
        tide_times,
        tide_values,
        linewidth=2.5,
        label="EOT20 tide model",
    )

    ax.scatter(
        [r["datetime"] for r in gnssr],
        [r["WL"] for r in gnssr],
        s=18,
        alpha=0.50,
        label="GNSS-R estimated water level",
    )

    ax.set_title(
        "Marconi GNSS-R Water Level vs EOT20 — Raw Overlay"
    )
    ax.set_xlabel("UTC")
    ax.set_ylabel(
        "Water-level elevation/anomaly (m)"
    )
    ax.grid(
        True,
        alpha=0.25,
    )
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "01_all_gnssr_vs_eot20_raw.png",
        dpi=200,
    )

    plt.close(fig)


def make_bias_aligned_plot(
    tide_times,
    tide_values,
    gnssr,
):
    # Interpolate EOT20 ONLY at GNSS-R observation times to determine
    # one purely visual median offset. The continuous tide curve itself
    # is still plotted from the original workbook values.
    diffs = []

    for r in gnssr:
        t = interpolate_tide(
            tide_times,
            tide_values,
            r["datetime"],
        )

        if math.isfinite(t):
            diffs.append(
                r["WL"] - t
            )

    offset = (
        float(np.median(diffs))
        if diffs
        else 0.0
    )

    aligned = np.asarray(
        [
            r["WL"] - offset
            for r in gnssr
        ],
        dtype=float,
    )

    fig, ax = plt.subplots(
        figsize=(16, 8),
    )

    ax.plot(
        tide_times,
        tide_values,
        linewidth=2.5,
        label="EOT20 tide model",
    )

    ax.scatter(
        [r["datetime"] for r in gnssr],
        aligned,
        s=18,
        alpha=0.50,
        label=(
            "GNSS-R estimated water level "
            f"(median offset removed: {offset:+.3f} m)"
        ),
    )

    ax.set_title(
        "Marconi GNSS-R Water Level vs EOT20 — Bias-Aligned Visualization"
    )
    ax.set_xlabel("UTC")
    ax.set_ylabel(
        "Relative water level (m)"
    )
    ax.grid(
        True,
        alpha=0.25,
    )
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "02_all_gnssr_vs_eot20_bias_aligned.png",
        dpi=200,
    )

    plt.close(fig)

    return offset


def make_top_tracks_plot(
    tide_times,
    tide_values,
    gnssr,
):
    fig, ax = plt.subplots(
        figsize=(16, 8),
    )

    ax.plot(
        tide_times,
        tide_values,
        linewidth=2.8,
        label="EOT20 tide model",
    )

    for sat, freq, rise, az, label in TRACKS:

        group = select_track(
            gnssr,
            sat,
            freq,
            rise,
            az,
        )

        if not group:
            continue

        ax.plot(
            [r["datetime"] for r in group],
            [r["WL"] for r in group],
            marker="o",
            markersize=3.5,
            linewidth=1.2,
            label=label,
        )

    ax.set_title(
        "Marconi Persistent GNSS-R Tracks vs EOT20"
    )
    ax.set_xlabel("UTC")
    ax.set_ylabel(
        "Water-level elevation/anomaly (m)"
    )
    ax.grid(
        True,
        alpha=0.25,
    )
    ax.legend(
        ncol=2,
        fontsize=8,
    )

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "03_top_tracks_vs_eot20.png",
        dpi=200,
    )

    plt.close(fig)


def make_single_track_plot(
    tide_times,
    tide_values,
    group,
    title,
    filename,
):
    fig, ax = plt.subplots(
        figsize=(16, 7),
    )

    ax.plot(
        tide_times,
        tide_values,
        linewidth=2.8,
        label="EOT20 tide model",
    )

    ax.plot(
        [r["datetime"] for r in group],
        [r["WL"] for r in group],
        marker="o",
        markersize=4,
        linewidth=1.4,
        label="GNSS-R estimated water level",
    )

    ax.set_title(title)
    ax.set_xlabel("UTC")
    ax.set_ylabel(
        "Water-level elevation/anomaly (m)"
    )
    ax.grid(
        True,
        alpha=0.25,
    )
    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_DIR / filename,
        dpi=200,
    )

    plt.close(fig)


# ---------------------------------------------------------------------
# SUMMARY
# ---------------------------------------------------------------------

def write_summary(
    tide_times,
    tide_values,
    gnssr,
    offset,
):
    lines = [
        "MARCONI CORRECT GNSS-R / EOT20 OVERLAY PLOTS",
        "=" * 90,
        "",
        "IMPORTANT:",
        "The EOT20 curve is plotted from every tide-workbook time point.",
        "GNSS-R observations are plotted as points and are never connected",
        "across different satellites or reflection tracks.",
        "",
        f"Primary tide model: {PRIMARY_MODEL}",
        f"H_ortho: {H_ORTHO_M:.3f} m",
        f"GNSS-R observations: {len(gnssr)}",
        (
            f"Model coverage plotted: "
            f"{START_DATE} through {END_DATE}"
        ),
        f"Median GNSS-R - EOT20 offset: {offset:+.4f} m",
        "",
        "Files:",
        "  01_all_gnssr_vs_eot20_raw.png",
        "  02_all_gnssr_vs_eot20_bias_aligned.png",
        "  03_top_tracks_vs_eot20.png",
        "  04_prn29_vs_eot20.png",
        "  05_prn9_vs_eot20.png",
        "  06_prn25_vs_eot20.png",
    ]

    SUMMARY.write_text(
        "\n".join(lines)
        + "\n"
    )


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():
    print()
    print("=" * 90)
    print(
        "MARCONI CORRECT GNSS-R / EOT20 OVERLAY PLOTS"
    )
    print("=" * 90)

    OUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    tide_times, tide_values = (
        load_tide_series()
    )

    gnssr = load_gnssr()

    if len(tide_times) == 0:
        raise SystemExit(
            "No tide points loaded."
        )

    if len(gnssr) == 0:
        raise SystemExit(
            "No GNSS-R observations loaded."
        )

    print(
        "Tide points plotted:",
        len(tide_times),
    )

    print(
        "GNSS-R observations plotted:",
        len(gnssr),
    )

    print(
        "Tide model:",
        PRIMARY_MODEL,
    )

    make_all_raw_plot(
        tide_times,
        tide_values,
        gnssr,
    )

    offset = make_bias_aligned_plot(
        tide_times,
        tide_values,
        gnssr,
    )

    make_top_tracks_plot(
        tide_times,
        tide_values,
        gnssr,
    )

    prn29 = select_track(
        gnssr,
        29,
        1,
        1,
        113.01,
    )

    prn9 = (
        select_track(
            gnssr,
            9,
            1,
            1,
            124.46,
        )
        + select_track(
            gnssr,
            9,
            1,
            -1,
            35.02,
        )
    )

    prn25 = select_track(
        gnssr,
        25,
        1,
        -1,
        47.61,
    )

    make_single_track_plot(
        tide_times,
        tide_values,
        prn29,
        "PRN 29 Rising ~113° vs EOT20",
        "04_prn29_vs_eot20.png",
    )

    make_single_track_plot(
        tide_times,
        tide_values,
        prn9,
        "PRN 9 Rising ~124° + Setting ~35° vs EOT20",
        "05_prn9_vs_eot20.png",
    )

    make_single_track_plot(
        tide_times,
        tide_values,
        prn25,
        "PRN 25 Setting ~48° vs EOT20",
        "06_prn25_vs_eot20.png",
    )

    write_summary(
        tide_times,
        tide_values,
        gnssr,
        offset,
    )

    print()
    print(
        "=" * 90
    )
    print("OUTPUTS")
    print("=" * 90)
    print(
        OUT_DIR.resolve()
    )
    print(
        "  01_all_gnssr_vs_eot20_raw.png"
    )
    print(
        "  02_all_gnssr_vs_eot20_bias_aligned.png"
    )
    print(
        "  03_top_tracks_vs_eot20.png"
    )
    print(
        "  04_prn29_vs_eot20.png"
    )
    print(
        "  05_prn9_vs_eot20.png"
    )
    print(
        "  06_prn25_vs_eot20.png"
    )
    print()
    print(
        "Median raw GNSS-R minus EOT20 offset:",
        f"{offset:+.4f} m",
    )
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
