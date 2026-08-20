#!/usr/bin/env python3
"""
plot_sats_15_26_30_vs_tide.py

Direct, unsmoothed comparison of satellites 15, 26, and 30 (freq 1
only -- freq 5/20 already shown to be unstable/ambiguity-prone at
this RH range) from the NReg-fixed ocean_test results, against the
real tide model.

Usage:
    python3 plot_sats_15_26_30_vs_tide.py marconi_tides_sherwood.xlsx
"""

import sys
from pathlib import Path
from datetime import datetime, timedelta

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook


def read_tide_models(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    time_col = header.index("time")
    model_cols = {}
    for i, name in enumerate(header):
        if name and str(name).endswith("_heightm"):
            model_cols[str(name)[: -len("_heightm")]] = i

    times = []
    values = {name: [] for name in model_cols}

    for row in rows[1:]:
        t = row[time_col]
        if not isinstance(t, datetime):
            continue
        row_valid = True
        row_vals = {}
        for name, col in model_cols.items():
            v = row[col]
            if v is None:
                row_valid = False
                break
            row_vals[name] = float(v)
        if not row_valid:
            continue
        times.append(t)
        for name in model_cols:
            values[name].append(row_vals[name])

    return times, values


def build_ensemble_interpolator(times, values):
    ensemble = np.mean(np.array([values[name] for name in values]), axis=0)
    epoch = np.array([(t - times[0]).total_seconds() for t in times], dtype=float)

    def interp(query_dt):
        q = (query_dt - times[0]).total_seconds()
        if q < epoch[0] or q > epoch[-1]:
            return float("nan")
        return float(np.interp(q, epoch, ensemble))

    return interp


def read_ocean_test_results(base_dir, doys, year=2026):
    records = []
    day_start_ref = datetime(year, 1, 1)

    for doy in doys:
        path = base_dir / f"{doy}.txt"
        if not path.exists():
            continue
        with open(path) as f:
            for line in f:
                if line.startswith("%") or not line.strip():
                    continue
                parts = line.split()
                rh = float(parts[2])
                sat = int(parts[3])
                utc_hours = float(parts[4])
                azim = float(parts[5])
                freq = int(parts[10])

                day_start = day_start_ref + timedelta(days=doy - 1)
                solution_dt = day_start + timedelta(hours=utc_hours)

                records.append(
                    {
                        "doy": doy,
                        "rh": rh,
                        "sat": sat,
                        "azim": azim,
                        "freq": freq,
                        "dt": solution_dt,
                    }
                )

    return records


def main():
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} tide_models.xlsx")
        sys.exit(1)

    tide_path = Path(sys.argv[1])

    print("Reading tide models...")
    times, values = read_tide_models(tide_path)
    tide_at = build_ensemble_interpolator(times, values)
    print(f"Parsed {len(times)} tide points, models: {list(values.keys())}")

    base_dir = Path("products/refl_code/2026/results/usgs/ocean_test")
    doys = [204, 205, 206, 207]

    records = read_ocean_test_results(base_dir, doys)

    # Freq 1 only -- freq 5/20 already shown unstable/jumping between
    # ambiguous peaks at this RH range.
    records = [r for r in records if r["freq"] == 1 and r["sat"] in (15, 26, 30)]
    print(f"Loaded {len(records)} freq-1 records for sats 15, 26, 30")

    for r in records:
        r["tide"] = tide_at(r["dt"])

    records = [r for r in records if np.isfinite(r["tide"])]

    fig, axes = plt.subplots(3, 1, figsize=(11, 13), sharex=False)

    sat_labels = {
        15: "Satellite 15 (~93 deg)",
        26: "Satellite 26 (~91 deg)",
        30: "Satellite 30 (~34 deg)",
    }

    for ax, sat in zip(axes, [15, 26, 30]):
        group = sorted([r for r in records if r["sat"] == sat], key=lambda r: r["dt"])

        if not group:
            ax.set_title(f"{sat_labels[sat]} -- no data")
            continue

        dts = [r["dt"] for r in group]
        rhs = [r["rh"] for r in group]
        tides = [r["tide"] for r in group]

        ax1 = ax
        ax1.scatter(dts, rhs, color="tab:blue", s=70, zorder=3, label="GNSS-IR RH (freq 1, individual arcs)")
        ax1.set_ylabel("Reflector height (m)", color="tab:blue")
        ax1.invert_yaxis()
        ax1.tick_params(axis="y", labelcolor="tab:blue")

        for r in group:
            ax1.annotate(
                f"{r['rh']:.2f}m",
                (r["dt"], r["rh"]),
                textcoords="offset points",
                xytext=(5, 5),
                fontsize=8,
                color="tab:blue",
            )

        ax2 = ax1.twinx()
        ax2.plot(dts, tides, "o--", color="tab:red", alpha=0.7, label="Tide model (ensemble mean, sampled at same times)")
        ax2.set_ylabel("Tide model height (m)", color="tab:red")
        ax2.tick_params(axis="y", labelcolor="tab:red")

        ax1.set_title(sat_labels[sat])
        ax1.grid(alpha=0.3)

        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper right", fontsize=8)

        if len(group) >= 2:
            rh_range = max(rhs) - min(rhs)
            tide_range = max(tides) - min(tides)
            print(f"Sat {sat}: RH range = {rh_range:.3f} m, Tide range over same points = {tide_range:.3f} m")

    fig.suptitle(
        "Satellites 15, 26, 30 (freq 1 only, NReg-fixed) vs Real Tide Model\n"
        "days 204-207 -- separate y-axes, compare shape/timing not absolute values",
        fontsize=12,
    )
    fig.autofmt_xdate()
    fig.tight_layout()

    out_path = Path("sats_15_26_30_vs_tide.png")
    fig.savefig(out_path, dpi=150)
    print(f"\nSaved plot to: {out_path}")


if __name__ == "__main__":
    main()
