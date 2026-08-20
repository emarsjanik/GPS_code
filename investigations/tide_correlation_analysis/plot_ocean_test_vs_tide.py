#!/usr/bin/env python3
"""
plot_ocean_test_vs_tide.py

Direct, unsmoothed comparison of the NReg-fixed ocean_test GNSS-IR
results (RH ~11-13m, PRN29, days 204-207) against the real tide
model -- individual points, not a spline, since this dataset is far
too sparse (4 points/day) for smoothing to be meaningful or trustworthy.

Split into the two distinct observed geometries (~33 deg and ~113 deg)
since they are clearly different reflectors and should not be mixed.

Usage:
    python3 plot_ocean_test_vs_tide.py marconi_tides_sherwood.xlsx
"""

import sys
import csv
from pathlib import Path
from datetime import datetime, timedelta

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook


def read_tide_models(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    time_col = header.index("time")
    model_cols = {}
    for i, name in enumerate(header):
        if name and str(name).endswith("_heightm"):
            model_cols[str(name)[: -len("_heightm")]] = i

    times = []
    values = {name: [] for name in model_cols}

    for row in rows[1:]:
        t = row[time_col]
        if not isinstance(t, datetime):
            continue
        row_valid = True
        row_vals = {}
        for name, col in model_cols.items():
            v = row[col]
            if v is None:
                row_valid = False
                break
            row_vals[name] = float(v)
        if not row_valid:
            continue
        times.append(t)
        for name in model_cols:
            values[name].append(row_vals[name])

    return times, values


def build_ensemble_interpolator(times, values):
    ensemble = np.mean(np.array([values[name] for name in values]), axis=0)
    epoch = np.array([(t - times[0]).total_seconds() for t in times], dtype=float)

    def interp(query_dt):
        q = (query_dt - times[0]).total_seconds()
        if q < epoch[0] or q > epoch[-1]:
            return float("nan")
        return float(np.interp(q, epoch, ensemble))

    return interp


def read_ocean_test_results(base_dir, doys, year=2026):
    """
    Reads the raw gnssrefl LSP result columns directly:
    year, doy, RH, sat, UTCtime(hrs), Azim, ...
    """
    records = []
    day_start_ref = datetime(year, 1, 1)

    for doy in doys:
        path = base_dir / f"{doy}.txt"
        if not path.exists():
            continue
        with open(path) as f:
            for line in f:
                if line.startswith("%") or not line.strip():
                    continue
                parts = line.split()
                rh = float(parts[2])
                sat = int(parts[3])
                utc_hours = float(parts[4])
                azim = float(parts[5])
                freq = int(parts[10])

                day_start = day_start_ref + timedelta(days=doy - 1)
                solution_dt = day_start + timedelta(hours=utc_hours)

                records.append(
                    {
                        "doy": doy,
                        "rh": rh,
                        "sat": sat,
                        "azim": azim,
                        "freq": freq,
                        "dt": solution_dt,
                    }
                )

    return records


def main():
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} tide_models.xlsx")
        sys.exit(1)

    tide_path = Path(sys.argv[1])

    print("Reading tide models...")
    times, values = read_tide_models(tide_path)
    tide_at = build_ensemble_interpolator(times, values)
    print(f"Parsed {len(times)} tide points, models: {list(values.keys())}")

    base_dir = Path("products/refl_code/2026/results/usgs/ocean_test")
    doys = [204, 205, 206, 207]

    records = read_ocean_test_results(base_dir, doys)
    print(f"Loaded {len(records)} ocean_test records")

    for r in records:
        r["tide"] = tide_at(r["dt"])

    records = [r for r in records if np.isfinite(r["tide"])]
    print(f"{len(records)} records with valid tide coverage")

    # Split by the two distinct observed geometries.
    az33 = [r for r in records if 25 <= r["azim"] <= 40]
    az113 = [r for r in records if 105 <= r["azim"] <= 120]

    print(f"Azimuth ~33 deg group: {len(az33)} records")
    print(f"Azimuth ~113 deg group: {len(az113)} records")

    fig, axes = plt.subplots(2, 1, figsize=(11, 10), sharex=False)

    for ax, group, label in [
        (axes[0], az33, "Azimuth ~33 deg (RH ~12.8-13.0m)"),
        (axes[1], az113, "Azimuth ~113 deg / PRN29 (RH ~11.2-11.3m)"),
    ]:
        if not group:
            ax.set_title(f"{label} -- no data")
            continue

        group = sorted(group, key=lambda r: r["dt"])

        dts = [r["dt"] for r in group]
        rhs = [r["rh"] for r in group]
        tides = [r["tide"] for r in group]

        ax1 = ax
        ax1.scatter(dts, rhs, color="tab:blue", s=60, zorder=3, label="GNSS-IR RH (individual arcs)")
        ax1.set_ylabel("Reflector height (m)", color="tab:blue")
        ax1.invert_yaxis()
        ax1.tick_params(axis="y", labelcolor="tab:blue")

        for r in group:
            ax1.annotate(
                f"f{r['freq']}",
                (r["dt"], r["rh"]),
                textcoords="offset points",
                xytext=(5, 5),
                fontsize=8,
                color="tab:blue",
            )

        ax2 = ax1.twinx()
        ax2.plot(dts, tides, "o--", color="tab:red", alpha=0.7, label="Tide model (ensemble mean, sampled at same times)")
        ax2.set_ylabel("Tide model height (m)", color="tab:red")
        ax2.tick_params(axis="y", labelcolor="tab:red")

        ax1.set_title(label)
        ax1.grid(alpha=0.3)

        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper right", fontsize=8)

    fig.suptitle(
        "GNSS-IR (NReg-fixed, unsmoothed individual arcs) vs Real Tide Model\n"
        "days 204-207, PRN29 -- separate y-axes, compare shape/timing not absolute values",
        fontsize=12,
    )
    fig.autofmt_xdate()
    fig.tight_layout()

    out_path = Path("ocean_test_vs_tide.png")
    fig.savefig(out_path, dpi=150)
    print(f"\nSaved plot to: {out_path}")

    # Quick numeric summary per group.
    for group, label in [(az33, "~33 deg"), (az113, "~113 deg")]:
        if len(group) < 2:
            continue
        rhs = np.array([r["rh"] for r in group])
        tides = np.array([r["tide"] for r in group])
        print(f"\n{label}: RH range = {rhs.max()-rhs.min():.3f} m, "
              f"Tide range over same points = {tides.max()-tides.min():.3f} m")


if __name__ == "__main__":
    main()
