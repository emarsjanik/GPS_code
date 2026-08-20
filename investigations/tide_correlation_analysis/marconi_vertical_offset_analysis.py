#!/usr/bin/env python3
"""
Marconi GNSS-R vs EOT20 vertical-offset analysis.

Purpose
-------
Quantify the vertical separation between:
    GNSS-R water level = H_ortho - RH
and
    EOT20 tide model evaluated at the EXACT GNSS-R observation time.

This test is specifically designed to investigate whether the apparent
~0.20-0.30 m vertical offset in the overlay can be treated as a largely
constant reference/datum difference.

It does NOT assume that the offset is a datum correction.

It reports:
  * mean GNSS-R - EOT20
  * median GNSS-R - EOT20
  * standard deviation
  * RMS
  * robust MAD
  * linear slope
  * correlation
  * statistics by persistent satellite/reflection track
  * daily median offsets
  * offset trend with tide level
  * offset trend with time

It also makes:
  1. the raw two-line overlay clipped exactly to the GNSS-R interval
  2. GNSS-R minus EOT20 residual time series
  3. residual versus EOT20 tide
  4. per-track offset box/scatter plot

Primary tide model:
    EOT20_heightm

GNSS-R water level:
    WL = 18.665 - RH

Track identity:
    satellite + frequency + rise/setting + azimuth cluster

Outputs:
    marconi_vertical_offset_analysis.csv
    marconi_vertical_offset_summary.txt
    marconi_vertical_offset_plots/
        01_raw_overlap_overlay.png
        02_gnssr_minus_eot20_time.png
        03_gnssr_minus_eot20_vs_tide.png
        04_track_offsets.png
"""

from __future__ import annotations

import csv
import math
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

OUT_CSV = Path(
    "marconi_vertical_offset_analysis.csv"
)

OUT_SUMMARY = Path(
    "marconi_vertical_offset_summary.txt"
)

OUT_DIR = Path(
    "marconi_vertical_offset_plots"
)

H_ORTHO_M = 18.665

PRIMARY_MODEL = "EOT20_heightm"

AZ_TOL_DEG = 3.0

MIN_TRACK_N = 4


# ---------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def utc_datetime(year, doy, utc_hours):
    return (
        datetime(
            year,
            1,
            1,
        )
        + timedelta(
            days=doy - 1,
            hours=float(utc_hours),
        )
    )


def pearson(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    if len(x) < 3:
        return math.nan

    if np.std(x) == 0 or np.std(y) == 0:
        return math.nan

    return float(
        np.corrcoef(x, y)[0, 1]
    )


def robust_mad(x):
    x = np.asarray(x, dtype=float)

    if len(x) == 0:
        return math.nan

    med = np.median(x)

    return float(
        1.4826
        * np.median(
            np.abs(
                x - med
            )
        )
    )


def rms(x):
    x = np.asarray(x, dtype=float)

    if len(x) == 0:
        return math.nan

    return float(
        np.sqrt(
            np.mean(
                x ** 2
            )
        )
    )


def mean_std(x):
    x = np.asarray(x, dtype=float)

    if len(x) == 0:
        return (
            math.nan,
            math.nan,
        )

    return (
        float(np.mean(x)),
        float(np.std(x)),
    )


def circular_az_diff(a, b):
    d = abs(a - b) % 360.0
    return min(
        d,
        360.0 - d,
    )


# ---------------------------------------------------------------------
# TIDE WORKBOOK
# ---------------------------------------------------------------------

def load_tide_series():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [
        c.value
        for c in ws[1]
    ]

    time_col = header.index(
        "time"
    )

    model_col = header.index(
        PRIMARY_MODEL
    )

    times = []
    values = []

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):

        t = row[time_col]
        v = finite(
            row[model_col]
        )

        if not isinstance(
            t,
            datetime,
        ):
            continue

        if v is None:
            continue

        times.append(t)
        values.append(v)

    wb.close()

    return (
        np.asarray(
            times,
            dtype="datetime64[ms]",
        ),
        np.asarray(
            values,
            dtype=float,
        ),
    )


def interpolate_tide(
    tide_times,
    tide_values,
    dt,
):
    x = (
        tide_times
        .astype("int64")
    )

    q = (
        np.datetime64(
            dt,
            "ms",
        )
        .astype("int64")
    )

    if (
        q < x[0]
        or q > x[-1]
    ):
        return math.nan

    return float(
        np.interp(
            q,
            x,
            tide_values,
        )
    )


# ---------------------------------------------------------------------
# GNSS-R
# ---------------------------------------------------------------------

def load_gnssr():
    rows = []

    for path in sorted(
        RESULT_DIR.glob("*.txt")
    ):

        try:
            int(path.stem)
        except Exception:
            continue

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(
                    float(c[0])
                )

                doy = int(
                    float(c[1])
                )

                rh = float(
                    c[2]
                )

                sat = int(
                    float(c[3])
                )

                utc_hours = float(
                    c[4]
                )

                az = float(
                    c[5]
                )

                amp = float(
                    c[6]
                )

                emin = float(
                    c[7]
                )

                emax = float(
                    c[8]
                )

                nobs = int(
                    float(c[9])
                )

                freq = int(
                    float(c[10])
                )

                rise = int(
                    float(c[11])
                )

                pkn = float(
                    c[13]
                )

                delt = float(
                    c[14]
                )

            except Exception:
                continue

            if freq != 1:
                continue

            dt = utc_datetime(
                year,
                doy,
                utc_hours,
            )

            wl = (
                H_ORTHO_M
                - rh
            )

            rows.append(
                {
                    "datetime": dt,
                    "doy": doy,
                    "sat": sat,
                    "freq": freq,
                    "rise": rise,
                    "az": az,
                    "RH_m": rh,
                    "GNSS_WL_m": wl,
                    "Amp": amp,
                    "PkNoise": pkn,
                    "emin": emin,
                    "emax": emax,
                    "NumbOf": nobs,
                    "DelT_min": delt,
                }
            )

    rows.sort(
        key=lambda r:
            r["datetime"]
    )

    return rows


# ---------------------------------------------------------------------
# TRACK CLUSTERING
# ---------------------------------------------------------------------

def cluster_tracks(rows):
    """
    Cluster observations by:
      satellite + frequency + rise/setting + azimuth

    Azimuth clusters are separated when adjacent azimuths differ by more
    than AZ_TOL_DEG.
    """
    base = defaultdict(list)

    for row in rows:
        base[
            (
                row["sat"],
                row["freq"],
                row["rise"],
            )
        ].append(row)

    tracks = []

    for key, group in base.items():

        group = sorted(
            group,
            key=lambda r:
                r["az"],
        )

        current = []
        previous = None

        for row in group:

            if (
                previous is None
                or circular_az_diff(
                    row["az"],
                    previous,
                ) <= AZ_TOL_DEG
            ):
                current.append(row)

            else:

                if len(current) >= MIN_TRACK_N:
                    tracks.append(
                        current
                    )

                current = [
                    row
                ]

            previous = row["az"]

        if len(current) >= MIN_TRACK_N:
            tracks.append(
                current
            )

    tracks.sort(
        key=lambda g: (
            g[0]["sat"],
            g[0]["rise"],
            np.mean(
                [r["az"] for r in g]
            ),
        )
    )

    return tracks


# ---------------------------------------------------------------------
# ANALYSIS
# ---------------------------------------------------------------------

def add_exact_tide_values(
    rows,
    tide_times,
    tide_values,
):
    for row in rows:

        row[
            "EOT20_at_obs_m"
        ] = interpolate_tide(
            tide_times,
            tide_values,
            row["datetime"],
        )

        if math.isfinite(
            row[
                "EOT20_at_obs_m"
            ]
        ):
            row[
                "GNSSR_minus_EOT20_m"
            ] = (
                row["GNSS_WL_m"]
                - row[
                    "EOT20_at_obs_m"
                ]
            )
        else:
            row[
                "GNSSR_minus_EOT20_m"
            ] = math.nan


def analyze_population(
    rows,
):
    valid = [
        r
        for r in rows
        if math.isfinite(
            r[
                "GNSSR_minus_EOT20_m"
            ]
        )
    ]

    residual = np.asarray(
        [
            r[
                "GNSSR_minus_EOT20_m"
            ]
            for r in valid
        ],
        dtype=float,
    )

    tide = np.asarray(
        [
            r[
                "EOT20_at_obs_m"
            ]
            for r in valid
        ],
        dtype=float,
    )

    wl = np.asarray(
        [
            r[
                "GNSS_WL_m"
            ]
            for r in valid
        ],
        dtype=float,
    )

    if len(valid) >= 3:

        corr = pearson(
            wl,
            tide,
        )

        slope = float(
            np.polyfit(
                tide,
                wl,
                1,
            )[0]
        )

    else:

        corr = math.nan
        slope = math.nan

    mean = float(
        np.mean(residual)
    )

    median = float(
        np.median(residual)
    )

    sd = float(
        np.std(residual)
    )

    return {
        "n": len(valid),
        "mean_m": mean,
        "median_m": median,
        "std_m": sd,
        "rms_m": rms(residual),
        "mae_m": float(
            np.mean(
                np.abs(
                    residual
                )
            )
        ),
        "MAD_m": robust_mad(
            residual
        ),
        "r": corr,
        "slope": slope,
    }


def analyze_track(
    group,
    track_id,
):
    valid = [
        r
        for r in group
        if math.isfinite(
            r[
                "GNSSR_minus_EOT20_m"
            ]
        )
    ]

    residual = np.asarray(
        [
            r[
                "GNSSR_minus_EOT20_m"
            ]
            for r in valid
        ],
        dtype=float,
    )

    tide = np.asarray(
        [
            r[
                "EOT20_at_obs_m"
            ]
            for r in valid
        ],
        dtype=float,
    )

    wl = np.asarray(
        [
            r[
                "GNSS_WL_m"
            ]
            for r in valid
        ],
        dtype=float,
    )

    if len(valid) >= 3:

        r = pearson(
            wl,
            tide,
        )

        slope = float(
            np.polyfit(
                tide,
                wl,
                1,
            )[0]
        )

    else:

        r = math.nan
        slope = math.nan

    days = sorted(
        {
            row["doy"]
            for row in group
        }
    )

    azs = np.asarray(
        [
            row["az"]
            for row in group
        ],
        dtype=float,
    )

    return {
        "track_id":
            track_id,

        "sat":
            group[0]["sat"],

        "freq":
            group[0]["freq"],

        "rise":
            group[0]["rise"],

        "n":
            len(valid),

        "days":
            len(days),

        "doy_first":
            min(days),

        "doy_last":
            max(days),

        "az_mean_deg":
            float(
                np.mean(azs)
            ),

        "az_std_deg":
            float(
                np.std(azs)
            ),

        "offset_mean_m":
            float(
                np.mean(residual)
            ),

        "offset_median_m":
            float(
                np.median(residual)
            ),

        "offset_sd_m":
            float(
                np.std(residual)
            ),

        "offset_rms_m":
            rms(residual),

        "offset_MAE_m":
            float(
                np.mean(
                    np.abs(
                        residual
                    )
                )
            ),

        "offset_MAD_m":
            robust_mad(
                residual
            ),

        "tide_r":
            r,

        "tide_slope":
            slope,
    }


# ---------------------------------------------------------------------
# PLOTS
# ---------------------------------------------------------------------

def plot_raw_overlap(
    tide_times,
    tide_values,
    rows,
):
    valid_rows = [
        r
        for r in rows
        if math.isfinite(
            r[
                "GNSSR_minus_EOT20_m"
            ]
        )
    ]

    if not valid_rows:
        return

    start = min(
        r["datetime"]
        for r in valid_rows
    )

    end = max(
        r["datetime"]
        for r in valid_rows
    )

    mask = (
        (
            tide_times.astype(
                "datetime64[ms]"
            )
            >= np.datetime64(
                start,
                "ms",
            )
        )
        &
        (
            tide_times.astype(
                "datetime64[ms]"
            )
            <= np.datetime64(
                end,
                "ms",
            )
        )
    )

    tt = tide_times[
        mask
    ]

    tv = tide_values[
        mask
    ]

    fig, ax = plt.subplots(
        figsize=(15, 7)
    )

    ax.plot(
        tt,
        tv,
        linewidth=2.5,
        label="EOT20 tide model",
    )

    ax.plot(
        [
            r["datetime"]
            for r in valid_rows
        ],
        [
            r["GNSS_WL_m"]
            for r in valid_rows
        ],
        linewidth=1.5,
        marker="o",
        markersize=2.5,
        label="GNSS-R estimated water level",
    )

    ax.set_title(
        "Marconi GNSS-R vs EOT20 — Exact Overlap Interval"
    )

    ax.set_xlabel(
        "UTC"
    )

    ax.set_ylabel(
        "Water-level elevation / anomaly (m)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "01_raw_overlap_overlay.png",
        dpi=200,
    )

    plt.close(fig)


def plot_residual_time(
    rows,
):
    valid = [
        r
        for r in rows
        if math.isfinite(
            r[
                "GNSSR_minus_EOT20_m"
            ]
        )
    ]

    fig, ax = plt.subplots(
        figsize=(15, 6)
    )

    ax.plot(
        [
            r["datetime"]
            for r in valid
        ],
        [
            100
            * r[
                "GNSSR_minus_EOT20_m"
            ]
            for r in valid
        ],
        linewidth=1.1,
        marker="o",
        markersize=2.2,
    )

    ax.axhline(
        0,
        linewidth=1.5,
    )

    ax.set_title(
        "GNSS-R minus EOT20 at Exact Observation Times"
    )

    ax.set_xlabel(
        "UTC"
    )

    ax.set_ylabel(
        "GNSS-R − EOT20 (cm)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "02_gnssr_minus_eot20_time.png",
        dpi=200,
    )

    plt.close(fig)


def plot_residual_vs_tide(
    rows,
):
    valid = [
        r
        for r in rows
        if math.isfinite(
            r[
                "GNSSR_minus_EOT20_m"
            ]
        )
    ]

    x = np.asarray(
        [
            r[
                "EOT20_at_obs_m"
            ]
            for r in valid
        ],
        dtype=float,
    )

    y = np.asarray(
        [
            r[
                "GNSSR_minus_EOT20_m"
            ]
            for r in valid
        ],
        dtype=float,
    )

    coeff = np.polyfit(
        x,
        y,
        1,
    )

    xx = np.linspace(
        np.min(x),
        np.max(x),
        200,
    )

    yy = (
        coeff[0]
        * xx
        + coeff[1]
    )

    fig, ax = plt.subplots(
        figsize=(8, 7)
    )

    ax.scatter(
        x,
        y * 100,
        s=18,
        alpha=0.55,
        label="GNSS-R observations",
    )

    ax.plot(
        xx,
        yy * 100,
        linewidth=2,
        label=(
            f"trend slope="
            f"{coeff[0]:+.3f}"
        ),
    )

    ax.axhline(
        np.median(y) * 100,
        linewidth=1.5,
        linestyle="--",
        label=(
            "median offset="
            f"{np.median(y)*100:+.1f} cm"
        ),
    )

    ax.set_title(
        "GNSS-R − EOT20 vs Tide Level"
    )

    ax.set_xlabel(
        "EOT20 at GNSS-R observation (m)"
    )

    ax.set_ylabel(
        "GNSS-R − EOT20 (cm)"
    )

    ax.grid(
        True,
        alpha=0.25,
    )

    ax.legend()

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "03_gnssr_minus_eot20_vs_tide.png",
        dpi=200,
    )

    plt.close(fig)


def plot_track_offsets(
    track_records,
):
    records = sorted(
        track_records,
        key=lambda r:
            r["offset_median_m"],
    )

    labels = [
        (
            f"PRN {r['sat']} "
            f"{'rise' if r['rise'] == 1 else 'set'} "
            f"{r['az_mean_deg']:.1f}°"
        )
        for r in records
    ]

    values = [
        100
        * r[
            "offset_median_m"
        ]
        for r in records
    ]

    fig, ax = plt.subplots(
        figsize=(12, 8)
    )

    ax.barh(
        labels,
        values,
    )

    ax.axvline(
        0,
        linewidth=1.5,
    )

    ax.set_title(
        "Persistent GNSS-R Track Median Vertical Offset from EOT20"
    )

    ax.set_xlabel(
        "GNSS-R − EOT20 (cm)"
    )

    ax.grid(
        True,
        axis="x",
        alpha=0.25,
    )

    fig.tight_layout()

    fig.savefig(
        OUT_DIR
        / "04_track_offsets.png",
        dpi=200,
    )

    plt.close(fig)


# ---------------------------------------------------------------------
# OUTPUT
# ---------------------------------------------------------------------

def write_csv(
    rows,
):
    fields = [
        "datetime",
        "sat",
        "freq",
        "rise",
        "az",
        "RH_m",
        "GNSS_WL_m",
        "EOT20_at_obs_m",
        "GNSSR_minus_EOT20_m",
        "Amp",
        "PkNoise",
        "emin",
        "emax",
        "NumbOf",
        "DelT_min",
    ]

    with open(
        OUT_CSV,
        "w",
        newline="",
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )

        writer.writeheader()

        for row in rows:

            out = {}

            for field in fields:

                value = row.get(
                    field,
                    "",
                )

                if isinstance(
                    value,
                    datetime,
                ):
                    value = value.isoformat()

                out[field] = value

            writer.writerow(
                out
            )


def write_summary(
    pop,
    track_records,
    rows,
):
    lines = []

    lines.append(
        "MARCONI GNSS-R vs EOT20 VERTICAL OFFSET ANALYSIS"
    )

    lines.append(
        "=" * 92
    )

    lines.append(
        f"Primary tide model: {PRIMARY_MODEL}"
    )

    lines.append(
        f"H_ortho: {H_ORTHO_M:.3f} m"
    )

    lines.append(
        ""
    )

    lines.append(
        "POPULATION RESULT"
    )

    lines.append(
        "-" * 92
    )

    for k, v in pop.items():
        if k in {
            "mean_m",
            "median_m",
            "std_m",
            "rms_m",
            "mae_m",
            "MAD_m",
        }:
            lines.append(
                f"{k:16s}: "
                f"{v:+.4f} m"
                f" ({v*100:+.2f} cm)"
            )
        else:
            lines.append(
                f"{k:16s}: {v}"
            )

    lines.append(
        ""
    )

    lines.append(
        "PERSISTENT TRACK OFFSETS"
    )

    lines.append(
        "-" * 92
    )

    for record in sorted(
        track_records,
        key=lambda r:
            r["offset_median_m"],
    ):

        lines.append(
            f"PRN {record['sat']:2d} "
            f"{'RISING' if record['rise'] == 1 else 'SETTING':7s} "
            f"Az={record['az_mean_deg']:7.2f}±"
            f"{record['az_std_deg']:.2f}° "
            f"N={record['n']:2d} "
            f"days={record['days']:2d} "
            f"median={record['offset_median_m']:+.4f} m "
            f"mean={record['offset_mean_m']:+.4f} m "
            f"SD={record['offset_sd_m']:.4f} m "
            f"RMS={record['offset_rms_m']:.4f} m "
            f"r={record['tide_r']:+.4f} "
            f"slope={record['tide_slope']:+.4f}"
        )

    lines.append(
        ""
    )

    lines.append(
        "DATUM INTERPRETATION"
    )

    lines.append(
        "-" * 92
    )

    lines.append(
        "This analysis quantifies the observed vertical separation."
    )

    lines.append(
        "It does NOT label that separation as a NAVD88-to-MSL datum shift."
    )

    lines.append(
        "A constant offset is consistent with a reference-level difference,"
    )

    lines.append(
        "but reflector-height bias, model reference conventions, local"
    )

    lines.append(
        "mean-water level, and other systematic effects can also contribute."
    )

    lines.append(
        ""
    )

    lines.append(
        "A useful next comparison is the observed population/track offset"
    )

    lines.append(
        "against an independently established Marconi local MSL/MLLW/"
    )

    lines.append(
        "NAVD88 relationship."
    )

    OUT_SUMMARY.write_text(
        "\n".join(lines)
        + "\n"
    )


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():
    print()
    print("=" * 92)
    print(
        "MARCONI GNSS-R / EOT20 VERTICAL OFFSET ANALYSIS"
    )
    print("=" * 92)

    OUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    tide_times, tide_values = (
        load_tide_series()
    )

    rows = load_gnssr()

    add_exact_tide_values(
        rows,
        tide_times,
        tide_values,
    )

    print(
        "GNSS-R observations:",
        len(rows),
    )

    valid = [
        r for r in rows
        if math.isfinite(
            r[
                "GNSSR_minus_EOT20_m"
            ]
        )
    ]

    if not valid:
        raise SystemExit(
            "No observations overlap tide coverage."
        )

    print(
        "Exact-time comparisons:",
        len(valid),
    )

    pop = analyze_population(
        rows
    )

    print()
    print(
        "=" * 92
    )
    print(
        "POPULATION RESULT"
    )
    print(
        "=" * 92
    )

    print(
        f"Mean offset    : {pop['mean_m']:+.4f} m"
        f" ({pop['mean_m']*100:+.2f} cm)"
    )

    print(
        f"Median offset  : {pop['median_m']:+.4f} m"
        f" ({pop['median_m']*100:+.2f} cm)"
    )

    print(
        f"Std deviation  : {pop['std_m']:.4f} m"
    )

    print(
        f"RMS            : {pop['rms_m']:.4f} m"
        f" ({pop['rms_m']*100:.2f} cm)"
    )

    print(
        f"MAE            : {pop['mae_m']:.4f} m"
    )

    print(
        f"Robust MAD     : {pop['MAD_m']:.4f} m"
    )

    print(
        f"Correlation    : {pop['r']:+.4f}"
    )

    print(
        f"Free slope     : {pop['slope']:+.4f}"
    )

    track_groups = cluster_tracks(
        valid
    )

    track_records = []

    for track_id, group in enumerate(
        track_groups,
        start=1,
    ):
        track_records.append(
            analyze_track(
                group,
                track_id,
            )
        )

    track_records.sort(
        key=lambda r:
            r["offset_median_m"],
    )

    print()
    print(
        "=" * 92
    )
    print(
        "PERSISTENT TRACK OFFSETS"
    )
    print(
        "=" * 92
    )

    for record in track_records:

        print(
            f"PRN {record['sat']:2d} "
            f"{'RISING' if record['rise'] == 1 else 'SETTING':7s} "
            f"Az={record['az_mean_deg']:7.2f} "
            f"N={record['n']:2d} "
            f"median={record['offset_median_m']:+.3f} m "
            f"mean={record['offset_mean_m']:+.3f} m "
            f"SD={record['offset_sd_m']:.3f} m "
            f"RMS={record['offset_rms_m']:.3f} m "
            f"r={record['tide_r']:+.4f} "
            f"slope={record['tide_slope']:+.4f}"
        )

    write_csv(
        valid
    )

    write_summary(
        pop,
        track_records,
        valid,
    )

    print()
    print(
        "Generating plots..."
    )

    plot_raw_overlap(
        tide_times,
        tide_values,
        valid,
    )

    plot_residual_time(
        valid
    )

    plot_residual_vs_tide(
        valid
    )

    if track_records:
        plot_track_offsets(
            track_records
        )

    print()
    print(
        "=" * 92
    )
    print(
        "OUTPUTS"
    )
    print(
        "=" * 92
    )

    print(
        "CSV:",
        OUT_CSV.resolve()
    )

    print(
        "Summary:",
        OUT_SUMMARY.resolve()
    )

    print(
        "Plots:",
        OUT_DIR.resolve()
    )

    print()
    print(
        "DONE"
    )


if __name__ == "__main__":
    main()
