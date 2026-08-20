#!/usr/bin/env python3
"""
Marconi tide-datum investigation + vertical sensitivity sweep.

This is a DIAGNOSTIC calibration experiment.

Questions answered:
  1. What vertical CRS/datum metadata are present in the 2021 USGS
     Marconi topobathy GeoTIFF?
  2. What metadata are present in the tide workbook?
  3. What are the four tide-model mean/min/max ranges over the GNSS-R
     observation interval?
  4. How does the 13-degree Fresnel footprint wet/dry classification
     change as an unknown vertical offset is swept from -0.50 to +0.50 m?
  5. For each candidate track, what offset would make the mean 13-degree
     wet fraction reach 80%?

IMPORTANT:
The tide-model columns in the workbook are tidal elevation signals
centered close to zero. Ocean tide model elevations are model tidal
corrections, not automatically NAVD88 absolute water-surface elevations.
Therefore this script NEVER claims that the raw tide column is an
absolute NAVD88 water elevation.

The topobathy raster is documented by USGS as NAD83(2011)/UTM Zone 19N
with NAVD88 vertical control. The USGS Marconi data release describes
the topobathy product as a merged 1-m bathymetry/DEM. The exact vertical
datum metadata are printed from the raster and the release is cited in
the accompanying analysis.

Outputs:
  marconi_datum_sensitivity.csv
  marconi_datum_sensitivity_summary.txt

Target tracks:
  PRN 26 / L1 / ~91°
  PRN 21 / L1 / ~93°
  PRN 16 / L1 / ~96°

The script reuses the exact GNSS-IR result rows from:
  ocean17_23_l1_e5_13
and reconstructs the same Fresnel footprints used in the preceding test.
"""

from __future__ import annotations

import csv
import math
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import rasterio
from rasterio.mask import mask
from pyproj import Transformer
from shapely.geometry import Polygon, mapping
from shapely.ops import transform as shapely_transform
from openpyxl import load_workbook
from gnssrefl.refl_zones import makeEllipse_latlon


# ---------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TOPO_FILE = Path(
    "2021022FA_Marconi_topobathy_1m.tif"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]

TARGETS = {
    (26, 1): "PRN26_AZ91",
    (21, 1): "PRN21_AZ93",
    (16, 1): "PRN16_AZ96",
}

DOYS = [204, 205, 206, 207]

LAT = 41.8928243333
LON = -69.9633227139

# The most diagnostic footprint is the highest elevation tested in
# the GNSS-IR experiment, where footprints are smaller.
ELEVATION_DEG = 13.0

# Vertical sensitivity range requested.
SHIFT_MIN = -0.50
SHIFT_MAX = +0.50
SHIFT_STEP = 0.05

WATER_FRACTION_TARGET = 0.80
NEAR_WATER_MARGIN_M = 0.15

OUT_CSV = Path(
    "marconi_datum_sensitivity.csv"
)

OUT_SUMMARY = Path(
    "marconi_datum_sensitivity_summary.txt"
)


# ---------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def load_tide_data():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [
        c.value for c in ws[1]
    ]

    tcol = header.index("time")
    mcols = {
        m: header.index(m)
        for m in MODELS
    }

    times = []
    values = {
        m: []
        for m in MODELS
    }

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[tcol]
        if not isinstance(t, datetime):
            continue

        tmp = {}
        good = True

        for m in MODELS:
            v = finite(
                row[mcols[m]]
            )
            if v is None:
                good = False
                break
            tmp[m] = v

        if not good:
            continue

        times.append(t)

        for m in MODELS:
            values[m].append(
                tmp[m]
            )

    # Workbook metadata.
    props = wb.properties

    wb.close()

    epoch = np.array(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    def interp(dt, model):
        x = (
            dt - times[0]
        ).total_seconds()

        if (
            x < epoch[0]
            or x > epoch[-1]
        ):
            return math.nan

        return float(
            np.interp(
                x,
                epoch,
                np.asarray(
                    values[model],
                    float,
                ),
            )
        )

    return (
        times,
        values,
        interp,
        props,
    )


def load_target_rows():
    rows = []

    for doy in DOYS:
        path = (
            RESULT_DIR
            / f"{doy}.txt"
        )

        if not path.exists():
            raise RuntimeError(
                f"Missing GNSS-IR result file: {path}"
            )

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy2 = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc_hours = float(c[4])
                az = float(c[5])
                freq = int(float(c[10]))
            except Exception:
                continue

            if (sat, freq) not in TARGETS:
                continue

            dt = (
                datetime(year, 1, 1)
                + timedelta(
                    days=doy2 - 1,
                    hours=utc_hours,
                )
            )

            rows.append(
                {
                    "doy": doy2,
                    "dt": dt,
                    "sat": sat,
                    "freq": freq,
                    "rh": rh,
                    "az": az,
                }
            )

    rows.sort(
        key=lambda r: (
            r["sat"],
            r["dt"],
        )
    )

    return rows


def build_fresnel_polygon(row):
    lon, lat = makeEllipse_latlon(
        row["freq"],
        ELEVATION_DEG,
        row["rh"],
        row["az"],
        LAT,
        LON,
    )

    return Polygon(
        list(
            zip(
                map(float, lon),
                map(float, lat),
            )
        )
    )


def raster_values_for_polygon(src, polygon):
    to_raster = Transformer.from_crs(
        "EPSG:4326",
        src.crs,
        always_xy=True,
    )

    poly_raster = shapely_transform(
        lambda x, y, z=None:
            to_raster.transform(x, y),
        polygon,
    )

    try:
        arr, _ = mask(
            src,
            [mapping(poly_raster)],
            crop=True,
            filled=False,
        )
    except ValueError:
        return np.array(
            [],
            dtype=float,
        )

    values = np.asarray(
        arr[0].compressed(),
        dtype=float,
    )

    return values


def fractions(values, water_elevation):
    if len(values) == 0:
        return (
            math.nan,
            math.nan,
            math.nan,
        )

    wet = (
        values
        <= water_elevation
    )

    near = (
        (values > water_elevation)
        & (
            values
            <= water_elevation
            + NEAR_WATER_MARGIN_M
        )
    )

    dry = (
        values
        > water_elevation
        + NEAR_WATER_MARGIN_M
    )

    return (
        float(np.mean(wet)),
        float(np.mean(near)),
        float(np.mean(dry)),
    )


def build_shift_grid():
    n = int(
        round(
            (
                SHIFT_MAX
                - SHIFT_MIN
            ) / SHIFT_STEP
        )
    )

    return [
        round(
            SHIFT_MIN
            + i * SHIFT_STEP,
            2,
        )
        for i in range(n + 1)
    ]


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():

    print()
    print("=" * 92)
    print(
        "MARCONI TIDE DATUM INVESTIGATION / ±0.50 m SENSITIVITY SWEEP"
    )
    print("=" * 92)

    if not TOPO_FILE.exists():
        raise SystemExit(
            f"Topobathy file not found: {TOPO_FILE}"
        )

    if not TIDE_FILE.exists():
        raise SystemExit(
            f"Tide workbook not found: {TIDE_FILE}"
        )

    tide_times, tide_values, tide_at, props = (
        load_tide_data()
    )

    rows = load_target_rows()

    print()
    print("TOPObathy metadata")
    print("-" * 92)

    with rasterio.open(
        TOPO_FILE
    ) as src:

        print(
            "CRS       :",
            src.crs,
        )
        print(
            "Resolution:",
            src.res,
        )
        print(
            "Bounds    :",
            src.bounds,
        )
        print(
            "Count     :",
            src.count,
        )
        print(
            "Dtype     :",
            src.dtypes,
        )
        print(
            "Nodata    :",
            src.nodata,
        )
        print(
            "Tags      :",
            src.tags(),
        )

        # Collect raw topobathy values for entire raster.
        data = src.read(
            1,
            masked=True,
        )

        vals = np.asarray(
            data.compressed(),
            dtype=float,
        )

        print(
            "Raster z min:",
            float(np.min(vals)),
        )
        print(
            "Raster z max:",
            float(np.max(vals)),
        )
        print(
            "Raster z median:",
            float(np.median(vals)),
        )

    print()
    print("TIDE WORKBOOK")
    print("-" * 92)

    print(
        "Title:",
        props.title,
    )
    print(
        "Subject:",
        props.subject,
    )
    print(
        "Creator:",
        props.creator,
    )
    print(
        "Description:",
        props.description,
    )

    print()
    print(
        "Tide time range:",
        tide_times[0],
        "through",
        tide_times[-1],
    )

    # Observation interval.
    obs_times = [
        r["dt"]
        for r in rows
    ]

    print(
        "GNSS-R target time range:",
        min(obs_times),
        "through",
        max(obs_times),
    )

    print()
    print("TIDE MODEL RANGES DURING TARGET OBSERVATIONS")
    print("-" * 92)

    for m in MODELS:
        sample = np.array(
            [
                tide_at(
                    r["dt"],
                    m,
                )
                for r in rows
            ],
            dtype=float,
        )

        print(
            f"{m:18s}"
            f" min={np.min(sample):+.4f} m"
            f" max={np.max(sample):+.4f} m"
            f" mean={np.mean(sample):+.4f} m"
            f" std={np.std(sample):.4f} m"
        )

    # -----------------------------------------------------------------
    # Build all footprints and raster values once.
    # -----------------------------------------------------------------

    prepared = []

    with rasterio.open(
        TOPO_FILE
    ) as src:

        for row in rows:

            polygon = build_fresnel_polygon(
                row
            )

            raster_values = (
                raster_values_for_polygon(
                    src,
                    polygon,
                )
            )

            tides = {
                m: tide_at(
                    row["dt"],
                    m,
                )
                for m in MODELS
            }

            prepared.append(
                {
                    "row": row,
                    "polygon": polygon,
                    "z": raster_values,
                    "tides": tides,
                }
            )

    shifts = build_shift_grid()

    output = []

    # -----------------------------------------------------------------
    # Sweep every shift for every model and target track.
    # -----------------------------------------------------------------

    for item in prepared:

        row = item["row"]
        z = item["z"]

        for model in MODELS:

            base_tide = item[
                "tides"
            ][model]

            for shift in shifts:

                water_elevation = (
                    base_tide
                    + shift
                )

                wet, near, dry = fractions(
                    z,
                    water_elevation,
                )

                output.append(
                    {
                        "sat":
                            row["sat"],
                        "freq":
                            row["freq"],
                        "track":
                            TARGETS[
                                (
                                    row["sat"],
                                    row["freq"],
                                )
                            ],
                        "doy":
                            row["doy"],
                        "datetime_utc":
                            row["dt"].isoformat(),
                        "RH_m":
                            row["rh"],
                        "az_deg":
                            row["az"],
                        "elevation_deg":
                            ELEVATION_DEG,
                        "tide_model":
                            model,
                        "raw_tide_m":
                            base_tide,
                        "datum_shift_m":
                            shift,
                        "water_elevation_test_m":
                            water_elevation,
                        "wet_fraction":
                            wet,
                        "near_water_fraction":
                            near,
                        "dry_fraction":
                            dry,
                        "raster_cells":
                            int(len(z)),
                    }
                )

    # -----------------------------------------------------------------
    # Write detailed CSV.
    # -----------------------------------------------------------------

    fields = list(
        output[0].keys()
    )

    with open(
        OUT_CSV,
        "w",
        newline="",
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fields,
        )
        writer.writeheader()
        writer.writerows(
            output
        )

    # -----------------------------------------------------------------
    # Track-level sweep summary.
    # -----------------------------------------------------------------

    summary = []

    summary.append(
        "MARCONI TIDE DATUM INVESTIGATION / ±0.50 m SENSITIVITY"
    )
    summary.append(
        "=" * 92
    )
    summary.append(
        f"Topobathy: {TOPO_FILE}"
    )
    summary.append(
        f"Tide workbook: {TIDE_FILE}"
    )
    summary.append(
        f"Footprint elevation: {ELEVATION_DEG:.1f} deg"
    )
    summary.append(
        f"Shift range: {SHIFT_MIN:+.2f} to {SHIFT_MAX:+.2f} m"
    )
    summary.append(
        ""
    )

    summary.append(
        "IMPORTANT DATUM FINDING"
    )
    summary.append(
        "The tide-model columns in the workbook are tidal signals"
    )
    summary.append(
        "centered near zero. They must not be interpreted as NAVD88"
    )
    summary.append(
        "absolute water elevations without an independently established"
    )
    summary.append(
        "local vertical datum offset."
    )
    summary.append(
        ""
    )

    summary.append(
        "TRACK / MODEL SHIFT SENSITIVITY"
    )
    summary.append(
        "-" * 92
    )

    for key, label in TARGETS.items():

        for model in MODELS:

            rows_model = [
                x for x in output
                if (
                    x["sat"] == key[0]
                    and x["freq"] == key[1]
                    and x["tide_model"] == model
                )
            ]

            # Track-level mean wet fraction for each shift.
            shift_means = []

            for shift in shifts:

                subset = [
                    x["wet_fraction"]
                    for x in rows_model
                    if x["datum_shift_m"] == shift
                ]

                subset = [
                    x for x in subset
                    if math.isfinite(x)
                ]

                if not subset:
                    mean = math.nan
                else:
                    mean = float(
                        np.mean(subset)
                    )

                shift_means.append(
                    (
                        shift,
                        mean,
                    )
                )

            eligible = [
                x for x in shift_means
                if math.isfinite(x[1])
            ]

            if eligible:

                # Shift closest to 80% mean wet.
                best = min(
                    eligible,
                    key=lambda x:
                        abs(
                            x[1]
                            - WATER_FRACTION_TARGET
                        ),
                )

                # Shifts where mean wet >= 80%.
                above = [
                    x for x in eligible
                    if x[1] >= WATER_FRACTION_TARGET
                ]

                if above:
                    min_shift = min(
                        x[0]
                        for x in above
                    )
                    max_shift = max(
                        x[0]
                        for x in above
                    )
                    band = (
                        f"{min_shift:+.2f} to "
                        f"{max_shift:+.2f} m"
                    )
                else:
                    band = "none in ±0.50 m"

                summary.append(
                    f"{label:16s} "
                    f"{model:18s} "
                    f"best80_shift={best[0]:+.2f} m "
                    f"mean_wet={best[1]:.3f} "
                    f">=80% band={band}"
                )

    summary.append(
        ""
    )
    summary.append(
        "INTERPRETATION"
    )
    summary.append(
        "The shift sweep is a sensitivity analysis, not a calibration."
    )
    summary.append(
        "Do not choose a datum shift solely because it maximizes water"
    )
    summary.append(
        "fraction. The final offset must come from the independently"
    )
    summary.append(
        "established relationship between the tide-model datum and the"
    )
    summary.append(
        "USGS NAVD88 topobathy datum / local water-level reference."
    )

    OUT_SUMMARY.write_text(
        "\n".join(summary)
        + "\n"
    )

    print()
    print("=" * 92)
    print(
        "KEY RESULT: VERTICAL DATUM SENSITIVITY"
    )
    print("=" * 92)

    # Print only EOT20 and ensemble-style interpretation here.
    # Full details are in the CSV/summary.
    for key, label in TARGETS.items():

        rows_model = [
            x for x in output
            if (
                x["sat"] == key[0]
                and x["freq"] == key[1]
                and x["tide_model"] == "EOT20_heightm"
            )
        ]

        print()
        print(label)

        for shift in shifts:

            wet = [
                x["wet_fraction"]
                for x in rows_model
                if x["datum_shift_m"] == shift
                and math.isfinite(
                    x["wet_fraction"]
                )
            ]

            if wet:
                print(
                    f"  shift={shift:+.2f} m"
                    f" mean_wet={np.mean(wet):.3f}"
                )

    print()
    print(
        "Outputs:"
    )
    print(
        f"  {OUT_CSV}"
    )
    print(
        f"  {OUT_SUMMARY}"
    )

    print()
    print("DONE")


if __name__ == "__main__":
    main()
