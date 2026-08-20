#!/usr/bin/env python3
"""
analyze_arc_error_vs_tide_change.py

USGS GNSS Reference Station Project
------------------------------------

Scientific diagnostic to test whether GNSS-IR reflector-height error
increases when the tide changes substantially DURING the same satellite
arc.

This is a diagnostic only. It does NOT modify production GNSS-IR
processing and does NOT apply a correction.

Primary scientific question:

    Do arcs with larger real tidal movement during their own duration
    have larger GNSS-IR errors?

The tide-model workbook and GNSS-IR reflector heights are not assumed
to share the same vertical datum.

Therefore we first fit a single global affine relationship:

    -RH = scale * tide + offset

The offset absorbs the unknown vertical datum difference.

The resulting residual is a SHAPE residual. It is not an absolute
water-level error.

Outputs:

    arc_error_vs_tide_change.csv
    arc_error_vs_tide_change.png
    arc_error_vs_tide_change_summary.txt

Usage:

    python3 analyze_arc_error_vs_tide_change.py \
        usgs 2026 204 207 marconi_tides_sherwood.xlsx

"""

from __future__ import annotations

import csv
import math
import sys
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import openpyxl


# ----------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------

MJD_EPOCH = datetime(1858, 11, 17)

# Same broad frequency set used by our existing arc diagnostics.
REAL_FREQUENCIES = [
    1,
    5,
    20,
    101,
    102,
    201,
    205,
    206,
    207,
    208,
    301,
    302,
    305,
    306,
    307,
    308,
]


# ----------------------------------------------------------------------
# Utility functions
# ----------------------------------------------------------------------

def finite(value):
    """Return a finite float or None."""
    try:
        value = float(value)
    except (TypeError, ValueError):
        return None

    if math.isfinite(value):
        return value

    return None


def pearson(x, y):
    """Calculate Pearson correlation coefficient."""
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    mask = np.isfinite(x) & np.isfinite(y)

    if np.count_nonzero(mask) < 3:
        return float("nan")

    x = x[mask]
    y = y[mask]

    if np.std(x) == 0 or np.std(y) == 0:
        return float("nan")

    return float(np.corrcoef(x, y)[0, 1])


def rankdata(x):
    """
    Simple average-rank implementation used for Spearman correlation.
    """
    x = np.asarray(x, dtype=float)

    order = np.argsort(x, kind="mergesort")

    ranks = np.empty(len(x), dtype=float)

    i = 0

    while i < len(x):

        j = i + 1

        while j < len(x) and x[order[j]] == x[order[i]]:
            j += 1

        ranks[order[i:j]] = (i + 1 + j) / 2.0

        i = j

    return ranks


def spearman(x, y):
    """Calculate Spearman rank correlation."""
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    mask = np.isfinite(x) & np.isfinite(y)

    if np.count_nonzero(mask) < 3:
        return float("nan")

    return pearson(
        rankdata(x[mask]),
        rankdata(y[mask]),
    )


def affine_fit(x, y):
    """
    Fit:

        y = scale*x + offset

    using least squares.
    """

    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    mask = np.isfinite(x) & np.isfinite(y)

    if np.count_nonzero(mask) < 3:
        raise ValueError(
            "Not enough valid points for affine fit."
        )

    x = x[mask]
    y = y[mask]

    A = np.column_stack(
        (
            x,
            np.ones_like(x),
        )
    )

    scale, offset = np.linalg.lstsq(
        A,
        y,
        rcond=None,
    )[0]

    fitted = scale * x + offset

    return (
        float(scale),
        float(offset),
        fitted,
    )


def seconds_to_datetime(day_start, seconds):
    """Convert seconds-of-day into datetime."""
    return day_start + timedelta(
        seconds=float(seconds)
    )


# ----------------------------------------------------------------------
# Tide model handling
# ----------------------------------------------------------------------

def read_tide_models(path):
    """
    Read the tide-model Excel workbook.

    Expected columns include:

        time
        EOT20_heightm
        GOT5.5_heightm
        GOT5.6_heightm
        FES2022_heightm

    Any column ending in '_heightm' is treated as a tide model.
    """

    print(
        f"Reading tide models from {path}..."
    )

    wb = openpyxl.load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    rows = list(
        ws.iter_rows(values_only=True)
    )

    if not rows:
        raise ValueError(
            "Tide-model workbook is empty."
        )

    header = rows[0]

    model_columns = {}

    for index, name in enumerate(header):

        if name is None:
            continue

        name = str(name)

        if name.endswith("_heightm"):

            model_name = name[
                :-len("_heightm")
            ]

            model_columns[
                model_name
            ] = index

    if not model_columns:
        raise ValueError(
            "No *_heightm tide-model columns found."
        )

    times = []

    model_values = {
        name: []
        for name in model_columns
    }

    for row in rows[1:]:

        if row[0] is None:
            continue

        if not isinstance(
            row[0],
            datetime,
        ):
            continue

        row_values = {}

        valid = True

        for model_name, column_index in model_columns.items():

            value = finite(
                row[column_index]
            )

            if value is None:
                valid = False
                break

            row_values[
                model_name
            ] = value

        if not valid:
            continue

        times.append(row[0])

        for model_name in model_columns:

            model_values[
                model_name
            ].append(
                row_values[model_name]
            )

    if len(times) < 2:
        raise ValueError(
            "Not enough valid tide-model points."
        )

    print(
        f"Parsed {len(times)} tide-model points."
    )

    print(
        "Models:",
        list(model_values.keys()),
    )

    print(
        f"Model time range: "
        f"{times[0]} through {times[-1]}"
    )

    return (
        times,
        model_values,
    )


def build_interpolators(
    times,
    model_values,
):
    """
    Build linear interpolators for each tide model and an ensemble mean.
    """

    x = np.array(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    interpolators = {}

    for model_name, values in model_values.items():

        y = np.asarray(
            values,
            dtype=float,
        )

        def make_interpolator(
            y_values
        ):

            def interpolate(
                query_time
            ):

                query_seconds = (
                    query_time - times[0]
                ).total_seconds()

                return float(
                    np.interp(
                        query_seconds,
                        x,
                        y_values,
                    )
                )

            return interpolate

        interpolators[
            model_name
        ] = make_interpolator(y)

    # Ensemble mean of all available models.
    ensemble = np.mean(
        np.vstack(
            [
                np.asarray(
                    model_values[name],
                    dtype=float,
                )
                for name in model_values
            ]
        ),
        axis=0,
    )

    interpolators[
        "ENSEMBLE"
    ] = make_interpolator(
        ensemble
    )

    return interpolators


# ----------------------------------------------------------------------
# Main analysis
# ----------------------------------------------------------------------

def main():

    if len(sys.argv) not in (5, 6):

        print(
            "Usage:"
        )

        print(
            f"  {sys.argv[0]} "
            "<station> <year> <doy1> "
            "[doy2] <tide_models.xlsx>"
        )

        sys.exit(1)

    station = sys.argv[1]

    year = int(
        sys.argv[2]
    )

    doy1 = int(
        sys.argv[3]
    )

    if len(sys.argv) == 6:

        doy2 = int(
            sys.argv[4]
        )

        tide_path = Path(
            sys.argv[5]
        )

    else:

        doy2 = doy1

        tide_path = Path(
            sys.argv[4]
        )

    if not tide_path.exists():

        print(
            f"ERROR: tide-model file not found:"
        )

        print(
            f"  {tide_path}"
        )

        sys.exit(1)

    # --------------------------------------------------------------
    # Import gnssrefl
    # --------------------------------------------------------------

    try:

        from gnssrefl.extract_arcs import (
            extract_arcs_from_station
        )

    except ImportError:

        print(
            "ERROR: could not import "
            "gnssrefl.extract_arcs."
        )

        print(
            "Make sure gnssrefl_venv is activated."
        )

        sys.exit(1)

    # --------------------------------------------------------------
    # Header
    # --------------------------------------------------------------

    print()
    print("=" * 70)
    print(
        "GNSS-IR ARC ERROR VS WITHIN-ARC TIDAL CHANGE"
    )
    print("=" * 70)

    print(
        f"Station       : {station}"
    )

    print(
        f"Year          : {year}"
    )

    print(
        f"DOY range     : {doy1}-{doy2}"
    )

    print(
        f"Tide workbook : {tide_path}"
    )

    print()

    print(
        "TIME CONVENTION"
    )

    print(
        "  GNSS-IR timestamps are treated as UTC."
    )

    print(
        "  Tide-model Excel timestamps are timezone-naive."
    )

    print(
        "  This analysis assumes they use the same clock basis."
    )

    print(
        "  That assumption is NOT independently proven here."
    )

    print()

    # --------------------------------------------------------------
    # Read tide models
    # --------------------------------------------------------------

    (
        tide_times,
        tide_models,
    ) = read_tide_models(
        tide_path
    )

    interpolators = build_interpolators(
        tide_times,
        tide_models,
    )

    # --------------------------------------------------------------
    # Extract arcs
    # --------------------------------------------------------------

    day_start_reference = datetime(
        year,
        1,
        1,
    )

    records = []

    total_extracted = 0

    total_qc = 0

    outside_model = 0

    for doy in range(
        doy1,
        doy2 + 1,
    ):

        print()

        print(
            f"Extracting all-frequency arcs "
            f"for {station} {year} doy {doy}..."
        )

        arcs = extract_arcs_from_station(
            station,
            year,
            doy,
            freq=REAL_FREQUENCIES,
            e1=5.0,
            e2=15.0,
            azlist=[
                100.0,
                130.0,
                150.0,
                215.0,
            ],
            polyV=2,
            pele=[
                5,
                30,
            ],
            attach_results=True,
            buffer_hours=2,
        )

        total_extracted += len(
            arcs
        )

        qc_arcs = []

        for meta, data in arcs:

            result = meta.get(
                "gnssir_processing_results"
            )

            if not isinstance(
                result,
                dict,
            ):
                continue

            rh = finite(
                result.get("RH")
            )

            if rh is None:
                continue

            qc_arcs.append(
                (
                    meta,
                    data,
                )
            )

        total_qc += len(
            qc_arcs
        )

        print(
            f"{len(arcs)} arcs extracted, "
            f"{len(qc_arcs)} passed QC"
        )

        day_start = (
            day_start_reference
            + timedelta(
                days=doy - 1
            )
        )

        # ----------------------------------------------------------
        # Process individual arcs
        # ----------------------------------------------------------

        for meta, data in qc_arcs:

            result = meta[
                "gnssir_processing_results"
            ]

            rh = finite(
                result.get("RH")
            )

            mjd = finite(
                result.get("MJD")
            )

            if rh is None:
                continue

            if mjd is None:
                continue

            # Arc start/end.
            try:

                arc_start = seconds_to_datetime(
                    day_start,
                    meta["time_start"],
                )

                arc_end = seconds_to_datetime(
                    day_start,
                    meta["time_end"],
                )

            except (
                KeyError,
                TypeError,
                ValueError,
            ):

                continue

            # GNSS-IR solution epoch.
            solution_time = (
                MJD_EPOCH
                + timedelta(
                    days=mjd
                )
            )

            # Make sure all required points are inside
            # the tide-model time span.
            if (
                arc_start < tide_times[0]
                or arc_end > tide_times[-1]
                or solution_time < tide_times[0]
                or solution_time > tide_times[-1]
            ):

                outside_model += 1

                continue

            tide_start = interpolators[
                "ENSEMBLE"
            ](
                arc_start
            )

            tide_solution = interpolators[
                "ENSEMBLE"
            ](
                solution_time
            )

            tide_end = interpolators[
                "ENSEMBLE"
            ](
                arc_end
            )

            tide_change_m = (
                tide_end
                - tide_start
            )

            tide_change_cm = (
                tide_change_m
                * 100.0
            )

            duration = finite(
                meta.get("delT")
            )

            if (
                duration is not None
                and duration > 0
            ):

                tide_rate = (
                    tide_change_cm
                    / duration
                )

            else:

                tide_rate = None

            # GNSS-IR sea-level-shaped quantity.
            #
            # RH increases upward from the reflecting surface.
            # For comparison with water level, use -RH.
            gnss_shape = -rh

            records.append(
                {
                    "doy": doy,

                    "sat": meta.get(
                        "sat"
                    ),

                    "freq": meta.get(
                        "freq"
                    ),

                    "delT_min": duration,

                    "arc_start_utc":
                        arc_start.isoformat(),

                    "solution_time_utc":
                        solution_time.isoformat(),

                    "arc_end_utc":
                        arc_end.isoformat(),

                    "rh_m": rh,

                    "gnss_shape_m":
                        gnss_shape,

                    "tide_start_m":
                        tide_start,

                    "tide_solution_m":
                        tide_solution,

                    "tide_end_m":
                        tide_end,

                    "tide_change_m":
                        tide_change_m,

                    "tide_change_cm":
                        tide_change_cm,

                    "abs_tide_change_cm":
                        abs(
                            tide_change_cm
                        ),

                    "tide_rate_cm_per_min":
                        tide_rate,

                    "cf_m":
                        finite(
                            meta.get("cf")
                        ),
                }
            )

    # --------------------------------------------------------------
    # Check data
    # --------------------------------------------------------------

    if len(records) < 10:

        print()

        print(
            "ERROR: only "
            f"{len(records)} usable arcs remained."
        )

        print(
            "Cannot perform a meaningful statistical test."
        )

        sys.exit(1)

    print()

    print("=" * 70)

    print(
        f"USABLE ARCS: {len(records)}"
    )

    print(
        f"Total extracted: {total_extracted}"
    )

    print(
        f"QC passing:     {total_qc}"
    )

    print(
        f"Outside model:  {outside_model}"
    )

    print("=" * 70)

    # --------------------------------------------------------------
    # Global affine fit
    # --------------------------------------------------------------

    tide = np.array(
        [
            r["tide_solution_m"]
            for r in records
        ],
        dtype=float,
    )

    gnss = np.array(
        [
            r["gnss_shape_m"]
            for r in records
        ],
        dtype=float,
    )

    scale, offset, fitted = affine_fit(
        tide,
        gnss,
    )

    residual = (
        gnss
        - fitted
    )

    for (
        record,
        fitted_value,
        residual_value,
    ) in zip(
        records,
        fitted,
        residual,
    ):

        record[
            "global_fit_gnss_m"
        ] = float(
            fitted_value
        )

        record[
            "shape_residual_m"
        ] = float(
            residual_value
        )

        record[
            "abs_shape_residual_cm"
        ] = abs(
            float(
                residual_value
            )
        ) * 100.0

    # --------------------------------------------------------------
    # Primary variables
    # --------------------------------------------------------------

    abs_tide_change = np.array(
        [
            r["abs_tide_change_cm"]
            for r in records
        ],
        dtype=float,
    )

    abs_error = np.array(
        [
            r[
                "abs_shape_residual_cm"
            ]
            for r in records
        ],
        dtype=float,
    )

    duration = np.array(
        [
            r["delT_min"]
            if r["delT_min"] is not None
            else np.nan
            for r in records
        ],
        dtype=float,
    )

    signed_tide_change = np.array(
        [
            r["tide_change_cm"]
            for r in records
        ],
        dtype=float,
    )

    tide_rate = np.array(
        [
            r[
                "tide_rate_cm_per_min"
            ]
            if r[
                "tide_rate_cm_per_min"
            ] is not None
            else np.nan
            for r in records
        ],
        dtype=float,
    )

    # --------------------------------------------------------------
    # Statistics
    # --------------------------------------------------------------

    primary_pearson = pearson(
        abs_error,
        abs_tide_change,
    )

    primary_spearman = spearman(
        abs_error,
        abs_tide_change,
    )

    duration_pearson = pearson(
        abs_error,
        duration,
    )

    duration_spearman = spearman(
        abs_error,
        duration,
    )

    rate_pearson = pearson(
        abs_error,
        np.abs(tide_rate),
    )

    shape_correlation = pearson(
        gnss,
        tide,
    )

    shape_rms_cm = (
        np.sqrt(
            np.mean(
                residual ** 2
            )
        )
        * 100.0
    )

    shape_mae_cm = (
        np.mean(
            np.abs(
                residual
            )
        )
        * 100.0
    )

    # --------------------------------------------------------------
    # Print results
    # --------------------------------------------------------------

    print()

    print(
        "--- GLOBAL SHAPE FIT ---"
    )

    print(
        "Model:"
    )

    print(
        "    -RH = scale * tide + offset"
    )

    print(
        f"Scale  : {scale:.6f}"
    )

    print(
        f"Offset : {offset:.6f} m"
    )

    print(
        f"Correlation (-RH, tide): "
        f"{shape_correlation:.4f}"
    )

    print(
        f"Shape residual RMS: "
        f"{shape_rms_cm:.2f} cm"
    )

    print(
        f"Shape residual MAE: "
        f"{shape_mae_cm:.2f} cm"
    )

    print()

    print(
        "--- PRIMARY TEST ---"
    )

    print(
        "Does GNSS-IR error increase when "
        "the tide moves more during the same arc?"
    )

    print()

    print(
        f"Pearson r(abs error, abs tide change): "
        f"{primary_pearson:.4f}"
    )

    print(
        f"Spearman rho(abs error, abs tide change): "
        f"{primary_spearman:.4f}"
    )

    print(
        f"Pearson r(abs error, arc duration): "
        f"{duration_pearson:.4f}"
    )

    print(
        f"Spearman rho(abs error, arc duration): "
        f"{duration_spearman:.4f}"
    )

    print(
        f"Pearson r(abs error, abs tide rate): "
        f"{rate_pearson:.4f}"
    )

    # --------------------------------------------------------------
    # Error versus tidal-movement bins
    # --------------------------------------------------------------

    print()

    print(
        "--- ERROR BY TIDAL-MOVEMENT BIN ---"
    )

    bins = [
        (
            "0-2 cm",
            0.0,
            2.0,
        ),

        (
            "2-5 cm",
            2.0,
            5.0,
        ),

        (
            "5-10 cm",
            5.0,
            10.0,
        ),

        (
            "10-20 cm",
            10.0,
            20.0,
        ),

        (
            ">20 cm",
            20.0,
            float("inf"),
        ),
    ]

    bin_rows = []

    for (
        label,
        low,
        high,
    ) in bins:

        mask = (
            (abs_tide_change >= low)
            &
            (abs_tide_change < high)
        )

        n = int(
            np.count_nonzero(
                mask
            )
        )

        if n > 0:

            mean_error = float(
                np.mean(
                    abs_error[mask]
                )
            )

            median_error = float(
                np.median(
                    abs_error[mask]
                )
            )

            rms_error = float(
                np.sqrt(
                    np.mean(
                        residual[mask] ** 2
                    )
                )
                * 100.0
            )

        else:

            mean_error = float("nan")

            median_error = float("nan")

            rms_error = float("nan")

        print(
            f"{label:>10}: "
            f"n={n:4d}  "
            f"mean abs error="
            f"{mean_error:7.2f} cm  "
            f"median="
            f"{median_error:7.2f} cm  "
            f"RMS="
            f"{rms_error:7.2f} cm"
        )

        bin_rows.append(
            (
                label,
                n,
                mean_error,
                median_error,
                rms_error,
            )
        )

    # --------------------------------------------------------------
    # Rising versus falling tide
    # --------------------------------------------------------------

    print()

    print(
        "--- ERROR BY TIDE DIRECTION ---"
    )

    for (
        label,
        mask,
    ) in (
        (
            "RISING",
            signed_tide_change > 0,
        ),

        (
            "FALLING",
            signed_tide_change < 0,
        ),

        (
            "ZERO",
            signed_tide_change == 0,
        ),
    ):

        n = int(
            np.count_nonzero(
                mask
            )
        )

        if n == 0:
            continue

        print(
            f"{label:>8}: "
            f"n={n:4d}  "
            f"mean signed residual="
            f"{np.mean(residual[mask]) * 100:8.3f} cm  "
            f"mean abs residual="
            f"{np.mean(abs_error[mask]):8.3f} cm"
        )

    # --------------------------------------------------------------
    # Frequency summary
    # --------------------------------------------------------------

    print()

    print(
        "--- ERROR BY FREQUENCY ---"
    )

    frequencies = sorted(
        {
            r["freq"]
            for r in records
            if r["freq"] is not None
        }
    )

    frequency_rows = []

    for freq in frequencies:

        mask = np.array(
            [
                r["freq"] == freq
                for r in records
            ],
            dtype=bool,
        )

        n = int(
            np.count_nonzero(
                mask
            )
        )

        mean_error = float(
            np.mean(
                abs_error[mask]
            )
        )

        rms_error = float(
            np.sqrt(
                np.mean(
                    residual[mask] ** 2
                )
            )
            * 100.0
        )

        print(
            f"freq={str(freq):>4}: "
            f"n={n:4d}  "
            f"mean abs error="
            f"{mean_error:7.2f} cm  "
            f"RMS="
            f"{rms_error:7.2f} cm"
        )

        frequency_rows.append(
            (
                freq,
                n,
                mean_error,
                rms_error,
            )
        )

    # --------------------------------------------------------------
    # Save detailed CSV
    # --------------------------------------------------------------

    csv_path = Path(
        "arc_error_vs_tide_change.csv"
    )

    fieldnames = list(
        records[0].keys()
    )

    with csv_path.open(
        "w",
        newline="",
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames,
        )

        writer.writeheader()

        writer.writerows(
            records
        )

    # --------------------------------------------------------------
    # Save text summary
    # --------------------------------------------------------------

    summary_path = Path(
        "arc_error_vs_tide_change_summary.txt"
    )

    with summary_path.open(
        "w"
    ) as f:

        f.write(
            "USGS GNSS-IR ARC ERROR VS WITHIN-ARC TIDAL CHANGE\n"
        )

        f.write(
            "=" * 60
            + "\n"
        )

        f.write(
            f"Station: {station}\n"
        )

        f.write(
            f"Year: {year}\n"
        )

        f.write(
            f"DOY: {doy1}-{doy2}\n"
        )

        f.write(
            f"Usable arcs: {len(records)}\n"
        )

        f.write(
            f"Total extracted: {total_extracted}\n"
        )

        f.write(
            f"QC passing: {total_qc}\n"
        )

        f.write(
            f"Outside model range: "
            f"{outside_model}\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "TIME CONVENTION WARNING\n"
        )

        f.write(
            "GNSS-IR timestamps are treated as UTC. "
            "Tide-model Excel timestamps are timezone-naive. "
            "This script assumes the same clock basis; "
            "that assumption is not independently proven.\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "GLOBAL SHAPE FIT\n"
        )

        f.write(
            f"scale={scale:.9f}\n"
        )

        f.write(
            f"offset_m={offset:.9f}\n"
        )

        f.write(
            f"correlation={shape_correlation:.9f}\n"
        )

        f.write(
            f"shape_rms_cm="
            f"{shape_rms_cm:.6f}\n"
        )

        f.write(
            f"shape_mae_cm="
            f"{shape_mae_cm:.6f}\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "PRIMARY CORRELATIONS\n"
        )

        f.write(
            "pearson_abs_error_vs_abs_tide_change="
            f"{primary_pearson:.9f}\n"
        )

        f.write(
            "spearman_abs_error_vs_abs_tide_change="
            f"{primary_spearman:.9f}\n"
        )

        f.write(
            "pearson_abs_error_vs_duration="
            f"{duration_pearson:.9f}\n"
        )

        f.write(
            "spearman_abs_error_vs_duration="
            f"{duration_spearman:.9f}\n"
        )

        f.write(
            "pearson_abs_error_vs_abs_tide_rate="
            f"{rate_pearson:.9f}\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "TIDAL-MOVEMENT BINS\n"
        )

        for row in bin_rows:

            f.write(
                f"{row[0]}: "
                f"n={row[1]}, "
                f"mean_abs_error_cm="
                f"{row[2]:.6f}, "
                f"median_abs_error_cm="
                f"{row[3]:.6f}, "
                f"rms_error_cm="
                f"{row[4]:.6f}\n"
            )

        f.write(
            "\n"
        )

        f.write(
            "FREQUENCY SUMMARY\n"
        )

        for row in frequency_rows:

            f.write(
                f"freq={row[0]}: "
                f"n={row[1]}, "
                f"mean_abs_error_cm="
                f"{row[2]:.6f}, "
                f"rms_error_cm="
                f"{row[3]:.6f}\n"
            )

    # --------------------------------------------------------------
    # Diagnostic plot
    # --------------------------------------------------------------

    fig = plt.figure(
        figsize=(12, 10)
    )

    # --------------------------------------------------------------
    # Plot 1
    # --------------------------------------------------------------

    ax = fig.add_subplot(
        2,
        2,
        1,
    )

    ax.scatter(
        abs_tide_change,
        abs_error,
        s=18,
        alpha=0.55,
    )

    ax.set_xlabel(
        "Absolute within-arc tidal change (cm)"
    )

    ax.set_ylabel(
        "Absolute GNSS-IR shape residual (cm)"
    )

    ax.set_title(
        "GNSS-IR error vs tidal movement\n"
        f"Pearson r={primary_pearson:.3f}, "
        f"Spearman={primary_spearman:.3f}"
    )

    ax.grid(
        alpha=0.3
    )

    # --------------------------------------------------------------
    # Plot 2
    # --------------------------------------------------------------

    ax = fig.add_subplot(
        2,
        2,
        2,
    )

    ax.scatter(
        duration,
        abs_error,
        s=18,
        alpha=0.55,
    )

    ax.set_xlabel(
        "Arc duration (minutes)"
    )

    ax.set_ylabel(
        "Absolute GNSS-IR shape residual (cm)"
    )

    ax.set_title(
        "GNSS-IR error vs arc duration\n"
        f"Pearson r={duration_pearson:.3f}"
    )

    ax.grid(
        alpha=0.3
    )

    # --------------------------------------------------------------
    # Plot 3
    # --------------------------------------------------------------

    ax = fig.add_subplot(
        2,
        2,
        3,
    )

    ax.scatter(
        tide_rate,
        residual * 100.0,
        s=18,
        alpha=0.55,
    )

    ax.axhline(
        0.0,
        linestyle="--",
    )

    ax.set_xlabel(
        "Signed tide rate during arc (cm/min)"
    )

    ax.set_ylabel(
        "Signed GNSS-IR shape residual (cm)"
    )

    ax.set_title(
        "Residual vs signed tidal rate"
    )

    ax.grid(
        alpha=0.3
    )

    # --------------------------------------------------------------
    # Plot 4
    # --------------------------------------------------------------

    ax = fig.add_subplot(
        2,
        2,
        4,
    )

    ax.scatter(
        tide,
        gnss,
        s=18,
        alpha=0.55,
    )

    line = np.linspace(
        float(
            np.min(tide)
        ),
        float(
            np.max(tide)
        ),
        100,
    )

    ax.plot(
        line,
        scale * line + offset,
        linewidth=2,
    )

    ax.set_xlabel(
        "Tide-model ensemble height (m)"
    )

    ax.set_ylabel(
        "-GNSS-IR RH (m)"
    )

    ax.set_title(
        f"Global shape fit: "
        f"scale={scale:.3f}, "
        f"r={shape_correlation:.3f}"
    )

    ax.grid(
        alpha=0.3
    )

    # --------------------------------------------------------------
    # Figure
    # --------------------------------------------------------------

    fig.suptitle(
        f"{station} {year} DOY {doy1}-{doy2}: "
        "Arc Error vs Within-Arc Tidal Change"
    )

    fig.tight_layout()

    plot_path = Path(
        "arc_error_vs_tide_change.png"
    )

    fig.savefig(
        plot_path,
        dpi=160,
    )

    plt.close(fig)

    # --------------------------------------------------------------
    # Finished
    # --------------------------------------------------------------

    print()

    print(
        "=" * 70
    )

    print(
        "OUTPUT FILES"
    )

    print(
        "=" * 70
    )

    print(
        f"CSV     : {csv_path}"
    )

    print(
        f"Plot    : {plot_path}"
    )

    print(
        f"Summary : {summary_path}"
    )

    print()

    print(
        "PRIMARY SCIENTIFIC TEST:"
    )

    print(
        "  Does absolute GNSS-IR shape error increase "
        "with absolute tidal movement during the same arc?"
    )

    print()

    print(
        "No production GNSS-IR processing was modified."
    )


if __name__ == "__main__":

    main()
