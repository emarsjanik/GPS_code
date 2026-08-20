#!/usr/bin/env python3
"""
Marconi GNSS-IR / TOPOBATHY Fresnel Validator

Scientific purpose
------------------
The previous geometry test used an 83.06 deg / 71.78 m half-plane.
This script replaces that simplified shoreline test with the actual
USGS Marconi Beach 1-m topobathy raster.

For each selected GNSS-IR arc it:
  1. Reads the actual RH and azimuth from the 17-23 m GPS L1 experiment.
  2. Builds the first Fresnel ellipse at the requested elevation.
  3. Samples the USGS topobathy grid under the ellipse.
  4. Interpolates the four tide-model water elevations at the arc epoch.
  5. Classifies raster cells as:
        WET          z <= tide + wet_margin
        NEAR_WATER   tide + wet_margin < z <= tide + dry_margin
        DRY          z > tide + dry_margin
  6. Calculates wet / near-water / dry fractions.
  7. Produces a CSV and KML.

IMPORTANT SCIENCE LIMITATIONS
-----------------------------
- The USGS topobathy survey is a 2021 site survey; it is NOT a 2026
  shoreline. It is used as an independently surveyed geomorphic baseline.
- Absolute water/land classification is only physically valid if the tide
  model and topobathy vertical datums are compatible. This script reports
  the input values and does not silently datum-shift them.
- The current Marconi tide workbook contains model heights but no embedded
  datum metadata. Therefore the first run should be treated as a geometry
  sensitivity test. Once the datum is verified, it becomes the physical
  water-footprint test.
- The raster is the actual terrain/bathymetry surface; unlike the previous
  half-plane, the classification varies with alongshore geometry and
  local beach/nearshore elevation.

USGS dataset used
-----------------
2021022FA_Marconi_topobathy_1m.tif

USGS describes this as a merged 1-m gridded bathymetry + digital elevation
model for Marconi Beach, Wellfleet, UTM Zone 19N. The same release documents
RTK beach elevation and nearshore single-beam bathymetry used to build the
continuous topobathy product.

The script will use an existing local TIFF first. If not present, it will
attempt to download the URL below with Python urllib. If the URL is not
available from the NUC, manually download the TIFF from the USGS data
release page and place it at the filename above.
"""

from __future__ import annotations

import csv
import math
import sys
import urllib.request
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import rasterio
from rasterio.mask import mask
from rasterio.warp import transform
from shapely.geometry import Polygon, mapping, Point
from shapely.ops import transform as shapely_transform
from pyproj import Transformer
from openpyxl import load_workbook
import simplekml

from gnssrefl.refl_zones import makeEllipse_latlon


# ---------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------

RESULT_DIR = Path(
    "products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13"
)

TOPO_FILE = Path(
    "2021022FA_Marconi_topobathy_1m.tif"
)

TOPO_URL = (
    "https://cmgds.marine.usgs.gov/"
    "data/field-activity-data/2021-022-FA/data/Topobathy/"
    "2021022FA_Marconi_topobathy_1m.tif"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

OUT_CSV = Path(
    "marconi_topobathy_fresnel_validation.csv"
)

OUT_KML = Path(
    "marconi_topobathy_fresnel_validation.kml"
)

OUT_SUMMARY = Path(
    "marconi_topobathy_fresnel_validation_summary.txt"
)

LAT = 41.8928243333
LON = -69.9633227139

MODELS = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]

# Leading physical candidate tracks.
TARGETS = {
    (26, 1): "PRN26_AZ91",
    (21, 1): "PRN21_AZ93",
    (16, 1): "PRN16_AZ96",
}

DOYS = [204, 205, 206, 207]

# Geometry sensitivity.
ELEVATIONS = [5.0, 9.0, 13.0]

# Classification tolerance.
# 0.05 m is intentionally conservative around the model waterline.
WET_MARGIN_M = 0.00
NEAR_WATER_MARGIN_M = 0.15

# If a tide-model / topobathy datum mismatch exists, this value can
# later be tested systematically. It is NOT automatically applied.
DATUM_SHIFT_M = 0.0


# ---------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def ensure_topobathy():
    if TOPO_FILE.exists():
        print(f"Topobathy found: {TOPO_FILE.resolve()}")
        return

    print()
    print("=" * 90)
    print("TOPobathy file not found")
    print("=" * 90)
    print("Attempting download:")
    print(TOPO_URL)

    try:
        urllib.request.urlretrieve(
            TOPO_URL,
            TOPO_FILE,
        )
    except Exception as exc:
        print()
        print("DOWNLOAD FAILED:")
        print(exc)
        print()
        print("Please download the USGS file manually and place:")
        print(f"  {TOPO_FILE.resolve()}")
        print()
        print("Then rerun this script.")
        raise SystemExit(2)

    print(
        f"Downloaded {TOPO_FILE.resolve()}"
    )


def utc_hours_to_datetime(year, doy, utc_hours):
    day = (
        datetime(year, 1, 1)
        + timedelta(days=doy - 1)
    )

    return day + timedelta(
        hours=float(utc_hours)
    )


def load_tides():
    wb = load_workbook(
        TIDE_FILE,
        data_only=True,
    )

    ws = wb[wb.sheetnames[0]]

    header = [
        c.value for c in ws[1]
    ]

    tcol = header.index("time")

    mcols = {
        m: header.index(m)
        for m in MODELS
    }

    times = []
    values = {
        m: [] for m in MODELS
    }

    for row in ws.iter_rows(
        min_row=2,
        values_only=True,
    ):
        t = row[tcol]
        if not isinstance(t, datetime):
            continue

        good = True
        temp = {}

        for m in MODELS:
            v = finite(
                row[mcols[m]]
            )
            if v is None:
                good = False
                break
            temp[m] = v

        if not good:
            continue

        times.append(t)

        for m in MODELS:
            values[m].append(
                temp[m]
            )

    wb.close()

    if len(times) < 2:
        raise RuntimeError(
            "Insufficient tide model points."
        )

    epoch = np.array(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    funcs = {}

    for m in MODELS:
        arr = np.asarray(
            values[m],
            dtype=float,
        )

        def make_interp(arr):
            def f(dt):
                x = (
                    dt - times[0]
                ).total_seconds()

                if (
                    x < epoch[0]
                    or x > epoch[-1]
                ):
                    return math.nan

                return float(
                    np.interp(
                        x,
                        epoch,
                        arr,
                    )
                )
            return f

        funcs[m] = make_interp(arr)

    return times, funcs


def load_target_rows():
    rows = []

    for doy in DOYS:
        path = (
            RESULT_DIR
            / f"{doy}.txt"
        )

        if not path.exists():
            raise RuntimeError(
                f"Missing GNSS-IR result file: {path}"
            )

        for line in path.read_text(
            errors="replace"
        ).splitlines():

            line = line.strip()

            if (
                not line
                or line.startswith("%")
            ):
                continue

            c = line.split()

            if len(c) < 17:
                continue

            try:
                year = int(float(c[0]))
                doy2 = int(float(c[1]))
                rh = float(c[2])
                sat = int(float(c[3]))
                utc_hours = float(c[4])
                az = float(c[5])
                amp = float(c[6])
                emin = float(c[7])
                emax = float(c[8])
                freq = int(float(c[10]))
                rise = int(float(c[11]))
                pkn = float(c[13])
                delt = float(c[14])
            except Exception:
                continue

            if (sat, freq) not in TARGETS:
                continue

            dt = utc_hours_to_datetime(
                year,
                doy2,
                utc_hours,
            )

            rows.append(
                {
                    "year": year,
                    "doy": doy2,
                    "dt": dt,
                    "sat": sat,
                    "freq": freq,
                    "rh": rh,
                    "az": az,
                    "amp": amp,
                    "pkn": pkn,
                    "emin": emin,
                    "emax": emax,
                    "rise": rise,
                    "delt": delt,
                }
            )

    return rows


def build_ellipse_polygon(freq, elev, rh, az):
    lon, lat = makeEllipse_latlon(
        int(freq),
        float(elev),
        float(rh),
        float(az),
        LAT,
        LON,
    )

    return Polygon(
        list(
            zip(
                map(float, lon),
                map(float, lat),
            )
        )
    )


def tide_values_at(dt, funcs):
    return {
        m: funcs[m](dt)
        for m in MODELS
    }


def rasterize_polygon_statistics(
    src,
    polygon_lonlat,
    tide_value_m,
):
    """
    Sample all raster cells whose centers intersect the Fresnel polygon,
    using rasterio.mask.

    Classification:
      wet:  z <= tide + WET_MARGIN
      near: tide+wet < z <= tide+near
      dry:  z > tide+near

    No interpolation beyond raster cell centers is applied.
    """

    # Convert polygon WGS84 -> raster CRS.
    to_raster = Transformer.from_crs(
        "EPSG:4326",
        src.crs,
        always_xy=True,
    )

    poly_raster = shapely_transform(
        lambda x, y, z=None:
            to_raster.transform(x, y),
        polygon_lonlat,
    )

    try:
        arr, _ = mask(
            src,
            [mapping(poly_raster)],
            crop=True,
            filled=False,
        )
    except ValueError:
        return {
            "n": 0,
            "wet_fraction": math.nan,
            "near_fraction": math.nan,
            "dry_fraction": math.nan,
            "z_min": math.nan,
            "z_max": math.nan,
            "z_median": math.nan,
        }

    values = np.asarray(
        arr[0].compressed(),
        dtype=float,
    )

    if len(values) == 0:
        return {
            "n": 0,
            "wet_fraction": math.nan,
            "near_fraction": math.nan,
            "dry_fraction": math.nan,
            "z_min": math.nan,
            "z_max": math.nan,
            "z_median": math.nan,
        }

    wet_limit = (
        tide_value_m
        + DATUM_SHIFT_M
        + WET_MARGIN_M
    )

    near_limit = (
        tide_value_m
        + DATUM_SHIFT_M
        + NEAR_WATER_MARGIN_M
    )

    wet = values <= wet_limit
    near = (
        (values > wet_limit)
        & (values <= near_limit)
    )
    dry = values > near_limit

    return {
        "n": int(len(values)),
        "wet_fraction": float(
            np.mean(wet)
        ),
        "near_fraction": float(
            np.mean(near)
        ),
        "dry_fraction": float(
            np.mean(dry)
        ),
        "z_min": float(
            np.min(values)
        ),
        "z_max": float(
            np.max(values)
        ),
        "z_median": float(
            np.median(values)
        ),
    }


def water_class(wet, near):
    if not math.isfinite(wet):
        return "NO_RASTER_COVERAGE"

    if wet >= 0.80:
        return "OPEN_WATER_DOMINANT"

    if wet + near >= 0.80:
        return "WATERLINE_OR_INTERTIDAL_DOMINANT"

    if wet <= 0.20:
        return "LAND_DOMINANT"

    return "MIXED"


def local_center_signed_distance(poly):
    """
    Diagnostic only: distance of footprint centroid from the old
    83.06°/71.78 m reference line, included so the new method can
    be compared with the old half-plane result.
    """
    cx, cy = poly.centroid.x, poly.centroid.y

    R = 6371000.0

    east = (
        math.radians(cx - LON)
        * R
        * math.cos(
            math.radians(LAT)
        )
    )

    north = (
        math.radians(cy - LAT)
        * R
    )

    b = math.radians(83.06)

    along = (
        east * math.sin(b)
        + north * math.cos(b)
    )

    return along - 71.78


def kml_color(classification):
    if classification == "OPEN_WATER_DOMINANT":
        return simplekml.Color.blue

    if classification == "WATERLINE_OR_INTERTIDAL_DOMINANT":
        return simplekml.Color.yellow

    if classification == "LAND_DOMINANT":
        return simplekml.Color.red

    return simplekml.Color.gray


# ---------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------

def main():

    print()
    print("=" * 92)
    print(
        "MARCONI TRUE TOPOBATHY / FRESNEL / TIDE VALIDATION"
    )
    print("=" * 92)

    print(
        f"GNSS-IR results: {RESULT_DIR}"
    )
    print(
        f"Topobathy: {TOPO_FILE}"
    )
    print(
        f"Tide workbook: {TIDE_FILE}"
    )

    ensure_topobathy()

    tide_times, tide_funcs = load_tides()

    rows = load_target_rows()

    if not rows:
        raise SystemExit(
            "No target GNSS-IR observations found."
        )

    print()
    print(
        f"Target GNSS-IR observations: {len(rows)}"
    )

    print()
    print(
        "Tide-model datum is NOT changed automatically."
    )
    print(
        f"DATUM_SHIFT_M = {DATUM_SHIFT_M:+.3f} m"
    )

    print()
    print(
        "Target tracks:"
    )

    for key, label in TARGETS.items():
        print(
            f"  SAT={key[0]} FREQ={key[1]} {label}"
        )

    # Open raster once.
    with rasterio.open(
        TOPO_FILE
    ) as src:

        print()
        print(
            "TOPOBATHY CRS:",
            src.crs
        )

        print(
            "TOPOBATHY resolution:",
            src.res
        )

        print(
            "TOPOBATHY bounds:",
            src.bounds
        )

        print(
            "TOPOBATHY nodata:",
            src.nodata
        )

        # KML
        kml = simplekml.Kml()

        station = kml.newpoint(
            name="Marconi GNSS station"
        )

        station.coords = [
            (LON, LAT)
        ]

        folders = {}

        for key, label in TARGETS.items():
            folders[key] = (
                kml.newfolder(
                    name=label
                )
            )

        out_rows = []

        for r in rows:

            tide = tide_values_at(
                r["dt"],
                tide_funcs,
            )

            tide_mean = float(
                np.mean(
                    list(tide.values())
                )
            )

            print()
            print(
                f"SAT={r['sat']}"
                f" Az={r['az']:.2f}"
                f" RH={r['rh']:.3f}"
                f" UTC={r['dt']}"
            )

            for elev in ELEVATIONS:

                poly = build_ellipse_polygon(
                    r["freq"],
                    elev,
                    r["rh"],
                    r["az"],
                )

                # KML footprint.
                # Classification below uses each individual tide model
                # and also their ensemble mean.
                for model in MODELS + ["ENSEMBLE_MEAN"]:

                    water_level = (
                        tide_mean
                        if model == "ENSEMBLE_MEAN"
                        else tide[model]
                    )

                    stats = (
                        rasterize_polygon_statistics(
                            src,
                            poly,
                            water_level,
                        )
                    )

                    classification = water_class(
                        stats["wet_fraction"],
                        stats["near_fraction"],
                    )

                    out_rows.append(
                        {
                            "datetime_utc":
                                r["dt"].isoformat(),

                            "doy":
                                r["doy"],

                            "sat":
                                r["sat"],

                            "freq":
                                r["freq"],

                            "track":
                                TARGETS[
                                    (
                                        r["sat"],
                                        r["freq"],
                                    )
                                ],

                            "az_deg":
                                r["az"],

                            "RH_m":
                                r["rh"],

                            "PkNoise":
                                r["pkn"],

                            "Amp":
                                r["amp"],

                            "eminO_deg":
                                r["emin"],

                            "emaxO_deg":
                                r["emax"],

                            "test_elevation_deg":
                                elev,

                            "tide_model":
                                model,

                            "tide_elevation_m":
                                water_level,

                            "datum_shift_m":
                                DATUM_SHIFT_M,

                            "raster_cell_count":
                                stats["n"],

                            "wet_fraction":
                                stats[
                                    "wet_fraction"
                                ],

                            "near_water_fraction":
                                stats[
                                    "near_fraction"
                                ],

                            "dry_fraction":
                                stats[
                                    "dry_fraction"
                                ],

                            "raster_z_min_m":
                                stats["z_min"],

                            "raster_z_median_m":
                                stats["z_median"],

                            "raster_z_max_m":
                                stats["z_max"],

                            "classification":
                                classification,

                            "old_halfplane_center_signed_m":
                                local_center_signed_distance(
                                    poly
                                ),
                        }
                    )

                    # KML name
                    p = folders[
                        (
                            r["sat"],
                            r["freq"],
                        )
                    ].newpolygon(
                        name=(
                            f"{r['dt'].strftime('%m-%d %H:%M')} "
                            f"EL={elev:.0f} "
                            f"{model} "
                            f"wet={stats['wet_fraction']:.2f}"
                        )
                    )

                    coords = [
                        (
                            float(x),
                            float(y)
                        )
                        for x, y in poly.exterior.coords
                    ]

                    p.outerboundaryis = coords

                    style = simplekml.Style()
                    color = kml_color(
                        classification
                    )

                    style.polystyle.color = (
                        simplekml.Color.changealphaint(
                            90,
                            color,
                        )
                    )

                    style.linestyle.color = color
                    style.linestyle.width = 2

                    p.style = style

    # Write CSV
    fields = list(
        out_rows[0].keys()
    )

    with open(
        OUT_CSV,
        "w",
        newline="",
    ) as f:
        w = csv.DictWriter(
            f,
            fieldnames=fields,
        )
        w.writeheader()
        w.writerows(out_rows)

    kml.save(
        str(OUT_KML)
    )

    # Summary focused on ensemble mean.
    summary = []

    summary.append(
        "MARCONI TRUE TOPOBATHY / FRESNEL / TIDE VALIDATION"
    )
    summary.append(
        "=" * 92
    )
    summary.append(
        f"Topobathy file: {TOPO_FILE}"
    )
    summary.append(
        f"DATUM_SHIFT_M: {DATUM_SHIFT_M:+.3f} m"
    )
    summary.append(
        ""
    )
    summary.append(
        "ENSEMBLE-MEAN TIDE RESULTS"
    )
    summary.append(
        "-" * 92
    )

    for key, label in TARGETS.items():

        for elev in ELEVATIONS:

            subset = [
                x for x in out_rows
                if (
                    x["sat"] == key[0]
                    and x["freq"] == key[1]
                    and x["test_elevation_deg"] == elev
                    and x["tide_model"] == "ENSEMBLE_MEAN"
                )
            ]

            if not subset:
                continue

            wet = np.array(
                [
                    x["wet_fraction"]
                    for x in subset
                ],
                dtype=float,
            )

            near = np.array(
                [
                    x["near_water_fraction"]
                    for x in subset
                ],
                dtype=float,
            )

            dry = np.array(
                [
                    x["dry_fraction"]
                    for x in subset
                ],
                dtype=float,
            )

            summary.append(
                f"{label:16s}"
                f" EL={elev:4.0f}°"
                f" N={len(subset):2d}"
                f" wet_mean={np.mean(wet):.3f}"
                f" wet_min={np.min(wet):.3f}"
                f" wet_max={np.max(wet):.3f}"
                f" near_mean={np.mean(near):.3f}"
                f" dry_mean={np.mean(dry):.3f}"
            )

    summary.append(
        ""
    )
    summary.append(
        "INTERPRETATION"
    )
    summary.append(
        "This is a baseline test using the 2021 USGS topobathy surface."
    )
    summary.append(
        "A high wet fraction means the Fresnel footprint lies mostly"
    )
    summary.append(
        "below the modeled water surface elevation in the raster."
    )
    summary.append(
        "A high near-water fraction indicates intertidal/waterline"
    )
    summary.append(
        "sensitivity and should not be called open ocean automatically."
    )
    summary.append(
        "A datum mismatch must be resolved before treating absolute"
    )
    summary.append(
        "wet/dry classification as final."
    )

    OUT_SUMMARY.write_text(
        "\n".join(summary)
        + "\n"
    )

    print()
    print("=" * 92)
    print("OUTPUTS")
    print("=" * 92)
    print(
        "CSV    :",
        OUT_CSV.resolve()
    )
    print(
        "KML    :",
        OUT_KML.resolve()
    )
    print(
        "Summary:",
        OUT_SUMMARY.resolve()
    )
    print()
    print("DONE")


if __name__ == "__main__":
    main()
