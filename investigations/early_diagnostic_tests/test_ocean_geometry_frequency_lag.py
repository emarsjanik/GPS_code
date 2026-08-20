#!/usr/bin/env python3

import csv
import math
import re
import datetime as dt
from pathlib import Path
from collections import defaultdict

import numpy as np
from openpyxl import load_workbook


# ================================================================
# OCEAN GNSS-IR GEOMETRY / FREQUENCY / LAG TEST
#
# Tests whether the apparent GNSS-IR/tide relationship is stronger
# for particular reflection azimuths, frequencies, and time lags.
#
# Lag convention:
#   lag = +30 min means evaluate the tide 30 minutes AFTER the
#   GNSS-IR solution time.
#   lag = -30 min means evaluate the tide 30 minutes BEFORE it.
#
# RH is used as the primary GNSS-IR variable.
# GNSS water level = H_reference - RH therefore has the opposite r.
# ================================================================


YEAR = 2026
DOY1 = 204
DOY2 = 207

CSV_PATH = Path("gnssir_tide_arc_analysis.csv")
LOG_BASE = Path("products/refl_code/logs/usgs/ocean90_150/2026")
TIDE_PATH = Path("marconi_tides_sherwood.xlsx")

MATCH_LIMIT_SEC = 90.0

SECTORS = [
    ("ALL", 90.0, 150.0),
    ("100-150", 100.0, 150.0),
    ("100-140", 100.0, 140.0),
    ("100-130", 100.0, 130.0),
    ("110-130", 110.0, 130.0),
    ("110-120", 110.0, 120.0),
    ("112-115", 112.0, 115.0),
]

# None means all frequencies.
FREQUENCIES = [
    ("ALL", None),
    ("1", 1),
    ("101", 101),
    ("201", 201),
    ("205", 205),
    ("207", 207),
    ("302", 302),
]

LAGS_MIN = list(range(-60, 61, 5))


# ---------------------------------------------------------------
# Utility
# ---------------------------------------------------------------

def pearson(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    valid = np.isfinite(x) & np.isfinite(y)
    x = x[valid]
    y = y[valid]

    if len(x) < 3:
        return float("nan")

    sx = np.std(x)
    sy = np.std(y)

    if sx == 0 or sy == 0:
        return float("nan")

    return float(np.corrcoef(x, y)[0, 1])


def median_abs(x):
    x = np.asarray(x, dtype=float)
    x = x[np.isfinite(x)]
    return float(np.median(np.abs(x))) if len(x) else float("nan")


# ---------------------------------------------------------------
# Read tide model
# ---------------------------------------------------------------

def load_tide_models(path):
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("Tide workbook is empty.")

    header = rows[0]

    time_col = header.index("time")

    model_cols = []
    for i, name in enumerate(header):
        if isinstance(name, str) and name.endswith("_heightm"):
            model_cols.append((i, name))

    if not model_cols:
        raise RuntimeError("No *_heightm tide model columns found.")

    times = []
    values = []

    for row in rows[1:]:
        if not row or row[time_col] is None:
            continue

        t = row[time_col]
        if not isinstance(t, dt.datetime):
            continue

        vals = []
        good = True

        for col, name in model_cols:
            try:
                v = float(row[col])
            except (TypeError, ValueError):
                good = False
                break

            if not math.isfinite(v):
                good = False
                break

            vals.append(v)

        if good:
            times.append(t)
            values.append(float(np.mean(vals)))

    if len(times) < 2:
        raise RuntimeError("Insufficient tide model data.")

    # Ensure chronological order.
    order = np.argsort(np.array(times, dtype="datetime64[ns]"))
    times = [times[i] for i in order]
    values = np.asarray([values[i] for i in order], dtype=float)

    t0 = times[0]
    seconds = np.array(
        [(t - t0).total_seconds() for t in times],
        dtype=float,
    )

    print("Tide model columns:")
    for _, name in model_cols:
        print(f"  {name}")

    print(f"Tide points loaded: {len(times)}")
    print(f"Tide coverage: {times[0]} through {times[-1]}")

    def interpolate(query_time):
        q = (query_time - t0).total_seconds()

        if q < seconds[0] or q > seconds[-1]:
            return float("nan")

        return float(np.interp(q, seconds, values))

    return interpolate


# ---------------------------------------------------------------
# Read GNSS-IR SUCCESS records
# ---------------------------------------------------------------

SUCCESS_RE = re.compile(
    r"SUCCESS\s+Azimuth\s+([0-9.]+)\s+"
    r"(?:Sat|Satellite)\s+(\d+)\s+"
    r"RH\s+([-+0-9.]+)\s+m\s+"
    r"PkNoise\s+([-+0-9.]+)\s+"
    r"Amp\s+([-+0-9.]+)\s+"
    r"Fr\s*(\d+)\s+"
    r"UTC\s+([0-9:]+)\s+"
    r"DT\s+([0-9.]+)"
)


def load_success_records():
    records = []

    for doy in range(DOY1, DOY2 + 1):
        path = LOG_BASE / f"{doy}_gnssir.txt"

        if not path.exists():
            print(f"WARNING: missing log: {path}")
            continue

        day = (
            dt.datetime(YEAR, 1, 1)
            + dt.timedelta(days=doy - 1)
        ).date()

        with open(path, encoding="utf-8", errors="replace") as f:
            for line in f:
                m = SUCCESS_RE.search(line)

                if not m:
                    continue

                # Accept both HH:MM:SS and HH:MM in gnssrefl logs.
                utc_text = m.group(7)
                utc_parts = utc_text.split(":")

                if len(utc_parts) == 3:
                    hh, mm, ss = map(int, utc_parts)
                elif len(utc_parts) == 2:
                    hh, mm = map(int, utc_parts)
                    ss = 0
                else:
                    continue

                if not (
                    0 <= hh <= 23
                    and 0 <= mm <= 59
                    and 0 <= ss <= 59
                ):
                    continue

                records.append({
                    "doy": doy,
                    "dt": dt.datetime.combine(
                        day,
                        dt.time(hh, mm, ss),
                    ),
                    "utc": utc_text,
                    "az": float(m.group(1)),
                    "sat": int(m.group(2)),
                    "rh": float(m.group(3)),
                    "pkn": float(m.group(4)),
                    "amp": float(m.group(5)),
                    "freq": int(m.group(6)),
                    "dt_arc": float(m.group(8)),
                })

    return records


# ---------------------------------------------------------------
# Read CSV and match each SUCCESS record to the corresponding
# same-day satellite/frequency tide-analysis record.
# ---------------------------------------------------------------

def load_csv_records(path):
    out = []

    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for row in reader:
            try:
                solution_time = dt.datetime.fromisoformat(
                    row["solution_time_utc"]
                )

                sat = int(float(row["sat"]))
                freq = int(float(row["freq"]))
                tide = float(row["tide_solution_m"])
                gnss = float(row["gnss_water_level_m"])

                if not all(math.isfinite(x) for x in (tide, gnss)):
                    continue

                out.append({
                    "dt": solution_time,
                    "sat": sat,
                    "freq": freq,
                    "tide": tide,
                    "gnss": gnss,
                })

            except (KeyError, TypeError, ValueError):
                continue

    return out


def match_success_to_csv(success, csv_records):
    matched = []
    unmatched = []

    for s in success:
        candidates = [
            r for r in csv_records
            if (
                r["sat"] == s["sat"]
                and r["freq"] == s["freq"]
                and r["dt"].date() == s["dt"].date()
            )
        ]

        if not candidates:
            unmatched.append(s)
            continue

        best = min(
            candidates,
            key=lambda r: abs(
                (r["dt"] - s["dt"]).total_seconds()
            ),
        )

        delta = abs(
            (best["dt"] - s["dt"]).total_seconds()
        )

        if delta <= MATCH_LIMIT_SEC:
            item = dict(s)
            item["csv_dt"] = best["dt"]
            item["match_delta_sec"] = delta
            item["tide_at_solution"] = best["tide"]
            item["gnss_csv"] = best["gnss"]
            matched.append(item)
        else:
            unmatched.append(s)

    return matched, unmatched


# ---------------------------------------------------------------
# Main
# ---------------------------------------------------------------

print()
print("=" * 80)
print("GNSS-IR OCEAN GEOMETRY / FREQUENCY / LAG TEST")
print("=" * 80)
print()
print(f"DOY range       : {DOY1}-{DOY2}")
print(f"SUCCESS logs    : {LOG_BASE}")
print(f"Tide workbook   : {TIDE_PATH}")
print(f"Match tolerance : {MATCH_LIMIT_SEC:.0f} sec")
print(f"Lag range       : {LAGS_MIN[0]} to {LAGS_MIN[-1]} min")
print(f"Lag increment   : {LAGS_MIN[1] - LAGS_MIN[0]} min")

if not CSV_PATH.exists():
    raise SystemExit(f"ERROR: CSV not found: {CSV_PATH}")

if not TIDE_PATH.exists():
    raise SystemExit(f"ERROR: tide workbook not found: {TIDE_PATH}")

tide_at = load_tide_models(TIDE_PATH)

success = load_success_records()
csv_records = load_csv_records(CSV_PATH)

matched, unmatched = match_success_to_csv(
    success,
    csv_records,
)

print()
print("=" * 80)
print("MATCHING")
print("=" * 80)
print(f"SUCCESS records : {len(success)}")
print(f"Matched         : {len(matched)}")
print(f"Unmatched       : {len(unmatched)}")

if len(matched) < 3:
    raise SystemExit("ERROR: fewer than 3 matched records.")

# Add the tide value at every tested lag.
for r in matched:
    r["tide_lag"] = {}

    for lag in LAGS_MIN:
        query_time = r["dt"] + dt.timedelta(minutes=lag)
        r["tide_lag"][lag] = tide_at(query_time)


# ---------------------------------------------------------------
# Summary of matched data
# ---------------------------------------------------------------

print()
print("=" * 80)
print("MATCHED DATASET")
print("=" * 80)

for r in sorted(matched, key=lambda x: x["dt"]):
    print(
        f"{r['dt'].strftime('%m-%d %H:%M:%S')} "
        f"Az={r['az']:6.1f} "
        f"sat={r['sat']:3d} "
        f"freq={r['freq']:3d} "
        f"RH={r['rh']:7.3f} "
        f"match={r['match_delta_sec']:5.1f}s"
    )


# ---------------------------------------------------------------
# Run all sector/frequency/lag combinations.
# ---------------------------------------------------------------

results = []

for sector_name, az_lo, az_hi in SECTORS:

    for freq_name, freq_value in FREQUENCIES:

        subset = [
            r for r in matched
            if (
                az_lo <= r["az"] <= az_hi
                and (
                    freq_value is None
                    or r["freq"] == freq_value
                )
            )
        ]

        if len(subset) < 3:
            continue

        zero_tide = [
            r["tide_lag"][0]
            for r in subset
        ]

        zero_r = pearson(
            [r["rh"] for r in subset],
            zero_tide,
        )

        for lag in LAGS_MIN:

            tide_values = [
                r["tide_lag"][lag]
                for r in subset
            ]

            r_value = pearson(
                [r["rh"] for r in subset],
                tide_values,
            )

            if not math.isfinite(r_value):
                continue

            results.append({
                "sector": sector_name,
                "freq": freq_name,
                "n": len(subset),
                "lag": lag,
                "r": r_value,
                "abs_r": abs(r_value),
                "zero_r": zero_r,
            })


# ---------------------------------------------------------------
# Best result for each sector/frequency combination
# ---------------------------------------------------------------

print()
print("=" * 80)
print("BEST LAG BY SECTOR / FREQUENCY")
print("=" * 80)
print()
print(
    f"{'SECTOR':<10} {'FREQ':>5} {'N':>4} "
    f"{'ZERO-r':>9} {'BEST LAG':>9} {'BEST r':>9}"
)
print("-" * 80)

best_groups = {}

for key in sorted(
    set((x["sector"], x["freq"]) for x in results)
):
    group = [
        x for x in results
        if x["sector"] == key[0]
        and x["freq"] == key[1]
    ]

    # Highest absolute correlation.
    best = max(
        group,
        key=lambda x: abs(x["r"]),
    )

    best_groups[key] = best

    print(
        f"{best['sector']:<10} "
        f"{best['freq']:>5} "
        f"{best['n']:>4} "
        f"{best['zero_r']:>9.4f} "
        f"{best['lag']:>+9d} "
        f"{best['r']:>9.4f}"
    )


# ---------------------------------------------------------------
# Sector-only results: all frequencies combined
# ---------------------------------------------------------------

print()
print("=" * 80)
print("SECTOR TEST — ALL FREQUENCIES")
print("=" * 80)
print()
print(
    f"{'SECTOR':<10} {'N':>4} "
    f"{'ZERO-r':>9} {'BEST LAG':>9} {'BEST r':>9}"
)
print("-" * 60)

sector_best = {}

for sector_name, _, _ in SECTORS:

    group = [
        x for x in results
        if x["sector"] == sector_name
        and x["freq"] == "ALL"
    ]

    if not group:
        continue

    best = max(
        group,
        key=lambda x: abs(x["r"]),
    )

    sector_best[sector_name] = best

    print(
        f"{sector_name:<10} "
        f"{best['n']:>4} "
        f"{best['zero_r']:>9.4f} "
        f"{best['lag']:>+9d} "
        f"{best['r']:>9.4f}"
    )


# ---------------------------------------------------------------
# Frequency-only results — all azimuths
# ---------------------------------------------------------------

print()
print("=" * 80)
print("FREQUENCY TEST — ALL AZIMUTHS")
print("=" * 80)
print()
print(
    f"{'FREQ':>6} {'N':>4} "
    f"{'ZERO-r':>9} {'BEST LAG':>9} {'BEST r':>9}"
)
print("-" * 55)

freq_best = {}

for freq_name, _ in FREQUENCIES:

    if freq_name == "ALL":
        continue

    group = [
        x for x in results
        if x["sector"] == "ALL"
        and x["freq"] == freq_name
    ]

    if not group:
        continue

    best = max(
        group,
        key=lambda x: abs(x["r"]),
    )

    freq_best[freq_name] = best

    print(
        f"{freq_name:>6} "
        f"{best['n']:>4} "
        f"{best['zero_r']:>9.4f} "
        f"{best['lag']:>+9d} "
        f"{best['r']:>9.4f}"
    )


# ---------------------------------------------------------------
# Global best results, but require n >= 8 to avoid overinterpreting
# tiny frequency groups.
# ---------------------------------------------------------------

print()
print("=" * 80)
print("STRONGEST RESULTS WITH n >= 8")
print("=" * 80)

large_results = [
    x for x in results
    if x["n"] >= 8
]

if large_results:

    top = sorted(
        large_results,
        key=lambda x: abs(x["r"]),
        reverse=True,
    )[:15]

    for x in top:
        print(
            f"sector={x['sector']:<10} "
            f"freq={x['freq']:>3} "
            f"n={x['n']:2d} "
            f"lag={x['lag']:+4d} min "
            f"r={x['r']:+.4f} "
            f"zero_r={x['zero_r']:+.4f}"
        )

else:
    print("No sector/frequency combination has n >= 8.")


# ---------------------------------------------------------------
# Explicit geometry comparison
# ---------------------------------------------------------------

print()
print("=" * 80)
print("GEOMETRY COMPARISON")
print("=" * 80)

for name in [
    "ALL",
    "100-130",
    "110-130",
    "110-120",
    "112-115",
]:

    key = (name, "ALL")

    if key not in best_groups:
        continue

    x = best_groups[key]

    print(
        f"{name:<10} "
        f"n={x['n']:2d} "
        f"zero-lag r={x['zero_r']:+.4f} "
        f"best lag={x['lag']:+4d} min "
        f"best r={x['r']:+.4f}"
    )


# ---------------------------------------------------------------
# Save machine-readable results
# ---------------------------------------------------------------

out_path = Path("ocean_geometry_frequency_lag_results.csv")

with open(out_path, "w", newline="") as f:
    writer = csv.writer(f)

    writer.writerow([
        "sector",
        "freq",
        "n",
        "lag_min",
        "r_rh_vs_tide",
        "abs_r",
        "zero_lag_r",
    ])

    for x in sorted(
        results,
        key=lambda x: (
            x["sector"],
            x["freq"],
            x["lag"],
        ),
    ):
        writer.writerow([
            x["sector"],
            x["freq"],
            x["n"],
            x["lag"],
            f"{x['r']:.8f}",
            f"{x['abs_r']:.8f}",
            f"{x['zero_r']:.8f}",
        ])


# ---------------------------------------------------------------
# Final interpretation
# ---------------------------------------------------------------

print()
print("=" * 80)
print("INTERPRETATION GUIDE")
print("=" * 80)

print("""
1. ZERO-r tells us the relationship at the actual GNSS-IR
   solution epoch.

2. BEST r searches +/-60 minutes in 5-minute increments.

3. A substantially stronger correlation in 110-120 degrees
   than in ALL/100-150 would support a reflection-geometry
   explanation.

4. A similar best lag across several azimuth sectors would
   provide more evidence for a timing/physical-lag effect.

5. A strong result confined to one frequency with only 3-5
   observations should NOT be treated as established.

6. The sign is reported for RH versus tide. The physical
   GNSS water-level proxy H-RH has the opposite sign.

7. This is an exploratory test. With only 19 matched
   observations, lag maxima can move substantially when a
   single observation is removed.
""")

print()
print(f"Results written to: {out_path}")
print()
print("=" * 80)
print("DONE")
print("=" * 80)
