import csv
import math
import re
from pathlib import Path
from datetime import datetime, timedelta

from openpyxl import load_workbook
import numpy as np


# ================================================================
# CONFIGURATION
# ================================================================

CSV_PATH = Path("gnssir_tide_arc_analysis.csv")

TIDE_PATH = Path(
    "marconi_tides_sherwood.xlsx"
)

LOG_BASE = Path(
    "products/refl_code/logs/usgs/ocean90_150/2026"
)

YEAR = 2026

DOY_START = 204
DOY_END = 207

AZ_MIN = 100.0
AZ_MAX = 150.0

MATCH_LIMIT_SECONDS = 90.0

LAG_MINUTES_START = -90
LAG_MINUTES_END = 90
LAG_STEP_MINUTES = 5


# ================================================================
# HEADER
# ================================================================

print()
print("=" * 80)
print("CORRECTED OCEAN-FACING GNSS-IR LAG TEST")
print("=" * 80)

print()
print("This test:")
print("  1. Reads the original 10-minute tide-model data.")
print("  2. Reads the ocean-facing SUCCESS records.")
print("  3. Matches SUCCESS records to the established CSV records.")
print("  4. Uses the continuous tide model for every lag.")
print("  5. Never extrapolates or clamps tide values.")
print()
print(
    f"Azimuth sector: {AZ_MIN:.0f}-{AZ_MAX:.0f} degrees"
)
print(
    f"DOY range     : {DOY_START}-{DOY_END}"
)
print(
    f"Lag range     : "
    f"{LAG_MINUTES_START:+d} to "
    f"{LAG_MINUTES_END:+d} minutes"
)
print()


# ================================================================
# UTILITY FUNCTIONS
# ================================================================

def parse_utc_time(text):
    """
    Parse either HH:MM or HH:MM:SS.
    """

    parts = text.strip().split(":")

    if len(parts) == 2:

        hh = int(parts[0])
        mm = int(parts[1])
        ss = 0

    elif len(parts) == 3:

        hh = int(parts[0])
        mm = int(parts[1])
        ss = int(float(parts[2]))

    else:

        raise ValueError(
            f"Unexpected UTC time format: {text}"
        )

    return hh, mm, ss


def doy_to_date(year, doy):

    return (
        datetime(
            year,
            1,
            1
        )
        + timedelta(
            days=doy - 1
        )
    ).date()


def correlation(x, y):

    if len(x) < 3:
        return float("nan")

    mx = sum(x) / len(x)
    my = sum(y) / len(y)

    numerator = sum(
        (a - mx) * (b - my)
        for a, b in zip(x, y)
    )

    sx = math.sqrt(
        sum(
            (a - mx) ** 2
            for a in x
        )
    )

    sy = math.sqrt(
        sum(
            (b - my) ** 2
            for b in y
        )
    )

    if sx == 0 or sy == 0:
        return float("nan")

    return numerator / (sx * sy)


# ================================================================
# READ TIDE MODEL
# ================================================================

print("=" * 80)
print("READING CONTINUOUS TIDE MODEL")
print("=" * 80)

if not TIDE_PATH.exists():

    raise FileNotFoundError(
        f"Tide file not found:\n{TIDE_PATH.resolve()}"
    )


wb = load_workbook(
    TIDE_PATH,
    data_only=True
)

print()
print("Workbook sheets:")

for name in wb.sheetnames:

    print(
        f"  {name}"
    )


ws = wb[wb.sheetnames[0]]

headers = [
    cell.value
    for cell in ws[1]
]


model_columns = []

for i, header in enumerate(headers):

    if (
        isinstance(header, str)
        and header.endswith("_heightm")
    ):

        model_columns.append(i)


if not model_columns:

    raise RuntimeError(
        "No *_heightm columns found."
    )


print()
print("Tide model columns:")

for i in model_columns:

    print(
        f"  {headers[i]}"
    )


tide_times = []
tide_values = []


for row in ws.iter_rows(
    min_row=2,
    values_only=True
):

    timestamp = row[0]

    if not isinstance(
        timestamp,
        datetime
    ):

        continue

    values = []

    valid = True

    for column in model_columns:

        value = row[column]

        try:

            value = float(value)

        except (
            ValueError,
            TypeError
        ):

            valid = False
            break

        if not math.isfinite(value):

            valid = False
            break

        values.append(value)


    if not valid:

        continue


    # Ensemble = arithmetic mean of the four tide models.

    ensemble = (
        sum(values)
        /
        len(values)
    )


    tide_times.append(
        timestamp
    )

    tide_values.append(
        ensemble
    )


if len(tide_times) < 2:

    raise RuntimeError(
        "Insufficient tide-model data."
    )


print()
print(
    f"Tide points loaded: "
    f"{len(tide_times)}"
)

print(
    f"Tide coverage: "
    f"{tide_times[0]} "
    f"through "
    f"{tide_times[-1]}"
)


# ================================================================
# PREPARE TIDE INTERPOLATOR
# ================================================================

tide_reference_time = tide_times[0]

tide_seconds = np.array(
    [
        (
            t
            -
            tide_reference_time
        ).total_seconds()
        for t in tide_times
    ],
    dtype=float
)


tide_array = np.array(
    tide_values,
    dtype=float
)


def tide_at(query_time):

    seconds = (
        query_time
        -
        tide_reference_time
    ).total_seconds()


    # IMPORTANT:
    # Do not extrapolate.
    # Do not clamp to an endpoint.

    if (
        seconds < tide_seconds[0]
        or
        seconds > tide_seconds[-1]
    ):

        return float("nan")


    return float(
        np.interp(
            seconds,
            tide_seconds,
            tide_array
        )
    )


# ================================================================
# READ OCEAN-FACING SUCCESS RECORDS
# ================================================================

print()
print("=" * 80)
print("READING OCEAN-FACING GNSS-IR SUCCESS RECORDS")
print("=" * 80)


success_pattern = re.compile(
    r"SUCCESS\s+"
    r"Azimuth\s+([0-9.]+)\s+"
    r"(?:Sat|Satellite)\s+(\d+)\s+"
    r"RH\s+([-+0-9.]+)\s+m\s+"
    r"PkNoise\s+([-+0-9.]+)\s+"
    r"Amp\s+([-+0-9.]+)\s+"
    r"Fr\s*(\d+)\s+"
    r"UTC\s+([0-9:]+)\s+"
    r"DT\s+([0-9.]+)"
)


success_records = []


for doy in range(
    DOY_START,
    DOY_END + 1
):

    log_path = (
        LOG_BASE
        /
        f"{doy}_gnssir.txt"
    )


    if not log_path.exists():

        print()
        print(
            "WARNING: log file not found:"
        )

        print(
            f"  {log_path}"
        )

        continue


    day = doy_to_date(
        YEAR,
        doy
    )


    with open(
        log_path,
        encoding="utf-8",
        errors="replace"
    ) as f:

        for line in f:

            match = (
                success_pattern.search(
                    line
                )
            )


            if not match:

                continue


            azimuth = float(
                match.group(1)
            )

            satellite = int(
                match.group(2)
            )

            rh = float(
                match.group(3)
            )

            frequency = int(
                match.group(6)
            )

            utc_text = match.group(7)

            dt_minutes = float(
                match.group(8)
            )


            # Ocean-facing sector.

            if not (
                AZ_MIN
                <= azimuth
                <= AZ_MAX
            ):

                continue


            hh, mm, ss = (
                parse_utc_time(
                    utc_text
                )
            )


            solution_time = datetime.combine(
                day,
                datetime.min.time()
            ) + timedelta(
                hours=hh,
                minutes=mm,
                seconds=ss
            )


            success_records.append({

                "doy": doy,

                "dt": solution_time,

                "utc": utc_text,

                "az": azimuth,

                "sat": satellite,

                "freq": frequency,

                "rh": rh,

                "dt_minutes": dt_minutes,
            })


print()
print(
    f"Ocean-facing SUCCESS records: "
    f"{len(success_records)}"
)


# ================================================================
# READ ESTABLISHED CSV DATASET
# ================================================================

print()
print("=" * 80)
print("MATCHING TO ESTABLISHED CSV DATASET")
print("=" * 80)


if not CSV_PATH.exists():

    raise FileNotFoundError(
        f"CSV not found:\n{CSV_PATH.resolve()}"
    )


csv_records = []


with open(
    CSV_PATH,
    newline="",
    encoding="utf-8",
    errors="replace"
) as f:

    reader = csv.DictReader(f)

    for row in reader:

        try:

            dt = datetime.fromisoformat(
                row[
                    "solution_time_utc"
                ]
            )

            sat = int(
                float(
                    row["sat"]
                )
            )

            freq = int(
                float(
                    row["freq"]
                )
            )

            az = float(
                row["Azim"]
            )

        except (
            ValueError,
            KeyError
        ):

            continue


        if not (
            AZ_MIN
            <= az
            <= AZ_MAX
        ):

            continue


        csv_records.append({

            "dt": dt,

            "sat": sat,

            "freq": freq,

            "az": az,
        })


# ================================================================
# MATCH SUCCESS -> CSV
# ================================================================

matched = []
unmatched = []


for success in success_records:

    candidates = [

        row

        for row in csv_records

        if (
            row["sat"]
            ==
            success["sat"]
        )

        and
        (
            row["freq"]
            ==
            success["freq"]
        )

        and
        (
            row["dt"].date()
            ==
            success["dt"].date()
        )

    ]


    if not candidates:

        unmatched.append(
            (
                success,
                None
            )
        )

        continue


    best = min(

        candidates,

        key=lambda row:
        abs(
            (
                row["dt"]
                -
                success["dt"]
            ).total_seconds()
        )

    )


    difference_seconds = abs(
        (
            best["dt"]
            -
            success["dt"]
        ).total_seconds()
    )


    if (
        difference_seconds
        <=
        MATCH_LIMIT_SECONDS
    ):

        matched.append({

            "dt": best["dt"],

            "rh": success["rh"],

            "az": success["az"],

            "sat": success["sat"],

            "freq": success["freq"],

            "doy": success["doy"],

            "utc": success["utc"],

            "match_delta":
                difference_seconds,
        })

    else:

        unmatched.append(
            (
                success,
                best
            )
        )


print()
print(
    f"SUCCESS records : "
    f"{len(success_records)}"
)

print(
    f"Matched         : "
    f"{len(matched)}"
)

print(
    f"Unmatched       : "
    f"{len(unmatched)}"
)


# ================================================================
# PRINT MATCHED RECORDS
# ================================================================

print()
print("=" * 80)
print("MATCHED RECORDS")
print("=" * 80)


for row in sorted(
    matched,
    key=lambda x: x["dt"]
):

    print(
        f"DOY={row['doy']} "
        f"UTC={row['utc']:>8} "
        f"Az={row['az']:6.1f} "
        f"Sat={row['sat']:3d} "
        f"Freq={row['freq']:3d} "
        f"RH={row['rh']:7.3f} "
        f"CSVdelta={row['match_delta']:5.1f}s"
    )


# ================================================================
# LAG TEST
# ================================================================

print()
print("=" * 80)
print("CONTINUOUS TIDE LAG TEST")
print("=" * 80)

print()
print(
    "Positive lag means:"
)

print(
    "  evaluate tide AFTER the GNSS solution time."
)

print()

print(
    "Negative lag means:"
)

print(
    "  evaluate tide BEFORE the GNSS solution time."
)

print()


results = []


for lag_minutes in range(
    LAG_MINUTES_START,
    LAG_MINUTES_END + 1,
    LAG_STEP_MINUTES
):

    lag = timedelta(
        minutes=lag_minutes
    )


    rh_values = []
    tide_values_at_lag = []


    for row in matched:

        shifted_time = (
            row["dt"]
            +
            lag
        )


        tide = tide_at(
            shifted_time
        )


        # No extrapolation.

        if not math.isfinite(
            tide
        ):

            continue


        rh_values.append(
            row["rh"]
        )

        tide_values_at_lag.append(
            tide
        )


    r_value = correlation(
        rh_values,
        tide_values_at_lag
    )


    results.append({

        "lag":
            lag_minutes,

        "r":
            r_value,

        "n":
            len(rh_values),
    })


# ================================================================
# PRINT LAG TABLE
# ================================================================

print(
    f"{'Lag(min)':>10} "
    f"{'N':>5} "
    f"{'Correlation':>14}"
)

print("-" * 35)


for result in results:

    print(
        f"{result['lag']:>+10d} "
        f"{result['n']:>5d} "
        f"{result['r']:>+14.4f}"
    )


# ================================================================
# BEST RESULTS
# ================================================================

valid_results = [

    result

    for result in results

    if math.isfinite(
        result["r"]
    )

]


best_positive = max(
    valid_results,
    key=lambda x: x["r"]
)


best_negative = min(
    valid_results,
    key=lambda x: x["r"]
)


best_absolute = max(
    valid_results,
    key=lambda x:
    abs(x["r"])
)


zero_lag = next(
    result
    for result in results
    if result["lag"] == 0
)


# ================================================================
# SUMMARY
# ================================================================

print()
print("=" * 80)
print("LAG TEST SUMMARY")
print("=" * 80)


print()
print("ZERO LAG")

print(
    f"  Lag = "
    f"{zero_lag['lag']:+d} min"
)

print(
    f"  r   = "
    f"{zero_lag['r']:+.4f}"
)

print(
    f"  n   = "
    f"{zero_lag['n']}"
)


print()
print("BEST POSITIVE CORRELATION")

print(
    f"  Lag = "
    f"{best_positive['lag']:+d} min"
)

print(
    f"  r   = "
    f"{best_positive['r']:+.4f}"
)

print(
    f"  n   = "
    f"{best_positive['n']}"
)


print()
print("MOST NEGATIVE CORRELATION")

print(
    f"  Lag = "
    f"{best_negative['lag']:+d} min"
)

print(
    f"  r   = "
    f"{best_negative['r']:+.4f}"
)

print(
    f"  n   = "
    f"{best_negative['n']}"
)


print()
print("STRONGEST ABSOLUTE CORRELATION")

print(
    f"  Lag = "
    f"{best_absolute['lag']:+d} min"
)

print(
    f"  r   = "
    f"{best_absolute['r']:+.4f}"
)

print(
    f"  n   = "
    f"{best_absolute['n']}"
)


# ================================================================
# CHECK THE SIGN-REVERSED RELATIONSHIP
# ================================================================

print()
print("=" * 80)
print("SIGN-REVERSED CHECK")
print("=" * 80)


print(
    "Because GNSS water level = H - RH,"
)

print(
    "RH and GNSS water level have opposite"
)

print(
    "correlations with the same tide series."
)

print()

print(
    f"At zero lag:"
)

print(
    f"  RH  vs tide = "
    f"{zero_lag['r']:+.4f}"
)

print(
    f"  RH  vs -tide = "
    f"{-zero_lag['r']:+.4f}"
)


# ================================================================
# FINAL DATA RANGE CHECK
# ================================================================

rh_all = [
    row["rh"]
    for row in matched
]


tide_zero = [

    tide_at(
        row["dt"]
    )

    for row in matched

]


tide_zero = [

    x

    for x in tide_zero

    if math.isfinite(x)

]


print()
print("=" * 80)
print("DATA RANGES")
print("=" * 80)


if rh_all:

    print(
        f"RH range       : "
        f"{min(rh_all):.3f} "
        f"to "
        f"{max(rh_all):.3f} m"
    )


if tide_zero:

    print(
        f"Tide range     : "
        f"{min(tide_zero):.3f} "
        f"to "
        f"{max(tide_zero):.3f} m"
    )


# ================================================================
# DONE
# ================================================================

print()
print("=" * 80)
print("DONE")
print("=" * 80)
print()
