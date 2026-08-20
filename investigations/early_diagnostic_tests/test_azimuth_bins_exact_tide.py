import csv
import math
import re
from pathlib import Path
from datetime import datetime, timedelta
from collections import Counter

from openpyxl import load_workbook


# ================================================================
# SETTINGS
# ================================================================

YEAR = 2026
DOY1 = 204
DOY2 = 207

LOG_DIR = Path(
    "products/refl_code/logs/usgs/ocean90_150/2026"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

SOURCE_CSV = Path(
    "gnssir_tide_arc_analysis.csv"
)

AZ_START = 100.0
AZ_END = 150.0
AZ_WIDTH = 5.0

MIN_N = 3


# ================================================================
# CORRELATION
# ================================================================

def correlation(x, y):

    if len(x) < 2:
        return float("nan")

    mx = sum(x) / len(x)
    my = sum(y) / len(y)

    dx = [v - mx for v in x]
    dy = [v - my for v in y]

    sx = math.sqrt(sum(v * v for v in dx))
    sy = math.sqrt(sum(v * v for v in dy))

    if sx == 0 or sy == 0:
        return float("nan")

    return (
        sum(a * b for a, b in zip(dx, dy))
        / (sx * sy)
    )


# ================================================================
# INTERPOLATED TIDE MODEL
# ================================================================

print()
print("=" * 80)
print("GNSS-IR AZIMUTH-BIN EXACT-TIME TIDE TEST")
print("=" * 80)

print()
print(f"DOY range       : {DOY1}-{DOY2}")
print(f"SUCCESS logs    : {LOG_DIR}")
print(f"Tide workbook   : {TIDE_FILE}")
print()
print(
    f"Azimuth bins    : "
    f"{AZ_START:.0f}-{AZ_END:.0f} degrees"
)
print(
    f"Bin width       : {AZ_WIDTH:.0f} degrees"
)


wb = load_workbook(
    TIDE_FILE,
    data_only=True
)

ws = wb[wb.sheetnames[0]]

rows = list(
    ws.iter_rows(values_only=True)
)

headers = list(rows[0])

model_columns = [
    i
    for i, name in enumerate(headers)
    if (
        isinstance(name, str)
        and name.endswith("_heightm")
    )
]

print()
print("Tide model columns:")

for i in model_columns:
    print(
        f"  {headers[i]}"
    )


tide_times = []
tide_values = []

for row in rows[1:]:

    if not row:
        continue

    if not isinstance(
        row[0],
        datetime
    ):
        continue

    values = []

    valid = True

    for i in model_columns:

        value = row[i]

        try:
            value = float(value)
        except:
            valid = False
            break

        if not math.isfinite(value):
            valid = False
            break

        values.append(value)

    if not valid:
        continue

    tide_times.append(row[0])

    tide_values.append(
        sum(values) / len(values)
    )


print()
print(
    f"Tide points loaded: "
    f"{len(tide_times)}"
)

print(
    f"Tide coverage: "
    f"{tide_times[0]} through "
    f"{tide_times[-1]}"
)


def tide_at(query_dt):

    if query_dt < tide_times[0]:
        return float("nan")

    if query_dt > tide_times[-1]:
        return float("nan")

    # Exact point
    for i in range(len(tide_times)):

        if query_dt == tide_times[i]:
            return tide_values[i]

    # Find enclosing interval
    for i in range(1, len(tide_times)):

        t1 = tide_times[i - 1]
        t2 = tide_times[i]

        if t1 <= query_dt <= t2:

            y1 = tide_values[i - 1]
            y2 = tide_values[i]

            total = (
                t2 - t1
            ).total_seconds()

            part = (
                query_dt - t1
            ).total_seconds()

            if total == 0:
                return y1

            return (
                y1
                + (y2 - y1)
                * part
                / total
            )

    return float("nan")


# ================================================================
# PARSE SUCCESS RECORDS
# ================================================================

pattern = re.compile(
    r"SUCCESS\s+Azimuth\s+([0-9.]+)\s+"
    r"(?:Sat|Satellite)\s+(\d+)\s+"
    r"RH\s+([-+0-9.]+)\s+m\s+"
    r"PkNoise\s+([-+0-9.]+)\s+"
    r"Amp\s+([-+0-9.]+)\s+"
    r"Fr\s*(\d+)\s+"
    r"UTC\s+([0-9:]+)\s+"
    r"DT\s+([0-9.]+)"
)


success = []

for doy in range(
    DOY1,
    DOY2 + 1
):

    path = (
        LOG_DIR /
        f"{doy}_gnssir.txt"
    )

    if not path.exists():
        continue

    with open(
        path,
        errors="ignore"
    ) as f:

        for line in f:

            m = pattern.search(line)

            if not m:
                continue

            az = float(m.group(1))
            sat = int(m.group(2))
            rh = float(m.group(3))
            freq = int(m.group(6))
            utc = m.group(7)

            parts = utc.split(":")

            if len(parts) == 2:

                hh = int(parts[0])
                mm = int(parts[1])
                ss = 0

            elif len(parts) == 3:

                hh = int(parts[0])
                mm = int(parts[1])
                ss = int(parts[2])

            else:
                continue

            day = (
                datetime(
                    YEAR,
                    1,
                    1
                )
                + timedelta(
                    days=doy - 1
                )
            )

            dt = datetime(
                day.year,
                day.month,
                day.day,
                hh,
                mm,
                ss
            )

            tide = tide_at(dt)

            if not math.isfinite(tide):
                continue

            success.append(
                {
                    "dt": dt,
                    "doy": doy,
                    "utc": utc,
                    "az": az,
                    "sat": sat,
                    "freq": freq,
                    "rh": rh,
                    "tide": tide,
                }
            )


print()
print(
    f"SUCCESS records in logs : "
    f"{len(success)}"
)


# ================================================================
# DISPLAY ALL RECORDS
# ================================================================

print()
print("=" * 80)
print("EXACT-TIME TIDE VALUES")
print("=" * 80)

for r in sorted(
    success,
    key=lambda x: x["dt"]
):

    print(
        f"{r['dt'].strftime('%m-%d %H:%M:%S')} "
        f"Az={r['az']:6.1f} "
        f"sat={r['sat']:3d} "
        f"freq={r['freq']:3d} "
        f"RH={r['rh']:7.3f} "
        f"TIDE={r['tide']:8.4f}"
    )


# ================================================================
# OVERALL RELATIONSHIP
# ================================================================

rh_all = [
    r["rh"]
    for r in success
]

tide_all = [
    r["tide"]
    for r in success
]

print()
print("=" * 80)
print("OVERALL ZERO-LAG RELATIONSHIP")
print("=" * 80)

print(
    f"n = {len(success)}"
)

print(
    f"RH vs tide         : "
    f"{correlation(rh_all, tide_all):+.4f}"
)

print(
    f"GNSS water vs tide : "
    f"{-correlation(rh_all, tide_all):+.4f}"
)


# ================================================================
# AZIMUTH BINS
# ================================================================

print()
print("=" * 80)
print("5-DEGREE AZIMUTH BINS — EXACT-TIME ZERO LAG")
print("=" * 80)

print()
print(
    f"{'BIN':>11} "
    f"{'N':>4} "
    f"{'RH-tide r':>11} "
    f"{'GNSS-tide r':>13} "
    f"{'Frequencies':>20}"
)

print("-" * 80)


bin_results = []

az = AZ_START

while az < AZ_END:

    lo = az
    hi = az + AZ_WIDTH

    subset = [
        r
        for r in success
        if (
            lo <= r["az"] < hi
        )
    ]

    n = len(subset)

    counts = Counter(
        r["freq"]
        for r in subset
    )

    freq_text = ",".join(
        f"{f}:{c}"
        for f, c in
        sorted(counts.items())
    )

    if n >= MIN_N:

        rh = [
            r["rh"]
            for r in subset
        ]

        tide = [
            r["tide"]
            for r in subset
        ]

        r_rh = correlation(
            rh,
            tide
        )

        r_gnss = -r_rh

        print(
            f"{lo:5.0f}-{hi:<5.0f} "
            f"{n:4d} "
            f"{r_rh:+11.4f} "
            f"{r_gnss:+13.4f} "
            f"{freq_text:>20}"
        )

        bin_results.append(
            {
                "lo": lo,
                "hi": hi,
                "n": n,
                "r_rh": r_rh,
                "r_gnss": r_gnss,
                "freq": freq_text,
            }
        )

    else:

        print(
            f"{lo:5.0f}-{hi:<5.0f} "
            f"{n:4d} "
            f"{'insufficient':>11} "
            f"{'':>13} "
            f"{freq_text:>20}"
        )

    az += AZ_WIDTH


# ================================================================
# IMPORTANT GEOMETRY COMPARISONS
# ================================================================

print()
print("=" * 80)
print("KEY GEOMETRY GROUPS")
print("=" * 80)


groups = [
    ("100-150", 100, 150),
    ("100-140", 100, 140),
    ("100-130", 100, 130),
    ("110-130", 110, 130),
    ("110-120", 110, 120),
    ("112-115", 112, 115),
]


for name, lo, hi in groups:

    subset = [
        r
        for r in success
        if lo <= r["az"] <= hi
    ]

    if len(subset) < 2:

        print(
            f"{name:10s} "
            f"n={len(subset):2d} "
            f"insufficient"
        )

        continue

    rh = [
        r["rh"]
        for r in subset
    ]

    tide = [
        r["tide"]
        for r in subset
    ]

    r = correlation(
        rh,
        tide
    )

    counts = Counter(
        r0["freq"]
        for r0 in subset
    )

    freq_text = ",".join(
        f"{f}:{n}"
        for f, n in
        sorted(counts.items())
    )

    print(
        f"{name:10s} "
        f"n={len(subset):2d} "
        f"r={r:+.4f} "
        f"freq={freq_text}"
    )


# ================================================================
# FREQUENCY STATISTICS
# ================================================================

print()
print("=" * 80)
print("FREQUENCY ZERO-LAG RESULTS")
print("=" * 80)

for freq in sorted(
    set(r["freq"] for r in success)
):

    subset = [
        r
        for r in success
        if r["freq"] == freq
    ]

    if len(subset) < 3:

        print(
            f"freq={freq:3d} "
            f"n={len(subset):2d} "
            f"insufficient"
        )

        continue

    rh = [
        r["rh"]
        for r in subset
    ]

    tide = [
        r["tide"]
        for r in subset
    ]

    r = correlation(
        rh,
        tide
    )

    print(
        f"freq={freq:3d} "
        f"n={len(subset):2d} "
        f"r={r:+.4f}"
    )


# ================================================================
# WRITE RESULTS
# ================================================================

output = Path(
    "azimuth_bins_exact_tide_results.csv"
)

with open(
    output,
    "w",
    newline=""
) as f:

    writer = csv.writer(f)

    writer.writerow(
        [
            "azimuth_low",
            "azimuth_high",
            "n",
            "rh_tide_r",
            "gnss_water_tide_r",
            "frequencies",
        ]
    )

    for r in bin_results:

        writer.writerow(
            [
                r["lo"],
                r["hi"],
                r["n"],
                r["r_rh"],
                r["r_gnss"],
                r["freq"],
            ]
        )


print()
print("=" * 80)
print("DONE")
print("=" * 80)

print()
print(
    f"Results written to: "
    f"{output}"
)
