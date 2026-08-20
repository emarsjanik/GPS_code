import csv
import math
import re
from pathlib import Path
from datetime import datetime, timedelta
from collections import Counter


# ================================================================
# SETTINGS
# ================================================================

DOY1 = 204
DOY2 = 207
YEAR = 2026

LOG_DIR = Path(
    "products/refl_code/logs/usgs/ocean90_150/2026"
)

TIDE_FILE = Path(
    "marconi_tides_sherwood.xlsx"
)

CSV_FILE = Path(
    "gnssir_tide_arc_analysis.csv"
)

MATCH_LIMIT_SEC = 90.0

REFERENCE_HEIGHT = 18.665

# 5-degree azimuth bins
AZ_START = 100.0
AZ_END = 150.0
AZ_WIDTH = 5.0

# Require at least this many observations for statistics
MIN_N = 3


# ================================================================
# HELPERS
# ================================================================

def correlation(x, y):

    if len(x) < 2:
        return float("nan")

    mx = sum(x) / len(x)
    my = sum(y) / len(y)

    dx = [v - mx for v in x]
    dy = [v - my for v in y]

    sx = math.sqrt(sum(v * v for v in dx))
    sy = math.sqrt(sum(v * v for v in dy))

    if sx == 0 or sy == 0:
        return float("nan")

    return sum(
        a * b for a, b in zip(dx, dy)
    ) / (sx * sy)


def rms(values):

    if not values:
        return float("nan")

    return math.sqrt(
        sum(v * v for v in values) / len(values)
    )


def median(values):

    if not values:
        return float("nan")

    x = sorted(values)
    n = len(x)

    if n % 2:
        return x[n // 2]

    return 0.5 * (
        x[n // 2 - 1] +
        x[n // 2]
    )


# ================================================================
# READ TIDE WORKBOOK
# ================================================================

from openpyxl import load_workbook

print()
print("=" * 80)
print("GNSS-IR AZIMUTH-BIN ZERO-LAG TEST")
print("=" * 80)

print()
print(f"DOY range       : {DOY1}-{DOY2}")
print(f"SUCCESS logs    : {LOG_DIR}")
print(f"Tide workbook   : {TIDE_FILE}")
print(f"Reference H     : {REFERENCE_HEIGHT:.3f} m")
print(f"Match tolerance : {MATCH_LIMIT_SEC:.0f} sec")
print()
print("Azimuth bins:")
print(
    f"{AZ_START:.0f}-{AZ_END:.0f} degrees "
    f"in {AZ_WIDTH:.0f}-degree bins"
)


wb = load_workbook(
    TIDE_FILE,
    data_only=True
)

ws = wb[wb.sheetnames[0]]

rows = list(
    ws.iter_rows(
        values_only=True
    )
)

headers = list(rows[0])

model_columns = []

for i, name in enumerate(headers):

    if (
        isinstance(name, str)
        and name.endswith("_heightm")
    ):
        model_columns.append(i)


print()
print("Tide model columns:")

for i in model_columns:
    print(
        " ",
        headers[i]
    )


tide_records = []

for row in rows[1:]:

    if not row:
        continue

    if not isinstance(
        row[0],
        datetime
    ):
        continue

    values = []

    for i in model_columns:

        value = row[i]

        if value is None:
            break

        try:
            value = float(value)
        except:
            break

        if not math.isfinite(value):
            break

        values.append(value)

    if len(values) != len(model_columns):
        continue

    ensemble = (
        sum(values) /
        len(values)
    )

    tide_records.append(
        {
            "dt": row[0],
            "tide": ensemble,
        }
    )


print(
    f"Tide points loaded: "
    f"{len(tide_records)}"
)

print(
    f"Tide coverage: "
    f"{tide_records[0]['dt']} "
    f"through "
    f"{tide_records[-1]['dt']}"
)


def tide_at(query_dt):

    if query_dt < tide_records[0]["dt"]:
        return float("nan")

    if query_dt > tide_records[-1]["dt"]:
        return float("nan")

    # Linear interpolation
    for i in range(
        1,
        len(tide_records)
    ):

        t1 = tide_records[i - 1]["dt"]
        t2 = tide_records[i]["dt"]

        if t1 <= query_dt <= t2:

            y1 = tide_records[i - 1]["tide"]
            y2 = tide_records[i]["tide"]

            total = (
                t2 - t1
            ).total_seconds()

            part = (
                query_dt - t1
            ).total_seconds()

            if total == 0:
                return y1

            return (
                y1
                + (y2 - y1)
                * part / total
            )

    return float("nan")


# ================================================================
# PARSE SUCCESS RECORDS
# ================================================================

pattern = re.compile(
    r"SUCCESS\s+Azimuth\s+([0-9.]+)\s+"
    r"(?:Sat|Satellite)\s+(\d+)\s+"
    r"RH\s+([-+0-9.]+)\s+m\s+"
    r"PkNoise\s+([-+0-9.]+)\s+"
    r"Amp\s+([-+0-9.]+)\s+"
    r"Fr\s*(\d+)\s+"
    r"UTC\s+([0-9:]+)\s+"
    r"DT\s+([0-9.]+)"
)

success = []

for doy in range(
    DOY1,
    DOY2 + 1
):

    path = (
        LOG_DIR /
        f"{doy}_gnssir.txt"
    )

    if not path.exists():
        continue

    with open(
        path,
        errors="ignore"
    ) as f:

        for line in f:

            m = pattern.search(line)

            if not m:
                continue

            az = float(
                m.group(1)
            )

            sat = int(
                m.group(2)
            )

            rh = float(
                m.group(3)
            )

            freq = int(
                m.group(6)
            )

            utc = m.group(7)

            parts = utc.split(":")

            if len(parts) == 2:

                hh = int(parts[0])
                mm = int(parts[1])
                ss = 0

            elif len(parts) == 3:

                hh = int(parts[0])
                mm = int(parts[1])
                ss = int(parts[2])

            else:
                continue

            day = (
                datetime(
                    YEAR,
                    1,
                    1
                )
                + timedelta(
                    days=doy - 1
                )
            )

            dt = datetime(
                day.year,
                day.month,
                day.day,
                hh,
                mm,
                ss
            )

            success.append(
                {
                    "dt": dt,
                    "doy": doy,
                    "az": az,
                    "sat": sat,
                    "freq": freq,
                    "rh": rh,
                }
            )


print()
print(
    f"SUCCESS records : "
    f"{len(success)}"
)


# ================================================================
# MATCH SUCCESS RECORDS TO TIDE DATASET
# ================================================================

matched = []
unmatched = []

for s in success:

    candidates = []

    for t in tide_records:

        if t["dt"].date() != s["dt"].date():
            continue

        delta = abs(
            (
                t["dt"] -
                s["dt"]
            ).total_seconds()
        )

        if delta <= MATCH_LIMIT_SEC:
            candidates.append(
                (delta, t)
            )

    if not candidates:

        unmatched.append(s)
        continue

    delta, best = min(
        candidates,
        key=lambda x: x[0]
    )

    tide = best["tide"]

    # Physical GNSS water-level proxy
    #
    # Higher RH = lower reflecting surface
    #
    # GNSS water level = H - RH
    #
    gnss_water = (
        REFERENCE_HEIGHT -
        s["rh"]
    )

    record = dict(s)

    record["tide"] = tide
    record["gnss_water"] = gnss_water
    record["match_delta"] = delta

    matched.append(record)


print(
    f"Matched         : "
    f"{len(matched)}"
)

print(
    f"Unmatched       : "
    f"{len(unmatched)}"
)


# ================================================================
# PRINT MATCHED DATA
# ================================================================

print()
print("=" * 80)
print("MATCHED RECORDS")
print("=" * 80)

for r in sorted(
    matched,
    key=lambda x: x["dt"]
):

    print(
        f"{r['dt'].strftime('%m-%d %H:%M:%S')} "
        f"Az={r['az']:6.1f} "
        f"sat={r['sat']:3d} "
        f"freq={r['freq']:3d} "
        f"RH={r['rh']:7.3f} "
        f"GNSS={r['gnss_water']:8.3f} "
        f"TIDE={r['tide']:7.3f}"
    )


# ================================================================
# OVERALL ZERO-LAG STATISTICS
# ================================================================

print()
print("=" * 80)
print("OVERALL ZERO-LAG RELATIONSHIP")
print("=" * 80)

x_rh = [
    r["rh"]
    for r in matched
]

x_gnss = [
    r["gnss_water"]
    for r in matched
]

y_tide = [
    r["tide"]
    for r in matched
]

print(
    f"n = {len(matched)}"
)

print(
    f"RH vs tide          : "
    f"{correlation(x_rh, y_tide):+.4f}"
)

print(
    f"GNSS water vs tide  : "
    f"{correlation(x_gnss, y_tide):+.4f}"
)


# ================================================================
# AZIMUTH BINS
# ================================================================

print()
print("=" * 80)
print("5-DEGREE AZIMUTH BINS — ZERO LAG")
print("=" * 80)

print()
print(
    f"{'BIN':>11} "
    f"{'N':>4} "
    f"{'RH-tide r':>11} "
    f"{'GNSS-tide r':>13} "
    f"{'RMS cm':>10} "
    f"{'RH mean':>10} "
    f"{'Tide mean':>10}"
)

print("-" * 80)


bin_results = []

az = AZ_START

while az < AZ_END:

    lo = az
    hi = az + AZ_WIDTH

    # Last bin includes upper endpoint
    if hi >= AZ_END:
        subset = [
            r for r in matched
            if lo <= r["az"] <= hi
        ]
    else:
        subset = [
            r for r in matched
            if lo <= r["az"] < hi
        ]

    n = len(subset)

    if n >= MIN_N:

        rh = [
            r["rh"]
            for r in subset
        ]

        gnss = [
            r["gnss_water"]
            for r in subset
        ]

        tide = [
            r["tide"]
            for r in subset
        ]

        r_rh = correlation(
            rh,
            tide
        )

        r_gnss = correlation(
            gnss,
            tide
        )

        # Remove constant bias before calculating RMS.
        #
        # This measures the scatter around the
        # best constant offset between GNSS and tide.
        errors = [
            (
                g - t
            )
            for g, t in zip(
                gnss,
                tide
            )
        ]

        bias = (
            sum(errors) /
            len(errors)
        )

        residuals = [
            e - bias
            for e in errors
        ]

        rms_cm = (
            rms(residuals)
            * 100.0
        )

        rh_mean = (
            sum(rh) /
            len(rh)
        )

        tide_mean = (
            sum(tide) /
            len(tide)
        )

        print(
            f"{lo:5.0f}-{hi:<5.0f} "
            f"{n:4d} "
            f"{r_rh:+11.4f} "
            f"{r_gnss:+13.4f} "
            f"{rms_cm:10.2f} "
            f"{rh_mean:10.3f} "
            f"{tide_mean:10.3f}"
        )

        bin_results.append(
            {
                "lo": lo,
                "hi": hi,
                "n": n,
                "r_rh": r_rh,
                "r_gnss": r_gnss,
                "rms_cm": rms_cm,
            }
        )

    else:

        print(
            f"{lo:5.0f}-{hi:<5.0f} "
            f"{n:4d} "
            f"{'insufficient':>11}"
        )

    az += AZ_WIDTH


# ================================================================
# BEST GEOMETRY
# ================================================================

if bin_results:

    best_rh = max(
        bin_results,
        key=lambda x: (
            -999
            if math.isnan(x["r_rh"])
            else x["r_rh"]
        )
    )

    best_abs = max(
        bin_results,
        key=lambda x: (
            -999
            if math.isnan(x["r_rh"])
            else abs(x["r_rh"])
        )
    )

    best_rms = min(
        bin_results,
        key=lambda x: x["rms_cm"]
    )

    print()
    print("=" * 80)
    print("BEST AZIMUTH BINS")
    print("=" * 80)

    print()
    print(
        "Highest RH-vs-tide correlation:"
    )

    print(
        f"  {best_rh['lo']:.0f}-"
        f"{best_rh['hi']:.0f} degrees"
    )

    print(
        f"  n = {best_rh['n']}"
    )

    print(
        f"  r = {best_rh['r_rh']:+.4f}"
    )

    print()
    print(
        "Strongest absolute RH-vs-tide correlation:"
    )

    print(
        f"  {best_abs['lo']:.0f}-"
        f"{best_abs['hi']:.0f} degrees"
    )

    print(
        f"  n = {best_abs['n']}"
    )

    print(
        f"  r = {best_abs['r_rh']:+.4f}"
    )

    print()
    print(
        "Lowest bias-removed RMS:"
    )

    print(
        f"  {best_rms['lo']:.0f}-"
        f"{best_rms['hi']:.0f} degrees"
    )

    print(
        f"  n = {best_rms['n']}"
    )

    print(
        f"  RMS = {best_rms['rms_cm']:.2f} cm"
    )


# ================================================================
# FREQUENCY-BY-GEOMETRY CHECK
# ================================================================

print()
print("=" * 80)
print("FREQUENCY DISTRIBUTION BY AZIMUTH BIN")
print("=" * 80)

for az in [
    AZ_START + i * AZ_WIDTH
    for i in range(
        int(
            (AZ_END - AZ_START)
            / AZ_WIDTH
        )
    )
]:

    lo = az
    hi = az + AZ_WIDTH

    subset = [
        r for r in matched
        if lo <= r["az"] < hi
    ]

    if not subset:
        continue

    counts = Counter(
        r["freq"]
        for r in subset
    )

    text = ", ".join(
        f"{freq}:{n}"
        for freq, n
        in sorted(counts.items())
    )

    print(
        f"{lo:.0f}-{hi:.0f}: "
        f"{text}"
    )


# ================================================================
# WRITE CSV
# ================================================================

output = Path(
    "azimuth_bins_zero_lag_results.csv"
)

with open(
    output,
    "w",
    newline=""
) as f:

    writer = csv.writer(f)

    writer.writerow(
        [
            "azimuth_low",
            "azimuth_high",
            "n",
            "rh_tide_r",
            "gnss_water_tide_r",
            "rms_cm",
        ]
    )

    for r in bin_results:

        writer.writerow(
            [
                r["lo"],
                r["hi"],
                r["n"],
                r["r_rh"],
                r["r_gnss"],
                r["rms_cm"],
            ]
        )


print()
print("=" * 80)
print("DONE")
print("=" * 80)

print()
print(
    f"Results written to: {output}"
)
