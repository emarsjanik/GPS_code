import csv
import math
from pathlib import Path
from datetime import datetime, timedelta

import numpy as np
from openpyxl import load_workbook


# ================================================================
# SETTINGS
# ================================================================

CSV_FILE = Path("gnssir_tide_arc_analysis.csv")
TIDE_FILE = Path("marconi_tides_sherwood.xlsx")

SAT = 29
FREQ = 1

REFERENCE_HEIGHT = 18.665


# ================================================================
# LOAD TIDE MODELS
# ================================================================

wb = load_workbook(
    TIDE_FILE,
    data_only=True
)

ws = wb[wb.sheetnames[0]]

headers = [
    cell.value
    for cell in ws[1]
]

TIME_COL = headers.index("time")

MODEL_NAMES = [
    "EOT20_heightm",
    "GOT5.5_heightm",
    "GOT5.6_heightm",
    "FES2022_heightm",
]

MODEL_COLS = [
    headers.index(name)
    for name in MODEL_NAMES
]


tide_times = []
tide_values = []


for row in ws.iter_rows(
    min_row=2,
    values_only=True
):

    t = row[TIME_COL]

    if not isinstance(t, datetime):
        continue

    vals = []

    valid = True

    for col in MODEL_COLS:

        v = row[col]

        if v is None:

            valid = False
            break

        try:
            v = float(v)
        except ValueError:
            valid = False
            break

        if not math.isfinite(v):

            valid = False
            break

        vals.append(v)

    if not valid:
        continue

    tide_times.append(t)
    tide_values.append(
        sum(vals) / len(vals)
    )


print()
print("=" * 80)
print("PRN 29 / FREQUENCY 1 ARC-TIDE TEST")
print("=" * 80)

print()
print("Tide points:", len(tide_times))

print(
    "Tide coverage:",
    tide_times[0],
    "through",
    tide_times[-1]
)


# ================================================================
# TIDE INTERPOLATOR
# ================================================================

epoch = np.array(
    [
        (t - tide_times[0]).total_seconds()
        for t in tide_times
    ],
    dtype=float
)

values = np.array(
    tide_values,
    dtype=float
)


def tide_at(dt):

    x = (
        dt - tide_times[0]
    ).total_seconds()

    if x < epoch[0] or x > epoch[-1]:

        return float("nan")

    return float(
        np.interp(
            x,
            epoch,
            values
        )
    )


# ================================================================
# LOAD PRN 29 / FREQ 1 RECORDS
# ================================================================

records = []

with open(
    CSV_FILE,
    newline=""
) as f:

    reader = csv.DictReader(f)

    for r in reader:

        try:

            sat = int(float(r["sat"]))
            freq = int(float(r["freq"]))

            if sat != SAT:
                continue

            if freq != FREQ:
                continue

            rh = float(r["RH_m"])

            start = datetime.fromisoformat(
                r["time_start_utc"]
            )

            solution = datetime.fromisoformat(
                r["solution_time_utc"]
            )

            end = datetime.fromisoformat(
                r["time_end_utc"]
            )

            az = float(r["Azim"])

        except (
            ValueError,
            KeyError
        ):

            continue

        records.append(
            {
                "start": start,
                "solution": solution,
                "end": end,
                "rh": rh,
                "az": az,
                "doy": int(
                    float(r["doy"])
                )
            }
        )


records.sort(
    key=lambda r: r["solution"]
)


print()
print(
    "PRN 29 / FREQ 1 records:",
    len(records)
)


# ================================================================
# ANALYZE EACH ARC
# ================================================================

results = []


for r in records:

    start = r["start"]
    solution = r["solution"]
    end = r["end"]

    tide_start = tide_at(start)
    tide_solution = tide_at(solution)
    tide_end = tide_at(end)

    if not all(
        math.isfinite(x)
        for x in [
            tide_start,
            tide_solution,
            tide_end
        ]
    ):
        continue

    # ------------------------------------------------------------
    # Sample tide throughout the actual GNSS-IR arc
    # ------------------------------------------------------------

    duration = (
        end - start
    ).total_seconds()

    sample_seconds = np.linspace(
        0,
        duration,
        100
    )

    tide_series = np.array(
        [
            tide_at(
                start +
                timedelta(
                    seconds=float(s)
                )
            )
            for s in sample_seconds
        ]
    )

    # ------------------------------------------------------------
    # Tide statistics
    # ------------------------------------------------------------

    tide_min = np.min(
        tide_series
    )

    tide_max = np.max(
        tide_series
    )

    tide_mean = np.mean(
        tide_series
    )

    tide_range = (
        tide_max -
        tide_min
    )

    tide_change = (
        tide_end -
        tide_start
    )

    # ------------------------------------------------------------
    # Where is GNSS solution inside arc?
    # ------------------------------------------------------------

    fraction = (
        solution - start
    ).total_seconds() / duration

    # ------------------------------------------------------------
    # Print
    # ------------------------------------------------------------

    print()
    print(
        f"DOY {r['doy']} "
        f"{start.strftime('%Y-%m-%d %H:%M:%S')} "
        f"-> "
        f"{end.strftime('%H:%M:%S')}"
    )

    print(
        f"  Azimuth        : {r['az']:.2f}"
    )

    print(
        f"  RH             : {r['rh']:.3f} m"
    )

    print(
        f"  Tide start     : {tide_start:+.4f} m"
    )

    print(
        f"  Tide solution  : {tide_solution:+.4f} m"
    )

    print(
        f"  Tide end       : {tide_end:+.4f} m"
    )

    print(
        f"  Tide change    : {tide_change:+.4f} m"
    )

    print(
        f"  Tide range     : {tide_range:.4f} m"
    )

    print(
        f"  Tide mean      : {tide_mean:+.4f} m"
    )

    print(
        f"  Solution frac. : {fraction:.3f}"
    )

    results.append(
        {
            "rh": r["rh"],
            "tide_solution": tide_solution,
            "tide_mean": tide_mean,
            "tide_change": tide_change,
            "tide_range": tide_range
        }
    )


# ================================================================
# CORRELATIONS
# ================================================================

def correlation(a, b):

    a = np.asarray(
        a,
        dtype=float
    )

    b = np.asarray(
        b,
        dtype=float
    )

    if len(a) < 3:
        return float("nan")

    if (
        np.std(a) == 0
        or np.std(b) == 0
    ):
        return float("nan")

    return float(
        np.corrcoef(
            a,
            b
        )[0, 1]
    )


rh = np.array(
    [
        r["rh"]
        for r in results
    ]
)

tide_solution = np.array(
    [
        r["tide_solution"]
        for r in results
    ]
)

tide_mean = np.array(
    [
        r["tide_mean"]
        for r in results
    ]
)

tide_change = np.array(
    [
        r["tide_change"]
        for r in results
    ]
)

tide_range = np.array(
    [
        r["tide_range"]
        for r in results
    ]
)


print()
print("=" * 80)
print("CORRELATIONS")
print("=" * 80)

print()
print(
    "RH vs tide at solution :",
    f"{correlation(rh, tide_solution):+.4f}"
)

print(
    "RH vs mean tide        :",
    f"{correlation(rh, tide_mean):+.4f}"
)

print(
    "RH vs tide change      :",
    f"{correlation(rh, tide_change):+.4f}"
)

print(
    "RH vs tide range       :",
    f"{correlation(rh, tide_range):+.4f}"
)


# ================================================================
# DIFFERENCE TEST
# ================================================================

if len(results) >= 2:

    drh = np.diff(rh)
    dtide = np.diff(
        tide_solution
    )

    print()
    print("=" * 80)
    print("SUCCESSIVE-OBSERVATION DIFFERENCES")
    print("=" * 80)

    print(
        "Delta RH vs delta tide:",
        f"{correlation(drh, dtide):+.4f}"
    )

    print()

    for i in range(
        len(drh)
    ):

        print(
            f"{i+1:2d} -> {i+2:2d} "
            f"dRH={drh[i]:+.4f} "
            f"dTide={dtide[i]:+.4f}"
        )


# ================================================================
# SAVE RESULTS
# ================================================================

out = Path(
    "prn29_arc_tide_test.csv"
)

with open(
    out,
    "w",
    newline=""
) as f:

    writer = csv.writer(f)

    writer.writerow(
        [
            "doy",
            "azimuth",
            "RH_m",
            "tide_start_m",
            "tide_solution_m",
            "tide_end_m",
            "tide_change_m",
            "tide_mean_m",
            "tide_range_m",
            "solution_fraction"
        ]
    )

    for r in records:

        start = r["start"]
        solution = r["solution"]
        end = r["end"]

        ts = tide_at(start)
        tm = tide_at(solution)
        te = tide_at(end)

        duration = (
            end - start
        ).total_seconds()

        fraction = (
            solution - start
        ).total_seconds() / duration

        writer.writerow(
            [
                r["doy"],
                r["az"],
                r["rh"],
                ts,
                tm,
                te,
                te - ts,
                tide_at(
                    start +
                    (end - start) / 2
                ),
                max(
                    ts,
                    tm,
                    te
                ) -
                min(
                    ts,
                    tm,
                    te
                ),
                fraction
            ]
        )


print()
print(
    "Results written to:",
    out
)

print()
print("=" * 80)
print("DONE")
print("=" * 80)
