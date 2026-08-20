#!/usr/bin/env python3
"""
diagnose_arc_timing_bias.py

Diagnostic script for the phase-lag investigation. Does NOT apply any
correction -- this is a first, honestly-scoped step to quantify how
exposed our real arcs actually are to a known, documented error
source, using real data from this site, before committing to building
a full correction (a substantially bigger undertaking).

Background (see REFERENCES.md for the full sources): gnssrefl's
standard processing, like every traditional LSP-based GNSS-IR tool,
treats the water level as CONSTANT across each satellite arc when
estimating its reflector height. A real arc takes real time (tens of
minutes here, per our own delTmax=40 setting) -- and during that
time, the tide is actually still moving. Chang et al. (2026) measured
a real, quantified 10-20 minute time bias from exactly this
assumption; Zhang et al. (2024) directly compared three different
correction methods for it, with real, quantified improvements (up to
~40% RMSE reduction depending on site and method).

What this script actually does, honestly:
    1. Extract every real satellite arc for the given day(s), matched
       to its gnssir processing result (using gnssrefl's own
       extract_arcs API -- see REFERENCES.md).
    2. For each arc that passed QC, look up the real tide model's
       predicted height at the arc's start time AND its end time
       (not just its midpoint) -- this tells us how much the tide
       genuinely moved during that specific arc's timespan, using
       data we already trust (the same tide model file used
       throughout tonight's comparisons).
    3. Report the real distribution of this "within-arc tidal
       change" across our actual arcs, and whether it correlates
       with arc duration.

What this script deliberately does NOT yet do: compare this against
our own RH deviation from the tide model. That requires working out
the vertical datum alignment between raw reflector height and the
tide model properly (compare_to_tide_models.py sidesteps this with a
dual-axis, shape-only comparison) -- a real next step, but a separate
one, best tackled once this first step confirms the error source is
large enough here to be worth correcting for at all.

Usage:
    python3 diagnose_arc_timing_bias.py <station> <year> <doy1> [doy2] <tide_models_xlsx>

Example:
    python3 diagnose_arc_timing_bias.py usgs 2026 204 207 marconi_tides_sherwood.xlsx

Requires the gnssrefl virtual environment to be active (for the
extract_arcs import) and the usual REFL_CODE/EXE/ORBITS environment
variables to be set, same as any other direct gnssrefl command.
"""

from __future__ import annotations

import sys
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import openpyxl


def read_tide_models(xlsx_path: Path):
    """
    Reads the real tide model Excel file directly. Returns
    (times, {model_name: [heights]}) for every column ending in
    '_heightm'. Matches compare_to_tide_models.py's own function
    exactly, so both scripts read this file identically.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    model_columns = {}
    for i, name in enumerate(header):
        if name and name.endswith("_heightm"):
            model_columns[name[: -len("_heightm")]] = i

    times = []
    model_values = {name: [] for name in model_columns}

    for row in rows[1:]:
        times.append(row[0])
        for model_name, col_index in model_columns.items():
            model_values[model_name].append(row[col_index])

    return times, model_values


def build_ensemble_interpolator(times, model_values):
    """
    Builds a simple linear interpolator over the ensemble mean of all
    tide models, so we can look up a predicted tide height at any
    arbitrary timestamp (an arc's exact start/end time), not just the
    discrete points in the spreadsheet.
    """
    ensemble = np.mean(
        np.array([model_values[name] for name in model_values]), axis=0
    )

    epoch_seconds = np.array(
        [(t - times[0]).total_seconds() for t in times], dtype=float
    )

    def interpolate(query_dt: datetime) -> float:
        query_seconds = (query_dt - times[0]).total_seconds()
        return float(np.interp(query_seconds, epoch_seconds, ensemble))

    return interpolate


def seconds_of_day_to_datetime(day_start: datetime, seconds: float) -> datetime:
    """Converts arc metadata's seconds-of-day (can be negative or >86400
    for midnight-crossing buffer data) into a real datetime."""
    return day_start + timedelta(seconds=seconds)


def main():
    if len(sys.argv) not in (5, 6):
        print(f"Usage: {sys.argv[0]} <station> <year> <doy1> [doy2] <tide_models_xlsx>")
        sys.exit(1)

    station = sys.argv[1]
    year = int(sys.argv[2])
    doy1 = int(sys.argv[3])

    if len(sys.argv) == 6:
        doy2 = int(sys.argv[4])
        xlsx_path = Path(sys.argv[5])
    else:
        doy2 = doy1
        xlsx_path = Path(sys.argv[4])

    if not xlsx_path.exists():
        print(f"ERROR: {xlsx_path} does not exist.")
        sys.exit(1)

    try:
        from gnssrefl.extract_arcs import extract_arcs_from_station
    except ImportError:
        print("ERROR: could not import gnssrefl.extract_arcs -- make sure the")
        print("gnssrefl virtual environment is activated.")
        sys.exit(1)

    print(f"Reading tide models from {xlsx_path}...")
    times, model_values = read_tide_models(xlsx_path)
    print(f"Parsed {len(times)} tide model points, models: {list(model_values.keys())}")
    tide_at = build_ensemble_interpolator(times, model_values)

    arc_details = []

    for doy in range(doy1, doy2 + 1):
        print(f"\nExtracting arcs for {station} {year} doy {doy}...")
        # Explicitly matching our real, production gnssir settings
        # (station.json / usgs.json) rather than extract_arcs's own
        # different defaults (e1=5.0/e2=25.0 vs our real 5.0/15.0) --
        # confirmed necessary so the arc metadata computed here (delT,
        # elevation range, etc.) genuinely matches what our actual
        # processing pipeline used, not a wider/different window.
        # Note: ediff is NOT a real parameter of this function (confirmed
        # against the actual installed signature, despite being listed
        # in the current online docs) -- it's applied later, inside
        # gnssir's own QC, which is already reflected correctly in
        # whether an arc has a matched gnssir_processing_results.
        arcs = extract_arcs_from_station(
            station, year, doy,
            e1=5.0, e2=15.0,
            azlist=[100.0, 130.0, 150.0, 215.0],
            polyV=2, pele=[5, 30],
            attach_results=True, buffer_hours=2,
        )

        day_start = datetime(year, 1, 1) + timedelta(days=doy - 1)

        matched = [(m, d) for m, d in arcs if m["gnssir_processing_results"] is not None]
        print(f"  {len(arcs)} arcs extracted, {len(matched)} passed QC")

        for meta, _data in matched:
            result = meta["gnssir_processing_results"]

            arc_start_dt = seconds_of_day_to_datetime(day_start, meta["time_start"])
            arc_end_dt = seconds_of_day_to_datetime(day_start, meta["time_end"])

            tide_start = tide_at(arc_start_dt)
            tide_end = tide_at(arc_end_dt)
            within_arc_change_cm = (tide_end - tide_start) * 100

            arc_details.append(
                {
                    "doy": doy,
                    "sat": meta["sat"],
                    "freq": meta["freq"],
                    "delT": meta["delT"],
                    "within_arc_change_cm": within_arc_change_cm,
                    "RH": result["RH"],
                }
            )

    if not arc_details:
        print("\nNo QC-passing arcs found in this range -- nothing to analyze.")
        sys.exit(1)

    print(f"\n{len(arc_details)} total QC-passing arcs analyzed across days {doy1}-{doy2}.")

    delTs = np.array([a["delT"] for a in arc_details])
    changes_abs = np.array([abs(a["within_arc_change_cm"]) for a in arc_details])

    print("\n--- Within-arc tidal change summary (real, computed from our own data) ---")
    print(f"  Arc duration (delT):        min={delTs.min():.1f}  "
          f"mean={delTs.mean():.1f}  max={delTs.max():.1f}  minutes")
    print(f"  Within-arc tide change:     min={changes_abs.min():.2f}  "
          f"mean={changes_abs.mean():.2f}  max={changes_abs.max():.2f}  cm")

    pct_over_2cm = 100 * np.mean(changes_abs > 2.0)
    pct_over_5cm = 100 * np.mean(changes_abs > 5.0)
    print(f"\n  {pct_over_2cm:.0f}% of arcs had more than 2 cm of real tidal")
    print(f"  change during their own timespan; {pct_over_5cm:.0f}% had more than 5 cm.")
    print("  (For reference, our best-case RMS against the tide models has been")
    print("  ~18-21 cm -- so if a meaningful share of arcs see multi-cm real tidal")
    print("  change within their own duration, this is a plausible, real")
    print("  contributor worth pursuing further, not just a borrowed assumption.)")

    correlation = np.corrcoef(delTs, changes_abs)[0, 1]
    print(f"\n  Correlation between arc duration and within-arc tidal change: "
          f"{correlation:.3f}")
    print("  (A positive correlation here would confirm our longer arcs are")
    print("  genuinely more exposed to this error source -- real evidence from")
    print("  our own data, not just an assumption borrowed from the literature.)")

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))

    ax1.scatter(delTs, changes_abs, alpha=0.5, s=15)
    ax1.set_xlabel("Arc duration, delT (minutes)")
    ax1.set_ylabel("Within-arc tidal change (cm, absolute)")
    ax1.set_title(f"Arc duration vs. tidal change during the arc\n(r = {correlation:.3f})")
    ax1.grid(alpha=0.3)

    ax2.hist(changes_abs, bins=30, alpha=0.7)
    ax2.axvline(2.0, color="orange", linestyle="--", label="2 cm")
    ax2.axvline(5.0, color="red", linestyle="--", label="5 cm")
    ax2.set_xlabel("Within-arc tidal change (cm, absolute)")
    ax2.set_ylabel("Number of arcs")
    ax2.set_title("Distribution of within-arc tidal change\nacross all analyzed arcs")
    ax2.legend()
    ax2.grid(alpha=0.3)

    fig.suptitle(f"{station} {year} doy {doy1}-{doy2}: arc-duration bias diagnostic "
                 f"({len(arc_details)} arcs)")
    fig.tight_layout()

    output_path = Path("arc_timing_bias_diagnostic.png")
    fig.savefig(output_path, dpi=150)
    print(f"\nSaved diagnostic plot to: {output_path}")


if __name__ == "__main__":
    main()
