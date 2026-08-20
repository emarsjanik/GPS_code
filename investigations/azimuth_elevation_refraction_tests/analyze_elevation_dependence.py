#!/usr/bin/env python3
"""
analyze_elevation_dependence.py

Tests whether the shared, azimuth-independent tide correlation found
throughout tonight's investigation depends on elevation angle instead.

Rationale: a fixed structural reflector (e.g. a specific piece of
terrain, a building, vegetation) would produce an effect tied to
AZIMUTH -- which direction the satellite is in -- roughly independent
of how high in the sky it is. An atmospheric or multipath-geometry
effect would more plausibly scale with ELEVATION angle, since
atmospheric path length and multipath delay both depend strongly on
how low the satellite sits above the horizon.

For each arc, gnssrefl's result file already reports eminO/emaxO (the
observed elevation range of that specific arc, columns 8/9). This
groups arcs by their mean observed elevation and checks whether
per-elevation-bin correlation strength or bias varies systematically
-- using data we already have, no new gnssir runs required.

Usage:
    python3 analyze_elevation_dependence.py
"""

import math
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timedelta

import numpy as np
from openpyxl import load_workbook

RESULT_DIR = Path("products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13")
TIDE_FILE = Path("marconi_tides_sherwood.xlsx")

DOYS = [196, 197, 198, 201, 203, 204, 205, 206, 207, 208, 209, 210,
        211, 212, 213, 214, 215, 216, 217, 218, 219, 220, 221]

MODEL = "EOT20_heightm"
H_ORTHO_M = 18.665


def finite(x):
    try:
        v = float(x)
        return v if math.isfinite(v) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if len(x) < 5:
        return float("nan")
    if np.std(x) == 0 or np.std(y) == 0:
        return float("nan")
    return float(np.corrcoef(x, y)[0, 1])


# Load tide model
wb = load_workbook(TIDE_FILE, data_only=True)
ws = wb[wb.sheetnames[0]]
header = [c.value for c in ws[1]]
time_col = header.index("time")
tide_col = header.index(MODEL)

times, values = [], []
for row in ws.iter_rows(min_row=2, values_only=True):
    t = row[time_col]
    if not isinstance(t, datetime):
        continue
    v = finite(row[tide_col])
    if v is None:
        continue
    times.append(t)
    values.append(v)
wb.close()

epoch = np.array([(t - times[0]).total_seconds() for t in times])
values = np.asarray(values, float)


def tide_at(dt):
    x = (dt - times[0]).total_seconds()
    if x < epoch[0] or x > epoch[-1]:
        return float("nan")
    return float(np.interp(x, epoch, values))


# Load all arcs (individual observations, not grouped by track this time --
# we want every single arc's own elevation range)
arcs = []
for doy in DOYS:
    path = RESULT_DIR / f"{doy}.txt"
    if not path.exists():
        continue
    with open(path, errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            cols = line.split()
            if len(cols) < 17:
                continue
            try:
                year = int(float(cols[0]))
                doy2 = int(float(cols[1]))
                rh = float(cols[2])
                sat = int(float(cols[3]))
                utc_hours = float(cols[4])
                az = float(cols[5])
                emin = float(cols[7])
                emax = float(cols[8])
                freq = int(float(cols[10]))
            except Exception:
                continue
            if freq != 1:
                continue
            day = datetime(year, 1, 1) + timedelta(days=doy2 - 1)
            dt = day + timedelta(hours=utc_hours)
            water_level = H_ORTHO_M - rh
            tide = tide_at(dt)
            if not math.isfinite(tide):
                continue
            arcs.append({
                "sat": sat, "az": az, "elev_mean": (emin + emax) / 2.0,
                "water_level": water_level, "tide": tide,
                "residual": water_level - tide,
            })

print(f"Total arcs loaded: {len(arcs)}")

elevs = np.array([a["elev_mean"] for a in arcs])
print(f"Elevation range observed: {elevs.min():.2f} to {elevs.max():.2f} deg")

# Bin by elevation (1-degree bins across the 5-13 deg tested range)
bin_edges = np.arange(5, 14, 1)
bin_stats = []

for i in range(len(bin_edges) - 1):
    lo, hi = bin_edges[i], bin_edges[i + 1]
    subset = [a for a in arcs if lo <= a["elev_mean"] < hi]
    if len(subset) < 10:
        continue
    residuals = np.array([a["residual"] for a in subset])
    wl = np.array([a["water_level"] for a in subset])
    tide = np.array([a["tide"] for a in subset])
    r = pearson(wl, tide)
    bin_stats.append({
        "elev_lo": lo, "elev_hi": hi, "n": len(subset),
        "mean_residual_m": float(np.mean(residuals)),
        "std_residual_m": float(np.std(residuals)),
        "r": r,
    })

print()
print("=" * 80)
print("ELEVATION-BINNED RESULTS")
print("=" * 80)
print(f"{'elev bin':>12} {'n':>5} {'mean_resid_m':>14} {'std_resid_m':>13} {'r':>8}")
for b in bin_stats:
    print(f"{b['elev_lo']:5.0f}-{b['elev_hi']:<5.0f} {b['n']:>5} "
          f"{b['mean_residual_m']:14.4f} {b['std_residual_m']:13.4f} {b['r']:8.4f}")

if len(bin_stats) >= 3:
    mid_elevs = np.array([(b["elev_lo"] + b["elev_hi"]) / 2 for b in bin_stats])
    mean_resids = np.array([b["mean_residual_m"] for b in bin_stats])
    corrs = np.array([abs(b["r"]) for b in bin_stats if math.isfinite(b["r"])])
    mid_elevs_for_corr = np.array([(b["elev_lo"] + b["elev_hi"]) / 2
                                     for b in bin_stats if math.isfinite(b["r"])])

    print()
    print("=" * 80)
    print("KEY DIAGNOSTIC: does the mean bias depend on elevation?")
    print("=" * 80)
    print("corr(elevation, mean_residual):", np.corrcoef(mid_elevs, mean_resids)[0, 1])
    if len(corrs) >= 3:
        print("corr(elevation, |tide correlation r|):", np.corrcoef(mid_elevs_for_corr, corrs)[0, 1])
    print()
    print("A strong relationship here would point toward an atmospheric/")
    print("multipath-geometry effect (which scales with elevation). Near-zero")
    print("would rule this out too, same as azimuth was ruled out.")
