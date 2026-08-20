#!/usr/bin/env python3
"""
MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V4

Uses ONLY the established production GNSS-IR result directory:
  products/refl_code/2026/results/usgs/ocean17_23_l1_e5_13/*.txt

It excludes failQC and all other test/configuration directories.

It:
  1. searches for local USGS RINEX files;
  2. generates missing SNR files;
  3. generates missing GNSS-IR daily results when possible;
  4. loads only the established production results;
  5. matches GNSS-R observations to the EOT20 tide model;
  6. clusters satellite/rise/azimuth tracks;
  7. applies the established GOOD-track screen;
  8. plots GOOD tracks only;
  9. reports population mean bias, mean absolute deviation,
     median absolute deviation, and RMS for raw and +0.242 m data.

The +0.242 m correction is diagnostic only and is NEVER used to
select GOOD tracks.
"""

from __future__ import annotations

import csv
import math
import os
import re
import subprocess
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required.")
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

BASE_DIR = Path.home() / "GNSS" / "v4.1"
REFL_CODE = Path(os.environ.get(
    "REFL_CODE", str(BASE_DIR / "products" / "refl_code")
))

YEAR = 2026
STATION = "usgs00usa"

# IMPORTANT: this is the ONLY production result set used.
GNSSIR_CONFIG = "ocean17_23_l1_e5_13"

PRIMARY_TIDE_MODEL = "EOT20_heightm"
DATUM_OFFSET_M = 0.242

SNR_TYPE = 66
ORBIT = "gnss"
STREAM = "R"
SAMPLE_RATE = 1

# Established GOOD-track criteria.
MIN_OBS = 14
MIN_DAYS = 14
MIN_TIDE_R = 0.90
SLOPE_MIN = 0.85
SLOPE_MAX = 1.15
MAX_UNIT_RMS_M = 0.30
MAX_AZ_SD_DEG = 1.0
AZ_CLUSTER_TOL_DEG = 3.0

RINEX_ROOTS = [
    BASE_DIR,
    Path.home(),
    Path("/mnt/I2Rgus_Data"),
]

PLOT_DIR = BASE_DIR / "marconi_longterm_plots_v4"
OUT_CSV = BASE_DIR / "marconi_longterm_track_stability_v4.csv"
OUT_SUMMARY = BASE_DIR / "marconi_longterm_track_stability_v4_summary.txt"


# ============================================================================
# BASIC UTILITIES
# ============================================================================

def banner(text):
    print("\n" + "=" * 92)
    print(text)
    print("=" * 92)


def run_command(cmd, cwd=None):
    print("COMMAND:", " ".join(str(x) for x in cmd))
    try:
        return subprocess.run([str(x) for x in cmd], cwd=str(cwd) if cwd else None)
    except FileNotFoundError:
        print("COMMAND NOT FOUND:", cmd[0])
        return None


def finite(x):
    try:
        return math.isfinite(float(x))
    except Exception:
        return False


def fnum(x):
    try:
        x = float(x)
        return x if math.isfinite(x) else math.nan
    except Exception:
        return math.nan


def normalize_dt(v):
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    if hasattr(v, "to_pydatetime"):
        try:
            return v.to_pydatetime().replace(tzinfo=None)
        except Exception:
            pass
    if isinstance(v, str):
        s = v.strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
        ):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        try:
            return datetime.fromisoformat(s.replace("Z", ""))
        except ValueError:
            pass
    return None


def azdiff(a, b):
    return abs((float(a) - float(b) + 180.0) % 360.0 - 180.0)


def azmean(values):
    a = np.asarray(values, float)
    a = a[np.isfinite(a)]
    if len(a) == 0:
        return math.nan
    r = np.deg2rad(a)
    return float(np.rad2deg(
        np.arctan2(np.mean(np.sin(r)), np.mean(np.cos(r)))
    ) % 360.0)


def azsd(values):
    a = np.asarray(values, float)
    a = a[np.isfinite(a)]
    if len(a) < 2:
        return 0.0
    m = azmean(a)
    return float(np.std([azdiff(x, m) for x in a], ddof=1))


def regression(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    ok = np.isfinite(x) & np.isfinite(y)
    x, y = x[ok], y[ok]
    if len(x) < 2 or np.std(x) == 0:
        return math.nan, math.nan, math.nan
    slope, intercept = np.polyfit(x, y, 1)
    r = float(np.corrcoef(x, y)[0, 1])
    return float(slope), float(intercept), r


def residual_stats(values):
    x = np.asarray(values, float)
    x = x[np.isfinite(x)]
    if len(x) == 0:
        return dict(n=0, mean_bias=math.nan,
                    mean_abs_deviation=math.nan,
                    median_abs_deviation=math.nan, rms=math.nan)
    return dict(
        n=len(x),
        mean_bias=float(np.mean(x)),
        mean_abs_deviation=float(np.mean(np.abs(x))),
        median_abs_deviation=float(np.median(np.abs(x))),
        rms=float(np.sqrt(np.mean(x * x))),
    )


# ============================================================================
# RINEX DISCOVERY
# ============================================================================

def parse_rinex_name(path):
    m = re.search(
        r"USGS00USA_R_(\d{4})(\d{3}).*01D_01S_MO",
        path.name, re.I
    )
    if not m:
        return None
    return int(m.group(1)), int(m.group(2))


def discover_rinex():
    banner("SEARCHING FOR ALL LOCAL USGS RINEX FILES")
    found = {}

    patterns = [
        "USGS00USA_R_*_01D_01S_MO.rnx",
        "USGS00USA_R_*_01D_01S_MO.rnx.gz",
        "USGS00USA_R_*_01D_01S_MO.rnx.Z",
        "USGS00USA_R_*_01D_01S_MO.crx",
        "USGS00USA_R_*_01D_01S_MO.crx.gz",
        "USGS00USA_R_*_01D_01S_MO.crx.Z",
    ]

    for root in RINEX_ROOTS:
        root = Path(root)
        if not root.exists():
            print("Searching:", root, "[not present]")
            continue

        print("Searching:", root)
        for pattern in patterns:
            try:
                for p in root.rglob(pattern):
                    if not p.is_file():
                        continue
                    parsed = parse_rinex_name(p)
                    if not parsed:
                        continue
                    year, doy = parsed
                    if year != YEAR:
                        continue
                    key = (year, doy)
                    # Prefer an uncompressed RINEX file if both exist.
                    old = found.get(key)
                    if old is None:
                        found[key] = p
                    elif old.name.endswith((".gz", ".Z")):
                        found[key] = p
            except (PermissionError, OSError) as exc:
                print("  search warning:", exc)

    print("Unique RINEX days discovered:", len(found))
    if found:
        print("DOYs:", " ".join(f"{d:03d}" for _, d in sorted(found)))
        for (y, d), p in sorted(found.items()):
            print(f"  {y} DOY {d:03d}: {p}")
    else:
        print("DOYs: NONE")
    return found


# ============================================================================
# SNR / GNSS-IR PROCESSING
# ============================================================================

def snr_path(doy):
    return REFL_CODE / str(YEAR) / "snr" / "usgs" / (
        f"usgs{doy:03d}0.26.snr{SNR_TYPE}.gz"
    )


def result_dir():
    return REFL_CODE / str(YEAR) / "results" / "usgs" / GNSSIR_CONFIG


def result_path(doy):
    return result_dir() / f"{doy:03d}.txt"


def ensure_snr(doy):
    target = snr_path(doy)
    if target.exists() and target.stat().st_size:
        print("SNR already exists:", target)
        return True

    banner(f"CONVERTING RINEX -> SNR: DOY {doy:03d}")
    cmd = [
        "rinex2snr", STATION, str(YEAR), str(doy),
        "-snr", str(SNR_TYPE),
        "-orb", ORBIT,
        "-stream", STREAM,
        "-samplerate", str(SAMPLE_RATE),
        "-nolook", "T",
        "-overwrite", "T",
    ]
    run_command(cmd, BASE_DIR)

    ok = target.exists() and target.stat().st_size > 0
    print("SUCCESS: SNR file created" if ok else
          f"WARNING: expected SNR not found: {target}")
    return ok


def process_day(doy):
    target = result_path(doy)
    if target.exists() and target.stat().st_size:
        print("GNSS-IR result already exists:", target)
        return True

    banner(f"PROCESSING GNSS-IR: DOY {doy:03d}")

    # Established command first. The second form is a compatibility fallback.
    commands = [
        [
            "gnssir", STATION, str(YEAR), str(doy),
            "-l1", "-e5", "-gnss", "1",
            "-extension", GNSSIR_CONFIG,
        ],
        [
            "gnssir", STATION, str(YEAR), str(doy),
            "-l1", "-e5", "-gnss", "1",
        ],
    ]

    for cmd in commands:
        run_command(cmd, BASE_DIR)
        if target.exists() and target.stat().st_size:
            print("SUCCESS: GNSS-IR result created:", target)
            return True

    print("WARNING: expected GNSS-IR result not found:", target)
    return False


# ============================================================================
# RESULT LOADING -- IMPORTANT: NO RECURSIVE SEARCH
# ============================================================================

def result_files():
    """
    ONLY select:
      results/usgs/ocean17_23_l1_e5_13/*.txt

    This intentionally excludes:
      results/usgs/*.txt
      results/usgs/failQC/*.txt
      ocean17_23_l1_e5_13/failQC/*.txt
      ocean90_150/*.txt
      ocean_test/*.txt
    """
    banner("LOADING ESTABLISHED GNSS-IR RESULT FILES")

    d = result_dir()
    print("Selected configuration:", GNSSIR_CONFIG)
    print("Selected directory:", d)

    if not d.exists():
        print("ERROR: directory does not exist.")
        return []

    files = sorted(
        p for p in d.glob("*.txt")
        if p.is_file() and p.stem.isdigit()
    )

    print("Daily result files selected:", len(files))
    if files:
        print("DOYs:", " ".join(p.stem for p in files))
    return files


def parse_result(path):
    rows = []

    try:
        with open(path, "r", errors="replace") as f:
            for line in f:
                s = line.strip()
                if not s or s.startswith("%"):
                    continue

                parts = s.split()
                if len(parts) < 12:
                    continue

                try:
                    year = int(parts[0])
                    doy = int(parts[1])
                    rh = float(parts[2])
                    sat = int(parts[3])
                    utc_hour = float(parts[4])
                    az = float(parts[5])
                    rise = int(float(parts[11]))
                except (ValueError, TypeError):
                    continue

                if year != YEAR:
                    continue

                dt = datetime(YEAR, 1, 1) + timedelta(
                    days=doy - 1, hours=utc_hour
                )

                rows.append({
                    "datetime_utc": dt,
                    "year": year,
                    "doy": doy,
                    "sat": sat,
                    "rise": rise,
                    "azimuth": az % 360.0,
                    "GNSS_WL_m": rh,
                    "source": str(path),
                })
    except OSError as exc:
        print("ERROR reading", path, exc)

    return rows


def load_results():
    files = result_files()
    rows = []

    for p in files:
        parsed = parse_result(p)
        if parsed:
            print(f"{p.stem:>4} : {len(parsed):4d} observations")
            rows.extend(parsed)

    # Exact-observation deduplication safety net.
    unique = {}
    for r in rows:
        key = (
            r["datetime_utc"], r["sat"], r["rise"],
            round(r["azimuth"], 5), round(r["GNSS_WL_m"], 6)
        )
        unique[key] = r

    rows = sorted(unique.values(), key=lambda r: r["datetime_utc"])
    print("Unique GNSS-R observations loaded:", len(rows))
    return rows


# ============================================================================
# TIDE LOADING
# ============================================================================

def load_tide():
    banner("LOADING MARCONI / SHERWOOD TIDE DATA")

    tide_file = BASE_DIR / "marconi_tides_sherwood.xlsx"
    if not tide_file.exists():
        raise FileNotFoundError(tide_file)

    wb = load_workbook(tide_file, data_only=True)

    for ws in wb.worksheets:
        header_row = None
        tide_col = None
        time_col = None

        for r in range(1, min(ws.max_row, 30) + 1):
            headers = [
                str(ws.cell(r, c).value).strip()
                if ws.cell(r, c).value is not None else ""
                for c in range(1, ws.max_column + 1)
            ]

            for c, h in enumerate(headers, 1):
                if h.lower() == PRIMARY_TIDE_MODEL.lower():
                    header_row = r
                    tide_col = c
                    break
            if header_row:
                break

        if header_row is None:
            continue

        headers = [
            str(ws.cell(header_row, c).value).strip()
            if ws.cell(header_row, c).value is not None else ""
            for c in range(1, ws.max_column + 1)
        ]

        # Prefer conventional time/date headers.
        for c, h in enumerate(headers, 1):
            if h.lower() in (
                "datetime", "datetime_utc", "date_time",
                "time", "date", "utc"
            ):
                time_col = c
                break

        # Otherwise identify the first column containing datetime values.
        if time_col is None:
            for c in range(1, ws.max_column + 1):
                n = 0
                for rr in range(header_row + 1,
                                min(ws.max_row, header_row + 20) + 1):
                    if normalize_dt(ws.cell(rr, c).value) is not None:
                        n += 1
                if n >= 3:
                    time_col = c
                    break

        if time_col is None:
            continue

        times = []
        values = []

        for rr in range(header_row + 1, ws.max_row + 1):
            dt = normalize_dt(ws.cell(rr, time_col).value)
            tv = fnum(ws.cell(rr, tide_col).value)
            if dt is not None and finite(tv):
                times.append(dt)
                values.append(tv)

        if times:
            order = np.argsort(
                np.array(times, dtype="datetime64[ns]")
            )
            times = [times[i] for i in order]
            values = np.asarray(values, float)[order]

            x = np.asarray(
                [(t - times[0]).total_seconds() for t in times],
                float
            )

            def tide_at(dt, x=x, times=times, values=values):
                q = (dt - times[0]).total_seconds()
                if q < x[0] or q > x[-1]:
                    return math.nan
                return float(np.interp(q, x, values))

            print("Tide sheet:", ws.title)
            print("Tide records:", len(times))
            print("Tide interval:", times[0], "through", times[-1])
            return times, tide_at

    raise RuntimeError(
        f"Could not locate {PRIMARY_TIDE_MODEL} and a usable time column."
    )


def add_tides(rows, tide_at):
    out = []
    for r in rows:
        tide = tide_at(r["datetime_utc"])
        if not finite(tide):
            continue
        q = dict(r)
        q["PRIMARY_TIDE_m"] = tide
        q["GNSS_PLUS_0242_m"] = q["GNSS_WL_m"] + DATUM_OFFSET_M
        out.append(q)
    return out


# ============================================================================
# TRACK CLUSTERING AND SCREENING
# ============================================================================

def cluster_tracks(rows):
    groups = defaultdict(list)
    for r in rows:
        groups[(r["sat"], r["rise"])].append(r)

    tracks = []

    for (sat, rise), group in sorted(groups.items()):
        clusters = []

        for r in sorted(group, key=lambda x: x["datetime_utc"]):
            placed = False
            for c in clusters:
                center = azmean([x["azimuth"] for x in c])
                if azdiff(r["azimuth"], center) <= AZ_CLUSTER_TOL_DEG:
                    c.append(r)
                    placed = True
                    break
            if not placed:
                clusters.append([r])

        for c in clusters:
            tracks.append({
                "sat": sat,
                "rise": rise,
                "rows": c,
                "azimuth": azmean([x["azimuth"] for x in c]),
            })

    return tracks


def track_statistics(track):
    rows = track["rows"]
    tide = np.asarray([r["PRIMARY_TIDE_m"] for r in rows], float)
    gnss = np.asarray([r["GNSS_WL_m"] for r in rows], float)
    ok = np.isfinite(tide) & np.isfinite(gnss)

    tide = tide[ok]
    gnss = gnss[ok]

    good_rows = [
        rows[i] for i in np.where(ok)[0]
    ]

    days = len({
        r["datetime_utc"].date() for r in good_rows
    })

    slope, intercept, r = regression(tide, gnss)

    raw_resid = gnss - tide
    plus_resid = gnss + DATUM_OFFSET_M - tide

    return {
        "sat": track["sat"],
        "rise": track["rise"],
        "N": len(gnss),
        "days": days,
        "azimuth": track["azimuth"],
        "az_sd": azsd([r["azimuth"] for r in good_rows]),
        "tide_r": r,
        "slope": slope,
        "intercept": intercept,
        "raw_rms": (
            float(np.sqrt(np.mean(raw_resid ** 2)))
            if len(raw_resid) else math.nan
        ),
        "plus_rms": (
            float(np.sqrt(np.mean(plus_resid ** 2)))
            if len(plus_resid) else math.nan
        ),
        "rows": good_rows,
    }


def is_good(s):
    return (
        s["N"] >= MIN_OBS
        and s["days"] >= MIN_DAYS
        and finite(s["tide_r"]) and s["tide_r"] >= MIN_TIDE_R
        and finite(s["slope"]) and SLOPE_MIN <= s["slope"] <= SLOPE_MAX
        and finite(s["raw_rms"]) and s["raw_rms"] <= MAX_UNIT_RMS_M
        and finite(s["az_sd"]) and s["az_sd"] <= MAX_AZ_SD_DEG
    )


# ============================================================================
# POPULATION STATISTICS
# ============================================================================

def population_residuals(good):
    raw = []
    plus = []

    for s in good:
        for r in s["rows"]:
            if finite(r["GNSS_WL_m"]) and finite(r["PRIMARY_TIDE_m"]):
                raw.append(r["GNSS_WL_m"] - r["PRIMARY_TIDE_m"])
                plus.append(
                    r["GNSS_WL_m"] + DATUM_OFFSET_M -
                    r["PRIMARY_TIDE_m"]
                )

    return np.asarray(raw), np.asarray(plus)


def print_population_statistics(good):
    raw, plus = population_residuals(good)
    rs = residual_stats(raw)
    ps = residual_stats(plus)

    print()
    print("=" * 92)
    print("GOOD-TRACK POPULATION STATISTICS")
    print("=" * 92)
    print("GOOD tracks:", len(good))
    print("Observations:", rs["n"])

    print()
    print(f"{'Statistic':30s}{'RAW GNSS-R':>18s}"
          f"{'GNSS-R + 24.2 cm':>22s}")
    print("-" * 72)

    for label, key in (
        ("Mean bias", "mean_bias"),
        ("Mean absolute deviation", "mean_abs_deviation"),
        ("Median absolute deviation", "median_abs_deviation"),
        ("RMS", "rms"),
    ):
        print(
            f"{label:30s}"
            f"{rs[key]*100:17.2f} cm"
            f"{ps[key]*100:21.2f} cm"
        )

    return rs, ps


# ============================================================================
# OUTPUTS / PLOTS
# ============================================================================

def write_csv(stats):
    fields = [
        "rank", "sat", "rise", "N", "days", "azimuth", "az_sd",
        "tide_r", "slope", "intercept", "raw_rms_m",
        "plus_0242_rms_m", "good"
    ]

    with open(OUT_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for rank, s in enumerate(stats, 1):
            w.writerow({
                "rank": rank,
                "sat": s["sat"],
                "rise": s["rise"],
                "N": s["N"],
                "days": s["days"],
                "azimuth": s["azimuth"],
                "az_sd": s["az_sd"],
                "tide_r": s["tide_r"],
                "slope": s["slope"],
                "intercept": s["intercept"],
                "raw_rms_m": s["raw_rms"],
                "plus_0242_rms_m": s["plus_rms"],
                "good": is_good(s),
            })


def write_summary(stats, good, raw, plus):
    with open(OUT_SUMMARY, "w") as f:
        f.write("MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V4\n")
        f.write("=" * 92 + "\n\n")
        f.write(f"GNSS-IR configuration: {GNSSIR_CONFIG}\n")
        f.write(f"Primary tide model: {PRIMARY_TIDE_MODEL}\n")
        f.write(f"Diagnostic offset: +{DATUM_OFFSET_M:.3f} m\n\n")

        f.write("GOOD-track criteria:\n")
        f.write(f"  observations >= {MIN_OBS}\n")
        f.write(f"  unique days   >= {MIN_DAYS}\n")
        f.write(f"  tide r        >= {MIN_TIDE_R}\n")
        f.write(f"  slope         = {SLOPE_MIN} to {SLOPE_MAX}\n")
        f.write(f"  unit RMS      <= {MAX_UNIT_RMS_M*100:.1f} cm\n")
        f.write(f"  azimuth SD    <= {MAX_AZ_SD_DEG:.1f} deg\n")
        f.write("  +24.2 cm is NOT used to select good tracks.\n\n")

        f.write(f"Total clustered tracks: {len(stats)}\n")
        f.write(f"GOOD tracks: {len(good)}\n\n")

        f.write("GOOD-TRACK POPULATION STATISTICS\n")
        f.write("-" * 92 + "\n")

        for label, key in (
            ("Mean bias", "mean_bias"),
            ("Mean absolute deviation", "mean_abs_deviation"),
            ("Median absolute deviation", "median_abs_deviation"),
            ("RMS", "rms"),
        ):
            f.write(
                f"{label:30s}"
                f"{raw[key]*100:17.2f} cm"
                f"{plus[key]*100:21.2f} cm\n"
            )


def make_plots(good):
    PLOT_DIR.mkdir(parents=True, exist_ok=True)

    rows = []
    for s in good:
        rows.extend(s["rows"])
    rows.sort(key=lambda r: r["datetime_utc"])

    if not rows:
        return

    t = [r["datetime_utc"] for r in rows]
    tide = np.asarray([r["PRIMARY_TIDE_m"] for r in rows])
    gnss = np.asarray([r["GNSS_WL_m"] for r in rows])
    plus = gnss + DATUM_OFFSET_M
    ok = np.isfinite(tide) & np.isfinite(gnss)

    # Raw good-track plot.
    fig, ax = plt.subplots(figsize=(14, 8))
    ax.plot(np.asarray(t)[ok], tide[ok], linewidth=2, label=PRIMARY_TIDE_MODEL)
    ax.scatter(np.asarray(t)[ok], gnss[ok], s=16, alpha=.65, label="GNSS-R raw")
    ax.set_xlabel("UTC")
    ax.set_ylabel("Height (m)")
    ax.set_title("Marconi GNSS-R vs EOT20 — GOOD tracks only")
    ax.grid(True, alpha=.25)
    ax.legend()
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "01_good_tracks_vs_tide_raw.png", dpi=180)
    plt.close(fig)

    # +24.2 cm plot.
    fig, ax = plt.subplots(figsize=(14, 8))
    ax.plot(np.asarray(t)[ok], tide[ok], linewidth=2, label=PRIMARY_TIDE_MODEL)
    ax.scatter(np.asarray(t)[ok], plus[ok], s=16, alpha=.65,
               label="GNSS-R + 0.242 m")
    ax.set_xlabel("UTC")
    ax.set_ylabel("Height (m)")
    ax.set_title("Marconi GNSS-R vs EOT20 — GOOD tracks + 24.2 cm")
    ax.grid(True, alpha=.25)
    ax.legend()
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "02_good_tracks_vs_tide_plus_0242m.png", dpi=180)
    plt.close(fig)

    # Residual histogram.
    raw, p24 = population_residuals(good)
    fig, ax = plt.subplots(figsize=(12, 7))
    ax.hist(raw * 100, bins=30, alpha=.55, label="Raw GNSS-R")
    ax.hist(p24 * 100, bins=30, alpha=.55, label="GNSS-R + 24.2 cm")
    ax.axvline(0, linewidth=1.5)
    ax.set_xlabel("GNSS-R − tide residual (cm)")
    ax.set_ylabel("Number of observations")
    ax.set_title("GOOD-track population residuals")
    ax.grid(True, alpha=.20)
    ax.legend()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "03_good_population_residuals.png", dpi=180)
    plt.close(fig)


# ============================================================================
# MAIN
# ============================================================================

def main():
    banner("MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V4")

    print("Base directory:", BASE_DIR)
    print("REFL_CODE     :", REFL_CODE)
    print("Primary tide  :", PRIMARY_TIDE_MODEL)
    print("GNSSIR config :", GNSSIR_CONFIG)
    print(f"Diagnostic offset: +{DATUM_OFFSET_M:.3f} m")

    print()
    print("GOOD-track criteria:")
    print(f"  observations >= {MIN_OBS}")
    print(f"  unique days   >= {MIN_DAYS}")
    print(f"  tide r        >= {MIN_TIDE_R}")
    print(f"  slope         = {SLOPE_MIN} to {SLOPE_MAX}")
    print(f"  unit RMS      <= {MAX_UNIT_RMS_M*100:.1f} cm")
    print(f"  azimuth SD    <= {MAX_AZ_SD_DEG:.1f} deg")
    print("  +24.2 cm is NOT used to select good tracks.")

    rinex = discover_rinex()

    if rinex:
        banner("CHECKING / GENERATING MISSING SNR FILES")
        for (_, doy), _path in sorted(rinex.items()):
            ensure_snr(doy)

        banner("CHECKING / GENERATING MISSING GNSS-IR RESULTS")
        for (_, doy), _path in sorted(rinex.items()):
            if snr_path(doy).exists():
                process_day(doy)
    else:
        print("No local RINEX files discovered.")
        print("Continuing with existing GNSS-IR production results.")

    rows = load_results()
    if not rows:
        raise SystemExit(
            "No observations found in " +
            str(result_dir())
        )

    _, tide_at = load_tide()
    rows = add_tides(rows, tide_at)

    print("GNSS-R observations matched to tide:", len(rows))
    if rows:
        print("Observation interval:", rows[0]["datetime_utc"],
              "through", rows[-1]["datetime_utc"])

    tracks = cluster_tracks(rows)
    stats = [track_statistics(t) for t in tracks]
    good = [s for s in stats if is_good(s)]
    good.sort(key=lambda s: (-s["tide_r"], s["raw_rms"]))

    print()
    print("=" * 92)
    print("GOOD TRACK SCREEN")
    print("=" * 92)
    print("Total clustered tracks:", len(stats))
    print("GOOD tracks:", len(good))

    for i, s in enumerate(good, 1):
        print(
            f"{i:2d} SAT={s['sat']:3d} rise={s['rise']:2d} "
            f"N={s['N']:3d} days={s['days']:2d} "
            f"Az={s['azimuth']:7.2f}+/-{s['az_sd']:.2f} "
            f"r={s['tide_r']:+.4f} slope={s['slope']:+.4f} "
            f"rawRMS={s['raw_rms']*100:.2f}cm "
            f"+24.2RMS={s['plus_rms']*100:.2f}cm "
            f"improvement={(s['raw_rms']-s['plus_rms'])*100:+.2f}cm"
        )

    raw, plus = print_population_statistics(good)

    write_csv(stats)
    write_summary(stats, good, raw, plus)
    make_plots(good)

    banner("OUTPUTS")
    print("CSV:", OUT_CSV)
    print("Summary:", OUT_SUMMARY)
    print("Plots:", PLOT_DIR)
    print("DONE")


if __name__ == "__main__":
    main()
