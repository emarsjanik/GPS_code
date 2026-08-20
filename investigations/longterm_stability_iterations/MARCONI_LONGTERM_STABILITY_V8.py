#!/usr/bin/env python3
"""
MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V8

Purpose
-------
Process the established Marconi GNSS-IR result files against the EOT20 tide
model, using the GNSS antenna/reference height correctly:

    GNSS-R water elevation = GNSS_REFERENCE_HEIGHT_M - RH

where RH is the reflector height reported by gnssrefl.

The +0.242 m datum offset is diagnostic only.  It is never used to decide
whether a track is GOOD.

Outputs
-------
marconi_longterm_v8/
    marconi_longterm_track_diagnostics_v8.csv
    marconi_longterm_population_statistics_v8.csv
    marconi_longterm_summary_v8.txt
    plots/

This version intentionally reads ONLY the established GNSS-IR result
configuration:
    ocean17_23_l1_e5_13

It does not attempt to rediscover or regenerate RINEX files.  The GNSS-IR
results already produced by gnssrefl are the authoritative input here.
"""

from __future__ import annotations

import csv
import math
import os
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required.")
    sys.exit(1)


# ============================================================================
# CONFIGURATION
# ============================================================================

BASE_DIR = Path.home() / "GNSS" / "v4.1"

REFL_CODE = Path(
    os.environ.get("REFL_CODE", str(BASE_DIR / "products" / "refl_code"))
)

YEAR = 2026
STATION = "usgs"

CONFIG = "ocean17_23_l1_e5_13"

RESULT_DIR = (
    REFL_CODE / str(YEAR) / "results" / STATION / CONFIG
)

TIDE_FILE = BASE_DIR / "marconi_tides_sherwood.xlsx"
TIDE_MODEL_COLUMN = "EOT20_heightm"

# This is the transformation required by the V7 diagnostic.
GNSS_REFERENCE_HEIGHT_M = 18.665

# Diagnostic datum test.
DATUM_OFFSET_M = 0.242

# Established GOOD-track criteria.
MIN_OBS = 14
MIN_DAYS = 14
MIN_TIDE_R = 0.90
SLOPE_MIN = 0.85
SLOPE_MAX = 1.15
MAX_UNIT_RMS_M = 0.30
MAX_AZ_SD_DEG = 1.0

AZ_CLUSTER_TOL_DEG = 3.0

OUT_DIR = BASE_DIR / "marconi_longterm_v8"
PLOT_DIR = OUT_DIR / "plots"

OUT_CSV = OUT_DIR / "marconi_longterm_track_diagnostics_v8.csv"
POP_CSV = OUT_DIR / "marconi_longterm_population_statistics_v8.csv"
SUMMARY = OUT_DIR / "marconi_longterm_summary_v8.txt"


# ============================================================================
# UTILITIES
# ============================================================================

def banner(text: str):
    print("\n" + "=" * 100)
    print(text)
    print("=" * 100)


def fnum(value):
    try:
        x = float(value)
        return x if math.isfinite(x) else math.nan
    except (TypeError, ValueError):
        return math.nan


def normalize_datetime(value):
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)

    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime().replace(tzinfo=None)
        except Exception:
            pass

    if isinstance(value, str):
        s = value.strip()
        formats = (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
        )
        for fmt in formats:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        try:
            return datetime.fromisoformat(s.replace("Z", ""))
        except ValueError:
            return None

    return None


def azdiff(a, b):
    return abs((a - b + 180.0) % 360.0 - 180.0)


def azmean(values):
    a = np.asarray(values, dtype=float)
    a = a[np.isfinite(a)]
    if len(a) == 0:
        return math.nan

    rad = np.deg2rad(a)
    return float(
        np.rad2deg(
            np.arctan2(np.mean(np.sin(rad)), np.mean(np.cos(rad)))
        ) % 360.0
    )


def azsd(values):
    a = np.asarray(values, dtype=float)
    a = a[np.isfinite(a)]

    if len(a) < 2:
        return 0.0

    m = azmean(a)
    return float(np.std([azdiff(x, m) for x in a], ddof=1))


def regression(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    ok = np.isfinite(x) & np.isfinite(y)
    x = x[ok]
    y = y[ok]

    if len(x) < 2 or np.std(x) == 0 or np.std(y) == 0:
        return math.nan, math.nan, math.nan

    slope, intercept = np.polyfit(x, y, 1)
    r = float(np.corrcoef(x, y)[0, 1])

    return float(slope), float(intercept), r


def residual_stats(values):
    a = np.asarray(values, dtype=float)
    a = a[np.isfinite(a)]

    if len(a) == 0:
        return {
            "n": 0,
            "mean_bias": math.nan,
            "mean_abs_deviation": math.nan,
            "median_abs_deviation": math.nan,
            "rms": math.nan,
        }

    return {
        "n": int(len(a)),
        "mean_bias": float(np.mean(a)),
        "mean_abs_deviation": float(np.mean(np.abs(a))),
        "median_abs_deviation": float(np.median(np.abs(a))),
        "rms": float(np.sqrt(np.mean(a * a))),
    }


# ============================================================================
# GNSS-IR RESULT PARSER
# ============================================================================

def parse_result(path: Path):
    """
    Parse a standard gnssrefl result text file.

    Columns used:
      year  doy  RH  sat  UTCtime  Azim ... rise ...
    """
    rows = []

    try:
        with path.open("r", errors="replace") as f:
            for line in f:
                s = line.strip()

                if not s or s.startswith("%") or s.startswith("#"):
                    continue

                parts = s.split()

                if len(parts) < 17:
                    continue

                try:
                    year = int(parts[0])
                    doy = int(parts[1])
                    rh = float(parts[2])
                    sat = int(parts[3])
                    utc_hours = float(parts[4])
                    az = float(parts[5])
                    rise = int(float(parts[11]))
                except (ValueError, IndexError):
                    continue

                if year != YEAR:
                    continue

                if not all(
                    math.isfinite(x)
                    for x in (rh, utc_hours, az)
                ):
                    continue

                midnight = datetime(year, 1, 1) + timedelta(days=doy - 1)

                seconds = utc_hours * 3600.0
                dt = midnight + timedelta(seconds=seconds)

                # CRITICAL V8 TRANSFORMATION:
                #
                # gnssrefl RH is reflector height above the GNSS reference.
                # Convert it to water-surface elevation before comparison
                # with the tide model.
                gnss_water = GNSS_REFERENCE_HEIGHT_M - rh

                rows.append(
                    {
                        "datetime_utc": dt,
                        "doy": doy,
                        "sat": sat,
                        "rise": rise,
                        "azimuth": az,
                        "RH_m": rh,
                        "GNSS_WL_m": gnss_water,
                        "source": str(path),
                    }
                )

    except OSError as e:
        print("WARNING: could not read", path, ":", e)

    return rows


def result_files():
    """
    Select only daily numeric result files from the requested configuration.

    Files such as usgs_2026_subdaily_all.txt are deliberately excluded.
    """
    if not RESULT_DIR.exists():
        return []

    files = []

    for p in RESULT_DIR.iterdir():
        if not p.is_file():
            continue

        if not re.fullmatch(r"\d{3}\.txt", p.name):
            continue

        try:
            doy = int(p.stem)
        except ValueError:
            continue

        if 1 <= doy <= 366:
            files.append(p)

    return sorted(files, key=lambda p: int(p.stem))


def load_results():
    banner("LOADING ESTABLISHED GNSS-IR RESULT FILES")

    print("Selected configuration:", CONFIG)
    print("Selected directory:", RESULT_DIR)

    files = result_files()

    print("Daily result files selected:", len(files))

    all_rows = []

    for p in files:
        rows = parse_result(p)

        if rows:
            print(f"{int(p.stem):4d} : {len(rows):5d} observations")
            all_rows.extend(rows)

    # De-duplicate exact observations.
    unique = {}

    for r in all_rows:
        key = (
            r["datetime_utc"],
            r["sat"],
            r["rise"],
            round(r["azimuth"], 5),
            round(r["RH_m"], 6),
        )
        unique[key] = r

    rows = sorted(
        unique.values(),
        key=lambda r: r["datetime_utc"]
    )

    print("Unique GPS L1 GNSS-R observations:", len(rows))

    return rows


# ============================================================================
# TIDE MODEL
# ============================================================================

def load_tide():
    banner("LOADING MARCONI / SHERWOOD TIDE DATA")

    if not TIDE_FILE.exists():
        raise FileNotFoundError(
            f"Tide workbook not found: {TIDE_FILE}"
        )

    wb = load_workbook(TIDE_FILE, data_only=True)

    if "in" in wb.sheetnames:
        ws = wb["in"]
    else:
        ws = wb[wb.sheetnames[0]]

    headers = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is not None:
            headers[str(value).strip()] = col

    # Find datetime column robustly.
    time_col = None

    for name, col in headers.items():
        low = name.lower()

        if (
            "time" in low
            or "date" in low
            or "datetime" in low
        ):
            time_col = col
            break

    if time_col is None:
        time_col = 1

    if TIDE_MODEL_COLUMN not in headers:
        # Fall back to exact case-insensitive match.
        match = None
        for name, col in headers.items():
            if name.lower() == TIDE_MODEL_COLUMN.lower():
                match = col
                break

        if match is None:
            raise RuntimeError(
                f"Could not find tide column {TIDE_MODEL_COLUMN!r}. "
                f"Columns found: {list(headers)}"
            )

        tide_col = match
    else:
        tide_col = headers[TIDE_MODEL_COLUMN]

    tide = []

    for row in ws.iter_rows(min_row=2):
        dt = normalize_datetime(row[time_col - 1].value)
        val = fnum(row[tide_col - 1].value)

        if dt is None or not math.isfinite(val):
            continue

        tide.append((dt, val))

    tide.sort()

    print("Tide sheet:", ws.title)
    print("Tide records:", len(tide))

    if tide:
        print("Tide interval:", tide[0][0], "through", tide[-1][0])

    print("Tide column:", TIDE_MODEL_COLUMN)

    return tide


def interpolate_tide(tide, dt):
    if not tide:
        return math.nan

    times = np.array(
        [(t - tide[0][0]).total_seconds() for t, _ in tide],
        dtype=float,
    )
    vals = np.array([v for _, v in tide], dtype=float)

    x = (dt - tide[0][0]).total_seconds()

    if x < times[0] or x > times[-1]:
        return math.nan

    return float(np.interp(x, times, vals))


def match_tide(rows, tide):
    matched = []

    for r in rows:
        model = interpolate_tide(tide, r["datetime_utc"])

        if not math.isfinite(model):
            continue

        rr = dict(r)

        rr["tide_m"] = model

        # RAW residual:
        # GNSS-R water elevation minus tide model.
        rr["residual_raw_m"] = (
            rr["GNSS_WL_m"] - rr["tide_m"]
        )

        # Diagnostic +24.2 cm residual.
        rr["residual_plus242_m"] = (
            rr["GNSS_WL_m"]
            + DATUM_OFFSET_M
            - rr["tide_m"]
        )

        matched.append(rr)

    print("GNSS-R observations matched to tide:", len(matched))

    if matched:
        print(
            "Observation interval:",
            matched[0]["datetime_utc"],
            "through",
            matched[-1]["datetime_utc"],
        )

    return matched


# ============================================================================
# TRACK CLUSTERING
# ============================================================================

def cluster_tracks(rows):
    """
    Cluster by satellite, rise/set flag, and azimuth.

    The azimuth is treated circularly.
    """
    groups = []

    # Sort to make behavior deterministic.
    ordered = sorted(
        rows,
        key=lambda r: (
            r["sat"],
            r["rise"],
            r["azimuth"],
            r["datetime_utc"],
        ),
    )

    for r in ordered:
        placed = False

        for g in groups:
            if r["sat"] != g["sat"]:
                continue

            if r["rise"] != g["rise"]:
                continue

            if azdiff(r["azimuth"], g["azmean"]) <= AZ_CLUSTER_TOL_DEG:
                g["rows"].append(r)
                g["azmean"] = azmean(
                    [x["azimuth"] for x in g["rows"]]
                )
                placed = True
                break

        if not placed:
            groups.append(
                {
                    "sat": r["sat"],
                    "rise": r["rise"],
                    "azmean": r["azimuth"],
                    "rows": [r],
                }
            )

    return groups


# ============================================================================
# TRACK SCREENING
# ============================================================================

def screen_track(group):
    rows = group["rows"]

    x = np.array(
        [r["tide_m"] for r in rows],
        dtype=float,
    )

    y = np.array(
        [r["GNSS_WL_m"] for r in rows],
        dtype=float,
    )

    ok = np.isfinite(x) & np.isfinite(y)

    x = x[ok]
    y = y[ok]

    used_rows = [
        r for r, good in zip(rows, ok) if good
    ]

    n = len(y)

    days = len(
        {
            r["datetime_utc"].date()
            for r in used_rows
        }
    )

    az = np.array(
        [r["azimuth"] for r in used_rows],
        dtype=float,
    )

    az_mean = azmean(az)
    az_sd = azsd(az)

    slope, intercept, rcoef = regression(x, y)

    residual = y - x

    rms = (
        float(np.sqrt(np.mean(residual * residual)))
        if len(residual)
        else math.nan
    )

    failures = []

    if n < MIN_OBS:
        failures.append("N")

    if days < MIN_DAYS:
        failures.append("DAYS")

    if not math.isfinite(rcoef) or rcoef < MIN_TIDE_R:
        failures.append("R")

    if (
        not math.isfinite(slope)
        or slope < SLOPE_MIN
        or slope > SLOPE_MAX
    ):
        failures.append("SLOPE")

    if (
        not math.isfinite(rms)
        or rms > MAX_UNIT_RMS_M
    ):
        failures.append("RMS")

    if (
        not math.isfinite(az_sd)
        or az_sd > MAX_AZ_SD_DEG
    ):
        failures.append("AZSD")

    good = len(failures) == 0

    if good:
        category = "GOOD"
    elif (
        n >= MIN_OBS
        and days >= MIN_DAYS
        and math.isfinite(rcoef)
        and rcoef >= 0.85
        and math.isfinite(slope)
        and 0.75 <= slope <= 1.25
    ):
        category = "NEAR-GOOD"
    else:
        category = "BAD"

    return {
        "sat": group["sat"],
        "rise": group["rise"],
        "n": n,
        "days": days,
        "azmean": az_mean,
        "azsd": az_sd,
        "r": rcoef,
        "slope": slope,
        "intercept": intercept,
        "rms_m": rms,
        "category": category,
        "fail": ",".join(failures),
        "rows": used_rows,
    }


# ============================================================================
# POPULATION STATISTICS
# ============================================================================

def population_stats(track_results):
    good_rows = []

    for tr in track_results:
        if tr["category"] == "GOOD":
            good_rows.extend(tr["rows"])

    raw = residual_stats(
        [r["residual_raw_m"] for r in good_rows]
    )

    plus = residual_stats(
        [r["residual_plus242_m"] for r in good_rows]
    )

    return good_rows, raw, plus


# ============================================================================
# CSV / SUMMARY OUTPUT
# ============================================================================

def write_track_csv(results):
    fields = [
        "track_id",
        "sat",
        "rise",
        "n",
        "days",
        "azmean_deg",
        "azsd_deg",
        "tide_r",
        "slope",
        "intercept_m",
        "raw_rms_cm",
        "plus242_rms_cm",
        "category",
        "fail",
    ]

    with OUT_CSV.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        for i, r in enumerate(results, 1):
            w.writerow(
                {
                    "track_id": i,
                    "sat": r["sat"],
                    "rise": r["rise"],
                    "n": r["n"],
                    "days": r["days"],
                    "azmean_deg": r["azmean"],
                    "azsd_deg": r["azsd"],
                    "tide_r": r["r"],
                    "slope": r["slope"],
                    "intercept_m": r["intercept"],
                    "raw_rms_cm": (
                        r["rms_m"] * 100
                        if math.isfinite(r["rms_m"])
                        else math.nan
                    ),
                    "plus242_rms_cm": (
                        float(
                            np.sqrt(
                                np.mean(
                                    [
                                        rr["residual_plus242_m"] ** 2
                                        for rr in r["rows"]
                                    ]
                                )
                            )
                            * 100
                        )
                        if r["rows"]
                        else math.nan
                    ),
                    "category": r["category"],
                    "fail": r["fail"],
                }
            )


def write_population_csv(raw, plus):
    fields = [
        "dataset",
        "n_observations",
        "mean_bias_cm",
        "mean_absolute_deviation_cm",
        "median_absolute_deviation_cm",
        "rms_cm",
    ]

    with POP_CSV.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        for name, s in (
            ("RAW", raw),
            ("GNSS-R +0.242 m", plus),
        ):
            w.writerow(
                {
                    "dataset": name,
                    "n_observations": s["n"],
                    "mean_bias_cm": s["mean_bias"] * 100,
                    "mean_absolute_deviation_cm":
                        s["mean_abs_deviation"] * 100,
                    "median_absolute_deviation_cm":
                        s["median_abs_deviation"] * 100,
                    "rms_cm": s["rms"] * 100,
                }
            )


def write_summary(track_results, good_rows, raw, plus, n_total):
    with SUMMARY.open("w") as f:
        f.write("MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V8\n")
        f.write("=" * 100 + "\n\n")

        f.write(f"Base directory: {BASE_DIR}\n")
        f.write(f"REFL_CODE: {REFL_CODE}\n")
        f.write(f"Result directory: {RESULT_DIR}\n")
        f.write(
            f"GNSS reference height: "
            f"{GNSS_REFERENCE_HEIGHT_M:.3f} m\n"
        )
        f.write(
            f"Diagnostic offset: +{DATUM_OFFSET_M:.3f} m\n\n"
        )

        f.write("GOOD-track criteria:\n")
        f.write(f"  observations >= {MIN_OBS}\n")
        f.write(f"  unique days   >= {MIN_DAYS}\n")
        f.write(f"  tide r        >= {MIN_TIDE_R}\n")
        f.write(
            f"  slope         = {SLOPE_MIN} to {SLOPE_MAX}\n"
        )
        f.write(
            f"  unit RMS      <= {MAX_UNIT_RMS_M * 100:.1f} cm\n"
        )
        f.write(
            f"  azimuth SD    <= {MAX_AZ_SD_DEG:.1f} deg\n"
        )
        f.write(
            "  +24.2 cm is NOT used to select good tracks.\n\n"
        )

        f.write(
            "GNSS-R transformation:\n"
            "  GNSS-R water elevation = 18.665 m - RH\n\n"
        )

        f.write(f"Total GNSS-R observations loaded: {n_total}\n")
        f.write(
            f"GOOD tracks: "
            f"{sum(r['category'] == 'GOOD' for r in track_results)}\n"
        )
        f.write(f"GOOD observations: {len(good_rows)}\n\n")

        for label, s in (
            ("RAW", raw),
            ("GNSS-R +0.242 m", plus),
        ):
            f.write(label + "\n")
            f.write("-" * len(label) + "\n")
            f.write(f"  observations             = {s['n']}\n")
            f.write(
                f"  mean bias                = "
                f"{s['mean_bias'] * 100:.3f} cm\n"
            )
            f.write(
                f"  mean absolute deviation = "
                f"{s['mean_abs_deviation'] * 100:.3f} cm\n"
            )
            f.write(
                f"  median absolute deviation = "
                f"{s['median_abs_deviation'] * 100:.3f} cm\n"
            )
            f.write(
                f"  RMS                      = "
                f"{s['rms'] * 100:.3f} cm\n\n"
            )

        f.write("TRACKS\n")
        f.write("=" * 100 + "\n")

        for i, r in enumerate(track_results, 1):
            f.write(
                f"{i:2d} SAT={r['sat']:2d} "
                f"rise={r['rise']:2d} "
                f"N={r['n']:3d} "
                f"days={r['days']:3d} "
                f"Az={r['azmean']:7.2f}+/-{r['azsd']:.2f} "
                f"r={r['r']:+.4f} "
                f"slope={r['slope']:+.4f} "
                f"rawRMS={r['rms_m'] * 100:.2f}cm "
                f"{r['category']}\n"
            )


# ============================================================================
# PLOTS
# ============================================================================

def make_plots(track_results, good_rows):
    PLOT_DIR.mkdir(parents=True, exist_ok=True)

    # 1. All good tracks: GNSS-R versus tide.
    if good_rows:
        tide = np.array(
            [r["tide_m"] for r in good_rows],
            dtype=float,
        )

        gnss = np.array(
            [r["GNSS_WL_m"] for r in good_rows],
            dtype=float,
        )

        order = np.argsort(
            [r["datetime_utc"] for r in good_rows]
        )

        dates = np.array(
            [r["datetime_utc"] for r in good_rows],
            dtype=object,
        )

        raw_resid_cm = (
            (gnss - tide) * 100.0
        )

        plus_resid_cm = (
            (gnss + DATUM_OFFSET_M - tide) * 100.0
        )

        fig = plt.figure(figsize=(12, 7))
        ax = fig.add_subplot(111)

        ax.plot(
            dates[order],
            gnss[order],
            ".",
            label="GNSS-R water elevation",
        )

        ax.plot(
            dates[order],
            tide[order],
            "-",
            label="EOT20 tide",
        )

        ax.set_xlabel("UTC")
        ax.set_ylabel("Elevation (m)")
        ax.set_title(
            "Marconi GNSS-R vs EOT20 — GOOD-track population"
        )
        ax.legend()
        ax.grid(True, alpha=0.3)
        fig.autofmt_xdate()
        fig.tight_layout()

        fig.savefig(
            PLOT_DIR / "marconi_good_population_timeseries_v8.png",
            dpi=180,
        )
        plt.close(fig)

        # Residual plot.
        fig = plt.figure(figsize=(12, 6))
        ax = fig.add_subplot(111)

        ax.plot(
            dates[order],
            raw_resid_cm[order],
            ".",
            label="RAW residual",
        )

        ax.plot(
            dates[order],
            plus_resid_cm[order],
            ".",
            label="GNSS-R +0.242 m residual",
        )

        ax.axhline(0.0, linewidth=1)
        ax.set_xlabel("UTC")
        ax.set_ylabel("GNSS-R minus EOT20 (cm)")
        ax.set_title(
            "Marconi GNSS-R residuals — GOOD-track population"
        )
        ax.legend()
        ax.grid(True, alpha=0.3)
        fig.autofmt_xdate()
        fig.tight_layout()

        fig.savefig(
            PLOT_DIR / "marconi_good_population_residuals_v8.png",
            dpi=180,
        )
        plt.close(fig)

    # 2. Per-track plots.
    for i, tr in enumerate(track_results, 1):
        if tr["category"] != "GOOD":
            continue

        rows = tr["rows"]

        if len(rows) < 2:
            continue

        x = np.array([r["tide_m"] for r in rows])
        y = np.array([r["GNSS_WL_m"] for r in rows])

        fig = plt.figure(figsize=(7, 6))
        ax = fig.add_subplot(111)

        ax.scatter(x, y)

        lo = min(np.min(x), np.min(y))
        hi = max(np.max(x), np.max(y))

        ax.plot([lo, hi], [lo, hi], linestyle="--")

        if math.isfinite(tr["slope"]) and math.isfinite(tr["intercept"]):
            xx = np.linspace(lo, hi, 100)
            yy = tr["intercept"] + tr["slope"] * xx
            ax.plot(xx, yy)

        ax.set_xlabel("EOT20 height (m)")
        ax.set_ylabel("GNSS-R water elevation (m)")
        ax.set_title(
            f"GOOD Track {i}: SAT {tr['sat']} rise {tr['rise']}"
        )
        ax.grid(True, alpha=0.3)
        fig.tight_layout()

        fig.savefig(
            PLOT_DIR / f"track_{i:02d}_sat{tr['sat']}_r{tr['rise']}.png",
            dpi=180,
        )
        plt.close(fig)


# ============================================================================
# MAIN
# ============================================================================

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PLOT_DIR.mkdir(parents=True, exist_ok=True)

    banner("MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V8")

    print("Base directory:", BASE_DIR)
    print("REFL_CODE:", REFL_CODE)
    print("Result directory:", RESULT_DIR)
    print(
        f"GNSS reference height: "
        f"{GNSS_REFERENCE_HEIGHT_M:.3f} m"
    )
    print(
        f"Diagnostic offset: "
        f"+{DATUM_OFFSET_M:.3f} m"
    )
    print(
        "CRITICAL V8 TRANSFORMATION: "
        "GNSS water elevation = 18.665 - RH"
    )
    print(
        "GOOD-track screen uses RAW transformed GNSS-R elevation."
    )
    print(
        "+0.242 m is diagnostic only."
    )

    rows = load_results()

    if not rows:
        print("\nERROR: no GNSS-IR observations were loaded.")
        return 1

    tide = load_tide()
    rows = match_tide(rows, tide)

    if not rows:
        print("\nERROR: no observations matched the tide model.")
        return 1

    banner("TRACK-BY-TRACK GOOD / NEAR-GOOD / BAD DIAGNOSTICS")

    groups = cluster_tracks(rows)
    results = [
        screen_track(g)
        for g in groups
    ]

    # Sort strongest correlations first, matching the useful V2 display.
    results.sort(
        key=lambda r: (
            -r["r"] if math.isfinite(r["r"]) else float("inf")
        )
    )

    print(
        "ID SAT R N DAYS AZmean AZsd r slope RMScm CATEGORY FAIL"
    )
    print("-" * 100)

    for i, r in enumerate(results, 1):
        print(
            f"{i:2d} "
            f"{r['sat']:3d} "
            f"{r['rise']:2d} "
            f"{r['n']:3d} "
            f"{r['days']:4d} "
            f"{r['azmean']:7.2f} "
            f"{r['azsd']:5.2f} "
            f"{r['r']:+.4f} "
            f"{r['slope']:+.4f} "
            f"{r['rms_m'] * 100:8.2f} "
            f"{r['category']:<9s} "
            f"{r['fail']}"
        )

    good_rows, raw, plus = population_stats(results)

    banner("GOOD-TRACK POPULATION STATISTICS")

    good_tracks = sum(
        r["category"] == "GOOD"
        for r in results
    )

    near_tracks = sum(
        r["category"] == "NEAR-GOOD"
        for r in results
    )

    bad_tracks = sum(
        r["category"] == "BAD"
        for r in results
    )

    print("GOOD tracks:", good_tracks)
    print("GOOD observations:", len(good_rows))
    print()

    print("RAW")
    print(
        f"  mean bias                  = "
        f"{raw['mean_bias'] * 100:.3f} cm"
    )
    print(
        f"  mean absolute deviation   = "
        f"{raw['mean_abs_deviation'] * 100:.3f} cm"
    )
    print(
        f"  median absolute deviation = "
        f"{raw['median_abs_deviation'] * 100:.3f} cm"
    )
    print(
        f"  RMS                        = "
        f"{raw['rms'] * 100:.3f} cm"
    )

    print()
    print("GNSS-R +0.242 m")
    print(
        f"  mean bias                  = "
        f"{plus['mean_bias'] * 100:.3f} cm"
    )
    print(
        f"  mean absolute deviation   = "
        f"{plus['mean_abs_deviation'] * 100:.3f} cm"
    )
    print(
        f"  median absolute deviation = "
        f"{plus['median_abs_deviation'] * 100:.3f} cm"
    )
    print(
        f"  RMS                        = "
        f"{plus['rms'] * 100:.3f} cm"
    )

    banner("GENERATING TRACK PLOTS")

    write_track_csv(results)
    write_population_csv(raw, plus)
    write_summary(
        results,
        good_rows,
        raw,
        plus,
        len(rows),
    )
    make_plots(results, good_rows)

    print("Track CSV:", OUT_CSV)
    print("Population CSV:", POP_CSV)
    print("Summary:", SUMMARY)
    print("Plots:", PLOT_DIR)

    banner("DONE")

    print(
        f"GOOD: {good_tracks}  "
        f"NEAR-GOOD: {near_tracks}  "
        f"BAD: {bad_tracks}"
    )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
