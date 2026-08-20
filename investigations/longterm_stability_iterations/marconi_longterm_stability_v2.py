#!/usr/bin/env python3
"""
MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V2

Discovers local USGS00USA RINEX files, makes missing SNR files with
rinex2snr, runs the established gnssrefl GNSS-IR configuration, loads the
Marconi/Sherwood tide workbook, screens satellite/rise/azimuth tracks, and
reports raw and +0.242 m population statistics.

The +0.242 m offset is diagnostic only; it is NOT used to select good tracks.
"""

from __future__ import annotations
import csv
import math
import os
import re
import subprocess
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: install openpyxl in the gnssrefl environment.")
    sys.exit(1)

# ---------------------------------------------------------------------------
# CONFIGURATION
# ---------------------------------------------------------------------------

BASE_DIR = Path.home() / "GNSS" / "v4.1"
REFL_CODE = Path(os.environ.get(
    "REFL_CODE", str(BASE_DIR / "products" / "refl_code")
))

TIDE_FILE = BASE_DIR / "marconi_tides_sherwood.xlsx"
STATION = "usgs00usa"
YEAR = 2026

PRIMARY_TIDE_MODEL = "EOT20_heightm"
GNSSIR_CONFIG = "ocean17_23_l1_e5_13"
DATUM_OFFSET_M = 0.242

SNR_TYPE = 66
ORBIT = "gnss"
STREAM = "R"
SAMPLE_RATE = 1

# Established good-track screen.
MIN_OBS = 14
MIN_DAYS = 14
MIN_TIDE_R = 0.90
SLOPE_MIN = 0.85
SLOPE_MAX = 1.15
MAX_UNIT_RMS_M = 0.30
MAX_AZ_SD_DEG = 1.0
AZ_CLUSTER_TOL_DEG = 3.0

RINEX_ROOTS = [
    BASE_DIR,
    Path.home(),
    Path("/mnt/I2Rgus_Data"),
]

PLOT_DIR = BASE_DIR / "marconi_longterm_plots_v2"
OUT_CSV = BASE_DIR / "marconi_longterm_track_stability_v2.csv"
OUT_SUMMARY = BASE_DIR / "marconi_longterm_track_stability_v2_summary.txt"


# ---------------------------------------------------------------------------
# UTILITIES
# ---------------------------------------------------------------------------

def banner(s):
    print("\n" + "=" * 92)
    print(s)
    print("=" * 92)


def run(cmd, cwd=None):
    try:
        return subprocess.run(cmd, cwd=str(cwd) if cwd else None)
    except FileNotFoundError:
        print("COMMAND NOT FOUND:", cmd[0])
        return None


def normalize_dt(v):
    if isinstance(v, datetime):
        return v.replace(tzinfo=None)
    if hasattr(v, "to_pydatetime"):
        return v.to_pydatetime().replace(tzinfo=None)
    if isinstance(v, str):
        s = v.strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
            "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M",
            "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M",
        ):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        try:
            return datetime.fromisoformat(s.replace("Z", ""))
        except ValueError:
            return None
    return None


def fnum(v):
    try:
        x = float(v)
        return x if math.isfinite(x) else math.nan
    except (TypeError, ValueError):
        return math.nan


def azdiff(a, b):
    return abs((a - b + 180.0) % 360.0 - 180.0)


def azmean(a):
    a = np.asarray(a, float)
    a = a[np.isfinite(a)]
    if not len(a):
        return math.nan
    r = np.deg2rad(a)
    return float(np.rad2deg(np.arctan2(
        np.mean(np.sin(r)), np.mean(np.cos(r))
    )) % 360.0)


def azsd(a):
    a = np.asarray(a, float)
    a = a[np.isfinite(a)]
    if len(a) < 2:
        return 0.0
    m = azmean(a)
    return float(np.std([azdiff(x, m) for x in a], ddof=1))


def regression(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    ok = np.isfinite(x) & np.isfinite(y)
    x, y = x[ok], y[ok]
    if len(x) < 2 or np.std(x) == 0:
        return math.nan, math.nan, math.nan
    slope, intercept = np.polyfit(x, y, 1)
    r = float(np.corrcoef(x, y)[0, 1])
    return float(slope), float(intercept), r


def residual_stats(v):
    v = np.asarray(v, float)
    v = v[np.isfinite(v)]
    if len(v) == 0:
        return dict(n=0, mean_bias=math.nan,
                    mean_abs_deviation=math.nan,
                    median_abs_deviation=math.nan, rms=math.nan)
    return dict(
        n=len(v),
        mean_bias=float(np.mean(v)),
        mean_abs_deviation=float(np.mean(np.abs(v))),
        median_abs_deviation=float(np.median(np.abs(v))),
        rms=float(np.sqrt(np.mean(v * v))),
    )


# ---------------------------------------------------------------------------
# RINEX DISCOVERY
# ---------------------------------------------------------------------------

RINEX_RX = re.compile(
    r"^USGS00USA_R_(?P<year>\d{4})(?P<doy>\d{3})0000_01D_01S_MO\."
    r"(?:rnx|crx)(?:\.(?:gz|Z))?$", re.I
)


def discover_rinex():
    banner("SEARCHING FOR ALL LOCAL USGS RINEX FILES")
    found = {}

    for root in RINEX_ROOTS:
        if not root.exists():
            print("Searching:", root, "[not present]")
            continue

        print("Searching:", root)
        cmd = [
            "find", str(root), "-type", "f",
            "(",
            "-iname", "USGS00USA_R_*_01D_01S_MO.rnx",
            "-o", "-iname", "USGS00USA_R_*_01D_01S_MO.rnx.gz",
            "-o", "-iname", "USGS00USA_R_*_01D_01S_MO.rnx.Z",
            "-o", "-iname", "USGS00USA_R_*_01D_01S_MO.crx",
            "-o", "-iname", "USGS00USA_R_*_01D_01S_MO.crx.gz",
            "-o", "-iname", "USGS00USA_R_*_01D_01S_MO.crx.Z",
            ")", "-print"
        ]
        try:
            p = subprocess.run(
                cmd, text=True, stdout=subprocess.PIPE,
                stderr=subprocess.PIPE
            )
        except FileNotFoundError:
            print("ERROR: find command unavailable.")
            continue

        for line in p.stdout.splitlines():
            pth = Path(line.strip())
            m = RINEX_RX.match(pth.name)
            if not m:
                continue
            y, d = int(m.group("year")), int(m.group("doy"))
            if y == YEAR:
                found[(y, d)] = pth

        if p.returncode not in (0, 1):
            print("  find returned", p.returncode)
        if p.stderr.strip():
            print("  find warning:", p.stderr.strip().splitlines()[0])

    print("Unique RINEX days discovered:", len(found))
    if found:
        print("DOYs:", " ".join(f"{d:03d}" for _, d in sorted(found)))
        for (y, d), pth in sorted(found.items()):
            print(f"  {y} DOY {d:03d}: {pth}")
    else:
        print("DOYs: NONE")
    return found


# ---------------------------------------------------------------------------
# SNR CONVERSION
# ---------------------------------------------------------------------------

def snr_path(doy):
    return REFL_CODE / "2026" / "snr" / "usgs" / (
        f"usgs{doy:03d}0.26.snr{SNR_TYPE}.gz"
    )


def ensure_snr(doy):
    target = snr_path(doy)
    if target.exists() and target.stat().st_size:
        print("  SNR already exists:", target)
        return True

    banner(f"CONVERTING RINEX -> SNR: DOY {doy:03d}")
    cmd = [
        "rinex2snr", STATION, str(YEAR), str(doy),
        "-snr", str(SNR_TYPE),
        "-orb", ORBIT,
        "-stream", STREAM,
        "-samplerate", str(SAMPLE_RATE),
        "-nolook", "T",
        "-overwrite", "T",
    ]
    p = run(cmd, BASE_DIR)
    ok = target.exists() and target.stat().st_size > 0
    if ok:
        print("SUCCESS:", target)
    else:
        print("WARNING: expected SNR not found:", target)
        if p is not None:
            print("  rinex2snr return code:", p.returncode)
    return ok


# ---------------------------------------------------------------------------
# GNSS-IR PROCESSING
# ---------------------------------------------------------------------------

def process_day(doy):
    banner(f"PROCESSING GNSS-IR: DOY {doy:03d}")

    # This is the established Marconi configuration.  Try the extension
    # form first, then the JSON form used by some gnssrefl 4.x installations.
    commands = [
        [
            "gnssir", STATION, str(YEAR), str(doy),
            "-plt", "F", "-extension", GNSSIR_CONFIG
        ],
        [
            "gnssir", STATION, str(YEAR), str(doy),
            "-plt", "F", "-json", GNSSIR_CONFIG
        ],
    ]

    for cmd in commands:
        p = run(cmd, BASE_DIR)
        if p is not None and p.returncode == 0:
            return True
    return False


# ---------------------------------------------------------------------------
# GNSS-IR RESULT PARSING
# ---------------------------------------------------------------------------

def result_files():
    out = []
    for root in (REFL_CODE, BASE_DIR):
        if not root.exists():
            continue
        try:
            for p in root.rglob("*.txt"):
                n = p.name.lower()
                if "translation" in n or "summary" in n:
                    continue
                if "usgs" in n or "results" in str(p.parent).lower():
                    out.append(p)
        except OSError:
            pass
    return sorted(set(out))


def parse_result(path):
    """
    Parse a gnssrefl 4.1.x GNSS-IR result file.

    Actual gnssrefl result format:

      year doy RH sat UTCtime Azim Amp eminO emaxO NumbOf freq rise ...

    Example:

      2026 190 19.244   5 19.445 116.79  36.07 ...
      2026 190 18.830  25 20.582 157.87   7.23 ...

    Fields used here:
      column 1  = year
      column 2  = day of year
      column 3  = reflector height (m)
      column 4  = satellite / PRN
      column 5  = UTC time in decimal hours
      column 6  = azimuth (degrees)
      column 12 = rise/set (-1 / +1)

    This parser intentionally uses the documented/observed numeric
    gnssrefl format rather than attempting to infer a CSV header.
    """

    rows = []

    try:
        lines = path.read_text(errors="replace").splitlines()
    except Exception as exc:
        print(f"  WARNING: cannot read {path}: {exc}")
        return rows

    for line in lines:
        line = line.strip()

        # Ignore comments, blank lines, and header material.
        if not line or line.startswith("%") or line.startswith("#"):
            continue

        parts = line.split()

        # A valid gnssrefl observation has at least 12 fields.
        if len(parts) < 12:
            continue

        try:
            year = int(parts[0])
            doy = int(parts[1])

            rh = float(parts[2])
            sat = int(float(parts[3]))
            utc_hours = float(parts[4])
            azimuth = float(parts[5])
            rise = int(float(parts[11]))
        except (ValueError, TypeError):
            continue

        # Basic sanity checks.
        if not (2000 <= year <= 2100):
            continue

        if not (1 <= doy <= 366):
            continue

        if not (1 <= sat <= 64):
            continue

        if not math.isfinite(rh):
            continue

        if not math.isfinite(utc_hours):
            continue

        if not math.isfinite(azimuth):
            continue

        if not (0.0 <= utc_hours < 24.0):
            continue

        if not (0.0 <= azimuth <= 360.0):
            continue

        if rise not in (-1, 1):
            continue

        try:
            date0 = datetime(year, 1, 1) + timedelta(days=doy - 1)

            total_seconds = utc_hours * 3600.0
            whole_seconds = int(total_seconds)
            microseconds = int(round(
                (total_seconds - whole_seconds) * 1_000_000
            ))

            # Handle rounding exactly to the next second.
            if microseconds >= 1_000_000:
                whole_seconds += 1
                microseconds -= 1_000_000

            dt = date0 + timedelta(
                seconds=whole_seconds,
                microseconds=microseconds
            )
        except (ValueError, OverflowError):
            continue

        rows.append(dict(
            datetime_utc=dt,
            sat=sat,
            rise=rise,
            azimuth=azimuth,
            GNSS_WL_m=rh,
            source=str(path),
        ))

    return rows


def load_results():
    banner("LOADING ALL GNSS-IR RESULTS")
    files = result_files()
    print("Candidate result text files:", len(files))
    rows = []
    for p in files:
        r = parse_result(p)
        if r:
            print(f"  {len(r):6d} observations: {p}")
            rows.extend(r)

    unique = {}
    for r in rows:
        key = (
            r["datetime_utc"], r["sat"], r["rise"],
            round(r["azimuth"], 5), round(r["GNSS_WL_m"], 6)
        )
        unique[key] = r
    rows = sorted(unique.values(), key=lambda x: x["datetime_utc"])
    print("Unique GNSS-R observations loaded:", len(rows))
    return rows


# ---------------------------------------------------------------------------
# TIDE
# ---------------------------------------------------------------------------

def load_tide():
    if not TIDE_FILE.exists():
        raise FileNotFoundError(TIDE_FILE)

    wb = load_workbook(TIDE_FILE, data_only=True)
    ws = None
    header_row = None
    headers = None

    for sh in wb.worksheets:
        for ri in range(1, min(sh.max_row, 100) + 1):
            vals = [
                sh.cell(ri, c).value
                for c in range(1, min(sh.max_column, 100) + 1)
            ]
            h = [str(v).strip().lower() if v is not None else "" for v in vals]
            if PRIMARY_TIDE_MODEL.lower() in h:
                ws, header_row, headers = sh, ri, h
                break
        if ws:
            break

    if ws is None:
        raise RuntimeError(
            f"{PRIMARY_TIDE_MODEL} not found in {TIDE_FILE}"
        )

    dt_col = None
    for i, h in enumerate(headers):
        if h in (
            "datetime", "date_time", "time", "date",
            "timestamp", "datetime_utc", "utc", "date/time"
        ):
            dt_col = i + 1
            break
    if dt_col is None:
        for i, h in enumerate(headers):
            if "time" in h or "date" in h:
                dt_col = i + 1
                break

    tide_col = headers.index(PRIMARY_TIDE_MODEL.lower()) + 1

    t, v = [], []
    for ri in range(header_row + 1, ws.max_row + 1):
        dt = normalize_dt(ws.cell(ri, dt_col).value)
        z = fnum(ws.cell(ri, tide_col).value)
        if dt is not None and math.isfinite(z):
            t.append(dt)
            v.append(z)

    if len(t) < 2:
        raise RuntimeError("Insufficient tide data.")

    d = {tt: zz for tt, zz in zip(t, v)}
    t = sorted(d)
    v = np.asarray([d[x] for x in t], float)
    t0 = t[0]
    x = np.asarray([(q - t0).total_seconds() for q in t])

    def tide_at(dt):
        if dt < t[0] or dt > t[-1]:
            return math.nan
        return float(np.interp((dt - t0).total_seconds(), x, v))

    print("Tide records:", len(t))
    print("Tide interval:", t[0], "through", t[-1])
    return tide_at


def add_tide(rows, tide_at):
    for r in rows:
        tide = tide_at(r["datetime_utc"])
        r["PRIMARY_TIDE_m"] = tide
        r["GNSS_PLUS_0242_m"] = r["GNSS_WL_m"] + DATUM_OFFSET_M
        r["RAW_RESIDUAL_m"] = r["GNSS_WL_m"] - tide
        r["PLUS_0242_RESIDUAL_m"] = r["GNSS_PLUS_0242_m"] - tide
    return rows


# ---------------------------------------------------------------------------
# TRACK CLUSTERING / SCREEN
# ---------------------------------------------------------------------------

def cluster_tracks(rows):
    base = defaultdict(list)
    for r in rows:
        if all(math.isfinite(r[k]) for k in (
            "azimuth", "GNSS_WL_m", "PRIMARY_TIDE_m"
        )):
            base[(r["sat"], r["rise"])].append(r)

    tracks = []
    for (sat, rise), obs in base.items():
        obs.sort(key=lambda r: r["azimuth"])
        clusters, cur = [], []
        for r in obs:
            if not cur:
                cur = [r]
                continue
            if azdiff(r["azimuth"], azmean([x["azimuth"] for x in cur])) <= AZ_CLUSTER_TOL_DEG:
                cur.append(r)
            else:
                clusters.append(cur)
                cur = [r]
        if cur:
            clusters.append(cur)

        for c in clusters:
            wl = np.asarray([r["GNSS_WL_m"] for r in c])
            tide = np.asarray([r["PRIMARY_TIDE_m"] for r in c])
            slope, intercept, rr = regression(tide, wl)
            if len(wl) < 2:
                continue
            fit = slope * tide + intercept if math.isfinite(slope) else np.full(len(wl), np.nan)
            unit_rms = float(np.sqrt(np.mean((wl - fit) ** 2))) if np.all(np.isfinite(fit)) else math.nan
            raw_rms = float(np.sqrt(np.mean((wl - tide) ** 2)))
            plus_rms = float(np.sqrt(np.mean((wl + DATUM_OFFSET_M - tide) ** 2)))
            tracks.append(dict(
                sat=sat, rise=rise, N=len(c),
                days=len({r["datetime_utc"].date() for r in c}),
                azimuth=azmean([r["azimuth"] for r in c]),
                az_sd=azsd([r["azimuth"] for r in c]),
                tide_r=rr, slope=slope, intercept_m=intercept,
                unit_rms_m=unit_rms, raw_rms_m=raw_rms,
                plus_rms_m=plus_rms,
                improvement_m=raw_rms - plus_rms, rows=c
            ))
    return tracks


def good(t):
    return (
        t["N"] >= MIN_OBS and
        t["days"] >= MIN_DAYS and
        math.isfinite(t["tide_r"]) and t["tide_r"] >= MIN_TIDE_R and
        math.isfinite(t["slope"]) and SLOPE_MIN <= t["slope"] <= SLOPE_MAX and
        math.isfinite(t["unit_rms_m"]) and t["unit_rms_m"] <= MAX_UNIT_RMS_M and
        math.isfinite(t["az_sd"]) and t["az_sd"] <= MAX_AZ_SD_DEG
    )


# ---------------------------------------------------------------------------
# POPULATION STATISTICS
# ---------------------------------------------------------------------------

def population_stats(tracks):
    raw, plus = [], []
    for t in tracks:
        for r in t["rows"]:
            if math.isfinite(r["RAW_RESIDUAL_m"]):
                raw.append(r["RAW_RESIDUAL_m"])
            if math.isfinite(r["PLUS_0242_RESIDUAL_m"]):
                plus.append(r["PLUS_0242_RESIDUAL_m"])
    return residual_stats(raw), residual_stats(plus)


# ---------------------------------------------------------------------------
# OUTPUTS
# ---------------------------------------------------------------------------

def write_csv(tracks):
    ranked = sorted(tracks, key=lambda t: (-t["tide_r"], t["raw_rms_m"]))
    with OUT_CSV.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "rank", "sat", "rise", "N", "days", "azimuth", "az_sd",
            "tide_r", "slope", "intercept_m", "unit_rms_cm",
            "raw_rms_cm", "plus_0242_rms_cm", "improvement_cm"
        ])
        for i, t in enumerate(ranked, 1):
            w.writerow([
                i, t["sat"], t["rise"], t["N"], t["days"],
                t["azimuth"], t["az_sd"], t["tide_r"], t["slope"],
                t["intercept_m"], t["unit_rms_m"] * 100,
                t["raw_rms_m"] * 100, t["plus_rms_m"] * 100,
                t["improvement_m"] * 100
            ])


def write_summary(rinex, snr_days, processed, rows, tracks, good_tracks, raw, plus):
    with OUT_SUMMARY.open("w") as f:
        f.write("MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V2\n")
        f.write("=" * 92 + "\n\n")
        f.write(f"Base directory: {BASE_DIR}\n")
        f.write(f"REFL_CODE: {REFL_CODE}\n")
        f.write(f"Station: {STATION}\n")
        f.write(f"Year: {YEAR}\n")
        f.write(f"Primary tide: {PRIMARY_TIDE_MODEL}\n")
        f.write(f"GNSSIR config: {GNSSIR_CONFIG}\n")
        f.write(f"Diagnostic offset: +{DATUM_OFFSET_M:.3f} m\n\n")

        f.write("GOOD-track criteria:\n")
        f.write(f"  observations >= {MIN_OBS}\n")
        f.write(f"  unique days   >= {MIN_DAYS}\n")
        f.write(f"  tide r        >= {MIN_TIDE_R}\n")
        f.write(f"  slope         = {SLOPE_MIN} to {SLOPE_MAX}\n")
        f.write(f"  unit RMS      <= {MAX_UNIT_RMS_M*100:.1f} cm\n")
        f.write(f"  azimuth SD    <= {MAX_AZ_SD_DEG:.1f} deg\n")
        f.write("  +24.2 cm is diagnostic only.\n\n")

        f.write("PROCESSING COUNTS\n")
        f.write(f"  RINEX days discovered: {len(rinex)}\n")
        f.write(f"  SNR days available/generated: {len(snr_days)}\n")
        f.write(f"  GNSS-IR days successfully processed: {len(processed)}\n")
        f.write(f"  GNSS-R observations loaded: {len(rows)}\n")
        f.write(f"  clustered tracks: {len(tracks)}\n")
        f.write(f"  good tracks: {len(good_tracks)}\n\n")

        f.write("ENTIRE GOOD-TRACK POPULATION\n")
        f.write("-" * 92 + "\n")
        f.write(f"{'Metric':32s}{'RAW':>14s}{'+24.2 cm':>14s}\n")
        f.write(f"{'N observations':32s}{raw['n']:14d}{plus['n']:14d}\n")
        for label, key in (
            ("Mean bias", "mean_bias"),
            ("Mean absolute deviation", "mean_abs_deviation"),
            ("Median absolute deviation", "median_abs_deviation"),
            ("RMS", "rms"),
        ):
            f.write(
                f"{label:32s}{raw[key]*100:14.3f}{plus[key]*100:14.3f} cm\n"
            )

        f.write("\nGOOD TRACKS\n")
        for i, t in enumerate(
            sorted(good_tracks, key=lambda t: (-t["tide_r"], t["raw_rms_m"])), 1
        ):
            f.write(
                f"{i:2d} SAT={t['sat']:2d} rise={t['rise']:2d} "
                f"N={t['N']:3d} days={t['days']:2d} "
                f"Az={t['azimuth']:7.2f}±{t['az_sd']:.2f} "
                f"r={t['tide_r']:+.4f} slope={t['slope']:+.4f} "
                f"rawRMS={t['raw_rms_m']*100:.2f}cm "
                f"+24.2RMS={t['plus_rms_m']*100:.2f}cm "
                f"improvement={t['improvement_m']*100:+.2f}cm\n"
            )


def make_plots(rows, good_tracks):
    PLOT_DIR.mkdir(parents=True, exist_ok=True)

    selected = {id(r) for t in good_tracks for r in t["rows"]}
    g = [
        r for r in rows if id(r) in selected and
        math.isfinite(r["PRIMARY_TIDE_m"]) and
        math.isfinite(r["GNSS_WL_m"])
    ]
    if not g:
        print("No good-track observations available for plotting.")
        return

    g.sort(key=lambda r: r["datetime_utc"])
    times = [r["datetime_utc"] for r in g]
    tide = np.asarray([r["PRIMARY_TIDE_m"] for r in g])
    gnss = np.asarray([r["GNSS_WL_m"] for r in g])
    plus = gnss + DATUM_OFFSET_M

    # 01: good tracks only, raw + offset + tide
    fig, ax = plt.subplots(figsize=(15, 8))
    ax.plot(times, tide, linewidth=2, label=PRIMARY_TIDE_MODEL)
    ax.scatter(times, gnss, s=12, alpha=.55, label="GNSS-R good tracks")
    ax.plot(times, plus, linewidth=1.2, alpha=.7, label="GNSS-R + 0.242 m")
    ax.set_title("Marconi — Good GNSS-R Tracks vs EOT20 Tide")
    ax.set_ylabel("Height (m)")
    ax.grid(True, alpha=.25)
    ax.legend()
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "01_good_tracks_vs_tide.png", dpi=180)
    plt.close(fig)

    raw_cm = (gnss - tide) * 100
    plus_cm = (gnss + DATUM_OFFSET_M - tide) * 100

    # 02: raw residual
    fig, ax = plt.subplots(figsize=(15, 7))
    ax.scatter(times, raw_cm, s=12, alpha=.55, label="Raw GNSS-R − EOT20")
    ax.axhline(0, linewidth=1.5)
    ax.set_title("Marconi — Good-Track Raw Residuals")
    ax.set_ylabel("Residual (cm)")
    ax.grid(True, alpha=.25)
    ax.legend()
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "02_good_track_residuals_raw.png", dpi=180)
    plt.close(fig)

    # 03: +24.2 cm
    fig, ax = plt.subplots(figsize=(15, 7))
    ax.scatter(times, plus_cm, s=12, alpha=.55, label="GNSS-R + 0.242 m − EOT20")
    ax.axhline(0, linewidth=1.5)
    ax.set_title("Marconi — Good-Track Residuals After +24.2 cm Diagnostic Offset")
    ax.set_ylabel("Residual (cm)")
    ax.grid(True, alpha=.25)
    ax.legend()
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "03_good_tracks_vs_tide_plus_0242m.png", dpi=180)
    plt.close(fig)

    # 04: residual distributions
    fig, ax = plt.subplots(figsize=(11, 7))
    ax.hist(raw_cm, bins=35, alpha=.55, label="Raw")
    ax.hist(plus_cm, bins=35, alpha=.55, label="+24.2 cm")
    ax.axvline(0, linewidth=1.5)
    ax.set_title("Marconi — Good-Track Residual Distribution")
    ax.set_xlabel("GNSS-R − EOT20 residual (cm)")
    ax.set_ylabel("Observations")
    ax.grid(True, alpha=.2)
    ax.legend()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "04_good_track_residual_histogram.png", dpi=180)
    plt.close(fig)

    # 05: track RMS
    ranked = sorted(good_tracks, key=lambda t: t["raw_rms_m"])
    labels = [
        f"S{t['sat']} {'R' if t['rise'] > 0 else 'F'}\n{t['azimuth']:.1f}°"
        for t in ranked
    ]
    x = np.arange(len(ranked))
    width = .38
    fig, ax = plt.subplots(figsize=(13, 8))
    ax.bar(x-width/2, [t["raw_rms_m"]*100 for t in ranked], width, label="Raw")
    ax.bar(x+width/2, [t["plus_rms_m"]*100 for t in ranked], width, label="+24.2 cm")
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=60, ha="right")
    ax.set_ylabel("RMS residual (cm)")
    ax.set_title("Marconi — Good Tracks: Raw vs +24.2 cm RMS")
    ax.grid(True, axis="y", alpha=.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "05_good_track_rms_comparison.png", dpi=180)
    plt.close(fig)

    # 06: scatter
    fig, ax = plt.subplots(figsize=(9, 9))
    ax.scatter(tide, gnss, s=14, alpha=.5, label="Good GNSS-R")
    lo = min(tide.min(), gnss.min())
    hi = max(tide.max(), gnss.max())
    ax.plot([lo, hi], [lo, hi], linewidth=1.5, label="1:1")
    ax.set_xlabel("EOT20 tide (m)")
    ax.set_ylabel("GNSS-R reflector height (m)")
    ax.set_title("Marconi — Good GNSS-R vs EOT20")
    ax.grid(True, alpha=.25)
    ax.legend()
    fig.tight_layout()
    fig.savefig(PLOT_DIR / "06_good_gnssr_vs_tide_scatter.png", dpi=180)
    plt.close(fig)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    banner("MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V2")
    print("Base directory:", BASE_DIR)
    print("REFL_CODE     :", REFL_CODE)
    print("RINEX search  :", ", ".join(map(str, RINEX_ROOTS)))
    print("Primary tide  :", PRIMARY_TIDE_MODEL)
    print("GNSSIR config :", GNSSIR_CONFIG)
    print(f"Diagnostic offset: +{DATUM_OFFSET_M:.3f} m")
    print()
    print("GOOD-track criteria:")
    print(f"  observations >= {MIN_OBS}")
    print(f"  unique days   >= {MIN_DAYS}")
    print(f"  tide r        >= {MIN_TIDE_R}")
    print(f"  slope         = {SLOPE_MIN} to {SLOPE_MAX}")
    print(f"  unit RMS      <= {MAX_UNIT_RMS_M*100:.1f} cm")
    print(f"  azimuth SD    <= {MAX_AZ_SD_DEG:.1f} deg")
    print("  +24.2 cm is NOT used to select good tracks.")

    rinex = discover_rinex()
    if not rinex:
        print("No local USGS RINEX files were discovered.")
        print("Add the actual RINEX mount/path to RINEX_ROOTS if necessary.")
        return 2

    banner("ENSURING SNR FILES EXIST FOR ALL RINEX DAYS")
    snr_days = [d for (_, d) in sorted(rinex) if ensure_snr(d)]
    print("SNR days available/generated:", len(snr_days))

    banner("PROCESSING ALL AVAILABLE GNSS-IR DAYS")
    processed = []
    for d in snr_days:
        if process_day(d):
            processed.append(d)
    print("GNSS-IR days successfully processed:", len(processed))

    rows = load_results()
    if not rows:
        print("No GNSS-IR observations were loaded.")
        return 3

    banner("LOADING TIDE MODEL")
    tide_at = load_tide()
    rows = add_tide(rows, tide_at)
    rows = [
        r for r in rows
        if math.isfinite(r["PRIMARY_TIDE_m"]) and
        math.isfinite(r["GNSS_WL_m"])
    ]
    print("Observations with valid tide:", len(rows))
    if not rows:
        return 4

    tracks = cluster_tracks(rows)
    good_tracks = [t for t in tracks if good(t)]
    good_tracks.sort(key=lambda t: (-t["tide_r"], t["raw_rms_m"]))

    banner("GOOD TRACK SCREEN")
    print("Total clustered tracks:", len(tracks))
    print("GOOD tracks:            ", len(good_tracks))
    for i, t in enumerate(good_tracks, 1):
        print(
            f"{i:2d} SAT={t['sat']:2d} rise={t['rise']:2d} "
            f"N={t['N']:3d} days={t['days']:2d} "
            f"Az={t['azimuth']:7.2f}±{t['az_sd']:.2f} "
            f"r={t['tide_r']:+.4f} slope={t['slope']:+.4f} "
            f"rawRMS={t['raw_rms_m']*100:.2f}cm "
            f"+24.2RMS={t['plus_rms_m']*100:.2f}cm "
            f"improvement={t['improvement_m']*100:+.2f}cm"
        )

    if not good_tracks:
        print("No tracks met the good-data criteria.")
        return 5

    raw, plus = population_stats(good_tracks)

    banner("ENTIRE GOOD-TRACK POPULATION STATISTICS")
    print("Residual = GNSS-R - EOT20")
    print()
    print(f"{'Metric':32s}{'RAW':>14s}{'+24.2 cm':>14s}")
    print("-" * 60)
    print(f"{'N observations':32s}{raw['n']:14d}{plus['n']:14d}")
    for label, key in (
        ("Mean bias", "mean_bias"),
        ("Mean absolute deviation", "mean_abs_deviation"),
        ("Median absolute deviation", "median_abs_deviation"),
        ("RMS", "rms"),
    ):
        print(
            f"{label:32s}{raw[key]*100:14.3f}{plus[key]*100:14.3f} cm"
        )

    banner("GENERATING PLOTS")
    make_plots(rows, good_tracks)
    write_csv(good_tracks)
    write_summary(
        rinex, snr_days, processed, rows, tracks, good_tracks, raw, plus
    )

    banner("OUTPUTS")
    print("CSV:     ", OUT_CSV)
    print("Summary: ", OUT_SUMMARY)
    print("Plots:   ", PLOT_DIR)
    print()
    print("Population statistics use ALL observations in ALL screened good tracks.")
    print("The +24.2 cm offset is diagnostic only.")
    print("DONE")
    return 0


if __name__ == "__main__":
    sys.exit(main())
