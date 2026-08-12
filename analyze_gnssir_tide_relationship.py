#!/usr/bin/env python3
"""
analyze_gnssir_tide_relationship.py

USGS GNSS Reference Station
GNSS-IR vs. Tide Model Scientific Diagnostic

Purpose
-------
Investigate why GNSS-IR reflector-height-derived water levels do not
match the verified tide models at the site.

This script does NOT modify the production GNSS-IR processing.

It performs four independent tests:

1. Reproduce the established within-arc tidal-change calculation
   using the exact 811-arc methodology already validated in
   diagnose_arc_timing_bias.py.

2. Compare GNSS-IR-derived relative water level against the ensemble
   tide model using a physically constrained unit slope and a freely
   estimated constant vertical offset.

3. Search for a systematic timing offset between GNSS-IR and the tide
   model from -300 to +300 minutes (widened from an original +/-120
   minute range -- that range could never have detected the specific
   ~4-hour EDT/UTC offset hypothesis this test exists to check; if a
   real offset that large were present, the true optimum would have
   sat right at or beyond the old boundary, invisible to the search).

4. Test whether GNSS-IR error increases with the amount of real tidal
   movement occurring during the same satellite arc.

Important
---------
The vertical relationship is:

    GNSS-derived water level = H_reference - RH

where H_reference is the antenna/reference height.

Only a CONSTANT vertical offset is fitted.

We deliberately do NOT fit an arbitrary scale factor to the tide
model because doing so can conceal a physically incorrect relationship.

Timestamps
----------
GNSS-IR solution time comes from:

    result["MJD"]

This was verified against result["UTCtime"] for the installed
gnssrefl version.

Arc start/end times come from:

    meta["time_start"]
    meta["time_end"]

QC
--
An arc is considered QC-passing exactly as in the established
diagnose_arc_timing_bias.py:

    meta["gnssir_processing_results"] is not None

No additional RH filter is imposed.

Production arc settings
-----------------------
    e1=5.0
    e2=15.0
    azlist=[100.0, 130.0, 150.0, 215.0]
    polyV=2
    pele=[5, 30]
    attach_results=True
    buffer_hours=2

Usage
-----

    source gnssrefl_venv/bin/activate

    python3 analyze_gnssir_tide_relationship.py \
        usgs 2026 204 207 marconi_tides_sherwood.xlsx

Optional reference height:

    python3 analyze_gnssir_tide_relationship.py \
        usgs 2026 204 207 marconi_tides_sherwood.xlsx 18.625

Outputs
-------
    gnssir_tide_arc_analysis.csv
    gnssir_tide_lag_analysis.csv
    gnssir_tide_analysis_summary.txt
    gnssir_tide_relationship.png
    gnssir_tide_lag_search.png
    gnssir_tide_smearing.png
"""

from __future__ import annotations

import csv
import math
import sys
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import openpyxl


# ---------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------

MJD_EPOCH = datetime(1858, 11, 17)

DEFAULT_REFERENCE_HEIGHT_M = 18.625

# Widened from the original +/-120 minutes: that range could never
# have detected the specific ~4-hour EDT/UTC offset hypothesis this
# test exists to check in the first place. +/-300 minutes (+/-5
# hours) gives real margin beyond a 4-hour offset in either direction,
# so a result pinned at the new boundary would itself be a real,
# actionable signal rather than an artifact of a too-narrow search.
LAG_MINUTES = np.arange(
    -300.0,
    300.0 + 5.0,
    5.0,
)


# ---------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------

def finite(value):
    """Return a finite float, otherwise None."""

    try:
        value = float(value)
    except (TypeError, ValueError):
        return None

    if math.isfinite(value):
        return value

    return None


def mjd_to_datetime(mjd):
    """Convert Modified Julian Date to naive UTC datetime."""

    return MJD_EPOCH + timedelta(
        days=float(mjd)
    )


def seconds_of_day_to_datetime(
    day_start,
    seconds,
):
    """
    Convert seconds-of-day from gnssrefl arc metadata into datetime.

    The value may be outside 0-86400 because extract_arcs can use
    buffered data around midnight.
    """

    return day_start + timedelta(
        seconds=float(seconds)
    )


def pearson(x, y):
    """Pearson correlation."""

    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    mask = np.isfinite(x) & np.isfinite(y)

    if np.count_nonzero(mask) < 3:
        return float("nan")

    x = x[mask]
    y = y[mask]

    if np.std(x) == 0 or np.std(y) == 0:
        return float("nan")

    return float(
        np.corrcoef(x, y)[0, 1]
    )


def rankdata(values):
    """Simple average-rank implementation."""

    values = np.asarray(
        values,
        dtype=float,
    )

    order = np.argsort(
        values,
        kind="mergesort",
    )

    ranks = np.empty(
        len(values),
        dtype=float,
    )

    i = 0

    while i < len(values):

        j = i + 1

        while (
            j < len(values)
            and values[order[j]]
            == values[order[i]]
        ):
            j += 1

        ranks[
            order[i:j]
        ] = (
            (i + 1) + j
        ) / 2.0

        i = j

    return ranks


def spearman(x, y):
    """Spearman rank correlation."""

    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    mask = (
        np.isfinite(x)
        & np.isfinite(y)
    )

    if np.count_nonzero(mask) < 3:
        return float("nan")

    return pearson(
        rankdata(x[mask]),
        rankdata(y[mask]),
    )


def fixed_slope_offset(
    gnss,
    tide,
):
    """
    Fit:

        gnss = tide + offset

    Only the constant offset is estimated.
    """

    gnss = np.asarray(
        gnss,
        dtype=float,
    )

    tide = np.asarray(
        tide,
        dtype=float,
    )

    mask = (
        np.isfinite(gnss)
        & np.isfinite(tide)
    )

    if np.count_nonzero(mask) == 0:
        raise ValueError(
            "No valid points for offset fit."
        )

    offset = np.mean(
        gnss[mask]
        - tide[mask]
    )

    return float(offset)


def rms(values):
    values = np.asarray(
        values,
        dtype=float,
    )

    return float(
        np.sqrt(
            np.mean(
                values ** 2
            )
        )
    )


# ---------------------------------------------------------------------
# Tide model handling
# ---------------------------------------------------------------------

def read_tide_models(path):
    """
    Read the tide-model workbook.

    Any column ending in '_heightm' is treated as a tide model.
    """

    print(
        f"Reading tide models from {path}..."
    )

    wb = openpyxl.load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    ws = wb[
        wb.sheetnames[0]
    ]

    rows = list(
        ws.iter_rows(
            values_only=True
        )
    )

    if not rows:
        raise RuntimeError(
            "Tide workbook is empty."
        )

    header = rows[0]

    model_columns = {}

    for index, name in enumerate(
        header
    ):

        if name is None:
            continue

        name = str(name)

        if name.endswith(
            "_heightm"
        ):

            model_name = name[
                :-len("_heightm")
            ]

            model_columns[
                model_name
            ] = index

    if not model_columns:
        raise RuntimeError(
            "No *_heightm tide-model "
            "columns found."
        )

    times = []

    values = {
        name: []
        for name in model_columns
    }

    for row in rows[1:]:

        if (
            not row
            or row[0] is None
        ):
            continue

        if not isinstance(
            row[0],
            datetime,
        ):
            continue

        row_valid = True

        row_values = {}

        for (
            model_name,
            column,
        ) in model_columns.items():

            value = finite(
                row[column]
            )

            if value is None:

                row_valid = False
                break

            row_values[
                model_name
            ] = value

        if not row_valid:
            continue

        times.append(
            row[0]
        )

        for model_name in model_columns:

            values[
                model_name
            ].append(
                row_values[
                    model_name
                ]
            )

    if len(times) < 2:
        raise RuntimeError(
            "Insufficient tide-model data."
        )

    print(
        f"Parsed {len(times)} tide-model points."
    )

    print(
        "Models:",
        list(values.keys()),
    )

    print(
        f"Model time range: "
        f"{times[0]} through {times[-1]}"
    )

    return times, values


def build_ensemble_interpolator(
    times,
    model_values,
):
    """
    Build a linear interpolator for the mean of all tide models.
    """

    matrix = np.array(
        [
            model_values[name]
            for name in model_values
        ],
        dtype=float,
    )

    ensemble = np.mean(
        matrix,
        axis=0,
    )

    epoch_seconds = np.array(
        [
            (
                t - times[0]
            ).total_seconds()
            for t in times
        ],
        dtype=float,
    )

    def interpolate(query_dt):

        query_seconds = (
            query_dt - times[0]
        ).total_seconds()

        if (
            query_seconds
            < epoch_seconds[0]
            or query_seconds
            > epoch_seconds[-1]
        ):
            return float("nan")

        return float(
            np.interp(
                query_seconds,
                epoch_seconds,
                ensemble,
            )
        )

    return interpolate


# ---------------------------------------------------------------------
# Arc extraction
# ---------------------------------------------------------------------

def extract_all_arcs(
    station,
    year,
    doy1,
    doy2,
):
    """
    Extract arcs using EXACTLY the established production settings.
    """

    try:

        from gnssrefl.extract_arcs import (
            extract_arcs_from_station
        )

    except ImportError:

        print(
            "ERROR: could not import "
            "gnssrefl.extract_arcs."
        )

        print(
            "Activate gnssrefl_venv first."
        )

        sys.exit(1)

    all_records = []

    total_extracted = 0
    total_qc = 0

    for doy in range(
        doy1,
        doy2 + 1,
    ):

        print()
        print(
            f"Extracting arcs for "
            f"{station} {year} DOY {doy}..."
        )

        arcs = extract_arcs_from_station(
            station,
            year,
            doy,
            e1=5.0,
            e2=15.0,
            azlist=[
                100.0,
                130.0,
                150.0,
                215.0,
            ],
            polyV=2,
            pele=[
                5,
                30,
            ],
            attach_results=True,
            buffer_hours=2,
        )

        total_extracted += len(
            arcs
        )

        matched = [
            (
                meta,
                data,
            )
            for meta, data in arcs
            if meta[
                "gnssir_processing_results"
            ] is not None
        ]

        total_qc += len(
            matched
        )

        print(
            f"  {len(arcs)} arcs extracted, "
            f"{len(matched)} passed QC"
        )

        day_start = (
            datetime(
                year,
                1,
                1,
            )
            + timedelta(
                days=doy - 1
            )
        )

        for meta, data in matched:

            result = meta[
                "gnssir_processing_results"
            ]

            rh = finite(
                result.get("RH")
            )

            mjd = finite(
                result.get("MJD")
            )

            utc_hours = finite(
                result.get("UTCtime")
            )

            if rh is None:
                print(
                    "WARNING: QC-passing arc "
                    "has invalid RH; skipping."
                )
                continue

            if mjd is None:
                print(
                    "WARNING: QC-passing arc "
                    "has no MJD; skipping."
                )
                continue

            arc_start = (
                seconds_of_day_to_datetime(
                    day_start,
                    meta["time_start"],
                )
            )

            arc_end = (
                seconds_of_day_to_datetime(
                    day_start,
                    meta["time_end"],
                )
            )

            solution_time = (
                mjd_to_datetime(mjd)
            )

            record = {
                "doy": doy,

                "sat": meta.get(
                    "sat"
                ),

                "freq": meta.get(
                    "freq"
                ),

                "arc_num": meta.get(
                    "arc_num"
                ),

                "arc_type": meta.get(
                    "arc_type"
                ),

                "delT_min": finite(
                    meta.get("delT")
                ),

                "time_start_utc":
                    arc_start.isoformat(),

                "solution_time_utc":
                    solution_time.isoformat(),

                "time_end_utc":
                    arc_end.isoformat(),

                "solution_mjd": mjd,

                "solution_UTCtime_hours":
                    utc_hours,

                "RH_m": rh,

                "Amp": finite(
                    result.get("Amp")
                ),

                "PkNoise": finite(
                    result.get("PkNoise")
                ),

                "Azim": finite(
                    result.get("Azim")
                ),

                "eminO": finite(
                    result.get("eminO")
                ),

                "emaxO": finite(
                    result.get("emaxO")
                ),

                "NumbOf": finite(
                    result.get("NumbOf")
                ),

                "rise": result.get(
                    "rise"
                ),

                "refr": result.get(
                    "refr"
                ),
            }

            all_records.append(
                record
            )

    return (
        all_records,
        total_extracted,
        total_qc,
    )


# ---------------------------------------------------------------------
# Main analysis
# ---------------------------------------------------------------------

def main():

    if len(sys.argv) not in (6, 7):

        print("Usage:")
        print(
            f"  {sys.argv[0]} "
            "station year doy1 doy2 "
            "tide_models.xlsx "
            "[reference_height_m]"
        )

        sys.exit(1)

    station = sys.argv[1]

    year = int(
        sys.argv[2]
    )

    doy1 = int(
        sys.argv[3]
    )

    doy2 = int(
        sys.argv[4]
    )

    tide_path = Path(
        sys.argv[5]
    )

    if len(sys.argv) == 7:

        reference_height = float(
            sys.argv[6]
        )

    else:

        reference_height = (
            DEFAULT_REFERENCE_HEIGHT_M
        )

    if not tide_path.exists():

        print(
            "ERROR: tide model file not found:"
        )

        print(
            f"  {tide_path}"
        )

        sys.exit(1)

    print()
    print("=" * 72)
    print(
        "USGS GNSS-IR / TIDE MODEL RELATIONSHIP ANALYSIS"
    )
    print("=" * 72)

    print(
        f"Station             : {station}"
    )

    print(
        f"Year                : {year}"
    )

    print(
        f"DOY                 : {doy1}-{doy2}"
    )

    print(
        f"Reference height    : "
        f"{reference_height:.3f} m"
    )

    print(
        f"Tide model file     : {tide_path}"
    )

    print()

    print(
        "Production arc settings:"
    )

    print(
        "  e1=5.0, e2=15.0"
    )

    print(
        "  azlist=[100,130,150,215]"
    )

    print(
        "  polyV=2, pele=[5,30]"
    )

    print(
        "  attach_results=True"
    )

    print(
        "  buffer_hours=2"
    )

    print()

    # ---------------------------------------------------------------
    # Tide models
    # ---------------------------------------------------------------

    (
        tide_times,
        tide_values,
    ) = read_tide_models(
        tide_path
    )

    tide_at = (
        build_ensemble_interpolator(
            tide_times,
            tide_values,
        )
    )

    # ---------------------------------------------------------------
    # Extract arcs
    # ---------------------------------------------------------------

    (
        records,
        total_extracted,
        total_qc,
    ) = extract_all_arcs(
        station,
        year,
        doy1,
        doy2,
    )

    print()
    print("=" * 72)

    print(
        f"Total arcs extracted : "
        f"{total_extracted}"
    )

    print(
        f"Total QC-passing     : "
        f"{total_qc}"
    )

    print(
        f"Usable for analysis  : "
        f"{len(records)}"
    )

    print("=" * 72)

    if not records:

        print(
            "ERROR: no usable arcs."
        )

        sys.exit(1)

    # ---------------------------------------------------------------
    # Calculate tide values for each arc
    # ---------------------------------------------------------------

    outside_model = 0
    timing_checks = []

    valid_records = []

    for record in records:

        start = datetime.fromisoformat(
            record[
                "time_start_utc"
            ]
        )

        solution = datetime.fromisoformat(
            record[
                "solution_time_utc"
            ]
        )

        end = datetime.fromisoformat(
            record[
                "time_end_utc"
            ]
        )

        tide_start = tide_at(
            start
        )

        tide_solution = tide_at(
            solution
        )

        tide_end = tide_at(
            end
        )

        if not all(
            np.isfinite(
                [
                    tide_start,
                    tide_solution,
                    tide_end,
                ]
            )
        ):

            outside_model += 1

            continue

        # Physical GNSS-IR water-level proxy.
        #
        # Higher RH means a lower reflecting surface, so:
        #
        #     water_level = H_reference - RH
        #
        gnss_water = (
            reference_height
            - record["RH_m"]
        )

        within_arc_change_m = (
            tide_end
            - tide_start
        )

        within_arc_change_cm = (
            within_arc_change_m
            * 100.0
        )

        # Difference between solution epoch and arc midpoint.
        midpoint = (
            start
            + (
                end - start
            ) / 2
        )

        solution_offset_sec = (
            solution - midpoint
        ).total_seconds()

        # Fraction through the arc.
        duration_seconds = (
            end - start
        ).total_seconds()

        if duration_seconds > 0:

            solution_fraction = (
                solution - start
            ).total_seconds() / (
                duration_seconds
            )

        else:

            solution_fraction = float(
                "nan"
            )

        record[
            "tide_start_m"
        ] = tide_start

        record[
            "tide_solution_m"
        ] = tide_solution

        record[
            "tide_end_m"
        ] = tide_end

        record[
            "within_arc_change_m"
        ] = within_arc_change_m

        record[
            "within_arc_change_cm"
        ] = within_arc_change_cm

        record[
            "abs_within_arc_change_cm"
        ] = abs(
            within_arc_change_cm
        )

        record[
            "gnss_water_level_m"
        ] = gnss_water

        record[
            "solution_minus_arc_midpoint_sec"
        ] = solution_offset_sec

        record[
            "solution_fraction_of_arc"
        ] = solution_fraction

        valid_records.append(
            record
        )

        timing_checks.append(
            solution_offset_sec
        )

    records = valid_records

    print()
    print(
        f"Arcs outside tide-model coverage: "
        f"{outside_model}"
    )

    print(
        f"Final paired arcs: {len(records)}"
    )

    if len(records) < 10:

        print(
            "ERROR: insufficient paired arcs."
        )

        sys.exit(1)

    # ---------------------------------------------------------------
    # Arrays
    # ---------------------------------------------------------------

    gnss_water = np.array(
        [
            r[
                "gnss_water_level_m"
            ]
            for r in records
        ],
        dtype=float,
    )

    tide_solution = np.array(
        [
            r[
                "tide_solution_m"
            ]
            for r in records
        ],
        dtype=float,
    )

    tide_change = np.array(
        [
            r[
                "within_arc_change_cm"
            ]
            for r in records
        ],
        dtype=float,
    )

    abs_tide_change = np.abs(
        tide_change
    )

    durations = np.array(
        [
            r["delT_min"]
            for r in records
        ],
        dtype=float,
    )

    # ---------------------------------------------------------------
    # Zero-lag physically constrained comparison
    # ---------------------------------------------------------------

    zero_offset = fixed_slope_offset(
        gnss_water,
        tide_solution,
    )

    zero_residual = (
        gnss_water
        - (
            tide_solution
            + zero_offset
        )
    )

    zero_abs_error_cm = (
        np.abs(
            zero_residual
        )
        * 100.0
    )

    zero_rms_cm = (
        rms(
            zero_residual
        )
        * 100.0
    )

    zero_mae_cm = (
        np.mean(
            zero_abs_error_cm
        )
    )

    zero_bias_cm = (
        np.mean(
            zero_residual
        )
        * 100.0
    )

    zero_correlation = pearson(
        gnss_water,
        tide_solution,
    )

    print()
    print(
        "--- ZERO-LAG PHYSICAL COMPARISON ---"
    )

    print(
        "Relationship:"
    )

    print(
        "  GNSS water level = tide + constant offset"
    )

    print(
        f"Best constant offset : "
        f"{zero_offset:.4f} m"
    )

    print(
        f"RMS residual         : "
        f"{zero_rms_cm:.2f} cm"
    )

    print(
        f"MAE residual         : "
        f"{zero_mae_cm:.2f} cm"
    )

    print(
        f"Mean residual        : "
        f"{zero_bias_cm:.2f} cm"
    )

    print(
        f"Correlation          : "
        f"{zero_correlation:.4f}"
    )

    # ---------------------------------------------------------------
    # Validate GNSS-IR solution timing
    # ---------------------------------------------------------------

    timing_offsets = np.array(
        timing_checks,
        dtype=float,
    )

    print()
    print(
        "--- GNSS-IR SOLUTION EPOCH CHECK ---"
    )

    print(
        "Solution epoch relative to arc midpoint:"
    )

    print(
        f"  min    = "
        f"{timing_offsets.min():.2f} sec"
    )

    print(
        f"  median = "
        f"{np.median(timing_offsets):.2f} sec"
    )

    print(
        f"  mean   = "
        f"{timing_offsets.mean():.2f} sec"
    )

    print(
        f"  max    = "
        f"{timing_offsets.max():.2f} sec"
    )

    # ---------------------------------------------------------------
    # Tide timing-lag search
    # ---------------------------------------------------------------

    print()
    print(
        "--- TIDE TIMING-LAG SEARCH ---"
    )

    lag_rows = []

    solution_times = [
        datetime.fromisoformat(
            r[
                "solution_time_utc"
            ]
        )
        for r in records
    ]

    for lag_minutes in LAG_MINUTES:

        shifted_tide = np.array(
            [
                tide_at(
                    t
                    + timedelta(
                        minutes=float(
                            lag_minutes
                        )
                    )
                )
                for t in solution_times
            ],
            dtype=float,
        )

        valid = np.isfinite(
            shifted_tide
        )

        if np.count_nonzero(
            valid
        ) < 10:

            continue

        offset = fixed_slope_offset(
            gnss_water[valid],
            shifted_tide[valid],
        )

        residual = (
            gnss_water[valid]
            - (
                shifted_tide[valid]
                + offset
            )
        )

        rms_cm = (
            rms(residual)
            * 100.0
        )

        mae_cm = (
            np.mean(
                np.abs(
                    residual
                )
            )
            * 100.0
        )

        corr = pearson(
            gnss_water[valid],
            shifted_tide[valid],
        )

        lag_rows.append(
            {
                "lag_minutes":
                    float(
                        lag_minutes
                    ),

                "offset_m":
                    float(offset),

                "rms_cm":
                    float(rms_cm),

                "mae_cm":
                    float(mae_cm),

                "correlation":
                    float(corr),
            }
        )

    if not lag_rows:

        print(
            "ERROR: no valid timing-lag results."
        )

        sys.exit(1)

    best_lag = min(
        lag_rows,
        key=lambda row:
        row["rms_cm"]
    )

    zero_lag_row = min(
        lag_rows,
        key=lambda row:
        abs(
            row["lag_minutes"]
        )
    )

    print(
        f"Best lag           : "
        f"{best_lag['lag_minutes']:.0f} min"
    )

    print(
        f"Best RMS            : "
        f"{best_lag['rms_cm']:.2f} cm"
    )

    print(
        f"Zero-lag RMS        : "
        f"{zero_lag_row['rms_cm']:.2f} cm"
    )

    print(
        f"RMS improvement     : "
        f"{zero_lag_row['rms_cm'] - best_lag['rms_cm']:.2f} cm"
    )

    # Confirmed necessary: flag loudly if the search still landed at
    # its own boundary, since that's a real sign the range may still
    # be too narrow to have found the true optimum, not evidence the
    # true lag is exactly at that value.
    if abs(best_lag["lag_minutes"] - LAG_MINUTES[0]) < 1e-6 or abs(
        best_lag["lag_minutes"] - LAG_MINUTES[-1]
    ) < 1e-6:
        print(
            "  WARNING: best lag landed exactly at the search "
            "boundary -- the true optimum may lie outside this "
            "range. Consider widening LAG_MINUTES further."
        )

    # ---------------------------------------------------------------
    # Within-arc tidal-smearing test
    # ---------------------------------------------------------------

    primary_pearson = pearson(
        zero_abs_error_cm,
        abs_tide_change,
    )

    primary_spearman = spearman(
        zero_abs_error_cm,
        abs_tide_change,
    )

    duration_pearson = pearson(
        zero_abs_error_cm,
        durations,
    )

    duration_spearman = spearman(
        zero_abs_error_cm,
        durations,
    )

    # Tide rate.
    tide_rate = np.full(
        len(records),
        np.nan,
        dtype=float,
    )

    valid_duration = (
        durations > 0
    )

    tide_rate[
        valid_duration
    ] = (
        tide_change[
            valid_duration
        ]
        / durations[
            valid_duration
        ]
    )

    rate_pearson = pearson(
        zero_abs_error_cm,
        np.abs(
            tide_rate
        ),
    )

    print()
    print(
        "--- WITHIN-ARC TIDAL-SMEARING TEST ---"
    )

    print(
        f"Pearson r(error, tidal change): "
        f"{primary_pearson:.4f}"
    )

    print(
        f"Spearman rho(error, tidal change): "
        f"{primary_spearman:.4f}"
    )

    print(
        f"Pearson r(error, arc duration): "
        f"{duration_pearson:.4f}"
    )

    print(
        f"Spearman rho(error, arc duration): "
        f"{duration_spearman:.4f}"
    )

    print(
        f"Pearson r(error, tidal rate): "
        f"{rate_pearson:.4f}"
    )

    print()
    print(
        "--- WITHIN-ARC TIDE DISTRIBUTION ---"
    )

    print(
        f"Minimum : "
        f"{abs_tide_change.min():.2f} cm"
    )

    print(
        f"Mean    : "
        f"{abs_tide_change.mean():.2f} cm"
    )

    print(
        f"Median  : "
        f"{np.median(abs_tide_change):.2f} cm"
    )

    print(
        f"Maximum : "
        f"{abs_tide_change.max():.2f} cm"
    )

    print(
        f">2 cm   : "
        f"{100*np.mean(abs_tide_change > 2):.1f}%"
    )

    print(
        f">5 cm   : "
        f"{100*np.mean(abs_tide_change > 5):.1f}%"
    )

    print(
        f">10 cm  : "
        f"{100*np.mean(abs_tide_change > 10):.1f}%"
    )

    # ---------------------------------------------------------------
    # Error by tidal movement bin
    # ---------------------------------------------------------------

    print()
    print(
        "--- ERROR BY WITHIN-ARC TIDAL MOVEMENT ---"
    )

    bins = [
        (
            "0-2 cm",
            0,
            2,
        ),
        (
            "2-5 cm",
            2,
            5,
        ),
        (
            "5-10 cm",
            5,
            10,
        ),
        (
            "10-20 cm",
            10,
            20,
        ),
        (
            ">20 cm",
            20,
            float("inf"),
        ),
    ]

    bin_rows = []

    for label, low, high in bins:

        mask = (
            (abs_tide_change >= low)
            & (
                abs_tide_change < high
            )
        )

        n = int(
            np.count_nonzero(
                mask
            )
        )

        if n:

            mean_error = float(
                np.mean(
                    zero_abs_error_cm[
                        mask
                    ]
                )
            )

            median_error = float(
                np.median(
                    zero_abs_error_cm[
                        mask
                    ]
                )
            )

            rms_error = float(
                rms(
                    zero_residual[
                        mask
                    ]
                )
                * 100.0
            )

        else:

            mean_error = float("nan")
            median_error = float("nan")
            rms_error = float("nan")

        print(
            f"{label:>10}: "
            f"n={n:4d} "
            f"mean={mean_error:7.2f} cm "
            f"median={median_error:7.2f} cm "
            f"RMS={rms_error:7.2f} cm"
        )

        bin_rows.append(
            {
                "bin": label,
                "n": n,
                "mean_abs_error_cm":
                    mean_error,
                "median_abs_error_cm":
                    median_error,
                "rms_error_cm":
                    rms_error,
            }
        )

    # ---------------------------------------------------------------
    # Rising / falling tide
    # ---------------------------------------------------------------

    print()
    print(
        "--- ERROR BY TIDE DIRECTION ---"
    )

    direction_rows = []

    for label, mask in [
        (
            "RISING",
            tide_change > 0,
        ),
        (
            "FALLING",
            tide_change < 0,
        ),
    ]:

        n = int(
            np.count_nonzero(
                mask
            )
        )

        if not n:
            continue

        signed_mean = (
            np.mean(
                zero_residual[
                    mask
                ]
            )
            * 100.0
        )

        abs_mean = np.mean(
            zero_abs_error_cm[
                mask
            ]
        )

        print(
            f"{label:>8}: "
            f"n={n:4d} "
            f"signed mean="
            f"{signed_mean:8.3f} cm "
            f"absolute mean="
            f"{abs_mean:8.3f} cm"
        )

        direction_rows.append(
            {
                "direction": label,
                "n": n,
                "signed_mean_cm":
                    signed_mean,
                "mean_abs_error_cm":
                    abs_mean,
            }
        )

    # ---------------------------------------------------------------
    # Frequency analysis
    # ---------------------------------------------------------------

    print()
    print(
        "--- ERROR BY FREQUENCY ---"
    )

    frequency_rows = []

    frequencies = sorted(
        {
            r["freq"]
            for r in records
            if r["freq"] is not None
        }
    )

    for freq in frequencies:

        mask = np.array(
            [
                r["freq"] == freq
                for r in records
            ],
            dtype=bool,
        )

        n = int(
            np.count_nonzero(
                mask
            )
        )

        if not n:
            continue

        mean_error = float(
            np.mean(
                zero_abs_error_cm[
                    mask
                ]
            )
        )

        rms_error = float(
            rms(
                zero_residual[
                    mask
                ]
            )
            * 100.0
        )

        correlation = pearson(
            zero_abs_error_cm[
                mask
            ],
            abs_tide_change[
                mask
            ],
        )

        print(
            f"freq={str(freq):>4}: "
            f"n={n:4d} "
            f"mean error="
            f"{mean_error:7.2f} cm "
            f"RMS="
            f"{rms_error:7.2f} cm "
            f"r="
            f"{correlation:.3f}"
        )

        frequency_rows.append(
            {
                "freq": freq,
                "n": n,
                "mean_abs_error_cm":
                    mean_error,
                "rms_error_cm":
                    rms_error,
                "error_tide_change_r":
                    correlation,
            }
        )

    # ---------------------------------------------------------------
    # Add calculated values to records
    # ---------------------------------------------------------------

    for index, record in enumerate(
        records
    ):

        record[
            "zero_lag_offset_m"
        ] = zero_offset

        record[
            "zero_lag_residual_cm"
        ] = (
            zero_residual[index]
            * 100.0
        )

        record[
            "zero_lag_abs_error_cm"
        ] = (
            zero_abs_error_cm[index]
        )

        record[
            "tide_rate_cm_per_min"
        ] = tide_rate[index]

    # ---------------------------------------------------------------
    # Save detailed CSV
    # ---------------------------------------------------------------

    csv_path = Path(
        "gnssir_tide_arc_analysis.csv"
    )

    fieldnames = list(
        records[0].keys()
    )

    with csv_path.open(
        "w",
        newline="",
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames,
        )

        writer.writeheader()

        writer.writerows(
            records
        )

    # ---------------------------------------------------------------
    # Save lag CSV
    # ---------------------------------------------------------------

    lag_csv_path = Path(
        "gnssir_tide_lag_analysis.csv"
    )

    with lag_csv_path.open(
        "w",
        newline="",
    ) as f:

        writer = csv.DictWriter(
            f,
            fieldnames=[
                "lag_minutes",
                "offset_m",
                "rms_cm",
                "mae_cm",
                "correlation",
            ],
        )

        writer.writeheader()

        writer.writerows(
            lag_rows
        )

    # ---------------------------------------------------------------
    # Save summary
    # ---------------------------------------------------------------

    summary_path = Path(
        "gnssir_tide_analysis_summary.txt"
    )

    with summary_path.open(
        "w"
    ) as f:

        f.write(
            "USGS GNSS-IR / TIDE MODEL ANALYSIS\n"
        )

        f.write(
            "=" * 70
            + "\n"
        )

        f.write(
            f"Station: {station}\n"
        )

        f.write(
            f"Year: {year}\n"
        )

        f.write(
            f"DOY: {doy1}-{doy2}\n"
        )

        f.write(
            f"Reference height: "
            f"{reference_height:.6f} m\n"
        )

        f.write(
            f"Total extracted: "
            f"{total_extracted}\n"
        )

        f.write(
            f"QC passing: "
            f"{total_qc}\n"
        )

        f.write(
            f"Final paired arcs: "
            f"{len(records)}\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "ARC SETTINGS\n"
        )

        f.write(
            "e1=5.0\n"
        )

        f.write(
            "e2=15.0\n"
        )

        f.write(
            "azlist=[100,130,150,215]\n"
        )

        f.write(
            "polyV=2\n"
        )

        f.write(
            "pele=[5,30]\n"
        )

        f.write(
            "attach_results=True\n"
        )

        f.write(
            "buffer_hours=2\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "QC DEFINITION\n"
        )

        f.write(
            "meta['gnssir_processing_results'] "
            "is not None\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "ZERO-LAG PHYSICAL COMPARISON\n"
        )

        f.write(
            "GNSS water level = H_reference - RH\n"
        )

        f.write(
            "Comparison = tide + constant offset\n"
        )

        f.write(
            f"offset_m={zero_offset:.9f}\n"
        )

        f.write(
            f"RMS_cm={zero_rms_cm:.6f}\n"
        )

        f.write(
            f"MAE_cm={zero_mae_cm:.6f}\n"
        )

        f.write(
            f"mean_bias_cm={zero_bias_cm:.6f}\n"
        )

        f.write(
            f"correlation={zero_correlation:.9f}\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "SOLUTION EPOCH RELATIVE TO ARC MIDPOINT\n"
        )

        f.write(
            f"min_sec={timing_offsets.min():.6f}\n"
        )

        f.write(
            f"median_sec={np.median(timing_offsets):.6f}\n"
        )

        f.write(
            f"mean_sec={timing_offsets.mean():.6f}\n"
        )

        f.write(
            f"max_sec={timing_offsets.max():.6f}\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "TIMING LAG SEARCH\n"
        )

        f.write(
            f"search_range_minutes=[{LAG_MINUTES[0]:.0f}, {LAG_MINUTES[-1]:.0f}]\n"
        )

        f.write(
            f"best_lag_minutes="
            f"{best_lag['lag_minutes']:.6f}\n"
        )

        f.write(
            f"best_rms_cm="
            f"{best_lag['rms_cm']:.6f}\n"
        )

        f.write(
            f"zero_lag_rms_cm="
            f"{zero_lag_row['rms_cm']:.6f}\n"
        )

        f.write(
            f"rms_improvement_cm="
            f"{zero_lag_row['rms_cm'] - best_lag['rms_cm']:.6f}\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "WITHIN-ARC TIDAL CHANGE\n"
        )

        f.write(
            f"minimum_cm="
            f"{abs_tide_change.min():.6f}\n"
        )

        f.write(
            f"mean_cm="
            f"{abs_tide_change.mean():.6f}\n"
        )

        f.write(
            f"median_cm="
            f"{np.median(abs_tide_change):.6f}\n"
        )

        f.write(
            f"maximum_cm="
            f"{abs_tide_change.max():.6f}\n"
        )

        f.write(
            f"percent_over_2cm="
            f"{100*np.mean(abs_tide_change > 2):.6f}\n"
        )

        f.write(
            f"percent_over_5cm="
            f"{100*np.mean(abs_tide_change > 5):.6f}\n"
        )

        f.write(
            f"percent_over_10cm="
            f"{100*np.mean(abs_tide_change > 10):.6f}\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "WITHIN-ARC TIDAL-SMEARING TEST\n"
        )

        f.write(
            f"pearson_error_vs_abs_tide_change="
            f"{primary_pearson:.9f}\n"
        )

        f.write(
            f"spearman_error_vs_abs_tide_change="
            f"{primary_spearman:.9f}\n"
        )

        f.write(
            f"pearson_error_vs_duration="
            f"{duration_pearson:.9f}\n"
        )

        f.write(
            f"spearman_error_vs_duration="
            f"{duration_spearman:.9f}\n"
        )

        f.write(
            f"pearson_error_vs_abs_tide_rate="
            f"{rate_pearson:.9f}\n"
        )

        f.write(
            "\n"
        )

        f.write(
            "TIDAL MOVEMENT BINS\n"
        )

        for row in bin_rows:

            f.write(
                f"{row['bin']}: "
                f"n={row['n']}, "
                f"mean_abs_error_cm="
                f"{row['mean_abs_error_cm']:.6f}, "
                f"median_abs_error_cm="
                f"{row['median_abs_error_cm']:.6f}, "
                f"rms_cm="
                f"{row['rms_error_cm']:.6f}\n"
            )

        f.write(
            "\n"
        )

        f.write(
            "RISING/FALLING\n"
        )

        for row in direction_rows:

            f.write(
                f"{row['direction']}: "
                f"n={row['n']}, "
                f"signed_mean_cm="
                f"{row['signed_mean_cm']:.6f}, "
                f"mean_abs_error_cm="
                f"{row['mean_abs_error_cm']:.6f}\n"
            )

        f.write(
            "\n"
        )

        f.write(
            "FREQUENCY SUMMARY\n"
        )

        for row in frequency_rows:

            f.write(
                f"freq={row['freq']}: "
                f"n={row['n']}, "
                f"mean_abs_error_cm="
                f"{row['mean_abs_error_cm']:.6f}, "
                f"rms_cm="
                f"{row['rms_error_cm']:.6f}, "
                f"error_tide_change_r="
                f"{row['error_tide_change_r']:.6f}\n"
            )

    # ---------------------------------------------------------------
    # Plot 1: physical relationship
    # ---------------------------------------------------------------

    fig, ax = plt.subplots(
        figsize=(10, 7)
    )

    ax.scatter(
        tide_solution,
        gnss_water,
        s=18,
        alpha=0.45,
    )

    lo = min(
        np.min(tide_solution),
        np.min(gnss_water),
    )

    hi = max(
        np.max(tide_solution),
        np.max(gnss_water),
    )

    line = np.linspace(
        lo,
        hi,
        100,
    )

    ax.plot(
        line,
        line + zero_offset,
        linewidth=2,
        label=(
            "unit-slope fit + offset"
        ),
    )

    ax.set_xlabel(
        "Ensemble tide model height (m)"
    )

    ax.set_ylabel(
        "GNSS-IR water-level proxy H - RH (m)"
    )

    ax.set_title(
        "GNSS-IR vs Tide Model — Physical Unit-Slope Comparison"
    )

    ax.legend()

    ax.grid(
        alpha=0.3
    )

    fig.tight_layout()

    relationship_plot = Path(
        "gnssir_tide_relationship.png"
    )

    fig.savefig(
        relationship_plot,
        dpi=160,
    )

    plt.close(fig)

    # ---------------------------------------------------------------
    # Plot 2: lag search
    # ---------------------------------------------------------------

    lag_x = np.array(
        [
            row[
                "lag_minutes"
            ]
            for row in lag_rows
        ],
        dtype=float,
    )

    lag_rms = np.array(
        [
            row["rms_cm"]
            for row in lag_rows
        ],
        dtype=float,
    )

    lag_corr = np.array(
        [
            row[
                "correlation"
            ]
            for row in lag_rows
        ],
        dtype=float,
    )

    fig, ax1 = plt.subplots(
        figsize=(10, 7)
    )

    ax1.plot(
        lag_x,
        lag_rms,
        marker="o",
        markersize=3,
    )

    ax1.axvline(
        0,
        linestyle="--",
    )

    ax1.axvline(
        best_lag[
            "lag_minutes"
        ],
        linestyle=":",
    )

    ax1.set_xlabel(
        "Tide-model time shift (minutes)"
    )

    ax1.set_ylabel(
        "RMS residual (cm)"
    )

    ax1.set_title(
        "GNSS-IR / Tide Model Timing-Lag Search"
    )

    ax1.grid(
        alpha=0.3
    )

    ax2 = ax1.twinx()

    ax2.plot(
        lag_x,
        lag_corr,
        linestyle="--",
        alpha=0.7,
    )

    ax2.set_ylabel(
        "Correlation"
    )

    fig.tight_layout()

    lag_plot = Path(
        "gnssir_tide_lag_search.png"
    )

    fig.savefig(
        lag_plot,
        dpi=160,
    )

    plt.close(fig)

    # ---------------------------------------------------------------
    # Plot 3: within-arc smearing
    # ---------------------------------------------------------------

    fig, ax = plt.subplots(
        figsize=(10, 7)
    )

    ax.scatter(
        abs_tide_change,
        zero_abs_error_cm,
        s=18,
        alpha=0.45,
    )

    ax.set_xlabel(
        "Absolute tidal change during arc (cm)"
    )

    ax.set_ylabel(
        "Absolute GNSS-IR residual (cm)"
    )

    ax.set_title(
        "GNSS-IR Error vs Within-Arc Tidal Movement"
        f"\nPearson r={primary_pearson:.3f}, "
        f"Spearman rho={primary_spearman:.3f}"
    )

    ax.grid(
        alpha=0.3
    )

    fig.tight_layout()

    smearing_plot = Path(
        "gnssir_tide_smearing.png"
    )

    fig.savefig(
        smearing_plot,
        dpi=160,
    )

    plt.close(fig)

    # ---------------------------------------------------------------
    # Final output
    # ---------------------------------------------------------------

    print()
    print("=" * 72)
    print("ANALYSIS COMPLETE")
    print("=" * 72)

    print()
    print(
        "Files:"
    )

    print(
        f"  {csv_path}"
    )

    print(
        f"  {lag_csv_path}"
    )

    print(
        f"  {summary_path}"
    )

    print(
        f"  {relationship_plot}"
    )

    print(
        f"  {lag_plot}"
    )

    print(
        f"  {smearing_plot}"
    )

    print()
    print(
        "KEY RESULTS"
    )

    print(
        f"  QC arcs              : {total_qc}"
    )

    print(
        f"  Paired arcs          : {len(records)}"
    )

    print(
        f"  Zero-lag RMS         : "
        f"{zero_rms_cm:.2f} cm"
    )

    print(
        f"  Best lag             : "
        f"{best_lag['lag_minutes']:.0f} min"
    )

    print(
        f"  Best-lag RMS         : "
        f"{best_lag['rms_cm']:.2f} cm"
    )

    print(
        f"  Error vs tidal move : "
        f"r={primary_pearson:.4f}"
    )

    print()
    print(
        "No production GNSS-IR processing was changed."
    )


if __name__ == "__main__":
    main()
