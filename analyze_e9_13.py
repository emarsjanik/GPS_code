#!/usr/bin/env python3
"""
analyze_norefr_azimuth_test.py

Runs the same track-clustering, tide-correlation, and azimuth-vs-
correlation-strength test used throughout tonight's investigation,
but on the ocean17_23_e9_13 dataset (refraction correction disabled)
instead of the normal ocean17_23_l1_e5_13 dataset (refraction
correction on, Bennett model).

Purpose: test whether disabling tropospheric refraction correction
changes the azimuth-independence of the tide correlation found
repeatedly in the corrected data. If refraction is driving the shared,
azimuth-independent pattern, removing its correction should measurably
change (most likely worsen) this pattern. If it makes no difference,
refraction correction is not the source.

Usage:
    python3 analyze_norefr_azimuth_test.py
"""

import math
from pathlib import Path
from collections import defaultdict
from datetime import datetime, timedelta

import numpy as np
from openpyxl import load_workbook

RESULT_DIR = Path("products/refl_code/2026/results/usgs/ocean17_23_e9_13")
TIDE_FILE = Path("marconi_tides_sherwood.xlsx")

DOYS = [190, 191, 192, 193, 194, 195, 196, 197, 198, 199, 200, 201, 203,
        204, 205, 206, 207, 208, 209, 210, 211, 212, 213, 214, 215, 216,
        217, 218, 219, 220, 221, 222]

MODEL = "EOT20_heightm"


def finite(x):
    try:
        v = float(x)
        return v if math.isfinite(v) else None
    except Exception:
        return None


def pearson(x, y):
    x = np.asarray(x, float)
    y = np.asarray(y, float)
    if len(x) < 5:
        return float("nan")
    if np.std(x) == 0 or np.std(y) == 0:
        return float("nan")
    return float(np.corrcoef(x, y)[0, 1])


# Load tide model
wb = load_workbook(TIDE_FILE, data_only=True)
ws = wb[wb.sheetnames[0]]
header = [c.value for c in ws[1]]
time_col = header.index("time")
tide_col = header.index(MODEL)

times, values = [], []
for row in ws.iter_rows(min_row=2, values_only=True):
    t = row[time_col]
    if not isinstance(t, datetime):
        continue
    v = finite(row[tide_col])
    if v is None:
        continue
    times.append(t)
    values.append(v)
wb.close()

epoch = np.array([(t - times[0]).total_seconds() for t in times])
values = np.asarray(values, float)


def tide_at(dt):
    x = (dt - times[0]).total_seconds()
    if x < epoch[0] or x > epoch[-1]:
        return float("nan")
    return float(np.interp(x, epoch, values))


# Load GNSS-IR results
tracks = defaultdict(list)
for doy in DOYS:
    path = RESULT_DIR / f"{doy}.txt"
    if not path.exists():
        continue
    with open(path, errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("%"):
                continue
            cols = line.split()
            if len(cols) < 17:
                continue
            try:
                year = int(float(cols[0]))
                doy2 = int(float(cols[1]))
                rh = float(cols[2])
                sat = int(float(cols[3]))
                utc_hours = float(cols[4])
                az = float(cols[5])
                freq = int(float(cols[10]))
            except Exception:
                continue
            if freq != 1:
                continue
            day = datetime(year, 1, 1) + timedelta(days=doy2 - 1)
            dt = day + timedelta(hours=utc_hours)
            tracks[sat].append({"dt": dt, "rh": rh, "az": az})

print(f"Tracks found (by satellite): {len(tracks)}")

results = []
for sat, observations in sorted(tracks.items()):
    if len(observations) < 5:
        continue
    observations.sort(key=lambda r: r["dt"])
    rh = np.array([r["rh"] for r in observations])
    az = np.array([r["az"] for r in observations])
    tide = np.array([tide_at(r["dt"]) for r in observations])
    valid = np.isfinite(tide)
    if np.sum(valid) < 5:
        continue
    r_corr = pearson(rh[valid], tide[valid])
    results.append({
        "sat": sat, "n": int(np.sum(valid)),
        "az_mean": float(np.mean(az)), "r": r_corr,
    })

results = [r for r in results if math.isfinite(r["r"])]
results.sort(key=lambda r: r["az_mean"])

print()
print("=" * 70)
print("NO-REFRACTION-CORRECTION DATASET: azimuth vs correlation")
print("=" * 70)
print(f"{'sat':>4} {'az':>7} {'n':>4} {'r':>8}")
for r in results:
    print(f"{r['sat']:>4} {r['az_mean']:7.2f} {r['n']:>4} {r['r']:8.4f}")

az = np.array([r["az_mean"] for r in results])
rvals = np.array([abs(r["r"]) for r in results])
print()
if len(results) < 6:
    print(f"Only {len(results)} tracks survived (need >=6 for a meaningful "
          f"azimuth-correlation estimate) -- NOT reporting a summary "
          f"statistic, since it would be a degenerate/near-degenerate "
          f"result (e.g. exactly +-1.0 with only 2 points) rather than "
          f"a real finding.")
else:
    print(f"corr(azimuth, |r|), this elevation window: {np.corrcoef(az, rvals)[0,1]:+.4f}")
print()
print("Compare directly against the full-window (5-13 deg) result found")
print("earlier tonight: corr(azimuth, |r|) = -0.4156 (N>=10 tracks)")
