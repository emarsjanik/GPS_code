#!/usr/bin/env python3
"""
Marconi Long-Term GNSS-R / Tide Pipeline V3

Purpose
-------
1. Search the NUC and mounted I2Rgus storage for ALL local USGS RINEX files.
2. Stage them into the canonical gnssrefl RINEX directory without unnecessary
   copying (symlinks are preferred).
3. Generate/re-generate SNR 66 files for every discovered RINEX day.
4. Run gnssir for EVERY available SNR day using the existing
   ocean17_23_l1_e5_13 experiment.
5. Load ALL successful GPS L1 GNSS-IR retrievals.
6. Cluster repeated satellite tracks by satellite + frequency + rise/set +
   azimuth.
7. Apply the established "good track" screen:
       observations >= 14
       unique days   >= 14
       tide r        >= 0.90
       slope         0.85 to 1.15
       fitted/unit RMS <= 0.30 m
       azimuth SD    <= 1.0 deg
8. Plot ONLY the good tracks.
9. Report pooled point-by-point residual statistics for the entire good-track
   population, both raw and after the diagnostic +0.242 m GNSS-R offset:
       mean bias
       mean absolute deviation
       median absolute deviation
       RMS
10. Also write per-track statistics and a complete processing inventory.

Important
---------
The +0.242 m value is a diagnostic datum offset. It is NOT used to decide
whether a track is good. Good-track selection is based on the raw GNSS-R
relationship with the tide model and the fitted/unit RMS.

This script intentionally re-processes every discovered SNR day rather than
only days that appear to be missing results. That prevents an old result file
from silently remaining in the long-term analysis after RINEX/SNR changes.

The script is designed for:
    ~/GNSS/v4.1
with:
    $REFL_CODE = ~/GNSS/v4.1/products/refl_code

gnssrefl version used during development:
    4.1.5
"""

from __future__ import annotations

import csv
import math
import os
import re
import shutil
import subprocess
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

try:
    from gnssrefl.extract_arcs import extract_arcs_from_station
except Exception as exc:
    raise SystemExit(
        "Could not import gnssrefl.extract_arcs. "
        "Make sure the gnssrefl_venv is active.\n"
        f"Import error: {exc}"
    )


# ============================================================================
# CONFIGURATION
# ============================================================================

YEAR = 2026
STATION_SHORT = "usgs"
STATION_RINEX = "usgs00usa"

BASE_DIR = Path.home() / "GNSS" / "v4.1"
REFL_CODE = BASE_DIR / "products" / "refl_code"

RINEX_DIR = REFL_CODE / str(YEAR) / "rinex" / STATION_SHORT
SNR_DIR = REFL_CODE / str(YEAR) / "snr" / STATION_SHORT
RESULT_DIR = (
    REFL_CODE
    / str(YEAR)
    / "results"
    / STATION_SHORT
    / "ocean17_23_l1_e5_13"
)
LOG_DIR = (
    REFL_CODE
    / "logs"
    / STATION_SHORT
    / "ocean17_23_l1_e5_13"
    / str(YEAR)
)

TIDE_FILE = BASE_DIR / "marconi_tides_sherwood.xlsx"
PLOT_DIR = BASE_DIR / "marconi_longterm_plots_v3"

OUT_CSV = BASE_DIR / "marconi_longterm_track_stability_v3.csv"
OUT_SUMMARY = BASE_DIR / "marconi_longterm_track_stability_v3_summary.txt"
OUT_METRICS = BASE_DIR / "marconi_good_population_metrics_v3.csv"
OUT_INVENTORY = BASE_DIR / "marconi_rinex_processing_inventory_v3.csv"

# Diagnostic datum test only.
DATUM_OFFSET_M = 0.242

PRIMARY_TIDE_MODEL = "EOT20_heightm"

# Established long-term track screen.
AZ_CLUSTER_TOL_DEG = 3.0
MIN_OBS = 14
MIN_DAYS = 14
MIN_R = 0.90
MIN_SLOPE = 0.85
MAX_SLOPE = 1.15
MAX_FIT_RMS_M = 0.30
MAX_AZ_SD_DEG = 1.0

# gnssrefl processing configuration already established for this project.
SNR_TYPE = "66"
GNSSIR_EXTENSION = "ocean17_23_l1_e5_13"
GNSSIR_FREQ = "1"
GNSSIR_ELEV_MIN = "5"
GNSSIR_ELEV_MAX = "30"

# Search roots. The /mnt tree is where the older RINEX archive was found.
SEARCH_ROOTS = [
    BASE_DIR,
    Path.home(),
    Path("/mnt/I2Rgus_Data"),
]

RINEX_RE = re.compile(
    r"^USGS00USA_R_(\d{4})(\d{3})00000_01D_01S_MO\.rnx$",
    re.IGNORECASE,
)


# ============================================================================
# GENERAL HELPERS
# ============================================================================

def finite(x):
    try:
        y = float(x)
        return y if math.isfinite(y) else None
    except Exception:
        return None


def run_command(cmd, env=None, cwd=None, label=None):
    print()
    print("-" * 92)
    if label:
        print(label)
    print("$ " + " ".join(str(x) for x in cmd))
    print("-" * 92)

    result = subprocess.run(
        [str(x) for x in cmd],
        cwd=str(cwd) if cwd else None,
        env=env,
        text=True,
    )

    if result.returncode != 0:
        print(f"COMMAND FAILED: return code {result.returncode}")
    return result.returncode


def ensure_directories():
    for p in [RINEX_DIR, SNR_DIR, RESULT_DIR, LOG_DIR, PLOT_DIR]:
        p.mkdir(parents=True, exist_ok=True)


def make_env():
    env = os.environ.copy()
    env["REFL_CODE"] = str(REFL_CODE)
    env["EXE"] = str(REFL_CODE / "exe")
    env["ORBITS"] = str(REFL_CODE / "orbits")
    return env


def doy_to_datetime(year, doy, utc_hours):
    day = datetime(year, 1, 1) + timedelta(days=int(doy) - 1)
    return day + timedelta(hours=float(utc_hours))


def circular_azimuth_difference(a, b):
    return abs((a - b + 180.0) % 360.0 - 180.0)


def circular_mean_deg(values):
    values = np.asarray(values, dtype=float)
    if len(values) == 0:
        return math.nan
    r = np.radians(values)
    s = np.mean(np.sin(r))
    c = np.mean(np.cos(r))
    return float(np.degrees(np.arctan2(s, c)) % 360.0)


# ============================================================================
# RINEX DISCOVERY / STAGING
# ============================================================================

def discover_rinex():
    """
    Search the local NUC and /mnt storage for exact USGS daily RINEX files.

    If duplicate copies exist, retain all paths in the inventory but select
    one preferred source per DOY. Preference:
      1. canonical RINEX directory
      2. /mnt/I2Rgus_Data
      3. other local path
    """
    candidates = defaultdict(list)

    print()
    print("=" * 92)
    print("SEARCHING FOR ALL LOCAL USGS RINEX FILES")
    print("=" * 92)

    seen_roots = set()

    for root in SEARCH_ROOTS:
        root = root.resolve()
        if not root.exists():
            print(f"Search root not found: {root}")
            continue

        # Avoid scanning the same physical root twice.
        if root in seen_roots:
            continue
        seen_roots.add(root)

        print(f"Searching: {root}")

        cmd = [
            "find",
            str(root),
            "-type",
            "f",
            "-iname",
            "USGS00USA_R_2026???00000_01D_01S_MO.rnx",
            "-print",
        ]

        # find can return status 1 when it encounters a protected/busy
        # directory even though it successfully found matching files.
        # Do NOT discard stdout in that case.
        result = subprocess.run(
            cmd,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )

        if result.returncode != 0:
            print(
                f"  find returned {result.returncode}; "
                "continuing with matches that were found"
            )
            if result.stderr.strip():
                print(f"  find warning: {result.stderr.strip()[:500]}")

        out = result.stdout

        for line in out.splitlines():
            p = Path(line.strip())
            m = RINEX_RE.match(p.name)
            if not m:
                continue

            y = int(m.group(1))
            doy = int(m.group(2))

            if y != YEAR:
                continue

            try:
                size = p.stat().st_size
            except OSError:
                continue

            # Ignore obviously empty/truncated files.
            if size < 10000:
                print(f"  Ignoring tiny RINEX: {p} ({size} bytes)")
                continue

            candidates[doy].append(p)

    def preference(p):
        s = str(p)
        if p.parent == RINEX_DIR:
            return (0, s)
        if "/mnt/I2Rgus_Data/" in s:
            return (1, s)
        return (2, s)

    preferred = {}
    for doy, paths in sorted(candidates.items()):
        paths = sorted(set(paths), key=preference)
        preferred[doy] = paths[0]

        print()
        print(f"DOY {doy}: {len(paths)} copy/copies")
        for p in paths:
            marker = "*" if p == paths[0] else " "
            print(f" {marker} {p}  {p.stat().st_size:,} bytes")

    print()
    print(f"Unique RINEX days discovered: {len(preferred)}")
    print(
        "DOYs:",
        ", ".join(str(x) for x in sorted(preferred))
        if preferred
        else "NONE",
    )

    return preferred, candidates


def stage_rinex(preferred):
    """
    Stage each preferred RINEX in the canonical gnssrefl directory.

    Existing canonical regular files are retained.
    Otherwise a symlink is created to avoid duplicating hundreds of MB.
    """
    print()
    print("=" * 92)
    print("STAGING RINEX INTO CANONICAL GNSSREFL DIRECTORY")
    print("=" * 92)

    staged = {}

    for doy, source in sorted(preferred.items()):
        target = RINEX_DIR / source.name

        if target.exists():
            try:
                same = target.resolve() == source.resolve()
            except Exception:
                same = False

            if same:
                staged[doy] = target
                print(f"DOY {doy}: already staged: {target}")
                continue

            # Never silently replace a real canonical file.
            if target.is_file() and not target.is_symlink():
                if target.stat().st_size == source.stat().st_size:
                    staged[doy] = target
                    print(
                        f"DOY {doy}: canonical copy already present "
                        f"({target.stat().st_size:,} bytes)"
                    )
                    continue

                print(
                    f"DOY {doy}: canonical file differs in size; "
                    f"leaving it unchanged: {target}"
                )
                staged[doy] = target
                continue

        try:
            target.symlink_to(source)
            print(f"DOY {doy}: symlinked")
            print(f"  {target} -> {source}")
        except Exception:
            # Fall back to a copy only if symlink creation is not possible.
            print(f"DOY {doy}: symlink failed; copying...")
            shutil.copy2(source, target)
            print(
                f"  copied {target.stat().st_size:,} bytes"
            )

        staged[doy] = target

    return staged


# ============================================================================
# RINEX -> SNR
# ============================================================================

def snr_path(doy):
    return SNR_DIR / f"{STATION_SHORT}{doy:03d}0.26.snr{SNR_TYPE}.gz"


def validate_snr(path):
    if not path.exists():
        return False
    try:
        if path.stat().st_size < 500:
            return False
        # Check that gzip is readable and contains actual records.
        p = subprocess.run(
            ["gzip", "-cd", str(path)],
            stdout=subprocess.PIPE,
            stderr=subprocess.DEVNULL,
            timeout=20,
        )
        if p.returncode != 0:
            return False
        lines = p.stdout.splitlines()
        return len(lines) >= 5
    except Exception:
        return False


def convert_all_rinex(staged, env):
    print()
    print("=" * 92)
    print("CONVERTING ALL RINEX -> SNR 66")
    print("=" * 92)

    inventory = []

    for doy, _path in sorted(staged.items()):
        out_snr = snr_path(doy)
        before = out_snr.stat().st_size if out_snr.exists() else 0

        rc = run_command(
            [
                "rinex2snr",
                STATION_RINEX,
                YEAR,
                doy,
                "-snr",
                SNR_TYPE,
                "-orb",
                "gnss",
                "-stream",
                "R",
                "-samplerate",
                "1",
                "-nolook",
                "T",
                "-overwrite",
                "T",
            ],
            env=env,
            cwd=BASE_DIR,
            label=f"RINEX -> SNR: DOY {doy}",
        )

        ok = validate_snr(out_snr)
        after = out_snr.stat().st_size if out_snr.exists() else 0

        status = "OK" if rc == 0 and ok else "FAILED"
        print(
            f"DOY {doy}: {status}  "
            f"SNR={out_snr}  size={after:,} bytes"
        )

        inventory.append(
            {
                "doy": doy,
                "rinex": str(staged[doy]),
                "snr": str(out_snr),
                "rinex_to_snr_returncode": rc,
                "snr_valid": ok,
                "snr_size_before": before,
                "snr_size_after": after,
            }
        )

    return inventory


# ============================================================================
# SNR -> GNSSIR
# ============================================================================

def discover_snr_days():
    days = []
    for p in sorted(SNR_DIR.glob(f"{STATION_SHORT}*.snr{SNR_TYPE}.gz")):
        m = re.search(rf"{STATION_SHORT}(\d{{3}})0\.26\.snr{SNR_TYPE}", p.name)
        if m:
            doy = int(m.group(1))
            if validate_snr(p):
                days.append(doy)

    return sorted(set(days))


def process_all_snr(env):
    print()
    print("=" * 92)
    print("PROCESSING ALL SNR DAYS WITH GNSSIR")
    print("=" * 92)

    days = discover_snr_days()

    print(f"Valid SNR days available: {len(days)}")
    print(
        "DOYs:",
        ", ".join(str(x) for x in days) if days else "NONE",
    )

    successful = []
    failed = []

    for doy in days:
        result_file = RESULT_DIR / f"{doy}.txt"

        rc = run_command(
            [
                "gnssir",
                STATION_SHORT,
                YEAR,
                doy,
                "-extension",
                GNSSIR_EXTENSION,
                "-fr",
                GNSSIR_FREQ,
                "-nooverwrite",
                "False",
            ],
            env=env,
            cwd=BASE_DIR,
            label=f"GNSSIR: DOY {doy}",
        )

        # gnssrefl deliberately does not create a result file when no
        # retrievals survive QC.
        exists = result_file.exists() and result_file.stat().st_size > 0

        if rc == 0 and exists:
            successful.append(doy)
            print(f"VERIFIED DOY {doy}: {result_file}")
        else:
            failed.append(doy)
            print(
                f"FAILED DOY {doy}: result file was not created "
                f"or command failed"
            )

    return days, successful, failed


# ============================================================================
# TIDE MODEL
# ============================================================================

def load_tide_interpolator():
    if not TIDE_FILE.exists():
        raise SystemExit(f"Missing tide workbook: {TIDE_FILE}")

    wb = load_workbook(TIDE_FILE, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [c.value for c in ws[1]]
    if "time" not in headers:
        raise SystemExit(
            f"'time' column not found in {TIDE_FILE}"
        )
    if PRIMARY_TIDE_MODEL not in headers:
        raise SystemExit(
            f"'{PRIMARY_TIDE_MODEL}' column not found in {TIDE_FILE}"
        )

    ti = headers.index("time")
    vi = headers.index(PRIMARY_TIDE_MODEL)

    times = []
    values = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[ti]
        v = finite(row[vi])
        if not isinstance(t, datetime) or v is None:
            continue
        times.append(t)
        values.append(v)

    wb.close()

    if len(times) < 2:
        raise SystemExit("Not enough tide model points.")

    x0 = times[0]
    x = np.array(
        [(t - x0).total_seconds() for t in times],
        dtype=float,
    )
    y = np.asarray(values, dtype=float)

    def tide_at(dt):
        xx = (dt - x0).total_seconds()
        if xx < x[0] or xx > x[-1]:
            return math.nan
        return float(np.interp(xx, x, y))

    return tide_at, times[0], times[-1]


# ============================================================================
# GNSSIR RESULT LOADING
# ============================================================================

def load_all_retrievals(days):
    """
    Use gnssrefl's supported extract_arcs API and attach the GNSSIR
    processing results. This avoids depending on a hard-coded result-file
    column order.
    """
    rows = []

    print()
    print("=" * 92)
    print("LOADING ALL SUCCESSFUL GPS L1 GNSS-IR RETRIEVALS")
    print("=" * 92)

    for doy in days:
        try:
            arcs = extract_arcs_from_station(
                STATION_SHORT,
                YEAR,
                doy,
                freq=1,
                e1=5.0,
                e2=30.0,
                min_pts=1,
                buffer_hours=2,
                filter_to_day=True,
                attach_results=True,
            )
        except Exception as exc:
            print(f"DOY {doy}: extract_arcs failed: {exc}")
            continue

        n_attached = 0

        for meta, data in arcs:
            result = meta.get("gnssir_processing_results")
            if not result:
                continue

            sat = int(meta.get("sat"))
            freq = int(meta.get("freq"))
            rise = int(result.get("rise", 0))

            if freq != 1:
                continue

            rh = finite(result.get("RH"))
            az = finite(result.get("Azim"))
            utc = finite(result.get("UTCtime"))
            mjd = finite(result.get("MJD"))
            amp = finite(result.get("Amp"))
            pkn = finite(result.get("PkNoise"))

            if None in (rh, az, utc, mjd):
                continue

            dt = doy_to_datetime(YEAR, doy, utc)

            rows.append(
                {
                    "doy": doy,
                    "datetime_utc": dt,
                    "sat": sat,
                    "freq": freq,
                    "rise": rise,
                    "rh": rh,
                    "az": az,
                    "utc_hours": utc,
                    "mjd": mjd,
                    "amp": amp if amp is not None else math.nan,
                    "pkn": pkn if pkn is not None else math.nan,
                    "delT": finite(result.get("DelT"))
                    if result.get("DelT") is not None
                    else math.nan,
                    "emin": finite(result.get("eminO"))
                    if result.get("eminO") is not None
                    else math.nan,
                    "emax": finite(result.get("emaxO"))
                    if result.get("emaxO") is not None
                    else math.nan,
                }
            )
            n_attached += 1

        print(f"DOY {doy}: {n_attached} GPS L1 retrievals")

    rows.sort(key=lambda r: r["datetime_utc"])

    print()
    print(f"TOTAL GPS L1 GNSS-R RETRIEVALS: {len(rows)}")
    return rows


def add_tides(rows, tide_at):
    good = []
    for r in rows:
        t = tide_at(r["datetime_utc"])
        if not math.isfinite(t):
            continue

        r["tide"] = t
        r["residual_raw"] = r["rh"] - t
        r["residual_plus0242"] = r["rh"] + DATUM_OFFSET_M - t
        good.append(r)

    return good


# ============================================================================
# TRACK CLUSTERING / STATISTICS
# ============================================================================

def split_azimuth_clusters(group):
    """
    Cluster a sat/freq/rise group by azimuth.

    This follows the established 3-degree clustering approach. The cluster
    center is computed from the arithmetic mean for consistency with the
    previous V2 implementation.
    """
    if not group:
        return []

    ordered = sorted(group, key=lambda r: r["az"])
    clusters = [[ordered[0]]]

    for r in ordered[1:]:
        previous = clusters[-1][-1]
        if circular_azimuth_difference(r["az"], previous["az"]) <= AZ_CLUSTER_TOL_DEG:
            clusters[-1].append(r)
        else:
            clusters.append([r])

    # Merge first and last clusters if they straddle 0/360 degrees.
    if len(clusters) > 1:
        first = clusters[0]
        last = clusters[-1]
        if circular_azimuth_difference(first[0]["az"], last[-1]["az"]) <= AZ_CLUSTER_TOL_DEG:
            clusters[0] = last + first
            clusters.pop()

    return clusters


def linear_fit_stats(x, y):
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    valid = np.isfinite(x) & np.isfinite(y)
    x = x[valid]
    y = y[valid]

    if len(x) < 3:
        return math.nan, math.nan, math.nan, math.nan

    slope, intercept = np.polyfit(x, y, 1)
    pred = slope * x + intercept
    residual = y - pred

    rms = float(np.sqrt(np.mean(residual ** 2)))

    if np.std(x) > 0 and np.std(y) > 0:
        r = float(np.corrcoef(x, y)[0, 1])
    else:
        r = math.nan

    return (
        float(slope),
        float(intercept),
        r,
        rms,
    )


def track_metrics(group):
    tide = np.asarray([r["tide"] for r in group], dtype=float)
    rh = np.asarray([r["rh"] for r in group], dtype=float)

    slope, intercept, r, fit_rms = linear_fit_stats(tide, rh)

    raw = rh - tide
    plus = rh + DATUM_OFFSET_M - tide

    az = np.asarray([r["az"] for r in group], dtype=float)

    return {
        "n": int(len(group)),
        "days": int(len(set(r["doy"] for r in group))),
        "sat": int(group[0]["sat"]),
        "freq": int(group[0]["freq"]),
        "rise": int(group[0]["rise"]),
        "az_mean": circular_mean_deg(az),
        "az_sd": float(np.std(az)),
        "r": r,
        "slope": slope,
        "intercept": intercept,
        "fit_rms_m": fit_rms,
        "raw_bias_m": float(np.mean(raw)),
        "raw_mad_m": float(np.mean(np.abs(raw))),
        "raw_median_abs_m": float(np.median(np.abs(raw))),
        "raw_rms_m": float(np.sqrt(np.mean(raw ** 2))),
        "plus_bias_m": float(np.mean(plus)),
        "plus_mad_m": float(np.mean(np.abs(plus))),
        "plus_median_abs_m": float(np.median(np.abs(plus))),
        "plus_rms_m": float(np.sqrt(np.mean(plus ** 2))),
        "group": group,
    }


def is_good_track(rec):
    return (
        rec["n"] >= MIN_OBS
        and rec["days"] >= MIN_DAYS
        and math.isfinite(rec["r"])
        and rec["r"] >= MIN_R
        and math.isfinite(rec["slope"])
        and MIN_SLOPE <= rec["slope"] <= MAX_SLOPE
        and math.isfinite(rec["fit_rms_m"])
        and rec["fit_rms_m"] <= MAX_FIT_RMS_M
        and math.isfinite(rec["az_sd"])
        and rec["az_sd"] <= MAX_AZ_SD_DEG
    )


def build_tracks(rows):
    by_key = defaultdict(list)

    for r in rows:
        by_key[
            (r["sat"], r["freq"], r["rise"])
        ].append(r)

    tracks = []

    for key, group in sorted(by_key.items()):
        clusters = split_azimuth_clusters(group)
        for cluster in clusters:
            if len(cluster) < 3:
                continue
            rec = track_metrics(cluster)
            tracks.append(rec)

    tracks.sort(
        key=lambda x: (
            -(x["r"] if math.isfinite(x["r"]) else -999),
            x["fit_rms_m"] if math.isfinite(x["fit_rms_m"]) else 999,
        )
    )

    return tracks


# ============================================================================
# POOLED GOOD-POPULATION METRICS
# ============================================================================

def pooled_metrics(rows):
    """
    Point-by-point pooled statistics.

    This is the primary answer to:
      "How far, on average, does GNSS-R differ from the tide model?"

    It is NOT the unweighted average of the individual track RMS values.
    Every observation contributes one residual.
    """
    raw = np.asarray(
        [r["residual_raw"] for r in rows],
        dtype=float,
    )
    plus = np.asarray(
        [r["residual_plus0242"] for r in rows],
        dtype=float,
    )

    def one(a):
        return {
            "n": int(len(a)),
            "mean_bias_m": float(np.mean(a)),
            "mean_abs_dev_m": float(np.mean(np.abs(a))),
            "median_abs_dev_m": float(np.median(np.abs(a))),
            "rms_m": float(np.sqrt(np.mean(a ** 2))),
        }

    return {
        "raw": one(raw),
        "plus0242": one(plus),
    }


# ============================================================================
# PLOTS
# ============================================================================

def savefig(path):
    plt.tight_layout()
    plt.savefig(path, dpi=180)
    plt.close()


def plot_good_tracks_vs_tide(good_tracks, offset=False):
    filename = (
        "03_good_tracks_vs_tide_plus_0242m.png"
        if offset
        else "02_good_tracks_vs_tide_raw.png"
    )

    fig, ax = plt.subplots(figsize=(16, 8))

    all_rows = []
    for rec in good_tracks:
        all_rows.extend(rec["group"])

    all_rows.sort(key=lambda r: r["datetime_utc"])

    times = [r["datetime_utc"] for r in all_rows]
    tide = np.asarray([r["tide"] for r in all_rows], dtype=float)

    ax.plot(
        times,
        tide,
        linewidth=2.5,
        label=PRIMARY_TIDE_MODEL,
    )

    # Plot each good track separately so the figure contains only the
    # observations that passed the established track screen.
    for i, rec in enumerate(good_tracks, start=1):
        g = sorted(rec["group"], key=lambda r: r["datetime_utc"])
        t = [r["datetime_utc"] for r in g]

        if offset:
            y = [
                r["rh"] + DATUM_OFFSET_M
                for r in g
            ]
        else:
            y = [r["rh"] for r in g]

        ax.plot(
            t,
            y,
            marker="o",
            markersize=2.5,
            linewidth=0.8,
            alpha=0.55,
            label=(
                f"{i}: S{rec['sat']} "
                f"{'rise' if rec['rise'] == 1 else 'set'} "
                f"Az {rec['az_mean']:.1f}°"
            ),
        )

    ax.set_xlabel("UTC")
    ax.set_ylabel("Water level / reflector height (m)")
    title = (
        "Marconi: GOOD GNSS-R Tracks + 0.242 m vs EOT20"
        if offset
        else "Marconi: GOOD GNSS-R Tracks vs EOT20"
    )
    ax.set_title(title)
    ax.grid(alpha=0.25)
    ax.legend(
        fontsize=7,
        ncol=3,
        loc="upper left",
    )

    savefig(PLOT_DIR / filename)


def plot_good_tracks_scatter(good_tracks, offset=False):
    filename = (
        "04_good_tracks_scatter_plus_0242m.png"
        if offset
        else "01_good_tracks_scatter_raw.png"
    )

    x = []
    y = []

    for rec in good_tracks:
        for r in rec["group"]:
            x.append(r["tide"])
            y.append(
                r["rh"] + DATUM_OFFSET_M
                if offset
                else r["rh"]
            )

    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)

    slope, intercept, rr, _ = linear_fit_stats(x, y)

    fig, ax = plt.subplots(figsize=(9, 9))
    ax.scatter(
        x,
        y,
        s=15,
        alpha=0.45,
    )

    lo = min(np.min(x), np.min(y))
    hi = max(np.max(x), np.max(y))
    xx = np.linspace(lo, hi, 200)

    ax.plot(
        xx,
        xx,
        linestyle="--",
        linewidth=1.5,
        label="1:1",
    )

    if math.isfinite(slope):
        ax.plot(
            xx,
            slope * xx + intercept,
            linewidth=2,
            label=f"fit slope={slope:.3f}, r={rr:.3f}",
        )

    ax.set_xlabel("EOT20 tide (m)")
    ax.set_ylabel(
        "GNSS-R + 0.242 m (m)"
        if offset
        else "GNSS-R (m)"
    )
    ax.set_title(
        "GOOD-track GNSS-R vs EOT20"
        + (" (+0.242 m)" if offset else " (raw)")
    )
    ax.grid(alpha=0.25)
    ax.legend()

    savefig(PLOT_DIR / filename)


def plot_good_track_residuals(good_tracks):
    fig, ax = plt.subplots(figsize=(16, 8))

    for i, rec in enumerate(good_tracks, start=1):
        g = sorted(rec["group"], key=lambda r: r["datetime_utc"])
        t = [r["datetime_utc"] for r in g]
        y = [100.0 * r["residual_raw"] for r in g]

        ax.plot(
            t,
            y,
            marker="o",
            markersize=2,
            linewidth=0.8,
            alpha=0.55,
            label=(
                f"{i}: S{rec['sat']} "
                f"{'rise' if rec['rise'] == 1 else 'set'} "
                f"Az {rec['az_mean']:.1f}°"
            ),
        )

    ax.axhline(0, linestyle="--", linewidth=1.5)
    ax.set_xlabel("UTC")
    ax.set_ylabel("GNSS-R − EOT20 residual (cm)")
    ax.set_title("GOOD-track residuals: GNSS-R − EOT20")
    ax.grid(alpha=0.25)
    ax.legend(fontsize=7, ncol=3)

    savefig(PLOT_DIR / "05_good_track_residuals_raw.png")


def plot_good_population_residual_comparison(good_rows):
    raw = np.asarray(
        [100.0 * r["residual_raw"] for r in good_rows],
        dtype=float,
    )
    plus = np.asarray(
        [100.0 * r["residual_plus0242"] for r in good_rows],
        dtype=float,
    )

    fig, ax = plt.subplots(figsize=(14, 7))

    ax.hist(
        raw,
        bins=35,
        alpha=0.55,
        label="Raw GNSS-R − EOT20",
    )
    ax.hist(
        plus,
        bins=35,
        alpha=0.55,
        label="GNSS-R + 0.242 m − EOT20",
    )

    ax.axvline(0, linestyle="--", linewidth=1.2)
    ax.set_xlabel("Residual (cm)")
    ax.set_ylabel("Number of observations")
    ax.set_title(
        "GOOD-track population residual distribution"
    )
    ax.grid(alpha=0.20)
    ax.legend()

    savefig(PLOT_DIR / "06_good_population_residual_distribution.png")


def plot_daily_good_population(good_rows):
    by_day = defaultdict(list)

    for r in good_rows:
        by_day[r["doy"]].append(r)

    days = sorted(by_day)

    tide_med = []
    gnss_med = []
    gnss_plus_med = []

    for d in days:
        g = by_day[d]
        tide_med.append(
            float(np.median([r["tide"] for r in g]))
        )
        gnss_med.append(
            float(np.median([r["rh"] for r in g]))
        )
        gnss_plus_med.append(
            float(np.median([r["rh"] for r in g]))
            + DATUM_OFFSET_M
        )

    day_dates = [
        datetime(YEAR, 1, 1) + timedelta(days=d - 1)
        for d in days
    ]

    fig, ax = plt.subplots(figsize=(16, 7))

    ax.plot(
        day_dates,
        tide_med,
        linewidth=2.5,
        label="EOT20 daily median",
    )
    ax.plot(
        day_dates,
        gnss_med,
        marker="o",
        linewidth=1.0,
        label="GOOD GNSS-R daily median",
    )
    ax.plot(
        day_dates,
        gnss_plus_med,
        marker="o",
        linewidth=1.0,
        label="GOOD GNSS-R +0.242 m daily median",
    )

    ax.set_xlabel("UTC date")
    ax.set_ylabel("Water level / reflector height (m)")
    ax.set_title(
        "Daily median: GOOD GNSS-R population vs EOT20"
    )
    ax.grid(alpha=0.25)
    ax.legend()

    savefig(PLOT_DIR / "07_daily_good_population_vs_tide.png")


# ============================================================================
# OUTPUT TABLES / SUMMARY
# ============================================================================

def write_track_csv(tracks):
    fields = [
        "rank",
        "sat",
        "freq",
        "rise",
        "n",
        "days",
        "az_mean",
        "az_sd",
        "r",
        "slope",
        "intercept",
        "fit_rms_m",
        "raw_bias_m",
        "raw_mad_m",
        "raw_median_abs_m",
        "raw_rms_m",
        "plus0242_bias_m",
        "plus0242_mad_m",
        "plus0242_median_abs_m",
        "plus0242_rms_m",
        "good",
    ]

    with OUT_CSV.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        for i, rec in enumerate(tracks, start=1):
            w.writerow(
                {
                    "rank": i,
                    "sat": rec["sat"],
                    "freq": rec["freq"],
                    "rise": rec["rise"],
                    "n": rec["n"],
                    "days": rec["days"],
                    "az_mean": rec["az_mean"],
                    "az_sd": rec["az_sd"],
                    "r": rec["r"],
                    "slope": rec["slope"],
                    "intercept": rec["intercept"],
                    "fit_rms_m": rec["fit_rms_m"],
                    "raw_bias_m": rec["raw_bias_m"],
                    "raw_mad_m": rec["raw_mad_m"],
                    "raw_median_abs_m": rec["raw_median_abs_m"],
                    "raw_rms_m": rec["raw_rms_m"],
                    "plus0242_bias_m": rec["plus_bias_m"],
                    "plus0242_mad_m": rec["plus_mad_m"],
                    "plus0242_median_abs_m": rec["plus_median_abs_m"],
                    "plus0242_rms_m": rec["plus_rms_m"],
                    "good": "YES" if is_good_track(rec) else "NO",
                }
            )


def write_population_metrics(good_rows):
    metrics = pooled_metrics(good_rows)

    fields = [
        "population",
        "n_observations",
        "mean_bias_m",
        "mean_absolute_deviation_m",
        "median_absolute_deviation_m",
        "rms_m",
    ]

    with OUT_METRICS.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        for name in ["raw", "plus0242"]:
            m = metrics[name]
            w.writerow(
                {
                    "population": (
                        "raw"
                        if name == "raw"
                        else "GNSS-R + 0.242 m"
                    ),
                    "n_observations": m["n"],
                    "mean_bias_m": m["mean_bias_m"],
                    "mean_absolute_deviation_m": m["mean_abs_dev_m"],
                    "median_absolute_deviation_m": m["median_abs_dev_m"],
                    "rms_m": m["rms_m"],
                }
            )

    return metrics


def write_inventory(inventory, successful, failed):
    with OUT_INVENTORY.open("w", newline="") as f:
        fields = [
            "doy",
            "rinex",
            "snr",
            "rinex_to_snr_returncode",
            "snr_valid",
            "snr_size_before",
            "snr_size_after",
            "gnssir_result_created",
        ]
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()

        by_doy = {x["doy"]: x for x in inventory}

        for doy in sorted(by_doy):
            x = dict(by_doy[doy])
            x["gnssir_result_created"] = (
                "YES" if doy in successful else "NO"
            )
            w.writerow(x)


def write_summary(
    rinex_days,
    snr_days,
    successful,
    failed,
    all_rows,
    good_tracks,
    good_rows,
    metrics,
):
    with OUT_SUMMARY.open("w") as f:
        f.write("MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V3\n")
        f.write("=" * 92 + "\n\n")

        f.write(f"Year: {YEAR}\n")
        f.write(f"RINEX days discovered: {len(rinex_days)}\n")
        f.write(f"SNR days available: {len(snr_days)}\n")
        f.write(f"GNSSIR result days created: {len(successful)}\n")
        f.write(f"GNSSIR result days failed: {len(failed)}\n\n")

        f.write(
            "RINEX DOYs:\n  "
            + ", ".join(str(x) for x in rinex_days)
            + "\n\n"
        )
        f.write(
            "GNSSIR successful DOYs:\n  "
            + ", ".join(str(x) for x in successful)
            + "\n\n"
        )
        f.write(
            "GNSSIR failed DOYs:\n  "
            + (", ".join(str(x) for x in failed) if failed else "NONE")
            + "\n\n"
        )

        f.write(f"All GPS L1 GNSS-R retrievals: {len(all_rows)}\n")
        f.write(f"Good tracks: {len(good_tracks)}\n")
        f.write(f"Observations in good tracks: {len(good_rows)}\n\n")

        f.write("GOOD-TRACK CRITERIA\n")
        f.write("-" * 92 + "\n")
        f.write(f"observations >= {MIN_OBS}\n")
        f.write(f"unique days   >= {MIN_DAYS}\n")
        f.write(f"tide r        >= {MIN_R:.2f}\n")
        f.write(f"slope         = {MIN_SLOPE:.2f} to {MAX_SLOPE:.2f}\n")
        f.write(f"fitted/unit RMS <= {MAX_FIT_RMS_M*100:.1f} cm\n")
        f.write(f"azimuth SD    <= {MAX_AZ_SD_DEG:.1f} deg\n\n")

        f.write(
            f"Diagnostic GNSS-R offset: +{DATUM_OFFSET_M:.3f} m\n"
        )
        f.write(
            "IMPORTANT: the +0.242 m offset is NOT used to select good tracks.\n\n"
        )

        f.write("GOOD TRACKS\n")
        f.write("-" * 92 + "\n")

        for i, rec in enumerate(good_tracks, start=1):
            f.write(
                f"{i:2d} SAT={rec['sat']:3d} "
                f"rise={rec['rise']:2d} "
                f"N={rec['n']:3d} "
                f"days={rec['days']:3d} "
                f"Az={rec['az_mean']:7.2f}±{rec['az_sd']:.2f} "
                f"r={rec['r']:+.4f} "
                f"slope={rec['slope']:+.4f} "
                f"fitRMS={100*rec['fit_rms_m']:.2f}cm "
                f"rawBias={100*rec['raw_bias_m']:+.2f}cm "
                f"rawMAD={100*rec['raw_mad_m']:.2f}cm "
                f"rawMedAbs={100*rec['raw_median_abs_m']:.2f}cm "
                f"rawRMS={100*rec['raw_rms_m']:.2f}cm "
                f"+24.2Bias={100*rec['plus_bias_m']:+.2f}cm "
                f"+24.2MAD={100*rec['plus_mad_m']:.2f}cm "
                f"+24.2MedAbs={100*rec['plus_median_abs_m']:.2f}cm "
                f"+24.2RMS={100*rec['plus_rms_m']:.2f}cm\n"
            )

        f.write("\n")
        f.write("POOLED GOOD-TRACK POPULATION METRICS\n")
        f.write("-" * 92 + "\n")
        f.write(
            "These are point-by-point pooled statistics across every observation "
            "in every good track. They are NOT an average of track RMS values.\n\n"
        )

        for label, key in [
            ("RAW GNSS-R", "raw"),
            ("GNSS-R + 0.242 m", "plus0242"),
        ]:
            m = metrics[key]
            f.write(f"{label}\n")
            f.write(f"  observations              : {m['n']}\n")
            f.write(
                f"  mean bias                 : "
                f"{100*m['mean_bias_m']:+.3f} cm\n"
            )
            f.write(
                f"  mean absolute deviation  : "
                f"{100*m['mean_abs_dev_m']:.3f} cm\n"
            )
            f.write(
                f"  median absolute deviation: "
                f"{100*m['median_abs_dev_m']:.3f} cm\n"
            )
            f.write(
                f"  RMS                       : "
                f"{100*m['rms_m']:.3f} cm\n\n"
            )


def print_final_summary(
    rinex_days,
    snr_days,
    successful,
    failed,
    all_rows,
    tracks,
    good_tracks,
    good_rows,
    metrics,
):
    print()
    print("=" * 92)
    print("FINAL MARCONI PROCESSING INVENTORY")
    print("=" * 92)

    print(f"RINEX days discovered : {len(rinex_days)}")
    print(f"SNR days available    : {len(snr_days)}")
    print(f"Result days created   : {len(successful)}")
    print(f"Failed result days    : {len(failed)}")
    print(f"Total L1 retrievals   : {len(all_rows)}")
    print(f"Clustered tracks      : {len(tracks)}")
    print(f"GOOD tracks            : {len(good_tracks)}")
    print(f"GOOD-track observations: {len(good_rows)}")

    if failed:
        print(
            "Failed DOYs:",
            ", ".join(str(x) for x in failed),
        )

    print()
    print("=" * 92)
    print("POOLED GOOD-TRACK POPULATION METRICS")
    print("=" * 92)
    print(
        "These are point-by-point statistics across all observations "
        "in the good tracks."
    )
    print()

    for label, key in [
        ("RAW GNSS-R", "raw"),
        ("GNSS-R + 0.242 m", "plus0242"),
    ]:
        m = metrics[key]
        print(label)
        print(f"  N observations             = {m['n']}")
        print(f"  Mean bias                  = {100*m['mean_bias_m']:+.3f} cm")
        print(f"  Mean absolute deviation   = {100*m['mean_abs_dev_m']:.3f} cm")
        print(f"  Median absolute deviation = {100*m['median_abs_dev_m']:.3f} cm")
        print(f"  RMS                        = {100*m['rms_m']:.3f} cm")
        print()

    print("=" * 92)
    print("OUTPUTS")
    print("=" * 92)
    print(f"Track CSV   : {OUT_CSV}")
    print(f"Metrics CSV : {OUT_METRICS}")
    print(f"Inventory   : {OUT_INVENTORY}")
    print(f"Summary     : {OUT_SUMMARY}")
    print(f"Plots       : {PLOT_DIR}")
    print("DONE")


# ============================================================================
# MAIN
# ============================================================================

def main():
    if not BASE_DIR.exists():
        raise SystemExit(f"Base directory does not exist: {BASE_DIR}")

    if not TIDE_FILE.exists():
        raise SystemExit(f"Tide workbook not found: {TIDE_FILE}")

    print()
    print("=" * 92)
    print("MARCONI LONG-TERM GNSS-R / TIDE PIPELINE V3")
    print("=" * 92)
    print(f"Base directory: {BASE_DIR}")
    print(f"REFL_CODE     : {REFL_CODE}")
    print(f"RINEX search  : {', '.join(str(x) for x in SEARCH_ROOTS)}")
    print(f"Primary tide  : {PRIMARY_TIDE_MODEL}")
    print(f"GNSSIR config : {GNSSIR_EXTENSION}")
    print(f"Diagnostic offset: +{DATUM_OFFSET_M:.3f} m")
    print()

    ensure_directories()
    env = make_env()

    preferred, all_copies = discover_rinex()

    if not preferred:
        raise SystemExit(
            "No local USGS RINEX files were discovered."
        )

    staged = stage_rinex(preferred)

    rinex_inventory = convert_all_rinex(staged, env)

    snr_days, successful, failed = process_all_snr(env)

    tide_at, tide_start, tide_end = load_tide_interpolator()

    all_rows = load_all_retrievals(successful)
    all_rows = add_tides(all_rows, tide_at)

    if not all_rows:
        raise SystemExit(
            "No GNSS-R retrievals with valid tide-model values were found."
        )

    tracks = build_tracks(all_rows)
    good_tracks = [r for r in tracks if is_good_track(r)]

    good_rows = []
    for rec in good_tracks:
        good_rows.extend(rec["group"])

    good_rows.sort(key=lambda r: r["datetime_utc"])

    metrics = pooled_metrics(good_rows)

    print()
    print("=" * 92)
    print("GOOD TRACK SCREEN")
    print("=" * 92)
    print(f"Total clustered tracks: {len(tracks)}")
    print(f"GOOD tracks:           {len(good_tracks)}")

    for i, rec in enumerate(good_tracks, start=1):
        print(
            f"{i:2d} SAT={rec['sat']:3d} "
            f"rise={rec['rise']:2d} "
            f"N={rec['n']:3d} "
            f"days={rec['days']:3d} "
            f"Az={rec['az_mean']:7.2f}±{rec['az_sd']:.2f} "
            f"r={rec['r']:+.4f} "
            f"slope={rec['slope']:+.4f} "
            f"fitRMS={100*rec['fit_rms_m']:.2f}cm "
            f"rawRMS={100*rec['raw_rms_m']:.2f}cm "
            f"+24.2RMS={100*rec['plus_rms_m']:.2f}cm"
        )

    print()
    print("=" * 92)
    print("GENERATING GOOD-TRACK PLOTS")
    print("=" * 92)

    plot_good_tracks_vs_tide(good_tracks, offset=False)
    plot_good_tracks_vs_tide(good_tracks, offset=True)
    plot_good_tracks_scatter(good_tracks, offset=False)
    plot_good_tracks_scatter(good_tracks, offset=True)
    plot_good_track_residuals(good_tracks)
    plot_good_population_residual_comparison(good_rows)
    plot_daily_good_population(good_rows)

    write_track_csv(tracks)
    write_population_metrics(good_rows)
    write_inventory(rinex_inventory, successful, failed)
    write_summary(
        sorted(preferred),
        snr_days,
        successful,
        failed,
        all_rows,
        good_tracks,
        good_rows,
        metrics,
    )

    print_final_summary(
        sorted(preferred),
        snr_days,
        successful,
        failed,
        all_rows,
        tracks,
        good_tracks,
        good_rows,
        metrics,
    )


if __name__ == "__main__":
    main()
